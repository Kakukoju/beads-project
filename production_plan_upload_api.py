# -*- coding: utf-8 -*-
"""Flask route for Production Plan watcher uploads."""

from __future__ import annotations

import logging
import os
import re
import tempfile
import traceback
from datetime import datetime

import openpyxl
from flask import jsonify, request
from sqlalchemy import text


SHEET_NAME = "P_plan Reagent"
HEADER_ROW = 2


def _normalize_date_header(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    value_text = str(value or "").strip()
    match = re.match(r"^(\d{4}-\d{2}-\d{2})(?:\s|$)", value_text)
    return match.group(1) if match else None


def _read_incremental_rows(workbook_path, since_date):
    workbook = openpyxl.load_workbook(
        workbook_path, read_only=True, data_only=True
    )
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"找不到工作表 {SHEET_NAME}")
        sheet = workbook[SHEET_NAME]
        row_iterator = sheet.iter_rows(min_row=HEADER_ROW, values_only=True)
        header_values = next(row_iterator)
        panel_column = next(
            (
                index
                for index, value in enumerate(header_values)
                if re.search(
                    r"panel.*no|panel_no", str(value or ""), re.IGNORECASE
                )
            ),
            None,
        )
        if panel_column is None:
            raise ValueError("找不到 Panel_NO 欄")

        date_columns = [
            (index, normalized)
            for index, value in enumerate(header_values)
            if (normalized := _normalize_date_header(value))
            and normalized >= since_date
        ]
        if not date_columns:
            raise ValueError(f"找不到 {since_date} 起的日期欄")

        rows = []
        for row_values in row_iterator:
            panel_value = row_values[panel_column]
            if panel_value is None or str(panel_value).strip() == "":
                continue
            if isinstance(panel_value, float) and panel_value.is_integer():
                panel_no = str(int(panel_value))
            else:
                panel_no = str(panel_value).strip()
            if re.search(r"panel.*no", panel_no, re.IGNORECASE):
                continue

            values = {}
            for column, normalized in date_columns:
                raw_value = row_values[column]
                if raw_value is None or raw_value == "":
                    values[normalized] = None
                    continue
                try:
                    numeric_value = float(raw_value)
                except (TypeError, ValueError):
                    values[normalized] = None
                else:
                    values[normalized] = (
                        str(numeric_value) if numeric_value > 0 else None
                    )
            rows.append((panel_no, values))
        return date_columns, rows
    finally:
        workbook.close()


def register_production_plan_upload_api(
    app, db, upload_api_key, excel_data_dir, set_sync_source
):
    endpoint = "upload_production_plan_incremental"
    if endpoint in app.view_functions:
        return

    @app.route(
        "/api/upload-production-plan-incremental",
        methods=["POST"],
        endpoint=endpoint,
    )
    def upload_production_plan_incremental():
        if request.headers.get("X-Api-Key") != upload_api_key:
            return jsonify({"ok": False, "error": "Unauthorized"}), 401

        uploaded = request.files.get("file")
        since_date = str(request.form.get("since_date", "")).strip()
        source_path = str(request.form.get("source_path", "")).strip()
        if not uploaded or not uploaded.filename:
            return jsonify({"ok": False, "error": "沒有 file 欄位"}), 400
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", since_date):
            return jsonify(
                {"ok": False, "error": "since_date 必須是 YYYY-MM-DD"}
            ), 400
        try:
            datetime.strptime(since_date, "%Y-%m-%d")
        except ValueError:
            return jsonify(
                {"ok": False, "error": "since_date 不是有效日期"}
            ), 400

        temporary = tempfile.NamedTemporaryFile(
            prefix=".production_plan.",
            suffix=".uploading.xlsm",
            dir=excel_data_dir,
            delete=False,
        )
        temporary_path = temporary.name
        temporary.close()
        save_path = os.path.join(excel_data_dir, "production_plan.xlsm")

        try:
            uploaded.save(temporary_path)
            date_columns, rows = _read_incremental_rows(
                temporary_path, since_date
            )

            with db.engine.begin() as connection:
                existing = [
                    row[0]
                    for row in connection.execute(
                        text(
                            "SELECT column_name "
                            "FROM information_schema.columns "
                            "WHERE table_schema='schedule' "
                            "AND table_name='production_Plan'"
                        )
                    ).fetchall()
                ]
                if not existing:
                    connection.execute(
                        text(
                            'CREATE TABLE schedule."production_Plan" ('
                            '"Plan" TEXT, "Panel_NO" TEXT)'
                        )
                    )
                    existing = ["Plan", "Panel_NO"]

                for _, date_column in date_columns:
                    if date_column not in existing:
                        connection.execute(
                            text(
                                'ALTER TABLE schedule."production_Plan" '
                                f'ADD COLUMN "{date_column}" TEXT'
                            )
                        )

                clear_clause = ", ".join(
                    f'"{date_column}" = NULL'
                    for _, date_column in date_columns
                )
                connection.execute(
                    text(
                        'UPDATE schedule."production_Plan" '
                        f"SET {clear_clause} WHERE \"Plan\" = 'Plan'"
                    )
                )

                updated_rows = 0
                inserted_rows = 0
                for panel_no, values in rows:
                    assignments = []
                    params = {"panel_no": panel_no}
                    for index, (_, date_column) in enumerate(date_columns):
                        param_name = f"value_{index}"
                        assignments.append(
                            f'"{date_column}" = :{param_name}'
                        )
                        params[param_name] = values[date_column]
                    result = connection.execute(
                        text(
                            'UPDATE schedule."production_Plan" '
                            f'SET {", ".join(assignments)} '
                            'WHERE "Panel_NO"::text = :panel_no '
                            "AND \"Plan\" = 'Plan'"
                        ),
                        params,
                    )
                    if result.rowcount:
                        updated_rows += result.rowcount
                        continue

                    columns = ['"Plan"', '"Panel_NO"'] + [
                        f'"{date_column}"'
                        for _, date_column in date_columns
                    ]
                    placeholders = [":plan", ":panel_no"] + [
                        f":value_{index}"
                        for index in range(len(date_columns))
                    ]
                    params["plan"] = "Plan"
                    connection.execute(
                        text(
                            'INSERT INTO schedule."production_Plan" '
                            f'({", ".join(columns)}) '
                            f'VALUES ({", ".join(placeholders)})'
                        ),
                        params,
                    )
                    inserted_rows += 1

                set_sync_source(
                    connection, "production_Plan", "watcher", len(rows)
                )

            os.replace(temporary_path, save_path)
            logging.info(
                "[ProductionPlan] %s -> %s, since=%s, rows=%s, dates=%s",
                source_path or uploaded.filename,
                save_path,
                since_date,
                len(rows),
                len(date_columns),
            )
            return jsonify(
                {
                    "ok": True,
                    "saved_as": "production_plan.xlsm",
                    "since_date": since_date,
                    "rows": len(rows),
                    "updated_rows": updated_rows,
                    "inserted_rows": inserted_rows,
                    "date_columns": len(date_columns),
                }
            )
        except Exception as exc:
            logging.error(
                "[ProductionPlan] incremental upload failed: %s",
                traceback.format_exc(),
            )
            return jsonify({"ok": False, "error": str(exc)}), 500
        finally:
            if os.path.exists(temporary_path):
                try:
                    os.remove(temporary_path)
                except OSError:
                    pass
