# -*- coding: utf-8 -*-
"""
scheduler_api.py  (V33.31-AWS)
AWS 遷移版：SQLite → PostgreSQL (SQLAlchemy)

V33.31 修正：
  [FIX-1] run_solver 加入 target_date 參數，供 staff off days weekday→offset 轉換
  [FIX-2] generate_schedule 預處理 resource_config：
            ① holidays → forbidden_day_offsets（weekday → day offset set）
            ② dryerMaintenance → 從 compatible_dryers 移除維修中機台
            ③ ivekMaintenance / ivekGlobalMaintenance → 過濾 IVEK job
  [FIX-3] run_solver 套用 CP-SAT 約束：
            ① forbidden_day_offsets → base_day != fd
            ④ staffOffDays → 特定人員在特定天不可被分配
  [FIX-4] 效能：max_time_in_seconds 60→30，加 num_search_workers=4
"""

import re
import pandas as pd
import collections
import math
import openpyxl
import os
import logging
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from ortools.sat.python import cp_model
from sqlalchemy import text

# ─────────────────────────────────────────────────────────────────────────────
# 常數
# ─────────────────────────────────────────────────────────────────────────────
START_HOUR          = 10
MINS_PER_GRID       = 30
DEFAULT_GRIDS       = 31
PREP_ONLY_ROW_START = 38
OUTPUT_DIR          = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILENAME     = os.path.join(OUTPUT_DIR,
    f"滴定排程結果_{datetime.now().strftime('%Y%m%d')}.xlsx")
FREEZE_BUFFER_MINS  = 150
PORT_GAP_GRIDS      = 1

# ── Na-IVEK 特例 ──────────────────────────────────────────────────────────────
NA_IVEK_PNS = frozenset({
    '5714400209', '5714400210',
    '5714400202', '5714400203',
    '5714400266', '5714400268',
    '5714400267', '5714400269',
})
NA_IVEK_FREEZE_MINS = round(20.42 * 60)

_NA_IVEK_UD_PAIRS = [
    ('5714400202', '5714400203'),
    ('5714400209', '5714400210'),
    ('5714400268', '5714400269'),
    ('5714400266', '5714400267'),
]

# ── tCREA / QCREA 三合一特例 ──────────────────────────────────────────────────
TCREA_PAIRED_PNS = frozenset({'5714400180', '5714400181'})
TCREA_SOLO_PN    = '5714400182'
QCREA_PAIRED_PNS = frozenset({'5714400236', '5714400237'})
QCREA_SOLO_PN    = '5714400238'

_TRIO_SOLO_REAGENT = {
    TCREA_SOLO_PN: 'tCREA_C',
    QCREA_SOLO_PN: 'QCREA_C',
}
_TRIO_GROUPS = [
    (TCREA_PAIRED_PNS, TCREA_SOLO_PN, 'tCREA_D_PAIR', 'tCREA_C'),
    (QCREA_PAIRED_PNS, QCREA_SOLO_PN, 'QCREA_D_PAIR', 'QCREA_C'),
]
_TRIO_ALL_PNS = TCREA_PAIRED_PNS | {TCREA_SOLO_PN} | QCREA_PAIRED_PNS | {QCREA_SOLO_PN}

FORCE_GROUP_MAP = {
    '5714400132': 'GGT_RGT',
    '5714400201': 'GGT_RGT',
    '5714400167': 'FRU',
    '5714400168': 'FRU',
    '5714400180': 'tCREA_D_PAIR',
    '5714400181': 'tCREA_D_PAIR',
    '5714400236': 'QCREA_D_PAIR',
    '5714400237': 'QCREA_D_PAIR',
}

GLIPA_PN_AU = '5714400221'
GLIPA_PN_AD = '5714400220'
QLIPA_PN_AU = '5714400293'
QLIPA_PN_AD = '5714400294'

DUAL_CONTAM_PNS = frozenset({
    '5714400116', '5714400117',
    '5714400216', '5714400226',
    '5714400214', '5714400215',
    '5714400199', '5714400222',
})
DUAL_CONTAM_DRYERS = frozenset({'5', '7'})
DUAL_BLOCK_DAYS    = 2
DUAL_PORT_GAP      = 1

_DUAL_PAIRS = [
    (GLIPA_PN_AU, GLIPA_PN_AD),
    (QLIPA_PN_AU, QLIPA_PN_AD),
]
_DUAL_AU_PNS  = frozenset(p[0] for p in _DUAL_PAIRS)
_DUAL_AD_PNS  = frozenset(p[1] for p in _DUAL_PAIRS)
_DUAL_ALL_PNS = _DUAL_AU_PNS | _DUAL_AD_PNS
_DUAL_PARTNER_MAP: dict = {}
for _au, _ad in _DUAL_PAIRS:
    _DUAL_PARTNER_MAP[_au] = (_au, _ad)
    _DUAL_PARTNER_MAP[_ad] = (_au, _ad)

GLIPA_CONTAM_PNS    = DUAL_CONTAM_PNS
GLIPA_CONTAM_DRYERS = DUAL_CONTAM_DRYERS
GLIPA_BLOCK_DAYS    = DUAL_BLOCK_DAYS

_GROUP_TO_PNS = collections.defaultdict(list)
for _pn, _grp in FORCE_GROUP_MAP.items():
    _GROUP_TO_PNS[_grp].append(_pn)

# ── BeadResource staffId → name fragment 對照表 ───────────────────────────────
# 與 BeadResource.tsx staffList 保持一致
_STAFF_ID_MAP = {
    "S01": "Suyo.Lin",
    "S02": "Crystal.Chen",
    "S03": "Angala Liu",
    "S04": "Dabby Wang",
    "S05": "Patty Li",
    "S06": "Wendy.HW.Chang",
    "S07": "Yayu.Huang",
}

_WD_MAP = {'一': 0, '二': 1, '三': 2, '四': 3, '五': 4, '六': 5, '日': 6}


# ─────────────────────────────────────────────────────────────────────────────
# 工具函數
# ─────────────────────────────────────────────────────────────────────────────

def safe_float(val):
    try:
        if pd.isna(val) or str(val).strip() == '':
            return 0.0
        return float(str(val).replace(',', ''))
    except Exception:
        return 0.0


def _compute_forbidden_day_offsets(holidays, target_date, horizon_days=14):
    """
    holidays: List[str]，例如 ['六', '日']
    回傳 Set[int]，例如 {5, 6, 12, 13}（距 target_date 的 day offset）
    """
    holiday_wdays = {_WD_MAP[h] for h in (holidays or []) if h in _WD_MAP}
    if not holiday_wdays:
        return set()
    base_dt = datetime.strptime(target_date, "%Y-%m-%d")
    return {
        offset for offset in range(horizon_days)
        if (base_dt + timedelta(days=offset)).weekday() in holiday_wdays
    }


def _compute_staff_off_offsets(staff_off_days_raw, target_date, horizon_days=14):
    """
    staff_off_days_raw: Dict[staffId, List[str]]，例如 {"S01": ["六"], "S03": ["日"]}
    回傳 Dict[staffId, Set[int]]，例如 {"S01": {5, 12}, "S03": {6, 13}}
    """
    if not staff_off_days_raw:
        return {}
    base_dt = datetime.strptime(target_date, "%Y-%m-%d")
    result = {}
    for staff_id, off_weekdays in staff_off_days_raw.items():
        off_wdays = {_WD_MAP[d] for d in off_weekdays if d in _WD_MAP}
        if not off_wdays:
            continue
        result[staff_id] = {
            offset for offset in range(horizon_days)
            if (base_dt + timedelta(days=offset)).weekday() in off_wdays
        }
    return result


# ─────────────────────────────────────────────────────────────────────────────
# load_db_data
# ─────────────────────────────────────────────────────────────────────────────

def load_db_data(db, target_date=None):
    try:
        engine = db.engine

        with engine.connect() as conn:
            df_rules = pd.read_sql_query(
                text('SELECT * FROM "配藥限制"'), conn)

        _today = datetime.now().strftime('%Y-%m-%d')
        with engine.connect() as conn:
            df_demand = pd.read_sql_query(
                text('SELECT * FROM "BeadNeed" WHERE date = :d'),
                conn, params={'d': _today})
            if df_demand.empty:
                df_demand = pd.read_sql_query(
                    text('SELECT * FROM "BeadNeed" '
                         'WHERE date = (SELECT MAX(date) FROM "BeadNeed")'), conn)
                _used_row = pd.read_sql_query(
                    text('SELECT MAX(date) AS d FROM "BeadNeed"'), conn)
                _used = _used_row.iloc[0]['d']
                print(f"[BeadNeed] 今日({_today})無資料，使用最新日期 {_used}")
            else:
                print(f"[BeadNeed] 使用日期: {_today}，共 {len(df_demand)} 筆")

        try:
            with engine.connect() as conn:
                df_forced = pd.read_sql_query(
                    text('SELECT * FROM "限制OR插單"'), conn)
        except Exception as e:
            logging.warning(f"[限制OR插單] 讀取失敗，略過：{e}")
            df_forced = pd.DataFrame()

        try:
            with engine.connect() as conn:
                df_plan = pd.read_sql_query(
                    text("SELECT * FROM \"production_Plan\" WHERE \"Plan\" = 'Plan'"), conn)
        except Exception as e:
            logging.warning(f"[production_Plan] 讀取失敗，略過：{e}")
            df_plan = pd.DataFrame()

    except Exception as e:
        logging.error(f"[load_db_data] RDS 連線/查詢失敗: {e}")
        return None, None, None, None, None, None

    rules, reagent_groups, all_staff = {}, collections.defaultdict(list), set()

    for _, row in df_rules.iterrows():
        pn = str(row.get('PN', '')).strip().replace('.0', '')
        if not pn:
            continue
        name     = str(row.get('Name', ''))
        reagent  = name.split('-')[0].strip()
        port_val = str(row.get('Port數', '2')).strip().upper()

        is_ivek   = (pn in NA_IVEK_PNS) or (port_val == 'IVEK')
        num_ports = 0 if is_ivek else (int(port_val) if port_val.isdigit() else 2)

        qty_raw      = str(row.get('數量', '0')).strip()
        is_prep_only = (qty_raw == '' or qty_raw == '0')
        qty_opts     = ([safe_float(x.strip()) for x in qty_raw.lower().replace('or', ',').split(',')
                         if x.strip()] if not is_prep_only else [0.0])
        staff = [str(row.get(f'配藥人-{i}', '')).strip() for i in range(1, 4)]
        staff = [s for s in staff if s and s.lower() != 'nan']
        for s in staff:
            all_staff.add(s)

        split_ud = (pn in NA_IVEK_PNS)
        if split_ud:
            _ud_suffix = '_U' if (name.endswith('-U') or name.endswith('-AU')
                                  or name.endswith('-BU')) else '_D'
            effective_reagent = FORCE_GROUP_MAP.get(pn, reagent + _ud_suffix)
        else:
            effective_reagent = FORCE_GROUP_MAP.get(pn, reagent)

        if pn in _TRIO_SOLO_REAGENT:
            effective_reagent = _TRIO_SOLO_REAGENT[pn]

        if pn in NA_IVEK_PNS:
            freeze_mins = NA_IVEK_FREEZE_MINS
        else:
            freeze_mins = int(safe_float(row.get('凍乾時間')) * 60) if row.get('凍乾時間') else 720

        rules[pn] = {
            'pn': pn, 'name': name, 'reagent': effective_reagent, 'num_ports': num_ports,
            'qty_options': sorted(qty_opts), 'is_prep_only': is_prep_only,
            'is_ivek': is_ivek, 'split_ud': split_ud,
            'duration_freeze': freeze_mins,
            'compatible_dryers': [d.strip() for d in str(row.get('可用凍乾機', '')).split(',')
                                  if d.strip()],
            'compatible_staff': staff,
            'delivery_days': str(row.get('交藥時間', '')).strip(),
        }
        if effective_reagent:
            reagent_groups[effective_reagent].append(pn)

    return rules, reagent_groups, sorted(list(all_staff)), df_demand, df_forced, df_plan


# ─────────────────────────────────────────────────────────────────────────────
# load_forced_jobs
# ─────────────────────────────────────────────────────────────────────────────

def load_forced_jobs(df_forced, rules, target_date):
    if df_forced is None or df_forced.empty:
        return []

    marker_to_pn = {}
    for pn, r in rules.items():
        marker_to_pn[r['name'].strip()] = pn

    base_dt = (datetime.strptime(target_date, "%Y-%m-%d") if target_date
               else datetime.now().replace(hour=0, minute=0, second=0, microsecond=0))

    REQUIRED_COLS = ['日期', 'Marker', '預計滴定時間', '滴定機', '凍乾機台', '配藥同仁', '數量']
    COL_ALIAS     = {'凍乾機台': ['凍乾機台', '可用凍乾機']}
    forced_jobs   = []

    for i, row in df_forced.iterrows():
        row_num = i + 2
        missing = []
        for col in REQUIRED_COLS:
            candidates = COL_ALIAS.get(col, [col])
            found = any(row.get(c) is not None and str(row.get(c)).strip() not in ('', 'nan')
                        for c in candidates)
            if not found:
                missing.append(col)
        if missing:
            print(f"⚠️  [限制OR插單] 第{row_num}筆缺少欄位 {missing}，略過此筆")
            continue

        date_str  = str(row.get('日期', '')).strip()
        marker    = str(row.get('Marker', '')).strip()
        time_str  = str(row.get('預計滴定時間', '')).strip()
        port_raw  = str(row.get('滴定機', '')).strip()
        dryer_raw = str(row.get('凍乾機台') or row.get('可用凍乾機', '')).strip()
        staff_raw = str(row.get('配藥同仁', '')).strip()
        qty       = safe_float(row.get('數量', 0))

        job_date = None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d", "%m-%d"):
            try:
                parsed = datetime.strptime(date_str, fmt)
                if fmt in ("%m/%d", "%m-%d"):
                    parsed = parsed.replace(year=base_dt.year)
                job_date = parsed
                break
            except ValueError:
                continue
        if job_date is None:
            m = re.match(r'(\d{1,2})月(\d{1,2})日', date_str)
            if m:
                try:
                    job_date = datetime(base_dt.year, int(m.group(1)), int(m.group(2)))
                except ValueError:
                    pass
        if job_date is None:
            print(f"⚠️  [限制OR插單] 第{row_num}筆日期格式錯誤 '{date_str}'，略過")
            continue

        try:
            t = datetime.strptime(time_str, "%H:%M")
        except ValueError:
            print(f"⚠️  [限制OR插單] 第{row_num}筆時間格式錯誤 '{time_str}'，略過")
            continue

        day_offset = (job_date - base_dt).days
        if day_offset < 0:
            print(f"⚠️  [限制OR插單] 第{row_num}筆日期 {date_str} 早於排程起始日，略過")
            continue

        time_grid_in_day = (t.hour * 60 + t.minute - START_HOUR * 60) // MINS_PER_GRID
        if time_grid_in_day < 0 or time_grid_in_day >= DEFAULT_GRIDS:
            print(f"⚠️  [限制OR插單] 第{row_num}筆時間 {time_str} 超出排程時間範圍，略過")
            continue

        pinned_grid = day_offset * DEFAULT_GRIDS + time_grid_in_day

        try:
            port_num_str = port_raw.upper().replace('PORT', '').strip()
            pinned_port  = int(float(port_num_str)) - 1
            if pinned_port < 0 or pinned_port > 11:
                raise ValueError
        except ValueError:
            print(f"⚠️  [限制OR插單] 第{row_num}筆滴定機 '{port_raw}' 無效，略過")
            continue

        pn = marker_to_pn.get(marker)
        if not pn:
            for name_key, pn_val in marker_to_pn.items():
                if marker.lower() in name_key.lower() or name_key.lower() in marker.lower():
                    pn = pn_val
                    break
        if not pn:
            pn = f"EXP_{marker}"
            if pn not in rules:
                rules[pn] = {
                    'pn': pn, 'name': marker, 'reagent': f"EXP_{marker}",
                    'num_ports': 1, 'qty_options': [qty], 'is_prep_only': False,
                    'is_ivek': False, 'split_ud': False,
                    'duration_freeze': 1200,
                    'compatible_dryers': [dryer_raw],
                    'compatible_staff': [staff_raw]
                }
            print(f"ℹ️   [限制OR插單] 第{row_num}筆 Marker '{marker}' 無對應 PN，以實驗插單處理")

        r        = rules[pn]
        divisor  = 1500 if r['num_ports'] > 0 else 800
        grid_dur = (math.ceil((qty / max(1, r['num_ports']) / divisor * 60 + 30) / 30)
                    if not r['is_prep_only'] else 2)

        if time_grid_in_day + grid_dur > DEFAULT_GRIDS:
            grid_dur = DEFAULT_GRIDS - time_grid_in_day
            print(f"ℹ️   [限制OR插單] 第{row_num}筆滴定時長超出當天邊界，截斷為 {grid_dur} grids")

        forced_jobs.append({
            **r,
            'id': f"FORCED_{i}_{pn}",
            'qty': qty, 'grid_duration': grid_dur, 'prio': 0,
            'forced': True,
            'pinned_grid': pinned_grid, 'pinned_port': pinned_port,
            'pinned_dryer': dryer_raw, 'pinned_staff': staff_raw,
            'pinned_day': day_offset,
        })
        print(f"✅  [限制OR插單] 第{row_num}筆插單：{marker}({pn}) {date_str} {time_str} "
              f"Port{pinned_port+1} 機:{dryer_raw} 人:{staff_raw}")

    return forced_jobs


# ─────────────────────────────────────────────────────────────────────────────
# build_ordered_prep_jobs
# ─────────────────────────────────────────────────────────────────────────────

def build_ordered_prep_jobs(rules, df_demand, df_plan, target_date,
                            holidays=None, max_horizon_days=14):
    if df_plan is None or df_plan.empty:
        return []
    base_dt = datetime.strptime(target_date, "%Y-%m-%d") if target_date else None
    if not base_dt:
        return []

    holiday_wdays: set = {_WD_MAP[h] for h in (holidays or []) if h in _WD_MAP}

    pp_cols  = list(df_plan.columns)
    plan_idx = next((i for i, c in enumerate(pp_cols) if c.strip().lower() == 'plan'), None)
    pn_col   = next((c for c in pp_cols if re.search(r'panel.*no', c, re.I)), None)
    if plan_idx is None or pn_col is None:
        return []
    date_cols = [c for c in pp_cols[plan_idx+1:]
                 if re.match(r'\d{4}[-/]\d{2}[-/]\d{2}', c)]

    demand_pns = set()
    if df_demand is not None and not df_demand.empty:
        for _, row in df_demand.iterrows():
            if (safe_float(row.get('w1')) > 0 or safe_float(row.get('w2')) > 0
                    or safe_float(row.get('w3')) > 0):
                demand_pns.add(str(row.get('pn', '')).strip().replace('.0', ''))

    ordered_jobs, job_counter = [], 0
    for pn, r in rules.items():
        delivery_rule = str(r.get('delivery_days', '')).strip()
        if delivery_rule not in ('前兩天下班', '前一天下班'):
            continue
        if pn not in demand_pns:
            continue
        offset_days = 2 if delivery_rule == '前兩天下班' else 1
        plan_rows = df_plan[df_plan[pn_col].astype(str).str.strip() == pn]
        if plan_rows.empty:
            continue
        for dc in date_cols:
            try:
                qty = safe_float(plan_rows[dc].values[0])
            except (IndexError, KeyError):
                continue
            if qty <= 0:
                continue
            demand_dt  = datetime.strptime(dc, "%Y-%m-%d")
            sched_dt   = demand_dt - timedelta(days=offset_days)
            day_offset = (sched_dt - base_dt).days
            if day_offset < 0:
                continue
            if max_horizon_days and day_offset >= max_horizon_days:
                print(f"[接單配藥] {r['name']}({pn}) 需求:{dc} 超出視窗({max_horizon_days}天)，略過")
                continue
            job_counter += 1
            holiday_conflict = sched_dt.weekday() in holiday_wdays
            ordered_jobs.append({
                **r,
                'id':              f"ORDERED_{job_counter}_{pn}_{dc}",
                'qty':             qty, 'grid_duration': DEFAULT_GRIDS, 'prio': 0,
                'is_prep_only':    True, 'num_ports': 0,
                'ordered_prep':    True, 'pinned_day': day_offset,
                'demand_date':     dc,
                'sched_date':      sched_dt.strftime("%Y-%m-%d"),
                'holiday_conflict': holiday_conflict,
            })
            flag = "⚠️ 排程日為休假日" if holiday_conflict else ""
            print(f"[接單配藥] {r['name']}({pn}) 需求:{dc} 排程:{sched_dt.strftime('%m/%d')} "
                  f"qty:{int(qty)} {flag}")
    return ordered_jobs


# ─────────────────────────────────────────────────────────────────────────────
# build_jobs
# ─────────────────────────────────────────────────────────────────────────────

def build_jobs(rules, df_demand, df_plan=None, selected_pns=None, forced_jobs=None):
    jobs = []

    for _, row in df_demand.iterrows():
        pn = str(row.get('pn', '')).strip().replace('.0', '')
        if pn not in rules:
            continue
        if selected_pns and pn not in selected_pns:
            continue
        r = rules[pn]
        if r.get('delivery_days', '') in ('前兩天下班', '前一天下班'):
            continue

        stock        = safe_float(row.get('stock_unstock'))
        w1           = safe_float(row.get('w1'))
        w2           = safe_float(row.get('w2'))
        w3           = safe_float(row.get('w3'))
        safety_stock = safe_float(row.get('safety_stock', 0))
        wb1 = int(safe_float(row.get('w1_batch', 0)))
        wb2 = int(safe_float(row.get('w2_batch', 0)))
        wb3 = int(safe_float(row.get('w3_batch', 0)))
        w_total_batches = wb1 + wb2 + wb3

        min_qty       = r['qty_options'][0] if r['qty_options'] else 0.0
        planned_total = w_total_batches * min_qty
        shortage      = max(0.0, (w1 + w2 + w3) - stock)
        replenish     = max(0.0, safety_stock - stock)
        true_need     = max(planned_total, shortage, replenish)

        if true_need <= 0:
            continue

        optimal_batches = w_total_batches if w_total_batches > 0 else 1
        optimal_qty     = min_qty
        for opt in sorted(r['qty_options'], reverse=True):
            if opt <= 0:
                continue
            n = math.ceil(true_need / opt)
            if n < optimal_batches:
                optimal_batches = n
                optimal_qty     = opt
                print(f"[批量優化] {r['name']}({pn}) {w_total_batches}批×{min_qty:.0f} "
                      f"→ {n}批×{opt:.0f} (true_need={true_need:.0f})")

        total_batches = max(1, optimal_batches)
        selected_qty  = optimal_qty
        base_prio     = 1 if wb1 > 0 else (2 if wb2 > 0 else 3)

        if r['is_prep_only']:
            grid_dur = 2
        elif r.get('is_ivek'):
            grid_dur = math.ceil((selected_qty / 2 / 1500 * 60 + 30) / 30)
        else:
            divisor  = 1500 if r['num_ports'] > 0 else 800
            grid_dur = math.ceil((selected_qty / max(1, r['num_ports']) / divisor * 60 + 30) / 30)

        for seq in range(total_batches):
            week_prio = (1 if seq < wb1 else (2 if seq < wb1 + wb2 else 3))
            jobs.append({
                **r,
                'id':            f"P{base_prio}_{pn}_B{seq + 1}",
                'qty':           selected_qty,
                'grid_duration': grid_dur,
                'prio':          week_prio,
                'batch_round':   seq + 1,
            })

        if total_batches > 1:
            print(f"[多批展開] {r['name']}({pn}) w1×{wb1} w2×{wb2} w3×{wb3} → {total_batches} jobs")

    job_pns = {j['pn'] for j in jobs}
    for grp_name, grp_pns in _GROUP_TO_PNS.items():
        if not any(pn in job_pns for pn in grp_pns):
            continue
        for pn_miss in grp_pns:
            if pn_miss in job_pns or pn_miss not in rules:
                continue
            r            = rules[pn_miss]
            selected_qty = r['qty_options'][0] if r['qty_options'] else 0.0
            grid_dur     = (2 if r['is_prep_only']
                            else math.ceil((selected_qty / max(1, r['num_ports']) / 1500 * 60 + 30) / 30))
            jobs.append({
                **r,
                'id': f"PAIR_{pn_miss}", 'qty': selected_qty,
                'grid_duration': grid_dur, 'prio': 99, 'batch_round': 1,
            })
            print(f"強制補入配對 [{grp_name}]: {pn_miss} ({r['name']})")

    reagent_pn_max: dict = {}
    for j in jobs:
        if j.get('forced') or j.get('ordered_prep'):
            continue
        rn = j['reagent']
        pn = j['pn']
        b  = j.get('batch_round', 1)
        reagent_pn_max.setdefault(rn, {})[pn] = max(reagent_pn_max.get(rn, {}).get(pn, 0), b)

    for reagent, pn_max_map in reagent_pn_max.items():
        if len(pn_max_map) <= 1:
            continue
        group_max_b = max(pn_max_map.values())
        for pn, cur_max_b in pn_max_map.items():
            if cur_max_b >= group_max_b:
                continue
            template = next(
                (j for j in jobs if j['pn'] == pn and j.get('batch_round', 1) == cur_max_b
                 and not j.get('forced') and not j.get('ordered_prep')), None)
            if not template:
                continue
            for extra_b in range(cur_max_b + 1, group_max_b + 1):
                jobs.append({**template, 'id': f"PAD_{pn}_B{extra_b}",
                             'batch_round': extra_b, 'prio': 3})
                other_pns = [p for p in pn_max_map if p != pn]
                print(f"[配對補批] {template['name']}({pn}) reagent={reagent} "
                      f"補入第{extra_b}批（配對{other_pns}）")

    job_batch_max: dict = {}
    for j in jobs:
        if j.get('forced') or j.get('ordered_prep'):
            continue
        pn = j['pn']
        job_batch_max[pn] = max(job_batch_max.get(pn, 0), j.get('batch_round', 1))

    for u_pn, d_pn in _NA_IVEK_UD_PAIRS:
        u_max = job_batch_max.get(u_pn, 0)
        d_max = job_batch_max.get(d_pn, 0)
        if u_max == 0 and d_max == 0:
            continue
        pair_max = max(u_max, d_max)
        for pn, cur_max in [(u_pn, u_max), (d_pn, d_max)]:
            if cur_max >= pair_max:
                continue
            if pn not in rules:
                continue
            template = next(
                (j for j in jobs if j['pn'] == pn
                 and not j.get('forced') and not j.get('ordered_prep')), None)
            if template is None:
                r = rules[pn]
                selected_qty = r['qty_options'][0] if r['qty_options'] else 12000.0
                grid_dur = math.ceil((selected_qty / 2 / 1500 * 60 + 30) / 30)
                template = {
                    **r,
                    'id': f"PAD_NA_{pn}_B1", 'qty': selected_qty,
                    'grid_duration': grid_dur, 'prio': 3, 'batch_round': 1,
                }
                jobs.append(template)
                print(f"[Na-IVEK 補批] {r['name']}({pn}) 無批次 → 補入第1批")
                cur_max = 1

            for extra_b in range(cur_max + 1, pair_max + 1):
                jobs.append({**template, 'id': f"PAD_NA_{pn}_B{extra_b}",
                             'batch_round': extra_b, 'prio': 3})
                partner = u_pn if pn == d_pn else d_pn
                r_name  = rules.get(pn, {}).get('name', pn)
                print(f"[Na-IVEK 補批] {r_name}({pn}) 補入第{extra_b}批（配對{partner}）")

    _trio_sets = [TCREA_PAIRED_PNS | {TCREA_SOLO_PN}, QCREA_PAIRED_PNS | {QCREA_SOLO_PN}]
    for trio_set in _trio_sets:
        trio_in_jobs = {j['pn'] for j in jobs
                        if j['pn'] in trio_set
                        and not j.get('forced') and not j.get('ordered_prep')}
        if not trio_in_jobs:
            continue

        trio_batch_max: dict = {}
        for j in jobs:
            if j['pn'] in trio_set and not j.get('forced') and not j.get('ordered_prep'):
                pn = j['pn']
                trio_batch_max[pn] = max(trio_batch_max.get(pn, 0), j.get('batch_round', 1))
        group_max = max(trio_batch_max.values())

        for pn in trio_set:
            if pn not in rules:
                continue
            cur_max = trio_batch_max.get(pn, 0)
            if cur_max >= group_max and cur_max > 0:
                continue
            r            = rules[pn]
            selected_qty = r['qty_options'][0] if r['qty_options'] else 0.0
            grid_dur     = (2 if r['is_prep_only'] else
                            math.ceil((selected_qty / max(1, r['num_ports']) / 1500 * 60 + 30) / 30))
            if cur_max == 0:
                template = {**r, 'id': f"TRIO_{pn}_B1", 'qty': selected_qty,
                            'grid_duration': grid_dur, 'prio': 3, 'batch_round': 1}
                jobs.append(template)
                print(f"[trio 補入] {r['name']}({pn}) 補入第1批（trio 組）")
                cur_max = 1
            else:
                template = next(
                    (j for j in jobs if j['pn'] == pn
                     and not j.get('forced') and not j.get('ordered_prep')), None)
            for extra_b in range(cur_max + 1, group_max + 1):
                if template:
                    jobs.append({**template, 'id': f"TRIO_{pn}_B{extra_b}",
                                 'batch_round': extra_b, 'prio': 3})
                    print(f"[trio 補批] {r['name']}({pn}) 補入第{extra_b}批（trio 組）")

    if forced_jobs:
        jobs = forced_jobs + jobs
    return jobs


# ─────────────────────────────────────────────────────────────────────────────
# _assign_job_vars
# ─────────────────────────────────────────────────────────────────────────────

def _assign_job_vars(model, j, reagent, base_s, base_day, base_st, job_dr,
                     p_grp_start, p_offset, sf_grp,
                     horizon_grids, max_gd,
                     time_intervals_2d, port_intervals_2d, job_vars):
    job_vars[j['id']] = {
        's': base_s, 'day': base_day, 'st': base_st, 'dr': job_dr,
        'port': p_grp_start, 'port_offset': p_offset, 'sf': sf_grp,
    }
    s_in_day = model.NewIntVar(0, DEFAULT_GRIDS, f"sid_{j['id']}")
    model.AddModuloEquality(s_in_day, base_s, DEFAULT_GRIDS)
    model.Add(s_in_day + j["grid_duration"] <= DEFAULT_GRIDS)
    model.AddLinearExpressionInDomain(s_in_day, cp_model.Domain.FromIntervals([(0, 3), (6, 34)]))

    if not j['is_prep_only']:
        f_g = math.ceil((j['duration_freeze'] + FREEZE_BUFFER_MINS) / MINS_PER_GRID)
        ef  = model.NewIntVar(0, horizon_grids + max_gd + f_g + 2, f"ef_{j['id']}")
        model.Add(ef == sf_grp + f_g)
        job_vars[j['id']].update({'ef': ef})

        if j.get('is_ivek'):
            return

        num_p         = max(1, j['num_ports'])
        port_start_iv = model.NewIntVar(0, 11, f"pstart_{j['id']}")
        model.Add(port_start_iv == p_grp_start + p_offset)
        time_iv = model.NewFixedSizeIntervalVar(base_s, j['grid_duration'] + PORT_GAP_GRIDS,
                                                f"tiv_{j['id']}")
        port_iv = model.NewFixedSizeIntervalVar(port_start_iv, num_p, f"portiv_{j['id']}")
        time_intervals_2d.append(time_iv)
        port_intervals_2d.append(port_iv)


# ─────────────────────────────────────────────────────────────────────────────
# run_solver
# [FIX-1] 加入 target_date 參數
# [FIX-3] 套用 forbidden_day_offsets 和 staffOffDays 約束
# [FIX-4] 效能調整
# ─────────────────────────────────────────────────────────────────────────────

def run_solver(jobs, reagent_groups, all_staff, horizon_days,
               resource_config=None, target_date=None):
    """
    [V33.31]
    resource_config 預期包含：
      forbidden_day_offsets : List[int]  ← generate_schedule 預處理後填入
      staff_off_offsets     : Dict[staffId, Set[int]]  ← generate_schedule 預處理後填入
    """
    model         = cp_model.CpModel()
    grids_per_day = DEFAULT_GRIDS
    horizon_grids = horizon_days * grids_per_day
    all_dryers    = sorted(list(set(
        d for j in jobs for d in j['compatible_dryers'] if not j['is_prep_only'])))

    rc = resource_config or {}

    job_vars             = {}
    staff_ints           = collections.defaultdict(list)
    dryer_exclusive_pool = collections.defaultdict(list)
    time_intervals_2d    = []
    port_intervals_2d    = []

    reagent_to_jobs = collections.defaultdict(list)
    for j in jobs:
        if not j.get('forced') and not j.get('ordered_prep'):
            group_key = f"{j['reagent']}__R{j.get('batch_round', 1)}"
            reagent_to_jobs[group_key].append(j)

    # ── 強制插單 ──────────────────────────────────────────────────────────────
    for j in jobs:
        if not j.get('forced') or j['is_prep_only']:
            continue
        pinned_g          = j['pinned_grid']
        pinned_p          = j['pinned_port']
        pinned_dryer_name = j['pinned_dryer']
        pinned_staff_name = j['pinned_staff']
        if pinned_dryer_name not in all_dryers:
            all_dryers.append(pinned_dryer_name)
        if pinned_staff_name not in all_staff:
            all_staff = sorted(list(set(all_staff) | {pinned_staff_name}))
        dr_idx   = all_dryers.index(pinned_dryer_name)
        st_idx   = all_staff.index(pinned_staff_name)
        base_s   = model.NewConstant(pinned_g)
        base_day = model.NewConstant(j['pinned_day'])
        base_st  = model.NewConstant(st_idx)
        base_dr  = model.NewConstant(dr_idx)
        print(f"  [強制排程] {j['name']} pinned_g={pinned_g} day={j['pinned_day']} "
              f"port={pinned_p} dryer={pinned_dryer_name} staff={pinned_staff_name}")
        sf = model.NewIntVar(0, horizon_grids + 400, f"sf_{j['id']}")
        ef = model.NewIntVar(0, horizon_grids + 500, f"ef_{j['id']}")
        model.Add(sf >= pinned_g + j['grid_duration'] + 1)
        f_g = math.ceil((j['duration_freeze'] + FREEZE_BUFFER_MINS) / 30)
        model.Add(ef == sf + f_g)
        _ef_mod = model.NewIntVar(0, 47, f"efmod_{j['id']}")
        model.AddModuloEquality(_ef_mod, ef, 48)
        model.AddLinearExpressionInDomain(_ef_mod, cp_model.Domain.FromIntervals([(0, 31), (42, 47)]))
        job_vars[j['id']] = {
            's': base_s, 'day': base_day, 'st': base_st, 'dr': base_dr,
            'port': model.NewConstant(pinned_p), 'port_offset': 0, 'sf': sf, 'ef': ef,
        }
        num_p         = max(1, j['num_ports'])
        port_start_iv = model.NewIntVar(pinned_p, pinned_p, f"pstart_{j['id']}")
        time_iv = model.NewFixedSizeIntervalVar(base_s, j['grid_duration'] + PORT_GAP_GRIDS,
                                                f"tiv_{j['id']}")
        port_iv = model.NewFixedSizeIntervalVar(port_start_iv, num_p, f"portiv_{j['id']}")
        time_intervals_2d.append(time_iv)
        port_intervals_2d.append(port_iv)
        lock_start = j['pinned_day'] * DEFAULT_GRIDS
        dryer_exclusive_pool[pinned_dryer_name].append(
            model.NewFixedSizeIntervalVar(model.NewConstant(lock_start), DEFAULT_GRIDS, ''))
        busy_s_val = max(0, pinned_g - 4)
        staff_ints[st_idx].append(
            model.NewFixedSizeIntervalVar(model.NewConstant(busy_s_val), 4, ''))

    # ── 接單配藥 ──────────────────────────────────────────────────────────────
    for j in jobs:
        if not j.get('ordered_prep'):
            continue
        pinned_day     = j.get('pinned_day', 0)
        staff_list     = j.get('compatible_staff', [])
        day_start_grid = pinned_day * DEFAULT_GRIDS
        for sname in staff_list:
            if sname not in all_staff:
                all_staff = sorted(list(set(all_staff) | {sname}))
            st_idx = all_staff.index(sname)
            staff_ints[st_idx].append(
                model.NewFixedSizeIntervalVar(
                    model.NewConstant(day_start_grid), DEFAULT_GRIDS,
                    f"ordered_{j['id']}_{sname}"))
        job_vars[j['id']] = {
            's':   model.NewConstant(day_start_grid),
            'day': model.NewConstant(pinned_day),
            'st':  model.NewConstant(all_staff.index(staff_list[0]) if staff_list else 0),
            'dr':  model.NewConstant(0),
            'port': model.NewConstant(0), 'port_offset': 0,
        }

    ivek_pn_round_day: dict = {}
    na_ivek_pn_days:   dict = {}
    reagent_dr_bools:  dict = {}
    dual_dryer_bools:  dict = {}
    reagent_base_day:  dict = {}
    reagent_base_st:   dict = {}
    reagent_base_dr:   dict = {}
    # [FIX-3-①-v2] 收集所有 reagent 的 base_day（含 dual dryer 路徑）
    all_reagent_base_days: dict = {}

    for reagent, group_jobs in reagent_to_jobs.items():
        base_s   = model.NewIntVar(0, horizon_grids, f"s_{reagent}")
        base_day = model.NewIntVar(0, horizon_days - 1, f"d_{reagent}")
        base_st  = model.NewIntVar(0, len(all_staff), f"st_{reagent}")
        model.AddDivisionEquality(base_day, base_s, grids_per_day)
        # [FIX-3-①-v2] 所有 reagent（含 dual dryer）都登記到 all_reagent_base_days
        all_reagent_base_days[reagent] = base_day

        is_prep_only_grp = all(j['is_prep_only'] for j in group_jobs)
        staff_pool = set()
        for j in group_jobs:
            staff_pool.update(j['compatible_staff'])

        st_bools = []
        for idx, sname in enumerate(all_staff):
            if sname in staff_pool:
                u = model.NewBoolVar(f"u_st_{reagent}_{idx}")
                st_bools.append(u)
                model.Add(base_st == idx).OnlyEnforceIf(u)
                if is_prep_only_grp:
                    staff_ints[idx].append(
                        model.NewOptionalFixedSizeIntervalVar(base_s, 6, u, ''))
                else:
                    busy_s = model.NewIntVar(-4, horizon_grids, '')
                    model.Add(busy_s == base_s - 4)
                    staff_ints[idx].append(
                        model.NewOptionalFixedSizeIntervalVar(busy_s, 4, u, ''))
        if st_bools:
            model.Add(sum(st_bools) == 1)

        non_prep_in_group = [j for j in group_jobs if not j['is_prep_only']]
        pns_in_grp        = {j['pn'] for j in non_prep_in_group}

        for j in group_jobs:
            if j.get('is_ivek') and j.get('pn') in NA_IVEK_PNS:
                na_ivek_pn_days.setdefault(j['pn'], [])
                if base_day not in na_ivek_pn_days[j['pn']]:
                    na_ivek_pn_days[j['pn']].append(base_day)
                break

        for j in group_jobs:
            if j.get('is_ivek') and not j.get('is_prep_only'):
                _key = (j['pn'], j.get('batch_round', 1))
                if _key not in ivek_pn_round_day:
                    ivek_pn_round_day[_key] = base_day
                break

        dual_pair     = next(((au, ad) for au, ad in _DUAL_PAIRS
                              if au in pns_in_grp or ad in pns_in_grp), None)
        is_dual_dryer = dual_pair is not None

        _non_prep_grp = [j for j in group_jobs if not j['is_prep_only']]
        _max_gd_grp   = max((j['grid_duration'] for j in _non_prep_grp), default=1)
        sf_grp = model.NewIntVar(0, horizon_grids + _max_gd_grp + 2, f"sf_grp_{reagent}")
        model.Add(sf_grp == base_s + _max_gd_grp + 1)

        if is_dual_dryer:
            au_pn, ad_pn = dual_pair
            base_dr_au   = model.NewIntVar(0, len(all_dryers) - 1, f"dr_au_{reagent}")
            base_dr_ad   = model.NewIntVar(0, len(all_dryers) - 1, f"dr_ad_{reagent}")
            model.Add(base_dr_au != base_dr_ad)

            for tag, pn_x, base_dr_x in [('au', au_pn, base_dr_au), ('ad', ad_pn, base_dr_ad)]:
                compat_d = set()
                for j in non_prep_in_group:
                    if j['pn'] == pn_x:
                        compat_d.update(j['compatible_dryers'])
                dr_bx = []
                for idx, dname in enumerate(all_dryers):
                    if dname in compat_d:
                        ud = model.NewBoolVar(f"ud_dr_{reagent}_{tag}_{idx}")
                        dr_bx.append(ud)
                        model.Add(base_dr_x == idx).OnlyEnforceIf(ud)
                        lock_start = base_day * grids_per_day
                        dryer_exclusive_pool[dname].append(
                            model.NewOptionalFixedSizeIntervalVar(
                                lock_start, grids_per_day, ud, ''))
                        dual_dryer_bools[(pn_x, dname)] = ud
                if dr_bx:
                    model.Add(sum(dr_bx) == 1)

            base_dr    = base_dr_au
            au_jobs    = [j for j in group_jobs if j['pn'] == au_pn]
            ad_jobs    = [j for j in group_jobs if j['pn'] == ad_pn]
            other_jobs = [j for j in group_jobs if j['pn'] not in (au_pn, ad_pn)]

            n_ports_au = sum(max(1, j['num_ports']) for j in au_jobs if not j['is_prep_only'])
            n_ports_ad = sum(max(1, j['num_ports']) for j in ad_jobs if not j['is_prep_only'])
            p_start_u  = model.NewIntVar(0, 11, f"psu_{reagent}")
            model.AddModuloEquality(0, p_start_u, 2)
            p_start_d  = model.NewIntVar(0, 11, f"psd_{reagent}")
            model.Add(p_start_d == p_start_u + n_ports_au + DUAL_PORT_GAP)
            model.Add(p_start_u + n_ports_au + DUAL_PORT_GAP + n_ports_ad <= 12)

            au_off = 0
            for j in au_jobs:
                _assign_job_vars(model, j, reagent, base_s, base_day, base_st, base_dr_au,
                                 p_start_u, au_off, sf_grp, horizon_grids, _max_gd_grp,
                                 time_intervals_2d, port_intervals_2d, job_vars)
                if not j['is_prep_only']:
                    au_off += max(1, j['num_ports'])
            ad_off = 0
            for j in ad_jobs:
                _assign_job_vars(model, j, reagent, base_s, base_day, base_st, base_dr_ad,
                                 p_start_d, ad_off, sf_grp, horizon_grids, _max_gd_grp,
                                 time_intervals_2d, port_intervals_2d, job_vars)
                if not j['is_prep_only']:
                    ad_off += max(1, j['num_ports'])
            for j in other_jobs:
                job_vars[j['id']] = {
                    's': base_s, 'day': base_day, 'st': base_st, 'dr': base_dr,
                    'port': p_start_u, 'port_offset': 0, 'sf': sf_grp,
                }

        else:
            base_dr  = model.NewIntVar(0, len(all_dryers) - 1, f"dr_{reagent}")
            dr_bools = []
            for idx, dname in enumerate(all_dryers):
                compat = (all(dname in j['compatible_dryers'] for j in non_prep_in_group)
                          if non_prep_in_group
                          else any(dname in j['compatible_dryers'] for j in group_jobs))
                if compat:
                    ud = model.NewBoolVar(f"ud_dr_{reagent}_{idx}")
                    dr_bools.append(ud)
                    model.Add(base_dr == idx).OnlyEnforceIf(ud)
                    lock_start = base_day * grids_per_day
                    dryer_exclusive_pool[dname].append(
                        model.NewOptionalFixedSizeIntervalVar(
                            lock_start, grids_per_day, ud, ''))
                    reagent_dr_bools[(reagent, dname)] = ud
            if dr_bools:
                model.Add(sum(dr_bools) == 1)

            reagent_base_day[reagent] = base_day
            reagent_base_st[reagent]  = base_st
            reagent_base_dr[reagent]  = base_dr

            p_grp_start = model.NewIntVar(0, 11, f"p_start_{reagent}")
            model.AddModuloEquality(0, p_grp_start, 2)

            def _ud_key(j):
                n = j.get('name', '')
                return 0 if (n.endswith('-U') or n.endswith('-AU') or n.endswith('-BU')) else 1
            ordered_grp = sorted(group_jobs, key=_ud_key)

            cur_off = 0
            for j in ordered_grp:
                _assign_job_vars(model, j, reagent, base_s, base_day, base_st, base_dr,
                                 p_grp_start, cur_off, sf_grp, horizon_grids, _max_gd_grp,
                                 time_intervals_2d, port_intervals_2d, job_vars)
                if not j['is_prep_only']:
                    cur_off += max(1, j['num_ports'])
            model.Add(p_grp_start + cur_off <= 12)

    # ── 汙染約束 ──────────────────────────────────────────────────────────────
    dual_job_infos = [
        (j, job_vars[j['id']]) for j in jobs
        if j.get('pn') in _DUAL_ALL_PNS
        and not j.get('forced') and not j.get('ordered_prep')
        and j['id'] in job_vars and not j.get('is_prep_only')
    ]
    contam_job_infos = [
        (j, job_vars[j['id']]) for j in jobs
        if j.get('pn') in DUAL_CONTAM_PNS
        and not j.get('forced') and not j.get('ordered_prep')
        and j['id'] in job_vars and not j.get('is_prep_only')
    ]
    for dual_j, dual_v in dual_job_infos:
        g_day     = dual_v['day']
        g_pn      = dual_j['pn']
        compat_cd = set(dual_j.get('compatible_dryers', [])) & DUAL_CONTAM_DRYERS
        for contam_j, contam_v in contam_job_infos:
            c_day     = contam_v['day']
            contam_rk = f"{contam_j['reagent']}__R{contam_j.get('batch_round', 1)}"
            shared_cd = compat_cd & (set(contam_j.get('compatible_dryers', [])) & DUAL_CONTAM_DRYERS)
            if not shared_cd:
                continue
            for dname in shared_cd:
                b_g = dual_dryer_bools.get((g_pn, dname))
                b_c = reagent_dr_bools.get((contam_rk, dname))
                if b_g is None or b_c is None:
                    continue
                b_ok_after = model.NewBoolVar(f"boka_{dual_j['id']}_{contam_j['id']}_{dname}")
                b_ok_far   = model.NewBoolVar(f"bokf_{dual_j['id']}_{contam_j['id']}_{dname}")
                model.Add(c_day >= g_day).OnlyEnforceIf(b_ok_after)
                model.Add(g_day - c_day >= DUAL_BLOCK_DAYS + 1).OnlyEnforceIf(b_ok_far)
                model.AddBoolOr([b_g.Not(), b_c.Not(), b_ok_after, b_ok_far])

    # ── Na-IVEK 直接配對約束 ──────────────────────────────────────────────────
    for _u_pn, _d_pn in _NA_IVEK_UD_PAIRS:
        _rounds = set(r for (pn, r) in ivek_pn_round_day if pn in (_u_pn, _d_pn))
        for _r in sorted(_rounds):
            _dv_u = ivek_pn_round_day.get((_u_pn, _r))
            _dv_d = ivek_pn_round_day.get((_d_pn, _r))
            if _dv_u is not None and _dv_d is not None:
                model.Add(_dv_d - _dv_u == 1)
                print(f"[IVEK UD/day] ({_u_pn}↔{_d_pn}) R{_r}: d_D = d_U + 1")
            elif _dv_u is None and _dv_d is not None:
                model.Add(_dv_d >= 1)

    # ── 全局 IVEK 每天最多一批 ────────────────────────────────────────────────
    _all_ivek_days = list(ivek_pn_round_day.values())
    if len(_all_ivek_days) > 1:
        model.AddAllDifferent(_all_ivek_days)
        print(f"[IVEK/day] {len(_all_ivek_days)} 批 IVEK 各排不同天")

    # ── tCREA/QCREA 三合一約束 ────────────────────────────────────────────────
    for _paired_pns, _solo_pn, _paired_base, _solo_base in _TRIO_GROUPS:
        _trio_rounds = set()
        for _key in reagent_base_day:
            _kb = re.sub(r'__R\d+$', '', _key)
            if _kb in (_paired_base, _solo_base):
                _m = re.search(r'__R(\d+)$', _key)
                if _m:
                    _trio_rounds.add(int(_m.group(1)))
        for _r in sorted(_trio_rounds):
            _pk = f"{_paired_base}__R{_r}"
            _sk = f"{_solo_base}__R{_r}"
            if _pk not in reagent_base_day or _sk not in reagent_base_day:
                continue
            _pd_day = reagent_base_day[_pk]
            _sd_day = reagent_base_day[_sk]
            _pd_st  = reagent_base_st[_pk]
            _sd_st  = reagent_base_st[_sk]
            _pd_dr  = reagent_base_dr.get(_pk)
            _sd_dr  = reagent_base_dr.get(_sk)
            model.Add(_pd_day == _sd_day)
            model.Add(_pd_st  == _sd_st)
            if _pd_dr is not None and _sd_dr is not None:
                _paired_jobs = [j for j in jobs if j['pn'] in _paired_pns and not j.get('forced')]
                _solo_jobs   = [j for j in jobs if j['pn'] == _solo_pn    and not j.get('forced')]
                _pd_dryers   = set(d for j in _paired_jobs for d in j.get('compatible_dryers', []))
                _sd_dryers   = set(d for j in _solo_jobs   for d in j.get('compatible_dryers', []))
                _union_dryers = _pd_dryers | _sd_dryers
                if len(_union_dryers) >= 2:
                    model.Add(_pd_dr != _sd_dr)
                    print(f"[trio] {_paired_base}+{_solo_base} R{_r}: 同天同人不同機 (可用:{sorted(_union_dryers)})")
                else:
                    print(f"[trio] {_paired_base}+{_solo_base} R{_r}: ⚠️ 凍乾機只有{sorted(_union_dryers)}，略過不同機約束")

    # ── Na-IVEK 每 PN 多批不同天 ──────────────────────────────────────────────
    for pn, day_vars in na_ivek_pn_days.items():
        if len(day_vars) > 1:
            model.AddAllDifferent(day_vars)
            print(f"[Na-IVEK/day] {pn}: {len(day_vars)} 批各排不同天")

    if time_intervals_2d:
        model.AddNoOverlap2D(time_intervals_2d, port_intervals_2d)
    for idx in staff_ints:
        model.AddNoOverlap(staff_ints[idx])
    for dname in dryer_exclusive_pool:
        model.AddNoOverlap(dryer_exclusive_pool[dname])

    # ══════════════════════════════════════════════════════════════════════════
    # [FIX-3-①] Holiday forbidden days
    # forbidden_day_offsets 由 generate_schedule 預處理填入 rc
    # ══════════════════════════════════════════════════════════════════════════
    forbidden_offsets = set(rc.get('forbidden_day_offsets', []))
    if forbidden_offsets:
        print(f"[Solver] 套用 forbidden day offsets: {sorted(forbidden_offsets)}")
        for reagent, b_day in all_reagent_base_days.items():  # [FIX-v2] 含 dual dryer
            for fd in forbidden_offsets:
                if fd < horizon_days:
                    model.Add(b_day != fd)
        # 警告：forced / ordered_prep job 若釘在休假日無法修改
        for j in jobs:
            if j.get('forced') or j.get('ordered_prep'):
                pd = j.get('pinned_day', j.get('day', 0))
                if pd in forbidden_offsets:
                    print(f"  ⚠️  [警告] {j['name']} 固定排在休假日 day={pd}，無法移動")

    # ══════════════════════════════════════════════════════════════════════════
    # [FIX-3-④] Staff off days
    # staff_off_offsets: Dict[staffId, Set[int]] 由 generate_schedule 預處理填入 rc
    # ══════════════════════════════════════════════════════════════════════════
    staff_off_offsets = rc.get('staff_off_offsets', {})
    if staff_off_offsets:
        for staff_id, off_offsets in staff_off_offsets.items():
            name_fragment = _STAFF_ID_MAP.get(staff_id, "")
            if not name_fragment:
                continue
            # 找出 all_staff 中對應的 index（partial match，不區分大小寫）
            matched_idxs = [
                i for i, s in enumerate(all_staff)
                if name_fragment.lower() in s.lower()
            ]
            if not matched_idxs:
                print(f"  ⚠️  [StaffOff] {staff_id}({name_fragment}) 未在 all_staff 中找到，略過")
                continue
            staff_idx = matched_idxs[0]
            valid_offs = {fd for fd in off_offsets if fd < horizon_days}
            if not valid_offs:
                continue
            print(f"  [StaffOff] {staff_id}({name_fragment}) idx={staff_idx} "
                  f"不可排 day={sorted(valid_offs)}")
            for reagent, b_day in all_reagent_base_days.items():  # [FIX-v2] 含 dual dryer
                b_st = reagent_base_st.get(reagent)
                if b_st is None:
                    continue
                for fd in valid_offs:
                    # 禁止「這個 reagent 排在休假 day AND 被分配給這位人員」同時成立
                    both_match = model.NewBoolVar(
                        f"soff_{staff_id}_{reagent}_{fd}")
                    model.Add(b_st == staff_idx).OnlyEnforceIf(both_match)
                    model.Add(b_day == fd).OnlyEnforceIf(both_match)
                    # both_match == True 代表同時成立 → 禁止
                    model.AddBoolOr([both_match.Not()])

    # ══════════════════════════════════════════════════════════════════════════
    model.Minimize(sum(v['day'] * 2000 + v['s'] for v in job_vars.values()))

    solver = cp_model.CpSolver()
    # [FIX-4] 效能：60s → 30s，開多核並行
    solver.parameters.max_time_in_seconds = 30.0
    solver.parameters.num_search_workers  = 4

    if solver.Solve(model) in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
        res = []
        for j in jobs:
            v = job_vars[j['id']]
            def val(var):
                try:
                    return solver.Value(var)
                except Exception:
                    return int(str(var))
            out = {
                **j,
                'day':            val(v['day']),
                'grid_start':     val(v['s']) % DEFAULT_GRIDS,
                'assigned_staff': all_staff[val(v['st'])],
            }
            if not j['is_prep_only']:
                out.update({
                    'p_idx':          val(v['port']) + v['port_offset'],
                    'assigned_dryer': all_dryers[val(v['dr'])],
                    'ef_abs':         val(v['ef']),
                    'sf_abs':         val(v['sf']),
                })
            res.append(out)

        group_max_f_g: dict = {}
        for r in res:
            if r.get('is_prep_only') or 'sf_abs' not in r:
                continue
            gk      = f"{r['reagent']}__R{r.get('batch_round', 1)}"
            f_g_val = math.ceil((r.get('duration_freeze', 720) + FREEZE_BUFFER_MINS) / MINS_PER_GRID)
            group_max_f_g[gk] = max(group_max_f_g.get(gk, 0), f_g_val)
        for r in res:
            if r.get('is_prep_only') or 'sf_abs' not in r:
                continue
            gk = f"{r['reagent']}__R{r.get('batch_round', 1)}"
            own_f_g = math.ceil((r.get('duration_freeze', 720) + FREEZE_BUFFER_MINS) / MINS_PER_GRID)
            r['dryer_avail_f_g'] = group_max_f_g.get(gk, own_f_g)

        return res
    return None


# ─────────────────────────────────────────────────────────────────────────────
# results_to_json
# ─────────────────────────────────────────────────────────────────────────────

def results_to_json(results, start_date=None, batch_num_start=1):
    base_dt = datetime.strptime(start_date, "%Y-%m-%d") if start_date else None
    if base_dt:
        iso_week     = base_dt.isocalendar()[1]
        year_2       = base_dt.strftime("%y")
        month_letter = chr(ord('A') + base_dt.month - 1)
        week_2       = f"{iso_week:02d}"
    else:
        iso_week, year_2, month_letter, week_2 = 1, "00", "A", "01"

    pn_occurrence: dict = {}
    batch_counter = batch_num_start
    final_data    = []

    # 先按日期+時間排序，確保同 PN 的 lot 序號按日期遞增
    results = sorted(results, key=lambda r: (r['day'], r['grid_start']))

    for r in results:
        day_offset = r['day']
        grid       = r['grid_start']
        rd_mins    = START_HOUR * 60 + grid * MINS_PER_GRID
        rd_dt      = datetime(2000, 1, 1) + timedelta(minutes=rd_mins)
        rd_str     = rd_dt.strftime("%H:%M")
        tit_str    = (rd_dt + timedelta(minutes=30)).strftime("%H:%M")
        date_str   = ((base_dt + timedelta(days=day_offset)).strftime("%m/%d")
                      if base_dt else f"D{day_offset + 1}")

        if r['is_prep_only']:
            finish_str = "—"
            port_str   = "純配藥"
            dryer_str  = "—"
            if not r.get('ordered_prep'):
                rd_str  = "16:00"
                tit_str = "—"
        else:
            sf_abs          = r.get('sf_abs', 0)
            sf_day          = sf_abs // DEFAULT_GRIDS
            sf_grid         = sf_abs % DEFAULT_GRIDS
            dryer_avail_f_g = r.get('dryer_avail_f_g',
                math.ceil((r.get('duration_freeze', 720) + FREEZE_BUFFER_MINS) / MINS_PER_GRID))
            if base_dt:
                sf_dt = base_dt + timedelta(days=sf_day, hours=START_HOUR,
                                            minutes=sf_grid * MINS_PER_GRID)
                ef_dt = sf_dt + timedelta(minutes=dryer_avail_f_g * MINS_PER_GRID)
                if 2 <= ef_dt.hour < 7:
                    ef_dt = ef_dt.replace(hour=7, minute=0, second=0)
                finish_str = ef_dt.strftime("%m/%d %H:%M")
            else:
                sf_mins    = START_HOUR * 60 + sf_grid * MINS_PER_GRID
                ef_mins    = sf_mins + dryer_avail_f_g * MINS_PER_GRID
                finish_str = (datetime(2000, 1, 1) + timedelta(minutes=ef_mins)).strftime("%H:%M")

            port_str  = "IVEK" if r.get('is_ivek') else str(r['p_idx'] + 1)
            dryer_str = r.get('assigned_dryer', '—')

        work_order = f"TMRA{week_2}{month_letter}{batch_counter:03d}"
        batch_counter += 1

        pn   = r['pn']
        pn_3 = pn[-3:] if len(pn) >= 3 else pn
        num_ports_actual = r['num_ports'] if not r['is_prep_only'] else 0

        if r.get('is_ivek') and not r['is_prep_only']:
            qty_each = int(r['qty']) // 2
            pn_occurrence[pn] = pn_occurrence.get(pn, 0) + 1
            lot1 = f"{pn_3}{year_2}{week_2}{pn_occurrence[pn]}"
            pn_occurrence[pn] += 1
            lot2 = f"{pn_3}{year_2}{week_2}{pn_occurrence[pn]}"
            base_row = {
                "day": r['day'] + 1, "date_str": date_str,
                "rd_time": rd_str, "start_time": tit_str, "finish_time": finish_str,
                "port": "IVEK", "num_ports": 0,
                "name": r['name'], "pn": pn,
                "qty": qty_each, "qty_total": int(r['qty']),
                "staff": r['assigned_staff'], "dryer": dryer_str,
                "is_prep_only": False, "ordered_prep": False,
                "holiday_conflict": r.get('holiday_conflict', False),
                "demand_date": r.get('demand_date', ''),
                "work_order": work_order, "lots": [lot1, lot2],
            }
            final_data.append({**base_row, "ivek_slot": 1, "lot": lot1})
            final_data.append({**base_row, "ivek_slot": 2, "lot": lot2})
            continue

        num_p = max(1, r['num_ports']) if not r['is_prep_only'] else 1
        lots  = []
        for _ in range(num_p):
            pn_occurrence[pn] = pn_occurrence.get(pn, 0) + 1
            lots.append(f"{pn_3}{year_2}{week_2}{pn_occurrence[pn]}")

        qty_per_port = (int(r['qty']) // num_ports_actual if num_ports_actual > 0
                        else int(r['qty']))
        final_data.append({
            "day": r['day'] + 1, "date_str": date_str,
            "rd_time": rd_str, "start_time": tit_str, "finish_time": finish_str,
            "port": port_str, "num_ports": r['num_ports'],
            "name": r['name'], "pn": pn,
            "qty": qty_per_port, "qty_total": int(r['qty']),
            "staff": r['assigned_staff'], "dryer": dryer_str,
            "is_prep_only": r['is_prep_only'],
            "ordered_prep": r.get('ordered_prep', False),
            "holiday_conflict": r.get('holiday_conflict', False),
            "demand_date": r.get('demand_date', ''),
            "work_order": work_order,
            "lot": lots[0], "lots": lots,
        })

    return sorted(final_data, key=lambda x: (x['day'], x['rd_time']))

# -*- coding: utf-8 -*-
# ─────────────────────────────────────────────────────────────────────────────
# export_excel  [V33.32]
# output_path=None → 回傳 BytesIO（Flask send_file 用）
# output_path=str  → 寫檔（向下相容）
# ─────────────────────────────────────────────────────────────────────────────

def export_excel(results, output_path=None, start_date=None):
    import io
    base_dt = datetime.strptime(start_date, "%Y-%m-%d") if start_date else None
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    by_day = collections.defaultdict(list)
    for j in results:
        by_day[j['day']].append(j)

    border    = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
    fill_h    = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    fill_prep = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
    colors    = ["E6B8AF","F4CCCC","FCE5CD","FFF2CC","D9EAD3",
                 "D0E0E3","C9DAF8","CFE2F3","D9D2E9","EAD1DC"]
    pn_colors = {}

    sheet_seq = 0
    for d in sorted(by_day.keys()):
        sheet_seq += 1
        sheet_label = (base_dt + timedelta(days=d)).strftime("%m-%d") if base_dt else f"{d+1}"
        ws = wb.create_sheet(f"Day {sheet_seq} ({sheet_label})")
        headers = ["時間", "IVEK 1", "IVEK 2"] + [f"Port {i}" for i in range(1, 13)] + ["純配藥區"]
        for c, txt in enumerate(headers, 1):
            cell = ws.cell(1, c, txt)
            cell.fill, cell.border, cell.alignment = fill_h, border, Alignment('center', 'center')
            ws.column_dimensions[get_column_letter(c)].width = 17
        for i in range(DEFAULT_GRIDS):
            m   = START_HOUR * 60 + i * MINS_PER_GRID
            txt = (f"{m//60:02d}:{m%60:02d}" if m < 1440
                   else f"隔{(m-1440)//60:02d}:{(m-1440)%60:02d}")
            ws.cell(i + 2, 1, txt).border = border

        p_idx_prep = 0
        for j in by_day[d]:
            if j['pn'] not in pn_colors:
                pn_colors[j['pn']] = colors[len(pn_colors) % len(colors)]
            c_fill = PatternFill(start_color=pn_colors[j['pn']],
                                 end_color=pn_colors[j['pn']], fill_type="solid")

            if j['is_prep_only']:
                c_s  = len(headers)
                r_s  = PREP_ONLY_ROW_START + p_idx_prep
                p_idx_prep += 1
                cell = ws.cell(r_s, c_s, f"{j['pn']}\n{j['name']}\n人:{j['assigned_staff']}")
                cell.fill, cell.border = fill_prep, border
                cell.alignment = Alignment('center', 'center', wrap_text=True)
            else:
                r_s = j['grid_start'] + 2
                r_e = j['grid_start'] + 2 + j['grid_duration'] - 1

                sf_abs          = j.get('sf_abs', 0)
                sf_day          = sf_abs // DEFAULT_GRIDS
                sf_grid         = sf_abs % DEFAULT_GRIDS
                dryer_avail_f_g = j.get('dryer_avail_f_g',
                    math.ceil((j.get('duration_freeze', 720) + FREEZE_BUFFER_MINS) / MINS_PER_GRID))
                if base_dt:
                    sf_dt   = base_dt + timedelta(days=sf_day, hours=START_HOUR,
                                                   minutes=sf_grid * MINS_PER_GRID)
                    ef_dt   = sf_dt + timedelta(minutes=dryer_avail_f_g * MINS_PER_GRID)
                    if 2 <= ef_dt.hour < 7:
                        ef_dt = ef_dt.replace(hour=7, minute=0, second=0)
                    end_str = ef_dt.strftime("%m/%d %H:%M")
                else:
                    sf_mins = START_HOUR * 60 + sf_grid * MINS_PER_GRID
                    ef_mins = sf_mins + dryer_avail_f_g * MINS_PER_GRID
                    end_str = f"{(ef_mins // 60) % 24:02d}:{ef_mins % 60:02d}"

                if j.get('is_ivek'):
                    qty_each = (int(j['qty_total']) if 'qty_total' in j else int(j['qty'])) // 2
                    txt = (f"{j['pn']}\n{j['name']}\n每IVEK:{qty_each}\n"
                           f"人:{j['assigned_staff']}\n機:{j['assigned_dryer']}\nDryer可用 {end_str}")
                    for col_ivek in (2, 3):
                        if r_e > r_s:
                            try:
                                ws.unmerge_cells(start_row=r_s, start_column=col_ivek,
                                                 end_row=r_e, end_column=col_ivek)
                            except Exception:
                                pass
                        try:
                            cell       = ws.cell(r_s, col_ivek)
                            cell.value = txt
                        except AttributeError:
                            pass
                        else:
                            cell.fill      = c_fill
                            cell.border    = border
                            cell.alignment = Alignment('center', 'center', wrap_text=True)
                        if r_e > r_s:
                            try:
                                ws.merge_cells(start_row=r_s, start_column=col_ivek,
                                               end_row=r_e, end_column=col_ivek)
                            except Exception:
                                pass
                        for r in range(r_s, r_e + 1):
                            ws.cell(r, col_ivek).border = border
                else:
                    c_s = j['p_idx'] + 4
                    c_e = j['p_idx'] + 4 + max(1, j['num_ports']) - 1
                    num_p        = max(1, j['num_ports'])
                    per_port_qty = (int(j['qty_total']) // num_p if 'qty_total' in j
                                    else int(j['qty']) // num_p)
                    txt = (f"{j['pn']}\n{j['name']}\n每port:{per_port_qty}\n"
                           f"人:{j['assigned_staff']}\n機:{j['assigned_dryer']}\nDryer可用 {end_str}")
                    cell = ws.cell(r_s, c_s, txt)
                    cell.fill, cell.border = c_fill, border
                    cell.alignment = Alignment('center', 'center', wrap_text=True)
                    if r_e > r_s or c_e > c_s:
                        try:
                            ws.merge_cells(start_row=r_s, start_column=c_s,
                                           end_row=r_e, end_column=c_e)
                        except Exception:
                            pass
                    for r in range(r_s, r_e + 1):
                        for c in range(c_s, c_e + 1):
                            ws.cell(r, c).border = border

    if output_path:
        wb.save(output_path)
        print(f"Excel 已輸出：{output_path}")
        return output_path
    else:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf


# ─────────────────────────────────────────────────────────────────────────────
# _filter_jobs_by_priority  [V33.32 新增]
# ─────────────────────────────────────────────────────────────────────────────

def _filter_jobs_by_priority(jobs, max_prio):
    """
    max_prio=3 → 全保留（W1+W2+W3）
    max_prio=2 → 放棄 W3（保留 W1+W2）
    max_prio=1 → 放棄 W2+W3（僅 W1）

    以下永遠保留：
      · forced（強制插單）
      · ordered_prep（接單配藥）
      · prio==99（PAIR 強制補入配對）
      · prio==0（forced 子類）
    """
    kept, dropped_names = [], []
    for j in jobs:
        p = j.get('prio', 1)
        if j.get('forced') or j.get('ordered_prep') or p == 99 or p == 0:
            kept.append(j)
        elif p <= max_prio:
            kept.append(j)
        else:
            dropped_names.append(j.get('name', j.get('pn', '?')))
    if dropped_names:
        sample = list(dict.fromkeys(dropped_names))[:5]  # unique, preserve order
        print(f"  [降級] 放棄 {len(dropped_names)} 個 job (prio>{max_prio})，"
              f"例：{sample}")
    return kept


# ─────────────────────────────────────────────────────────────────────────────
# generate_schedule  [V33.32]
#
# Horizon：target_date 所在週的工作日（到週日為止的 calendar days）
#   e.g. target=週一 → horizon=7（Mon~Sun）
#        target=週三 → horizon=5（Wed~Sun）
#        中間 holidays 由 forbidden_day_offsets 讓 solver 跳過，
#        horizon 本身不縮短（維持完整週曆範圍）
#
# 三段降級重試（同一 horizon，solver 各跑一次）：
#   ① 全需求 W1+W2+W3  → run_solver
#   ② 找不到解 → 放棄 W3  → run_solver
#   ③ 找不到解 → 放棄 W2+W3（僅 W1）→ run_solver
#   全部失敗 → return error
#
# export_excel 回傳 BytesIO，不寫 EC2 硬碟
# ─────────────────────────────────────────────────────────────────────────────

def generate_schedule(db, target_date, resource_config=None,
                      selected_pns=None, batch_num_start=1):
    try:
        result = load_db_data(db, target_date)
        if result[0] is None:
            return {"status": "error", "message": "RDS 資料讀取失敗或規則表為空"}

        rules, groups, staff, df_demand, df_forced, df_plan = result

        if df_demand is None or df_demand.empty:
            return {"status": "error", "message": "找不到 BeadNeed 需求資料"}

        forced_jobs = load_forced_jobs(df_forced, rules, target_date)

        # ── rc：標準化 resource_config ────────────────────────────────────────
        rc = dict(resource_config) if resource_config else {}
        _holidays = rc.get('holidays', [])

        # ── Horizon：target_date 到本週日（含）的 calendar days ──────────────
        # 六日預設為 holidays（BeadResource 預設勾選），但加班時使用者可取消勾選
        # → horizon 永遠跑到週日，實際能不能排由 forbidden_day_offsets 決定
        start_dt     = datetime.strptime(target_date, "%Y-%m-%d")
        start_wd     = start_dt.weekday()            # Mon=0 … Sun=6
        days_to_sun  = 6 - start_wd                  # 距本週日天數
        horizon_days = days_to_sun + 1               # +1 包含 target 當天
        end_dt       = start_dt + timedelta(days=horizon_days - 1)
        wd_names = ['一','二','三','四','五','六','日']
        print(f"[排程] {target_date}（{wd_names[start_wd]}）"
              f" ~ {end_dt.strftime('%m/%d')}（{wd_names[end_dt.weekday()]}），"
              f"horizon={horizon_days} 天")

        # ── forbidden_day_offsets：完全依照 BeadResource 傳來的 holidays ──────
        # 六日預設在前端已勾選為休假，加班時使用者取消勾選 → 不在 holidays 內 → 可排
        forbidden_day_offsets = _compute_forbidden_day_offsets(
            _holidays, target_date, horizon_days=horizon_days)
        rc['forbidden_day_offsets'] = sorted(forbidden_day_offsets)
        print(f"[Resource] forbidden offsets: {sorted(forbidden_day_offsets)} "
              f"(holidays={_holidays})")

        # ── staff_off_offsets ─────────────────────────────────────────────────
        staff_off_offsets = _compute_staff_off_offsets(
            rc.get('staffOffDays', {}), target_date, horizon_days=horizon_days)
        rc['staff_off_offsets'] = staff_off_offsets
        if staff_off_offsets:
            print(f"[Resource] Staff off: "
                  f"{ {k: sorted(v) for k, v in staff_off_offsets.items()} }")

        # ── 設備維修 ──────────────────────────────────────────────────────────
        dryer_maint      = set(str(x) for x in rc.get('dryerMaintenance', []))
        ivek_maint_ports = set(str(x) for x in rc.get('ivekMaintenance', []))
        ivek_global      = bool(rc.get('ivekGlobalMaintenance', False))
        if dryer_maint:
            print(f"[Resource] 凍乾機維修中: {dryer_maint}")
        if ivek_global:
            print(f"[Resource] IVEK 全機維修")
        elif ivek_maint_ports:
            print(f"[Resource] IVEK port 維修: {ivek_maint_ports}")

        # ── build jobs ────────────────────────────────────────────────────────
        ordered_jobs = build_ordered_prep_jobs(
            rules, df_demand, df_plan, target_date,
            holidays=_holidays, max_horizon_days=horizon_days)

        jobs = build_jobs(rules, df_demand, df_plan=None,
                          selected_pns=selected_pns, forced_jobs=forced_jobs)

        if ordered_jobs:
            jobs = ordered_jobs + [j for j in jobs if not j.get('ordered_prep')]

        if not jobs:
            return {"status": "error", "message": "篩選後無有效待生產任務"}

        # ── 設備過濾 ──────────────────────────────────────────────────────────
        filtered_jobs = []
        skipped_count = 0
        for j in jobs:
            if j.get('is_ivek') and not j.get('is_prep_only') and ivek_global:
                print(f"  [略過-IVEK全修] {j['name']}({j['pn']})")
                skipped_count += 1
                continue
            if not j.get('is_prep_only') and not j.get('is_ivek') and dryer_maint:
                original  = j.get('compatible_dryers', [])
                available = [d for d in original if d not in dryer_maint]
                if not available:
                    print(f"  [略過-無可用凍乾機] {j['name']}({j['pn']}) "
                          f"原:{original} 維修:{dryer_maint}")
                    skipped_count += 1
                    continue
                if set(available) != set(original):
                    print(f"  [縮減凍乾機] {j['name']}({j['pn']}) {original}→{available}")
                j = {**j, 'compatible_dryers': available}
            filtered_jobs.append(j)

        jobs = filtered_jobs
        if skipped_count:
            print(f"[Resource] 共略過 {skipped_count} 個 job（設備維修中）")
        if not jobs:
            return {"status": "error",
                    "message": "所有任務因設備維修而無法排程，請確認維修設定"}

        # ══════════════════════════════════════════════════════════════════════
        # 三段降級重試（同一 horizon，每段跑一次 solver）
        # ══════════════════════════════════════════════════════════════════════
        attempts = [
            (3, "全需求 W1+W2+W3"),
            (2, "放棄 W3，保留 W1+W2"),
            (1, "放棄 W2+W3，僅保留 W1"),
        ]
        res          = None
        used_label   = ""

        for max_prio, label in attempts:
            attempt_jobs = _filter_jobs_by_priority(jobs, max_prio)
            if not attempt_jobs:
                print(f"[排程嘗試] {label} → 無任何 job，跳過")
                continue

            print(f"\n[排程嘗試] {label}，jobs={len(attempt_jobs)}，"
                  f"horizon={horizon_days}天...")
            res = run_solver(
                attempt_jobs, groups, staff, horizon_days,
                resource_config=rc,
                target_date=target_date,
            )
            if res:
                used_label = label
                print(f"[排程成功] {label}")
                break
            else:
                print(f"[排程失敗] {label}，嘗試降級...")

        if not res:
            return {"status": "error",
                    "message": f"即使放棄 W2+W3 需求，仍無法在本週（{horizon_days}天）內找到可行解"}

        # ── export Excel → BytesIO ────────────────────────────────────────────
        excel_bytes = None
        try:
            excel_bytes = export_excel(res, output_path=None, start_date=target_date)
        except Exception as e:
            import traceback
            print(f"Excel 輸出失敗：{traceback.format_exc()}")

        return {
            "status":      "success",
            "schedule_label": used_label,
            "data":        results_to_json(res, start_date=target_date,
                                           batch_num_start=batch_num_start),
            "excel_bytes": excel_bytes,   # BytesIO，不序列化進 JSON
        }

    except Exception as e:
        import traceback
        print(f"generate_schedule 例外: {traceback.format_exc()}")
        return {"status": "error", "message": str(e)}


# ─────────────────────────────────────────────────────────────────────────────
# main()
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("AWS 版 scheduler_api.py 不支援直接執行（無本地 SQLite）。")
    print("請透過 Flask app.py 的 /api/run-production-schedule 路由觸發排程。")


if __name__ == '__main__':
    main()