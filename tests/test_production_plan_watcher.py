from datetime import datetime

from openpyxl import Workbook

from production_plan_watcher import (
    earliest_changed_date,
    file_signature,
    find_latest_workbook,
    read_sheet_snapshot,
)


def make_workbook(path, quantities):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "P_plan Reagent"
    sheet.cell(2, 2, "Panel_NO")
    sheet.cell(2, 5, "Plan")
    sheet.cell(2, 6, datetime(2026, 6, 9))
    sheet.cell(2, 7, datetime(2026, 6, 10))
    for offset, (panel_no, values) in enumerate(quantities.items(), start=3):
        sheet.cell(offset, 2, panel_no)
        sheet.cell(offset, 5, "Plan")
        sheet.cell(offset, 6, values[0])
        sheet.cell(offset, 7, values[1])
    workbook.save(path)


def test_snapshot_finds_earliest_changed_date(tmp_path):
    workbook_path = tmp_path / "Production plan-20260609.xlsm"
    make_workbook(workbook_path, {"3620101301": (100, 200)})
    previous = read_sheet_snapshot(workbook_path)

    make_workbook(workbook_path, {"3620101301": (150, 200)})
    current = read_sheet_snapshot(workbook_path)

    assert earliest_changed_date(
        previous, current, "2026-06-09"
    ) == "2026-06-09"


def test_snapshot_ignores_unchanged_sheet(tmp_path):
    workbook_path = tmp_path / "Production plan-20260609.xlsm"
    make_workbook(workbook_path, {"3620101301": (100, 200)})
    snapshot = read_sheet_snapshot(workbook_path)

    assert earliest_changed_date(snapshot, snapshot, "2026-06-09") is None


def test_first_sync_starts_at_configured_date(tmp_path):
    workbook_path = tmp_path / "Production plan-20260609.xlsm"
    make_workbook(workbook_path, {"3620101301": (100, 200)})
    snapshot = read_sheet_snapshot(workbook_path)

    assert earliest_changed_date(None, snapshot, "2026-06-10") == "2026-06-10"


def test_latest_workbook_uses_filename_date(tmp_path):
    make_workbook(
        tmp_path / "Production plan-20260608.xlsm",
        {"3620101301": (100, 200)},
    )
    latest = tmp_path / "Production plan-20260609.xlsm"
    make_workbook(latest, {"3620101301": (100, 200)})

    assert find_latest_workbook(tmp_path) == latest


def test_deleted_monitored_workbook_does_not_fall_back(tmp_path):
    old = tmp_path / "Production plan-20260608.xlsm"
    make_workbook(old, {"3620101301": (100, 200)})

    assert find_latest_workbook(
        tmp_path, "Production plan-20260609.xlsm"
    ) is None


def test_newer_workbook_replaces_deleted_monitored_workbook(tmp_path):
    make_workbook(
        tmp_path / "Production plan-20260608.xlsm",
        {"3620101301": (100, 200)},
    )
    newer = tmp_path / "Production plan-20260610.xlsm"
    make_workbook(newer, {"3620101301": (100, 200)})

    assert find_latest_workbook(
        tmp_path, "Production plan-20260609.xlsm"
    ) == newer


def test_existing_monitored_workbook_is_used_until_newer_file_exists(tmp_path):
    make_workbook(
        tmp_path / "Production plan-20260608.xlsm",
        {"3620101301": (100, 200)},
    )
    monitored = tmp_path / "Production plan-20260609.xlsm"
    make_workbook(monitored, {"3620101301": (100, 200)})

    assert find_latest_workbook(tmp_path, monitored.name) == monitored


def test_file_signature_changes_after_workbook_update(tmp_path):
    workbook_path = tmp_path / "Production plan-20260609.xlsm"
    make_workbook(workbook_path, {"3620101301": (100, 200)})
    previous = file_signature(workbook_path)

    make_workbook(
        workbook_path,
        {"3620101301": (100, 200), "3620102301": (300, 400)},
    )

    assert file_signature(workbook_path) != previous
