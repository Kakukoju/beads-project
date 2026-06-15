"""
BEADS 庫存 Excel 監控程式 (Windows 本地端)
- 每 10 分鐘檢查檔案是否變動 (檔名/時間/大小)
- 有變動才讀取 Excel → 上傳 JSON → push to RDS
- 不開啟檔案鎖，不影響工作者

使用方式:
  pip install openpyxl requests
  python beads_inventory_monitor.py

可用 Windows Task Scheduler 設為開機啟動
"""

import os, re, glob, time as time_mod, json, logging, threading
from datetime import datetime, date, time as dt_time
import openpyxl
import requests

try:
    from watchdog.observers.polling import PollingObserver
    from watchdog.events import FileSystemEventHandler
    HAS_WATCHDOG = True
except ImportError:
    HAS_WATCHDOG = False

# ─── 設定 ───
WATCH_DIR = r"\\fls341\MBBU_FAB\MB_PD\生管自動化\滴定\祐銓"
FILE_PATTERN = "*BEADS庫存*-*NEW.xlsm"
API_URL = "https://52-192-28-39.sslip.io/api/upload-beads-json"
API_KEY = "beadsops-upload-key"
INTERVAL = 600  # 10 分鐘
STATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".beads_inv_state.json")

SHEET_PREFIX = "BEADS庫存表("
HEADER_ROW = 5
START_COL = 1
END_COL = 15  # A:O

LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "beads_inventory_monitor.log")
LOG_MAX_DAYS = 30

def _rotate_log():
    """log 檔超過 30 天就刪除重建"""
    if os.path.exists(LOG_FILE):
        age = (datetime.now() - datetime.fromtimestamp(os.path.getctime(LOG_FILE))).days
        if age >= LOG_MAX_DAYS:
            os.remove(LOG_FILE)

_rotate_log()
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)


def find_target_file():
    """找符合 *BEADS庫存YYYYMMDD-YYYYMMDD* 的檔案，多個則取後面日期最新的"""
    pattern = os.path.join(WATCH_DIR, FILE_PATTERN)
    files = glob.glob(pattern)
    if not files:
        return None

    # 提取日期對 YYYYMMDD-YYYYMMDD
    date_re = re.compile(r"(\d{8})--?(\d{8})")
    candidates = []
    for f in files:
        m = date_re.search(os.path.basename(f))
        if m:
            candidates.append((m.group(2), f))  # 取後面日期排序

    if not candidates:
        return files[0]  # fallback

    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def get_file_state(filepath):
    """取得檔名+修改時間+大小"""
    stat = os.stat(filepath)
    return {
        "path": filepath,
        "name": os.path.basename(filepath),
        "mtime": stat.st_mtime,
        "size": stat.st_size,
    }


def load_prev_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def save_state(state):
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False)


def find_sheet(wb):
    """找 BEADS庫存表(YYYYMM~ 的 sheet"""
    for name in wb.sheetnames:
        if name.startswith(SHEET_PREFIX):
            return name
    return None


def read_excel_data(filepath):
    """讀取 Excel A:O, header=row5, start=row6, 過濾全空列"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_name = find_sheet(wb)
    if not sheet_name:
        wb.close()
        raise ValueError(f"找不到 sheet 開頭為 '{SHEET_PREFIX}'")

    ws = wb[sheet_name]
    # 讀 header
    headers = []
    for cell in list(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW,
                                   min_col=START_COL, max_col=END_COL))[0]:
        headers.append(str(cell.value).strip() if cell.value else f"col{cell.column}")

    # 讀資料
    rows = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, min_col=START_COL, max_col=END_COL):
        values = [cell.value for cell in row]
        # 跳過全空列 (所有欄位為 None 或空字串)
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in values):
            continue
        record = {}
        for i, v in enumerate(values):
            h = headers[i]
            if v is None or (isinstance(v, str) and v.strip() == ""):
                record[h] = None
            elif isinstance(v, datetime):
                record[h] = v.strftime("%Y-%m-%d")
            elif isinstance(v, date):
                record[h] = v.strftime("%Y-%m-%d")
            elif isinstance(v, dt_time):
                record[h] = str(v)
            else:
                record[h] = v
            # 確保中文欄位名包含引號格式 (pandas to_sql 會自動處理)
        rows.append(record)

    wb.close()
    log.info(f"讀取 sheet '{sheet_name}': {len(rows)} 筆")
    return rows


def upload_json(data):
    """上傳 JSON 到 API → push to RDS schedule.beads_Inventory"""
    resp = requests.post(
        API_URL,
        json=data,
        headers={
            "Content-Type": "application/json; charset=utf-8",
            "X-Api-Key": API_KEY,
        },
        timeout=60,
    )
    resp.raise_for_status()
    result = resp.json()
    if not result.get("ok"):
        raise RuntimeError(f"API error: {result.get('error')}")
    return result


# ─── 核心: 檢查並上傳 ───
def check_and_upload():
    """檢查檔案變動，有變動就上傳"""
    try:
        filepath = find_target_file()
        if not filepath:
            log.warning("找不到符合條件的檔案")
            return

        current = get_file_state(filepath)
        prev = load_prev_state()

        changed = (
            prev is None
            or prev.get("name") != current["name"]
            or prev.get("mtime") != current["mtime"]
            or prev.get("size") != current["size"]
        )

        if not changed:
            return

        log.info(f"偵測到變動: {current['name']} (size={current['size']}, mtime={current['mtime']})")
        data = read_excel_data(filepath)
        result = upload_json(data)
        log.info(f"上傳成功: {result.get('rows')} 筆 → schedule.\"beads_Inventory\"")
        save_state(current)

    except PermissionError:
        log.warning("檔案被鎖定中，下次再試")
    except Exception as e:
        log.error(f"錯誤: {e}", exc_info=True)


# ─── Watchdog handler ───
class BeadsFileHandler(FileSystemEventHandler):
    """檔案事件觸發時延遲 5 秒後執行 check（避免存檔過程中讀取）"""
    def __init__(self):
        self._timer = None

    def _debounce(self):
        if self._timer:
            self._timer.cancel()
        self._timer = threading.Timer(5.0, check_and_upload)
        self._timer.start()

    def on_modified(self, event):
        if not event.is_directory and "BEADS庫存" in os.path.basename(event.src_path):
            log.info(f"[watchdog] 偵測到修改: {os.path.basename(event.src_path)}")
            self._debounce()

    def on_created(self, event):
        if not event.is_directory and "BEADS庫存" in os.path.basename(event.src_path):
            log.info(f"[watchdog] 偵測到新檔: {os.path.basename(event.src_path)}")
            self._debounce()


def main():
    log.info("BEADS 庫存監控啟動")
    log.info(f"監控目錄: {WATCH_DIR}")
    log.info(f"檢查間隔: {INTERVAL}s")
    log.info(f"Watchdog: {'啟用 (PollingObserver)' if HAS_WATCHDOG else '未安裝，僅用 polling'}")

    # 啟動 watchdog (網路磁碟用 PollingObserver 較穩定)
    if HAS_WATCHDOG and os.path.isdir(WATCH_DIR):
        observer = PollingObserver(timeout=30)
        observer.schedule(BeadsFileHandler(), WATCH_DIR, recursive=False)
        observer.daemon = True
        observer.start()
        log.info("[watchdog] observer 已啟動")

    # polling fallback: 保底每 10 分鐘檢查一次
    while True:
        check_and_upload()
        time_mod.sleep(INTERVAL)


if __name__ == "__main__":
    main()
