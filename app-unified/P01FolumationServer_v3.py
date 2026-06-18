# app.py — Unified Flask: (1) CSV→SQLite 上傳, (2) DB 查詢/匯出 Excel
# -*- coding: utf-8 -*-
from flask import Flask, request, jsonify, Response, send_file
from flask_cors import CORS
from pathlib import Path
from openpyxl import load_workbook
import sqlite3, csv, io, re, logging, os
import qrcode


app = Flask(__name__)
CORS(app)
logging.basicConfig(level=logging.INFO)
log = logging.getLogger("unified")

# ==================== Config ====================
# 可用環境變數覆蓋，否則用預設值；你可改成你現在的實際路徑
app.config.update(
    DB_PATH=os.environ.get(
        "DB_PATH",
        r"D:\OneDrive - 天亮醫療器材股份有限公司\.vscode\資料庫\配藥\P01_formualte_schedule.db"
    ),
    AUTH_TOKEN=os.environ.get("AUTH_TOKEN", ""),  # 空字串=不驗證
    TEMPLATE_XLSX=os.environ.get(
        "TEMPLATE_XLSM",
        r"\\fls341\Reagent RD\配藥端 -配製紀錄表\配藥紀錄Temp.xlsm"
    ),
    # Excel 對欄位用的預設：第 8 列為表頭（依你的模板可調）
    HEADER_ROW=int(os.environ.get("HEADER_ROW", "8")),
    # Fallback 欄位：B / P / S（1-based col index）
    COL_BOM_FALLBACK=int(os.environ.get("COL_BOM_FALLBACK", "2")),
    COL_TOTAL_FALLBACK=int(os.environ.get("COL_TOTAL_FALLBACK", "16")),
    COL_LOT_FALLBACK=int(os.environ.get("COL_LOT_FALLBACK", "19")),
)

# 確保 DB 目錄存在
Path(app.config["DB_PATH"]).parent.mkdir(parents=True, exist_ok=True)

# ================== Common helpers ==================
def auth_ok() -> bool:
    token = app.config.get("AUTH_TOKEN", "")
    return (not token) or (request.headers.get("Authorization", "") == ("Bearer " + token))

def qident(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'

def q(conn, sql, args=()):
    cur = conn.execute(sql, args)
    cols = [c[0] for c in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]

def safe_table_name(name: str) -> str:
    """僅允許字母/數字/底線，避免 SQL 注入；如需更嚴格可改白名單。"""
    if not name:
        raise ValueError("table is required")
    if not re.fullmatch(r"[A-Za-z0-9_]+", name):
        raise ValueError("invalid table name")
    return name

def get_existing_cols(con: sqlite3.Connection, table: str):
    cur = con.execute(f'PRAGMA table_info({qident(table)});')
    return [r[1] for r in cur.fetchall()]

def ensure_table_and_columns(con: sqlite3.Connection, table: str, headers: list[str]) -> None:
    cur = con.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?;", (table,))
    exists = cur.fetchone() is not None
    if not exists:
        cols_def = ", ".join(f'{qident(h)} TEXT' for h in headers)
        con.execute(f'CREATE TABLE {qident(table)} ({cols_def});')
        return
    existing = set(get_existing_cols(con, table))
    for h in headers:
        if h not in existing:
            con.execute(f'ALTER TABLE {qident(table)} ADD COLUMN {qident(h)} TEXT;')

def ensure_unique_index(con: sqlite3.Connection, table: str, pk_cols: list[str]) -> None:
    if not pk_cols:
        return
    idx = "ux_" + re.sub(r'[^A-Za-z0-9_]', '_', table) + "__" + "__".join(
        re.sub(r'[^A-Za-z0-9_]', '_', c) for c in pk_cols
    )
    cols = ", ".join(qident(c) for c in pk_cols)
    con.execute(f'CREATE UNIQUE INDEX IF NOT EXISTS {qident(idx)} ON {qident(table)} ({cols});')

def count_distinct_pk(con: sqlite3.Connection, table: str, pk_cols: list[str]) -> int:
    if not pk_cols:
        cur = con.execute(f'SELECT COUNT(*) FROM {qident(table)};')
    else:
        cols = ", ".join(qident(c) for c in pk_cols)
        cur = con.execute(f'SELECT COUNT(DISTINCT {cols}) FROM {qident(table)};')
    return int(cur.fetchone()[0])

def upsert_rows(con: sqlite3.Connection, table: str, headers: list[str], rows: list[tuple], pk_cols: list[str]) -> tuple[int,int]:
    if not rows:
        return (0, 0)
    cols = ", ".join(qident(h) for h in headers)
    qmarks = ", ".join(["?"] * len(headers))
    inserted = 0
    updated = 0
    upd_cols = [h for h in headers if h not in pk_cols]
    if upd_cols:
        upd = ", ".join(f'{qident(c)}=excluded.{qident(c)}' for c in upd_cols)
        sql = f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks}) ' \
              f'ON CONFLICT({", ".join(qident(c) for c in pk_cols)}) DO UPDATE SET {upd};'
    else:
        sql = f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks}) ' \
              f'ON CONFLICT({", ".join(qident(c) for c in pk_cols)}) DO NOTHING;'
    try:
        before = count_distinct_pk(con, table, pk_cols)
    except Exception:
        before = None
    con.executemany(sql, rows)
    try:
        after = count_distinct_pk(con, table, pk_cols)
        if before is not None:
            inserted = max(0, after - before)
            updated = max(0, len(rows) - inserted)
    except Exception:
        pass
    return (inserted, updated)

def insert_rows(con: sqlite3.Connection, table: str, headers: list[str], rows: list[tuple]) -> int:
    if not rows:
        return 0
    cols = ", ".join(qident(h) for h in headers)
    qmarks = ", ".join(["?"] * len(headers))
    con.executemany(f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks});', rows)
    return len(rows)

def replace_rows(con: sqlite3.Connection, table: str, headers: list[str], rows: list[tuple]) -> int:
    con.execute(f'DELETE FROM {qident(table)};')
    return insert_rows(con, table, headers, rows)

def parse_csv_text(text: str) -> tuple[list[str], list[tuple]]:
    """以 UTF-8(-sig) 解析 CSV 文字，回傳 (headers, rows)"""
    if text and text[:1] == '\ufeff':
        text = text.lstrip('\ufeff')
    reader = csv.reader(io.StringIO(text))
    try:
        headers = next(reader)
    except StopIteration:
        headers = []
    headers = [h.strip() if h else f"col_{i+1}" for i, h in enumerate(headers)]
    rows = []
    for r in reader:
        if len(r) < len(headers):
            r = r + [""] * (len(headers) - len(r))
        elif len(r) > len(headers):
            r = r[:len(headers)]
        rows.append(tuple(r))
    return headers, rows

def get_mode() -> str:
    mode = (request.form.get("mode") or request.args.get("mode") or "replace").strip().lower()
    return mode if mode in ("append", "replace", "upsert") else "replace"

def want_csv_response() -> bool:
    """是否回傳純文字 CSV 結果（用於前端 VBA 逐工單畫勾/叉）"""
    return (request.args.get("return") or "").strip().lower() == "csv"

def extract_work_orders(headers: list[str], rows: list[tuple]) -> list[str]:
    """從上傳內容中蒐集「工單號碼」欄位的唯一清單。"""
    candidates = ["工單號碼"]
    idx = -1
    for name in candidates:
        try:
            idx = headers.index(name)
            break
        except ValueError:
            continue
    if idx < 0:
        return []
    seen = set()
    out = []
    for r in rows:
        if idx < len(r):
            wo = str(r[idx]).strip()
            if wo and wo not in seen:
                seen.add(wo)
                out.append(wo)
    return out

def to_csv_response_line_list(work_orders: list[str], status: str) -> str:
    buf = io.StringIO()
    buf.write("work_order,status\n")
    for wo in work_orders:
        buf.write(f"{wo},{status}\n")
    return buf.getvalue()

def read_cells_as_text(xlsx_path, cells):
    wb = load_workbook(xlsx_path, data_only=True, keep_vba=True)
    ws = wb.active  # 或指定工作表
    values = []
    for addr in cells:
        # 支援 Sheet!A1
        if "!" in addr:
            sheet, cell = addr.split("!", 1)
            ws2 = wb[sheet]
            values.append("" if ws2[cell].value is None else str(ws2[cell].value))
        else:
            values.append("" if ws[addr].value is None else str(ws[addr].value))
    return values

# ================== CSV upload endpoints ==================
@app.post("/upload_csv")
def upload_csv():
    if not auth_ok():
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    table = (request.form.get("table") or request.args.get("table") or "").strip()
    pk_raw = (request.form.get("pk") or request.args.get("pk") or "").strip()
    pk_cols = [c.strip() for c in pk_raw.split(",") if c.strip()]
    mode = get_mode()

    f = request.files.get("file") or request.files.get("csv")
    if not f:
        return jsonify({
            "ok": False,
            "error": "missing file",
            "debug": {
                "form": list(request.form.keys()),
                "files": list(request.files.keys()),
                "content_type": request.headers.get("Content-Type", ""),
            }
        }), 400

    raw = f.read()
    text = raw.decode("utf-8-sig", errors="replace")
    headers, rows = parse_csv_text(text)
    return _handle_upload(table, pk_cols, headers, rows, mode)

@app.post("/upload_csv_raw")
def upload_csv_raw():
    if not auth_ok():
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    table = (request.args.get("table") or "").strip()
    pk_raw = (request.args.get("pk") or "").strip()
    mode = get_mode()
    pk_cols = [c.strip() for c in pk_raw.split(",") if c.strip()]

    text = request.get_data(cache=False, as_text=True)
    if not text:
        return jsonify({"ok": False, "error": "empty body"}), 400

    headers, rows = parse_csv_text(text)
    return _handle_upload(table, pk_cols, headers, rows, mode)

def _handle_upload(table: str, pk_cols: list[str], headers: list[str], rows: list[tuple], mode: str):
    if not table:
        return jsonify({"ok": False, "error": "missing table"}), 400

    # 若未指定 pk，且表頭有「工單號碼,料號」或「工單號碼」可自動採用
    if not pk_cols:
        if "工單號碼" in headers and "料號" in headers:
            pk_cols = ["工單號碼", "料號"]
        elif "工單號碼" in headers:
            pk_cols = ["工單號碼"]

    batch_work_orders = extract_work_orders(headers, rows)

    con = sqlite3.connect(app.config["DB_PATH"])
    # 加速（與你原設定一致）
    con.execute("PRAGMA journal_mode=WAL;")
    con.execute("PRAGMA synchronous=OFF;")
    con.execute("PRAGMA temp_store=MEMORY;")

    try:
        with con:
            ensure_table_and_columns(con, table, headers)
            if pk_cols:
                ensure_unique_index(con, table, pk_cols)

            if mode == "append":
                if pk_cols:
                    upsert_rows(con, table, headers, rows, pk_cols)  # 視為 upsert-append
                else:
                    insert_rows(con, table, headers, rows)
            elif mode == "upsert":
                if not pk_cols:
                    return jsonify({"ok": False, "error": "pk required for upsert"}), 400
                upsert_rows(con, table, headers, rows, pk_cols)
            else:  # replace
                if pk_cols:
                    upsert_rows(con, table, headers, rows, pk_cols)
                else:
                    replace_rows(con, table, headers, rows)
    except Exception as e:
        con.close()
        if want_csv_response():
            csv_text = to_csv_response_line_list(batch_work_orders, "FAIL")
            return Response(csv_text, status=500, mimetype="text/csv; charset=utf-8")
        else:
            log.exception("upload failed")
            return jsonify({"ok": False, "error": str(e)}), 500

    con.close()

    if want_csv_response():
        csv_text = to_csv_response_line_list(batch_work_orders, "OK")
        return Response(csv_text, status=200, mimetype="text/csv; charset=utf-8")
    else:
        return Response(status=204)

# ================== Query / Export endpoints ==================
@app.get("/api/formulation")
def get_formulation():
    wo = (request.args.get("work_order") or "").strip()
    table = (request.args.get("table") or "beads").strip()
    try:
        table = safe_table_name(table)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    if not wo:
        return jsonify({"error": "work_order is required"}), 400

    sql = f"""
      SELECT
        工單號碼              AS work_order,
        BeadsLot              AS bead_lot,
        試劑配製日期           AS date_made,
        配製人員              AS maker,
        料號                  AS material,
        化學品名              AS chemical_name,
        FillerName            AS filler_name,
        重量紀錄              AS weight,
        Filler_Lot            AS lot_no,
        L1OD, L2OD, 起始L1OD, 起始L2OD,
        總重量                AS total_weight,
        配製備註              AS remarks
      FROM {table}
      WHERE 工單號碼 = ?
      ORDER BY 料號
    """
    with sqlite3.connect(app.config["DB_PATH"]) as conn:
        rows = q(conn, sql, (wo,))
    return jsonify(rows)

@app.get("/api/formulation/export_excel")
def export_excel():
    wo = (request.args.get("work_order") or "").strip()
    table = (request.args.get("table") or "beads").strip()
    try:
        table = safe_table_name(table)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    if not wo:
        return jsonify({"error": "work_order is required"}), 400

    template = app.config["TEMPLATE_XLSX"]
    if not os.path.exists(template):
        return jsonify({"error": f"template not found: {template}"}), 500

    sql = f"""
      SELECT
        工單號碼    AS work_order,
        BeadsLot    AS bead_lot,
        料號        AS material,     -- → BOM P/N
        重量紀錄    AS weight,       -- → Total Qty
        Filler_Lot  AS lot_no        -- → Lot No.
      FROM {table}
      WHERE 工單號碼 = ?
      ORDER BY 料號
    """
    with sqlite3.connect(app.config["DB_PATH"]) as conn:
        rows = q(conn, sql, (wo,))

    wb = load_workbook(template)
    if "製程紀錄表" not in wb.sheetnames:
        return jsonify({"error": "template has no sheet: 製程紀錄表"}), 500
    ws = wb["製程紀錄表"]

    # 工單號碼寫入 V6
    ws["V6"] = wo

    HEADER_ROW = app.config["HEADER_ROW"]
    COL_BOM_FALLBACK   = app.config["COL_BOM_FALLBACK"]     # B
    COL_TOTAL_FALLBACK = app.config["COL_TOTAL_FALLBACK"]   # P
    COL_LOT_FALLBACK   = app.config["COL_LOT_FALLBACK"]     # S

    def norm(s): return (str(s or "")).strip().lower().replace(" ", "")

    # 建立「表頭文字→欄位 index」對照
    title2col = {}
    max_col = ws.max_column if ws.max_column else 1
    for c in range(1, max_col + 1):
        title2col[norm(ws.cell(HEADER_ROW, c).value)] = c

    def get_col(names, fallback):
        for n in names:
            col = title2col.get(norm(n))
            if col:
                return col
        return fallback

    # 允許中英文多種拼法
    col_bom   = get_col(["bom p/n", "bompn", "料號", "bom"], COL_BOM_FALLBACK)
    col_total = get_col(["total qty", "totalqty", "重量紀錄"], COL_TOTAL_FALLBACK)
    col_lot   = get_col(["lot no.", "lotno", "filler_lot", "lot no", "lotno."], COL_LOT_FALLBACK)

    # 清掉表頭下方的舊資料列（保留格式）
    start_row = HEADER_ROW + 1
    end_row = ws.max_row if ws.max_row > start_row else start_row
    if end_row >= start_row:
        ws.delete_rows(start_row, end_row - start_row + 1)

    r = start_row
    for it in rows:  # rows 來源：DB 查詢
        ws.cell(r, col_bom).value   = it.get("material")
        ws.cell(r, col_total).value = it.get("weight")
        ws.cell(r, col_lot).value   = it.get("lot_no")
        r += 1

    bio = io.BytesIO()
    wb.save(bio); bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name=f"製程紀錄表_{wo}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ================== Entrypoint ==================
if __name__ == "__main__":
    # 開發直接跑 Flask；正式可改用 waitress/gunicorn 搭反代
    port = int(os.environ.get("PORT", "8055"))
    app.run(host="0.0.0.0", port=port, debug=False)
    # from waitress import serve
    # serve(app, host="0.0.0.0", port=port)
