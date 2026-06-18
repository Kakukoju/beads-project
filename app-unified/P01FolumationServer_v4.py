# -*- coding: utf-8 -*-
from flask import Flask, request, jsonify, Response, send_file
from flask_cors import CORS
from pathlib import Path
from openpyxl import load_workbook
import sqlite3, csv, io, re, logging, os, uuid, shutil, pythoncom
from qrcode.constants import ERROR_CORRECT_M
import qrcode
from win32com.client import Dispatch

app = Flask(__name__)
CORS(app)
logging.basicConfig(level=logging.INFO)
log = logging.getLogger("formulation")

# ==================== Config ====================
app.config.update(
    # SQLite 路徑
    DB_PATH=os.environ.get("DB_PATH", r"D:\配藥表\資料庫\P01_formualte_schedule.db"),
    # Excel 範本（配藥紀錄Temp.xlsm）
    TEMPLATE_XLSM=os.environ.get("TEMPLATE_XLSM", r"D:\配藥表\配藥紀錄\配藥紀錄Temp.xlsm"),
    # 上傳暫存
    UPLOAD_DIR=os.environ.get("UPLOAD_DIR", r"D:\配藥表\_temp"),
    # 資料列標題所在行（活頁簿→DB 與 填表 都共用）
    HEADER_ROW=int(os.environ.get("HEADER_ROW", "8")),
    # 填表 fallback 欄位（B、P、S）
    COL_BOM_FALLBACK=int(os.environ.get("COL_BOM_FALLBACK", "2")),
    COL_TOTAL_FALLBACK=int(os.environ.get("COL_TOTAL_FALLBACK", "16")),
    COL_LOT_FALLBACK=int(os.environ.get("COL_LOT_FALLBACK", "19")),
    # UNC 預設
    DEFAULT_SERVER_DIR=os.environ.get("DEFAULT_SERVER_DIR", r"\\fls341\Reagent RD\配藥端 -配製紀錄表\\"),
)

Path(app.config["UPLOAD_DIR"]).mkdir(parents=True, exist_ok=True)
Path(app.config["DB_PATH"]).parent.mkdir(parents=True, exist_ok=True)

# ============== Helpers: DB & naming ==============
def qident(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'

def safe_table_name(name: str) -> str:
    if not name:
        raise ValueError("table is required")
    if not re.fullmatch(r"[A-Za-z0-9_]+", name):
        raise ValueError(f"invalid table name: {name}")
    return name

def ensure_table_and_columns(con: sqlite3.Connection, table: str, headers: list[str]) -> None:
    cur = con.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?;", (table,))
    exists = cur.fetchone() is not None
    if not exists:
        cols_def = ", ".join(f'{qident(h)} TEXT' for h in headers)
        con.execute(f'CREATE TABLE {qident(table)} ({cols_def});')
        return
    cur = con.execute(f'PRAGMA table_info({qident(table)});')
    existing = {r[1] for r in cur.fetchall()}
    for h in headers:
        if h not in existing:
            con.execute(f'ALTER TABLE {qident(table)} ADD COLUMN {qident(h)} TEXT;')

def upsert_rows(con: sqlite3.Connection, table: str, headers: list[str], rows: list[tuple], pk_cols: list[str]) -> None:
    cols = ", ".join(qident(h) for h in headers)
    qmarks = ", ".join(["?"] * len(headers))
    if pk_cols:
        upd_cols = [h for h in headers if h not in pk_cols]
        if upd_cols:
            upd = ", ".join(f'{qident(c)}=excluded.{qident(c)}' for c in upd_cols)
            sql = f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks}) ' \
                  f'ON CONFLICT({", ".join(qident(c) for c in pk_cols)}) DO UPDATE SET {upd};'
        else:
            sql = f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks}) ' \
                  f'ON CONFLICT({", ".join(qident(c) for c in pk_cols)}) DO NOTHING;'
    else:
        sql = f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks});'
    con.executemany(sql, rows)

def normalize_table_name(sheet_name: str) -> str:
    # 允許中英文工作表名 → 規範成英數與底線（但會盡量保留原意）
    name = re.sub(r'\s+', '_', sheet_name.strip())
    name = re.sub(r'[^\w]', '_', name)  # 非 [A-Za-z0-9_]
    name = re.sub(r'_+', '_', name).strip('_')
    if not name:
        name = "sheet"
    return name

def read_sheet_to_db(con: sqlite3.Connection, ws, table_name: str, header_row: int) -> dict:
    """
    依 header_row 讀 headers，將 header_row+1 到最後一列的資料寫入 DB。
    空值以空字串寫入；全部轉成文字儲存。
    PK：若同時有「工單號碼」「料號」→ upsert by (工單號碼,料號)；若只有工單號碼 → by (工單號碼)；否則單純 append。
    """
    max_col = ws.max_column or 1
    headers = [str(ws.cell(header_row, c).value or "").strip() for c in range(1, max_col + 1)]
    # 修正空白表頭
    headers = [h if h else f"col_{i+1}" for i, h in enumerate(headers)]

    rows = []
    r = header_row + 1
    max_r = ws.max_row or r
    for rr in range(r, max_r + 1):
        row = []
        nonempty = False
        for c in range(1, max_col + 1):
            v = ws.cell(rr, c).value
            if v not in (None, ""):
                nonempty = True
            row.append("" if v is None else str(v))
        if nonempty:
            rows.append(tuple(row))

    ensure_table_and_columns(con, table_name, headers)
    pk_cols = []
    if "工單號碼" in headers and "料號" in headers:
        pk_cols = ["工單號碼", "料號"]
    elif "工單號碼" in headers:
        pk_cols = ["工單號碼"]
    upsert_rows(con, table_name, headers, rows, pk_cols)
    return {"table": table_name, "headers": headers, "rows": len(rows), "pk": pk_cols}

# ============== Excel COM 宏 ==============
def run_excel_macro(xlsm_path: str, macro_name: str) -> tuple[bool, str | None]:
    try:
        pythoncom.CoInitialize()
        excel = Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            excel.AutomationSecurity = 1  # msoAutomationSecurityLow
        except Exception:
            pass

        wb = excel.Workbooks.Open(xlsm_path)
        try:
            excel.Run(macro_name)  # 例："manuinsert.InsertNewByWorkOrderArrays"
            wb.Save()
            return True, None
        except Exception as e:
            return False, f"Run macro failed: {e}"
        finally:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
            excel.Quit()
    except Exception as e:
        return False, f"Excel automation error: {e}"
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass

# ============== In-memory 暫存 ==============
UPLOAD_DIR = Path(app.config["UPLOAD_DIR"])
TEMP_IN_MAP: dict[str, Path] = {}   # 上傳（且已跑過巨集）的原始檔
TEMP_OUT_MAP: dict[str, Path] = {}  # 以 TEMPLATE_XLSM 產生的填寫後暫存檔

# ============== 健康檢查 ==============
@app.get("/api/health")
def health():
    return {"ok": True}

@app.get("/")
def root():
    return "Backend OK. Try /api/health.", 200

# ============== 上傳 → 執行巨集 → 導入 SQLite（前兩個工作表） ==============
@app.post("/api/upload_excel")
def api_upload_excel():
    f = request.files.get("file")
    if not f:
        return jsonify(ok=False, message="missing file"), 400

    fname_raw = f.filename or "uploaded.xlsm"
    if "配藥表" not in fname_raw:
        return jsonify(ok=False, message="所選檔名未包含「配藥表」，請重新選擇正確的配藥表 Excel。"), 400

    ext = Path(fname_raw).suffix.lower() or ".xlsm"
    if ext not in (".xlsm", ".xlsx", ".xls"):
        ext = ".xlsm"

    temp_id = uuid.uuid4().hex[:8]
    temp_in_path = UPLOAD_DIR / f"upload_{temp_id}{ext}"
    f.save(str(temp_in_path))
    TEMP_IN_MAP[temp_id] = temp_in_path

    # a) 先跑巨集
    ok_macro, err_macro = run_excel_macro(str(temp_in_path), "manuinsert.InsertNewByWorkOrderArrays")
    if not ok_macro:
        return jsonify(ok=False, message=f"Excel 宏執行失敗：{err_macro}"), 500

    # b) 匯入 DB：取第 1~2 工作表
    wb = load_workbook(temp_in_path, keep_vba=True, data_only=True)
    sheets = wb.sheetnames[:2]  # 最多兩張
    ingested = []

    con = sqlite3.connect(app.config["DB_PATH"])
    con.execute("PRAGMA journal_mode=WAL;")
    con.execute("PRAGMA synchronous=OFF;")
    con.execute("PRAGMA temp_store=MEMORY;")
    try:
        with con:
            for s in sheets:
                ws = wb[s]
                table = normalize_table_name(s)
                info = read_sheet_to_db(con, ws, table, app.config["HEADER_ROW"])
                ingested.append(info)
    finally:
        con.close()

    return jsonify(ok=True, temp_id=temp_id, filename=fname_raw, sheets=sheets, ingested=ingested)

# ============== 建立：依 table(=sheet) + 工單號 → 填入 TEMPLATE_XLSM ==============
@app.post("/api/create_record")
def api_create_record():
    data = request.get_json(force=True, silent=True) or {}
    work_order = (data.get("work_order") or "").strip()
    temp_id = (data.get("temp_id") or "").strip()
    table_raw = (data.get("table") or "").strip()
    if not work_order or not temp_id:
        return jsonify(ok=False, message="work_order & temp_id are required"), 400

    # table 預設：取剛剛上傳檔的第 1 個工作表
    table_norm = None
    if table_raw:
        table_norm = normalize_table_name(table_raw)

    # 從上傳檔讀取實際第一張表名（以便預設）
    in_path = TEMP_IN_MAP.get(temp_id)
    if not in_path or not in_path.exists():
        return jsonify(ok=False, message=f"temp_id not found: {temp_id}"), 404
    wb_in = load_workbook(in_path, keep_vba=True, data_only=True)
    first_sheet = normalize_table_name(wb_in.sheetnames[0]) if wb_in.sheetnames else None
    table = table_norm or first_sheet
    if not table:
        return jsonify(ok=False, message="no sheet to use as table"), 400

    try:
        table = safe_table_name(table)
    except ValueError as e:
        return jsonify(ok=False, message=str(e)), 400

    # c) 查 DB → 填 TEMPLATE_XLSM
    sql = f"""
      SELECT 工單號碼 AS work_order, BeadsLot AS bead_lot,
             料號 AS material, 重量紀錄 AS weight, Filler_Lot AS lot_no
      FROM {qident(table)}
      WHERE 工單號碼 = ?
      ORDER BY 料號
    """
    with sqlite3.connect(app.config["DB_PATH"]) as conn:
        cur = conn.execute(sql, (work_order,))
        rows = [dict(zip([c[0] for c in cur.description], r)) for r in cur.fetchall()]

    template = app.config["TEMPLATE_XLSM"]
    if not os.path.exists(template):
        return jsonify(ok=False, message=f"template not found: {template}"), 500

    # 以 TEMPLATE 複製一份輸出檔（與 temp_id 綁定）
    out_path = UPLOAD_DIR / f"record_{temp_id}.xlsm"
    shutil.copy2(template, out_path)

    wb = load_workbook(out_path, keep_vba=True, data_only=False)
    if "製程紀錄表" not in wb.sheetnames:
        return jsonify(ok=False, message="template has no sheet: 製程紀錄表"), 500
    ws = wb["製程紀錄表"]

    # 工單號碼 → V6
    ws["V6"] = work_order

    HEADER_ROW = app.config["HEADER_ROW"]
    COL_BOM_FALLBACK   = app.config["COL_BOM_FALLBACK"]
    COL_TOTAL_FALLBACK = app.config["COL_TOTAL_FALLBACK"]
    COL_LOT_FALLBACK   = app.config["COL_LOT_FALLBACK"]

    def norm(s): return (str(s or "")).strip().lower().replace(" ", "")

    # 表頭對照
    title2col = {}
    max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        title2col[norm(ws.cell(HEADER_ROW, c).value)] = c

    def get_col(names, fallback):
        for n in names:
            col = title2col.get(norm(n))
            if col:
                return col
        return fallback

    col_bom   = get_col(["bom p/n","bompn","料號","bom"], COL_BOM_FALLBACK)
    col_total = get_col(["total qty","totalqty","重量紀錄"], COL_TOTAL_FALLBACK)
    col_lot   = get_col(["lot no.","lotno","filler_lot","lot no","lotno."], COL_LOT_FALLBACK)

    # 清資料列（保留格式）
    start_row = HEADER_ROW + 1
    end_row = ws.max_row if ws.max_row > start_row else start_row
    if end_row >= start_row:
        ws.delete_rows(start_row, end_row - start_row + 1)

    # 寫入
    r = start_row
    for it in rows:
        ws.cell(r, col_bom).value   = it.get("material")
        ws.cell(r, col_total).value = it.get("weight")
        ws.cell(r, col_lot).value   = it.get("lot_no")
        r += 1

    wb.save(out_path)
    TEMP_OUT_MAP[temp_id] = out_path
    return jsonify(ok=True, filled=len(rows), out_path=str(out_path), table=table)

# ============== 預覽 / 下載 / 儲存 ==============
@app.get("/api/template_preview")
def api_template_preview():
    temp_id = (request.args.get("temp_id") or "").strip()
    out_path = TEMP_OUT_MAP.get(temp_id)
    if not out_path or not Path(out_path).exists():
        return jsonify(ok=False, message=f"no generated file for temp_id: {temp_id}"), 404

    wb = load_workbook(out_path, keep_vba=True, data_only=True)
    if "製程紀錄表" not in wb.sheetnames:
        return jsonify(ok=False, message="template has no sheet: 製程紀錄表"), 500
    ws = wb["製程紀錄表"]

    HEADER_ROW = app.config["HEADER_ROW"]
    max_col = ws.max_column or 1
    headers = [str(ws.cell(HEADER_ROW, c).value or "") for c in range(1, max_col + 1)]

    rows = []
    r = HEADER_ROW + 1
    max_r = min(ws.max_row or r, r + 50)
    for rr in range(r, max_r + 1):
        row = []
        for c in range(1, max_col + 1):
            v = ws.cell(rr, c).value
            row.append("" if v is None else str(v))
        rows.append(row)
    return jsonify(ok=True, headers=headers, rows=rows)

@app.get("/api/template_file")
def api_template_file():
    temp_id = (request.args.get("temp_id") or "").strip()
    out_path = TEMP_OUT_MAP.get(temp_id)
    if not out_path or not Path(out_path).exists():
        return jsonify(ok=False, message=f"no generated file for temp_id: {temp_id}"), 404
    return send_file(
        out_path,
        as_attachment=True,
        download_name=Path(out_path).name,
        mimetype="application/vnd.ms-excel.sheet.macroEnabled.12",
    )

@app.post("/api/save_template")
def api_save_template():
    data = request.get_json(force=True, silent=True) or {}
    temp_id = (data.get("temp_id") or "").strip()
    work_order = (data.get("work_order") or "").strip()
    server_dir = (data.get("server_dir") or "").strip()
    filename   = (data.get("filename") or "").strip()

    out_src = TEMP_OUT_MAP.get(temp_id)
    if not out_src or not Path(out_src).exists():
        return jsonify(ok=False, message=f"no generated file for temp_id: {temp_id}"), 404

    if not server_dir or not filename:
        return jsonify(ok=False, message="server_dir, filename are required"), 400
    if not filename.endswith(".xlsm"):
        filename += ".xlsm"

    # 若要覆寫 V6
    if work_order:
        wb = load_workbook(out_src, keep_vba=True, data_only=False)
        if "製程紀錄表" in wb.sheetnames:
            wb["製程紀錄表"]["V6"] = work_order
            wb.save(out_src)

    out_dir = Path(server_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / filename
    shutil.copy2(out_src, out_path)
    return jsonify(ok=True, saved_path=str(out_path))
# ============== CSV 相容上傳（upload_csv_raw / upload_csv） ==============
# 舊 VBA/工具若還在打 /upload_csv_raw，就用這一段接住。
# 同時提供 /api/upload_csv_raw 與 /upload_csv_raw 兩個路由以便相容。

def _parse_csv_text(text: str) -> tuple[list[str], list[tuple]]:
    """解析 CSV，回傳 (headers, rows)。空 header 以 col_1, col_2 補上。"""
    if text and text[:1] == '\ufeff':
        text = text.lstrip('\ufeff')
    reader = csv.reader(io.StringIO(text))
    try:
        headers = next(reader)
    except StopIteration:
        headers = []
    headers = [h.strip() if h else f"col_{i+1}" for i, h in enumerate(headers)]
    rows = []
    for r in reader:
        if len(r) < len(headers):
            r = r + [""] * (len(headers) - len(r))
        elif len(r) > len(headers):
            r = r[:len(headers)]
        rows.append(tuple(r))
    return headers, rows

def ensure_unique_index(con: sqlite3.Connection, table: str, pk_cols: list[str]) -> None:
    if not pk_cols:
        return
    idx = "ux_" + re.sub(r'[^A-Za-z0-9_]', '_', table) + "__" + "__".join(
        re.sub(r'[^A-Za-z0-9_]', '_', c) for c in pk_cols
    )
    cols = ", ".join(qident(c) for c in pk_cols)
    con.execute(f'CREATE UNIQUE INDEX IF NOT EXISTS {qident(idx)} ON {qident(table)} ({cols});')

def _get_mode() -> str:
    mode = (request.form.get("mode") or request.args.get("mode") or "replace").strip().lower()
    return mode if mode in ("append", "replace", "upsert") else "replace"

def _want_csv_response() -> bool:
    return (request.args.get("return") or "").strip().lower() == "csv"

def _insert_rows(con: sqlite3.Connection, table: str, headers: list[str], rows: list[tuple]) -> int:
    if not rows:
        return 0
    cols = ", ".join(qident(h) for h in headers)
    qmarks = ", ".join(["?"] * len(headers))
    con.executemany(f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks});', rows)
    return len(rows)

def _replace_rows(con: sqlite3.Connection, table: str, headers: list[str], rows: list[tuple]) -> int:
    con.execute(f'DELETE FROM {qident(table)};')
    return _insert_rows(con, table, headers, rows)

def _handle_csv_upload(table: str, headers: list[str], rows: list[tuple], pk_cols: list[str], mode: str):
    """建立/補欄後依 mode 寫入：append/upsert/replace。"""
    con = sqlite3.connect(app.config["DB_PATH"])
    con.execute("PRAGMA journal_mode=WAL;")
    con.execute("PRAGMA synchronous=OFF;")
    con.execute("PRAGMA temp_store=MEMORY;")
    with con:
        ensure_table_and_columns(con, table, headers)
        # 未明確給 pk 時，若含「工單號碼,料號」或「工單號碼」則採用
        if not pk_cols:
            if "工單號碼" in headers and "料號" in headers:
                pk_cols = ["工單號碼", "料號"]
            elif "工單號碼" in headers:
                pk_cols = ["工單號碼"]
        if pk_cols:
            ensure_unique_index(con, table, pk_cols)
        if mode == "append":
            if pk_cols:
                upsert_rows(con, table, headers, rows, pk_cols)
            else:
                _insert_rows(con, table, headers, rows)
        elif mode == "upsert":
            if not pk_cols:
                # 沒 pk 無法 upsert，退回 append 邏輯
                _insert_rows(con, table, headers, rows)
            else:
                upsert_rows(con, table, headers, rows, pk_cols)
        else:  # replace
            if pk_cols:
                # replace + pk：等同 upsert 全量
                upsert_rows(con, table, headers, rows, pk_cols)
            else:
                _replace_rows(con, table, headers, rows)

# 文字原文 CSV（raw body）
@app.post("/api/upload_csv_raw")
@app.post("/upload_csv_raw")
def api_upload_csv_raw():
    table = (request.args.get("table") or "").strip()
    if not table:
        return jsonify(ok=False, message="missing table"), 400
    # 允許中文表名，正規化後再套 safe_table_name
    table = normalize_table_name(table)
    try:
        table = safe_table_name(table)
    except ValueError as e:
        return jsonify(ok=False, message=str(e)), 400

    pk_raw = (request.args.get("pk") or "").strip()
    mode = _get_mode()
    pk_cols = [c.strip() for c in pk_raw.split(",") if c.strip()]

    text = request.get_data(cache=False, as_text=True)
    if not text:
        return jsonify(ok=False, message="empty body"), 400

    headers, rows = _parse_csv_text(text)
    try:
        _handle_csv_upload(table, headers, rows, pk_cols, mode)
    except Exception as e:
        app.logger.exception("upload_csv_raw failed")
        if _want_csv_response():
            return Response("status,FAIL\n", status=500, mimetype="text/csv; charset=utf-8")
        return jsonify(ok=False, message=str(e)), 500

    if _want_csv_response():
        return Response("status,OK\n", mimetype="text/csv; charset=utf-8")
    return jsonify(ok=True)

# 檔案表單 CSV（multipart/form-data）
@app.post("/api/upload_csv")
@app.post("/upload_csv")
def api_upload_csv():
    table = (request.form.get("table") or request.args.get("table") or "").strip()
    if not table:
        return jsonify(ok=False, message="missing table"), 400
    table = normalize_table_name(table)
    try:
        table = safe_table_name(table)
    except ValueError as e:
        return jsonify(ok=False, message=str(e)), 400

    pk_raw = (request.form.get("pk") or request.args.get("pk") or "").strip()
    mode = _get_mode()
    pk_cols = [c.strip() for c in pk_raw.split(",") if c.strip()]

    f = request.files.get("file") or request.files.get("csv")
    if not f:
        return jsonify(ok=False, message="missing file"), 400

    raw = f.read()
    # 以 UTF-8(-sig) 解析
    try:
        text = raw.decode("utf-8-sig", errors="replace")
    except Exception:
        text = raw.decode(errors="replace")

    headers, rows = _parse_csv_text(text)
    try:
        _handle_csv_upload(table, headers, rows, pk_cols, mode)
    except Exception as e:
        app.logger.exception("upload_csv failed")
        if _want_csv_response():
            return Response("status,FAIL\n", status=500, mimetype="text/csv; charset=utf-8")
        return jsonify(ok=False, message=str(e)), 500

    if _want_csv_response():
        return Response("status,OK\n", mimetype="text/csv; charset=utf-8")
    return jsonify(ok=True)

# ============== 產生 QR（以已儲存或暫存成品檔為來源） ==============
@app.post("/api/qr_png_from_cells")
def api_qr_png_from_cells():
    data = request.get_json(force=True)
    work_order = (data.get("work_order") or "").strip()
    cells = data.get("cells") or []
    joiner = (data.get("joiner") or "|")
    file_path = (data.get("file_path") or "").strip()
    temp_id = (data.get("temp_id") or "").strip()

    if not (file_path or temp_id):
        return jsonify(ok=False, message="file_path or temp_id is required"), 400
    if not cells:
        return jsonify(ok=False, message="cells is required"), 400

    xlsm_path = None
    if file_path:
        xlsm_path = file_path
    elif temp_id:
        xlsm_path = TEMP_OUT_MAP.get(temp_id)

    if not xlsm_path or not os.path.exists(xlsm_path):
        return jsonify(ok=False, message=f"file not found: {xlsm_path}"), 404

    wb = load_workbook(xlsm_path, data_only=True, keep_vba=True)
    ws = wb.active
    vals = []
    for addr in cells:
        if "!" in addr:
            sheet, cell = addr.split("!", 1)
            vals.append("" if wb[sheet][cell].value is None else str(wb[sheet][cell].value))
        else:
            vals.append("" if ws[addr].value is None else str(ws[addr].value))
    payload = joiner.join((v or "").strip() for v in vals)

    qr = qrcode.QRCode(version=None, error_correction=ERROR_CORRECT_M, box_size=10, border=2)
    qr.add_data(payload)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png")

# ================== Entrypoint ==================
if __name__ == "__main__":
    print(app.url_map)
    port = int(os.environ.get("PORT", "8055"))
    app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False)
