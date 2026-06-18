# -*- coding: utf-8 -*-
# app_unified.py — 單一 Flask（port 5011）
# DropletSchedule(日期/時間修正) + Excel 巨集/兩表導入 + CSV 相容上傳 + WorkOrder 查詢/儲存 + QR

from flask import Flask, request, jsonify, Response, send_file
from flask_cors import CORS
from pathlib import Path
from openpyxl import load_workbook
import sqlite3, csv, io, os, re, json, datetime, time, logging, uuid, shutil, pythoncom, socket
import qrcode
from qrcode.constants import ERROR_CORRECT_M
from win32com.client import Dispatch

# ==================== CONFIG ====================
DB_PATH = os.environ.get("DB_PATH", r"D:\配藥表\資料庫\P01_formualte_schedule.db")
ALLOW_ORIGINS = ["*"]

# DropletSchedule 專用
DROPLET_TABLE = "DropletSchedule"
HEADER_MAP = {
    "Pump": "Pump",
    "Marker": "Marker",
    "凍乾機台": "Lyophilizer",
    "可用凍乾機": "AvailableLyophilizer",
    "數量": "Quantity",
    "配藥同仁": "Preparer",
    "日期": "Date",
    "RD給藥時間": "DrugGivenAt",
    "預計滴定時間": "ExpectedTitrationStart",
    "預計結束": "ExpectedTitrationEnd",
    "工單編號": "WorkOrder",
    "Lot": "Lot",
    "備註": "Remark",
}

# Excel/CSV/模板
APP_CONF = {
    "DB_PATH": DB_PATH,
    "TEMPLATE_XLSM": os.environ.get("TEMPLATE_XLSM", r"D:\配藥表\配藥紀錄\配藥紀錄Temp.xlsm"),
    "UPLOAD_DIR": os.environ.get("UPLOAD_DIR", r"D:\配藥表\_temp"),
    "HEADER_ROW": int(os.environ.get("HEADER_ROW", "8")),
    "COL_BOM_FALLBACK": int(os.environ.get("COL_BOM_FALLBACK", "2")),
    "COL_TOTAL_FALLBACK": int(os.environ.get("COL_TOTAL_FALLBACK", "16")),
    "COL_LOT_FALLBACK": int(os.environ.get("COL_LOT_FALLBACK", "19")),
    "DEFAULT_SERVER_DIR": os.environ.get("DEFAULT_SERVER_DIR", r"\\fls341\Reagent RD\配藥端 -配製紀錄表\\"),
}

# ==================== APP INIT ====================
app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": ALLOW_ORIGINS}})
logging.basicConfig(level=logging.INFO)
log = logging.getLogger("unified")

Path(APP_CONF["UPLOAD_DIR"]).mkdir(parents=True, exist_ok=True)
Path(DB_PATH).parent.mkdir(parents=True, exist_ok=True)

# ==================== DB CONNECT（WAL + busy_timeout + thread-safe） ====================
def connect():
    conn = sqlite3.connect(DB_PATH, timeout=20, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.execute("PRAGMA busy_timeout = 15000;")
    return conn

# ========= 共用小工具 =========
def remove_cn_spaces(s):
    if s is None:
        return ""
    return str(s).replace("\u3000", "").replace("\xa0", "").strip()

def sanitize_for_ident(name):
    s = remove_cn_spaces(name).replace(" ", "")
    s = re.sub(r"[^0-9A-Za-z_]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    if not s:
        return "col"
    if s[0].isdigit():
        s = "_" + s
    return s

def zh_header_to_en(h):
    base = remove_cn_spaces(h).replace(" ", "")
    if base in HEADER_MAP:
        return sanitize_for_ident(HEADER_MAP[base])
    return sanitize_for_ident(base)

def normalize_headers(headers):
    seen, out = set(), []
    for idx, h in enumerate(headers):
        name = zh_header_to_en(h) or f"col_{idx+1}"
        orig = name
        i = 2
        while name in seen:
            name = f"{orig}_{i}"
            i += 1
        seen.add(name)
        out.append(name)
    return out

def normalize_table_name(sheet_name: str) -> str:
    name = re.sub(r'\s+', '_', (sheet_name or "").strip())
    name = re.sub(r'[^\w]', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    return name or "sheet"

def qident(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'

def safe_table_name(name: str) -> str:
    if not name:
        raise ValueError("table is required")
    if not re.fullmatch(r"[A-Za-z0-9_]+", name):
        raise ValueError(f"invalid table name: {name}")
    return name

# ========= DropletSchedule 專用：建表/補欄 =========
def ensure_droplet_table_and_columns(conn, headers):
    cols = normalize_headers(headers)
    cur = conn.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=? LIMIT 1", (DROPLET_TABLE,))
    exists = cur.fetchone() is not None
    if not exists:
        col_defs = ", ".join(f'"{c}" TEXT' for c in cols)
        conn.execute(f'CREATE TABLE "{DROPLET_TABLE}" ({col_defs});')
        conn.commit()
        return cols
    existing = {row[1] for row in conn.execute(f'PRAGMA table_info("{DROPLET_TABLE}")')}
    for c in cols:
        if c not in existing:
            conn.execute(f'ALTER TABLE "{DROPLET_TABLE}" ADD COLUMN "{c}" TEXT;')
    conn.commit()
    return cols

# ========= CSV parse =========
def parse_csv_text(text):
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        return [], []
    headers = [remove_cn_spaces(h) for h in rows[0]]
    return headers, rows[1:]

def parse_csv_file(file_storage):
    text = file_storage.read().decode("utf-8-sig", errors="replace")
    return parse_csv_text(text)

# ========= DropletSchedule：日期/時間修正 =========
def fix_date_field(rows, headers):
    if "Date" not in headers:
        return rows
    idx = headers.index("Date")
    fixed = []
    for r in rows:
        r = list(r)
        val = remove_cn_spaces(r[idx])
        if re.fullmatch(r"\d{4,6}", val or ""):
            try:
                base = datetime.datetime(1899, 12, 30)
                d = base + datetime.timedelta(days=int(val))
                r[idx] = d.strftime("%Y/%m/%d")
            except Exception:
                r[idx] = val
        else:
            parsed = None
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%m/%d/%Y"):
                try:
                    parsed = datetime.datetime.strptime(val, fmt)
                    break
                except Exception:
                    continue
            r[idx] = parsed.strftime("%Y/%m/%d") if parsed else val
        fixed.append(r)
    return fixed

def fix_time_fields(rows, headers):
    time_fields = {"DrugGivenAt", "ExpectedTitrationStart", "ExpectedTitrationEnd"}
    idxs = {h: i for i, h in enumerate(headers) if h in time_fields}
    if not idxs:
        return rows
    fixed = []
    for r in rows:
        r = list(r)
        for h, i in idxs.items():
            val = str(r[i] or "").strip()
            if not val:
                continue
            try:
                f = float(val)
                if 0 <= f < 1:
                    seconds = int(round(f * 86400))
                    hh = seconds // 3600
                    mm = (seconds % 3600) // 60
                    r[i] = f"{hh:02d}:{mm:02d}"
                    continue
            except Exception:
                pass
            try:
                if "1899" in val or "1900" in val:
                    dt = datetime.datetime.fromisoformat(val)
                    r[i] = dt.strftime("%H:%M")
                    continue
            except Exception:
                pass
            parsed = None
            for fmt in ("%H:%M", "%I:%M %p", "%H:%M:%S"):
                try:
                    parsed = datetime.datetime.strptime(val, fmt)
                    break
                except Exception:
                    continue
            if parsed:
                r[i] = parsed.strftime("%H:%M")
        fixed.append(r)
    return fixed

# ==================== 健康檢查 / Root ====================
@app.get("/api/health")
def health():
    return {"ok": True, "db": DB_PATH}

@app.get("/")
def root():
    return "Backend OK. Try /api/health.", 200

# ==================== DropletSchedule：上傳/預覽 ====================
@app.post("/api/upload_droplet_schedule")
def upload_droplet_schedule():
    mode = (request.args.get("mode") or "").lower()
    headers, rows = [], []
    ctype = (request.content_type or "").lower()

    if "multipart/form-data" in ctype:
        if "file" not in request.files:
            return Response("missing file", 400)
        headers, rows = parse_csv_file(request.files["file"])

    elif "application/json" in ctype:
        payload = request.get_json(silent=True) or {}
        headers = [remove_cn_spaces(h) for h in (payload.get("headers") or [])]
        rows = payload.get("rows") or []

    elif "text/csv" in ctype or request.data:
        data = request.get_data()
        if not data:
            return Response("empty request body", 400)
        try:
            text = data.decode("utf-8-sig", errors="replace")
        except UnicodeDecodeError:
            return Response("CSV decoding failed.", 400)
        headers, rows = parse_csv_text(text)

    else:
        return Response("unsupported content-type", 415)

    if not headers:
        return Response("empty headers or data", 400)

    try:
        with connect() as conn:
            if mode == "replace":
                conn.execute(f'DELETE FROM "{DROPLET_TABLE}";')
                conn.commit()

            cols = ensure_droplet_table_and_columns(conn, headers)
            rows = fix_date_field(rows, cols)
            rows = fix_time_fields(rows, cols)

            clean_rows = []
            for r in rows:
                row = [remove_cn_spaces(v) for v in r]
                row += [""] * max(0, len(cols) - len(row))
                clean_rows.append(row[:len(cols)])

            placeholders = ", ".join("?" for _ in cols)
            collist = ", ".join(f'"{c}"' for c in cols)
            sql = f'INSERT INTO "{DROPLET_TABLE}" ({collist}) VALUES ({placeholders})'
            for attempt in range(5):
                try:
                    conn.executemany(sql, clean_rows)
                    conn.commit()
                    break
                except sqlite3.OperationalError as e:
                    if "locked" in str(e).lower() and attempt < 4:
                        time.sleep(0.3)
                        continue
                    raise

        return Response(status=204)

    except sqlite3.Error as e:
        app.logger.error(f"Database Error: {e}")
        return Response(f"Database operation failed: {e}", 500)
    except Exception as e:
        app.logger.error(f"Unexpected Server Error during upload: {e}")
        return Response("Internal Server Error during data processing.", 500)

@app.get("/api/preview")
def preview():
    limit = int(request.args.get("limit", "50"))
    with connect() as conn:
        cur = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?;", (DROPLET_TABLE,))
        if cur.fetchone() is None:
            return jsonify({"columns": [], "rows": []})
        cols = [r[1] for r in conn.execute(f'PRAGMA table_info("{DROPLET_TABLE}")')]
        cur = conn.execute(f'SELECT * FROM "{DROPLET_TABLE}" LIMIT {limit}')
        rows = [dict(r) for r in cur.fetchall()]
        return jsonify({"columns": cols, "rows": rows})

# ==================== Excel 宏 / 兩表導入 DB ====================
TEMP_IN_MAP: dict[str, Path] = {}
TEMP_OUT_MAP: dict[str, Path] = {}

def ensure_table_and_columns(con: sqlite3.Connection, table: str, headers: list[str]) -> None:
    cur = con.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?;", (table,))
    exists = cur.fetchone() is not None
    if not exists:
        cols_def = ", ".join(f'{qident(h)} TEXT' for h in headers)
        con.execute(f'CREATE TABLE {qident(table)} ({cols_def});')
        return
    cur = con.execute(f'PRAGMA table_info({qident(table)});')
    existing = {r[1] for r in cur.fetchall()}
    for h in headers:
        if h not in existing:
            con.execute(f'ALTER TABLE {qident(table)} ADD COLUMN {qident(h)} TEXT;')

def upsert_rows(con: sqlite3.Connection, table: str, headers: list[str], rows: list[tuple], pk_cols: list[str]) -> None:
    cols = ", ".join(qident(h) for h in headers)
    qmarks = ", ".join(["?"] * len(headers))
    if pk_cols:
        upd_cols = [h for h in headers if h not in pk_cols]
        if upd_cols:
            upd = ", ".join(f'{qident(c)}=excluded.{qident(c)}' for c in upd_cols)
            sql = f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks}) ' \
                  f'ON CONFLICT({", ".join(qident(c) for c in pk_cols)}) DO UPDATE SET {upd};'
        else:
            sql = f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks}) ' \
                  f'ON CONFLICT({", ".join(qident(c) for c in pk_cols)}) DO NOTHING;'
    else:
        sql = f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks});'
    for attempt in range(5):
        try:
            con.executemany(sql, rows)
            return
        except sqlite3.OperationalError as e:
            if "locked" in str(e).lower() and attempt < 4:
                time.sleep(0.3); continue
            raise

def read_sheet_to_db(con: sqlite3.Connection, ws, table_name: str, header_row: int) -> dict:
    max_col = ws.max_column or 1
    headers = [str(ws.cell(header_row, c).value or "").strip() for c in range(1, max_col + 1)]
    headers = [h if h else f"col_{i+1}" for i, h in enumerate(headers)]
    rows = []
    r0 = header_row + 1
    max_r = ws.max_row or r0
    for rr in range(r0, max_r + 1):
        row, nonempty = [], False
        for c in range(1, max_col + 1):
            v = ws.cell(rr, c).value
            if v not in (None, ""): nonempty = True
            row.append("" if v is None else str(v))
        if nonempty:
            rows.append(tuple(row))

    ensure_table_and_columns(con, table_name, headers)
    pk_cols = []
    if "工單號碼" in headers and "料號" in headers:
        pk_cols = ["工單號碼", "料號"]
    elif "工單號碼" in headers:
        pk_cols = ["工單號碼"]
    upsert_rows(con, table_name, headers, rows, pk_cols)
    return {"table": table_name, "headers": headers, "rows": len(rows), "pk": pk_cols}

def run_excel_macro(xlsm_path: str, macro_name: str) -> tuple[bool, str | None]:
    try:
        pythoncom.CoInitialize()
        excel = Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            excel.AutomationSecurity = 1  # msoAutomationSecurityLow
        except Exception:
            pass
        wb = excel.Workbooks.Open(xlsm_path)
        try:
            excel.Run(macro_name)
            wb.Save()
            return True, None
        except Exception as e:
            return False, f"Run macro failed: {e}"
        finally:
            try: wb.Close(SaveChanges=False)
            except Exception: pass
            excel.Quit()
    except Exception as e:
        return False, f"Excel automation error: {e}"
    finally:
        try: pythoncom.CoUninitialize()
        except Exception: pass

@app.post("/api/upload_excel")
def api_upload_excel():
    f = request.files.get("file")
    if not f:
        return jsonify(ok=False, message="missing file"), 400

    fname_raw = f.filename or "uploaded.xlsm"
    if "配藥表" not in fname_raw:
        return jsonify(ok=False, message="所選檔名未包含「配藥表」，請重新選擇正確的配藥表 Excel。"), 400

    ext = Path(fname_raw).suffix.lower() or ".xlsm"
    if ext not in (".xlsm", ".xlsx", ".xls"):
        ext = ".xlsm"

    temp_id = uuid.uuid4().hex[:8]
    temp_in_path = Path(APP_CONF["UPLOAD_DIR"]) / f"upload_{temp_id}{ext}"
    f.save(str(temp_in_path))
    TEMP_IN_MAP[temp_id] = temp_in_path

    ok_macro, err_macro = run_excel_macro(str(temp_in_path), "manuinsert.InsertNewByWorkOrderArrays")
    if not ok_macro:
        return jsonify(ok=False, message=f"Excel 宏執行失敗：{err_macro}"), 500

    wb = load_workbook(temp_in_path, keep_vba=True, data_only=True)
    sheets = wb.sheetnames[:2]
    ingested = []

    con = connect()
    try:
        with con:
            con.execute("PRAGMA temp_store=MEMORY;")
            for s in sheets:
                ws = wb[s]
                table = normalize_table_name(s)
                info = read_sheet_to_db(con, ws, table, APP_CONF["HEADER_ROW"])
                ingested.append(info)
    finally:
        con.close()

    return jsonify(ok=True, temp_id=temp_id, filename=fname_raw, sheets=sheets, ingested=ingested)

@app.post("/api/create_record")
def api_create_record():
    data = request.get_json(force=True, silent=True) or {}
    work_order = (data.get("work_order") or "").strip()
    temp_id = (data.get("temp_id") or "").strip()
    table_raw = (data.get("table") or "").strip()
    if not work_order or not temp_id:
        return jsonify(ok=False, message="work_order & temp_id are required"), 400

    in_path = TEMP_IN_MAP.get(temp_id)
    if not in_path or not in_path.exists():
        return jsonify(ok=False, message=f"temp_id not found: {temp_id}"), 404
    wb_in = load_workbook(in_path, keep_vba=True, data_only=True)
    first_sheet = normalize_table_name(wb_in.sheetnames[0]) if wb_in.sheetnames else None
    table = normalize_table_name(table_raw) if table_raw else first_sheet
    if not table:
        return jsonify(ok=False, message="no sheet to use as table"), 400

    sql = f"""
      SELECT 工單號碼 AS work_order, BeadsLot AS bead_lot,
             料號 AS material, 重量紀錄 AS weight, Filler_Lot AS lot_no
      FROM {qident(table)}
      WHERE 工單號碼 = ?
      ORDER BY 料號
    """
    with connect() as conn:
        cur = conn.execute(sql, (work_order,))
        rows = [dict(zip([c[0] for c in cur.description], r)) for r in cur.fetchall()]

    template = APP_CONF["TEMPLATE_XLSM"]
    if not os.path.exists(template):
        return jsonify(ok=False, message=f"template not found: {template}"), 500

    out_path = Path(APP_CONF["UPLOAD_DIR"]) / f"record_{temp_id}.xlsm"
    shutil.copy2(template, out_path)

    wb = load_workbook(out_path, keep_vba=True, data_only=False)
    if "製程紀錄表" not in wb.sheetnames:
        return jsonify(ok=False, message="template has no sheet: 製程紀錄表"), 500
    ws = wb["製程紀錄表"]

    ws["V6"] = work_order

    HEADER_ROW = APP_CONF["HEADER_ROW"]
    COL_BOM_FALLBACK   = APP_CONF["COL_BOM_FALLBACK"]
    COL_TOTAL_FALLBACK = APP_CONF["COL_TOTAL_FALLBACK"]
    COL_LOT_FALLBACK   = APP_CONF["COL_LOT_FALLBACK"]

    def norm(s): return (str(s or "")).strip().lower().replace(" ", "")

    title2col = {}
    max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        title2col[norm(ws.cell(HEADER_ROW, c).value)] = c

    def get_col(names, fallback):
        for n in names:
            col = title2col.get(norm(n))
            if col:
                return col
        return fallback

    col_bom   = get_col(["bom p/n","bompn","料號","bom"], COL_BOM_FALLBACK)
    col_total = get_col(["total qty","totalqty","重量紀錄"], COL_TOTAL_FALLBACK)
    col_lot   = get_col(["lot no.","lotno","filler_lot","lot no","lotno."], COL_LOT_FALLBACK)

    start_row = HEADER_ROW + 1
    end_row = ws.max_row if ws.max_row > start_row else start_row
    if end_row >= start_row:
        ws.delete_rows(start_row, end_row - start_row + 1)

    r = start_row
    for it in rows:
        ws.cell(r, col_bom).value   = it.get("material")
        ws.cell(r, col_total).value = it.get("weight")
        ws.cell(r, col_lot).value   = it.get("lot_no")
        r += 1

    wb.save(out_path)
    TEMP_OUT_MAP[temp_id] = out_path
    return jsonify(ok=True, filled=len(rows), out_path=str(out_path), table=table)

@app.get("/api/template_preview")
def api_template_preview():
    temp_id = (request.args.get("temp_id") or "").strip()
    out_path = TEMP_OUT_MAP.get(temp_id)
    if not out_path or not Path(out_path).exists():
        return jsonify(ok=False, message=f"no generated file for temp_id: {temp_id}"), 404

    wb = load_workbook(out_path, keep_vba=True, data_only=True)
    if "製程紀錄表" not in wb.sheetnames:
        return jsonify(ok=False, message="template has no sheet: 製程紀錄表"), 500
    ws = wb["製程紀錄表"]

    HEADER_ROW = APP_CONF["HEADER_ROW"]
    max_col = ws.max_column or 1
    headers = [str(ws.cell(HEADER_ROW, c).value or "") for c in range(1, max_col + 1)]

    rows = []
    r = HEADER_ROW + 1
    max_r = min(ws.max_row or r, r + 50)
    for rr in range(r, max_r + 1):
        row = []
        for c in range(1, max_col + 1):
            v = ws.cell(rr, c).value
            row.append("" if v is None else str(v))
        rows.append(row)
    return jsonify(ok=True, headers=headers, rows=rows)

@app.get("/api/template_file")
def api_template_file():
    temp_id = (request.args.get("temp_id") or "").strip()
    out_path = TEMP_OUT_MAP.get(temp_id)
    if not out_path or not Path(out_path).exists():
        return jsonify(ok=False, message=f"no generated file for temp_id: {temp_id}"), 404
    return send_file(
        out_path,
        as_attachment=True,
        download_name=Path(out_path).name,
        mimetype="application/vnd.ms-excel.sheet.macroEnabled.12",
    )

@app.post("/api/save_template")
def api_save_template():
    data = request.get_json(force=True, silent=True) or {}
    temp_id = (data.get("temp_id") or "").strip()
    work_order = (data.get("work_order") or "").strip()
    server_dir = (data.get("server_dir") or "").strip()
    filename   = (data.get("filename") or "").strip()

    out_src = TEMP_OUT_MAP.get(temp_id)
    if not out_src or not Path(out_src).exists():
        return jsonify(ok=False, message=f"no generated file for temp_id: {temp_id}"), 404

    if not server_dir or not filename:
        return jsonify(ok=False, message="server_dir, filename are required"), 400
    if not filename.endswith(".xlsm"):
        filename += ".xlsm"

    if work_order:
        wb = load_workbook(out_src, keep_vba=True, data_only=False)
        if "製程紀錄表" in wb.sheetnames:
            wb["製程紀錄表"]["V6"] = work_order
            wb.save(out_src)

    out_dir = Path(server_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / filename
    shutil.copy2(out_src, out_path)
    return jsonify(ok=True, saved_path=str(out_path))

# ==================== CSV 相容上傳 (raw/form-data) ====================
def _parse_csv_text(text: str):
    if text and text[:1] == '\ufeff':
        text = text.lstrip('\ufeff')
    reader = csv.reader(io.StringIO(text))
    try:
        headers = next(reader)
    except StopIteration:
        headers = []
    headers = [h.strip() if h else f"col_{i+1}" for i, h in enumerate(headers)]
    rows = []
    for r in reader:
        if len(r) < len(headers): r = r + [""] * (len(headers) - len(r))
        elif len(r) > len(headers): r = r[:len(headers)]
        rows.append(tuple(r))
    return headers, rows

def ensure_unique_index(con: sqlite3.Connection, table: str, pk_cols: list[str]) -> None:
    if not pk_cols:
        return
    idx = "ux_" + re.sub(r'[^A-Za-z0-9_]', '_', table) + "__" + "__".join(
        re.sub(r'[^A-Za-z0-9_]', '_', c) for c in pk_cols
    )
    cols = ", ".join(qident(c) for c in pk_cols)
    con.execute(f'CREATE UNIQUE INDEX IF NOT EXISTS {qident(idx)} ON {qident(table)} ({cols});')

def _get_mode() -> str:
    mode = (request.form.get("mode") or request.args.get("mode") or "replace").strip().lower()
    return mode if mode in ("append", "replace", "upsert") else "replace"

def _want_csv_response() -> bool:
    return (request.args.get("return") or "").strip().lower() == "csv"

def _insert_rows(con: sqlite3.Connection, table: str, headers: list[str], rows: list[tuple]) -> int:
    if not rows:
        return 0
    cols = ", ".join(qident(h) for h in headers)
    qmarks = ", ".join(["?"] * len(headers))
    con.executemany(f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks});', rows)
    return len(rows)

def _replace_rows(con: sqlite3.Connection, table: str, headers: list[str], rows: list[tuple]) -> int:
    con.execute(f'DELETE FROM {qident(table)};')
    return _insert_rows(con, table, headers, rows)

def _handle_csv_upload(table: str, headers: list[str], rows: list[tuple], pk_cols: list[str], mode: str):
    con = connect()
    try:
        with con:
            con.execute("PRAGMA temp_store=MEMORY;")
            ensure_table_and_columns(con, table, headers)
            if not pk_cols:
                if "工單號碼" in headers and "料號" in headers:
                    pk_cols = ["工單號碼", "料號"]
                elif "工單號碼" in headers:
                    pk_cols = ["工單號碼"]
            if pk_cols:
                ensure_unique_index(con, table, pk_cols)
            if mode == "append":
                if pk_cols:
                    upsert_rows(con, table, headers, rows, pk_cols)
                else:
                    _insert_rows(con, table, headers, rows)
            elif mode == "upsert":
                if not pk_cols:
                    _insert_rows(con, table, headers, rows)
                else:
                    upsert_rows(con, table, headers, rows, pk_cols)
            else:  # replace
                if pk_cols:
                    upsert_rows(con, table, headers, rows, pk_cols)
                else:
                    _replace_rows(con, table, headers, rows)
    finally:
        con.close()

@app.post("/api/upload_csv_raw")
@app.post("/upload_csv_raw")
def api_upload_csv_raw():
    table = (request.args.get("table") or "").strip()
    if not table:
        return jsonify(ok=False, message="missing table"), 400
    table = normalize_table_name(table)
    try:
        table = safe_table_name(table)
    except ValueError as e:
        return jsonify(ok=False, message=str(e)), 400

    pk_raw = (request.args.get("pk") or "").strip()
    mode = _get_mode()
    pk_cols = [c.strip() for c in pk_raw.split(",") if c.strip()]

    text = request.get_data(cache=False, as_text=True)
    if not text:
        return jsonify(ok=False, message="empty body"), 400

    headers, rows = _parse_csv_text(text)
    try:
        _handle_csv_upload(table, headers, rows, pk_cols, mode)
    except Exception as e:
        app.logger.exception("upload_csv_raw failed")
        if _want_csv_response():
            return Response("status,FAIL\n", status=500, mimetype="text/csv; charset=utf-8")
        return jsonify(ok=False, message=str(e)), 500

    if _want_csv_response():
        return Response("status,OK\n", mimetype="text/csv; charset=utf-8")
    return jsonify(ok=True)

@app.post("/api/upload_csv")
@app.post("/upload_csv")
def api_upload_csv():
    table = (request.form.get("table") or request.args.get("table") or "").strip()
    if not table:
        return jsonify(ok=False, message="missing table"), 400
    table = normalize_table_name(table)
    try:
        table = safe_table_name(table)
    except ValueError as e:
        return jsonify(ok=False, message=str(e)), 400

    pk_raw = (request.form.get("pk") or request.args.get("pk") or "").strip()
    mode = _get_mode()
    pk_cols = [c.strip() for c in pk_raw.split(",") if c.strip()]

    f = request.files.get("file") or request.files.get("csv")
    if not f:
        return jsonify(ok=False, message="missing file"), 400

    raw = f.read()
    try:
        text = raw.decode("utf-8-sig", errors="replace")
    except Exception:
        text = raw.decode(errors="replace")

    headers, rows = _parse_csv_text(text)
    try:
        _handle_csv_upload(table, headers, rows, pk_cols, mode)
    except Exception as e:
        app.logger.exception("upload_csv failed")
        if _want_csv_response():
            return Response("status,FAIL\n", status=500, mimetype="text/csv; charset=utf-8")
        return jsonify(ok=False, message=str(e)), 500

    if _want_csv_response():
        return Response("status,OK\n", mimetype="text/csv; charset=utf-8")
    return jsonify(ok=True)

# ==================== QR 產生（從 xlsm cells） ====================
@app.post("/api/qr_png_from_cells")
def api_qr_png_from_cells():
    data = request.get_json(force=True)
    cells = data.get("cells") or []
    joiner = (data.get("joiner") or "|")
    file_path = (data.get("file_path") or "").strip()
    temp_id = (data.get("temp_id") or "").strip()

    if not (file_path or temp_id):
        return jsonify(ok=False, message="file_path or temp_id is required"), 400
    if not cells:
        return jsonify(ok=False, message="cells is required"), 400

    xlsm_path = file_path or str(TEMP_OUT_MAP.get(temp_id) or "")
    if not xlsm_path or not os.path.exists(xlsm_path):
        return jsonify(ok=False, message=f"file not found: {xlsm_path}"), 404

    wb = load_workbook(xlsm_path, data_only=True, keep_vba=True)
    ws = wb.active
    vals = []
    for addr in cells:
        if "!" in addr:
            sheet, cell = addr.split("!", 1)
            vals.append("" if wb[sheet][cell].value is None else str(wb[sheet][cell].value))
        else:
            vals.append("" if ws[addr].value is None else str(ws[addr].value))
    payload = joiner.join((v or "").strip() for v in vals)

    qr = qrcode.QRCode(version=None, error_correction=ERROR_CORRECT_M, box_size=10, border=2)
    qr.add_data(payload)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png")

# ==================== WorkOrder 查詢 / 儲存 ====================

def query_db(sql, args=()):
    with connect() as conn:
        cur = conn.execute(sql, args)
        return [dict(row) for row in cur.fetchall()]

def safe_float(v):
    try:
        return float(v) if v not in (None, "") else 0.0
    except:
        return 0.0

def safe_int(v):
    try:
        return int(float(v))
    except:
        return 0

def parse_range(text):
    if not text:
        return None
    m = re.match(r"([\d.]+)\s*[-~]\s*([\d.]+)", str(text))
    if not m:
        return None
    return float(m.group(1)), float(m.group(2))

def check_in_range(value, text_range):
    if not text_range:
        return None
    lo, hi = text_range
    return lo <= value <= hi

def norm_name(s: str) -> str:
    if not s:
        return ""
    s = str(s).strip()
    s = re.sub(r"[-_].*$", "", s)
    return s.strip().upper()

import sqlite3

def find_table_contains_workorder(conn: sqlite3.Connection, work_order: str):
    """回傳第一個包含該工單的 table 名稱，找不到回傳 None"""

    # 1. 清理工單號碼參數，並開始追蹤
    work_order = work_order.strip()
    print(f"\n--- 🐛 追蹤: 尋找工單 '{work_order}' (優化版) ---")

    # 2. 篩選資料表：只檢查名稱中包含 '571' 的資料表 (提高效率)
    cur = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE '%571%'")
    tables = [r[0] for r in cur.fetchall()]
    print(f"📚 找到包含 '571' 的資料表總數: {len(tables)}")

    # 設置模糊匹配模式
    like_pattern = f"%{work_order}%"

    for t in tables:
        print(f"--- 檢查資料表: {t} ---")
        try:
            # 3. 檢查欄位名稱（確保該欄位存在，避免 SQL 錯誤）
            # 由於這是優化版，我們假設所有 '571' 表都有 '工單號碼' 欄位，但仍應加 TRY-EXCEPT。

            # 4. 執行寬鬆匹配查詢：
            #    - 使用 TRIM(工單號碼) 移除資料庫中欄位的前後空格。
            #    - 使用 LIKE ? (參數為 %工單號碼%) 進行模糊匹配，解決隱藏字元問題。
            sql_like = f"SELECT 1 FROM '{t}' WHERE TRIM(工單號碼) LIKE ? LIMIT 1"
            rows = conn.execute(sql_like, (like_pattern,)).fetchall()

            if rows:
                print(f"✅ 找到工單 {work_order} 位於表 {t}，成功。")
                print(f"--- 🐛 追蹤結束 ---")
                return t

        except sqlite3.OperationalError as e:
            # 如果資料表根本沒有 '工單號碼' 欄位，會在這裡報錯
            print(f"❌ {t}: 查詢發生 SQLite 錯誤 (可能缺少 '工單號碼' 欄位): {e}")
        except Exception as e:
            print(f"❌ {t}: 發生未預期的錯誤: {e}")

    print(f"❌ 總結：未在任何包含 '571' 的表找到工單 {work_order}")
    print(f"--- 🐛 追蹤結束 ---")
    return None

# ==================== WorkOrder 查詢 / 儲存 ====================

# ... (query_db, safe_float, safe_int, parse_range, check_in_range, norm_name, find_table_contains_workorder... 這些都保留不變) ...


def check_suspension(v):
    """(懸浮物) '混濁' = True, 其他 = False"""
    result = "混濁" in str(v or "")
    print(f"  check_suspension({v!r}) → {result}")
    return result

def check_yes(v):
    """'Yes' (不區分大小寫) = True, 其他 = False"""
    result = str(v or "").strip().lower() == "yes"
    print(f"  check_yes({v!r}) → {result}")
    return result

def check_not_room_temp(v):
    """
    '室溫' 或空值 = False
    '4℃' 或其他非空值 = True
    """
    val = str(v or "").strip()
    # ✅ 修正：空值和 '室溫' 都返回 False
    result = bool(val) and val != "室溫"
    print(f"  check_not_room_temp({v!r}) → val={val!r} → {result}")
    return result

def check_not_no(v):
    """
    'No' (不區分大小寫) 或空值 = False
    'Yes' 或其他非空值 = True
    """
    val = str(v or "").strip()
    # ✅ 修正：空值和 'No' 都返回 False
    result = bool(val) and val.lower() != "no"
    print(f"  check_not_no({v!r}) → val={val!r} → {result}")
    return result


@app.get("/api/get_workorder")
def api_get_workorder():
    work_order = request.args.get("work_order")
    if not work_order:
        return jsonify({"error": "缺少工單號碼參數"}), 400

    print(f"\n{'='*60}")
    print(f"🔍 收到查詢工單請求: {work_order}")
    print(f"{'='*60}\n")

    try:
        with connect() as conn:
            target_table = find_table_contains_workorder(conn, work_order)
            if not target_table:
                return jsonify({"message": f"查無工單 {work_order}"}), 404

            cur = conn.execute(
                f"SELECT * FROM '{target_table}' WHERE TRIM(工單號碼) LIKE ?",
                (f"%{work_order.strip()}%",),
            )
            cols = [d[0] for d in cur.description]
            rows = cur.fetchall()
            result_rows = [dict(zip(cols, row)) | {"來源表": target_table} for row in rows]

            if not result_rows:
                return jsonify({"message": f"查無工單 {work_order} 的詳細記錄"}), 404

            first = result_rows[0]

            # DropletSchedule 和 pump 邏輯（保持不變）
            dispose_lots, maker_name, product_quantity = [], None, 0
            cur = conn.execute(
                "SELECT Lot, Pump, Lyophilizer, Marker, Quantity FROM DropletSchedule WHERE TRIM(WorkOrder) LIKE ?",
                (f"%{work_order.strip()}%",),
            )
            for r in cur.fetchall():
                r = dict(r)
                maker_name = r.get("Marker") or maker_name
                product_quantity += safe_int(r.get("Quantity"))
                dispose_lots.append({
                    "id": r.get("Lot"),
                    "port": r.get("Pump"),
                    "freezeDry": r.get("Lyophilizer"),
                    "pump": None,
                })

            pump_ids = []
            if maker_name:
                cur2 = conn.execute("SELECT * FROM 'pump No.'")
                for row in cur2.fetchall():
                    row = dict(row)
                    for k in ["可滴定之試劑-1", "可滴定之試劑-2", "可滴定之試劑-3", "可滴定之試劑-4"]:
                        val = row.get(k)
                        if val and maker_name.lower() in str(val).strip().lower():
                            pump_ids.append(row.get("Pump編號"))
                            break
            if pump_ids:
                for i, lot in enumerate(dispose_lots):
                    lot["pump"] = pump_ids[i % len(pump_ids)]

            # ✅ 取得產品型號（從 table 名稱提取）
            match_571 = re.search(r"571\d{5,}", target_table or "")
            product_model = match_571.group(0) if match_571 else None
            
            print(f"📍 工單表: {target_table}")
            print(f"📍 產品型號: {product_model}")
            print(f"📍 Marker 名稱: {maker_name}")

            # ✅ 1. 查詢 [Liquid form QC] - 使用標準化名稱
            base_name = norm_name(maker_name or "")
            print(f"📍 標準化名稱: {base_name}")
            
            print(f"\n--- 查詢 [Liquid form QC] ---")
            qc_all = query_db("""
                SELECT Name,[L1-OD],[L2-OD],[L1-起始OD],[L2-起始OD],[懸浮物]
                FROM [Liquid form QC]
            """)
            
            selected_qc = None
            for row in qc_all:
                name = norm_name(row.get("Name", ""))
                if name == base_name:
                    selected_qc = row
                    print(f"✅ 找到 QC 資料")
                    break
            
            if not selected_qc:
                selected_qc = {
                    "L1-OD": None, "L2-OD": None, 
                    "L1-起始OD": None, "L2-起始OD": None, 
                    "懸浮物": None
                }
                print(f"⚠️ 未找到 QC 資料")

            # ✅ 2. 查詢滴定條件 - 使用產品型號（PN）
            print(f"\n--- 查詢滴定條件 ---")
            
            selected_titration = {}
            
            # 嘗試不同的表名
            for table_name in ["適定條件", "滴定條件"]:
                try:
                    # 檢查表是否存在
                    cur = conn.execute(
                        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                        (table_name,)
                    )
                    if not cur.fetchone():
                        print(f"  ⚠️ 表 [{table_name}] 不存在")
                        continue
                    
                    print(f"  🔍 查詢表: [{table_name}]")
                    
                    # 檢查欄位
                    cur = conn.execute(f"PRAGMA table_info('{table_name}')")
                    table_cols = [r[1] for r in cur.fetchall()]
                    print(f"    欄位: {table_cols}")
                    
                    # ✅ 優先使用產品型號（PN）查詢
                    if "PN" in table_cols and product_model:
                        print(f"    嘗試用 PN = '{product_model}' 查詢...")
                        result = query_db(f"""
                            SELECT * FROM [{table_name}] WHERE PN = ?
                        """, (product_model,))
                        
                        if result:
                            selected_titration = result[0]
                            print(f"    ✅ 找到資料！")
                            print(f"    完整資料: {selected_titration}")
                            break
                    
                    # 如果 PN 查不到，再嘗試用 Name
                    if "Name" in table_cols and base_name:
                        print(f"    嘗試用 Name = '{base_name}' 查詢...")
                        result = query_db(f"""
                            SELECT * FROM [{table_name}] WHERE Name = ?
                        """, (base_name,))
                        
                        if result:
                            selected_titration = result[0]
                            print(f"    ✅ 找到資料！")
                            print(f"    完整資料: {selected_titration}")
                            break
                    
                except Exception as e:
                    print(f"  ❌ 查詢表 [{table_name}] 失敗: {e}")
                    continue
            
            if not selected_titration:
                print(f"⚠️ 未找到滴定條件")

            # ✅ 3. 建立 confirm 資料
            print(f"\n{'='*60}")
            print(f"開始建立 confirm_data")
            print(f"{'='*60}")
            
            print(f"\n📋 原始資料:")
            print(f"  懸浮物: {selected_qc.get('懸浮物')!r}")
            print(f"  儲存時避光: {selected_titration.get('儲存時避光')!r}")
            print(f"  儲存時冰浴: {selected_titration.get('儲存時冰浴')!r}")
            print(f"  滴定時避光: {selected_titration.get('滴定時避光')!r}")
            print(f"  滴定時冰浴: {selected_titration.get('滴定時冰浴')!r}")
            print(f"  滴定時攪拌: {selected_titration.get('滴定時攪拌')!r}")
            print(f"  滴定_Mixing: {selected_titration.get('滴定_Mixing')!r}")
            
            print(f"\n🔄 轉換過程:")
            
            suspension_val = check_suspension(selected_qc.get("懸浮物"))
            storeLight_val = check_yes(selected_titration.get("儲存時避光"))
            storeIce_val = check_not_room_temp(selected_titration.get("儲存時冰浴"))
            dyeing_val = check_yes(selected_titration.get("滴定時避光"))
            washing_val = check_yes(selected_titration.get("滴定時冰浴"))
            
            stir_raw = selected_titration.get("滴定時攪拌") or selected_titration.get("滴定_Mixing")
            stir_val = check_not_no(stir_raw)
            
            confirm_data = {
                "suspension": suspension_val,
                "storeLight": storeLight_val,
                "storeIce": storeIce_val,
                "dyeing": dyeing_val,
                "washing": washing_val,
                "stir": stir_val,
            }

            print(f"\n✅ 最終 confirm_data: {confirm_data}")
            print(f"{'='*60}\n")

            # 其餘邏輯保持不變
            L1OD = safe_float(first.get("L1 OD") or first.get("L1OD"))
            L2OD = safe_float(first.get("L2 OD") or first.get("L2OD"))
            L1StartOD = safe_float(first.get("L1 起始 OD") or first.get("起始L1OD"))
            L2StartOD = safe_float(first.get("L2 起始 OD") or first.get("起始L2OD"))

            qc_map = {"L1OD":"L1-OD","L2OD":"L2-OD","L1StartOD":"L1-起始OD","L2StartOD":"L2-起始OD"}
            values = {"L1OD": L1OD, "L2OD": L2OD, "L1StartOD": L1StartOD, "L2StartOD": L2StartOD}
            check_results = {}
            for k, qc_col in qc_map.items():
                rng = parse_range(selected_qc.get(qc_col))
                val = values[k]
                check_results[k] = {
                    "value": val, 
                    "qc_range": selected_qc.get(qc_col), 
                    "pass": check_in_range(val, rng)
                }

            data = {
                "workOrderNo": work_order,
                "productModel": product_model or target_table,
                "markerName": maker_name or (target_table or "").split("_")[0],
                "productQuantity": product_quantity,
                "date": first.get("試劑配製日期", ""),
                "beads": [
                    {
                        "beadName": r.get("化學品名", ""),
                        "beadPN": r.get("料號", ""),
                        "unit": "g",
                        "qtyPerBead": r.get("總重量", 0),
                        "totalQty": r.get("重量紀錄", 0),
                        "lotNo": r.get("Filler_Lot", ""),
                    }
                    for r in result_rows
                ],
                "reagent": {
                    "preparedBy": first.get("配製人員", ""),
                    "confirm": confirm_data,
                },
                "bufferBase": {
                    "L1OD": L1OD, "L2OD": L2OD, 
                    "L1StartOD": L1StartOD, "L2StartOD": L2StartOD,
                },
                "qcRanges": {
                    "L1-OD": selected_qc.get("L1-OD"),
                    "L2-OD": selected_qc.get("L2-OD"),
                    "L1-起始OD": selected_qc.get("L1-起始OD"),
                    "L2-起始OD": selected_qc.get("L2-起始OD"),
                },
                "qcCheckResult": check_results,
                "disposeLots": dispose_lots,
            }
            return jsonify(data)

    except Exception as e:
        print("❌ 查詢錯誤:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.post("/api/save_workorder")
def save_workorder():
    try:
        payload = request.get_json(force=True)
        work_order = payload.get("workOrderNo")
        if not work_order:
            return jsonify({"error": "workOrderNo missing"}), 400

        print(f"💾 收到儲存請求: {work_order}")

        beads = payload.get("beads", [])
        reagent = payload.get("reagent", {})
        buffer = payload.get("bufferBase", {})
        confirm = reagent.get("confirm", {})  # ✅ 取得 confirm 資料

        with connect() as conn:
            # 找到真的含該工單的 table
            table = find_table_contains_workorder(conn, work_order)
            if not table:
                return jsonify({"error": f"No table contains work order {work_order}"}), 404

            updated = 0
            # 確保有 備註 欄（若有 remark 要寫）
            need_remark = any("remark" in (b or {}) for b in beads)
            if need_remark:
                cur = conn.execute(f"PRAGMA table_info('{table}')")
                cols = [r[1] for r in cur.fetchall()]
                if "備註" not in cols:
                    conn.execute(f"ALTER TABLE '{table}' ADD COLUMN 備註 TEXT")
                    conn.commit()

            for bead in beads:
                remark = bead.get("remark")
                sql = f"""
                    UPDATE '{table}'
                    SET L1OD=?, L2OD=?, 起始L1OD=?, 起始L2OD=?, 配製人員=?, 總重量=?, 重量紀錄=?"""
                args = [
                    buffer.get("L1OD", ""),
                    buffer.get("L2OD", ""),
                    buffer.get("L1StartOD", ""),
                    buffer.get("L2StartOD", ""),
                    reagent.get("preparedBy", ""),
                    bead.get("totalQty", 0),
                    bead.get("qtyPerBead", 0),
                ]
                if remark is not None:
                    sql += ", 備註=?"
                    args.append(remark)
                sql += " WHERE 工單號碼=? AND 料號=?"
                args.extend([work_order, bead.get("beadPN", "")])

                conn.execute(sql, args)
                conn.commit()
                updated += conn.total_changes

        return jsonify({"message": f"✅ 已更新 {updated} 筆資料", "table": table})

    except Exception as e:
        print("❌ save_workorder error:", e)
        return jsonify({"error": str(e)}), 500

# ==================== MAIN ====================
if __name__ == "__main__":
    local_ip = socket.gethostbyname(socket.gethostname())
    print("🚀 Flask 啟動成功！")
    print(f"🌐 同網段裝置可訪問： http://{local_ip}:5011")
    print(f"🧠 本機： http://localhost:5011")
    port = 5011  # ✅ 單一 port
    app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False)
