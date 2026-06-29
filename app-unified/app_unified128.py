# -*- coding: utf-8 -*-
# app_unified.py — 單一 Flask（port 5011）
# DropletSchedule(日期/時間修正) + Excel 巨集/兩表導入 + CSV 相容上傳 + WorkOrder 查詢/儲存 + QR

from flask import Flask, request, jsonify, Response, send_file
from flask_cors import CORS
from pathlib import Path
from openpyxl import load_workbook
import sqlite3, csv, io, os, re, json, datetime, time, logging, uuid, shutil, socket
import qrcode
from qrcode.constants import ERROR_CORRECT_M
# from win32com.client import Dispatch  # disabled on Linux

# ==================== CONFIG ====================
DB_PATH = os.environ.get("DB_PATH", "/opt/beadsops/data/P01_formualte_schedule.db")
ALLOW_ORIGINS = ["*"]

# DropletSchedule 專用
DROPLET_TABLE = "DropletSchedule"
HEADER_MAP = {
    "Pump": "Pump",
    "Marker": "Marker",
    "凍乾機台": "Lyophilizer",
    "可用凍乾機": "AvailableLyophilizer",
    "數量": "Quantity",
    "配藥同仁": "Preparer",
    "日期": "Date",
    "RD給藥時間": "DrugGivenAt",
    "預計滴定時間": "ExpectedTitrationStart",
    "預計結束": "ExpectedTitrationEnd",
    "工單編號": "WorkOrder",
    "Lot": "Lot",
    "備註": "Remark",
}

# Excel/CSV/模板
APP_CONF = {
    "DB_PATH": DB_PATH,
    "TEMPLATE_XLSM": os.environ.get("TEMPLATE_XLSM", r"D:\配藥表\配藥紀錄\配藥紀錄Temp.xlsm"),
    "UPLOAD_DIR": os.environ.get("UPLOAD_DIR", r"D:\配藥表\_temp"),
    "HEADER_ROW": int(os.environ.get("HEADER_ROW", "8")),
    "COL_BOM_FALLBACK": int(os.environ.get("COL_BOM_FALLBACK", "2")),
    "COL_TOTAL_FALLBACK": int(os.environ.get("COL_TOTAL_FALLBACK", "16")),
    "COL_LOT_FALLBACK": int(os.environ.get("COL_LOT_FALLBACK", "19")),
    "DEFAULT_SERVER_DIR": os.environ.get("DEFAULT_SERVER_DIR", r"\\fls341\Reagent RD\配藥端 -配製紀錄表\\"),
}

# ==================== APP INIT ====================
app = Flask(__name__)
CORS(app, resources={r"/api/*": {
    "origins": "*",
    "methods": ["GET", "POST", "OPTIONS", "PUT", "DELETE"],
    "allow_headers": ["Content-Type", "Authorization"]
}})
logging.basicConfig(level=logging.INFO)
log = logging.getLogger("unified")

Path(APP_CONF["UPLOAD_DIR"]).mkdir(parents=True, exist_ok=True)
Path(DB_PATH).parent.mkdir(parents=True, exist_ok=True)

# ==================== DB CONNECT（WAL + busy_timeout + thread-safe） ====================
def connect():
    conn = sqlite3.connect(DB_PATH, timeout=20, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.execute("PRAGMA busy_timeout = 15000;")
    return conn

# ========= 共用小工具 =========
def remove_cn_spaces(s):
    if s is None:
        return ""
    return str(s).replace("\u3000", "").replace("\xa0", "").strip()

def sanitize_for_ident(name):
    s = remove_cn_spaces(name).replace(" ", "")
    s = re.sub(r"[^0-9A-Za-z_]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    if not s:
        return "col"
    if s[0].isdigit():
        s = "_" + s
    return s

def zh_header_to_en(h):
    base = remove_cn_spaces(h).replace(" ", "")
    if base in HEADER_MAP:
        return sanitize_for_ident(HEADER_MAP[base])
    return sanitize_for_ident(base)

def normalize_headers(headers):
    seen, out = set(), []
    for idx, h in enumerate(headers):
        name = zh_header_to_en(h) or f"col_{idx+1}"
        orig = name
        i = 2
        while name in seen:
            name = f"{orig}_{i}"
            i += 1
        seen.add(name)
        out.append(name)
    return out

def normalize_table_name(sheet_name: str) -> str:
    name = re.sub(r'\s+', '_', (sheet_name or "").strip())
    name = re.sub(r'[^\w]', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    return name or "sheet"

def qident(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'

def safe_table_name(name: str) -> str:
    if not name:
        raise ValueError("table is required")
    if not re.fullmatch(r"[A-Za-z0-9_]+", name):
        raise ValueError(f"invalid table name: {name}")
    return name

# ========= DropletSchedule 專用：建表/補欄 =========
def ensure_droplet_table_and_columns(conn, headers):
    cols = normalize_headers(headers)
    cur = conn.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=? LIMIT 1", (DROPLET_TABLE,))
    exists = cur.fetchone() is not None
    if not exists:
        col_defs = ", ".join(f'"{c}" TEXT' for c in cols)
        conn.execute(f'CREATE TABLE "{DROPLET_TABLE}" ({col_defs});')
        conn.commit()
        return cols
    existing = {row[1] for row in conn.execute(f'PRAGMA table_info("{DROPLET_TABLE}")')}
    for c in cols:
        if c not in existing:
            conn.execute(f'ALTER TABLE "{DROPLET_TABLE}" ADD COLUMN "{c}" TEXT;')
    conn.commit()
    return cols

# ========= CSV parse =========
def parse_csv_text(text):
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        return [], []
    headers = [remove_cn_spaces(h) for h in rows[0]]
    return headers, rows[1:]

def parse_csv_file(file_storage):
    text = file_storage.read().decode("utf-8-sig", errors="replace")
    return parse_csv_text(text)

# ========= DropletSchedule：日期/時間修正 =========
def fix_date_field(rows, headers):
    if "Date" not in headers:
        return rows
    idx = headers.index("Date")
    fixed = []
    for r in rows:
        r = list(r)
        val = remove_cn_spaces(r[idx])
        if re.fullmatch(r"\d{4,6}", val or ""):
            try:
                base = datetime.datetime(1899, 12, 30)
                d = base + datetime.timedelta(days=int(val))
                r[idx] = d.strftime("%Y/%m/%d")
            except Exception:
                r[idx] = val
        else:
            parsed = None
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%m/%d/%Y"):
                try:
                    parsed = datetime.datetime.strptime(val, fmt)
                    break
                except Exception:
                    continue
            r[idx] = parsed.strftime("%Y/%m/%d") if parsed else val
        fixed.append(r)
    return fixed

def fix_time_fields(rows, headers):
    time_fields = {"DrugGivenAt", "ExpectedTitrationStart", "ExpectedTitrationEnd"}
    idxs = {h: i for i, h in enumerate(headers) if h in time_fields}
    if not idxs:
        return rows
    fixed = []
    for r in rows:
        r = list(r)
        for h, i in idxs.items():
            val = str(r[i] or "").strip()
            if not val:
                continue
            try:
                f = float(val)
                if 0 <= f < 1:
                    seconds = int(round(f * 86400))
                    hh = seconds // 3600
                    mm = (seconds % 3600) // 60
                    r[i] = f"{hh:02d}:{mm:02d}"
                    continue
            except Exception:
                pass
            try:
                if "1899" in val or "1900" in val:
                    dt = datetime.datetime.fromisoformat(val)
                    r[i] = dt.strftime("%H:%M")
                    continue
            except Exception:
                pass
            parsed = None
            for fmt in ("%H:%M", "%I:%M %p", "%H:%M:%S"):
                try:
                    parsed = datetime.datetime.strptime(val, fmt)
                    break
                except Exception:
                    continue
            if parsed:
                r[i] = parsed.strftime("%H:%M")
        fixed.append(r)
    return fixed
# ==========================================
# Sync Trigger (Flask side)
# ==========================================

import subprocess
import sys
from pathlib import Path
import time
import uuid

SYNC_SCRIPT = Path("/opt/beadsops/dropfreeze/sync_droplet_record.py")
SYNC_LOCK = Path("/tmp/droplet_sync.lock")

SYNC_LOCK_TTL = 120  # seconds

def trigger_droplet_sync_safe():
    """
    防止短時間重複觸發，啟動 sync 程式。
    Lock 超過 TTL 自動過期，避免卡死。
    """
    if SYNC_LOCK.exists():
        age = time.time() - SYNC_LOCK.stat().st_mtime
        if age < SYNC_LOCK_TTL:
            return
        SYNC_LOCK.unlink(missing_ok=True)

    try:
        SYNC_LOCK.parent.mkdir(parents=True, exist_ok=True)
        lock_token = f"{time.time()}|{uuid.uuid4()}"
        SYNC_LOCK.write_text(lock_token)

        subprocess.Popen(
            [sys.executable, str(SYNC_SCRIPT)],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
    except Exception as e:
        print(f"⚠️ Sync trigger failed: {e}")

# ==================== 健康檢查 / Root ====================
@app.get("/api/health")
def health():
    return {"ok": True, "db": DB_PATH}

@app.get("/")
def root():
    return "Backend OK. Try /api/health.", 200

# ==================== DropletSchedule：上傳/預覽 ====================
@app.post("/api/upload_droplet_schedule")
def upload_droplet_schedule():
    # 1. 接收與解析請求 (Parsing)
    headers, rows = [], []
    ctype = (request.content_type or "").lower()

    if "multipart/form-data" in ctype:
        if "file" not in request.files:
            return Response("missing file", 400)
        headers, rows = parse_csv_file(request.files["file"])

    elif "application/json" in ctype:
        payload = request.get_json(silent=True) or {}
        headers = [remove_cn_spaces(h) for h in (payload.get("headers") or [])]
        rows = payload.get("rows") or []

    elif "text/csv" in ctype or request.data:
        data = request.get_data()
        if not data:
            return Response("empty request body", 400)
        try:
            text = data.decode("utf-8-sig", errors="replace")
        except UnicodeDecodeError:
            return Response("CSV decoding failed.", 400)
        headers, rows = parse_csv_text(text)

    else:
        return Response("unsupported content-type", 415)

    if not headers or not rows:
        return Response("empty headers or data", 400)

    try:
        with connect() as conn:
            # 2. 確保資料表結構 (Schema)
            # 這一步會回傳標準化後的欄位清單 (cols)
            cols = ensure_droplet_table_and_columns(conn, headers)

            # 設定識別日期的欄位名稱 (請依據您資料庫實際欄位修改，例如 "Date" 或 "日期")
            target_date_col = "Date"

            # 3. 資料預處理 (Data Preparation in Python)
            # 先在記憶體中處理好，減少資料庫鎖定時間
            try:
                # 假設這兩個函數會回傳處理後的 list
                rows = fix_date_field(rows, cols)
                rows = fix_time_fields(rows, cols)

                # 找出 Date 欄位在 cols 中的索引位置
                if target_date_col not in cols:
                     return Response(f"Missing required column: {target_date_col}", 400)

                date_col_index = cols.index(target_date_col)

                clean_rows = []
                involved_dates = set() # 用來儲存 CSV 中出現過的日期

                col_count = len(cols)
                for r in rows:
                    # 轉為 list 並清洗字串
                    row_data = [remove_cn_spaces(str(v)) for v in r]

                    # 補齊或截斷欄位長度
                    if len(row_data) < col_count:
                        row_data += [""] * (col_count - len(row_data))
                    final_row = row_data[:col_count]

                    clean_rows.append(final_row)

                    # 記錄這筆資料的日期
                    date_val = final_row[date_col_index]
                    if date_val:
                        involved_dates.add(date_val)

                if not involved_dates:
                    return Response("No valid dates found in upload data.", 400)

            except Exception as e:
                app.logger.error(f"Data processing error: {e}")
                return Response(f"Data processing failed: {e}", 400)

            # 4. 資料庫寫入 (Transaction)
            # 邏輯：只刪除 involved_dates 裡的那些日期的舊資料，然後插入新資料

            # 準備 SQL 語句
            placeholders_dates = ", ".join("?" for _ in involved_dates)
            delete_sql = f'DELETE FROM "{DROPLET_TABLE}" WHERE "{target_date_col}" IN ({placeholders_dates})'

            placeholders_vals = ", ".join("?" for _ in cols)
            collist = ", ".join(f'"{c}"' for c in cols)
            insert_sql = f'INSERT INTO "{DROPLET_TABLE}" ({collist}) VALUES ({placeholders_vals})'

            # Retry 機制 (處理 SQLite Locked)
            for attempt in range(5):
                try:
                    # [關鍵] 使用 SQLite 事務特性，不立即 Commit

                    # 步驟 A: 刪除舊資料 (Scope Delete)
                    conn.execute(delete_sql, list(involved_dates))

                    # 步驟 B: 寫入新資料 (Batch Insert)
                    conn.executemany(insert_sql, clean_rows)

                    # 步驟 C: 全部成功才提交 (Commit)
                    conn.commit()
                    trigger_droplet_sync_safe()
                    break # 成功則跳出迴圈


                except sqlite3.OperationalError as e:
                    if "locked" in str(e).lower() and attempt < 4:
                        conn.rollback() # 鎖定時回滾，準備重試
                        time.sleep(0.3)
                        continue
                    else:
                        raise # 其他錯誤或重試耗盡
                except Exception:
                    conn.rollback() # 其他錯誤，安全回滾
                    raise

        return Response(status=204)

    except sqlite3.Error as e:
        app.logger.error(f"Database Error: {e}")
        return Response(f"Database operation failed: {e}", 500)
    except Exception as e:
        app.logger.error(f"Unexpected Server Error during upload: {e}")
        return Response("Internal Server Error during data processing.", 500)

@app.get("/api/sync_status")  # DropletSchedule 同步狀態查詢
def api_sync_status():
    with connect() as conn:
        row = conn.execute(
            "SELECT status, start_time, end_time, duration_sec, last_error FROM sync_status WHERE id=1"
        ).fetchone()
        if not row:
            return jsonify({"status": "unknown"})
        return jsonify(dict(row))

@app.get("/api/preview")
def preview():
    limit = int(request.args.get("limit", "50"))
    with connect() as conn:
        cur = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?;", (DROPLET_TABLE,))
        if cur.fetchone() is None:
            return jsonify({"columns": [], "rows": []})
        cols = [r[1] for r in conn.execute(f'PRAGMA table_info("{DROPLET_TABLE}")')]
        cur = conn.execute(f'SELECT * FROM "{DROPLET_TABLE}" LIMIT {limit}')
        rows = [dict(r) for r in cur.fetchall()]
        return jsonify({"columns": cols, "rows": rows})





# ===========================
# 日期格式規範化
# ===========================
from flask import Flask, request, jsonify
# ===========================
# 日期格式規範化
# ===========================
# 注意：这里不需要额外的导入，因为文件顶部已经有 import datetime

def normalize_date(date_str: str) -> str:
    """
    将各种日期格式统一转换为 yyyy-mm-dd
    支持: yyyy-mm-dd, yyyy/m/dd, yyyy/mm/dd
    """
    if not date_str:
        return None

    # 尝试 yyyy-mm-dd 格式
    try:
        date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d')
        return date_obj.strftime('%Y-%m-%d')
    except ValueError:
        pass

    # 尝试 yyyy/m/dd 或 yyyy/mm/dd 格式
    try:
        date_obj = datetime.datetime.strptime(date_str, '%Y/%m/%d')
        return date_obj.strftime('%Y-%m-%d')
    except ValueError:
        pass

    return None


def generate_date_formats(normalized_date: str) -> list:
    """
    根据标准化日期生成数据库可能存在的多种格式

    输入: 2026-01-09
    输出: ['2026-1-9', '2026/1/9', '2026-01-09', '2026/01/09']
    """
    parts = normalized_date.split('-')
    year, month, day = parts[0], parts[1], parts[2]

    # 去掉前导零的版本
    month_no_zero = str(int(month))
    day_no_zero = str(int(day))

    # 保留前导零的版本
    month_zero = month
    day_zero = day

    return [
        f"{year}-{month_no_zero}-{day_no_zero}",      # yyyy-m-d (如: 2026-1-9)
        f"{year}/{month_no_zero}/{day_no_zero}",      # yyyy/m/d (如: 2026/1/9) ← 数据库实际格式
        f"{year}-{month_zero}-{day_zero}",            # yyyy-mm-dd (如: 2026-01-09)
        f"{year}/{month_zero}/{day_zero}",            # yyyy/mm/dd (如: 2026/01/09)
    ]


def query_droplet_records_sqlite(date_formats: list):
    """
    查询 SQLite 数据库，支持多种日期格式
    """
    con = connect()
    try:
        # 检查表是否存在
        cur = con.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
            (DROPLET_TABLE,)
        )
        if not cur.fetchone():
            return []

        # 构建 IN 子句
        placeholders = ','.join(['?' for _ in date_formats])
        query = f"""
            SELECT * FROM "{DROPLET_TABLE}"
            WHERE "Date" IN ({placeholders})
            ORDER BY "Date" DESC
        """

        cur = con.execute(query, date_formats)
        cols = [d[0] for d in cur.description]
        rows = cur.fetchall()
        records = [dict(zip(cols, row)) for row in rows]

        return records
    finally:
        con.close()


# ===========================
# Flask API 路由
# ===========================

@app.route('/api/droplet-records', methods=['GET'])
def get_droplet_records():
    """
    获取滴定条件记录

    查询参数: date
    示例: /api/droplet-records?date=2026-01-19
    """

    # 获取 date 参数
    date_param = request.args.get('date')

    if not date_param:
        return jsonify({'error': '缺少 date 参数'}), 400

    # 标准化日期参数
    normalized_date = normalize_date(date_param)

    if not normalized_date:
        return jsonify({'error': '日期格式不正确'}), 400

    # 生成数据库可能存在的多种格式
    date_formats = generate_date_formats(normalized_date)

    print(f"🔍 查询日期: {date_param}")
    print(f"📅 标准化为: {normalized_date}")
    print(f"🔎 搜索格式: {date_formats}")

    try:
        # 直接查询 SQLite 数据库
        con = connect()
        try:
            # 检查表是否存在
            cur = con.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
                ("dropletRecord",)
            )
            if not cur.fetchone():
                print(f"❌ 表 'dropletRecord' 不存在")
                return jsonify({'error': '表不存在'}), 500

            print(f"✅ 找到表 'dropletRecord'")

            # 构建 IN 子句
            placeholders = ','.join(['?' for _ in date_formats])
            query = f"""
                SELECT * FROM "dropletRecord"
                WHERE "record_date" IN ({placeholders})
                ORDER BY "record_date" DESC
            """

            print(f"📋 执行SQL: {query}")
            print(f"📍 参数: {date_formats}")

            cur = con.execute(query, date_formats)
            cols = [d[0] for d in cur.description]
            rows = cur.fetchall()

            print(f"📊 查询到 {len(rows)} 行")
            print(f"📋 字段: {cols}")

            # 转换为字典
            records = [dict(zip(cols, row)) for row in rows]

            if records:
                print(f"📝 第一条数据: {records[0]}")

            # 返回给前端
            return jsonify({'rows': records}), 200

        finally:
            con.close()

    except Exception as e:
        print(f"❌ 查询错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': '数据库查询失败', 'detail': str(e)}), 500

# ==================== Excel 宏 / 兩表導入 DB ====================
TEMP_IN_MAP: dict[str, Path] = {}
TEMP_OUT_MAP: dict[str, Path] = {}

def ensure_table_and_columns(con: sqlite3.Connection, table: str, headers: list[str]) -> None:
    cur = con.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?;", (table,))
    exists = cur.fetchone() is not None
    if not exists:
        cols_def = ", ".join(f'{qident(h)} TEXT' for h in headers)
        con.execute(f'CREATE TABLE {qident(table)} ({cols_def});')
        return
    cur = con.execute(f'PRAGMA table_info({qident(table)});')
    existing = {r[1] for r in cur.fetchall()}
    for h in headers:
        if h not in existing:
            con.execute(f'ALTER TABLE {qident(table)} ADD COLUMN {qident(h)} TEXT;')

def upsert_rows(con: sqlite3.Connection, table: str, headers: list[str], rows: list[tuple], pk_cols: list[str]) -> None:
    cols = ", ".join(qident(h) for h in headers)
    qmarks = ", ".join(["?"] * len(headers))
    if pk_cols:
        upd_cols = [h for h in headers if h not in pk_cols]
        if upd_cols:
            upd = ", ".join(f'{qident(c)}=excluded.{qident(c)}' for c in upd_cols)
            sql = f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks}) ' \
                  f'ON CONFLICT({", ".join(qident(c) for c in pk_cols)}) DO UPDATE SET {upd};'
        else:
            sql = f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks}) ' \
                  f'ON CONFLICT({", ".join(qident(c) for c in pk_cols)}) DO NOTHING;'
    else:
        sql = f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks});'
    for attempt in range(5):
        try:
            con.executemany(sql, rows)
            return
        except sqlite3.OperationalError as e:
            if "locked" in str(e).lower() and attempt < 4:
                time.sleep(0.3); continue
            raise

def read_sheet_to_db(con: sqlite3.Connection, ws, table_name: str, header_row: int) -> dict:
    max_col = ws.max_column or 1
    headers = [str(ws.cell(header_row, c).value or "").strip() for c in range(1, max_col + 1)]
    headers = [h if h else f"col_{i+1}" for i, h in enumerate(headers)]
    rows = []
    r0 = header_row + 1
    max_r = ws.max_row or r0
    for rr in range(r0, max_r + 1):
        row, nonempty = [], False
        for c in range(1, max_col + 1):
            v = ws.cell(rr, c).value
            if v not in (None, ""): nonempty = True
            row.append("" if v is None else str(v))
        if nonempty:
            rows.append(tuple(row))

    ensure_table_and_columns(con, table_name, headers)
    pk_cols = []
    if "工單號碼" in headers and "料號" in headers:
        pk_cols = ["工單號碼", "料號"]
    elif "工單號碼" in headers:
        pk_cols = ["工單號碼"]
    upsert_rows(con, table_name, headers, rows, pk_cols)
    return {"table": table_name, "headers": headers, "rows": len(rows), "pk": pk_cols}

def run_excel_macro(xlsm_path: str, macro_name: str) -> tuple[bool, str | None]:
    try:
        pythoncom.CoInitialize()
        excel = Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            excel.AutomationSecurity = 1  # msoAutomationSecurityLow
        except Exception:
            pass
        wb = excel.Workbooks.Open(xlsm_path)
        try:
            excel.Run(macro_name)
            wb.Save()
            return True, None
        except Exception as e:
            return False, f"Run macro failed: {e}"
        finally:
            try: wb.Close(SaveChanges=False)
            except Exception: pass
            excel.Quit()
    except Exception as e:
        return False, f"Excel automation error: {e}"
    finally:
        try: pythoncom.CoUninitialize()
        except Exception: pass

@app.post("/api/upload_excel")
def api_upload_excel():
    f = request.files.get("file")
    if not f:
        return jsonify(ok=False, message="missing file"), 400

    fname_raw = f.filename or "uploaded.xlsm"
    if "配藥表" not in fname_raw:
        return jsonify(ok=False, message="所選檔名未包含「配藥表」，請重新選擇正確的配藥表 Excel。"), 400

    ext = Path(fname_raw).suffix.lower() or ".xlsm"
    if ext not in (".xlsm", ".xlsx", ".xls"):
        ext = ".xlsm"

    temp_id = uuid.uuid4().hex[:8]
    temp_in_path = Path(APP_CONF["UPLOAD_DIR"]) / f"upload_{temp_id}{ext}"
    f.save(str(temp_in_path))
    TEMP_IN_MAP[temp_id] = temp_in_path

    ok_macro, err_macro = run_excel_macro(str(temp_in_path), "manuinsert.InsertNewByWorkOrderArrays")
    if not ok_macro:
        return jsonify(ok=False, message=f"Excel 宏執行失敗：{err_macro}"), 500

    wb = load_workbook(temp_in_path, keep_vba=True, data_only=True)
    sheets = wb.sheetnames[:2]
    ingested = []

    con = connect()
    try:
        with con:
            con.execute("PRAGMA temp_store=MEMORY;")
            for s in sheets:
                ws = wb[s]
                table = normalize_table_name(s)
                info = read_sheet_to_db(con, ws, table, APP_CONF["HEADER_ROW"])
                ingested.append(info)
    finally:
        con.close()

    return jsonify(ok=True, temp_id=temp_id, filename=fname_raw, sheets=sheets, ingested=ingested)

@app.post("/api/create_record")
def api_create_record():
    data = request.get_json(force=True, silent=True) or {}
    work_order = (data.get("work_order") or "").strip()
    temp_id = (data.get("temp_id") or "").strip()
    table_raw = (data.get("table") or "").strip()
    if not work_order or not temp_id:
        return jsonify(ok=False, message="work_order & temp_id are required"), 400

    in_path = TEMP_IN_MAP.get(temp_id)
    if not in_path or not in_path.exists():
        return jsonify(ok=False, message=f"temp_id not found: {temp_id}"), 404
    wb_in = load_workbook(in_path, keep_vba=True, data_only=True)
    first_sheet = normalize_table_name(wb_in.sheetnames[0]) if wb_in.sheetnames else None
    table = normalize_table_name(table_raw) if table_raw else first_sheet
    if not table:
        return jsonify(ok=False, message="no sheet to use as table"), 400

    sql = f"""
      SELECT 工單號碼 AS work_order, BeadsLot AS bead_lot,
             料號 AS material, 重量紀錄 AS weight, Filler_Lot AS lot_no
      FROM {qident(table)}
      WHERE 工單號碼 = ?
      ORDER BY 料號
    """
    with connect() as conn:
        cur = conn.execute(sql, (work_order,))
        rows = [dict(zip([c[0] for c in cur.description], r)) for r in cur.fetchall()]

    template = APP_CONF["TEMPLATE_XLSM"]
    if not os.path.exists(template):
        return jsonify(ok=False, message=f"template not found: {template}"), 500

    out_path = Path(APP_CONF["UPLOAD_DIR"]) / f"record_{temp_id}.xlsm"
    shutil.copy2(template, out_path)

    wb = load_workbook(out_path, keep_vba=True, data_only=False)
    if "製程紀錄表" not in wb.sheetnames:
        return jsonify(ok=False, message="template has no sheet: 製程紀錄表"), 500
    ws = wb["製程紀錄表"]

    ws["V6"] = work_order

    HEADER_ROW = APP_CONF["HEADER_ROW"]
    COL_BOM_FALLBACK   = APP_CONF["COL_BOM_FALLBACK"]
    COL_TOTAL_FALLBACK = APP_CONF["COL_TOTAL_FALLBACK"]
    COL_LOT_FALLBACK   = APP_CONF["COL_LOT_FALLBACK"]

    def norm(s): return (str(s or "")).strip().lower().replace(" ", "")

    title2col = {}
    max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        title2col[norm(ws.cell(HEADER_ROW, c).value)] = c

    def get_col(names, fallback):
        for n in names:
            col = title2col.get(norm(n))
            if col:
                return col
        return fallback

    col_bom   = get_col(["bom p/n","bompn","料號","bom"], COL_BOM_FALLBACK)
    col_total = get_col(["total qty","totalqty","重量紀錄"], COL_TOTAL_FALLBACK)
    col_lot   = get_col(["lot no.","lotno","filler_lot","lot no","lotno."], COL_LOT_FALLBACK)

    start_row = HEADER_ROW + 1
    end_row = ws.max_row if ws.max_row > start_row else start_row
    if end_row >= start_row:
        ws.delete_rows(start_row, end_row - start_row + 1)

    r = start_row
    for it in rows:
        ws.cell(r, col_bom).value   = it.get("material")
        ws.cell(r, col_total).value = it.get("weight")
        ws.cell(r, col_lot).value   = it.get("lot_no")
        r += 1

    wb.save(out_path)
    TEMP_OUT_MAP[temp_id] = out_path
    return jsonify(ok=True, filled=len(rows), out_path=str(out_path), table=table)

@app.get("/api/template_preview")
def api_template_preview():
    temp_id = (request.args.get("temp_id") or "").strip()
    out_path = TEMP_OUT_MAP.get(temp_id)
    if not out_path or not Path(out_path).exists():
        return jsonify(ok=False, message=f"no generated file for temp_id: {temp_id}"), 404

    wb = load_workbook(out_path, keep_vba=True, data_only=True)
    if "製程紀錄表" not in wb.sheetnames:
        return jsonify(ok=False, message="template has no sheet: 製程紀錄表"), 500
    ws = wb["製程紀錄表"]

    HEADER_ROW = APP_CONF["HEADER_ROW"]
    max_col = ws.max_column or 1
    headers = [str(ws.cell(HEADER_ROW, c).value or "") for c in range(1, max_col + 1)]

    rows = []
    r = HEADER_ROW + 1
    max_r = min(ws.max_row or r, r + 50)
    for rr in range(r, max_r + 1):
        row = []
        for c in range(1, max_col + 1):
            v = ws.cell(rr, c).value
            row.append("" if v is None else str(v))
        rows.append(row)
    return jsonify(ok=True, headers=headers, rows=rows)

@app.get("/api/template_file")
def api_template_file():
    temp_id = (request.args.get("temp_id") or "").strip()
    out_path = TEMP_OUT_MAP.get(temp_id)
    if not out_path or not Path(out_path).exists():
        return jsonify(ok=False, message=f"no generated file for temp_id: {temp_id}"), 404
    return send_file(
        out_path,
        as_attachment=True,
        download_name=Path(out_path).name,
        mimetype="application/vnd.ms-excel.sheet.macroEnabled.12",
    )

@app.post("/api/save_template")
def api_save_template():
    data = request.get_json(force=True, silent=True) or {}
    temp_id = (data.get("temp_id") or "").strip()
    work_order = (data.get("work_order") or "").strip()
    server_dir = (data.get("server_dir") or "").strip()
    filename   = (data.get("filename") or "").strip()

    out_src = TEMP_OUT_MAP.get(temp_id)
    if not out_src or not Path(out_src).exists():
        return jsonify(ok=False, message=f"no generated file for temp_id: {temp_id}"), 404

    if not server_dir or not filename:
        return jsonify(ok=False, message="server_dir, filename are required"), 400
    if not filename.endswith(".xlsm"):
        filename += ".xlsm"

    if work_order:
        wb = load_workbook(out_src, keep_vba=True, data_only=False)
        if "製程紀錄表" in wb.sheetnames:
            wb["製程紀錄表"]["V6"] = work_order
            wb.save(out_src)

    out_dir = Path(server_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / filename
    shutil.copy2(out_src, out_path)
    return jsonify(ok=True, saved_path=str(out_path))

# ==================== CSV 相容上傳 (raw/form-data) ====================
def _parse_csv_text(text: str):
    if text and text[:1] == '\ufeff':
        text = text.lstrip('\ufeff')
    reader = csv.reader(io.StringIO(text))
    try:
        headers = next(reader)
    except StopIteration:
        headers = []
    headers = [h.strip() if h else f"col_{i+1}" for i, h in enumerate(headers)]
    rows = []
    for r in reader:
        if len(r) < len(headers): r = r + [""] * (len(headers) - len(r))
        elif len(r) > len(headers): r = r[:len(headers)]
        rows.append(tuple(r))
    return headers, rows

def ensure_unique_index(con: sqlite3.Connection, table: str, pk_cols: list[str]) -> None:
    if not pk_cols:
        return
    idx = "ux_" + re.sub(r'[^A-Za-z0-9_]', '_', table) + "__" + "__".join(
        re.sub(r'[^A-Za-z0-9_]', '_', c) for c in pk_cols
    )
    cols = ", ".join(qident(c) for c in pk_cols)
    con.execute(f'CREATE UNIQUE INDEX IF NOT EXISTS {qident(idx)} ON {qident(table)} ({cols});')

def _get_mode() -> str:
    mode = (request.form.get("mode") or request.args.get("mode") or "replace").strip().lower()
    return mode if mode in ("append", "replace", "upsert") else "replace"

def _want_csv_response() -> bool:
    return (request.args.get("return") or "").strip().lower() == "csv"

def _insert_rows(con: sqlite3.Connection, table: str, headers: list[str], rows: list[tuple]) -> int:
    if not rows:
        return 0
    cols = ", ".join(qident(h) for h in headers)
    qmarks = ", ".join(["?"] * len(headers))
    con.executemany(f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks});', rows)
    return len(rows)

def _replace_rows(con: sqlite3.Connection, table: str, headers: list[str], rows: list[tuple]) -> int:
    con.execute(f'DELETE FROM {qident(table)};')
    return _insert_rows(con, table, headers, rows)

def _handle_csv_upload(table: str, headers: list[str], rows: list[tuple], pk_cols: list[str], mode: str):
    con = connect()
    try:
        with con:
            con.execute("PRAGMA temp_store=MEMORY;")
            ensure_table_and_columns(con, table, headers)
            if not pk_cols:
                if "工單號碼" in headers and "料號" in headers:
                    pk_cols = ["工單號碼", "料號"]
                elif "工單號碼" in headers:
                    pk_cols = ["工單號碼"]
            if pk_cols:
                ensure_unique_index(con, table, pk_cols)
            if mode == "append":
                if pk_cols:
                    upsert_rows(con, table, headers, rows, pk_cols)
                else:
                    _insert_rows(con, table, headers, rows)
            elif mode == "upsert":
                if not pk_cols:
                    _insert_rows(con, table, headers, rows)
                else:
                    upsert_rows(con, table, headers, rows, pk_cols)
            else:  # replace
                if pk_cols:
                    upsert_rows(con, table, headers, rows, pk_cols)
                else:
                    _replace_rows(con, table, headers, rows)
    finally:
        con.close()

@app.post("/api/upload_csv_raw")
@app.post("/upload_csv_raw")
def api_upload_csv_raw():
    table = (request.args.get("table") or "").strip()
    if not table:
        return jsonify(ok=False, message="missing table"), 400
    table = normalize_table_name(table)
    try:
        table = safe_table_name(table)
    except ValueError as e:
        return jsonify(ok=False, message=str(e)), 400

    pk_raw = (request.args.get("pk") or "").strip()
    mode = _get_mode()
    pk_cols = [c.strip() for c in pk_raw.split(",") if c.strip()]

    text = request.get_data(cache=False, as_text=True)
    if not text:
        return jsonify(ok=False, message="empty body"), 400

    headers, rows = _parse_csv_text(text)
    try:
        _handle_csv_upload(table, headers, rows, pk_cols, mode)
    except Exception as e:
        app.logger.exception("upload_csv_raw failed")
        if _want_csv_response():
            return Response("status,FAIL\n", status=500, mimetype="text/csv; charset=utf-8")
        return jsonify(ok=False, message=str(e)), 500

    if _want_csv_response():
        return Response("status,OK\n", mimetype="text/csv; charset=utf-8")
    return jsonify(ok=True)

@app.post("/api/upload_csv")
@app.post("/upload_csv")
def api_upload_csv():
    table = (request.form.get("table") or request.args.get("table") or "").strip()
    if not table:
        return jsonify(ok=False, message="missing table"), 400
    table = normalize_table_name(table)
    try:
        table = safe_table_name(table)
    except ValueError as e:
        return jsonify(ok=False, message=str(e)), 400

    pk_raw = (request.form.get("pk") or request.args.get("pk") or "").strip()
    mode = _get_mode()
    pk_cols = [c.strip() for c in pk_raw.split(",") if c.strip()]

    f = request.files.get("file") or request.files.get("csv")
    if not f:
        return jsonify(ok=False, message="missing file"), 400

    raw = f.read()
    try:
        text = raw.decode("utf-8-sig", errors="replace")
    except Exception:
        text = raw.decode(errors="replace")

    headers, rows = _parse_csv_text(text)
    try:
        _handle_csv_upload(table, headers, rows, pk_cols, mode)
    except Exception as e:
        app.logger.exception("upload_csv failed")
        if _want_csv_response():
            return Response("status,FAIL\n", status=500, mimetype="text/csv; charset=utf-8")
        return jsonify(ok=False, message=str(e)), 500

    if _want_csv_response():
        return Response("status,OK\n", mimetype="text/csv; charset=utf-8")
    return jsonify(ok=True)

# ==================== QR 產生（從 xlsm cells） ====================
@app.post("/api/qr_png_from_cells")
def api_qr_png_from_cells():
    data = request.get_json(force=True)
    cells = data.get("cells") or []
    joiner = (data.get("joiner") or "|")
    file_path = (data.get("file_path") or "").strip()
    temp_id = (data.get("temp_id") or "").strip()

    if not (file_path or temp_id):
        return jsonify(ok=False, message="file_path or temp_id is required"), 400
    if not cells:
        return jsonify(ok=False, message="cells is required"), 400

    xlsm_path = file_path or str(TEMP_OUT_MAP.get(temp_id) or "")
    if not xlsm_path or not os.path.exists(xlsm_path):
        return jsonify(ok=False, message=f"file not found: {xlsm_path}"), 404

    wb = load_workbook(xlsm_path, data_only=True, keep_vba=True)
    ws = wb.active
    vals = []
    for addr in cells:
        if "!" in addr:
            sheet, cell = addr.split("!", 1)
            vals.append("" if wb[sheet][cell].value is None else str(wb[sheet][cell].value))
        else:
            vals.append("" if ws[addr].value is None else str(ws[addr].value))
    payload = joiner.join((v or "").strip() for v in vals)

    qr = qrcode.QRCode(version=None, error_correction=ERROR_CORRECT_M, box_size=10, border=2)
    qr.add_data(payload)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png")

# ==================== WorkOrder 查詢 / 儲存 ====================

def query_db(sql, args=()):
    with connect() as conn:
        cur = conn.execute(sql, args)
        return [dict(row) for row in cur.fetchall()]

def safe_float(v):
    try:
        return float(v) if v not in (None, "") else 0.0
    except:
        return 0.0

def safe_int(v):
    try:
        return int(float(v))
    except:
        return 0

def parse_range(text):
    if not text:
        return None
    m = re.match(r"([\d.]+)\s*[-~]\s*([\d.]+)", str(text))
    if not m:
        return None
    return float(m.group(1)), float(m.group(2))

def check_in_range(value, text_range):
    if not text_range:
        return None
    lo, hi = text_range
    return lo <= value <= hi

def norm_name(s: str) -> str:
    if not s:
        return ""
    s = str(s).strip()
    s = re.sub(r"[-_].*$", "", s)
    return s.strip().upper()

import sqlite3

def find_table_contains_workorder(conn: sqlite3.Connection, work_order: str):
    """回傳第一個包含該工單的 table 名稱，找不到回傳 None"""

    # 1. 清理工單號碼參數，並開始追蹤
    work_order = work_order.strip()
    print(f"\n--- 🐛 追蹤: 尋找工單 '{work_order}' (含 Rdtest 表) ---")

    # 2. 篩選資料表：
    #    條件 A: 名稱包含 '571'
    #    條件 B: 名稱是 'RdtestURD2026_001' 或 'RdtestDRD2026_002'
    sql_get_tables = """
        SELECT name FROM sqlite_master
        WHERE type='table'
        AND (
            name LIKE '%571%'
            OR name IN ('RdtestURD2026_001', 'RdtestDRD2026_002')
        )
    """

    cur = conn.execute(sql_get_tables)
    tables = [r[0] for r in cur.fetchall()]
    print(f"📚 找到符合搜尋範圍的資料表總數: {len(tables)}")
    print(f"   列表: {tables}")  # 印出來確認一下有沒有抓到 Rdtest 表

    # 設置模糊匹配模式
    like_pattern = f"%{work_order}%"

    for t in tables:
        print(f"--- 檢查資料表: {t} ---")
        try:
            # 3. 檢查資料表是否有資料
            #    使用 TRIM(工單號碼) 移除資料庫中欄位的前後空格。
            #    使用 LIKE ? (參數為 %工單號碼%) 進行模糊匹配。

            # 注意：這裡假設這兩張新表也有 '工單號碼' 這個欄位
            # 如果新表的欄位名稱不同 (例如叫 'WorkOrder')，這裡會報 OperationalError
            sql_like = f"SELECT 1 FROM '{t}' WHERE TRIM(工單號碼) LIKE ? LIMIT 1"
            rows = conn.execute(sql_like, (like_pattern,)).fetchall()

            if rows:
                print(f"✅ 找到工單 {work_order} 位於表 {t}，成功。")
                print(f"--- 🐛 追蹤結束 ---")
                return t

        except sqlite3.OperationalError as e:
            # 如果資料表根本沒有 '工單號碼' 欄位，會在這裡報錯
            print(f"❌ {t}: 查詢發生 SQLite 錯誤 (可能缺少 '工單號碼' 欄位): {e}")
        except Exception as e:
            print(f"❌ {t}: 發生未預期的錯誤: {e}")

    print(f"❌ 總結：未在任何指定範圍的表 (571 或 Rdtest) 找到工單 {work_order}")
    print(f"--- 🐛 追蹤結束 ---")
    return None

# ==================== WorkOrder 查詢 / 儲存 ====================

# ... (query_db, safe_float, safe_int, parse_range, check_in_range, norm_name, find_table_contains_workorder... 這些都保留不變) ...


def check_suspension(v):
    """(懸浮物) '混濁' = True, 其他 = False"""
    result = "混濁" in str(v or "")
    print(f"  check_suspension({v!r}) → {result}")
    return result

def check_yes(v):
    """'Yes' (不區分大小寫) = True, 其他 = False"""
    result = str(v or "").strip().lower() == "yes"
    print(f"  check_yes({v!r}) → {result}")
    return result

def check_not_room_temp(v):
    """
    '室溫' 或空值 = False
    '4℃' 或其他非空值 = True
    """
    val = str(v or "").strip()
    # ✅ 修正：空值和 '室溫' 都返回 False
    result = bool(val) and val != "室溫"
    print(f"  check_not_room_temp({v!r}) → val={val!r} → {result}")
    return result

def check_not_no(v):
    """
    'No' (不區分大小寫) 或空值 = False
    'Yes' 或其他非空值 = True
    """
    val = str(v or "").strip()
    # ✅ 修正：空值和 'No' 都返回 False
    result = bool(val) and val.lower() != "no"
    print(f"  check_not_no({v!r}) → val={val!r} → {result}")
    return result


@app.get("/api/get_workorder")
def api_get_workorder():
    work_order = request.args.get("work_order")
    if not work_order:
        return jsonify({"error": "缺少工單號碼參數"}), 400

    print(f"\n{'='*60}")
    print(f"🔍 收到查詢工單請求: {work_order}")
    print(f"{'='*60}\n")

    try:
        with connect() as conn:
            target_table, dbg = resolve_target_571_table_fast(conn, work_order)
            print(f"🧭 resolve_target_571_table_fast: {dbg}")

            # ✅ fallback：原本全表掃描（保底）
            if not target_table:
                target_table = find_table_contains_workorder(conn, work_order)

            if not target_table:
                return jsonify({"message": f"查無工單 {work_order}"}), 404

            cur = conn.execute(
                f"SELECT * FROM '{target_table}' WHERE TRIM(工單號碼) LIKE ?",
                (f"%{work_order.strip()}%",),
            )
            cols = [d[0] for d in cur.description]
            rows = cur.fetchall()
            result_rows = [dict(zip(cols, row)) | {"來源表": target_table} for row in rows]

            if not result_rows:
                fallback_table = find_table_contains_workorder(conn, work_order)
                if fallback_table and fallback_table != target_table:
                    print(
                        f"⚠️ {target_table} 無工單明細，改用 fallback 表 {fallback_table}"
                    )
                    target_table = fallback_table
                    cur = conn.execute(
                        f"SELECT * FROM '{target_table}' WHERE TRIM(工單號碼) LIKE ?",
                        (f"%{work_order.strip()}%",),
                    )
                    cols = [d[0] for d in cur.description]
                    rows = cur.fetchall()
                    result_rows = [
                        dict(zip(cols, row)) | {"來源表": target_table}
                        for row in rows
                    ]

                if not result_rows:
                    return jsonify({"message": f"查無工單 {work_order} 的詳細記錄"}), 404

            first = result_rows[0]

            # DropletSchedule 和 pump 邏輯
            maker_name = None
            titration_lots = []      # 滴定日記錄（有 Pump/Lyophilizer）
            formulation_lots = []    # 配藥日記錄（無 Pump/Lyophilizer）
            titration_qty = 0
            formulation_qty = 0
            formulation_date = None
            cur = conn.execute(
                "SELECT Lot, Pump, Lyophilizer, Marker, Quantity, Date, Remark FROM DropletSchedule WHERE TRIM(WorkOrder) LIKE ?",
                (f"%{work_order.strip()}%",),
            )
            for r in cur.fetchall():
                r = dict(r)
                maker_name = r.get("Marker") or maker_name
                pump_val = (r.get("Pump") or "").strip()
                lyo_val  = (r.get("Lyophilizer") or "").strip()
                lot_val  = (r.get("Lot") or "").strip()
                remark   = (r.get("Remark") or "").strip()
                if pump_val or lyo_val:
                    # 滴定日：有 Pump/Lyophilizer
                    titration_qty += safe_int(r.get("Quantity"))
                    titration_lots.append({
                        "id": lot_val,
                        "port": pump_val,
                        "freezeDry": lyo_val,
                        "pump": None,
                    })
                else:
                    # 配藥日
                    formulation_qty += safe_int(r.get("Quantity"))
                    if not formulation_date:
                        formulation_date = r.get("Date")
                    if lot_val:
                        formulation_lots.append({"id": lot_val, "port": "", "freezeDry": "", "pump": None})
                    elif remark and "lot" in remark.lower():
                        lot_match = re.search(r"[Ll]ot\s*([\w/\-]+)", remark)
                        lot_id = lot_match.group(1) if lot_match else remark
                        formulation_lots.append({"id": lot_id, "port": "", "freezeDry": "", "pump": None})
            # 優先使用滴定日；僅純配藥工單（無任何滴定日記錄）才 fallback 用配藥日
            if titration_lots:
                dispose_lots = titration_lots
                product_quantity = titration_qty
            else:
                dispose_lots = formulation_lots
                product_quantity = formulation_qty

            pump_ids = []
            if maker_name:
                cur2 = conn.execute('SELECT * FROM "pump No."')
                for row in cur2.fetchall():
                    row = dict(row)
                    for i in range(1, 11):
                        k = f"可滴定之試劑-{i}"
                        val = row.get(k)
                        if val and maker_name.lower() in str(val).strip().lower():
                            pid = row.get("pump編號")  # ✅ 你的欄位名是 pump編號
                            if pid not in (None, ""):
                                pump_ids.append(str(pid).strip())
                            break

            # 去重（保序）
            seen = set()
            pump_ids = [x for x in pump_ids if not (x in seen or seen.add(x))]

            if pump_ids:
                for i, lot in enumerate(dispose_lots):
                    lot["pump"] = pump_ids[i % len(pump_ids)]

            # ✅ 取得產品型號（從 table 名稱提取）
            match_571 = re.search(r"571\d{5,}", target_table or "")
            product_model = match_571.group(0) if match_571 else None

            print(f"📍 工單表: {target_table}")
            print(f"📍 產品型號: {product_model}")
            print(f"📍 Marker 名稱: {maker_name}")
            base_name = norm_name(maker_name or "")
            print(f"📍 標準化名稱: {base_name}")  # 建議加上這行 log 確認

            # ✅ 1. 查詢 [Liquid form QC] (修改後版本)
            print(f"\n--- 查詢 [Liquid form QC] ---")
            selected_qc = None

            # --- 策略 A: 優先使用 PN (產品型號) 精確查詢 (最準確) ---
            if product_model:
                print(f" 🔍 嘗試使用 PN='{product_model}' 查詢 QC...")
                # 查詢所有欄位，以免漏掉 L1-OD, 版本 等資訊
                qc_rows = query_db("SELECT * FROM [Liquid form QC] WHERE PN = ?", (product_model,))

                # 如果找不到，嘗試處理 Excel 匯入常見的 '.0' 問題 (例如 5715600047.0)
                if not qc_rows:
                    try:
                        pn_clean = str(int(float(product_model)))
                        qc_rows = query_db("SELECT * FROM [Liquid form QC] WHERE PN = ?", (pn_clean,))
                    except:
                        pass

                # ✅ PN 命中後，用 Marker name 嚴格過濾（區分 TG vs Hi TG 等）
                if qc_rows and maker_name:
                    mk_norm = norm_name(maker_name)
                    filtered = [r for r in qc_rows if norm_name(r.get("Marker name", "")) == mk_norm]
                    if filtered:
                        selected_qc = filtered[0]
                        print(f" ✅ PN + Marker name 匹配成功！(Name: {selected_qc.get('Name')}, Marker: {selected_qc.get('Marker name')})")
                    else:
                        print(f" ⚠️ PN 命中但 Marker name 不符 (期望={mk_norm})，跳過繼續比對")
                elif qc_rows:
                    selected_qc = qc_rows[0]
                    print(f" ✅ 成功透過 PN 找到 QC 資料！(Name: {selected_qc.get('Name')})")

            # --- 策略 B: 用完整原始名稱精確匹配 Marker name（不經 norm_name 截斷）---
            if not selected_qc:
                raw_name = (maker_name or "").strip().upper()
                print(f" ⚠️ PN 查無資料，改用完整 Marker name='{raw_name}' 精確比對...")
                for row in query_db("SELECT * FROM [Liquid form QC]"):
                    db_marker_raw = (row.get("Marker name") or "").strip().upper()
                    if db_marker_raw and db_marker_raw == raw_name:
                        selected_qc = row
                        print(f" ✅ 成功透過完整 Marker name 精確比對找到 QC 資料！(Match: {db_marker_raw})")
                        break

            # --- 策略 C: 用完整原始名稱精確匹配 Name ---
            if not selected_qc:
                raw_name = (maker_name or "").strip().upper()
                print(f" ⚠️ 完整 Marker name 查無，改用完整 Name='{raw_name}' 精確比對...")
                for row in query_db("SELECT * FROM [Liquid form QC]"):
                    db_name_raw = (row.get("Name") or "").strip().upper()
                    if db_name_raw and db_name_raw == raw_name:
                        selected_qc = row
                        print(f" ✅ 成功透過完整 Name 精確比對找到 QC 資料！(Match: {db_name_raw})")
                        break

            # --- 策略 D: fallback 用 norm_name 模糊比對 ---
            if not selected_qc:
                print(f" ⚠️ 完整名稱查無，改用 norm_name='{base_name}' 模糊比對...")
                for row in query_db("SELECT * FROM [Liquid form QC]"):
                    db_marker = norm_name(row.get("Marker name", ""))
                    if db_marker and db_marker == base_name:
                        selected_qc = row
                        print(f" ✅ 成功透過 norm_name Marker 比對找到 QC 資料！(Match: {db_marker})")
                        break

            if not selected_qc:
                for row in query_db("SELECT * FROM [Liquid form QC]"):
                    db_name = norm_name(row.get("Name", ""))
                    if db_name and (db_name == base_name or db_name in base_name or base_name in db_name):
                        selected_qc = row
                        print(f" ✅ 成功透過 norm_name Name 模糊比對找到 QC 資料！(Match: {db_name})")
                        break

            # --- 最終檢查 ---
            if not selected_qc:
                selected_qc = {
                    "L1-OD": None, "L2-OD": None,
                    "L1-起始OD": None, "L2-起始OD": None,
                    "懸浮物": None,
                    "版本": None
                }
                print(f" ❌ 最終仍未找到 QC 資料")
            else:
                # 確保 L1-OD 等欄位能被後續讀取 (處理可能的欄位名稱差異)
                print(f" 📋 QC 資料預覽: L1-OD={selected_qc.get('L1-OD')}, 版本={selected_qc.get('版本')}")

            # ✅ 2. 查詢滴定條件 - 使用產品型號（PN）
            print(f"\n--- 查詢滴定條件 ---")

            selected_titration = {}

            # 嘗試不同的表名
            for table_name in ["適定條件", "滴定條件"]:
                try:
                    # 檢查表是否存在
                    cur = conn.execute(
                        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                        (table_name,)
                    )
                    if not cur.fetchone():
                        print(f"  ⚠️ 表 [{table_name}] 不存在")
                        continue

                    print(f"  🔍 查詢表: [{table_name}]")

                    # 檢查欄位
                    cur = conn.execute(f"PRAGMA table_info('{table_name}')")
                    table_cols = [r[1] for r in cur.fetchall()]
                    print(f"    欄位: {table_cols}")

                    # ✅ 優先使用產品型號（PN）查詢
                    if "PN" in table_cols and product_model:
                        print(f"    嘗試用 PN = '{product_model}' 查詢...")
                        result = query_db(f"""
                            SELECT * FROM [{table_name}] WHERE PN = ?
                        """, (product_model,))

                        if result:
                            selected_titration = result[0]
                            print(f"    ✅ 找到資料！")
                            print(f"    完整資料: {selected_titration}")
                            break

                    # 如果 PN 查不到，再嘗試用 Name
                    if "Name" in table_cols and base_name:
                        print(f"    嘗試用 Name = '{base_name}' 查詢...")
                        result = query_db(f"""
                            SELECT * FROM [{table_name}] WHERE Name = ?
                        """, (base_name,))

                        if result:
                            selected_titration = result[0]
                            print(f"    ✅ 找到資料！")
                            print(f"    完整資料: {selected_titration}")
                            break

                except Exception as e:
                    print(f"  ❌ 查詢表 [{table_name}] 失敗: {e}")
                    continue

            if not selected_titration:
                print(f"⚠️ 未找到滴定條件")

            # ✅ 3. 建立 confirm 資料
            print(f"\n{'='*60}")
            print(f"開始建立 confirm_data")
            print(f"{'='*60}")

            print(f"\n📋 原始資料:")
            print(f"  懸浮物: {selected_qc.get('懸浮物')!r}")
            print(f"  版本: {selected_qc.get('版本')!r}")
            print(f"  儲存時避光: {selected_titration.get('儲存時避光')!r}")
            print(f"  儲存時冰浴: {selected_titration.get('儲存時冰浴')!r}")
            print(f"  滴定時避光: {selected_titration.get('滴定時避光')!r}")
            print(f"  滴定時冰浴: {selected_titration.get('滴定時冰浴')!r}")
            print(f"  滴定時攪拌: {selected_titration.get('滴定時攪拌')!r}")
            print(f"  滴定_Mixing: {selected_titration.get('滴定_Mixing')!r}")

            print(f"\n🔄 轉換過程:")

            suspension_val = check_suspension(selected_qc.get("懸浮物"))
            storeLight_val = check_yes(selected_titration.get("儲存時避光"))
            storeIce_val = check_not_room_temp(selected_titration.get("儲存時冰浴"))
            dyeing_val = check_yes(selected_titration.get("滴定時避光"))
            washing_val = check_yes(selected_titration.get("滴定時冰浴"))

            stir_raw = selected_titration.get("滴定時攪拌") or selected_titration.get("滴定_Mixing")
            stir_val = check_not_no(stir_raw)

            confirm_data = {
                "suspension": suspension_val,
                "storeLight": storeLight_val,
                "storeIce": storeIce_val,
                "dyeing": dyeing_val,
                "washing": washing_val,
                "stir": stir_val,
            }

            print(f"\n✅ 最終 confirm_data: {confirm_data}")
            print(f"{'='*60}\n")

            # 其餘邏輯保持不變
            # ✅ 修正：增加讀取 "L1-OD" (連字號) 與 "L1-起始OD" 的支援
            L1OD = safe_float(first.get("L1-OD") or first.get("L1OD") or first.get("L1 OD"))
            L2OD = safe_float(first.get("L2-OD") or first.get("L2OD") or first.get("L2 OD"))

            # ✅ 修正：增加對應常見的中文欄位變體
            L1StartOD = safe_float(first.get("L1-起始OD") or first.get("起始L1OD") or first.get("L1 起始 OD") or first.get("L1StartOD"))
            L2StartOD = safe_float(first.get("L2-起始OD") or first.get("起始L2OD") or first.get("L2 起始 OD") or first.get("L2StartOD"))

            qc_map = {"L1OD":"L1-OD","L2OD":"L2-OD","L1StartOD":"L1-起始OD","L2StartOD":"L2-起始OD"}
            values = {"L1OD": L1OD, "L2OD": L2OD, "L1StartOD": L1StartOD, "L2StartOD": L2StartOD}
            check_results = {}
            for k, qc_col in qc_map.items():
                rng = parse_range(selected_qc.get(qc_col))
                val = values[k]
                check_results[k] = {
                    "value": val,
                    "qc_range": selected_qc.get(qc_col),
                    "pass": check_in_range(val, rng)
                }

            data = {
                "workOrderNo": work_order,
                "productModel": product_model or target_table,
                "markerName": maker_name or (target_table or "").split("_")[0],
                "productQuantity": product_quantity,
                "date": formulation_date or first.get("試劑配製日期", ""),
                "version": selected_qc.get("版本", ""),  # ✅ 加入版本資訊
                "beads": [
                    {
                        "beadName": r.get("化學品名", ""),
                        "beadPN": r.get("料號", ""),
                        "unit": "g",
                        "qtyPerBead": r.get("總重量", 0),
                        "totalQty": r.get("重量紀錄", 0),
                        "lotNo": r.get("Filler_Lot", ""),
                    }
                    for r in result_rows
                ],
                "reagent": {
                    "preparedBy": first.get("配製人員", ""),
                    "confirm": confirm_data,
                },
                "bufferBase": {
                    "L1OD": L1OD, "L2OD": L2OD,
                    "L1StartOD": L1StartOD, "L2StartOD": L2StartOD,
                },
                "qcRanges": {
                    "L1-OD": selected_qc.get("L1-OD"),
                    "L2-OD": selected_qc.get("L2-OD"),
                    "L1-起始OD": selected_qc.get("L1-起始OD"),
                    "L2-起始OD": selected_qc.get("L2-起始OD"),
                },
                "qcCheckResult": check_results,
                "disposeLots": dispose_lots,
            }
            return jsonify(data)

    except Exception as e:
        print("❌ 查詢錯誤:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
# ==================== 新增路由：搜尋 571 表(資料庫) ====================
import traceback
import re

def norm_key(s: str) -> str:
    """用來比對 Marker 名稱：去空白、去連字號、轉大寫"""
    return re.sub(r"[\s\-_/]+", "", (s or "").strip()).upper()

def get_marker_candidates_from_dropletschedule(conn, work_order: str) -> list[str]:
    cur = conn.execute(
        "SELECT DISTINCT Marker FROM DropletSchedule WHERE TRIM(WorkOrder) LIKE ?",
        (f"%{work_order.strip()}%",),
    )
    markers = []
    for (m,) in cur.fetchall():
        if m and str(m).strip():
            markers.append(str(m).strip())
    return markers

def resolve_pn_from_qc_by_marker(conn, marker: str) -> str | None:
    """
    用 DropletSchedule.Marker → Liquid form QC.Name → PN(571*)
    規則：
    1) norm(QC.Name) == norm(marker)  ← 最優先
    2) norm(QC.Name).startswith(norm(marker))
    """
    if not marker:
        return None

    mk = norm_key(marker)

    cur = conn.execute("SELECT PN, Name FROM [Liquid form QC]")
    for pn, name in cur.fetchall():
        if not name or not pn:
            continue

        nk = norm_key(name)

        # ✅ 嚴格規則（避免 Na_BD ⊂ QNa_BD 這種誤中）
        if nk == mk or nk.startswith(mk):
            pn_str = str(pn).strip()
            try:
                pn_str = str(int(float(pn_str)))  # 處理 571xxxx.0
            except:
                pass

            if pn_str.startswith("571"):
                return pn_str

    return None



def resolve_571_table_by_pn(conn, pn: str) -> str | None:
    """用 PN 找出包含該 PN 的 571 table 名稱"""
    if not pn:
        return None
    cur = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE ?",
        (f"%{pn}%",),
    )
    # 若有多張，挑最像你命名規則的（包含 571 開頭數字且較長）
    candidates = [r[0] for r in cur.fetchall()]
    if not candidates:
        return None
    # 先優先包含 571\d 的
    candidates.sort(key=lambda x: (0 if re.search(r"571\d+", x) else 1, len(x)))
    return candidates[0]

def resolve_target_571_table_fast(conn, work_order: str) -> tuple[str | None, dict]:
    """
    主要加速器：
    work_order → DropletSchedule.Marker candidates → QC PN → 571 table
    回傳：(target_table, debug_info)
    """
    dbg = {"markers": [], "pn": None, "table": None, "reason": ""}

    markers = get_marker_candidates_from_dropletschedule(conn, work_order)
    dbg["markers"] = markers

    # 逐一用 marker 嘗試解 PN（遇到 BU/BD 混在一起也沒關係）
    for m in markers:
        pn = resolve_pn_from_qc_by_marker(conn, m)
        if pn:
            dbg["pn"] = pn
            table = resolve_571_table_by_pn(conn, pn)
            if table:
                dbg["table"] = table
                dbg["reason"] = f"resolved by DropletSchedule.Marker='{m}' → QC.PN='{pn}' → table='{table}'"
                return table, dbg

    dbg["reason"] = "cannot resolve by marker→qc→pn, fallback to find_table_contains_workorder()"
    return None, dbg


@app.get("/api/search_571_tables")
def api_search_571_tables():
    work_order = request.args.get("work_order", "").strip()

    if not work_order:
        return jsonify({"ok": False, "message": "No work order provided"}), 400

    try:
        with connect() as conn:
            # 🛑 修正重點 1：
            # 在這裡還不要設定 conn.row_factory = dict_factory
            # 讓 find_table_contains_workorder 使用它習慣的預設模式 (Tuple) 運作

            # 1. 先找出是哪一張表
            target_table = find_table_contains_workorder(conn, work_order)

            if target_table:
                # ✅ 修正重點 2：
                # 確定找到表之後，現在我們才把模式切換成「字典模式」，
                # 這樣等一下 SELECT * 抓出來的就會是前端要的 JSON 格式
                conn.row_factory = dict_factory

                # 2. 為了保險，先抓取該表的欄位名稱，找出哪個是用來存工單的
                cursor = conn.execute(f"PRAGMA table_info({target_table})")
                columns = [info['name'] for info in cursor.fetchall()]

                # 判斷欄位名稱 (相容舊資料表)
                wo_col_name = None
                possible_names = ['工單號碼', 'WorkOrder', 'work_order', '工單編號']
                for name in possible_names:
                    if name in columns:
                        wo_col_name = name
                        break

                rows = []
                if wo_col_name:
                    # 3. 執行查詢 (這時候 conn.row_factory 已經生效，會回傳 dict)
                    sql = f"SELECT * FROM {target_table} WHERE {wo_col_name} = ?"
                    cursor = conn.execute(sql, (work_order,))
                    rows = cursor.fetchall()

                return jsonify({
                    "ok": True,
                    "table": target_table,
                    "rows": rows,
                    "message": f"Found in {target_table}"
                })
            else:
                return jsonify({
                    "ok": False,
                    "rows": [],
                    "table": None,
                    "message": "查無資料"
                })

    except Exception as e:
        # 顯示更詳細的錯誤，方便除錯
        traceback.print_exc()
        print(f"❌ api_search_571_tables error: {e}")
        return jsonify({"ok": False, "message": str(e)}), 500

# 補充：如果您的程式碼沒有這個 helper function，請補上
def dict_factory(cursor, row):
    d = {}
    for idx, col in enumerate(cursor.description):
        d[col[0]] = row[idx]
    return d

# ============================================================
# 整合型儲存 API (支援多個路徑與多種資料格式)
# ============================================================
@app.route("/api/save_workorder", methods=["POST", "OPTIONS"])
@app.route("/api/update_571_table", methods=["POST", "OPTIONS"])
def unified_save_api():
    # 1. CORS 預檢
    if request.method == "OPTIONS":
        return Response(status=200)

    try:
        payload = request.get_json(force=True, silent=True) or {}

        # 2. 相容性解析：找出 table, work_order 與要更新的 rows
        # 優先從 payload 抓取，如果沒有則從 nested 結構中提取
        work_order = (payload.get("workOrderNo") or payload.get("work_order") or "").strip()
        table = payload.get("table", "").strip()

        # 如果沒有傳 table 名稱，自動去資料庫尋找
        with connect() as conn:
            if not table and work_order:
                table = find_table_contains_workorder(conn, work_order)

            if not table:
                return jsonify({"ok": False, "message": f"無法定位工單 {work_order} 的資料表"}), 404

            # 準備要更新的資料列
            # 如果是 update_571_table 格式，資料在 rows
            # 如果是 save_workorder 格式，資料在 beads
            rows_to_update = payload.get("rows") or payload.get("beads") or []

            # 取得額外資訊 (針對 save_workorder 格式)
            reagent = payload.get("reagent", {})
            buffer = payload.get("bufferBase", {})
            prepared_by = reagent.get("preparedBy", "")

            # 3. 取得資料庫欄位資訊
            cur = conn.execute(f'PRAGMA table_info("{table}")')
            existing_cols = {r[1] for r in cur.fetchall()}

            # 4. 開始更新
            updated_count = 0
            # 先取得該工單所有 rowid，按順序對應前端的 beads 陣列
            cur_rows = conn.execute(
                f'SELECT rowid FROM "{table}" WHERE TRIM("工單號碼") = ? ORDER BY rowid',
                (work_order,)
            )
            db_rowids = [r[0] for r in cur_rows.fetchall()]

            for idx, row_data in enumerate(rows_to_update):
                material_val = (row_data.get("beadPN") or row_data.get("料號") or row_data.get("material") or "")

                update_pairs = []
                update_args = []

                # 定義前端與資料庫的欄位對應
                mapping = {
                    "beadName": "化學品名",
                    "totalQty": "總重量",
                    "qtyPerBead": "重量紀錄",
                    "lotNo": "Filler_Lot",
                    "remark": "配製備註"
                }

                # (A) 處理通用 rows 的內容
                for k, v in row_data.items():
                    db_col = mapping.get(k, k)
                    if db_col in existing_cols and db_col not in ("工單號碼", "料號"):
                        update_pairs.append(f'"{db_col}" = ?')
                        update_args.append(v)

                # (B) 如果是 save_workorder 格式，額外塞入 buffer 與人員資訊
                extra_data = {
                    "L1OD": buffer.get("L1OD"),
                    "L2OD": buffer.get("L2OD"),
                    "起始L1OD": buffer.get("L1StartOD"),
                    "起始L2OD": buffer.get("L2StartOD"),
                    "配製人員": prepared_by
                }
                for col, val in extra_data.items():
                    if col in existing_cols and val is not None:
                        # 避免重複加入
                        if f'"{col}" = ?' not in update_pairs:
                            update_pairs.append(f'"{col}" = ?')
                            update_args.append(val)

                if not update_pairs: continue

                # (C) 執行 SQL - 優先用 rowid 匹配，確保無料號的 row 也能更新
                if idx < len(db_rowids):
                    sql = f'UPDATE "{table}" SET {", ".join(update_pairs)} WHERE rowid = ?'
                    update_args.append(db_rowids[idx])
                else:
                    sql = f'UPDATE "{table}" SET {", ".join(update_pairs)} ' \
                          f'WHERE TRIM("工單號碼") = ? AND TRIM("料號") = ?'
                    update_args.extend([work_order, str(material_val).strip()])

                res = conn.execute(sql, update_args)
                updated_count += res.rowcount

            conn.commit()
            print(f"✅ [Unified Save] 表: {table} | 工單: {work_order} | 更新筆數: {updated_count}")

        return jsonify({"ok": True, "message": f"成功更新 {updated_count} 筆資料", "table": table})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "message": str(e)}), 500

# ==================== 新增：Heatmap 統計 API ====================
from collections import defaultdict

@app.get("/api/heatmap_usage")
def api_heatmap_usage():
    try:
        mode = request.args.get("mode", "week").strip().lower()
        top_n = int(request.args.get("top", "24"))
    except:
        top_n = 24

    # 定義時間桶 (Bucket) 轉換邏輯
    def get_bucket(date_str, mode):
        try:
            # 嘗試解析日期，支援常見格式
            for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y%m%d"):
                try:
                    dt = datetime.datetime.strptime(date_str, fmt)
                    break
                except ValueError:
                    continue
            else:
                return None # 解析失敗

            if mode == "month":
                return dt.strftime("%Y-%m") # 2024-01
            elif mode == "quarter":
                q = (dt.month - 1) // 3 + 1
                return f"{dt.year}-Q{q}"    # 2024-Q1
            else: # week (default)
                # ISO 週曆: (year, week, weekday)
                y, w, _ = dt.isocalendar()
                return f"{y}-W{w:02d}"      # 2024-W05
        except:
            return None

    usage_map = defaultdict(lambda: defaultdict(int)) # {marker: {bucket: count}}
    all_buckets = set()
    marker_totals = defaultdict(int)

    with connect() as conn:
        # 讀取 DropletSchedule 表中的 Marker 和 Date
        # 排除 Marker 為空或 Date 為空的資料
        sql = f'SELECT Marker, "Date" FROM "{DROPLET_TABLE}" WHERE Marker IS NOT NULL AND "Date" IS NOT NULL'
        try:
            cur = conn.execute(sql)
            rows = cur.fetchall()
        except sqlite3.OperationalError:
            # 如果表不存在
            return jsonify({
                "markers": [], "buckets": [], "matrix": [], "mode": mode
            })

        for r in rows:
            marker = (r["Marker"] or "").strip()
            date_str = (r["Date"] or "").strip()

            if not marker or not date_str:
                continue

            # 統一 Marker 名稱 (轉大寫)
            marker = marker.upper()

            bucket = get_bucket(date_str, mode)
            if bucket:
                usage_map[marker][bucket] += 1
                marker_totals[marker] += 1
                all_buckets.add(bucket)

    # 1. 篩選 Top N 的 Marker
    sorted_markers = sorted(marker_totals.keys(), key=lambda k: marker_totals[k], reverse=True)[:top_n]

    # 2. 排序 Buckets (時間軸)
    sorted_buckets = sorted(list(all_buckets))

    # 3. 建立 Matrix (二維陣列) [marker_index][bucket_index]
    # Nivo Heatmap 格式其實是 Object array，但在 API 回傳矩陣較省流量，前端再組裝
    # 這裡我們依照前端 format 的邏輯：
    # matrix[i][j] 對應 sorted_markers[i] 在 sorted_buckets[j] 的數值

    matrix = []
    for m in sorted_markers:
        row_data = []
        for b in sorted_buckets:
            val = usage_map[m].get(b, 0)
            row_data.append(val)
        matrix.append(row_data)

    return jsonify({
        "markers": sorted_markers,
        "buckets": sorted_buckets,
        "matrix": matrix,
        "mode": mode
    })

@app.get("/api/debug_find_workorder")
def debug_find_workorder():
    wo = request.args.get("work_order", "").strip()
    if not wo:
        return jsonify({"error": "missing work_order"}), 400

    results = []
    with connect() as conn:
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()]

        for t in tables:
            try:
                # 先確認有沒有工單號碼欄位
                cols = [r[1] for r in conn.execute(f'PRAGMA table_info("{t}")').fetchall()]
                if "工單號碼" not in cols:
                    continue
                hit = conn.execute(
                    f'SELECT 1 FROM "{t}" WHERE TRIM(工單號碼) LIKE ? LIMIT 1',
                    (f"%{wo}%",)
                ).fetchone()
                if hit:
                    results.append(t)
            except Exception as e:
                results.append(f"[ERROR] {t}: {e}")

    return jsonify({"work_order": wo, "found_in_tables": results})
# ==================== MAIN ====================
if __name__ == "__main__":
    local_ip = socket.gethostbyname(socket.gethostname())
    print("🚀 Flask 啟動成功！")
    print(f"🌐 同網段裝置可訪問： http://{local_ip}:5011")
    print(f"🧠 本機： http://localhost:5011")
    port = 5011  # ✅ 單一 port
    app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False)
