# -*- coding: utf-8 -*-
# app_unified.py — 單一 Flask（port 5011）
# DropletSchedule(日期/時間修正) + Excel 巨集/兩表導入 + CSV 相容上傳 + WorkOrder 查詢/儲存 + QR
# v3 — pg_dual_write 改為 HTTP POST 到 EC2 relay（本機不直連 RDS）

from flask import Flask, request, jsonify, Response, send_file
from flask_cors import CORS
from pathlib import Path
from openpyxl import load_workbook
import sqlite3, csv, io, os, re, json, datetime, time, logging, uuid, shutil, pythoncom, socket
import requests as http_requests   # 用於 pg_dual_write → EC2 relay
import qrcode
from qrcode.constants import ERROR_CORRECT_M
from win32com.client import Dispatch

# ==================== CONFIG ====================
DB_PATH = os.environ.get("DB_PATH", r"D:\配藥表\資料庫\P01_formualte_schedule.db")
ALLOW_ORIGINS = ["*"]

# DropletSchedule 專用
DROPLET_TABLE = "DropletSchedule"
HEADER_MAP = {
    "Pump": "Pump",
    "Marker": "Marker",
    "凍乾機台": "Lyophilizer",
    "可用凍乾機": "AvailableLyophilizer",
    "數量": "Quantity",
    "配藥同仁": "Preparer",
    "日期": "Date",
    "RD給藥時間": "DrugGivenAt",
    "預計滴定時間": "ExpectedTitrationStart",
    "預計結束": "ExpectedTitrationEnd",
    "工單編號": "WorkOrder",
    "Lot": "Lot",
    "備註": "Remark",
}

# Excel/CSV/模板
APP_CONF = {
    "DB_PATH": DB_PATH,
    "TEMPLATE_XLSM": os.environ.get("TEMPLATE_XLSM", r"D:\配藥表\配藥紀錄\配藥紀錄Temp.xlsm"),
    "UPLOAD_DIR": os.environ.get("UPLOAD_DIR", r"D:\配藥表\_temp"),
    "HEADER_ROW": int(os.environ.get("HEADER_ROW", "8")),
    "COL_BOM_FALLBACK": int(os.environ.get("COL_BOM_FALLBACK", "2")),
    "COL_TOTAL_FALLBACK": int(os.environ.get("COL_TOTAL_FALLBACK", "16")),
    "COL_LOT_FALLBACK": int(os.environ.get("COL_LOT_FALLBACK", "19")),
    "DEFAULT_SERVER_DIR": os.environ.get("DEFAULT_SERVER_DIR", r"\\fls341\Reagent RD\配藥端 -配製紀錄表\\"),
}

# ==================== EC2 Relay CONFIG ====================
EC2_RELAY_URL   = os.environ.get(
    "EC2_RELAY_URL",
    "http://54.250.30.179:5012/api/relay_droplet"
)
EC2_RELAY_TOKEN = os.environ.get("EC2_RELAY_TOKEN", "change-me")
PG_ENABLED      = os.environ.get("PG_ENABLED", "1") == "1"   # 設 "0" 可臨時停用

# ==================== APP INIT ====================
app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": ALLOW_ORIGINS}})
logging.basicConfig(level=logging.INFO)
log = logging.getLogger("unified")

Path(APP_CONF["UPLOAD_DIR"]).mkdir(parents=True, exist_ok=True)
Path(DB_PATH).parent.mkdir(parents=True, exist_ok=True)

# ==================== DB CONNECT（WAL + busy_timeout + thread-safe） ====================
def connect():
    conn = sqlite3.connect(DB_PATH, timeout=20, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.execute("PRAGMA busy_timeout = 15000;")
    return conn

# ==================== EC2 Relay：雙寫 RDS ====================
def pg_dual_write(sqlite_cols: list, clean_rows: list, involved_dates: set):
    """
    透過 EC2 relay API 寫入 RDS PostgreSQL。
    判斷基準：col.Date
      - 相同 Date 已存在 → EC2 端 DELETE 舊資料再 INSERT
      - 不存在           → EC2 端直接 INSERT（append）
    本機不直連 RDS，HTTP POST 失敗只記 log，不影響 SQLite 主流程。
    """
    if not PG_ENABLED:
        return

    payload = {
        "dates": sorted(involved_dates),
        "cols":  sqlite_cols,
        "rows":  clean_rows,
    }

    try:
        resp = http_requests.post(
            EC2_RELAY_URL,
            json=payload,
            headers={
                "Content-Type":  "application/json",
                "X-Relay-Token": EC2_RELAY_TOKEN,
            },
            timeout=15,
        )
        if resp.ok:
            result = resp.json()
            log.info(
                f"[PG relay] OK — deleted={result.get('deleted')}, "
                f"inserted={result.get('inserted')}, "
                f"dates={result.get('dates')}"
            )
        else:
            log.error(f"[PG relay] FAILED {resp.status_code}: {resp.text}")

    except Exception as e:
        log.error(f"[PG relay] request exception (SQLite unaffected): {e}")

# ========= 共用小工具 =========
def remove_cn_spaces(s):
    if s is None:
        return ""
    return str(s).replace("\u3000", "").replace("\xa0", "").strip()

def sanitize_for_ident(name):
    s = remove_cn_spaces(name).replace(" ", "")
    s = re.sub(r"[^0-9A-Za-z_]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    if not s:
        return "col"
    if s[0].isdigit():
        s = "_" + s
    return s

def zh_header_to_en(h):
    base = remove_cn_spaces(h).replace(" ", "")
    if base in HEADER_MAP:
        return sanitize_for_ident(HEADER_MAP[base])
    return sanitize_for_ident(base)

def normalize_headers(headers):
    seen, out = set(), []
    for idx, h in enumerate(headers):
        name = zh_header_to_en(h) or f"col_{idx+1}"
        orig = name
        i = 2
        while name in seen:
            name = f"{orig}_{i}"
            i += 1
        seen.add(name)
        out.append(name)
    return out

def normalize_table_name(sheet_name: str) -> str:
    name = re.sub(r'\s+', '_', (sheet_name or "").strip())
    name = re.sub(r'[^\w]', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    return name or "sheet"

def qident(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'

def safe_table_name(name: str) -> str:
    if not name:
        raise ValueError("table is required")
    if not re.fullmatch(r"[A-Za-z0-9_]+", name):
        raise ValueError(f"invalid table name: {name}")
    return name

# ========= DropletSchedule 專用：建表/補欄 =========
def ensure_droplet_table_and_columns(conn, headers):
    cols = normalize_headers(headers)
    cur = conn.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=? LIMIT 1", (DROPLET_TABLE,))
    exists = cur.fetchone() is not None
    if not exists:
        col_defs = ", ".join(f'"{c}" TEXT' for c in cols)
        conn.execute(f'CREATE TABLE "{DROPLET_TABLE}" ({col_defs});')
        conn.commit()
        return cols
    existing = {row[1] for row in conn.execute(f'PRAGMA table_info("{DROPLET_TABLE}")')}
    for c in cols:
        if c not in existing:
            conn.execute(f'ALTER TABLE "{DROPLET_TABLE}" ADD COLUMN "{c}" TEXT;')
    conn.commit()
    return cols

# ========= CSV parse =========
def parse_csv_text(text):
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        return [], []
    headers = [remove_cn_spaces(h) for h in rows[0]]
    return headers, rows[1:]

def parse_csv_file(file_storage):
    text = file_storage.read().decode("utf-8-sig", errors="replace")
    return parse_csv_text(text)

# ========= DropletSchedule：日期/時間修正 =========
def fix_date_field(rows, headers):
    if "Date" not in headers:
        return rows
    idx = headers.index("Date")
    fixed = []
    for r in rows:
        r = list(r)
        val = remove_cn_spaces(r[idx])
        if re.fullmatch(r"\d{4,6}", val or ""):
            try:
                base = datetime.datetime(1899, 12, 30)
                d = base + datetime.timedelta(days=int(val))
                r[idx] = d.strftime("%Y/%m/%d")
            except Exception:
                r[idx] = val
        else:
            parsed = None
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%m/%d/%Y"):
                try:
                    parsed = datetime.datetime.strptime(val, fmt)
                    break
                except Exception:
                    continue
            r[idx] = parsed.strftime("%Y/%m/%d") if parsed else val
        fixed.append(r)
    return fixed

def fix_time_fields(rows, headers):
    time_fields = {"DrugGivenAt", "ExpectedTitrationStart", "ExpectedTitrationEnd"}
    idxs = {h: i for i, h in enumerate(headers) if h in time_fields}
    if not idxs:
        return rows
    fixed = []
    for r in rows:
        r = list(r)
        for h, i in idxs.items():
            val = str(r[i] or "").strip()
            if not val:
                continue
            try:
                f = float(val)
                if 0 <= f < 1:
                    seconds = int(round(f * 86400))
                    hh = seconds // 3600
                    mm = (seconds % 3600) // 60
                    r[i] = f"{hh:02d}:{mm:02d}"
                    continue
            except Exception:
                pass
            try:
                if "1899" in val or "1900" in val:
                    dt = datetime.datetime.fromisoformat(val)
                    r[i] = dt.strftime("%H:%M")
                    continue
            except Exception:
                pass
            parsed = None
            for fmt in ("%H:%M", "%I:%M %p", "%H:%M:%S"):
                try:
                    parsed = datetime.datetime.strptime(val, fmt)
                    break
                except Exception:
                    continue
            if parsed:
                r[i] = parsed.strftime("%H:%M")
        fixed.append(r)
    return fixed

# ==========================================
# Sync Trigger (Flask side)
# ==========================================
import subprocess
import sys

SYNC_SCRIPT = Path(
    r"D:\OneDrive - 天亮醫療器材股份有限公司\.vscode\Bead-record AN\WebApp\dropfreeze\sync_droplet_record.py"
)
SYNC_LOCK = Path(r"D:\temp\droplet_sync.lock")

def trigger_droplet_sync_safe():
    if SYNC_LOCK.exists():
        return
    try:
        SYNC_LOCK.parent.mkdir(parents=True, exist_ok=True)
        lock_token = f"{time.time()}|{uuid.uuid4()}"
        SYNC_LOCK.write_text(lock_token)
        subprocess.Popen(
            [sys.executable, str(SYNC_SCRIPT)],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
    except Exception as e:
        print(f"⚠️ Sync trigger failed: {e}")

# ==================== 健康檢查 / Root ====================
@app.get("/api/health")
def health():
    return {"ok": True, "db": DB_PATH, "pg_enabled": PG_ENABLED, "ec2_relay": EC2_RELAY_URL}

@app.get("/")
def root():
    return "Backend OK. Try /api/health.", 200

# ==================== DropletSchedule：上傳/預覽 ====================
@app.post("/api/upload_droplet_schedule")
def upload_droplet_schedule():
    # 1. 接收與解析請求
    headers, rows = [], []
    ctype = (request.content_type or "").lower()

    if "multipart/form-data" in ctype:
        if "file" not in request.files:
            return Response("missing file", 400)
        headers, rows = parse_csv_file(request.files["file"])

    elif "application/json" in ctype:
        payload = request.get_json(silent=True) or {}
        headers = [remove_cn_spaces(h) for h in (payload.get("headers") or [])]
        rows = payload.get("rows") or []

    elif "text/csv" in ctype or request.data:
        data = request.get_data()
        if not data:
            return Response("empty request body", 400)
        try:
            text = data.decode("utf-8-sig", errors="replace")
        except UnicodeDecodeError:
            return Response("CSV decoding failed.", 400)
        headers, rows = parse_csv_text(text)

    else:
        return Response("unsupported content-type", 415)

    if not headers or not rows:
        return Response("empty headers or data", 400)

    try:
        with connect() as conn:
            # 2. 確保資料表結構
            cols = ensure_droplet_table_and_columns(conn, headers)
            target_date_col = "Date"

            # 3. 資料預處理
            try:
                rows = fix_date_field(rows, cols)
                rows = fix_time_fields(rows, cols)

                if target_date_col not in cols:
                    return Response(f"Missing required column: {target_date_col}", 400)

                date_col_index = cols.index(target_date_col)
                clean_rows     = []
                involved_dates = set()
                col_count      = len(cols)

                for r in rows:
                    row_data = [remove_cn_spaces(str(v)) for v in r]
                    if len(row_data) < col_count:
                        row_data += [""] * (col_count - len(row_data))
                    final_row = row_data[:col_count]
                    clean_rows.append(final_row)
                    date_val = final_row[date_col_index]
                    if date_val:
                        involved_dates.add(date_val)

                if not involved_dates:
                    return Response("No valid dates found in upload data.", 400)

            except Exception as e:
                app.logger.error(f"Data processing error: {e}")
                return Response(f"Data processing failed: {e}", 400)

            # 4. SQLite 寫入（delete by date → insert）
            placeholders_dates = ", ".join("?" for _ in involved_dates)
            delete_sql = f'DELETE FROM "{DROPLET_TABLE}" WHERE "{target_date_col}" IN ({placeholders_dates})'
            placeholders_vals = ", ".join("?" for _ in cols)
            collist    = ", ".join(f'"{c}"' for c in cols)
            insert_sql = f'INSERT INTO "{DROPLET_TABLE}" ({collist}) VALUES ({placeholders_vals})'

            for attempt in range(5):
                try:
                    conn.execute(delete_sql, list(involved_dates))
                    conn.executemany(insert_sql, clean_rows)
                    conn.commit()

                    trigger_droplet_sync_safe()

                    # ===== 雙寫 RDS（透過 EC2 relay，失敗不阻擋主流程）=====
                    try:
                        pg_dual_write(cols, clean_rows, involved_dates)
                    except Exception as pg_err:
                        log.error(f"[PG relay] exception: {pg_err}")

                    break

                except sqlite3.OperationalError as e:
                    if "locked" in str(e).lower() and attempt < 4:
                        conn.rollback()
                        time.sleep(0.3)
                        continue
                    else:
                        raise
                except Exception:
                    conn.rollback()
                    raise

        return Response(status=204)

    except sqlite3.Error as e:
        app.logger.error(f"Database Error: {e}")
        return Response(f"Database operation failed: {e}", 500)
    except Exception as e:
        app.logger.error(f"Unexpected Server Error during upload: {e}")
        return Response("Internal Server Error during data processing.", 500)


@app.get("/api/sync_status")
def api_sync_status():
    with connect() as conn:
        row = conn.execute(
            "SELECT status, start_time, end_time, duration_sec, last_error FROM sync_status WHERE id=1"
        ).fetchone()
        if not row:
            return jsonify({"status": "unknown"})
        return jsonify(dict(row))

@app.get("/api/preview")
def preview():
    limit = int(request.args.get("limit", "50"))
    with connect() as conn:
        cur = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?;", (DROPLET_TABLE,))
        if cur.fetchone() is None:
            return jsonify({"columns": [], "rows": []})
        cols = [r[1] for r in conn.execute(f'PRAGMA table_info("{DROPLET_TABLE}")')]
        cur  = conn.execute(f'SELECT * FROM "{DROPLET_TABLE}" LIMIT {limit}')
        rows = [dict(r) for r in cur.fetchall()]
        return jsonify({"columns": cols, "rows": rows})

# ===========================
# 日期格式規範化
# ===========================
def normalize_date(date_str: str) -> str:
    if not date_str:
        return None
    try:
        return datetime.datetime.strptime(date_str, '%Y-%m-%d').strftime('%Y-%m-%d')
    except ValueError:
        pass
    try:
        return datetime.datetime.strptime(date_str, '%Y/%m/%d').strftime('%Y-%m-%d')
    except ValueError:
        pass
    return None

def generate_date_formats(normalized_date: str) -> list:
    parts = normalized_date.split('-')
    year, month, day = parts[0], parts[1], parts[2]
    month_no_zero = str(int(month))
    day_no_zero   = str(int(day))
    return [
        f"{year}-{month_no_zero}-{day_no_zero}",
        f"{year}/{month_no_zero}/{day_no_zero}",
        f"{year}-{month}-{day}",
        f"{year}/{month}/{day}",
    ]

@app.route('/api/droplet-records', methods=['GET'])
def get_droplet_records():
    date_param = request.args.get('date')
    if not date_param:
        return jsonify({'error': '缺少 date 參數'}), 400
    normalized_date = normalize_date(date_param)
    if not normalized_date:
        return jsonify({'error': '日期格式不正確'}), 400
    date_formats = generate_date_formats(normalized_date)
    try:
        con = connect()
        try:
            cur = con.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", ("dropletRecord",)
            )
            if not cur.fetchone():
                return jsonify({'error': '表不存在'}), 500
            placeholders = ','.join(['?' for _ in date_formats])
            cur     = con.execute(
                f'SELECT * FROM "dropletRecord" WHERE "record_date" IN ({placeholders}) ORDER BY "record_date" DESC',
                date_formats
            )
            cols    = [d[0] for d in cur.description]
            records = [dict(zip(cols, row)) for row in cur.fetchall()]
            return jsonify({'rows': records}), 200
        finally:
            con.close()
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': '資料庫查詢失敗', 'detail': str(e)}), 500

# ==================== Excel 宏 / 兩表導入 DB ====================
TEMP_IN_MAP:  dict[str, Path] = {}
TEMP_OUT_MAP: dict[str, Path] = {}

def ensure_table_and_columns(con: sqlite3.Connection, table: str, headers: list[str]) -> None:
    cur = con.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?;", (table,))
    exists = cur.fetchone() is not None
    if not exists:
        cols_def = ", ".join(f'{qident(h)} TEXT' for h in headers)
        con.execute(f'CREATE TABLE {qident(table)} ({cols_def});')
        return
    cur = con.execute(f'PRAGMA table_info({qident(table)});')
    existing = {r[1] for r in cur.fetchall()}
    for h in headers:
        if h not in existing:
            con.execute(f'ALTER TABLE {qident(table)} ADD COLUMN {qident(h)} TEXT;')

def upsert_rows(con: sqlite3.Connection, table: str, headers: list[str], rows: list[tuple], pk_cols: list[str]) -> None:
    cols   = ", ".join(qident(h) for h in headers)
    qmarks = ", ".join(["?"] * len(headers))
    if pk_cols:
        upd_cols = [h for h in headers if h not in pk_cols]
        if upd_cols:
            upd = ", ".join(f'{qident(c)}=excluded.{qident(c)}' for c in upd_cols)
            sql = (f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks}) '
                   f'ON CONFLICT({", ".join(qident(c) for c in pk_cols)}) DO UPDATE SET {upd};')
        else:
            sql = (f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks}) '
                   f'ON CONFLICT({", ".join(qident(c) for c in pk_cols)}) DO NOTHING;')
    else:
        sql = f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks});'
    for attempt in range(5):
        try:
            con.executemany(sql, rows)
            return
        except sqlite3.OperationalError as e:
            if "locked" in str(e).lower() and attempt < 4:
                time.sleep(0.3); continue
            raise

def read_sheet_to_db(con: sqlite3.Connection, ws, table_name: str, header_row: int) -> dict:
    max_col = ws.max_column or 1
    headers = [str(ws.cell(header_row, c).value or "").strip() for c in range(1, max_col + 1)]
    headers = [h if h else f"col_{i+1}" for i, h in enumerate(headers)]
    rows = []
    r0   = header_row + 1
    max_r = ws.max_row or r0
    for rr in range(r0, max_r + 1):
        row, nonempty = [], False
        for c in range(1, max_col + 1):
            v = ws.cell(rr, c).value
            if v not in (None, ""): nonempty = True
            row.append("" if v is None else str(v))
        if nonempty:
            rows.append(tuple(row))
    ensure_table_and_columns(con, table_name, headers)
    pk_cols = []
    if "工單號碼" in headers and "料號" in headers:
        pk_cols = ["工單號碼", "料號"]
    elif "工單號碼" in headers:
        pk_cols = ["工單號碼"]
    upsert_rows(con, table_name, headers, rows, pk_cols)
    return {"table": table_name, "headers": headers, "rows": len(rows), "pk": pk_cols}

def run_excel_macro(xlsm_path: str, macro_name: str) -> tuple[bool, str | None]:
    try:
        pythoncom.CoInitialize()
        excel = Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            excel.AutomationSecurity = 1
        except Exception:
            pass
        wb = excel.Workbooks.Open(xlsm_path)
        try:
            excel.Run(macro_name)
            wb.Save()
            return True, None
        except Exception as e:
            return False, f"Run macro failed: {e}"
        finally:
            try: wb.Close(SaveChanges=False)
            except Exception: pass
            excel.Quit()
    except Exception as e:
        return False, f"Excel automation error: {e}"
    finally:
        try: pythoncom.CoUninitialize()
        except Exception: pass

@app.post("/api/upload_excel")
def api_upload_excel():
    f = request.files.get("file")
    if not f:
        return jsonify(ok=False, message="missing file"), 400
    fname_raw = f.filename or "uploaded.xlsm"
    if "配藥表" not in fname_raw:
        return jsonify(ok=False, message="所選檔名未包含「配藥表」，請重新選擇正確的配藥表 Excel。"), 400
    ext = Path(fname_raw).suffix.lower() or ".xlsm"
    if ext not in (".xlsm", ".xlsx", ".xls"):
        ext = ".xlsm"
    temp_id      = uuid.uuid4().hex[:8]
    temp_in_path = Path(APP_CONF["UPLOAD_DIR"]) / f"upload_{temp_id}{ext}"
    f.save(str(temp_in_path))
    TEMP_IN_MAP[temp_id] = temp_in_path
    ok_macro, err_macro = run_excel_macro(str(temp_in_path), "manuinsert.InsertNewByWorkOrderArrays")
    if not ok_macro:
        return jsonify(ok=False, message=f"Excel 宏執行失敗：{err_macro}"), 500
    wb     = load_workbook(temp_in_path, keep_vba=True, data_only=True)
    sheets = wb.sheetnames[:2]
    ingested = []
    con = connect()
    try:
        with con:
            con.execute("PRAGMA temp_store=MEMORY;")
            for s in sheets:
                ws    = wb[s]
                table = normalize_table_name(s)
                info  = read_sheet_to_db(con, ws, table, APP_CONF["HEADER_ROW"])
                ingested.append(info)
    finally:
        con.close()
    return jsonify(ok=True, temp_id=temp_id, filename=fname_raw, sheets=sheets, ingested=ingested)

@app.post("/api/create_record")
def api_create_record():
    data       = request.get_json(force=True, silent=True) or {}
    work_order = (data.get("work_order") or "").strip()
    temp_id    = (data.get("temp_id") or "").strip()
    table_raw  = (data.get("table") or "").strip()
    if not work_order or not temp_id:
        return jsonify(ok=False, message="work_order & temp_id are required"), 400
    in_path = TEMP_IN_MAP.get(temp_id)
    if not in_path or not in_path.exists():
        return jsonify(ok=False, message=f"temp_id not found: {temp_id}"), 404
    wb_in       = load_workbook(in_path, keep_vba=True, data_only=True)
    first_sheet = normalize_table_name(wb_in.sheetnames[0]) if wb_in.sheetnames else None
    table       = normalize_table_name(table_raw) if table_raw else first_sheet
    if not table:
        return jsonify(ok=False, message="no sheet to use as table"), 400
    sql = f"""
      SELECT 工單號碼 AS work_order, BeadsLot AS bead_lot,
             料號 AS material, 重量紀錄 AS weight, Filler_Lot AS lot_no
      FROM {qident(table)} WHERE 工單號碼 = ? ORDER BY 料號
    """
    with connect() as conn:
        cur  = conn.execute(sql, (work_order,))
        rows = [dict(zip([c[0] for c in cur.description], r)) for r in cur.fetchall()]
    template = APP_CONF["TEMPLATE_XLSM"]
    if not os.path.exists(template):
        return jsonify(ok=False, message=f"template not found: {template}"), 500
    out_path = Path(APP_CONF["UPLOAD_DIR"]) / f"record_{temp_id}.xlsm"
    shutil.copy2(template, out_path)
    wb = load_workbook(out_path, keep_vba=True, data_only=False)
    if "製程紀錄表" not in wb.sheetnames:
        return jsonify(ok=False, message="template has no sheet: 製程紀錄表"), 500
    ws       = wb["製程紀錄表"]
    ws["V6"] = work_order
    HEADER_ROW         = APP_CONF["HEADER_ROW"]
    COL_BOM_FALLBACK   = APP_CONF["COL_BOM_FALLBACK"]
    COL_TOTAL_FALLBACK = APP_CONF["COL_TOTAL_FALLBACK"]
    COL_LOT_FALLBACK   = APP_CONF["COL_LOT_FALLBACK"]
    def norm(s): return (str(s or "")).strip().lower().replace(" ", "")
    title2col = {}
    max_col   = ws.max_column or 1
    for c in range(1, max_col + 1):
        title2col[norm(ws.cell(HEADER_ROW, c).value)] = c
    def get_col(names, fallback):
        for n in names:
            col = title2col.get(norm(n))
            if col: return col
        return fallback
    col_bom   = get_col(["bom p/n","bompn","料號","bom"], COL_BOM_FALLBACK)
    col_total = get_col(["total qty","totalqty","重量紀錄"], COL_TOTAL_FALLBACK)
    col_lot   = get_col(["lot no.","lotno","filler_lot","lot no","lotno."], COL_LOT_FALLBACK)
    start_row = HEADER_ROW + 1
    end_row   = ws.max_row if ws.max_row > start_row else start_row
    if end_row >= start_row:
        ws.delete_rows(start_row, end_row - start_row + 1)
    r = start_row
    for it in rows:
        ws.cell(r, col_bom).value   = it.get("material")
        ws.cell(r, col_total).value = it.get("weight")
        ws.cell(r, col_lot).value   = it.get("lot_no")
        r += 1
    wb.save(out_path)
    TEMP_OUT_MAP[temp_id] = out_path
    return jsonify(ok=True, filled=len(rows), out_path=str(out_path), table=table)

@app.get("/api/template_preview")
def api_template_preview():
    temp_id  = (request.args.get("temp_id") or "").strip()
    out_path = TEMP_OUT_MAP.get(temp_id)
    if not out_path or not Path(out_path).exists():
        return jsonify(ok=False, message=f"no generated file for temp_id: {temp_id}"), 404
    wb = load_workbook(out_path, keep_vba=True, data_only=True)
    if "製程紀錄表" not in wb.sheetnames:
        return jsonify(ok=False, message="template has no sheet: 製程紀錄表"), 500
    ws         = wb["製程紀錄表"]
    HEADER_ROW = APP_CONF["HEADER_ROW"]
    max_col    = ws.max_column or 1
    headers    = [str(ws.cell(HEADER_ROW, c).value or "") for c in range(1, max_col + 1)]
    rows  = []
    r     = HEADER_ROW + 1
    max_r = min(ws.max_row or r, r + 50)
    for rr in range(r, max_r + 1):
        row = [("" if ws.cell(rr, c).value is None else str(ws.cell(rr, c).value)) for c in range(1, max_col + 1)]
        rows.append(row)
    return jsonify(ok=True, headers=headers, rows=rows)

@app.get("/api/template_file")
def api_template_file():
    temp_id  = (request.args.get("temp_id") or "").strip()
    out_path = TEMP_OUT_MAP.get(temp_id)
    if not out_path or not Path(out_path).exists():
        return jsonify(ok=False, message=f"no generated file for temp_id: {temp_id}"), 404
    return send_file(
        out_path, as_attachment=True,
        download_name=Path(out_path).name,
        mimetype="application/vnd.ms-excel.sheet.macroEnabled.12",
    )

@app.post("/api/save_template")
def api_save_template():
    data       = request.get_json(force=True, silent=True) or {}
    temp_id    = (data.get("temp_id") or "").strip()
    work_order = (data.get("work_order") or "").strip()
    server_dir = (data.get("server_dir") or "").strip()
    filename   = (data.get("filename") or "").strip()
    out_src    = TEMP_OUT_MAP.get(temp_id)
    if not out_src or not Path(out_src).exists():
        return jsonify(ok=False, message=f"no generated file for temp_id: {temp_id}"), 404
    if not server_dir or not filename:
        return jsonify(ok=False, message="server_dir, filename are required"), 400
    if not filename.endswith(".xlsm"):
        filename += ".xlsm"
    if work_order:
        wb = load_workbook(out_src, keep_vba=True, data_only=False)
        if "製程紀錄表" in wb.sheetnames:
            wb["製程紀錄表"]["V6"] = work_order
            wb.save(out_src)
    out_dir  = Path(server_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(out_src, out_dir / filename)
    return jsonify(ok=True, saved_path=str(out_dir / filename))

# ==================== CSV 相容上傳 (raw/form-data) ====================
def _parse_csv_text(text: str):
    if text and text[:1] == '\ufeff':
        text = text.lstrip('\ufeff')
    reader = csv.reader(io.StringIO(text))
    try:
        headers = next(reader)
    except StopIteration:
        headers = []
    headers = [h.strip() if h else f"col_{i+1}" for i, h in enumerate(headers)]
    rows = []
    for r in reader:
        if len(r) < len(headers): r = r + [""] * (len(headers) - len(r))
        elif len(r) > len(headers): r = r[:len(headers)]
        rows.append(tuple(r))
    return headers, rows

def ensure_unique_index(con: sqlite3.Connection, table: str, pk_cols: list[str]) -> None:
    if not pk_cols:
        return
    idx  = "ux_" + re.sub(r'[^A-Za-z0-9_]', '_', table) + "__" + "__".join(
        re.sub(r'[^A-Za-z0-9_]', '_', c) for c in pk_cols
    )
    cols = ", ".join(qident(c) for c in pk_cols)
    con.execute(f'CREATE UNIQUE INDEX IF NOT EXISTS {qident(idx)} ON {qident(table)} ({cols});')

def _get_mode() -> str:
    mode = (request.form.get("mode") or request.args.get("mode") or "replace").strip().lower()
    return mode if mode in ("append", "replace", "upsert") else "replace"

def _want_csv_response() -> bool:
    return (request.args.get("return") or "").strip().lower() == "csv"

def _insert_rows(con: sqlite3.Connection, table: str, headers: list[str], rows: list[tuple]) -> int:
    if not rows: return 0
    cols   = ", ".join(qident(h) for h in headers)
    qmarks = ", ".join(["?"] * len(headers))
    con.executemany(f'INSERT INTO {qident(table)} ({cols}) VALUES ({qmarks});', rows)
    return len(rows)

def _replace_rows(con: sqlite3.Connection, table: str, headers: list[str], rows: list[tuple]) -> int:
    con.execute(f'DELETE FROM {qident(table)};')
    return _insert_rows(con, table, headers, rows)

def _handle_csv_upload(table: str, headers: list[str], rows: list[tuple], pk_cols: list[str], mode: str):
    con = connect()
    try:
        with con:
            con.execute("PRAGMA temp_store=MEMORY;")
            ensure_table_and_columns(con, table, headers)
            if not pk_cols:
                if "工單號碼" in headers and "料號" in headers:
                    pk_cols = ["工單號碼", "料號"]
                elif "工單號碼" in headers:
                    pk_cols = ["工單號碼"]
            if pk_cols:
                ensure_unique_index(con, table, pk_cols)
            if mode == "append":
                upsert_rows(con, table, headers, rows, pk_cols) if pk_cols else _insert_rows(con, table, headers, rows)
            elif mode == "upsert":
                upsert_rows(con, table, headers, rows, pk_cols) if pk_cols else _insert_rows(con, table, headers, rows)
            else:  # replace
                upsert_rows(con, table, headers, rows, pk_cols) if pk_cols else _replace_rows(con, table, headers, rows)
    finally:
        con.close()

@app.post("/api/upload_csv_raw")
@app.post("/upload_csv_raw")
def api_upload_csv_raw():
    table = (request.args.get("table") or "").strip()
    if not table:
        return jsonify(ok=False, message="missing table"), 400
    table = normalize_table_name(table)
    try: table = safe_table_name(table)
    except ValueError as e: return jsonify(ok=False, message=str(e)), 400
    pk_raw  = (request.args.get("pk") or "").strip()
    mode    = _get_mode()
    pk_cols = [c.strip() for c in pk_raw.split(",") if c.strip()]
    text    = request.get_data(cache=False, as_text=True)
    if not text:
        return jsonify(ok=False, message="empty body"), 400
    headers, rows = _parse_csv_text(text)
    try:
        _handle_csv_upload(table, headers, rows, pk_cols, mode)
    except Exception as e:
        app.logger.exception("upload_csv_raw failed")
        return Response("status,FAIL\n", status=500, mimetype="text/csv; charset=utf-8") if _want_csv_response() else jsonify(ok=False, message=str(e)), 500
    return Response("status,OK\n", mimetype="text/csv; charset=utf-8") if _want_csv_response() else jsonify(ok=True)

@app.post("/api/upload_csv")
@app.post("/upload_csv")
def api_upload_csv():
    table = (request.form.get("table") or request.args.get("table") or "").strip()
    if not table:
        return jsonify(ok=False, message="missing table"), 400
    table = normalize_table_name(table)
    try: table = safe_table_name(table)
    except ValueError as e: return jsonify(ok=False, message=str(e)), 400
    pk_raw  = (request.form.get("pk") or request.args.get("pk") or "").strip()
    mode    = _get_mode()
    pk_cols = [c.strip() for c in pk_raw.split(",") if c.strip()]
    f = request.files.get("file") or request.files.get("csv")
    if not f:
        return jsonify(ok=False, message="missing file"), 400
    raw = f.read()
    try: text = raw.decode("utf-8-sig", errors="replace")
    except Exception: text = raw.decode(errors="replace")
    headers, rows = _parse_csv_text(text)
    try:
        _handle_csv_upload(table, headers, rows, pk_cols, mode)
    except Exception as e:
        app.logger.exception("upload_csv failed")
        return Response("status,FAIL\n", status=500, mimetype="text/csv; charset=utf-8") if _want_csv_response() else jsonify(ok=False, message=str(e)), 500
    return Response("status,OK\n", mimetype="text/csv; charset=utf-8") if _want_csv_response() else jsonify(ok=True)

# ==================== QR 產生（從 xlsm cells） ====================
@app.post("/api/qr_png_from_cells")
def api_qr_png_from_cells():
    data      = request.get_json(force=True)
    cells     = data.get("cells") or []
    joiner    = (data.get("joiner") or "|")
    file_path = (data.get("file_path") or "").strip()
    temp_id   = (data.get("temp_id") or "").strip()
    if not (file_path or temp_id):
        return jsonify(ok=False, message="file_path or temp_id is required"), 400
    if not cells:
        return jsonify(ok=False, message="cells is required"), 400
    xlsm_path = file_path or str(TEMP_OUT_MAP.get(temp_id) or "")
    if not xlsm_path or not os.path.exists(xlsm_path):
        return jsonify(ok=False, message=f"file not found: {xlsm_path}"), 404
    wb   = load_workbook(xlsm_path, data_only=True, keep_vba=True)
    ws   = wb.active
    vals = []
    for addr in cells:
        if "!" in addr:
            sheet, cell = addr.split("!", 1)
            vals.append("" if wb[sheet][cell].value is None else str(wb[sheet][cell].value))
        else:
            vals.append("" if ws[addr].value is None else str(ws[addr].value))
    payload = joiner.join((v or "").strip() for v in vals)
    qr = qrcode.QRCode(version=None, error_correction=ERROR_CORRECT_M, box_size=10, border=2)
    qr.add_data(payload)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png")

# ==================== WorkOrder 查詢 / 儲存 ====================
def query_db(sql, args=()):
    with connect() as conn:
        cur = conn.execute(sql, args)
        return [dict(row) for row in cur.fetchall()]

def safe_float(v):
    try: return float(v) if v not in (None, "") else 0.0
    except: return 0.0

def safe_int(v):
    try: return int(float(v))
    except: return 0

def parse_range(text):
    if not text: return None
    m = re.match(r"([\d.]+)\s*[-~]\s*([\d.]+)", str(text))
    return (float(m.group(1)), float(m.group(2))) if m else None

def check_in_range(value, text_range):
    if not text_range: return None
    lo, hi = text_range
    return lo <= value <= hi

def norm_name(s: str) -> str:
    if not s: return ""
    s = re.sub(r"[-_].*$", "", str(s).strip())
    return s.strip().upper()

def find_table_contains_workorder(conn: sqlite3.Connection, work_order: str):
    work_order = work_order.strip()
    cur    = conn.execute("""
        SELECT name FROM sqlite_master WHERE type='table'
        AND (name LIKE '%571%' OR name IN ('RdtestURD2026_001', 'RdtestDRD2026_002'))
    """)
    tables = [r[0] for r in cur.fetchall()]
    like_pattern = f"%{work_order}%"
    for t in tables:
        try:
            if conn.execute(f"SELECT 1 FROM '{t}' WHERE TRIM(工單號碼) LIKE ? LIMIT 1", (like_pattern,)).fetchall():
                return t
        except Exception:
            pass
    return None

def check_suspension(v):    return "混濁" in str(v or "")
def check_yes(v):           return str(v or "").strip().lower() == "yes"
def check_not_room_temp(v):
    val = str(v or "").strip(); return bool(val) and val != "室溫"
def check_not_no(v):
    val = str(v or "").strip(); return bool(val) and val.lower() != "no"

@app.get("/api/get_workorder")
def api_get_workorder():
    work_order = request.args.get("work_order")
    if not work_order:
        return jsonify({"error": "缺少工單號碼參數"}), 400
    try:
        with connect() as conn:
            target_table = find_table_contains_workorder(conn, work_order)
            if not target_table:
                return jsonify({"message": f"查無工單 {work_order}"}), 404
            cur         = conn.execute(
                f"SELECT * FROM '{target_table}' WHERE TRIM(工單號碼) LIKE ?",
                (f"%{work_order.strip()}%",),
            )
            cols        = [d[0] for d in cur.description]
            result_rows = [dict(zip(cols, row)) | {"來源表": target_table} for row in cur.fetchall()]
            if not result_rows:
                return jsonify({"message": f"查無工單 {work_order} 的詳細記錄"}), 404
            first = result_rows[0]

            dispose_lots, maker_name, product_quantity = [], None, 0
            formulation_date = None
            for r in conn.execute(
                "SELECT Lot, Pump, Lyophilizer, Marker, Quantity, Date, Remark FROM DropletSchedule WHERE TRIM(WorkOrder) LIKE ?",
                (f"%{work_order.strip()}%",),
            ).fetchall():
                r = dict(r)
                maker_name = r.get("Marker") or maker_name
                pump_val = (r.get("Pump") or "").strip()
                lyo_val  = (r.get("Lyophilizer") or "").strip()
                lot_val  = (r.get("Lot") or "").strip()
                remark   = (r.get("Remark") or "").strip()
                if pump_val or lyo_val:
                    # 滴定日：有 Pump/Lyophilizer 的才計入製令數量和 disposeLots
                    product_quantity += safe_int(r.get("Quantity"))
                    dispose_lots.append({"id": lot_val, "port": pump_val, "freezeDry": lyo_val, "pump": None})
                else:
                    # 配藥日：Diluent 等無滴定/凍乾欄位的工單，數量仍來自排程表
                    product_quantity += safe_int(r.get("Quantity"))
                    if not formulation_date:
                        formulation_date = r.get("Date")
                    # 若 Lot 或 Remark 有批次資訊，仍建立 disposeLots 供標籤使用
                    if lot_val:
                        dispose_lots.append({"id": lot_val, "port": "", "freezeDry": "", "pump": None})
                    elif remark and "lot" in remark.lower():
                        # 從 Remark 解析 Lot（格式如 "Lot 211/211"）
                        lot_match = re.search(r"[Ll]ot\s*([\w/\-]+)", remark)
                        lot_id = lot_match.group(1) if lot_match else remark
                        dispose_lots.append({"id": lot_id, "port": "", "freezeDry": "", "pump": None})

            pump_ids = []
            if maker_name:
                for row in conn.execute("SELECT * FROM 'pump No.'").fetchall():
                    row = dict(row)
                    for i in range(1, 11):
                        val = row.get(f"可滴定之試劑-{i}")
                        if val and maker_name.lower() in str(val).strip().lower():
                            pump_ids.append(row.get("Pump編號")); break
            if pump_ids:
                for i, lot in enumerate(dispose_lots):
                    lot["pump"] = pump_ids[i % len(pump_ids)]

            match_571     = re.search(r"571\d{5,}", target_table or "")
            product_model = match_571.group(0) if match_571 else None
            base_name     = norm_name(maker_name or "")

            selected_qc = None
            if product_model:
                qc_rows = query_db("SELECT * FROM [Liquid form QC] WHERE PN = ?", (product_model,))
                if not qc_rows:
                    try:
                        qc_rows = query_db("SELECT * FROM [Liquid form QC] WHERE PN = ?", (str(int(float(product_model))),))
                    except: pass
                # ✅ PN 命中後，用 Marker name 嚴格過濾（區分 TG vs Hi TG 等）
                if qc_rows and maker_name:
                    mk_norm = norm_name(maker_name)
                    filtered = [r for r in qc_rows if norm_name(r.get("Marker name", "")) == mk_norm]
                    if filtered:
                        selected_qc = filtered[0]
                    # Marker name 不符時不 fallback，繼續往下走策略 B/C
                elif qc_rows:
                    selected_qc = qc_rows[0]
            # 策略 B: 用完整原始名稱精確匹配 Marker name（不經 norm_name 截斷）
            if not selected_qc:
                raw_name = (maker_name or "").strip().upper()
                for row in query_db("SELECT * FROM [Liquid form QC]"):
                    db_marker_raw = (row.get("Marker name") or "").strip().upper()
                    if db_marker_raw and db_marker_raw == raw_name:
                        selected_qc = row; break
            # 策略 C: 用完整原始名稱精確匹配 Name
            if not selected_qc:
                raw_name = (maker_name or "").strip().upper()
                for row in query_db("SELECT * FROM [Liquid form QC]"):
                    db_name_raw = (row.get("Name") or "").strip().upper()
                    if db_name_raw and db_name_raw == raw_name:
                        selected_qc = row; break
            # 策略 D: fallback 用 norm_name 模糊匹配
            if not selected_qc:
                for row in query_db("SELECT * FROM [Liquid form QC]"):
                    db_marker = norm_name(row.get("Marker name", ""))
                    if db_marker and db_marker == base_name:
                        selected_qc = row; break
            if not selected_qc:
                for row in query_db("SELECT * FROM [Liquid form QC]"):
                    db_name = norm_name(row.get("Name", ""))
                    if db_name and (db_name == base_name or db_name in base_name or base_name in db_name):
                        selected_qc = row; break
            if not selected_qc:
                selected_qc = {"L1-OD": None, "L2-OD": None, "L1-起始OD": None, "L2-起始OD": None, "懸浮物": None, "版本": None}

            selected_titration = {}
            for tname in ["適定條件", "滴定條件"]:
                try:
                    if not conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (tname,)).fetchone():
                        continue
                    t_cols = [r[1] for r in conn.execute(f"PRAGMA table_info('{tname}')").fetchall()]
                    if "PN" in t_cols and product_model:
                        result = query_db(f"SELECT * FROM [{tname}] WHERE PN = ?", (product_model,))
                        if result: selected_titration = result[0]; break
                    if "Name" in t_cols and base_name:
                        result = query_db(f"SELECT * FROM [{tname}] WHERE Name = ?", (base_name,))
                        if result: selected_titration = result[0]; break
                except Exception: continue

            confirm_data = {
                "suspension": check_suspension(selected_qc.get("懸浮物")),
                "storeLight": check_yes(selected_titration.get("儲存時避光")),
                "storeIce":   check_not_room_temp(selected_titration.get("儲存時冰浴")),
                "dyeing":     check_yes(selected_titration.get("滴定時避光")),
                "washing":    check_yes(selected_titration.get("滴定時冰浴")),
                "stir":       check_not_no(selected_titration.get("滴定時攪拌") or selected_titration.get("滴定_Mixing")),
            }
            L1OD      = safe_float(first.get("L1-OD") or first.get("L1OD") or first.get("L1 OD"))
            L2OD      = safe_float(first.get("L2-OD") or first.get("L2OD") or first.get("L2 OD"))
            L1StartOD = safe_float(first.get("L1-起始OD") or first.get("起始L1OD") or first.get("L1StartOD"))
            L2StartOD = safe_float(first.get("L2-起始OD") or first.get("起始L2OD") or first.get("L2StartOD"))
            qc_map    = {"L1OD":"L1-OD","L2OD":"L2-OD","L1StartOD":"L1-起始OD","L2StartOD":"L2-起始OD"}
            values    = {"L1OD": L1OD, "L2OD": L2OD, "L1StartOD": L1StartOD, "L2StartOD": L2StartOD}
            check_results = {
                k: {"value": values[k], "qc_range": selected_qc.get(qc_col),
                    "pass": check_in_range(values[k], parse_range(selected_qc.get(qc_col)))}
                for k, qc_col in qc_map.items()
            }
            return jsonify({
                "workOrderNo": work_order, "productModel": product_model or target_table,
                "markerName":  maker_name or (target_table or "").split("_")[0],
                "productQuantity": product_quantity, "date": formulation_date or first.get("試劑配製日期",""),
                "version": selected_qc.get("版本",""),
                "beads": [{"beadName": r.get("化學品名",""), "beadPN": r.get("料號",""), "unit": "g",
                            "qtyPerBead": r.get("總重量",0), "totalQty": r.get("重量紀錄",0), "lotNo": r.get("Filler_Lot","")}
                           for r in result_rows],
                "reagent":       {"preparedBy": first.get("配製人員",""), "confirm": confirm_data},
                "bufferBase":    {"L1OD": L1OD, "L2OD": L2OD, "L1StartOD": L1StartOD, "L2StartOD": L2StartOD},
                "qcRanges":      {"L1-OD": selected_qc.get("L1-OD"), "L2-OD": selected_qc.get("L2-OD"),
                                  "L1-起始OD": selected_qc.get("L1-起始OD"), "L2-起始OD": selected_qc.get("L2-起始OD")},
                "qcCheckResult": check_results,
                "disposeLots":   dispose_lots,
            })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500

def dict_factory(cursor, row):
    return {col[0]: row[idx] for idx, col in enumerate(cursor.description)}

@app.get("/api/search_571_tables")
def api_search_571_tables():
    import traceback
    work_order = request.args.get("work_order", "").strip()
    if not work_order:
        return jsonify({"ok": False, "message": "No work order provided"}), 400
    try:
        with connect() as conn:
            target_table = find_table_contains_workorder(conn, work_order)
            if target_table:
                conn.row_factory = dict_factory
                columns     = [info['name'] for info in conn.execute(f"PRAGMA table_info({target_table})").fetchall()]
                wo_col_name = next((n for n in ['工單號碼','WorkOrder','work_order','工單編號'] if n in columns), None)
                rows = conn.execute(f"SELECT * FROM {target_table} WHERE {wo_col_name} = ?", (work_order,)).fetchall() if wo_col_name else []
                return jsonify({"ok": True, "table": target_table, "rows": rows, "message": f"Found in {target_table}"})
            return jsonify({"ok": False, "rows": [], "table": None, "message": "查無資料"})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "message": str(e)}), 500

@app.post("/api/save_workorder")
def save_workorder():
    try:
        payload    = request.get_json(force=True)
        work_order = payload.get("workOrderNo")
        if not work_order:
            return jsonify({"error": "workOrderNo missing"}), 400
        beads   = payload.get("beads", [])
        reagent = payload.get("reagent", {})
        buffer  = payload.get("bufferBase", {})
        with connect() as conn:
            table = find_table_contains_workorder(conn, work_order)
            if not table:
                return jsonify({"error": f"No table contains work order {work_order}"}), 404
            if any("remark" in (b or {}) for b in beads):
                cols = [r[1] for r in conn.execute(f"PRAGMA table_info('{table}')").fetchall()]
                if "備註" not in cols:
                    conn.execute(f"ALTER TABLE '{table}' ADD COLUMN 備註 TEXT"); conn.commit()
            updated = 0
            for bead in beads:
                remark = bead.get("remark")
                sql    = f"UPDATE '{table}' SET L1OD=?, L2OD=?, 起始L1OD=?, 起始L2OD=?, 配製人員=?, 總重量=?, 重量紀錄=?"
                args   = [buffer.get("L1OD",""), buffer.get("L2OD",""),
                          buffer.get("L1StartOD",""), buffer.get("L2StartOD",""),
                          reagent.get("preparedBy",""), bead.get("totalQty",0), bead.get("qtyPerBead",0)]
                if remark is not None:
                    sql += ", 備註=?"; args.append(remark)
                sql += " WHERE 工單號碼=? AND 料號=?"
                args.extend([work_order, bead.get("beadPN","")])
                conn.execute(sql, args); conn.commit()
                updated += conn.total_changes
        return jsonify({"message": f"✅ 已更新 {updated} 筆資料", "table": table})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==================== Heatmap 統計 API ====================
from collections import defaultdict

@app.get("/api/heatmap_usage")
def api_heatmap_usage():
    try:
        mode  = request.args.get("mode", "week").strip().lower()
        top_n = int(request.args.get("top", "24"))
    except: top_n = 24

    def get_bucket(date_str, mode):
        try:
            for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y%m%d"):
                try: dt = datetime.datetime.strptime(date_str, fmt); break
                except ValueError: continue
            else: return None
            if mode == "month":    return dt.strftime("%Y-%m")
            elif mode == "quarter": q = (dt.month-1)//3+1; return f"{dt.year}-Q{q}"
            else: y, w, _ = dt.isocalendar(); return f"{y}-W{w:02d}"
        except: return None

    usage_map     = defaultdict(lambda: defaultdict(int))
    all_buckets   = set()
    marker_totals = defaultdict(int)
    with connect() as conn:
        try:
            rows = conn.execute(
                f'SELECT Marker, "Date" FROM "{DROPLET_TABLE}" WHERE Marker IS NOT NULL AND "Date" IS NOT NULL'
            ).fetchall()
        except sqlite3.OperationalError:
            return jsonify({"markers": [], "buckets": [], "matrix": [], "mode": mode})
        for r in rows:
            marker   = (r["Marker"] or "").strip().upper()
            date_str = (r["Date"] or "").strip()
            if not marker or not date_str: continue
            bucket = get_bucket(date_str, mode)
            if bucket:
                usage_map[marker][bucket] += 1
                marker_totals[marker]     += 1
                all_buckets.add(bucket)

    sorted_markers = sorted(marker_totals, key=lambda k: marker_totals[k], reverse=True)[:top_n]
    sorted_buckets = sorted(all_buckets)
    matrix = [[usage_map[m].get(b, 0) for b in sorted_buckets] for m in sorted_markers]
    return jsonify({"markers": sorted_markers, "buckets": sorted_buckets, "matrix": matrix, "mode": mode})

# ==================== MAIN ====================
if __name__ == "__main__":
    local_ip = socket.gethostbyname(socket.gethostname())
    print("🚀 Flask 啟動成功！")
    print(f"🌐 同網段裝置可訪問： http://{local_ip}:5011")
    print(f"🧠 本機： http://localhost:5011")
    print(f"🐘 RDS 雙寫：{'✅ 啟用' if PG_ENABLED else '⏭️ 停用'}  →  {EC2_RELAY_URL}")
    app.run(host="0.0.0.0", port=5011, debug=False, use_reloader=False)
