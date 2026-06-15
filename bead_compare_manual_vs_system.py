# -*- coding: utf-8 -*-
"""
Bead 人工 vs 系統需求/庫存比較工具

用途：
  input 2 個 Excel：
    1) 人工計算檔，例如 Beads 庫存統計26W23.xlsx
    2) 系統計算檔，例如 BeadNeed-yyyymmddhhmmss.xlsx
  套用固定規則後輸出比較表，格式接近 bead_manual_vs_system_stock_take_rule_AMY_alias_fixed.xlsx。

安裝：
  pip install openpyxl

執行方式：
  方式 1：直接雙擊或執行以下指令，開啟 GUI 手動選檔
    python bead_compare_manual_vs_system.py

  方式 2：命令列模式
    python bead_compare_manual_vs_system.py ^
      --manual "Beads 庫存統計26W23.xlsx" ^
      --system "BeadNeed-20260608084551.xlsx" ^
      --output "bead_compare_output.xlsx"

重要規則：
  - 舊檔/人工檔 = 人工計算
  - 新檔/BeadNeed = 系統計算
  - 兩劑品項：需求取大，庫存取小
    例如 ALP-D / ALP-U -> ALP；需求 MAX(D,U)，庫存 MIN(D,U)
  - Q 開頭獨立，不與非 Q 合併：QTG != TG
  - T4 DB 獨立，不與 T4 合併：T4 DB != T4
  - Bead's Need 的 AMY-A = AMY
  - CREA 不單獨拉出 tCREA-U 搭配量；Summary 只看 CREA
    CREA 需求 = MAX(tCRE-D, tCREA-D)；tCREA-U 只用於庫存/搭配判斷，不列為獨立需求列
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# -----------------------------
# 可依現場規則微調的設定
# -----------------------------
PAIR_SUFFIXES = {
    "-AD": "D",
    "-AU": "U",
    "-BD": "D",
    "-BU": "U",
    "-D": "D",
    "-U": "U",
}

# 配方版本 suffix，去除後歸為同一品項 (plain)
VERSION_SUFFIXES = ["-A", "-B", "-C"]

EXACT_ALIAS = {
    "AMY-A": "AMY",
    "AMY A": "AMY",
    "GLU-B": "GLU",
    "TDBIL": "TBIL",
    "TBIL-U": "TBIL",  # 人工檔工作表2有時出現；彙總表通常已是 TBIL
    "TBIL U": "TBIL",
    "CK": "CPK",
    "CK-AD": "CPK-AD",
    "CK-U": "CPK-U",
    "TAST": "AST",
    "TASTI": "AST",
    "TCREA": "CREA",
    "T-CREA": "CREA",
    "CREA": "CREA",
    "NT4-D": "T4-D",
    "NT4-U": "T4-U",
    "NT4": "T4",
    "TCO2-D": "TCO-2-D",
    "TCO2-U": "TCO-2-U",
    "TCO2": "TCO-2",
    "QTCO2-D": "QTCO-2-D",
    "QTCO2-U": "QTCO-2-U",
    "QTCO2": "QTCO-2",
    "GLIPA-AD": "LIPA-AD",
    "GLIPA-AU": "LIPA-AU",
    "GLIPA AD": "LIPA",
    "GLIPA": "LIPA",
    "RGT-D": "GGT-D",
    "RGT": "GGT",
}

# CREA 三劑組合特殊辨識
CREA_D_ALIASES = {
    "TCRE-D", "T-CRE-D", "CRE-D", "T CRE D",
    "TCREA-D", "T-CREA-D", "CREA-D", "T CREA D",
}
CREA_U_ALIASES = {
    "TCREA-U", "T-CREA-U", "CREA-U", "T CREA U",
}
CREA_PLAIN_ALIASES = {"CREA", "TCREA", "T-CREA"}

MANUAL_WEEK_LABELS = ["下週需求", "下周需求", "下下週需求", "下下周需求", "下下下週需求", "下下下周需求"]
SYSTEM_WEEK_LABELS = ["第一周需求", "第一週需求", "第二周需求", "第二週需求", "第三周需求", "第三週需求"]
STOCK_LABELS = ["庫存+滴定", "庫存+滴定(stock)", "庫存", "stock"]


@dataclass
class RawRow:
    source: str
    sheet: str
    row_no: int
    item_raw: str
    item_norm: str
    weeks: List[float]
    stock: float
    detail: str = ""


@dataclass
class AggItem:
    item: str
    weeks: List[float] = field(default_factory=lambda: [0.0, 0.0, 0.0])
    stock: float = 0.0
    details: List[str] = field(default_factory=list)


def to_number(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    if not s:
        return 0.0
    # 只取數字部分，避免 '<安全水位' 這類文字
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else 0.0


def norm_text(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    s = s.replace("–", "-").replace("—", "-").replace("－", "-")
    s = re.sub(r"\s+", " ", s)
    return s


def normalize_item_name(name: str) -> str:
    """保守正規化：不移除 Q，不把 T4 DB 併入 T4。"""
    s = norm_text(name).upper()
    s = s.replace("_", "-")
    s = re.sub(r"\s*-\s*", "-", s)
    s = re.sub(r"\s+", " ", s).strip()

    # T4 DB 要保留完整名稱，不能變 T4
    if s in {"T4 DB", "T4-DB", "T4DB"}:
        return "T4 DB"

    # CREA special names 先保留，後續由 aggregate_crea() 處理
    if s in CREA_D_ALIASES or s in CREA_U_ALIASES or s in CREA_PLAIN_ALIASES:
        return s

    if s in EXACT_ALIAS:
        return EXACT_ALIAS[s]

    return s


def find_header_index(headers: List[str], candidates: Iterable[str]) -> Optional[int]:
    h_norm = [norm_text(h).lower() for h in headers]
    for cand in candidates:
        c = norm_text(cand).lower()
        for i, h in enumerate(h_norm):
            if h == c or c in h:
                return i
    return None


def read_system_file(path: Path) -> List[RawRow]:
    wb = load_workbook(path, data_only=True, read_only=True)
    rows: List[RawRow] = []

    for ws in wb.worksheets:
        # 找 header row：同時包含品名、第一周、第二周、第三周
        header_row_no = None
        headers = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = [norm_text(v) for v in row]
            joined = "|".join(cells)
            if ("品名" in joined or "Bead Name" in joined or "Bead" in joined) and \
               any(x in joined for x in ["第一周需求", "第一週需求"]) and \
               any(x in joined for x in ["第二周需求", "第二週需求"]) and \
               any(x in joined for x in ["第三周需求", "第三週需求"]):
                header_row_no = r_idx
                headers = cells
                break

        if not header_row_no:
            continue

        item_idx = find_header_index(headers, ["品名(Bead Name)", "品名", "Bead Name", "Beads", "項目"])
        stock_idx = find_header_index(headers, STOCK_LABELS)
        w1_idx = find_header_index(headers, ["第一周需求", "第一週需求"])
        w2_idx = find_header_index(headers, ["第二周需求", "第二週需求"])
        w3_idx = find_header_index(headers, ["第三周需求", "第三週需求"])

        required = [item_idx, stock_idx, w1_idx, w2_idx, w3_idx]
        if any(x is None for x in required):
            continue

        for r_idx, row in enumerate(ws.iter_rows(min_row=header_row_no + 1, values_only=True), start=header_row_no + 1):
            vals = list(row)
            item_raw = norm_text(vals[item_idx]) if item_idx < len(vals) else ""
            if not item_raw:
                continue
            weeks = [to_number(vals[i]) if i < len(vals) else 0.0 for i in [w1_idx, w2_idx, w3_idx]]
            stock = to_number(vals[stock_idx]) if stock_idx < len(vals) else 0.0
            rows.append(RawRow(
                source="系統",
                sheet=ws.title,
                row_no=r_idx,
                item_raw=item_raw,
                item_norm=normalize_item_name(item_raw),
                weeks=weeks,
                stock=stock,
                detail=f"{ws.title}!R{r_idx}",
            ))
    return rows


def read_manual_file(path: Path) -> List[RawRow]:
    wb = load_workbook(path, data_only=True, read_only=True)
    rows: List[RawRow] = []

    for ws in wb.worksheets:
        # 人工檔常見橫向彙總表：A欄列名，品項在上一列，列名包含 庫存+滴定/下週需求/...
        data = list(ws.iter_rows(values_only=True))
        for r_idx, row in enumerate(data, start=1):
            first = norm_text(row[0]) if row else ""
            if first not in STOCK_LABELS:
                continue
            header_row_idx = r_idx - 1
            if header_row_idx < 1:
                continue
            header = data[header_row_idx - 1]

            # 尋找需求列
            label_to_row = {first: r_idx}
            for rr in range(r_idx + 1, min(r_idx + 8, len(data)) + 1):
                label = norm_text(data[rr - 1][0]) if data[rr - 1] else ""
                if label:
                    label_to_row[label] = rr

            w_rows = []
            for labels in [["下週需求", "下周需求"], ["下下週需求", "下下周需求"], ["下下下週需求", "下下下周需求"]]:
                found = None
                for label in labels:
                    if label in label_to_row:
                        found = label_to_row[label]
                        break
                if found is None:
                    w_rows = []
                    break
                w_rows.append(found)
            if not w_rows:
                continue

            # 從第 2 欄開始，每欄是一個品項
            max_cols = max(len(header), len(row))
            for c_idx in range(2, max_cols + 1):
                item_raw = norm_text(header[c_idx - 1]) if c_idx - 1 < len(header) else ""
                if not item_raw:
                    continue
                stock = to_number(row[c_idx - 1]) if c_idx - 1 < len(row) else 0.0
                weeks = []
                for wr in w_rows:
                    rr = data[wr - 1]
                    weeks.append(to_number(rr[c_idx - 1]) if c_idx - 1 < len(rr) else 0.0)
                rows.append(RawRow(
                    source="人工",
                    sheet=ws.title,
                    row_no=r_idx,
                    item_raw=item_raw,
                    item_norm=normalize_item_name(item_raw),
                    weeks=weeks,
                    stock=stock,
                    detail=f"{ws.title}!C{c_idx}: header R{header_row_idx}, data R{r_idx}-{w_rows[-1]}",
                ))
    return rows


def pair_base_and_side(item: str) -> Tuple[str, Optional[str]]:
    """回傳 (base, side)。side 為 D/U 表示兩劑；None 表示非兩劑。
    兩劑 suffix: -D, -U, -AD, -AU, -BD, -BU
    配方版本 suffix: -A, -B, -C (去除後歸為同一 base, side=None)
    """
    # CREA 交給特殊規則，不走一般 D/U
    if item in CREA_D_ALIASES or item in CREA_U_ALIASES:
        return item, None
    # 先檢查兩劑 suffix
    for suffix, side in sorted(PAIR_SUFFIXES.items(), key=lambda x: len(x[0]), reverse=True):
        if item.endswith(suffix):
            return item[: -len(suffix)], side
    # 再檢查配方版本 suffix (去除後歸為同一 base, 不配對)
    for suffix in VERSION_SUFFIXES:
        if item.endswith(suffix) and len(item) > len(suffix):
            return item[: -len(suffix)], None
    return item, None


def aggregate_rows(raw_rows: List[RawRow]) -> Dict[str, AggItem]:
    """套用規則後彙總。"""
    plain: Dict[str, List[RawRow]] = defaultdict(list)
    pair_groups: Dict[str, Dict[str, List[RawRow]]] = defaultdict(lambda: defaultdict(list))
    crea_d: List[RawRow] = []
    crea_u: List[RawRow] = []
    crea_plain: List[RawRow] = []

    for rr in raw_rows:
        item = rr.item_norm
        if item in CREA_D_ALIASES:
            crea_d.append(rr)
            continue
        if item in CREA_U_ALIASES:
            crea_u.append(rr)
            continue
        if item in CREA_PLAIN_ALIASES:
            crea_plain.append(rr)
            continue

        base, side = pair_base_and_side(item)
        if side:
            pair_groups[base][side].append(rr)
        else:
            plain[base].append(rr)

    out: Dict[str, AggItem] = {}

    def add_plain_item(item: str, rows: List[RawRow]):
        weeks = [sum(r.weeks[i] for r in rows) for i in range(3)]
        stock = sum(r.stock for r in rows)
        details = [f"{r.item_raw}({r.detail})" for r in rows]
        out[item] = AggItem(item=item, weeks=weeks, stock=stock, details=details)

    for item, rows in plain.items():
        add_plain_item(item, rows)

    # 兩劑規則：需求取大，庫存取小
    for base, sides in pair_groups.items():
        d_rows = sides.get("D", [])
        u_rows = sides.get("U", [])
        if d_rows and u_rows:
            d_weeks = [sum(r.weeks[i] for r in d_rows) for i in range(3)]
            u_weeks = [sum(r.weeks[i] for r in u_rows) for i in range(3)]
            weeks = [max(d_weeks[i], u_weeks[i]) for i in range(3)]
            d_stock = sum(r.stock for r in d_rows)
            u_stock = sum(r.stock for r in u_rows)
            stock = min(d_stock, u_stock)
            details = (
                [f"D:{r.item_raw}({r.detail})" for r in d_rows] +
                [f"U:{r.item_raw}({r.detail})" for r in u_rows] +
                [f"規則=兩劑需求取大/庫存取小；D需求={d_weeks}, U需求={u_weeks}, D庫存={d_stock:g}, U庫存={u_stock:g}"]
            )
        else:
            only_rows = d_rows or u_rows
            side = "D" if d_rows else "U"
            weeks = [sum(r.weeks[i] for r in only_rows) for i in range(3)]
            stock = sum(r.stock for r in only_rows)
            details = [f"{side}:{r.item_raw}({r.detail})" for r in only_rows] + ["只有單側資料，直接採用"]
        out[base] = AggItem(item=base, weeks=weeks, stock=stock, details=details)

    # CREA 特殊規則
    if crea_plain or crea_d or crea_u:
        if crea_d:
            # D 類多個替代，需求取大；庫存取大後再與 U/2 取小
            d_week_candidates = [[r.weeks[i] for r in crea_d] for i in range(3)]
            d_weeks = [max(cands) if cands else 0.0 for cands in d_week_candidates]
            d_stock = max([r.stock for r in crea_d] or [0.0])
            u_stock_total = sum(r.stock for r in crea_u)
            stock = min(d_stock, u_stock_total / 2.0) if crea_u else d_stock
            weeks = d_weeks
            details = (
                [f"D替代:{r.item_raw}({r.detail})" for r in crea_d] +
                [f"U搭配:{r.item_raw}({r.detail})" for r in crea_u] +
                ["規則=CREA需求取 MAX(tCRE-D,tCREA-D)；tCREA-U 不單獨列需求，只用於庫存/搭配判斷",
                 f"D庫存取大={d_stock:g}；U庫存/2={u_stock_total/2.0:g}；CREA有效庫存={stock:g}"]
            )
        else:
            weeks = [sum(r.weeks[i] for r in crea_plain) for i in range(3)]
            stock = sum(r.stock for r in crea_plain)
            details = [f"{r.item_raw}({r.detail})" for r in crea_plain] + ["人工彙總 CREA，直接採用"]
        out["CREA"] = AggItem(item="CREA", weeks=weeks, stock=stock, details=details)

    return out


def sorted_items(manual: Dict[str, AggItem], system: Dict[str, AggItem]) -> List[str]:
    def key_func(x: str):
        # 常用品項優先，其餘字母排序
        priority = ["TBIL", "TP", "ALP", "BUN", "UA", "GLU", "ALB", "TG", "ALT", "K", "HDL", "CREA", "AST", "CPK", "CA", "NA", "AMY", "GGT", "PHOS", "CL", "LIPA"]
        try:
            return (0, priority.index(x), x)
        except ValueError:
            return (1, x)
    return sorted(set(manual) | set(system), key=key_func)


def fmt_num(v: float):
    if abs(v - round(v)) < 1e-9:
        return int(round(v))
    return round(v, 2)


def build_comparison_rows(manual: Dict[str, AggItem], system: Dict[str, AggItem]) -> List[List]:
    headers = [
        "品項",
        "人工第一周需求", "人工第二周需求", "人工第三周需求", "人工三周需求",
        "系統第一周需求", "系統第二周需求", "系統第三周需求", "系統三周需求",
        "系統-人工需求差異",
        "人工有效庫存", "系統有效庫存", "系統-人工庫存差異",
        "狀態", "人工計算明細", "系統計算明細",
    ]
    rows = [headers]
    for item in sorted_items(manual, system):
        m = manual.get(item, AggItem(item=item))
        s = system.get(item, AggItem(item=item))
        m_total = sum(m.weeks)
        s_total = sum(s.weeks)
        demand_diff = s_total - m_total
        stock_diff = s.stock - m.stock
        if demand_diff > 0:
            status = "系統計算較高"
        elif demand_diff < 0:
            status = "系統計算較低"
        else:
            status = "需求一致"
        rows.append([
            item,
            *[fmt_num(x) for x in m.weeks], fmt_num(m_total),
            *[fmt_num(x) for x in s.weeks], fmt_num(s_total),
            fmt_num(demand_diff),
            fmt_num(m.stock), fmt_num(s.stock), fmt_num(stock_diff),
            status,
            "；".join(m.details),
            "；".join(s.details),
        ])
    return rows


def style_header(ws, row=1, fill="1F4E78"):
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color="999999"))


def auto_width(ws, max_width=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col[:200]:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def add_table_style(ws):
    thin = Side(style="thin", color="DDDDDD")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"


def write_output(output_path: Path, manual_rows: List[RawRow], system_rows: List[RawRow], manual_agg: Dict[str, AggItem], system_agg: Dict[str, AggItem]):
    wb = Workbook()
    # 刪除預設 sheet
    ws = wb.active
    ws.title = "Summary"

    comp_rows = build_comparison_rows(manual_agg, system_agg)
    headers = comp_rows[0]
    data_rows = comp_rows[1:]

    # Summary
    total_manual = sum(r[4] for r in data_rows)
    total_system = sum(r[8] for r in data_rows)
    total_diff = total_system - total_manual
    total_manual_stock = sum(r[10] for r in data_rows)
    total_system_stock = sum(r[11] for r in data_rows)
    total_stock_diff = total_system_stock - total_manual_stock

    ws["A1"] = "Bead 人工 vs 系統比較 Summary"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
    summary = [
        ["項目", "數值", "說明"],
        ["人工三周需求合計", fmt_num(total_manual), "人工計算檔套用規則後合計"],
        ["系統三周需求合計", fmt_num(total_system), "Bead's Need 套用規則後合計"],
        ["系統-人工需求差異", fmt_num(total_diff), "正數表示系統計算較高"],
        ["人工有效庫存合計", fmt_num(total_manual_stock), "人工檔有效庫存合計"],
        ["系統有效庫存合計", fmt_num(total_system_stock), "系統檔有效庫存合計"],
        ["系統-人工庫存差異", fmt_num(total_stock_diff), "正數表示系統庫存較高"],
        ["CREA 說明", "不單獨列 tCREA-U", "CREA 需求看 CREA 列；tCREA-U 只作搭配/庫存判斷"],
    ]
    for r_idx, row in enumerate(summary, start=3):
        for c_idx, v in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, v)
    style_header(ws, 3)
    add_table_style(ws)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 70

    # Top diff table
    top = sorted(data_rows, key=lambda r: abs(to_number(r[9])), reverse=True)[:15]
    start = 13
    ws.cell(start, 1, "Top 15 需求差異品項")
    ws.cell(start, 1).font = Font(bold=True, size=13, color="1F4E78")
    top_headers = ["品項", "人工三周需求", "系統三周需求", "系統-人工差異", "人工庫存", "系統庫存"]
    for c, v in enumerate(top_headers, 1):
        ws.cell(start + 1, c, v)
    for i, r in enumerate(top, start + 2):
        vals = [r[0], r[4], r[8], r[9], r[10], r[11]]
        for c, v in enumerate(vals, 1):
            ws.cell(i, c, v)
    style_header(ws, start + 1)
    for row in ws.iter_rows(min_row=start + 2, max_row=start + 1 + len(top), min_col=1, max_col=6):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    if top:
        chart = BarChart()
        chart.title = "Top 15 系統-人工需求差異"
        chart.y_axis.title = "差異數量"
        chart.x_axis.title = "品項"
        cats = Reference(ws, min_col=1, min_row=start + 2, max_row=start + 1 + len(top))
        data = Reference(ws, min_col=4, min_row=start + 1, max_row=start + 1 + len(top))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 20
        ws.add_chart(chart, "H13")

    # 彙總比對
    ws2 = wb.create_sheet("彙總比對_含庫存")
    for r_idx, row in enumerate(comp_rows, start=1):
        for c_idx, v in enumerate(row, start=1):
            ws2.cell(r_idx, c_idx, v)
    style_header(ws2, 1)
    add_table_style(ws2)
    ws2.freeze_panes = "B2"
    ws2.auto_filter.ref = ws2.dimensions
    auto_width(ws2)
    for col in range(2, 14):
        for cell in ws2.iter_cols(min_col=col, max_col=col, min_row=2, max_row=ws2.max_row):
            for c in cell:
                c.number_format = "#,##0"
    # 差異欄醒目
    for r in range(2, ws2.max_row + 1):
        diff = to_number(ws2.cell(r, 10).value)
        if diff > 0:
            ws2.cell(r, 10).fill = PatternFill("solid", fgColor="FCE4D6")
        elif diff < 0:
            ws2.cell(r, 10).fill = PatternFill("solid", fgColor="DDEBF7")

    # 系統計算明細
    ws3 = wb.create_sheet("系統計算明細")
    raw_headers = ["來源", "Sheet", "Row", "原始品名", "正規化品名", "第一周需求", "第二周需求", "第三周需求", "庫存+滴定", "位置"]
    ws3.append(raw_headers)
    for rr in system_rows:
        ws3.append([rr.source, rr.sheet, rr.row_no, rr.item_raw, rr.item_norm, *[fmt_num(x) for x in rr.weeks], fmt_num(rr.stock), rr.detail])
    style_header(ws3, 1)
    add_table_style(ws3)
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions
    auto_width(ws3)

    # 規則說明
    ws4 = wb.create_sheet("規則說明")
    rules = [
        ["規則", "說明"],
        ["人工/系統", "人工計算檔稱為人工；Bead's Need 檔稱為系統。差異欄位皆為 系統 - 人工。"],
        ["三周需求", "人工：下週/下下週/下下下週需求；系統：第一周/第二周/第三周需求。"],
        ["兩劑需求", "同一品項的 -D / -U / -AU 每周需求取 MAX，不相加。"],
        ["兩劑庫存", "同一品項的 -D / -U / -AU 庫存取 MIN，避免其中一劑不足時高估可用量。"],
        ["Q 品項", "保留 Q 前綴，例如 QTG 不等於 TG。"],
        ["T4 DB", "T4 DB 不等於 T4，兩者分開列。"],
        ["AMY", "Bead's Need 內 AMY-A 視為 AMY。"],
        ["CREA", "CREA 為組合測試；需求取 MAX(tCRE-D,tCREA-D)。tCREA-U 不單獨列需求，只用於搭配與庫存判斷。"],
        ["CREA庫存", "若有 tCREA-U，CREA有效庫存 = MIN(D側庫存取大, tCREA-U庫存/2)。"],
    ]
    for r_idx, row in enumerate(rules, start=1):
        for c_idx, v in enumerate(row, start=1):
            ws4.cell(r_idx, c_idx, v)
    style_header(ws4, 1)
    add_table_style(ws4)
    ws4.column_dimensions["A"].width = 20
    ws4.column_dimensions["B"].width = 90

    # T4_DB檢查
    ws5 = wb.create_sheet("T4_DB檢查")
    check_headers = headers
    ws5.append(check_headers)
    for row in data_rows:
        if row[0] in {"T4", "T4 DB"}:
            ws5.append(row)
    style_header(ws5, 1)
    add_table_style(ws5)
    auto_width(ws5)

    # AMY檢查
    ws6 = wb.create_sheet("AMY檢查")
    ws6.append(check_headers)
    for row in data_rows:
        if row[0] == "AMY":
            ws6.append(row)
    ws6.append([])
    ws6.append(["原始 AMY/AMY-A 明細"])
    ws6.append(raw_headers)
    for rr in manual_rows + system_rows:
        if normalize_item_name(rr.item_raw) == "AMY" or "AMY" in normalize_item_name(rr.item_raw):
            ws6.append([rr.source, rr.sheet, rr.row_no, rr.item_raw, rr.item_norm, *[fmt_num(x) for x in rr.weeks], fmt_num(rr.stock), rr.detail])
    style_header(ws6, 1)
    style_header(ws6, 5 if ws6.max_row >= 5 else 1, fill="70AD47")
    add_table_style(ws6)
    auto_width(ws6)

    # 全工作表美化
    for wsx in wb.worksheets:
        for row in wsx.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        wsx.sheet_view.showGridLines = False

    wb.save(output_path)


def run_compare(manual_path: Path, system_path: Path, output_path: Path) -> Tuple[int, str]:
    """執行比較，供 CLI 與 GUI 共用。回傳 (exit_code, message)。"""
    manual_path = Path(manual_path)
    system_path = Path(system_path)
    output_path = Path(output_path)

    if not manual_path.exists():
        return 2, f"找不到人工檔：{manual_path}"
    if not system_path.exists():
        return 2, f"找不到系統檔：{system_path}"
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")

    manual_rows = read_manual_file(manual_path)
    system_rows = read_system_file(system_path)

    if not manual_rows:
        return 3, "人工檔沒有讀到資料：請確認有 橫向彙總表，且列名包含 庫存+滴定/下週需求/下下週需求/下下下週需求。"
    if not system_rows:
        return 3, "系統檔沒有讀到資料：請確認有品名、庫存+滴定、第一/二/三周需求欄位。"

    manual_agg = aggregate_rows(manual_rows)
    system_agg = aggregate_rows(system_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_output(output_path, manual_rows, system_rows, manual_agg, system_agg)
    msg = (
        f"完成：{output_path}\n"
        f"人工原始列：{len(manual_rows)}\n"
        f"系統原始列：{len(system_rows)}\n"
        f"彙總品項：{len(set(manual_agg) | set(system_agg))}"
    )
    return 0, msg


def default_output_path(system_file: str) -> str:
    """依系統檔位置自動建議輸出檔名。"""
    if not system_file:
        return ""
    p = Path(system_file)
    return str(p.with_name(f"bead_manual_vs_system_compare_{p.stem}.xlsx"))


def launch_gui() -> int:
    """Tkinter 手動選檔 UI。Windows 內建 tkinter，不需額外安裝。"""
    try:
        import threading
        import tkinter as tk
        from tkinter import filedialog, messagebox, ttk
    except Exception as exc:
        print(f"無法啟動 GUI：{exc}", file=sys.stderr)
        return 10

    root = tk.Tk()
    root.title("Bead 人工 vs 系統比較工具")
    root.geometry("760x420")
    root.minsize(720, 390)

    manual_var = tk.StringVar()
    system_var = tk.StringVar()
    output_var = tk.StringVar()
    status_var = tk.StringVar(value="請選擇人工檔與系統檔。")

    def pick_excel(target_var: tk.StringVar, is_system: bool = False):
        filename = filedialog.askopenfilename(
            title="選擇 Excel 檔",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")],
        )
        if filename:
            target_var.set(filename)
            if is_system and not output_var.get().strip():
                output_var.set(default_output_path(filename))

    def pick_output():
        initial = output_var.get().strip() or default_output_path(system_var.get().strip()) or "bead_compare_output.xlsx"
        filename = filedialog.asksaveasfilename(
            title="選擇輸出 Excel 檔",
            defaultextension=".xlsx",
            initialfile=Path(initial).name,
            initialdir=str(Path(initial).parent) if Path(initial).parent.exists() else None,
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if filename:
            output_var.set(filename)

    def set_busy(is_busy: bool):
        run_btn.config(state="disabled" if is_busy else "normal")
        manual_btn.config(state="disabled" if is_busy else "normal")
        system_btn.config(state="disabled" if is_busy else "normal")
        output_btn.config(state="disabled" if is_busy else "normal")
        progress.config(mode="indeterminate" if is_busy else "determinate")
        if is_busy:
            progress.start(10)
        else:
            progress.stop()
            progress["value"] = 0

    def append_log(text: str):
        log_box.config(state="normal")
        log_box.insert("end", text + "\n")
        log_box.see("end")
        log_box.config(state="disabled")

    def run_clicked():
        manual = manual_var.get().strip()
        system = system_var.get().strip()
        output = output_var.get().strip() or default_output_path(system)
        if not manual:
            messagebox.showwarning("缺少人工檔", "請先選擇人工計算 Excel。")
            return
        if not system:
            messagebox.showwarning("缺少系統檔", "請先選擇系統計算 BeadNeed Excel。")
            return
        if not output:
            messagebox.showwarning("缺少輸出檔", "請選擇輸出 Excel 路徑。")
            return
        output_var.set(output)
        set_busy(True)
        status_var.set("執行中，請稍候...")
        append_log("開始比較...")
        append_log(f"人工檔：{manual}")
        append_log(f"系統檔：{system}")
        append_log(f"輸出檔：{output}")

        def worker():
            try:
                code, msg = run_compare(Path(manual), Path(system), Path(output))
            except Exception as exc:
                code, msg = 99, f"執行失敗：{exc}"
            def done():
                set_busy(False)
                append_log(msg)
                if code == 0:
                    status_var.set("完成。")
                    messagebox.showinfo("完成", msg)
                else:
                    status_var.set("失敗，請查看訊息。")
                    messagebox.showerror("失敗", msg)
            root.after(0, done)
        threading.Thread(target=worker, daemon=True).start()

    outer = ttk.Frame(root, padding=18)
    outer.pack(fill="both", expand=True)

    title = ttk.Label(outer, text="Bead 人工 vs 系統需求/庫存比較", font=("Microsoft JhengHei UI", 16, "bold"))
    title.grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 14))

    ttk.Label(outer, text="人工計算檔 manual Excel：").grid(row=1, column=0, sticky="w", pady=6)
    ttk.Entry(outer, textvariable=manual_var).grid(row=1, column=1, sticky="ew", pady=6, padx=8)
    manual_btn = ttk.Button(outer, text="選擇檔案", command=lambda: pick_excel(manual_var, False))
    manual_btn.grid(row=1, column=2, sticky="ew", pady=6)

    ttk.Label(outer, text="系統計算檔 system Excel：").grid(row=2, column=0, sticky="w", pady=6)
    ttk.Entry(outer, textvariable=system_var).grid(row=2, column=1, sticky="ew", pady=6, padx=8)
    system_btn = ttk.Button(outer, text="選擇檔案", command=lambda: pick_excel(system_var, True))
    system_btn.grid(row=2, column=2, sticky="ew", pady=6)

    ttk.Label(outer, text="輸出比較檔 output Excel：").grid(row=3, column=0, sticky="w", pady=6)
    ttk.Entry(outer, textvariable=output_var).grid(row=3, column=1, sticky="ew", pady=6, padx=8)
    output_btn = ttk.Button(outer, text="另存新檔", command=pick_output)
    output_btn.grid(row=3, column=2, sticky="ew", pady=6)

    run_btn = ttk.Button(outer, text="開始比較並輸出 Excel", command=run_clicked)
    run_btn.grid(row=4, column=1, sticky="ew", pady=(14, 8), padx=8)

    progress = ttk.Progressbar(outer)
    progress.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(2, 8))

    ttk.Label(outer, textvariable=status_var).grid(row=6, column=0, columnspan=3, sticky="w", pady=(0, 8))

    log_box = tk.Text(outer, height=8, wrap="word", state="disabled")
    log_box.grid(row=7, column=0, columnspan=3, sticky="nsew")

    outer.columnconfigure(1, weight=1)
    outer.rowconfigure(7, weight=1)

    notes = (
        "規則：兩劑需求取大、庫存取小；QTG≠TG；T4 DB≠T4；AMY-A=AMY；CREA 不單獨列 tCREA-U。"
    )
    append_log(notes)
    root.mainloop()
    return 0


def main(argv=None) -> int:
    if argv is None:
        argv = sys.argv[1:]
    # 沒帶參數時直接開 GUI，方便雙擊使用。
    if not argv:
        return launch_gui()

    parser = argparse.ArgumentParser(description="比較人工計算 Excel 與 Bead's Need 系統 Excel，輸出需求/庫存差異表。")
    parser.add_argument("--gui", action="store_true", help="啟動手動選檔 GUI")
    parser.add_argument("--manual", help="人工計算 Excel，例如 Beads 庫存統計26W23.xlsx")
    parser.add_argument("--system", help="系統計算 Excel，例如 BeadNeed-20260608084551.xlsx")
    parser.add_argument("--output", help="輸出 Excel 路徑")
    args = parser.parse_args(argv)

    if args.gui:
        return launch_gui()
    missing = [name for name, value in [("--manual", args.manual), ("--system", args.system), ("--output", args.output)] if not value]
    if missing:
        parser.error("命令列模式需要參數：" + ", ".join(missing) + "；或直接執行不帶參數開啟 GUI。")

    code, msg = run_compare(Path(args.manual), Path(args.system), Path(args.output))
    if code == 0:
        print(msg)
    else:
        print(msg, file=sys.stderr)
    return code


if __name__ == "__main__":
    raise SystemExit(main())
