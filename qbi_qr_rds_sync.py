# -*- coding: utf-8 -*-
"""Sync Qbi QR Excel lookup tables into RDS beadsdb.

Source workbook:
    /home/ubuntu/qc-web-ipqc/diskQR/Qbi_QR格式_V3_22.xlsx

Destination:
    qbi_qr.disc_types
    qbi_qr.markers
    qbi_qr.panels
"""

from __future__ import annotations

import logging
import os
import threading
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import text
from sqlalchemy.dialects.postgresql import insert as pg_insert

try:
    from watchdog.events import FileSystemEventHandler
    from watchdog.observers import Observer
except Exception:  # pragma: no cover - allows manual sync without watchdog installed
    FileSystemEventHandler = None
    Observer = None


DEFAULT_QBI_QR_EXCEL_PATH = "/home/ubuntu/qc-web-ipqc/diskQR/Qbi_QR格式_V3_22.xlsx"
QBI_QR_SCHEMA = "qbi_qr"


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r", "\n").strip()


def _split_lines(value: Any) -> list[str]:
    return [part.strip() for part in _clean(value).split("\n") if part.strip()]


def _split_markers(value: Any) -> list[str]:
    return [
        part.strip()
        for part in _clean(value).replace("\n", ", ").split(",")
        if part.strip()
    ]


def _qr_disc_type(disc_type_no: Any) -> str:
    return _clean(disc_type_no).ljust(2, "0")


def _upsert_many(conn, table, rows: list[dict[str, Any]], index_cols: list[str]) -> None:
    if not rows:
        return
    stmt = pg_insert(table).values(rows)
    update_cols = {
        col.name: getattr(stmt.excluded, col.name)
        for col in table.c
        if col.name not in index_cols and col.name != "updated_at"
    }
    update_cols["updated_at"] = text("now()")
    conn.execute(stmt.on_conflict_do_update(index_elements=index_cols, set_=update_cols))


def build_qbi_qr_records(excel_path: str | Path) -> dict[str, list[dict[str, Any]]]:
    workbook_path = Path(excel_path)
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    source_file = workbook_path.name

    disc_types: dict[str, dict[str, Any]] = {}
    ws_disc = wb["Disc Type"]
    for row_idx in range(5, ws_disc.max_row + 1):
        category_name = _clean(ws_disc.cell(row_idx, 1).value)
        category_name_zh = _clean(ws_disc.cell(row_idx, 2).value)
        disc_type_no = _clean(ws_disc.cell(row_idx, 3).value)
        if not category_name_zh or not disc_type_no.isdigit():
            continue
        if not category_name and disc_type_no == "0":
            category_name = "CHEM"
        if not category_name:
            category_name = category_name_zh
        qr_disc_type = _qr_disc_type(disc_type_no)
        disc_types[qr_disc_type] = {
            "qr_disc_type": qr_disc_type,
            "disc_type_no": disc_type_no,
            "category_name": category_name,
            "category_name_zh": category_name_zh,
            "source_file": source_file,
            "source_sheet": "Disc Type",
            "source_row": row_idx,
        }

    markers: dict[tuple[str, str], dict[str, Any]] = {}
    ws_marker = wb["Marker No(總表)"]

    def add_marker_block(
        *,
        source_group: str,
        target_disc_types: list[str] | None,
        old_name_col: int,
        marker_col: int,
        species_col: int,
        category_col: int,
        num_col: int,
        note_col: int,
        category_to_disc_types=None,
    ) -> None:
        for row_idx in range(4, ws_marker.max_row + 1):
            marker_name = _clean(ws_marker.cell(row_idx, marker_col).value)
            marker_number = _clean(ws_marker.cell(row_idx, num_col).value)
            if (
                not marker_name
                or not marker_number.isdigit()
                or marker_name in {"尿液", "Others"}
            ):
                continue
            category = _clean(ws_marker.cell(row_idx, category_col).value)
            disc_type_keys = (
                category_to_disc_types(category)
                if category_to_disc_types
                else target_disc_types
            )
            for disc_type_key in disc_type_keys or []:
                marker_number_key = marker_number.zfill(3)
                markers[(disc_type_key, marker_number_key)] = {
                    "qr_disc_type": disc_type_key,
                    "marker_number": marker_number_key,
                    "marker_name": marker_name,
                    "old_name": _clean(ws_marker.cell(row_idx, old_name_col).value),
                    "species_ref_range": _clean(
                        ws_marker.cell(row_idx, species_col).value
                    ),
                    "category": category,
                    "note": _clean(ws_marker.cell(row_idx, note_col).value),
                    "source_group": source_group,
                    "source_file": source_file,
                    "source_sheet": "Marker No(總表)",
                    "source_row": row_idx,
                }

    add_marker_block(
        source_group="Vet (生化)",
        target_disc_types=["00"],
        old_name_col=1,
        marker_col=2,
        species_col=3,
        category_col=4,
        num_col=5,
        note_col=6,
    )
    add_marker_block(
        source_group="Vet (免疫)",
        target_disc_types=None,
        old_name_col=8,
        marker_col=9,
        species_col=10,
        category_col=11,
        num_col=12,
        note_col=13,
        category_to_disc_types=lambda category: ["20"]
        if "ELISA" in category
        else ["10"],
    )
    add_marker_block(
        source_group="Vet (凝血)",
        target_disc_types=["30"],
        old_name_col=15,
        marker_col=16,
        species_col=17,
        category_col=18,
        num_col=19,
        note_col=20,
    )
    add_marker_block(
        source_group="Vet (快篩)",
        target_disc_types=["40"],
        old_name_col=22,
        marker_col=23,
        species_col=24,
        category_col=25,
        num_col=26,
        note_col=27,
    )
    add_marker_block(
        source_group="Vet (校正)",
        target_disc_types=["90"],
        old_name_col=28,
        marker_col=29,
        species_col=30,
        category_col=31,
        num_col=32,
        note_col=33,
    )
    for disc_type_key in ["00", "10", "20", "30", "40", "90"]:
        markers[(disc_type_key, "000")] = {
            "qr_disc_type": disc_type_key,
            "marker_number": "000",
            "marker_name": "Blank",
            "old_name": "-",
            "species_ref_range": "x",
            "category": "x",
            "note": "",
            "source_group": "System",
            "source_file": source_file,
            "source_sheet": "Marker No(總表)",
            "source_row": None,
        }

    panels: dict[str, dict[str, Any]] = {}
    ws_panel = wb["Panel Type"]
    carry = {
        "panel_name": "",
        "panel_name_cn": "",
        "marker_list_raw": "",
        "product_code": "",
        "one_piece_box_panel_type": "",
        "label_version": "",
    }
    for row_idx in range(2, ws_panel.max_row + 1):
        # This sheet has a leading blank openpyxl column.
        # Actual Excel letters: B=3, C=4, D=5, L=13, O=16, P=17, Q=18, R=19.
        row_panel_name = _clean(ws_panel.cell(row_idx, 3).value)
        row_panel_name_cn = _clean(ws_panel.cell(row_idx, 4).value)
        row_marker_list = _clean(ws_panel.cell(row_idx, 5).value)
        row_product_code = _clean(ws_panel.cell(row_idx, 13).value)
        row_one_piece = _clean(ws_panel.cell(row_idx, 17).value)
        row_label_version = _clean(ws_panel.cell(row_idx, 19).value)
        if row_panel_name:
            carry["panel_name"] = row_panel_name
        if row_panel_name_cn:
            carry["panel_name_cn"] = row_panel_name_cn
        if row_marker_list:
            carry["marker_list_raw"] = row_marker_list
        if row_product_code:
            carry["product_code"] = row_product_code
        if row_one_piece:
            carry["one_piece_box_panel_type"] = row_one_piece
        if row_label_version:
            carry["label_version"] = row_label_version

        is_chem_row = 2 <= row_idx <= 26
        raw_disc_type_no = "0" if is_chem_row else _clean(ws_panel.cell(row_idx, 16).value)
        raw_sub_panel_types = _clean(ws_panel.cell(row_idx, 18).value)
        panel_name = row_panel_name or carry["panel_name"]
        if not panel_name or not raw_disc_type_no.isdigit() or not raw_sub_panel_types:
            continue

        qr_disc_type = _qr_disc_type(raw_disc_type_no)
        if is_chem_row:
            disc_category = "CHEM"
            disc_category_zh = "生化"
        else:
            disc_info = disc_types.get(qr_disc_type)
            disc_category = disc_info["category_name"] if disc_info else "Unknown"
            disc_category_zh = (
                disc_info["category_name_zh"] if disc_info else "Unknown"
            )

        for raw_sub_panel_type in _split_lines(raw_sub_panel_types):
            if not raw_sub_panel_type.isdigit():
                continue
            sub_panel_type = raw_sub_panel_type.zfill(3)
            panel_key = f"{qr_disc_type}-{sub_panel_type}"
            if panel_key in panels:
                continue
            panels[panel_key] = {
                "panel_key": panel_key,
                "qr_disc_type": qr_disc_type,
                "sub_panel_type": sub_panel_type,
                "panel_name": panel_name,
                "panel_name_cn": row_panel_name_cn or carry["panel_name_cn"],
                "product_code": row_product_code or carry["product_code"] or None,
                "one_piece_box_panel_type": (
                    row_one_piece or carry["one_piece_box_panel_type"]
                )
                .replace(" ", "")
                .replace("\n", ""),
                "disc_category": disc_category,
                "disc_category_zh": disc_category_zh,
                "marker_list": _split_markers(row_marker_list or carry["marker_list_raw"]),
                "label_version": row_label_version or carry["label_version"] or None,
                "source_file": source_file,
                "source_sheet": "Panel Type",
                "source_row": row_idx,
            }

    return {
        "disc_types": list(disc_types.values()),
        "markers": list(markers.values()),
        "panels": list(panels.values()),
    }


def ensure_qbi_qr_tables(db) -> dict[str, Any]:
    metadata = db.MetaData(schema=QBI_QR_SCHEMA)
    disc_types = db.Table(
        "disc_types",
        metadata,
        db.Column("qr_disc_type", db.Text, primary_key=True),
        db.Column("disc_type_no", db.Text, nullable=False),
        db.Column("category_name", db.Text, nullable=False),
        db.Column("category_name_zh", db.Text, nullable=False),
        db.Column("source_file", db.Text, nullable=False),
        db.Column("source_sheet", db.Text, nullable=False),
        db.Column("source_row", db.Integer),
        db.Column("updated_at", db.DateTime(timezone=True), server_default=text("now()")),
    )
    markers = db.Table(
        "markers",
        metadata,
        db.Column("qr_disc_type", db.Text, primary_key=True),
        db.Column("marker_number", db.Text, primary_key=True),
        db.Column("marker_name", db.Text, nullable=False),
        db.Column("old_name", db.Text),
        db.Column("species_ref_range", db.Text),
        db.Column("category", db.Text),
        db.Column("note", db.Text),
        db.Column("source_group", db.Text, nullable=False),
        db.Column("source_file", db.Text, nullable=False),
        db.Column("source_sheet", db.Text, nullable=False),
        db.Column("source_row", db.Integer),
        db.Column("updated_at", db.DateTime(timezone=True), server_default=text("now()")),
    )
    panels = db.Table(
        "panels",
        metadata,
        db.Column("panel_key", db.Text, primary_key=True),
        db.Column("qr_disc_type", db.Text, nullable=False),
        db.Column("sub_panel_type", db.Text, nullable=False),
        db.Column("panel_name", db.Text, nullable=False),
        db.Column("panel_name_cn", db.Text),
        db.Column("product_code", db.Text),
        db.Column("one_piece_box_panel_type", db.Text),
        db.Column("disc_category", db.Text, nullable=False),
        db.Column("disc_category_zh", db.Text, nullable=False),
        db.Column("marker_list", db.JSON, nullable=False, server_default=text("'[]'::jsonb")),
        db.Column("label_version", db.Text),
        db.Column("source_file", db.Text, nullable=False),
        db.Column("source_sheet", db.Text, nullable=False),
        db.Column("source_row", db.Integer),
        db.Column("updated_at", db.DateTime(timezone=True), server_default=text("now()")),
    )

    with db.engine.begin() as conn:
        conn.execute(text(f"CREATE SCHEMA IF NOT EXISTS {QBI_QR_SCHEMA}"))
        metadata.create_all(conn)

    return {
        "disc_types": disc_types,
        "markers": markers,
        "panels": panels,
    }


def sync_qbi_qr_excel_to_rds(db, excel_path: str | Path | None = None) -> dict[str, Any]:
    workbook_path = Path(excel_path or os.getenv("QBI_QR_EXCEL_PATH", DEFAULT_QBI_QR_EXCEL_PATH))
    if not workbook_path.exists():
        raise FileNotFoundError(f"Qbi QR Excel not found: {workbook_path}")

    tables = ensure_qbi_qr_tables(db)
    records = build_qbi_qr_records(workbook_path)
    with db.engine.begin() as conn:
        _upsert_many(conn, tables["disc_types"], records["disc_types"], ["qr_disc_type"])
        _upsert_many(conn, tables["markers"], records["markers"], ["qr_disc_type", "marker_number"])
        _upsert_many(conn, tables["panels"], records["panels"], ["panel_key"])

    result = {
        "ok": True,
        "excel_path": str(workbook_path),
        "disc_types": len(records["disc_types"]),
        "markers": len(records["markers"]),
        "panels": len(records["panels"]),
    }
    logging.info("[QbiQR] synced %s", result)
    return result


class QbiQrExcelChangeHandler(FileSystemEventHandler if FileSystemEventHandler else object):
    def __init__(self, app, db, excel_path: Path, debounce_seconds: float = 2.0):
        self.app = app
        self.db = db
        self.excel_path = excel_path.resolve()
        self.debounce_seconds = debounce_seconds
        self._timer: threading.Timer | None = None
        self._lock = threading.Lock()

    def on_any_event(self, event):  # noqa: D401 - watchdog callback
        if getattr(event, "is_directory", False):
            return
        candidates = [Path(getattr(event, "src_path", ""))]
        dest_path = getattr(event, "dest_path", None)
        if dest_path:
            candidates.append(Path(dest_path))
        if not any(path.resolve() == self.excel_path for path in candidates if str(path)):
            return
        self.schedule_sync()

    def schedule_sync(self) -> None:
        with self._lock:
            if self._timer:
                self._timer.cancel()
            self._timer = threading.Timer(self.debounce_seconds, self._sync)
            self._timer.daemon = True
            self._timer.start()

    def _sync(self) -> None:
        try:
            with self.app.app_context():
                sync_qbi_qr_excel_to_rds(self.db, self.excel_path)
        except Exception:
            logging.exception("[QbiQR] sync failed after Excel change")


_observer = None


def start_qbi_qr_excel_watcher(app, db, excel_path: str | Path | None = None):
    global _observer
    if _observer is not None:
        return _observer
    if Observer is None:
        logging.warning("[QbiQR] watchdog is not installed; Excel watcher disabled")
        return None

    workbook_path = Path(excel_path or os.getenv("QBI_QR_EXCEL_PATH", DEFAULT_QBI_QR_EXCEL_PATH))
    watch_dir = workbook_path.parent
    watch_dir.mkdir(parents=True, exist_ok=True)
    handler = QbiQrExcelChangeHandler(app, db, workbook_path)
    observer = Observer()
    observer.schedule(handler, str(watch_dir), recursive=False)
    observer.daemon = True
    observer.start()
    _observer = observer
    logging.info("[QbiQR] watching %s", workbook_path)
    return observer
