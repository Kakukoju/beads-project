# -*- coding: utf-8 -*-
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text
from werkzeug.utils import secure_filename
import subprocess
import os
import io
import glob
import re
import json
import logging
import base64
import traceback
import importlib
import math
import requests as http_requests
from datetime import datetime
import openpyxl
import pandas as pd
from qbi_qr_rds_sync import (
    DEFAULT_QBI_QR_EXCEL_PATH,
    start_qbi_qr_excel_watcher,
    sync_qbi_qr_excel_to_rds,
)

# 引入排程 API 模組
import scheduler_api
importlib.reload(scheduler_api)

# 引入 AI 排程分析模組
from ai_schedule import ai_schedule_bp

app = Flask(__name__)
CORS(app)

# ---------- 1. 配置 AWS RDS PostgreSQL 連線 ----------
DB_USER = "harryguo"
DB_PASS = "skyla168"
DB_HOST = "database-1.cfutwrwyrxts.ap-northeast-1.rds.amazonaws.com"
DB_PORT = "5432"
DB_NAME = "beadsdb"

app.config['SQLALCHEMY_DATABASE_URI'] = f"postgresql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    "pool_pre_ping": True,
    "pool_recycle": 300,
    "connect_args": {"options": "-csearch_path=P01_formualte_schedule,panel_production,schedule,public"}
}

db = SQLAlchemy(app)

# ---------- Initialize panel_production schema and tutti_work_orders table ----------
with app.app_context():
    try:
        db.session.execute(text("""
            CREATE SCHEMA IF NOT EXISTS panel_production;
        """))
        db.session.execute(text("""
            DROP TABLE IF EXISTS panel_production.tutti_production;
        """))
        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS panel_production.tutti_work_orders (
                id              SERIAL PRIMARY KEY,
                work_order_no   VARCHAR(50) NOT NULL,
                lot_no          VARCHAR(50),
                form_data       JSONB NOT NULL,
                created_at      TIMESTAMP DEFAULT NOW(),
                updated_at      TIMESTAMP DEFAULT NOW(),
                created_by      VARCHAR(50)
            );
        """))
        db.session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_tutti_wo_work_order
                ON panel_production.tutti_work_orders (work_order_no);
        """))
        db.session.execute(text("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_tutti_wo_composite
                ON panel_production.tutti_work_orders (work_order_no, lot_no);
        """))
        db.session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_tutti_wo_lot_no
                ON panel_production.tutti_work_orders (lot_no);
        """))
        db.session.commit()
        logging.info("panel_production schema and tutti_work_orders table initialized successfully.")
    except Exception as e:
        db.session.rollback()
        logging.error(f"Failed to initialize panel_production schema/table: {e}")

# ---------- Register AI Schedule Blueprint ----------
app.register_blueprint(ai_schedule_bp)

# ---------- Initialize P01_formualte_schedule schema for AI schedule tables ----------
with app.app_context():
    try:
        db.session.execute(text("""
            CREATE SCHEMA IF NOT EXISTS "P01_formualte_schedule";
        """))
        db.session.commit()
        # Import models so SQLAlchemy metadata knows about them
        from ai_schedule import models  # noqa: F401
        # Create all tables defined in ai_schedule/models.py (schema-qualified)
        db.create_all()
        logging.info("P01_formualte_schedule schema and AI schedule tables initialized successfully.")
    except Exception as e:
        db.session.rollback()
        logging.error(f"Failed to initialize AI schedule tables: {e}")

# ---------- 2. 路徑配置 ----------
BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
CALC_DIR       = os.path.join(BASE_DIR, "calculation")
OUTPUT_EXCEL   = os.path.join(BASE_DIR, "outputs", "beads_needs.xlsx")
EXPORTS_DIR    = os.path.join(BASE_DIR, "exports")
EXCEL_DATA_DIR = os.path.join(BASE_DIR, "excelData")

for d in [CALC_DIR, os.path.join(BASE_DIR, "outputs"), EXPORTS_DIR, EXCEL_DATA_DIR]:
    os.makedirs(d, exist_ok=True)

# ---------- 3. API Key（上傳保護）----------
UPLOAD_API_KEY = os.environ.get('UPLOAD_API_KEY', 'beadsops-upload-key')

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")


# ---------- Qbi QR Excel → RDS watchdog sync ----------
QBI_QR_EXCEL_PATH = os.environ.get("QBI_QR_EXCEL_PATH", DEFAULT_QBI_QR_EXCEL_PATH)


def start_qbi_qr_watcher_once():
    if os.environ.get("QBI_QR_WATCH_ENABLED", "1") != "1":
        logging.info("[QbiQR] Excel watcher disabled by QBI_QR_WATCH_ENABLED")
        return None
    try:
        return start_qbi_qr_excel_watcher(app, db, QBI_QR_EXCEL_PATH)
    except Exception as e:
        logging.error(f"[QbiQR] Failed to start Excel watcher: {e}")
        return None


start_qbi_qr_watcher_once()


@app.route("/api/qbi-qr/sync", methods=["POST"])
def sync_qbi_qr_lookup_tables():
    try:
        result = sync_qbi_qr_excel_to_rds(db, QBI_QR_EXCEL_PATH)
        return jsonify(result)
    except Exception as e:
        logging.error(f"[QbiQR] Manual sync failed: {traceback.format_exc()}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/qbi-qr/status", methods=["GET"])
def qbi_qr_lookup_status():
    try:
        rows = db.session.execute(text("""
            SELECT 'disc_types' AS table_name, COUNT(*)::int AS count FROM qbi_qr.disc_types
            UNION ALL
            SELECT 'markers' AS table_name, COUNT(*)::int AS count FROM qbi_qr.markers
            UNION ALL
            SELECT 'panels' AS table_name, COUNT(*)::int AS count FROM qbi_qr.panels
            ORDER BY table_name
        """)).fetchall()
        return jsonify({
            "ok": True,
            "excel_path": QBI_QR_EXCEL_PATH,
            "counts": {row[0]: row[1] for row in rows},
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ---------- 核心工具：從 RDS 獲取名稱字典 ----------
def get_mrp_name_map():
    name_map = {}
    try:
        query = text('SELECT "Finished_PartNo", "panelName" FROM "PNandName"')
        result = db.session.execute(query)
        for row in result:
            pn = str(row[0]).strip().split('.')[0]
            name_map[pn] = str(row[1]).strip()
    except Exception as e:
        logging.error(f"無法從 RDS 載入名稱字典: {e}")
    return name_map


# ---------- 路由 1: 取得生產排程 ----------
@app.route('/api/get-production-plan', methods=['GET'])
def get_production_plan():
    start_date = request.args.get('start')
    end_date   = request.args.get('end')
    if not start_date or not end_date:
        return jsonify({"ok": False, "error": "缺少日期範圍參數"}), 400
    try:
        name_dict = get_mrp_name_map()
        col_query = text("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = 'production_Plan' ORDER BY ordinal_position
        """)
        actual_cols = [r[0] for r in db.session.execute(col_query)]
        pn_col_name   = next((c for c in actual_cols if re.search(r'panel.*no', c, re.I)), None)
        desc_col_name = next((c for c in actual_cols if re.search(r'desc', c, re.I)), None)
        date_cols     = [c for c in actual_cols if re.match(r'\d{4}[-/]\d{2}[-/]\d{2}', c)]
        query   = text('SELECT * FROM "production_Plan" WHERE "Plan" = \'Plan\'')
        rows    = db.session.execute(query).fetchall()
        results = []
        for row in rows:
            row_dict = dict(zip(actual_cols, row))
            raw_pn   = row_dict.get(pn_col_name)
            if not raw_pn:
                continue
            pn_key             = str(raw_pn).strip().split('.')[0]
            raw_desc           = row_dict.get(desc_col_name)
            panel_display_name = name_dict.get(pn_key, raw_desc if raw_desc else pn_key)
            for col in date_cols:
                if start_date <= col <= end_date:
                    qty = row_dict.get(col)
                    if qty and float(qty) > 0:
                        results.append({
                            "panelName": panel_display_name,
                            "PN":        pn_key,
                            "qty":       int(float(qty)),
                            "date":      col
                        })
        results.sort(key=lambda x: x['date'])
        return jsonify(results)
    except Exception as e:
        logging.error(f"❌ Get Production Plan 錯誤: {traceback.format_exc()}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ---------- 路由 2: 取得插單限制資料 ----------
@app.route('/api/get-rush-orders', methods=['GET'])
def get_rush_orders():
    try:
        query   = text('SELECT "日期", "滴定機", "Marker", "數量", "備註" FROM "限制OR插單"')
        rows    = db.session.execute(query).fetchall()
        results = [{
            "date": r[0], "titrator": r[1], "marker": r[2], "qty": r[3], "note": r[4]
        } for r in rows]
        return jsonify({"ok": True, "data": results})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ---------- 路由 3: 儲存 Beads Need 全景表 ----------
@app.route('/api/save-bead-need', methods=['POST'])
def save_bead_need():
    try:
        data       = request.json
        save_date  = data.get('date')
        table_data = data.get('tableData', [])
        if not save_date or not isinstance(table_data, list):
            return jsonify({"ok": False, "error": "資料格式錯誤或缺少日期"}), 400
        current_month = save_date.split('-')[1] if '-' in save_date else "01"

        def _safe(val):
            try:
                if val is None or val == '': return 0.0
                return float(val)
            except (ValueError, TypeError):
                return 0.0

        db.session.execute(text('DELETE FROM "BeadNeed" WHERE "date" = :d'), {"d": save_date})
        for row in table_data:
            pn = str(row.get('pn', ''))
            db.session.execute(text("""
                INSERT INTO "BeadNeed"
                    ("date","pn","name","min_batch","safety_stock","stock_unstock",
                     "w1","w2","w3","w1_batch","w2_batch","w3_batch")
                VALUES
                    (:date,:pn,:name,:mb,:ss,:su,:w1,:w2,:w3,:wb1,:wb2,:wb3)
            """), {
                "date": save_date, "pn": pn, "name": row.get('name', ''),
                "mb": _safe(row.get('minBatch')),   "ss": _safe(row.get('safetyStock')),
                "su": _safe(row.get('stockUnstock')),"w1": _safe(row.get('w1')),
                "w2": _safe(row.get('w2')),          "w3": _safe(row.get('w3')),
                "wb1": _safe(row.get('w1Batch')),    "wb2": _safe(row.get('w2Batch')),
                "wb3": _safe(row.get('w3Batch')),
            })
            db.session.execute(text("""
                INSERT INTO "beads_seasonal_safety_level"
                    ("PN","Name","Month","Seasonal_Safety_Level","Min_Batch")
                VALUES (:pn,:name,:month,:ss,:mb)
                ON CONFLICT ("PN","Month") DO UPDATE SET
                    "Seasonal_Safety_Level" = EXCLUDED."Seasonal_Safety_Level",
                    "Min_Batch"             = EXCLUDED."Min_Batch",
                    "Name"                  = EXCLUDED."Name"
            """), {
                "pn": pn, "name": row.get('name',''), "month": current_month,
                "ss": _safe(row.get('safetyStock')), "mb": _safe(row.get('minBatch')),
            })
        db.session.commit()
        logging.info(f"✅ BeadNeed 寫入成功（{save_date}，{len(table_data)} 筆）")
        return jsonify({"ok": True, "message": "RDS 數據儲存成功"})
    except Exception as e:
        db.session.rollback()
        logging.error(f"❌ save_bead_need 例外: {traceback.format_exc()}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ---------- 路由 4: 執行排程引擎 ----------
@app.route('/api/run-production-schedule', methods=['POST'])
def run_production_schedule():
    data        = request.get_json()
    target_date = data.get('date')
    batch_start = int(data.get('batch_num_start', 1))
    try:
        result = scheduler_api.generate_schedule(
            db, target_date,
            resource_config=data.get('resource'),
            batch_num_start=batch_start,
        )
        if result.get('status') != 'success':
            return jsonify({'ok': False, 'error': result.get('message')}), 400

        excel_wb    = result.get('excel_wb')
        excel_bytes = result.get('excel_bytes')

        # ── 1. 存獨立的 2D-plot Excel ────────────────────────────────────────
        if excel_bytes:
            plot_path = os.path.join(EXPORTS_DIR, f"滴定排程結果_{target_date}.xlsx")
            with open(plot_path, 'wb') as f:
                f.write(excel_bytes.getvalue())
            print(f"[Excel] 2D-plot 已存：{plot_path}")

        # ── 2. 把 2D-plot sheets 合併到 W{week} Excel ─────────────────────────
        if excel_wb:
            dt       = datetime.strptime(target_date, '%Y-%m-%d')
            week_num = dt.isocalendar()[1]
            w_path   = os.path.join(EXPORTS_DIR, f"排程_{dt.year}-W{week_num:02d}.xlsx")
            if os.path.exists(w_path):
                w_wb = openpyxl.load_workbook(w_path)
            else:
                w_wb = openpyxl.Workbook()
                if 'Sheet' in w_wb.sheetnames:
                    del w_wb['Sheet']
            for sheet_name in excel_wb.sheetnames:
                src_ws = excel_wb[sheet_name]
                if sheet_name in w_wb.sheetnames:
                    del w_wb[sheet_name]
                dst_ws = w_wb.create_sheet(title=sheet_name)
                for row in src_ws.iter_rows():
                    for cell in row:
                        dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            dst_cell.font          = cell.font.copy()
                            dst_cell.border        = cell.border.copy()
                            dst_cell.fill          = cell.fill.copy()
                            dst_cell.number_format = cell.number_format
                            dst_cell.alignment     = cell.alignment.copy()
                for col_letter, col_dim in src_ws.column_dimensions.items():
                    dst_ws.column_dimensions[col_letter].width = col_dim.width
                for merge in src_ws.merged_cells.ranges:
                    dst_ws.merge_cells(str(merge))
            w_wb.save(w_path)
            print(f"[Excel] 2D-plot sheets 已合併至：{w_path}")

        return jsonify({
            'ok':             True,
            'data':           result['data'],
            'schedule_label': result.get('schedule_label', ''),
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


# ---------- 路由 4b: 下載 excelData 檔案 ----------
@app.route('/api/download-excel', methods=['GET'])
def download_excel():
    fname = request.args.get('file', '')
    allowed = ['beads_inventory.xlsm', 'production_plan.xlsm',
               'panel_detail.xlsm', 'schedule_limit.xlsm', 'titration_limit.xlsm']
    if fname not in allowed:
        return jsonify({'ok': False, 'error': '檔案不存在'}), 404
    fpath = os.path.join(EXCEL_DATA_DIR, fname)
    if not os.path.exists(fpath):
        return jsonify({'ok': False, 'error': '檔案尚未上傳'}), 404
    return send_file(fpath, as_attachment=True, download_name=fname)


# ---------- 路由 5: VBA 上傳 Excel 檔案到 EC2 ----------
@app.route('/api/upload-excel', methods=['POST'])
def upload_excel():
    if request.headers.get('X-Api-Key') != UPLOAD_API_KEY:
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 401

    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': '沒有 file 欄位'}), 400

    f = request.files['file']
    if not f.filename:
        return jsonify({'ok': False, 'error': '檔名為空'}), 400

    # ── 用原始檔名判斷類型（secure_filename 會去掉中文）──────────────
    original_name = f.filename
    fn_upper      = original_name.upper()

    if 'BEADS' in fn_upper or '庫存' in original_name:
        save_name = 'beads_inventory.xlsm'
    elif 'PRODUCTION' in fn_upper or 'PLAN' in fn_upper:
        save_name = 'production_plan.xlsm'
    elif 'PANEL' in fn_upper or '明細' in original_name:
        save_name = 'panel_detail.xlsm'
    elif '排程' in original_name:                        # ← 先判斷排程限制
        save_name = 'schedule_limit.xlsm'
    elif '限制' in original_name or 'LIMIT' in fn_upper: # ← 再判斷配藥限制
        save_name = 'titration_limit.xlsm'
    else:
        save_name = secure_filename(original_name) or 'unknown.xlsm'

    save_path = os.path.join(EXCEL_DATA_DIR, save_name)
    f.save(save_path)
    print(f"[Upload] {original_name} → {save_path}")

    # 上傳後自動觸發 sync
    try:
        with app.test_request_context():
            sync_resp = trigger_sync()
            sync_data = sync_resp[0].get_json() if isinstance(sync_resp, tuple) else sync_resp.get_json()
    except Exception as e:
        sync_data = {'sync_error': str(e)}

    return jsonify({'ok': True, 'saved_as': save_name, 'sync': sync_data})


# ---------- 路由 6b: 輕量上傳 beads_inventory JSON（VBA 直傳資料）----------
@app.route('/api/upload-beads-json', methods=['POST'])
def upload_beads_json():
    """接收 VBA 直接傳來的 JSON 陣列，寫入 RDS，跳過檔案上傳+pandas解析。"""
    if request.headers.get('X-Api-Key') != UPLOAD_API_KEY:
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 401

    data = request.get_json(silent=True)
    if not data or not isinstance(data, list):
        return jsonify({'ok': False, 'error': '需要 JSON 陣列'}), 400

    try:
        df = pd.DataFrame(data)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how='all')

        with db.engine.begin() as conn:
            conn.execute(text('TRUNCATE TABLE schedule."beads_Inventory"'))
        df.to_sql(
            'beads_Inventory', db.engine,
            schema='schedule', if_exists='append', index=False,
        )
        return jsonify({'ok': True, 'rows': len(df)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ---------- 路由 6: SYNC：讀 excelData/ → 解析 → 寫 RDS ----------
@app.route('/api/trigger-sync', methods=['POST'])
def trigger_sync():
    errors  = []
    results = {}

    # ══════════════════════════════════════════════════════════════════════
    # 1. beads_Inventory
    # ══════════════════════════════════════════════════════════════════════
    inv_path = os.path.join(EXCEL_DATA_DIR, 'beads_inventory.xlsm')
    if not os.path.exists(inv_path):
        errors.append('beads_inventory.xlsm 尚未上傳（請先在 Excel 存檔觸發 VBA 上傳）')
    else:
        try:
            df_inv = pd.read_excel(
                inv_path,
                sheet_name='BEADS庫存表(202405~',
                header=4,         # row 5，index=4
                usecols='A:O',
                engine='openpyxl',
            )
            df_inv.columns = [str(c).strip() for c in df_inv.columns]
            df_inv = df_inv.dropna(how='all')

            with db.engine.begin() as conn:
                conn.execute(text('TRUNCATE TABLE schedule."beads_Inventory"'))
            df_inv.to_sql(
                'beads_Inventory', db.engine,
                schema='schedule', if_exists='append', index=False,
            )
            results['beads_Inventory'] = f"OK，{len(df_inv)} 筆"
            print(f"[Sync] beads_Inventory {len(df_inv)} 筆")
        except Exception as e:
            msg = f"beads_Inventory 失敗：{e}"
            errors.append(msg)
            print(f"[Sync] {msg}\n{traceback.format_exc()}")

    # ══════════════════════════════════════════════════════════════════════
    # 2. production_Plan
    # ══════════════════════════════════════════════════════════════════════
    plan_path = os.path.join(EXCEL_DATA_DIR, 'production_plan.xlsm')
    if not os.path.exists(plan_path):
        errors.append('production_plan.xlsm 尚未上傳（請先在 Excel 存檔觸發 VBA 上傳）')
    else:
        try:
            df_raw = pd.read_excel(
                plan_path,
                sheet_name='P_plan Reagent',
                header=1,         # row 2，index=1
                engine='openpyxl',
            )
            df_raw.columns = [str(c).strip() for c in df_raw.columns]

            # 找 Panel_NO 欄
            pn_col = next(
                (c for c in df_raw.columns if re.search(r'panel.*no|panel_no', c, re.I)),
                df_raw.columns[1]
            )
            # 找 Plan 欄 index
            plan_col_idx = next(
                (i for i, c in enumerate(df_raw.columns)
                 if str(c).strip().lower() == 'plan'),
                4   # fallback E欄
            )
            date_cols_raw = df_raw.columns[plan_col_idx + 1:]
            current_year  = datetime.now().year

            def parse_date_col(col_name):
                import datetime as dt_module
                # datetime 物件
                if isinstance(col_name, dt_module.datetime):
                    return col_name.strftime('%Y-%m-%d')
                s = str(col_name).strip()
                # str(datetime) 轉出的格式："2026-03-23 00:00:00"
                m = re.match(r'^(\d{4}-\d{2}-\d{2})\s', s)
                if m:
                    return m.group(1)
                # M/D 格式
                m = re.match(r'^(\d{1,2})/(\d{1,2})$', s)
                if m:
                    mo, dy = int(m.group(1)), int(m.group(2))
                    yr = datetime.now().year
                    if mo < datetime.now().month - 3:
                        yr = datetime.now().year + 1
                    try:
                        return datetime(yr, mo, dy).strftime('%Y-%m-%d')
                    except ValueError:
                        return None
                # 已是 YYYY-MM-DD
                if re.match(r'^\d{4}-\d{2}-\d{2}$', s):
                    return s
                return None

            date_col_map = {}
            for c in date_cols_raw:
                parsed = parse_date_col(c)
                if parsed:
                    date_col_map[c] = parsed
            if not date_col_map:
                raise ValueError("找不到任何有效日期欄位")

            # 過濾有效列（Panel_NO 有值，非標題列）
            df_data = df_raw[
                df_raw[pn_col].notna() &
                (df_raw[pn_col].astype(str).str.strip() != '') &
                (~df_raw[pn_col].astype(str).str.contains(
                    'Panel_NO|Panel NO', case=False, na=False))
            ].copy()

            # 建輸出 DataFrame
            out_rows = []
            for _, row in df_data.iterrows():
                new_row = {
                    'Plan':     'Plan',
                    'Panel_NO': str(row[pn_col]).strip(),
                }
                for raw_col, iso_date in date_col_map.items():
                    try:
                        qty = float(row.get(raw_col, 0) or 0)
                        new_row[iso_date] = qty if qty > 0 else None
                    except (TypeError, ValueError):
                        new_row[iso_date] = None
                out_rows.append(new_row)

            df_plan = pd.DataFrame(out_rows)

            # 自動新增缺少的日期欄位
            with db.engine.begin() as conn:
                existing = [r[0] for r in conn.execute(text(
                    "SELECT column_name FROM information_schema.columns "
                    "WHERE table_schema='schedule' AND table_name='production_Plan'"
                )).fetchall()]
                for col in df_plan.columns:
                    if col not in existing:
                        conn.execute(text(
                            f'ALTER TABLE schedule."production_Plan" ADD COLUMN "{col}" DOUBLE PRECISION'
                        ))
                        print(f"[Sync] ADD COLUMN: {col}")
                conn.execute(text('TRUNCATE TABLE schedule."production_Plan"'))
            df_plan.to_sql(
                'production_Plan', db.engine,
                schema='schedule', if_exists='append', index=False,
            )
            results['production_Plan'] = (
                f"OK，{len(df_plan)} 筆，{len(date_col_map)} 個日期欄"
            )
            print(f"[Sync] production_Plan {len(df_plan)} 筆")

        except Exception as e:
            msg = f"production_Plan 失敗：{e}"
            errors.append(msg)
            print(f"[Sync] {msg}\n{traceback.format_exc()}")
    # ══════════════════════════════════════════════════════════════════════
    # 3. 配藥限制 → schedule."配藥限制"
    # ══════════════════════════════════════════════════════════════════════
    limit_path = os.path.join(EXCEL_DATA_DIR, 'titration_limit.xlsm')
    if not os.path.exists(limit_path):
        errors.append('titration_limit.xlsm 尚未上傳')
    else:
        try:
            df_limit = pd.read_excel(
                limit_path,
                sheet_name='配藥限制',
                header=0,        # row 1
                engine='openpyxl',
            )
            df_limit.columns = [str(c).strip() for c in df_limit.columns]
            df_limit = df_limit.dropna(how='all')

            # PN 欄清理（去除 .0）
            if 'PN' in df_limit.columns:
                df_limit['PN'] = df_limit['PN'].apply(
                    lambda x: str(int(x)) if pd.notna(x) and str(x).replace('.','').isdigit()
                    else str(x).strip() if pd.notna(x) else None
                )
                df_limit = df_limit[df_limit['PN'].notna()]

            with db.engine.begin() as conn:
                conn.execute(text('TRUNCATE TABLE schedule."配藥限制"'))
            df_limit.to_sql(
                '配藥限制', db.engine,
                schema='schedule', if_exists='append', index=False,
            )
            results['配藥限制'] = f"OK，{len(df_limit)} 筆"
            print(f"[Sync] 配藥限制 {len(df_limit)} 筆")

        except Exception as e:
            msg = f"配藥限制 失敗：{e}"
            errors.append(msg)
            print(f"[Sync] {msg}\n{traceback.format_exc()}")
    # ══════════════════════════════════════════════════════════════════════
    # 4. 限制OR插單 → schedule."限制OR插單"
    # ══════════════════════════════════════════════════════════════════════
    sched_limit_path = os.path.join(EXCEL_DATA_DIR, 'schedule_limit.xlsm')
    if not os.path.exists(sched_limit_path):
        errors.append('schedule_limit.xlsm 尚未上傳')
    else:
        try:
            df_sl = pd.read_excel(
                sched_limit_path,
                sheet_name='限制OR插單',
                header=0,         # row 1
                engine='openpyxl',
            )
            # 欄位名稱清理（去空白）
            df_sl.columns = [str(c).strip() for c in df_sl.columns]
            df_sl = df_sl.dropna(how='all')

            # 日期欄轉 YYYY-MM-DD 字串
            if '日期' in df_sl.columns:
                df_sl['日期'] = df_sl['日期'].apply(
                    lambda x: x.strftime('%Y-%m-%d')
                    if hasattr(x, 'strftime') else str(x).strip()
                )

            # 時間欄轉 HH:MM 字串
            for time_col in ['RD給藥時間', '預計滴定時間', '預計結束']:
                if time_col in df_sl.columns:
                    df_sl[time_col] = df_sl[time_col].apply(
                        lambda x: x.strftime('%H:%M')
                        if hasattr(x, 'strftime') else (str(x).strip() if pd.notna(x) else None)
                    )

            # 數量欄清理
            qty_col = next((c for c in df_sl.columns if '數量' in c), None)
            if qty_col and qty_col != '數量':
                df_sl = df_sl.rename(columns={qty_col: '數量'})

            with db.engine.begin() as conn:
                conn.execute(text('TRUNCATE TABLE schedule."限制OR插單"'))
            df_sl.to_sql(
                '限制OR插單', db.engine,
                schema='schedule', if_exists='append', index=False,
            )
            results['限制OR插單'] = f"OK，{len(df_sl)} 筆"
            print(f"[Sync] 限制OR插單 {len(df_sl)} 筆")

        except Exception as e:
            msg = f"限制OR插單 失敗：{e}"
            errors.append(msg)
            print(f"[Sync] {msg}\n{traceback.format_exc()}")
    # ══════════════════════════════════════════════════════════════════════
    # 5. BOM_Details ← Panel_明細.xlsm
    # ══════════════════════════════════════════════════════════════════════
    panel_path = os.path.join(EXCEL_DATA_DIR, 'panel_detail.xlsm')
    if not os.path.exists(panel_path):
        errors.append('panel_detail.xlsm 尚未上傳')
    else:
        try:
            wb_panel = openpyxl.load_workbook(panel_path, read_only=True)
            ws_panel = wb_panel.worksheets[0]
            panel_rows = list(ws_panel.iter_rows(values_only=True))

            COMP_START = 3   # D欄 0-based index

            def _clean_bom(v):
                if v is None: return ''
                s = str(v).strip().replace('\n', ' ')
                m = re.fullmatch(r'(\d+)\.0+', s)
                return m.group(1) if m else s

            bom_records = []
            total_r = len(panel_rows)
            for i in range(0, total_r - 2, 3):
                r1, r2, r3 = panel_rows[i], panel_rows[i+1], panel_rows[i+2]
                part_nos = [_clean_bom(r1[0]),
                            _clean_bom(r2[0]),
                            _clean_bom(r3[0])]
                part_nos = [p for p in part_nos if p]
                last_col = max(len(r1), len(r2), len(r3))
                for j in range(COMP_START, last_col):
                    comp_no   = _clean_bom(r1[j]) if j < len(r1) else ''
                    comp_name = _clean_bom(r2[j]) if j < len(r2) else ''
                    qty_raw   = r3[j]              if j < len(r3) else None
                    if not comp_no or qty_raw is None:
                        continue
                    try:
                        qty = float(qty_raw)
                    except (TypeError, ValueError):
                        continue
                    if qty == 0:
                        continue
                    for pn in part_nos:
                        bom_records.append({
                            'Finished_PartNo':  pn,
                            'Component_No':     comp_no,
                            'Component_Name':   comp_name,
                            'Quantity':         qty,
                        })

            df_bom = pd.DataFrame(bom_records)

            with db.engine.begin() as conn:
                conn.execute(text('TRUNCATE TABLE schedule."BOM_Details"'))
            df_bom.to_sql(
                'BOM_Details', db.engine,
                schema='schedule', if_exists='append', index=False,
            )
            results['BOM_Details'] = f"OK，{len(df_bom)} 筆"
            print(f"[Sync] BOM_Details {len(df_bom)} 筆")

        except Exception as e:
            msg = f"BOM_Details 失敗：{e}"
            errors.append(msg)
            print(f"[Sync] {msg}\n{traceback.format_exc()}")
            
    if errors:
        return jsonify({'ok': False, 'results': results, 'errors': errors}), 500
    return jsonify({'ok': True, 'results': results})


# ---------- 路由 7: 儲存插單 ----------
@app.route('/api/save-rush-orders', methods=['POST'])
def save_rush_orders():
    try:
        data   = request.json
        orders = data.get('orders', [])
        db.session.execute(text('DELETE FROM "限制OR插單"'))
        for o in orders:
            if not o.get('marker'):
                continue
            db.session.execute(text("""
                INSERT INTO "限制OR插單" ("日期","滴定機","Marker","數量","備註")
                VALUES (:date,:titrator,:marker,:qty,:note)
            """), {
                "date":     o.get('date', ''),   "titrator": o.get('titrator', ''),
                "marker":   o.get('marker', ''), "qty":      o.get('qty', 0),
                "note":     o.get('note', ''),
            })
        db.session.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


# ---------- 路由 8: run-beads-analysis ----------
@app.route('/api/run-beads-analysis', methods=['POST'])
def run_beads_analysis():
    try:
        data            = request.get_json()
        target_date_str = data.get('date')
        resource        = data.get('resource', {})
        batch_num_start = int(data.get('batch_num_start', 1))

        from datetime import date as date_type, timedelta
        from collections import defaultdict

        def _to_number(x):
            if x is None: return 0
            if isinstance(x, (int, float)): return x
            s = str(x).strip().replace(',', '')
            try: return float(s)
            except: return 0

        def _extract_numbers(x):
            if x is None: return [0.0]
            if isinstance(x, (int, float)): return [float(x)]
            nums = re.findall(r'\d+(?:\.\d+)?', str(x))
            return [float(n) for n in nums] if nums else [0.0]

        def _normalize_pn(x):
            if x is None: return ""
            s = str(x).strip()
            m = re.fullmatch(r'(\d+)(?:\.0+)?', s)
            return m.group(1) if m else s

        def _pick_batch(candidates, demand_3w):
            above = [c for c in candidates if c >= demand_3w]
            return min(above) if above else max(candidates)

        d_day         = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        current_month = d_day.strftime('%m')
        engine        = db.engine

        safe_dict = defaultdict(lambda: defaultdict(float))
        with engine.connect() as conn:
            rows = conn.execute(text(
                'SELECT "PN","Month","Seasonal_Safety_Level" FROM "beads_seasonal_safety_level"'
            )).fetchall()
        for r in rows:
            safe_dict[_normalize_pn(r[0])][str(r[1]).zfill(2)] = _to_number(r[2])

        batch_dict = defaultdict(lambda: [0.0])
        with engine.connect() as conn:
            rows = conn.execute(text('SELECT "PN","數量" FROM "配藥限制"')).fetchall()
        for r in rows:
            pn = _normalize_pn(r[0])
            if pn: batch_dict[pn] = _extract_numbers(r[1])

        inv_data = defaultdict(lambda: {'real_stock': 0.0, 'unstocked': 0.0})
        with engine.connect() as conn:
            rows = conn.execute(text(
                'SELECT "PN","工單數","入庫","可使用庫存","累計領用" FROM "beads_Inventory"'
            )).fetchall()
        for r in rows:
            pn     = _normalize_pn(r[0])
            avail  = _to_number(r[3])
            wo_qty = _to_number(r[1])
            accum  = _to_number(r[4])
            in_st  = str(r[2] or '').strip()
            if avail >= 500:
                inv_data[pn]['real_stock'] += avail
            in_st_num = _to_number(in_st) if in_st else 0
            if avail == 0 and in_st_num > 0 and accum == 0:
                inv_data[pn]['unstocked'] += wo_qty

        bom_map = defaultdict(list)
        with engine.connect() as conn:
            rows = conn.execute(text(
                'SELECT "Finished_PartNo","Component_No","Quantity" FROM "BOM_Details"'
            )).fetchall()
        for r in rows:
            bom_map[_normalize_pn(r[0])].append({
                'bpn': _normalize_pn(r[1]), 'qty': _to_number(r[2])
            })

        bead_to_name = {}
        with engine.connect() as conn:
            rows = conn.execute(text('SELECT "PN","BEADS別" FROM "beads_Inventory"')).fetchall()
        for r in rows:
            pn = _normalize_pn(r[0])
            if pn: bead_to_name[pn] = str(r[1] or '').strip()
        with engine.connect() as conn:
            rows = conn.execute(text(
                'SELECT "Component_No","Component_Name" FROM "BOM_Details"'
            )).fetchall()
        for r in rows:
            pn = _normalize_pn(r[0])
            if pn and pn not in bead_to_name:
                bead_to_name[pn] = str(r[1] or '').strip()

        with engine.connect() as conn:
            actual_cols = [r[0] for r in conn.execute(text(
                "SELECT column_name FROM information_schema.columns "
                "WHERE table_name='production_Plan' ORDER BY ordinal_position"
            )).fetchall()]
            plans = conn.execute(text(
                'SELECT * FROM "production_Plan" WHERE "Plan"=\'Plan\''
            )).fetchall()

        pn_col   = next((c for c in actual_cols if re.search(r'panel.*no', c, re.I)), 'Panel_NO')
        plan_idx = next((i for i, c in enumerate(actual_cols) if c.strip().lower() == 'plan'), 0)
        date_cols = sorted([c for c in actual_cols[plan_idx+1:]
                            if re.match(r'\d{4}[-/]\d{2}[-/]\d{2}', c)])

        bead_needs      = defaultdict(lambda: [0, 0, 0])
        bead_earliest   = {}
        shortage_events = []
        sim_inv         = {pn: d['real_stock'] + d['unstocked'] for pn, d in inv_data.items()}

        for d_col in date_cols:
            try:
                target_dt = datetime.strptime(d_col, '%Y-%m-%d').date()
            except:
                continue
            if target_dt < d_day:
                continue
            diff = (target_dt - d_day).days
            for row in plans:
                row_dict = dict(zip(actual_cols, row))
                p_pn = _normalize_pn(row_dict.get(pn_col))
                qty  = _to_number(row_dict.get(d_col))
                if qty <= 0:
                    continue
                for comp in bom_map.get(p_pn, []):
                    bpn    = comp['bpn']
                    needed = comp['qty'] * qty
                    if diff < 7:    bead_needs[bpn][0] += needed
                    elif diff < 14: bead_needs[bpn][1] += needed
                    elif diff < 21: bead_needs[bpn][2] += needed
                    if diff < 21:
                        stock = sim_inv.get(bpn, 0.0)
                        if stock < needed:
                            shortage_events.append({
                                'date':         target_dt.strftime('%Y-%m-%d'),
                                'bead_pn':      bpn,
                                'bead_name':    bead_to_name.get(bpn, 'Unknown'),
                                'shortage_qty': int(needed - stock)
                            })
                            if bpn not in bead_earliest:
                                bead_earliest[bpn] = target_dt.strftime('%Y-%m-%d')
                        sim_inv[bpn] = stock - needed

        KIT_DEFS = [
            {'name': 'CREA套組',      'members': {'5714400180':1,'5714400181':1,'5714400182':2}},
            {'name': 'GGT/RGT(1:1)', 'members': {'5714400132':1,'5714400201':1}},
        ]
        auto_pairs = defaultdict(list)
        for bpn, name in bead_to_name.items():
            if any(bpn in k['members'] for k in KIT_DEFS): continue
            clean = str(name).strip().upper()
            if re.search(r'-[A-Z]*[UD]$', clean):
                base = re.sub(r'-[A-Z]*[UD]$', '', clean).strip()
                auto_pairs[base].append(bpn)
        for base, bpns in auto_pairs.items():
            if len(bpns) > 1:
                KIT_DEFS.append({'name': f'{base}套組', 'members': {b:1 for b in bpns}})

        active_beads = set(bead_needs.keys())
        for bpn, d in inv_data.items():
            if d['real_stock'] < safe_dict[bpn].get(current_month, 0):
                active_beads.add(bpn)
        for kit in KIT_DEFS:
            if any(b in active_beads for b in kit['members']):
                for b in kit['members']: active_beads.add(b)

        res = []
        for bpn in active_beads:
            n          = bead_needs[bpn]
            real_stock = inv_data[bpn]['real_stock']
            unstocked  = inv_data[bpn]['unstocked']
            total_inv  = real_stock + unstocked
            safe_qty   = safe_dict[bpn].get(current_month, 0)
            demand_3w  = n[0] + n[1] + n[2]
            m_batch    = _pick_batch(batch_dict.get(bpn, [0.0]), demand_3w)
            projected_w1   = total_inv - n[0]
            status         = 'SAFE'
            suggested_prod = 0
            if projected_w1 < 0:
                status         = 'CRITICAL'
                suggested_prod = m_batch if m_batch > 0 else int(abs(projected_w1) + safe_qty)
            elif projected_w1 < safe_qty:
                status         = 'WARNING'
                suggested_prod = m_batch if m_batch > 0 else safe_qty
            res.append({
                'partNumber':    bpn,
                'description':   bead_to_name.get(bpn, 'Unknown Item'),
                'inventory':     real_stock,
                'unstocked':     unstocked,
                'w1_demand':     int(n[0]),
                'w2_demand':     int(n[1]),
                'w3_demand':     int(n[2]),
                'neededQuantity': int(demand_3w),
                'earliest_need': bead_earliest.get(bpn, 'Future'),
                'safetyStock':   safe_qty,
                'minBatch':      m_batch,
                'status':        status,
                'suggestedProd': suggested_prod,
            })

        res_dict = {x['partNumber']: x for x in res}
        for kit in KIT_DEFS:
            max_mult = 0
            trigger  = False
            for bpn, ratio in kit['members'].items():
                if bpn in res_dict and res_dict[bpn]['status'] in ('WARNING','CRITICAL'):
                    trigger   = True
                    base_mult = math.ceil(res_dict[bpn]['suggestedProd'] / ratio) if ratio else 0
                    if base_mult > max_mult: max_mult = base_mult
            if trigger:
                for bpn, ratio in kit['members'].items():
                    if bpn in res_dict:
                        sync_qty = max_mult * ratio
                        if res_dict[bpn]['suggestedProd'] < sync_qty:
                            res_dict[bpn]['suggestedProd'] = sync_qty
                        if res_dict[bpn]['status'] == 'SAFE':
                            res_dict[bpn]['status']      = 'WARNING'
                            res_dict[bpn]['description'] = f"🔗[齊套] {res_dict[bpn]['description']}"

        critical_count = sum(1 for x in res if x['status'] == 'CRITICAL')
        warning_count  = sum(1 for x in res if x['status'] == 'WARNING')

        sorted_res = sorted(res, key=lambda x: (
            x['status'] != 'CRITICAL', x['status'] != 'WARNING', x['earliest_need']))

        return jsonify({
            'ok':              True,
            'components':      sorted_res,
            'shortage_events': sorted(shortage_events, key=lambda x: x['date']),
            'aiSummary': {
                'totalItems':    len(res),
                'criticalCount': critical_count,
                'warningCount':  warning_count,
            }
        })

    except Exception as e:
        logging.error(f"run_beads_analysis 例外: {traceback.format_exc()}")
        return jsonify({'ok': False, 'error': str(e)}), 500


# ---------- SkylaiCloud API Config ----------
SKYLAI_API_BASE = os.environ.get('SKYLAI_API_BASE_URL', 'https://api.skylaicloud.com.tw')
SKYLAI_API_TOKEN = os.environ.get('SKYLAI_API_TOKEN', '43488|TnN58D0tNwbwrRjFYXbavtBvrmzuDtATunXP3Jwy')
SKYLAI_PROJECT_ID = int(os.environ.get('SKYLAI_PROJECT_ID', '1'))

# ---------- Initialize panel_dispatch table ----------
with app.app_context():
    try:
        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS panel_production.panel_dispatch (
                id              SERIAL PRIMARY KEY,
                device_sn       VARCHAR(50) NOT NULL,
                position        INTEGER NOT NULL,
                work_order_no   VARCHAR(50),
                lot_no          VARCHAR(50),
                line            VARCHAR(10),
                well_assignments JSONB,
                dispatched_at   TIMESTAMP DEFAULT NOW(),
                UNIQUE(device_sn, position, dispatched_at)
            );
        """))
        db.session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_panel_dispatch_device
                ON panel_production.panel_dispatch (device_sn, dispatched_at DESC);
        """))
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logging.error(f"Failed to initialize panel_dispatch table: {e}")


# ---------- Tutti: Upload Excel → Parse → RDS ----------
TEMP_DIR = os.path.join(BASE_DIR, 'temp')
TEMP_LOG_DIR = os.path.join(TEMP_DIR, 'log')
os.makedirs(TEMP_LOG_DIR, exist_ok=True)

def _parse_tutti_excel(path):
    """Parse Qbi製程記錄表 Excel into tutti_work_orders form_data format.
    Returns a list of form_data dicts, one per OW/ow sheet."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ow_sheets = [s for s in wb.sheetnames if s.lower().startswith('ow')]
    if not ow_sheets:
        wb.close()
        raise ValueError('Excel 中找不到 OW/ow 開頭的 sheet')
    results = []
    for sheet_name in ow_sheets:
        results.append(_parse_tutti_sheet(wb, sheet_name))
    wb.close()
    return results


def _parse_tutti_sheet(wb, sheet_name):
    """Parse a single OW sheet."""
    import re
    ws = wb[sheet_name]

    def cv(row, col):
        v = ws.cell(row, col).value
        if v is None: return ''
        if hasattr(v, 'isoformat'): return v.isoformat()[:10]
        s = str(v).strip()
        # Parse "生產日期： 2026 年 3 月 5 日" format
        m = re.search(r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日', s)
        if m:
            return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'
        return s

    # Header
    header = {
        'lotNo': cv(5, 3),
        'workOrderNumber': cv(6, 3),
        'codeA': cv(7, 3),
        'productionOrderQty': cv(8, 3),
        'modelPn': cv(9, 3),
        'date': cv(5, 15),
        'formTitle': cv(5, 6) or cv(2, 6) or 'Panel 製程紀錄表',
    }

    # Materials (rows 12-22)
    materials = []
    for r in range(12, 23):
        materials.append({
            'operationName': cv(r, 1),
            'itemName': cv(r, 2),
            'partNumber': cv(r, 5),
            'version': cv(r, 8),
            'holeNumber': cv(r, 9),
            'batchNumber': cv(r, 11),
        })

    # Well rows helper
    def parse_wells(start_row, line_label, default_slots):
        rows = []
        for i in range(10):
            r = start_row + i
            rows.append({
                'sheetName': 'RL' if i == 0 else '',
                'line': line_label if i == 0 else '',
                'wellPosition': str(i + 1),
                'slot1': cv(r, 5) or (default_slots[i][0] if i < len(default_slots) else ''),
                'reagentName1': cv(r, 6),
                'batch1': cv(r, 7),
                'qty1': cv(r, 8),
                'slot2': cv(r, 9) or (default_slots[i][1] if i < len(default_slots) else ''),
                'reagentName2': cv(r, 10),
                'batch2': cv(r, 11),
                'qty2': cv(r, 12),
                'formulaNumber': cv(start_row, 13) if i == 0 else '',
                'weldingParam': cv(start_row, 14) if i == 0 else '',
                'productionQty': cv(start_row, 15) if i == 0 else '',
                'defectQty': cv(start_row, 16) if i == 0 else '',
                'qaInspection': cv(start_row, 17) if i == 0 else '',
            })
        return rows

    L1_SLOTS = [('1A1','1D1'),('1C1','1F1'),('1B1','1E1'),('1A2','1D2'),('1C2','1F2'),('1B2','1E2'),('1A3','1D3'),('1C3','1F3'),('1B3','1E3'),('1A4','1D4')]
    L2_SLOTS = [('2A1','2D1'),('2C1','2F1'),('2B1','2E1'),('2A2','2D2'),('2C2','2F2'),('2B2','2E2'),('2A3','2D3'),('2C3','2F3'),('2B3','2E3'),('2A4','2D4')]

    wellL1 = parse_wells(25, 'L1', L1_SLOTS)
    wellL2 = parse_wells(35, 'L2', L2_SLOTS)
    wellL3 = parse_wells(45, 'L3', [])

    # Post-process (rows 57-61)
    post = []
    for r in range(57, 62):
        post.append({
            'operationName': cv(r, 1),
            'fileName': cv(r, 4),
            'formulaNumber': cv(r, 6),
            'productBatch': cv(57, 7) if r == 57 else '',
            'productExpiry': cv(56, 10) if r == 57 else '',
            'confirmedBy': cv(r, 13),
            'confirmedDate': cv(r, 16),
            'remark': cv(59, 13) if r == 59 else '',
        })

    return {
        'header': header,
        'materials': materials,
        'wells': {'L1': wellL1, 'L2': wellL2, 'L3': wellL3},
        'postProcess': post,
        'sheetName': sheet_name,
    }


@app.route('/api/tutti-production/upload-excel', methods=['POST'])
def upload_tutti_excel():
    """Upload 製程記錄表 Excel → parse → push to RDS → delete file."""
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': '缺少 file 欄位'}), 400

    f = request.files['file']
    if not f.filename:
        return jsonify({'ok': False, 'error': '檔名為空'}), 400

    original_name = f.filename
    save_path = os.path.join(TEMP_DIR, original_name)
    f.save(save_path)

    try:
        form_data_list = _parse_tutti_excel(save_path)
        results = []

        for form_data in form_data_list:
            work_order_no = form_data['header']['workOrderNumber']
            lot_no = form_data['header']['lotNo']

            if not work_order_no and not lot_no:
                results.append({'sheet': form_data.get('sheetName', ''), 'skipped': True, 'reason': '工單號碼與 LOT NO 皆為空'})
                continue

            new_json = json.dumps(form_data, sort_keys=True, ensure_ascii=False)

            # Lookup priority: 1) lot_no + work_order_no  2) lot_no  3) work_order_no
            existing = None
            if lot_no and work_order_no:
                existing = db.session.execute(
                    text("SELECT id, form_data FROM panel_production.tutti_work_orders WHERE lot_no = :lot AND work_order_no = :wo"),
                    {"lot": lot_no, "wo": work_order_no}
                ).fetchone()
            if not existing and lot_no:
                existing = db.session.execute(
                    text("SELECT id, form_data FROM panel_production.tutti_work_orders WHERE lot_no = :lot"),
                    {"lot": lot_no}
                ).fetchone()
            if not existing and work_order_no:
                existing = db.session.execute(
                    text("SELECT id, form_data FROM panel_production.tutti_work_orders WHERE work_order_no = :wo"),
                    {"wo": work_order_no}
                ).fetchone()

            if existing:
                old_json = json.dumps(existing[1], sort_keys=True, ensure_ascii=False) if isinstance(existing[1], dict) else json.dumps(json.loads(existing[1]), sort_keys=True, ensure_ascii=False)
                if old_json == new_json:
                    results.append({'sheet': form_data.get('sheetName', ''), 'work_order_no': work_order_no, 'lot_no': lot_no, 'skipped': True, 'reason': '內容無變更'})
                    continue
                db.session.execute(text("""
                    UPDATE panel_production.tutti_work_orders
                    SET form_data = :form_data, work_order_no = :wo, lot_no = :lot, updated_at = NOW()
                    WHERE id = :id
                """), {"form_data": new_json, "wo": work_order_no or None, "lot": lot_no or None, "id": existing[0]})
                db.session.commit()
                results.append({'sheet': form_data.get('sheetName', ''), 'work_order_no': work_order_no, 'lot_no': lot_no, 'updated': True})
            else:
                db.session.execute(text("""
                    INSERT INTO panel_production.tutti_work_orders (work_order_no, lot_no, form_data)
                    VALUES (:wo, :lot_no, :form_data)
                """), {"wo": work_order_no or None, "lot_no": lot_no or None, "form_data": new_json})
                db.session.commit()
                results.append({'sheet': form_data.get('sheetName', ''), 'work_order_no': work_order_no, 'lot_no': lot_no, 'inserted': True})

            log_line = f"{datetime.now().isoformat()} | UPLOAD | {original_name} | SHEET={form_data.get('sheetName','')} | WO={work_order_no} | LOT={lot_no} | OK\n"
            with open(os.path.join(TEMP_LOG_DIR, 'upload.log'), 'a') as lf:
                lf.write(log_line)

        os.remove(save_path)
        return jsonify({'ok': True, 'sheets': len(form_data_list), 'results': results})

    except Exception as e:
        db.session.rollback()
        # Log error
        log_line = f"{datetime.now().isoformat()} | UPLOAD | {original_name} | ERROR: {str(e)}\n"
        with open(os.path.join(TEMP_LOG_DIR, 'upload.log'), 'a') as lf:
            lf.write(log_line)
        # Clean up temp file
        if os.path.exists(save_path):
            os.remove(save_path)
        return jsonify({'ok': False, 'error': str(e)}), 500


# ---------- Tutti Production Endpoints (JSONB form storage) ----------

@app.route('/api/tutti-production', methods=['GET'])
def list_tutti_work_orders():
    """List all tutti work orders (summary only)."""
    try:
        rows = db.session.execute(text("""
            SELECT id, work_order_no, lot_no, created_at, updated_at
            FROM panel_production.tutti_work_orders
            ORDER BY updated_at DESC
        """)).fetchall()

        results = []
        for row in rows:
            results.append({
                "id": row[0],
                "work_order_no": row[1],
                "lot_no": row[2],
                "created_at": row[3].isoformat() if row[3] else None,
                "updated_at": row[4].isoformat() if row[4] else None,
            })
        return jsonify(results)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/tutti-production/form', methods=['POST'])
def save_tutti_form():
    """Save the entire Tutti work order form as JSONB."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"ok": False, "error": "缺少請求資料"}), 400

        header = data.get('header', {})
        work_order_no = header.get('workOrderNumber', '').strip()
        lot_no = header.get('lotNo', '').strip()

        if not work_order_no:
            return jsonify({"ok": False, "error": "工單號碼為必填欄位"}), 400

        if not lot_no:
            return jsonify({"ok": False, "error": "LOT NO 為必填欄位"}), 400

        # Check if record exists (upsert by work_order_no + lot_no composite key)
        existing = db.session.execute(
            text("SELECT id FROM panel_production.tutti_work_orders WHERE work_order_no = :wo AND lot_no = :lot"),
            {"wo": work_order_no, "lot": lot_no}
        ).fetchone()

        if existing:
            # Update
            db.session.execute(text("""
                UPDATE panel_production.tutti_work_orders
                SET form_data = :form_data, updated_at = NOW()
                WHERE work_order_no = :wo AND lot_no = :lot
            """), {"form_data": json.dumps(data), "wo": work_order_no, "lot": lot_no})
        else:
            # Insert
            db.session.execute(text("""
                INSERT INTO panel_production.tutti_work_orders (work_order_no, lot_no, form_data)
                VALUES (:wo, :lot_no, :form_data)
            """), {"wo": work_order_no, "lot_no": lot_no, "form_data": json.dumps(data)})

        db.session.commit()
        return jsonify({"ok": True, "work_order_no": work_order_no, "lot_no": lot_no})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/tutti-production/form', methods=['GET'])
def load_tutti_form():
    """Load a Tutti work order form by work_order, lot_no, or mfg_lot_no."""
    try:
        work_order = request.args.get('work_order', '').strip()
        lot_no = request.args.get('lot_no', '').strip()
        mfg_lot_no = request.args.get('mfg_lot_no', '').strip()

        if not work_order and not lot_no and not mfg_lot_no:
            return jsonify({"ok": False, "error": "請提供 work_order 或 mfg_lot_no 參數"}), 400

        row = None
        if mfg_lot_no:
            # Search by lot_no field (mfg_lot_no)
            row = db.session.execute(
                text("SELECT form_data FROM panel_production.tutti_work_orders WHERE lot_no = :lot ORDER BY updated_at DESC LIMIT 1"),
                {"lot": mfg_lot_no}
            ).fetchone()
        elif work_order and lot_no:
            row = db.session.execute(
                text("SELECT form_data FROM panel_production.tutti_work_orders WHERE work_order_no = :wo AND lot_no = :lot"),
                {"wo": work_order, "lot": lot_no}
            ).fetchone()
        else:
            row = db.session.execute(
                text("SELECT form_data FROM panel_production.tutti_work_orders WHERE work_order_no = :wo ORDER BY updated_at DESC LIMIT 1"),
                {"wo": work_order}
            ).fetchone()

        if not row:
            return jsonify({"ok": False, "error": "找不到該工單記錄"}), 404

        form_data = row[0]
        if isinstance(form_data, str):
            form_data = json.loads(form_data)

        return jsonify(form_data)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/tutti-production/assay-records', methods=['POST'])
def save_assay_records():
    """Save AssayProcess test results linked to a work order."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"ok": False, "error": "缺少請求資料"}), 400

        work_order_no = data.get('work_order_no', '').strip()
        lot_no = data.get('lot_no', '').strip()
        records = data.get('records', [])

        if not work_order_no:
            return jsonify({"ok": False, "error": "work_order_no 為必填"}), 400
        if not records:
            return jsonify({"ok": False, "error": "records 不可為空"}), 400

        # Find work order
        wo = db.session.execute(
            text("SELECT id FROM panel_production.tutti_work_orders WHERE work_order_no = :wo AND (:lot = '' OR lot_no = :lot) ORDER BY updated_at DESC LIMIT 1"),
            {"wo": work_order_no, "lot": lot_no}
        ).fetchone()
        if not wo:
            return jsonify({"ok": False, "error": f"找不到工單 {work_order_no}"}), 404
        work_order_id = wo[0]

        inserted = 0
        for rec in records:
            db.session.execute(text("""
                INSERT INTO panel_production.assay_process_records
                (work_order_id, work_order_no, lot_no, device_sn, panel_name,
                 analyze_date, analyze_time, sample_type, species, patient_id,
                 lot_code, mfg_lot_no, analyze_item, analyze_result, unit,
                 test_zone, test_well, baseline, baseline_equation,
                 final_delta_od, cal_od, equation, eq_type, raw_data)
                VALUES
                (:work_order_id, :work_order_no, :lot_no, :device_sn, :panel_name,
                 :analyze_date, :analyze_time, :sample_type, :species, :patient_id,
                 :lot_code, :mfg_lot_no, :analyze_item, :analyze_result, :unit,
                 :test_zone, :test_well, :baseline, :baseline_equation,
                 :final_delta_od, :cal_od, :equation, :eq_type, :raw_data)
                ON CONFLICT (work_order_id, device_sn, panel_name, analyze_date, analyze_time, patient_id, analyze_item, test_well)
                DO UPDATE SET
                  analyze_result = EXCLUDED.analyze_result,
                  baseline = EXCLUDED.baseline,
                  baseline_equation = EXCLUDED.baseline_equation,
                  raw_data = EXCLUDED.raw_data
            """), {
                "work_order_id": work_order_id,
                "work_order_no": work_order_no,
                "lot_no": lot_no or rec.get('lot_no', ''),
                "device_sn": rec.get('device_sn', ''),
                "panel_name": rec.get('panel_name', ''),
                "analyze_date": rec.get('analyze_date', ''),
                "analyze_time": rec.get('analyze_time', ''),
                "sample_type": rec.get('sample_type', ''),
                "species": rec.get('species', ''),
                "patient_id": rec.get('patient_id', ''),
                "lot_code": rec.get('lot_code', ''),
                "mfg_lot_no": rec.get('mfg_lot_no', ''),
                "analyze_item": rec.get('analyze_item', ''),
                "analyze_result": rec.get('analyze_result', ''),
                "unit": rec.get('unit', ''),
                "test_zone": rec.get('test_zone', ''),
                "test_well": rec.get('test_well', ''),
                "baseline": rec.get('baseline', ''),
                "baseline_equation": rec.get('baseline_equation', ''),
                "final_delta_od": rec.get('final_delta_od', ''),
                "cal_od": rec.get('cal_od', ''),
                "equation": rec.get('equation', ''),
                "eq_type": rec.get('eq_type', ''),
                "raw_data": json.dumps(rec.get('raw_data')) if rec.get('raw_data') else None,
            })
            inserted += 1

        db.session.commit()
        return jsonify({"ok": True, "inserted": inserted, "work_order_id": work_order_id})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/tutti-production/assay-records', methods=['GET'])
def get_assay_records():
    """Get AssayProcess records for a work order."""
    try:
        work_order_no = request.args.get('work_order', '').strip()
        lot_no = request.args.get('lot_no', '').strip()
        if not work_order_no:
            return jsonify({"ok": False, "error": "請提供 work_order 參數"}), 400

        lot_clause = "AND lot_no = :lot" if lot_no else ""
        rows = db.session.execute(text(f"""
            SELECT id, device_sn, panel_name, analyze_date, analyze_time,
                   patient_id, lot_code, mfg_lot_no, analyze_item,
                   analyze_result, unit, test_well, baseline, baseline_equation, created_at
            FROM panel_production.assay_process_records
            WHERE work_order_no = :wo {lot_clause}
            ORDER BY analyze_date, analyze_time, analyze_item
        """), {"wo": work_order_no, "lot": lot_no} if lot_no else {"wo": work_order_no}).fetchall()

        results = [{
            "id": r[0], "device_sn": r[1], "panel_name": r[2],
            "analyze_date": r[3], "analyze_time": r[4], "patient_id": r[5],
            "lot_code": r[6], "mfg_lot_no": r[7], "analyze_item": r[8],
            "analyze_result": r[9], "unit": r[10], "test_well": r[11],
            "baseline": r[12], "baseline_equation": r[13],
            "created_at": r[14].isoformat() if r[14] else None,
        } for r in rows]

        return jsonify({"ok": True, "records": results, "total": len(results)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=3001, debug=False)


# ---------- Confirm Build Line (Panel 管理 → 確認建線) ----------
@app.route('/api/tutti-production/confirm-build-line', methods=['POST'])
def confirm_build_line():
    """Write panel_dispatch + pre-create baseline=true placeholder records."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'ok': False, 'error': '缺少請求資料'}), 400

        machine_id = data.get('machineId', '').strip()
        panels = data.get('panels', [])

        if not machine_id:
            return jsonify({'ok': False, 'error': '缺少 machineId'}), 400
        if not panels:
            return jsonify({'ok': False, 'error': '缺少 panels 資料'}), 400

        dispatched_at = datetime.now()
        dispatch_count = 0
        baseline_count = 0

        for panel in panels:
            pos = panel.get('position')
            work_order_no = panel.get('workOrder', '')
            lot_no = panel.get('lotNo', '')
            line = panel.get('line', '')
            assignments = panel.get('assignments', [])

            if not pos or not lot_no:
                continue

            # Write panel_dispatch
            db.session.execute(text("""
                INSERT INTO panel_production.panel_dispatch
                    (device_sn, position, work_order_no, lot_no, line, well_assignments, dispatched_at)
                VALUES (:device_sn, :pos, :wo, :lot, :line, :assignments, :dispatched_at)
            """), {
                'device_sn': machine_id,
                'pos': pos,
                'wo': work_order_no,
                'lot': lot_no,
                'line': line,
                'assignments': json.dumps(assignments),
                'dispatched_at': dispatched_at,
            })
            dispatch_count += 1

            # Pre-create baseline placeholder records in assay_process_records
            # Find work_order_id from tutti_work_orders
            wo_row = db.session.execute(
                text("SELECT id FROM panel_production.tutti_work_orders WHERE lot_no = :lot ORDER BY updated_at DESC LIMIT 1"),
                {'lot': lot_no}
            ).fetchone()
            work_order_id = wo_row[0] if wo_row else None

            # Get panel_name from form_data
            panel_name = ''
            if wo_row:
                fd_row = db.session.execute(
                    text("SELECT form_data->'header'->>'formTitle' FROM panel_production.tutti_work_orders WHERE id = :id"),
                    {'id': work_order_id}
                ).fetchone()
                panel_name = fd_row[0] if fd_row and fd_row[0] else ''

            for assignment in assignments:
                # Extract analyze_item (marker) from beadName or reagentName1
                bead_name = assignment.get('beadName', '')
                reagent1 = assignment.get('reagentName1', '')
                reagent2 = assignment.get('reagentName2', '')
                well = assignment.get('well', '')
                # well format: "W1" -> test_well = "1"
                test_well = well.replace('W', '').strip() if well else ''

                markers = []
                if reagent1:
                    markers.append(reagent1)
                if reagent2:
                    markers.append(reagent2)

                for marker in markers:
                    db.session.execute(text("""
                        INSERT INTO panel_production.assay_process_records
                            (work_order_id, work_order_no, lot_no, device_sn, panel_name,
                             analyze_date, analyze_time, sample_type, species, patient_id,
                             lot_code, mfg_lot_no, analyze_item, analyze_result, unit,
                             test_zone, test_well, baseline, baseline_equation, created_at)
                        VALUES
                            (:woid, :wo, :lot, :device_sn, :panel_name,
                             :analyze_date, :analyze_time, 'Control', 'control', :patient_id,
                             :lot, :lot, :marker, '', '',
                             :position, :test_well, 'true', '', :created_at)
                    """), {
                        'woid': work_order_id,
                        'wo': work_order_no,
                        'lot': lot_no,
                        'device_sn': machine_id,
                        'panel_name': panel_name,
                        'analyze_date': dispatched_at.strftime('%Y-%m-%d'),
                        'analyze_time': dispatched_at.strftime('%H:%M:%S'),
                        'patient_id': f'baseline-{dispatched_at.strftime("%Y%m%d%H%M%S")}',
                        'marker': marker,
                        'position': str(pos),
                        'test_well': test_well,
                        'created_at': dispatched_at,
                    })
                    baseline_count += 1

        db.session.commit()
        return jsonify({
            'ok': True,
            'dispatch_count': dispatch_count,
            'baseline_count': baseline_count,
            'dispatched_at': dispatched_at.isoformat(),
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500


# ---------- AWS IoT Webhook: Analyze Status Report ----------

@app.route('/iot/analyze-status/report', methods=['POST'])
def iot_analyze_status_report():
    """Receive AWS IoT notification, fetch data from SkylaiCloud, match baseline placeholders."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'ok': False, 'error': 'empty payload'}), 400

        device_sn = data.get('device_sn', '').strip()
        analyze_date = data.get('analyze_date', '').strip()

        if not device_sn:
            return jsonify({'ok': False, 'error': 'missing device_sn'}), 400
        if not analyze_date:
            analyze_date = datetime.now().strftime('%Y-%m-%d')

        logging.info(f"[IoT] Received analyze report: device={device_sn}, date={analyze_date}")

        # Find latest dispatch for this device
        dispatch_row = db.session.execute(text("""
            SELECT dispatched_at FROM panel_production.panel_dispatch
            WHERE device_sn = :sn ORDER BY dispatched_at DESC LIMIT 1
        """), {'sn': device_sn}).fetchone()

        if not dispatch_row:
            logging.warning(f"[IoT] No dispatch found for device {device_sn}")
            return jsonify({'ok': True, 'matched': 0, 'reason': 'no dispatch found'})

        dispatched_at = dispatch_row[0]

        # Fetch data from SkylaiCloud API
        headers = {'Authorization': f'Bearer {SKYLAI_API_TOKEN}', 'Content-Type': 'application/json'}
        resp = http_requests.post(f'{SKYLAI_API_BASE}/api/get_device_data', headers=headers, json={
            'command': 'get_device_data',
            'project_id': SKYLAI_PROJECT_ID,
            'device_sn': device_sn,
            'start_date': analyze_date,
            'end_date': analyze_date,
            'per_page': 100,
            'current_page': 1,
            'sort': ['-analyze_date'],
        }, timeout=30)

        if resp.status_code != 200:
            logging.error(f"[IoT] SkylaiCloud API error: {resp.status_code} {resp.text[:200]}")
            return jsonify({'ok': False, 'error': f'SkylaiCloud API error: {resp.status_code}'}), 502

        api_data = resp.json().get('data', [])
        matched = 0

        for session in api_data:
            session_date = session.get('analyze_date', '')
            session_time = session.get('analyze_time', '')
            markers = session.get('markers', [])

            for marker_data in markers:
                analyze_item = marker_data.get('analyze_item', '')
                analyze_result = marker_data.get('analyze_result', '')
                unit = marker_data.get('unit', '')
                test_well = marker_data.get('test_well', '')
                final_delta_od = marker_data.get('final_delta_od', '')
                cal_od = marker_data.get('cal_od_sec_rfu', '')
                equation = marker_data.get('equation', '')
                eq_type = marker_data.get('eq_type', '')
                test_zone = marker_data.get('test_zone', '')

                if not analyze_item or not test_well:
                    continue

                # Match baseline placeholder: device_sn + test_well + analyze_item + time >= dispatched_at
                placeholder = db.session.execute(text("""
                    SELECT id FROM panel_production.assay_process_records
                    WHERE device_sn = :sn
                      AND test_well = CAST(:well AS VARCHAR)
                      AND analyze_item = :item
                      AND baseline = 'true'
                      AND (analyze_result IS NULL OR analyze_result = '')
                      AND created_at >= :dispatched_at
                    ORDER BY created_at ASC
                    LIMIT 1
                """), {
                    'sn': device_sn,
                    'well': test_well,
                    'item': analyze_item,
                    'dispatched_at': dispatched_at,
                }).fetchone()

                if placeholder:
                    # Update placeholder with actual test results
                    db.session.execute(text("""
                        UPDATE panel_production.assay_process_records
                        SET analyze_result = :result,
                            unit = :unit,
                            analyze_date = :date,
                            analyze_time = :time,
                            final_delta_od = :final_od,
                            cal_od = :cal_od,
                            equation = :equation,
                            eq_type = :eq_type,
                            test_zone = :test_zone,
                            raw_data = :raw_data
                        WHERE id = :id
                    """), {
                        'result': str(analyze_result),
                        'unit': unit,
                        'date': session_date,
                        'time': session_time,
                        'final_od': str(final_delta_od) if final_delta_od is not None else '',
                        'cal_od': str(cal_od) if cal_od is not None else '',
                        'equation': str(equation) if equation else '',
                        'eq_type': str(eq_type) if eq_type else '',
                        'test_zone': str(test_zone) if test_zone else '',
                        'raw_data': json.dumps(marker_data),
                        'id': placeholder[0],
                    })
                    matched += 1

        db.session.commit()
        logging.info(f"[IoT] Matched {matched} records for device {device_sn}")

        # Sync matched baseline records to SQLite for pre-assignment inspections/build-lines
        if matched > 0:
            try:
                import sqlite3
                SQLITE_DB = '/home/ubuntu/qc-web-ipqc/tutti-qc-assayprocess/data/Tutti_QC_assayprocess.db'
                dispatches = db.session.execute(text("""
                    SELECT lot_no, work_order_no, position FROM panel_production.panel_dispatch
                    WHERE device_sn = :sn AND dispatched_at = :dat
                """), {'sn': device_sn, 'dat': dispatched_at}).fetchall()

                rds_rows = db.session.execute(text("""
                    SELECT analyze_item, analyze_result, unit, analyze_date, analyze_time,
                           test_zone, test_well, final_delta_od, cal_od, equation, eq_type,
                           mfg_lot_no, panel_name, device_sn, patient_id, raw_data
                    FROM panel_production.assay_process_records
                    WHERE device_sn = :sn AND baseline = 'true' AND created_at >= :dat
                      AND analyze_result IS NOT NULL AND analyze_result != ''
                """), {'sn': device_sn, 'dat': dispatched_at}).fetchall()

                conn = sqlite3.connect(SQLITE_DB)
                cur = conn.cursor()
                for row in rds_rows:
                    raw = row[15] if isinstance(row[15], dict) else (json.loads(row[15]) if row[15] else {})
                    cur.execute('''
                        INSERT INTO assay_process_records (
                            source_file, source_file_name, imported_at, row_index,
                            panel_name, analyze_date, analyze_time, sample_type, "Species",
                            patient_id, "Lot code", analyze_item, "Disc_result", analyze_result,
                            unit, "Test Zone", "Test Well", "Final Delta OD", "Cal. OD/Sec/RFU",
                            "Equation", "Eq Type", "Analyzer Serial", "Serial Number",
                            baseline, mfg_lot_no, device_sn, "Cal. Run"
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        'iot-webhook',
                        f'iot-{device_sn}-{row[3]}',
                        datetime.now().isoformat(),
                        0,
                        row[12],
                        row[3],
                        row[4],
                        'Control',
                        'Control',
                        'control-1',
                        row[11],
                        row[0],
                        str(raw.get('disc_result', '')),
                        str(row[1]),
                        row[2],
                        row[5],
                        row[6],
                        str(row[7]) if row[7] else '',
                        str(row[8]) if row[8] else '',
                        row[9] or '',
                        row[10] or '',
                        device_sn,
                        device_sn,
                        'true',
                        row[11],
                        device_sn,
                        str(raw.get('cal_run', '')),
                    ))
                conn.commit()
                conn.close()
                logging.info(f"[IoT] Synced {len(rds_rows)} records to SQLite")
            except Exception as sqlite_err:
                logging.error(f"[IoT] SQLite sync error: {sqlite_err}")

        return jsonify({'ok': True, 'matched': matched, 'device_sn': device_sn})

    except Exception as e:
        db.session.rollback()
        logging.error(f"[IoT] Error: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500

# ---------- Panel Mechanical Materials API ----------
# Direct integration (no proxy needed — imports calc_engine from mechanic_materials)

import sys as _sys
_sys.path.insert(0, '/home/ubuntu/mechanic_materials')

@app.route('/api/panel-materials', methods=['GET'])
def api_panel_materials():
    """Calculate and return panel material requirements."""
    try:
        from calc_engine import calculate_all_materials
        start_date = request.args.get('start_date')
        result = calculate_all_materials(start_date)
        # Sanitize NaN/Inf values that aren't valid JSON
        clean = json.loads(json.dumps(_sanitize_for_json(result)))
        return jsonify({'ok': True, **clean})
    except Exception as e:
        logging.error(f"[Panel Materials] {e}")
        import traceback; traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


def _sanitize_for_json(obj):
    """Recursively replace NaN/Inf with None for JSON serialization."""
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    elif isinstance(obj, dict):
        return {k: _sanitize_for_json(v) for k, v in obj.items()}
    elif isinstance(obj, (list, tuple)):
        return [_sanitize_for_json(i) for i in obj]
    return obj


@app.route('/api/panel-materials/generate-excel', methods=['POST'])
def api_panel_materials_gen_excel():
    """Generate output Excel file."""
    try:
        from calc_engine import calculate_all_materials
        from excel_writer import generate_output_excel
        data = request.get_json() or {}
        start_date = data.get('start_date')
        result = calculate_all_materials(start_date)
        output_path = generate_output_excel(result)
        filename = os.path.basename(output_path)
        return jsonify({'ok': True, 'filename': filename})
    except Exception as e:
        logging.error(f"[Panel Materials Excel] {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/panel-materials/download/<filename>', methods=['GET'])
def api_panel_materials_download(filename):
    """Download generated Excel file."""
    output_dir = '/home/ubuntu/mechanic_materials/Exceldata/output'
    filepath = os.path.join(output_dir, filename)
    if not os.path.exists(filepath):
        return jsonify({'ok': False, 'error': 'File not found'}), 404
    return send_file(filepath, as_attachment=True, download_name=filename)


@app.route('/api/panel-materials/output-files', methods=['GET'])
def api_panel_materials_list_files():
    """List available output files."""
    output_dir = '/home/ubuntu/mechanic_materials/Exceldata/output'
    os.makedirs(output_dir, exist_ok=True)
    files = []
    for f in sorted(glob.glob(os.path.join(output_dir, '*.xlsx')), reverse=True):
        stat = os.stat(f)
        files.append({
            'filename': os.path.basename(f),
            'size': stat.st_size,
            'modified': datetime.fromtimestamp(stat.st_mtime).isoformat(),
        })
    return jsonify({'ok': True, 'files': files})


@app.route('/api/panel-materials/sync', methods=['POST'])
def api_panel_materials_sync():
    """Sync Excel reference data to RDS."""
    try:
        from scripts.sync_excel_to_rds import sync_excel_to_rds
        sync_excel_to_rds()
        return jsonify({'ok': True, 'message': 'Sync complete'})
    except Exception as e:
        logging.error(f"[Panel Materials Sync] {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/panel-materials/upload-stock', methods=['POST'])
def api_panel_materials_upload_stock():
    """Upload a stock EXPORT_YYYYMMDD.xlsx file."""
    import re as _re
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': '未選擇檔案'}), 400
    file = request.files['file']
    filename = file.filename or ''
    # Validate filename format: EXPORT_YYYYMMDD.xlsx
    if not _re.match(r'^EXPORT_\d{8}\.xlsx$', filename):
        return jsonify({'ok': False, 'error': f'檔名格式錯誤，必須為 EXPORT_YYYYMMDD.xlsx（收到: {filename}）'}), 400
    dest = os.path.join('/home/ubuntu/mechanic_materials/Exceldata', filename)
    file.save(dest)
    return jsonify({'ok': True, 'filename': filename, 'message': f'已上傳: {filename}'})
