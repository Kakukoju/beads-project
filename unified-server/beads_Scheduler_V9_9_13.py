import pandas as pd
import sqlite3
import os
import re
import math
import sys      # ✅ 新增
import json     # ✅ 新增
import argparse   # ✅ 新增
from datetime import datetime, timedelta, time
import traceback
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ====================================================================
# V9.9.2 修正版排程系統 (新增休假日功能)
# ====================================================================

# --------------------------------------------------------------------
# 1. 設定
# --------------------------------------------------------------------
MAIN_DB_PATH = r"\\fls341\MBBU_FAB\MB_PD\生管自動化\滴定\資料庫\beads_sync.db"
SCHEDULE_DB_PATH = r"\\fls341\MBBU_FAB\MB_PD\生管自動化\滴定\資料庫\Beads_Schedule.db"
output_dir = r"\\fls341\MBBU_FAB\MB_PD\生管自動化\滴定\beadsSchedule"
CURRENT_VERSION = "v9.9.12 (凌晨不收藥)"

SCHEDULING_DAYS = 5
ALLOW_EXTEND_TO_6_DAYS = True 
MAX_PORTS = 12 
DOSING_PREP_TIME_MIN = 30 
HARVEST_TIME_MIN = 30 
# ========================================
# V9.9.12: 禁止時段設定
# ========================================
FORBIDDEN_END_TIME_START = time(3, 0)   # 禁止結束時間起始：03:00
FORBIDDEN_END_TIME_END = time(8, 0)     # 禁止結束時間結束：08:00

# ===== MODIFIED RULES (2025-11-04) =====
DOSING_RATE_PER_HR = 1500
MAX_PORTS_PER_PN = 4
PORT_OPTIONS = [4, 2]
DOSING_TIME_MIN = 3.0
DOSING_TIME_MAX = 8.0
# ========================================

PREP_DURATION_MIN = 30 
WORK_START_TIME = time(8, 0)
WORK_END_TIME_LATE = time(3, 0)
WORK_END_TIME_ALERT = time(3, 0)
IVEK_COMPONENTS = ["5714400202", "5714400203", "5714400209", "5714400210"]

# ========================================
# V9.9.2: 全域休假日設定
# ========================================
HOLIDAYS = set()  # 格式: datetime.date 物件
BATCH_START_NUMBER = 1  # 工單批次起始編號
VACATION_STAFF = {}  # {date_str: set(["人員1", "人員2", ...])}
# ✅ 新增：儲存排程起始日期
SCHEDULE_START_DATE = None

# --------------------------------------------------------------------
# 2. 輔助工具
# --------------------------------------------------------------------
def time_str_to_float(time_str: str) -> float:
    try:
        h, m = map(int, time_str.split(':'))
        return h + m / 60.0
    except Exception: 
        return 0.0

def float_to_time_str(hours: float) -> str:
    total_minutes = int(hours * 60)
    h = (total_minutes // 60) % 24 
    m = total_minutes % 60
    return f"{h:02d}:{m:02d}"

def add_hours_to_time(start_datetime: datetime, hours: float) -> datetime:
    return start_datetime + timedelta(minutes=int(hours * 60))

def parse_date_columns(df):
    """解析 Production Plan 的日期欄位"""
    date_cols = {}
    non_date_cols = [
        "PN", "料號", "品名", "Customer", "Description", "Plan", 
        "JAN", "FEB", "MAR", "APR", "MAY", "JUN", 
        "JUL", "AUG", "SEP", "OCT", "NOV", "DEC", "Unnamed: 0"
    ]
    non_date_cols.extend([f"WK{i:02d}" for i in range(1, 54)])
    
    date_regex = re.compile(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}')
        
    for col in df.columns:
        if col in non_date_cols:
            continue
        
        match = date_regex.match(col)
        if match:
            try:
                col_date = pd.to_datetime(match.group(0)) 
                date_cols[col] = col_date
                continue
            except (ValueError, TypeError):
                pass 

        try:
            col_clean = re.sub(r'\(.*\)', '', col).strip() 
            col_date = pd.to_datetime(col_clean)
            date_cols[col] = col_date
        except (ValueError, TypeError):
            pass
            
    return date_cols

def standardize_pn(pn_series: pd.Series) -> pd.Series:
    """標準化 PN"""
    return pn_series.astype(str).str.split('.').str[0]

# ========================================
# V9.9.2: 休假日輸入功能
# ========================================
def get_holidays_from_user():
    """
    讓使用者輸入休假日
    返回: set of datetime.date 物件
    """
    global HOLIDAYS
    
    print("\n" + "="*70, file=sys.stderr)
    print("🗓️  休假日設定", file=sys.stderr)
    print("="*70, file=sys.stderr)
    print("預計排程週是否有休假日？", file=sys.stderr)
    print("  [Y] 是，有休假日", file=sys.stderr)
    print("  [N] 否，無休假日（預設）", file=sys.stderr)
    
    has_holiday = input("\n請選擇 (Y/N，直接按 Enter 預設為 N): ").strip().upper()
    
    if has_holiday != 'Y':
        print("✅ 無休假日，繼續排程\n", file=sys.stderr)
        return set()
    
    print("\n請輸入休假日期:", file=sys.stderr)
    print("  格式: MM-DD 或 MM/DD", file=sys.stderr)
    print("  多日請用逗號分隔，例如: 11-11,11-12 或 11/11,11/12", file=sys.stderr)
    print("  年份將自動使用當前年份", file=sys.stderr)
    
    holiday_input = input("\n休假日期: ").strip()
    
    if not holiday_input:
        print("✅ 無休假日輸入，繼續排程\n", file=sys.stderr)
        return set()
    
    holidays = set()
    current_year = datetime.now().year
    
    # 解析輸入
    date_strings = [d.strip() for d in holiday_input.split(',') if d.strip()]
    
    for date_str in date_strings:
        try:
            # 替換 - 為 /
            date_str = date_str.replace('-', '/')
            
            # 嘗試解析 MM/DD
            parts = date_str.split('/')
            if len(parts) == 2:
                month = int(parts[0])
                day = int(parts[1])
                holiday_date = datetime(current_year, month, day).date()
                holidays.add(holiday_date)
                print(f"  ✅ 已加入休假日: {holiday_date.strftime('%Y-%m-%d')}", file=sys.stderr)
            else:
                print(f"  ⚠️ 格式錯誤，跳過: {date_str}", file=sys.stderr)
                
        except Exception as e:
            print(f"  ⚠️ 無法解析日期 '{date_str}': {e}", file=sys.stderr)
    
    if holidays:
        print(f"\n✅ 共設定 {len(holidays)} 個休假日:", file=sys.stderr)
        for h in sorted(holidays):
            print(f"      {h.strftime('%Y-%m-%d (%A)')}", file=sys.stderr)
    else:
        print("\n⚠️ 未成功設定任何休假日", file=sys.stderr)
    
    print("="*70 + "\n", file=sys.stderr)
    return holidays

def is_holiday(date_obj):
    """檢查日期是否為休假日"""
    if isinstance(date_obj, datetime):
        date_obj = date_obj.date()
    return date_obj in HOLIDAYS
# ========================================
# V9.9.9: 命令行參數解析
# ========================================
def parse_arguments():
    """解析命令行參數"""
    parser = argparse.ArgumentParser(description='Beads 排程系統（前端整合版）')
    
    # 必要參數
    parser.add_argument('--date', required=True, 
                        help='排程起始日期（週一），格式：MM/DD，例如：11/10')
    parser.add_argument('--need', required=True,
                        help='需求檔路徑（Excel 檔案）')
    
    # 可選參數
    parser.add_argument('--outdir', 
                        default=r"\\fls341\MBBU_FAB\MB_PD\生管自動化\滴定\beadsSchedule",
                        help='輸出資料夾路徑')
    parser.add_argument('--holidays', default='',
                        help='休假日，逗號分隔，格式：MM/DD，例如：11/10,11/11')
    parser.add_argument('--batch-numbers', default='',
                        help='生產工單批次編號起始值，例如：111')
    parser.add_argument('--vacation-staff', default='',
                        help='休假人員，格式：MM/DD-人名，例如：11/10-張三,11/11-李四')
    parser.add_argument('--dry-run', action='store_true',
                        help='Dry Run 模式（僅預覽，不輸出檔案）')
    
    return parser.parse_args()

def setup_holidays_from_args(holidays_str, start_date):
    """從參數設定休假日"""
    global HOLIDAYS
    
    if not holidays_str:
        return set()
    
    holidays = set()
    current_year = start_date.year
    
    date_strings = [d.strip() for d in holidays_str.split(',') if d.strip()]
    
    for date_str in date_strings:
        try:
            date_str = date_str.replace('-', '/')
            parts = date_str.split('/')
            
            if len(parts) == 2:
                month = int(parts[0])
                day = int(parts[1])
                holiday_date = datetime(current_year, month, day).date()
                holidays.add(holiday_date)
                print(f"  ✅ 已加入休假日: {holiday_date.strftime('%Y-%m-%d')}", file=sys.stderr)
            else:
                print(f"  ⚠️ 格式錯誤，跳過: {date_str}", file=sys.stderr)
                
        except Exception as e:
            print(f"  ⚠️ 無法解析日期 '{date_str}': {e}", file=sys.stderr)
    
    HOLIDAYS = holidays
    return holidays

def setup_batch_start_from_args(batch_str):
    """從參數設定批次編號"""
    global BATCH_START_NUMBER
    
    if not batch_str:
        return 1
    
    try:
        batch_start = int(batch_str)
        if 1 <= batch_start <= 999:
            BATCH_START_NUMBER = batch_start
            print(f"  ✅ 工單批次起始編號: {batch_start}", file=sys.stderr)
            return batch_start
    except:
        pass
    
    print(f"  ⚠️ 批次編號無效，使用預設值 1", file=sys.stderr)
    return 1

def setup_vacation_staff_from_args(vacation_str, start_date):
    """從參數設定休假人員"""
    global VACATION_STAFF
    
    if not vacation_str:
        VACATION_STAFF = {}
        return {}
    
    vacation_staff = {}
    current_year = start_date.year
    
    entries = [e.strip() for e in vacation_str.split(',') if e.strip()]
    
    for entry in entries:
        try:
            if '-' not in entry:
                print(f"  ⚠️ 休假人員格式錯誤，跳過: {entry}", file=sys.stderr)
                print(f"      正確格式：MM/DD-人名，例如：11/10-張三", file=sys.stderr)
                continue
            
            date_part, person = entry.split('-', 1)
            date_part = date_part.strip().replace('/', '/')
            person = person.strip().lower() # ✅ 修正：儲存時轉為小寫
            
            parts = date_part.split('/')
            if len(parts) == 2:
                month = int(parts[0])
                day = int(parts[1])
                vacation_date = datetime(current_year, month, day).date()
                date_str = vacation_date.strftime('%Y-%m-%d')
                
                if date_str not in vacation_staff:
                    vacation_staff[date_str] = set()
                vacation_staff[date_str].add(person)
                
                print(f"  ✅ 已記錄休假: {date_str} - {person}", file=sys.stderr)
            else:
                print(f"  ⚠️ 日期格式錯誤，跳過: {date_part}", file=sys.stderr)
                
        except Exception as e:
            print(f"  ⚠️ 無法解析休假人員資訊 '{entry}': {e}", file=sys.stderr)
    
    VACATION_STAFF = vacation_staff
    
    if vacation_staff:
        print(f"\n  📊 休假人員統計:", file=sys.stderr)
        for date_str in sorted(vacation_staff.keys()):
            people = vacation_staff[date_str]
            print(f"      {date_str}: {', '.join(sorted(people))} (共 {len(people)} 人)", file=sys.stderr)
    
    return vacation_staff

def is_person_on_vacation(person, date_obj):
    """檢查人員在指定日期是否請假"""
    if isinstance(date_obj, datetime):
        date_obj = date_obj.date()
    
    date_str = date_obj.strftime('%Y-%m-%d')
    
    if date_str in VACATION_STAFF:
        return person.lower() in VACATION_STAFF[date_str] # ✅ 修正：比對時轉為小寫
    
    return False

# --------------------------------------------------------------------
# 3. 資料載入, 加入藥名
# --------------------------------------------------------------------
# --------------------------------------------------------------------
# 3. 資料載入
# --------------------------------------------------------------------
def load_all_data(demand_xlsx_path):
    """載入所有資料"""
    print(f"Step 1: 正在載入資料 ({CURRENT_VERSION})...", file=sys.stderr)
    db_tables = {} 
    df_demand = pd.DataFrame()
    df_constraints = pd.DataFrame()

    # 載入 Dosing_Constraints
    try:
        with sqlite3.connect(MAIN_DB_PATH) as conn:
            df_constraints_raw = pd.read_sql("SELECT * FROM '配藥限制'", conn)
            
        df_constraints = df_constraints_raw.copy()
        
        def norm_col(s: str) -> str:
            return (str(s).strip().replace(" ", "").replace("_blank", "").lower())

        rename_map = {}
        for c in df_constraints.columns:
            nz = norm_col(c)
            if nz in ("可用凍乾機", "可用凍乾機台", "dryers", "dryer"):
                rename_map[c] = "可用凍乾機"
            elif nz == "pn":
                rename_map[c] = "PN"
            elif nz in ("配藥人-1","人員1","staff1"):
                rename_map[c] = "配藥人-1"
            elif nz in ("配藥人-2","人員2","staff2"):
                rename_map[c] = "配藥人-2"
            elif nz in ("配藥人-3","人員3","staff3"):
                rename_map[c] = "配藥人-3"
            elif nz in ("交藥時間","給藥時間","rd給藥時間"):
                rename_map[c] = "交藥時間"
            elif nz in ("凍乾時間", "freezehrs", "freezehour", "freezehr"):
                rename_map[c] = "凍乾時間"
            elif nz in ("滴定後凍乾時間差距(hr)", "滴定後凍乾時間差距", "waithr", "waithrs"):
                rename_map[c] = "滴定後凍乾時間差距 (HR)"

        df_constraints = df_constraints.rename(columns=rename_map)
        df_constraints = df_constraints[df_constraints['PN'].notna() & (df_constraints['PN'] != "")]
        df_constraints["PN"] = standardize_pn(df_constraints["PN"])
        
        num_cols = ["凍乾時間", "滴定後凍乾時間差距 (HR)"]
        for col in num_cols:
            if col in df_constraints.columns:
                df_constraints[col] = pd.to_numeric(df_constraints[col], errors='coerce').fillna(0)
            else:
                df_constraints[col] = 0 
        
        def count_csv(s):
            if not isinstance(s, str) or s == "": return 99
            return len([d for d in s.split(',') if d.strip()])
        
        df_constraints["V9_Dryer_Count"] = df_constraints["可用凍乾機"].apply(count_csv)

        def count_people(row):
            count = 0
            if isinstance(row.get("配藥人-1"), str) and row.get("配藥人-1", "").strip(): count += 1
            if isinstance(row.get("配藥人-2"), str) and row.get("配藥人-2", "").strip(): count += 1
            if isinstance(row.get("配藥人-3"), str) and row.get("配藥人-3", "").strip(): count += 1
            return count if count > 0 else 99
            
        df_constraints["V9_Person_Count"] = df_constraints.apply(count_people, axis=1)
        
        def get_slots(s):
            if not isinstance(s, str) or s == "": return ["AM", "PM"] 
            times = [t.strip() for t in s.split(',') if t.strip()]
            slots = set()
            
            for t in times:
                try:
                    hour = int(t.split(':')[0])
                    if 9 <= hour <= 14:
                        slots.add("AM")
                    elif 15 <= hour <= 23:
                        slots.add("PM")
                except:
                    pass

            return list(slots) if slots else ["AM", "PM"]

        df_constraints["V9_Slot_List"] = df_constraints["交藥時間"].apply(get_slots)
        df_constraints["V9_Single_Slot"] = df_constraints["交藥時間"].apply(
            lambda s: isinstance(s, str) and len(s.split(',')) == 1 and ":" in s
        )

        print("  ✅ '配藥限制' 載入完成", file=sys.stderr)

    except Exception as e:
        print(f"❌ 無法載入 '配藥限制': {e}", file=sys.stderr)
        return None 

    # 載入 BDC
    df_bdc = pd.DataFrame()
    try:
        with sqlite3.connect(MAIN_DB_PATH) as conn:
            df_bdc_raw = pd.read_sql("SELECT * FROM Beads_Dry_Count", conn)
            
            if len(df_bdc_raw.columns) < 4:
                raise ValueError(f"Beads_Dry_Count 欄位不足")
            
            col_map = {
                df_bdc_raw.columns[0]: "BDC_Group_Name", 
                df_bdc_raw.columns[1]: "料號",          
                df_bdc_raw.columns[2]: "BDC_Marker_Name",
                df_bdc_raw.columns[3]: "BDC_Prod_Qty"   
            }
            df_bdc = df_bdc_raw.rename(columns=col_map)
            df_bdc["料號"] = standardize_pn(df_bdc["料號"])
            df_bdc["BDC_Prod_Qty"] = pd.to_numeric(df_bdc["BDC_Prod_Qty"], errors='coerce').fillna(0)
            select_cols = ["BDC_Group_Name", "料號", "BDC_Marker_Name", "BDC_Prod_Qty"]
            df_bdc = df_bdc[select_cols]
            # ✅ 新增：標準化 BDC 的藥名 (去除空白)
            df_bdc["BDC_Group_Name"] = df_bdc["BDC_Group_Name"].astype(str).str.strip()
            print(f"  ✅ BDC 載入完成 ({len(df_bdc)} 筆)", file=sys.stderr)
            
    except Exception as e:
        print(f"❌ 無法載入 BDC: {e}", file=sys.stderr)
        return None 

    # 載入需求表
    try:
        print(f"  ⏳ 正在載入需求檔: {demand_xlsx_path}", file=sys.stderr)
        # ✅ 修正：使用 header=1 (因為標題在 Excel 的第二行)
        df_demand = pd.read_excel(demand_xlsx_path, sheet_name="滴定排程需求表", header=1, engine='openpyxl')
        
        def norm(s: str) -> str:
            if s is None: return ""
            return str(s).strip().replace(" ", "").replace("　", "").replace("＋", "").replace("週", "周")

        def find_col(*keywords):
            for raw in df_demand.columns:
                nz = norm(raw)
                if all(k in nz for k in keywords):
                    return raw
            return None

        col_stock = find_col("庫存", "滴定") or find_col("庫存")
        col_w1 = find_col("第一","周","需求") or "第一周需求"
        col_w2 = find_col("第二","周","需求") or "第二周需求"
        col_w3 = find_col("第三","周","需求") or "第三周需求"
        col_pn = find_col("料號") or "料號"
        col_group = find_col("藥名") or find_col("Group") or "BDC_Group_Name" # ✅ 新增

        df_demand[col_pn] = standardize_pn(df_demand[col_pn])
        for c in [col_stock, col_w1, col_w2, col_w3]:
            df_demand[c] = pd.to_numeric(df_demand[c], errors="coerce").fillna(0)

        df_demand = df_demand.rename(columns={
            col_pn: "料號", 
            col_group: "BDC_Group_Name", # ✅ 新增
            col_stock: "Stock_plus_Dosing",
            col_w1: "第一周需求", col_w2: "第二周需求", col_w3: "第三周需求"
        })
        # ✅ 修改：新增 BDC_Group_Name 到欄位列表中
        df_demand = df_demand[["料號", "BDC_Group_Name", "Stock_plus_Dosing", "第一周需求", "第二周需求", "第三周需求"]]
        # ✅ 新增：標準化藥名 (去除空白)
        df_demand["BDC_Group_Name"] = df_demand["BDC_Group_Name"].astype(str).str.strip()
        
        print(f"  ✅ 需求表載入完成 ({len(df_demand)} 筆)", file=sys.stderr)

    except Exception as e:
        print(f"❌ 無法載入需求表: {e}", file=sys.stderr)
        return None

    # 合併
    if df_demand is None or df_bdc is None or df_constraints is None:
        return None
        
    # ✅ 修改：使用 ["料號", "BDC_Group_Name"] 作為合併的 KEY
    df_merged = pd.merge(df_demand, df_bdc, on=["料號", "BDC_Group_Name"], how="left")
    
    df_constraints_v9 = df_constraints[[
        "PN", "V9_Dryer_Count", "V9_Person_Count", "V9_Slot_List", "V9_Single_Slot",
        "凍乾時間", "滴定後凍乾時間差距 (HR)", "可用凍乾機", "配藥人-1", "配藥人-2", "配藥人-3",
        "交藥時間",
        "備註",
        "Port數"  # ✅ 新增：讀取 Port 數欄位
        ]].rename(columns={"PN": "料號"})
    
    df_merged = pd.merge(df_merged, df_constraints_v9, on="料號", how="left")
    df_merged = df_merged.dropna(subset=["BDC_Group_Name"])
    
    print(f"  ✅ 合併完成 ({len(df_merged)} 筆有效需求)\n", file=sys.stderr)

    # 載入其他資料庫
    try:
        with sqlite3.connect(MAIN_DB_PATH) as conn:
            df_plan = pd.read_sql("SELECT * FROM production_Plan", conn)
            date_cols_map = parse_date_columns(df_plan)
            if date_cols_map:
                for col in date_cols_map.keys():
                    df_plan[col] = pd.to_numeric(df_plan[col], errors='coerce').fillna(0)

            db_tables["production_Plan"] = df_plan
            db_tables["BOM_Details"] = pd.read_sql("SELECT * FROM BOM_Details", conn)

        db_tables["BOM_Details"]["Finished_PartNo"] = standardize_pn(db_tables["BOM_Details"]["Finished_PartNo"])
        db_tables["BOM_Details"]["Component_No"] = standardize_pn(db_tables["BOM_Details"]["Component_No"])
        db_tables["production_Plan"]["PN"] = standardize_pn(db_tables["production_Plan"]["PN"])
        
        # ✅ 修改：不再使用 set_index
        db_tables["Beads_Dry_Count_Info"] = df_bdc.copy()
        db_tables["Dosing_Constraints"] = df_constraints.set_index("PN")

        # === V9.9: 載入 P0 (限制 OR 插單) ===
        try:
            df_p0_orders = pd.read_sql("SELECT * FROM '限制_OR_插單'", conn)
            print(f"  ... P0 原始欄位: {df_p0_orders.columns.to_list()}", file=sys.stderr)
            
            def clean_col_name(col):
                return str(col).strip().replace(':', '')
            
            df_p0_orders.columns = [clean_col_name(c) for c in df_p0_orders.columns]
            print(f"  ... P0 清理後欄位: {df_p0_orders.columns.to_list()}", file=sys.stderr)
            
            if "日期" in df_p0_orders.columns:
                df_p0_orders["日期"] = df_p0_orders["日期"].astype(str)
            
            db_tables["P0_Orders"] = df_p0_orders
            print(f"  ✅ '限制 OR 插單' 載入完成 ({len(df_p0_orders)} 筆)", file=sys.stderr)
        except Exception as e:
            print(f"  ⚠️ 警告: 無法載入 '限制 OR 插單' 表: {e}", file=sys.stderr)
            db_tables["P0_Orders"] = pd.DataFrame()

        print("  ✅ 資料庫載入完成\n", file=sys.stderr)

    except Exception as e:
        print(f"⚠️ 資料庫載入警告: {e}", file=sys.stderr)

    return {
        "demand": df_merged, 
        "db": db_tables,
        "output_template": pd.DataFrame() 
    }

# --------------------------------------------------------------------
# 4. 需求計算
# --------------------------------------------------------------------
def calculate_demand_queue(data: dict):
    print("Step 2: 計算 P1-P3 (缺料) 需求優先級...", file=sys.stderr)

    df = data["demand"] 
    db = data["db"]

    df_prod_plan = db.get("production_Plan", pd.DataFrame())
    bom = db.get("BOM_Details", pd.DataFrame())
    
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    monday = (today + timedelta(days=-today.weekday(), weeks=0)) 
    next_wednesday = monday + timedelta(days=2) 
    
    w1_urgent_wed_pn = set()
    w1_urgent_other_pn = set()
    
    if not df_prod_plan.empty and not bom.empty:
        date_cols_map = parse_date_columns(df_prod_plan)
        
        if date_cols_map and "PN" in df_prod_plan.columns:
            urgent_cols = [col for col, dt in date_cols_map.items() if dt <= next_wednesday]
            other_cols = [col for col, dt in date_cols_map.items() if dt > next_wednesday]
            
            df_plan_copy = df_prod_plan.set_index("PN")
            
            if urgent_cols:
                urgent_qty = df_plan_copy[urgent_cols].sum(axis=1)
                urgent_pns = set(urgent_qty[urgent_qty > 0].index)
                w1_urgent_wed_pn = set(bom[bom["Finished_PartNo"].isin(urgent_pns)]["Component_No"])
            
            if other_cols:
                other_qty = df_plan_copy[other_cols].sum(axis=1)
                other_pns = set(other_qty[other_qty > 0].index)
                w1_urgent_other_pn = set(bom[bom["Finished_PartNo"].isin(other_pns)]["Component_No"])

            print(f"  半品: {len(w1_urgent_wed_pn)} (週三前), {len(w1_urgent_other_pn)} (其他)", file=sys.stderr)

    # ==========================================================
    # ✅ 修正：按 BDC_Group_Name (藥名) 分組計算總需求 (1:1 邏輯)
    # ==========================================================
    
    agg_cols = {
        'Stock_plus_Dosing': 'min', 
        '第一周需求': 'max',       
        '第二周需求': 'max',       
        '第三周需求': 'max',       
        'BDC_Prod_Qty': 'first', 
        'BDC_Marker_Name': 'first',
        'V9_Dryer_Count': 'first',
        'V9_Person_Count': 'first'
    }
    
    def agg_pn_list(pn_series):
        return list(pn_series.unique())
    agg_cols['料號'] = agg_pn_list

    df_clean = df[df['BDC_Group_Name'].notna() & (df['BDC_Group_Name'] != "")]
    
    # ==========================================================
    # ✅ 移除 tCREA 特例，所有藥品使用統一的 1:1 邏輯
    # ==========================================================
    df_grouped = df_clean.groupby('BDC_Group_Name').agg(agg_cols).reset_index()

    # ==========================================================
    # 4. 遍歷分組後的 df_grouped (P1, P2, P3)
    # ==========================================================
    task_queue = [] # P1-P5 任務
    p6_p7_tasks = [] # P6/P7 任務 (填補空缺用)
    
    for _, r in df_grouped.iterrows():
        group_name = r["BDC_Group_Name"]
        prod_qty = r["BDC_Prod_Qty"] 
        marker_name = r["BDC_Marker_Name"] 
        dryer_prio = r.get("V9_Dryer_Count", 99)
        person_prio = r.get("V9_Person_Count", 99)
        pns_in_group = r.get("料號", [])

        if pd.isna(prod_qty) or prod_qty == 0: 
            print(f"  ⚠️ 警告: 群組 {group_name} 的 BDC_Prod_Qty 為 0 或 NaN (可能未在 BDC 資料庫中定義)，跳過此群組。", file=sys.stderr)
            continue

        w1_balance = r["Stock_plus_Dosing"] - r["第一周需求"]
        w1_short = abs(w1_balance) if w1_balance < 0 else 0
        w1_prod_count = math.ceil(w1_short / prod_qty) if w1_short > 0 else 0
        w1_planned = w1_prod_count * prod_qty

        w2_balance = w1_balance + w1_planned - r["第二周需求"]
        w2_short = abs(w2_balance) if w2_balance < 0 else 0
        w2_prod_count = math.ceil(w2_short / prod_qty) if w2_short > 0 else 0
        w2_planned = w2_prod_count * prod_qty

        w3_balance = w2_balance + w2_planned - r["第三周需求"]
        w3_short = abs(w3_balance) if w3_balance < 0 else 0
        
        has_urgent_wed = any(pn in w1_urgent_wed_pn for pn in pns_in_group)
        has_urgent_other = any(pn in w1_urgent_other_pn for pn in pns_in_group)
        representative_pn = pns_in_group[0] if pns_in_group else "UNKNOWN"
        
        if w1_short > 0:
            if has_urgent_wed: plan_prio = 0
            elif has_urgent_other: plan_prio = 1
            else: plan_prio = 2
            subp = (dryer_prio, person_prio, plan_prio)
            task_queue.append((1, subp, representative_pn, w1_short, "W1", group_name, marker_name, prod_qty))
        
        if w2_short > 0:
            subp = (dryer_prio, person_prio, 0)
            task_queue.append((2, subp, representative_pn, w2_short, "W2", group_name, marker_name, prod_qty))
        
        if w3_short > 0:
            subp = (dryer_prio, person_prio, 0)
            task_queue.append((3, subp, representative_pn, w3_short, "W3", group_name, marker_name, prod_qty))

    # ==========================================================
    # ✅ P3 特殊庫存規則
    # ==========================================================
    print("Step 2a: 檢查 P3 特殊庫存規則...", file=sys.stderr)
    existing_task_groups = set(t[5] for t in task_queue)
    df_all_demand_data = data["demand"].set_index('料號')
    
    def create_p3_stock_task(pn, group_name, marker_name, prod_qty, dryer_prio, person_prio):
        if pd.isna(prod_qty) or prod_qty == 0:
            print(f"  ⚠️ 警告: P3 庫存規則 {group_name} 凍乾數為 0 或 NaN，無法建立任務", file=sys.stderr)
            return None
        print(f"  ✅ 觸發 P3 庫存規則: {group_name} (PN: {pn})", file=sys.stderr)
        subp = (dryer_prio, person_prio, 0) # 0 = stock priority
        return (3, subp, pn, 0, "P3_Stock", group_name, marker_name, prod_qty)

    # --- 規則 1: CREA (5714400197, 5714400198) < 4000 ---
    rule1_pns = ['5714400197', '5714400198']
    rule1_threshold = 4000
    try:
        rule1_stocks = df_all_demand_data.loc[rule1_pns]['Stock_plus_Dosing']
        min_stock_r1 = rule1_stocks.min()
        if min_stock_r1 < rule1_threshold:
            pn_info_r1 = df_all_demand_data.loc['5714400197']
            rule1_group_name = pn_info_r1['BDC_Group_Name']
            if rule1_group_name not in existing_task_groups:
                task = create_p3_stock_task('5714400197', rule1_group_name, pn_info_r1['BDC_Marker_Name'], pn_info_r1['BDC_Prod_Qty'], pn_info_r1.get('V9_Dryer_Count', 99), pn_info_r1.get('V9_Person_Count', 99))
                if task: 
                    task_queue.append(task)
                    existing_task_groups.add(rule1_group_name)
    except KeyError:
        print(f"  ⚠️ 警告: P3 庫存規則 PN {rule1_pns} 不在需求檔中，跳過規則 1", file=sys.stderr)
    except Exception as e:
        print(f"  ⚠️ 警告: P3 庫存規則 1 失敗: {e}", file=sys.stderr)

    # --- 規則 2: IVEK-CREA (5714400202, 5714400203) < 30000 ---
    rule2_pns = ['5714400202', '5714400203']
    rule2_threshold = 30000
    try:
        rule2_stocks = df_all_demand_data.loc[rule2_pns]['Stock_plus_Dosing']
        min_stock_r2 = rule2_stocks.min()
        if min_stock_r2 < rule2_threshold:
            pn_info_r2 = df_all_demand_data.loc['5714400202']
            rule2_group_name = pn_info_r2['BDC_Group_Name']
            if rule2_group_name not in existing_task_groups:
                task = create_p3_stock_task('5714400202', rule2_group_name, pn_info_r2['BDC_Marker_Name'], pn_info_r2['BDC_Prod_Qty'], pn_info_r2.get('V9_Dryer_Count', 99), pn_info_r2.get('V9_Person_Count', 99))
                if task: 
                    task_queue.append(task)
                    existing_task_groups.add(rule2_group_name)
    except KeyError:
        print(f"  ⚠️ 警告: P3 庫存規則 PN {rule2_pns} 不在需求檔中，跳過規則 2", file=sys.stderr)
    except Exception as e:
        print(f"  ⚠️ 警告: P3 庫存規則 2 失敗: {e}", file=sys.stderr)

    # ==========================================================
    # === P4/P5 補庫存任務 ===
    # ==========================================================
    print("Step 2b: 計算 P4/P5 (補庫存) 任務...", file=sys.stderr)
    constraints_df = data["db"].get("Dosing_Constraints", pd.DataFrame())
    IDLE_COL_NAME = "空閒產能時可優先安排生產"
    idle_pns = set()
    
    if IDLE_COL_NAME in constraints_df.columns:
        idle_filter = constraints_df[IDLE_COL_NAME].notna() & (constraints_df[IDLE_COL_NAME] != "")
        idle_pns = set(constraints_df[idle_filter].index.unique())
        print(f"  ✅ 找到 {len(idle_pns)} 個 P4/P5 補庫存標記", file=sys.stderr)
    else:
        print(f"⚠️ 找不到 '{IDLE_COL_NAME}' 欄位，跳過 P4/P5 任務", file=sys.stderr)

    added_groups_p4 = set()
    added_groups_p5 = set()
    df_all_demand_data_no_index = data["demand"] 

    if idle_pns:
        shortage_pns_in_groups = existing_task_groups 
        df_idle_tasks_data = df_all_demand_data_no_index[
            df_all_demand_data_no_index['料號'].isin(idle_pns) &
            ~df_all_demand_data_no_index['BDC_Group_Name'].isin(shortage_pns_in_groups)
        ]
        df_idle_tasks_data = df_idle_tasks_data.sort_values(by='Stock_plus_Dosing', ascending=True)
        
        new_idle_tasks_p4 = 0
        for _, r in df_idle_tasks_data.iterrows():
            group_name = r["BDC_Group_Name"]
            if group_name in added_groups_p4: continue
            pn = str(r["料號"]).strip()
            prod_qty = r["BDC_Prod_Qty"]
            if pd.isna(prod_qty) or prod_qty == 0: continue 
            marker_name = r["BDC_Marker_Name"]
            stock_level = r["Stock_plus_Dosing"]
            dryer_prio = r.get("V9_Dryer_Count", 99)
            person_prio = r.get("V9_Person_Count", 99)
            priority = 4; tag = "Idle"
            subp = (dryer_prio, person_prio, stock_level)
            task_queue.append((priority, subp, pn, 0, tag, group_name, marker_name, prod_qty))
            added_groups_p4.add(group_name)
            new_idle_tasks_p4 += 1
        print(f"  ✅ 新增 {new_idle_tasks_p4} 個 P4 任務", file=sys.stderr)

        new_idle_tasks_p5 = 0
        for _, r in df_idle_tasks_data.iterrows():
            group_name = r["BDC_Group_Name"]
            if group_name in added_groups_p5: continue
            pn = str(r["料號"]).strip()
            prod_qty = r["BDC_Prod_Qty"]
            if pd.isna(prod_qty) or prod_qty == 0: continue 
            marker_name = r["BDC_Marker_Name"]
            stock_level = r["Stock_plus_Dosing"]
            dryer_prio = r.get("V9_Dryer_Count", 99)
            person_prio = r.get("V9_Person_Count", 99)
            priority = 5; tag = "Idle_R2"
            subp = (dryer_prio, person_prio, stock_level)
            task_queue.append((priority, subp, pn, 0, tag, group_name, marker_name, prod_qty))
            added_groups_p5.add(group_name)
            new_idle_tasks_p5 += 1
        print(f"  ✅ 新增 {new_idle_tasks_p5} 個 P5 任務", file=sys.stderr)

    # ==========================================================
    # ✅ 修改：P6/P7 庫存填補任務 (分離儲存, 供 fill_gaps 使用)
    # ==========================================================
    print("Step 2c: 計算 P6/P7 (自動補庫存) 任務...", file=sys.stderr)
    
    # 1. 取得所有已在 P1-P5 清單中的群組
    all_p1_p5_groups = existing_task_groups.union(added_groups_p4, added_groups_p5)
    
    p6_p7_tasks = [] # ✅ 這是 P6/P7 填補任務的清單
    
    p6_rule_cols = ['Port數']
    if all(col in df_all_demand_data_no_index.columns for col in p6_rule_cols):
        
        # 3. 篩選 P6/P7 候選 PNs
        df_p6_p7_candidates = df_all_demand_data_no_index[
            (~df_all_demand_data_no_index['BDC_Group_Name'].isin(all_p1_p5_groups))
        ].copy()
        
        # 規則 3.1: Port數 必須為 1 或 2
        df_p6_p7_candidates = df_p6_p7_candidates[
            df_p6_p7_candidates['Port數'].astype(str).isin(['1', '2'])
        ]
            
        # 規則 3.2: 庫存 < 100,000 (P6)
        df_p6_tasks = df_p6_p7_candidates[
            df_p6_p7_candidates['Stock_plus_Dosing'] < 100000
        ]
        
        # 規則 3.3: 庫存 >= 100,000 (P7)
        df_p7_tasks = df_p6_p7_candidates[
            df_p6_p7_candidates['Stock_plus_Dosing'] >= 100000
        ]

        # 4. 排序 - 庫存最低的優先
        df_p6_tasks = df_p6_tasks.sort_values(by='Stock_plus_Dosing', ascending=True)
        df_p7_tasks = df_p7_tasks.sort_values(by='Stock_plus_Dosing', ascending=True)

        # 5. 加入 P6 任務
        added_groups_p6 = set()
        new_idle_tasks_p6 = 0
        for _, r in df_p6_tasks.iterrows():
            group_name = r["BDC_Group_Name"]
            if group_name in added_groups_p6: continue
            
            pn = str(r["料號"]).strip()
            prod_qty = r["BDC_Prod_Qty"]
            if pd.isna(prod_qty) or prod_qty == 0: continue 
            marker_name = r["BDC_Marker_Name"]
            stock_level = r["Stock_plus_Dosing"]
            dryer_prio = r.get("V9_Dryer_Count", 99)
            person_prio = r.get("V9_Person_Count", 99)
            priority = 6; tag = "P6_Stock_L"
            subp = (dryer_prio, person_prio, stock_level)
            p6_p7_tasks.append((priority, subp, pn, 0, tag, group_name, marker_name, prod_qty))
            added_groups_p6.add(group_name)
            new_idle_tasks_p6 += 1
        print(f"  ✅ 新增 {new_idle_tasks_p6} 個 P6 任務 (<100k)", file=sys.stderr)

        # 6. 加入 P7 任務
        added_groups_p7 = set()
        new_idle_tasks_p7 = 0
        for _, r in df_p7_tasks.iterrows():
            group_name = r["BDC_Group_Name"]
            if group_name in added_groups_p7: continue
                
            pn = str(r["料號"]).strip()
            prod_qty = r["BDC_Prod_Qty"]
            if pd.isna(prod_qty) or prod_qty == 0: continue 
            marker_name = r["BDC_Marker_Name"]
            stock_level = r["Stock_plus_Dosing"]
            dryer_prio = r.get("V9_Dryer_Count", 99)
            person_prio = r.get("V9_Person_Count", 99)
            priority = 7; tag = "P7_Stock_H"
            subp = (dryer_prio, person_prio, stock_level)
            p6_p7_tasks.append((priority, subp, pn, 0, tag, group_name, marker_name, prod_qty))
            added_groups_p7.add(group_name)
            new_idle_tasks_p7 += 1
        print(f"  ✅ 新增 {new_idle_tasks_p7} 個 P7 任務 (>=100k, 填補空缺)", file=sys.stderr)

    else:
        print(f"  ⚠️ 警告: 找不到 'Port數' 欄位，跳過 P6/P7 任務", file=sys.stderr)
    
    # ==========================================================
    # ✅ P1-P5 任務清單 (主佇列)
    # ==========================================================
    task_queue.sort(key=lambda x: (x[0], x[5], x[1][0], x[1][1], x[1][2], x[2])) # ✅ Rule 2: 按藥名 (x[5]) 排序
    task_dict = {(t[5], t[2]): t for t in task_queue} 

    print(f"✅ 需求計算完成: {len(task_queue)} 個 P1-P5 任務, {len(p6_p7_tasks)} 個 P6/P7 任務\n", file=sys.stderr)

    return task_queue, task_dict, p6_p7_tasks

class ResourceManager:
    """V9.9.1: 資源管理器 (支援 P0 實驗/非實驗模式)"""
    
    def __init__(self):
        # P1-P5 資源
        self.person_usage = {}
        self.person_day_limit = {}
        self.dryer_usage = {}
        
        # V9.9: P0 專用資源 (非實驗模式：強制，P1-P5 都不能衝突)
        self.p0_strict_person_usage = {}
        self.p0_strict_dryer_usage = {}
        self.p0_strict_port_usage = {}
        self.p0_strict_ivek_usage = {}
        
        # V9.9.1: P0 實驗模式資源 (P1/P2 允許衝突，P3-P5 不能衝突)
        self.p0_experiment_person_usage = {}
        self.p0_experiment_dryer_usage = {}
        self.p0_experiment_port_usage = {}
        self.p0_experiment_ivek_usage = {}
        
        # ✅ 新增：凍乾機維修狀態追蹤
        self.dryer_maintenance = {}  # {day_key: set(['LD-3', 'LD-5', ...])}

    def reset_shift(self, shift_key):
        """重置某個班次的 P1-P5 資源"""
        self.person_usage[shift_key] = set()
        
    def reset_day(self, day_key):
        """重置某天的 P1-P5 資源"""
        self.person_day_limit[day_key] = set()
        self.dryer_usage[day_key] = set()

    # ========================================
    # V9.9: P0 非實驗任務預訂 (強制，P1-P5 都不能衝突)
    # ========================================

    def book_p0_strict_person(self, person, shift_key, day_key):
        """P0 (非實驗) 強制預訂人員"""
        person_lower = person.lower() # ✅ 修正：使用小寫
        if shift_key not in self.p0_strict_person_usage:
            self.p0_strict_person_usage[shift_key] = set()
        if shift_key not in self.person_usage:
            self.person_usage[shift_key] = set()

        if person_lower in self.p0_strict_person_usage[shift_key]: # ✅ 修正：使用小寫
            print(f"    ⚠️ P0 警告: 人員 {person} 在 {shift_key} 已被其他 P0 任務佔用", file=sys.stderr)
        
        if person_lower in self.person_usage[shift_key]: # ✅ 修正：使用小寫
            print(f"    ⚠️ P0 警告: 人員 {person} 在 {shift_key} 已被 P1-P5 任務佔用", file=sys.stderr)

        self.p0_strict_person_usage[shift_key].add(person_lower) # ✅ 修正：使用小寫
        self.person_usage[shift_key].add(person_lower) # ✅ 修正：使用小寫
        print(f"    🟢 P0 預訂人員: {person} ({shift_key})", file=sys.stderr) # Log 仍顯示原始大小寫

    def book_p0_strict_dryer(self, dryer, day_key):
        """P0 (非實驗) 強制預訂凍乾機"""
        if day_key not in self.p0_strict_dryer_usage:
            self.p0_strict_dryer_usage[day_key] = set()
        if day_key not in self.dryer_usage:
            self.dryer_usage[day_key] = set()
            
        if dryer in self.p0_strict_dryer_usage[day_key]:
            print(f"    ⚠️ P0 警告: 凍乾機 {dryer} 在 {day_key} 已被其他 P0 任務佔用", file=sys.stderr)
        
        if dryer in self.dryer_usage[day_key]:
            print(f"    ⚠️ P0 警告: 凍乾機 {dryer} 在 {day_key} 已被 P1-P5 任務佔用", file=sys.stderr)

        self.p0_strict_dryer_usage[day_key].add(dryer)
        self.dryer_usage[day_key].add(dryer)
        print(f"    🟢 P0 預訂凍乾機: {dryer} ({day_key})", file=sys.stderr)

    def book_p0_strict_port_pair(self, shift_key, port_num):
        """P0 (非實驗) 強制預訂 Port"""
        if shift_key not in self.p0_strict_port_usage:
            self.p0_strict_port_usage[shift_key] = set()

        if port_num % 2 == 0:
            pair = (port_num - 1, port_num)
        else:
            pair = (port_num, port_num + 1)
        
        port1, port2 = pair
        
        if (port1 in self.p0_strict_port_usage[shift_key]) or (port2 in self.p0_strict_port_usage[shift_key]):
             print(f"    ⚠️ P0 警告: Port {port1}/{port2} 在 {shift_key} 已被其他 P0 任務佔用", file=sys.stderr)
        
        self.p0_strict_port_usage[shift_key].add(port1)
        self.p0_strict_port_usage[shift_key].add(port2)
        print(f"    🟢 P0 預訂 Ports: {port1} 和 {port2} ({shift_key})", file=sys.stderr)

    def book_p0_strict_ivek(self, day_key):
        """P0 (非實驗) 強制預訂 IVEK"""
        if day_key not in self.p0_strict_ivek_usage:
            self.p0_strict_ivek_usage[day_key] = set()
            
        if "IVEK" in self.p0_strict_ivek_usage[day_key]:
            print(f"    ⚠️ P0 警告: IVEK 在 {day_key} 已被其他 P0 任務佔用", file=sys.stderr)
            
        self.p0_strict_ivek_usage[day_key].add("IVEK")
        print(f"    🟢 P0 預訂 IVEK ({day_key})", file=sys.stderr)

    
    # ========================================
    # V9.9.1: P0 實驗任務預訂 (P1/P2 允許衝突，P3-P5 不能衝突)
    # ========================================

    def book_p0_experiment_person(self, person, shift_key):
        """P0 (實驗) 預訂人員"""
        person_lower = person.lower() # ✅ 修正：使用小寫
        if shift_key not in self.p0_experiment_person_usage:
            self.p0_experiment_person_usage[shift_key] = set()
        
        if person_lower in self.p0_experiment_person_usage[shift_key]: # ✅ 修正：使用小寫
            print(f"    ⚠️ P0 (實驗) 警告: 人員 {person} 在 {shift_key} 已被其他實驗 P0 佔用", file=sys.stderr)
        
        self.p0_experiment_person_usage[shift_key].add(person_lower) # ✅ 修正：使用小寫
        print(f"    🧪 P0 (實驗) 預訂人員: {person} ({shift_key})", file=sys.stderr) # Log 仍顯示原始大小寫

    def book_p0_experiment_dryer(self, dryer, day_key):
        """P0 (實驗) 預訂凍乾機"""
        if day_key not in self.p0_experiment_dryer_usage:
            self.p0_experiment_dryer_usage[day_key] = set()
        
        if dryer in self.p0_experiment_dryer_usage[day_key]:
            print(f"    ⚠️ P0 (實驗) 警告: 凍乾機 {dryer} 在 {day_key} 已被其他實驗 P0 佔用", file=sys.stderr)
        
        self.p0_experiment_dryer_usage[day_key].add(dryer)
        print(f"    🧪 P0 (實驗) 預訂凍乾機: {dryer} ({day_key})", file=sys.stderr)

    def book_p0_experiment_port_pair(self, shift_key, port_num):
        """P0 (實驗) 預訂 Port"""
        if shift_key not in self.p0_experiment_port_usage:
            self.p0_experiment_port_usage[shift_key] = set()

        if port_num % 2 == 0:
            pair = (port_num - 1, port_num)
        else:
            pair = (port_num, port_num + 1)
        
        port1, port2 = pair
        
        if (port1 in self.p0_experiment_port_usage[shift_key]) or (port2 in self.p0_experiment_port_usage[shift_key]):
             print(f"    ⚠️ P0 (實驗) 警告: Port {port1}/{port2} 在 {shift_key} 已被其他實驗 P0 佔用", file=sys.stderr)
        
        self.p0_experiment_port_usage[shift_key].add(port1)
        self.p0_experiment_port_usage[shift_key].add(port2)
        print(f"    🧪 P0 (實驗) 預訂 Ports: {port1} 和 {port2} ({shift_key})", file=sys.stderr)

    def book_p0_experiment_ivek(self, day_key):
        """P0 (實驗) 預訂 IVEK"""
        if day_key not in self.p0_experiment_ivek_usage:
            self.p0_experiment_ivek_usage[day_key] = set()
            
        if "IVEK" in self.p0_experiment_ivek_usage[day_key]:
            print(f"    ⚠️ P0 (實驗) 警告: IVEK 在 {day_key} 已被其他實驗 P0 佔用", file=sys.stderr)
            
        self.p0_experiment_ivek_usage[day_key].add("IVEK")
        print(f"    🧪 P0 (實驗) 預訂 IVEK ({day_key})", file=sys.stderr)
        
    # ========================================
    # V9.9.1: P1-P5 任務檢查 (需考慮優先級)
    # ========================================

    def can_use_person(self, person, shift_key, day_key, is_single_slot, priority=None):
        """檢查人員是否可用 (P1-P5)"""
        # ========================================
        # V9.9.9: [新增] 檢查人員是否請假
        # ========================================
        try:
            date_obj = datetime.strptime(day_key, '%Y-%m-%d').date()
            if is_person_on_vacation(person, date_obj): # ✅ 修正：is_person_on_vacation 內部會處理 .lower()
                return False, f"{person} 在 {day_key} 請假"
        except Exception as e:
            print(f"              ⚠️ 無法檢查 {person} 的請假狀態: {e}", file=sys.stderr)
        
        person_lower = person.lower() # ✅ 修正：為所有比對建立小寫版本

        # 檢查非實驗 P0 衝突（所有 P1-P5 都不能衝突）
        if person_lower in self.p0_strict_person_usage.get(shift_key, set()): # ✅ 修正：使用小寫
            return False, f"{person} 被 P0 (非實驗) 佔用"
        
        # 檢查實驗 P0 衝突（僅 P3-P5 不能衝突）
        if priority is not None and priority >= 3:
            if person_lower in self.p0_experiment_person_usage.get(shift_key, set()): # ✅ 修正：使用小寫
                return False, f"{person} 被 P0 (實驗) 佔用"

        # 檢查 P1-P5 衝突
        if person_lower in self.person_usage.get(shift_key, set()): # ✅ 修正：使用小寫
            return False, f"{person} 在 {shift_key} 已被佔用"
        
        # 檢查單次限制
        if is_single_slot and person_lower in self.person_day_limit.get(day_key, set()): # ✅ 修正：使用小寫
            return False, f"{person} 在 {day_key} 已配過（單次限制）"
        
        return True, None
    
    def book_person(self, person, shift_key, day_key, is_single_slot, priority=None):
        """預訂人員 (P1-P5)"""
        can_use, reason = self.can_use_person(person, shift_key, day_key, is_single_slot, priority)
        
        if not can_use:
             raise Exception(reason)
        
        person_lower = person.lower() # ✅ 修正：使用小寫
        if shift_key not in self.person_usage:
            self.person_usage[shift_key] = set()
        self.person_usage[shift_key].add(person_lower) # ✅ 修正：使用小寫
        
        if is_single_slot:
            if day_key not in self.person_day_limit:
                self.person_day_limit[day_key] = set()
            self.person_day_limit[day_key].add(person_lower) # ✅ 修正：使用小寫
        
        print(f"        🟢 預訂人員: {person} ({shift_key})", file=sys.stderr) # Log 仍顯示原始大小寫

    def can_use_dryer(self, dryer, day_key, priority=None):
        """檢查凍乾機是否可用 (P1-P5)"""
        # ✅ 新增：檢查維修狀態（最優先檢查）
        if dryer in self.dryer_maintenance.get(day_key, set()):
            return False, f"{dryer} 在 {day_key} 維修中"
        
        # 檢查非實驗 P0 衝突（所有 P1-P5 都不能衝突）
        if dryer in self.p0_strict_dryer_usage.get(day_key, set()):
            return False, f"{dryer} 被 P0 (非實驗) 佔用"
        
        # 檢查實驗 P0 衝突（僅 P3-P5 不能衝突）
        if priority is not None and priority >= 3:
            if dryer in self.p0_experiment_dryer_usage.get(day_key, set()):
                return False, f"{dryer} 被 P0 (實驗) 佔用"

        # 檢查 P1-P5 衝突
        if dryer in self.dryer_usage.get(day_key, set()):
            return False, f"{dryer} 在 {day_key} 已被佔用"
            
        return True, None

    def book_dryer(self, dryer, day_key, priority=None):
        """預訂凍乾機 (P1-P5)"""
        can_use, reason = self.can_use_dryer(dryer, day_key, priority)
        
        if not can_use:
            raise Exception(reason)
            
        if day_key not in self.dryer_usage:
            self.dryer_usage[day_key] = set()
        self.dryer_usage[day_key].add(dryer)
        print(f"        🟢 預訂凍乾機: {dryer} ({day_key})", file=sys.stderr)

    def mark_dryer_maintenance(self, dryer, day_key):
        """標記凍乾機在指定日期維修中"""
        if day_key not in self.dryer_maintenance:
            self.dryer_maintenance[day_key] = set()
        self.dryer_maintenance[day_key].add(dryer)
        print(f"    🔧 標記凍乾機維修: {dryer} ({day_key})", file=sys.stderr)
    
    # ========================================
    # V9.9.1: P1-P5 檢查 P0 Port/IVEK 佔用
    # ========================================

    def is_p0_strict_port_booked(self, shift_key, port_num):
        """P1-P5 檢查非實驗 P0 是否佔用此 Port"""
        return port_num in self.p0_strict_port_usage.get(shift_key, set())

    def is_p0_strict_ivek_booked(self, day_key):
        """P1-P5 檢查非實驗 P0 是否佔用 IVEK"""
        return "IVEK" in self.p0_strict_ivek_usage.get(day_key, set())

    def is_p0_experiment_port_booked(self, shift_key, port_num):
        """P3-P5 檢查實驗 P0 是否佔用此 Port"""
        return port_num in self.p0_experiment_port_usage.get(shift_key, set())

    def is_p0_experiment_ivek_booked(self, day_key):
        """P3-P5 檢查實驗 P0 是否佔用 IVEK"""
        return "IVEK" in self.p0_experiment_ivek_usage.get(day_key, set())

    def get_status(self, shift_key=None, day_key=None):
        """獲取資源使用狀態"""
        status = {}
        if shift_key:
            status['people'] = list(self.person_usage.get(shift_key, set()))
            status['people_p0_strict'] = list(self.p0_strict_person_usage.get(shift_key, set()))
            status['people_p0_experiment'] = list(self.p0_experiment_person_usage.get(shift_key, set()))
        if day_key:
            status['dryers'] = list(self.dryer_usage.get(day_key, set()))
            status['dryers_p0_strict'] = list(self.p0_strict_dryer_usage.get(day_key, set()))
            status['dryers_p0_experiment'] = list(self.p0_experiment_dryer_usage.get(day_key, set()))
            status['day_limited'] = list(self.person_day_limit.get(day_key, set()))
        return status
# --------------------------------------------------------------------
# 5. 排程器
# --------------------------------------------------------------------
class Scheduler:
    def __init__(self, data):
        self.data = data
        self.db_constraints = data["db"]
        self.schedule_df = pd.DataFrame()  
       # ✅ 修改：接收 P6/P7 任務清單
        self.task_queue, self.task_dict, self.p6_p7_fill_tasks = calculate_demand_queue(data)
        self.constraints = self.db_constraints.get("Dosing_Constraints", pd.DataFrame())
        self.beads_dry_info = self.db_constraints.get("Beads_Dry_Count_Info", pd.DataFrame())
        
        self.scheduled_pns = set() 
        self.days_to_schedule = SCHEDULING_DAYS
        self.w1_needs_met = False
        self.simulated_inventory = data["demand"].set_index('料號')['Stock_plus_Dosing'].to_dict()
        
        self.shift_port_counter = {}
        self.resources = ResourceManager()
        
        self.no_dryer_records = []
        self.p0_strict_records = []      # V9.9: 非實驗 P0
        self.p0_experiment_records = []  # V9.9: 實驗 P0
        # ========================================
        # ✅ 新增：CK/PHOS/K 會 "污染" 凍乾機 3 天, GLIPA 不可使用
        # ========================================
        # "被限制的" 群組 (GLIPA)
        self.RULE_GLIPA_PNs = {'5714400220', '5714400221'} 
        # "執行限制的" 群組 (CK/PHOS/K...)
        self.RULE_BLOCKING_PNs = {
            '5714400222', '5714400199', # CK
            '5714400214', '5714400215', # PHOS (VB)
            '5714400116', '5714400226', '5714400216', '5714400117' # K1, K2, K3
        }
        self.RULE_BLOCK_DAYS = 2 # 污染 3 天 (Day 0, 1, 2)。第 3 天 (Day 3) 才可用。
        
        # 用來追蹤哪台凍乾機被污染到哪一天
        # 格式: { 'LD-5': date(2025, 11, 19), ... }
        self.dryer_block_until_date = {}

    def _create_dummy_task(self, pn, group_name): # ✅ 修改參數
        try:
            # ✅ 修改查詢邏輯
            # 1. 根據 PN 查詢
            bdc_info_all = self.beads_dry_info[self.beads_dry_info["料號"] == pn]
            
            # 2. 根據 group_name 篩選出唯一一筆
            bdc_info_group = bdc_info_all[bdc_info_all["BDC_Group_Name"] == group_name]
            
            if bdc_info_group.empty:
                return None # 找不到對應的組合
            
            bdc_info = bdc_info_group.iloc[0]
            
            marker_name = bdc_info["BDC_Marker_Name"]
            prod_qty = bdc_info["BDC_Prod_Qty"]
            
            if pn in self.constraints.index:
                c_data = self.constraints.loc[pn]
                if isinstance(c_data, pd.DataFrame):
                    c = c_data.iloc[0]
                else:
                    c = c_data
                subp = (c.get("V9_Dryer_Count", 99), c.get("V9_Person_Count", 99), 99)
            else:
                subp = (99, 99, 99)
                
            return (99, subp, pn, 0, "NoDemand", group_name, marker_name, prod_qty)
        except Exception:
            return None

    def find_batch(self, current_task_info):
        pn = current_task_info[2]
        group_name = current_task_info[5]
        task_priority = current_task_info[0]

        if not group_name or group_name == 'UNKNOWN':
            return [current_task_info]

        try:
            # ✅ 修改：查詢這個群組的所有 PN (從 '料號' 欄位, 而非 index)
            pns_in_group = self.beads_dry_info[
                self.beads_dry_info["BDC_Group_Name"] == group_name
            ]["料號"].tolist()
        except:
            return [current_task_info]
            
        batch_tasks = []
        for pn_in_group in pns_in_group:
            
            # ✅ 新增：使用組合鍵 (group_name, pn)
            task_key = (group_name, pn_in_group)
            
            if (pn_in_group in self.scheduled_pns) and (task_priority != 5):
                continue 

            # ✅ 修改：使用組合鍵查詢 BDC_Prod_Qty
            bdc_rows = self.beads_dry_info[
                (self.beads_dry_info["料號"] == pn_in_group) &
                (self.beads_dry_info["BDC_Group_Name"] == group_name)
            ]
            
            if not bdc_rows.empty:
                prod_qty = bdc_rows.iloc[0]["BDC_Prod_Qty"]
                if prod_qty == 0:
                    continue
            else:
                continue 
            
            # ✅ 修改：使用組合鍵 (task_key) 查詢
            if task_key in self.task_dict:
                batch_tasks.append(self.task_dict[task_key])
            else:
                # ✅ 修改：傳入 group_name 參數
                dummy = self._create_dummy_task(pn_in_group, group_name)
                if dummy:
                    batch_tasks.append(dummy)
        
        unique_batch = list({t[2]: t for t in batch_tasks}.values())
        
        if not unique_batch:
            return []

        if len(unique_batch) > 2:
            urgent_tasks = [t for t in unique_batch if t[1][2] in (0, 1)] 
            if urgent_tasks:
                return urgent_tasks
            return unique_batch
        
        return unique_batch

    def _normalize_constraints(self, batch_tasks):
        constraints_raw = []
        for task in batch_tasks:
            pn = task[2]
            prod_qty = task[7]
            
            constr_dict = {}
            if pn in self.constraints.index:
                constr_data = self.constraints.loc[pn]
                if isinstance(constr_data, pd.DataFrame):
                    constr = constr_data.iloc[0]
                else:
                    constr = constr_data
                
                if isinstance(constr, pd.Series):
                    constr_dict = constr.to_dict()
                elif isinstance(constr, dict):
                    constr_dict = constr.copy()

            if "PN" not in constr_dict:
                constr_dict["PN"] = pn
                
            constr_dict["數量"] = prod_qty 
            
            if pn in self.constraints.index:
                c_data_v9 = self.constraints.loc[pn]
                if isinstance(c_data_v9, pd.DataFrame):
                    c_v9 = c_data_v9.iloc[0]
                else:
                    c_v9 = c_data_v9
                        
                constr_dict["V9_Slot_List"] = c_v9.get("V9_Slot_List", ["AM", "PM"])
                constr_dict["V9_Single_Slot"] = c_v9.get("V9_Single_Slot", False)
                # ✅ V9.9.10: 保留原始交藥時間
                constr_dict["交藥時間"] = c_v9.get("交藥時間", "")
                # ✅ V9.9.11: 保留備註（週五限制）
                constr_dict["備註"] = c_v9.get("備註", "")
            
            constraints_raw.append(constr_dict)
                
        return constraints_raw
    
    def get_delivery_time_from_constraints(self, constraints_list, slot_name, current_day, day_batch_times, day_key):
        """
        V9.9.10: 根据配藥限制的交藥時間决定 RD給藥時間
        
        规则：
        - 如果交藥時間有两组时间或 *,*：可以排 AM/PM，使用动态时间
        - 如果只有一个时间（如 16:00）：必须使用该时间
        """
        AM_START = time(10, 30)
        PM_START = time(15, 30)
        PM_END = time(17, 30)
        BATCH_INTERVAL_MINUTES = 30
        
        # 检查批次中所有 PN 的交藥時間
        single_time = None
        has_single_time = False
        
        for c in constraints_list:
            delivery_time_str = str(c.get("交藥時間", "")).strip()
            
            if not delivery_time_str or delivery_time_str == "":
                continue
            
            # 分割时间（逗号分隔）
            times = [t.strip() for t in delivery_time_str.split(',') if t.strip() and t.strip() != '*']
            
            # 如果只有一个时间，且包含 ':'
            if len(times) == 1 and ':' in times[0]:
                try:
                    time_parts = times[0].split(':')
                    hour = int(time_parts[0])
                    minute = int(time_parts[1]) if len(time_parts) > 1 else 0
                    
                    # 验证时间是否符合当前班次
                    if slot_name == "AM" and 9 <= hour <= 14:
                        single_time = time(hour, minute)
                        has_single_time = True
                        break
                    elif slot_name == "PM" and 15 <= hour <= 23:
                        single_time = time(hour, minute)
                        has_single_time = True
                        break
                except Exception as e:
                    print(f"        ⚠️ 解析交藥時間失敗: {times[0]} - {e}", file=sys.stderr)
                    pass
        
        # 如果有单一指定时间，直接使用
        if has_single_time and single_time:
            print(f"        📌 使用指定交藥時間: {single_time.strftime('%H:%M')}", file=sys.stderr)
            return single_time
        
        # 否则使用默认的动态分配策略
        if slot_name == "AM":
            if not day_batch_times[day_key]["AM"]:
                return AM_START
            else:
                last_time = max(day_batch_times[day_key]["AM"])
                last_dt = datetime.combine(current_day, last_time)
                next_dt = last_dt + timedelta(minutes=BATCH_INTERVAL_MINUTES)
                next_time = next_dt.time()
                
                if next_time >= time(15, 0):
                    return None  # AM 时段已满
                return next_time
        else:  # PM
            if not day_batch_times[day_key]["PM"]:
                return PM_START
            else:
                last_time = max(day_batch_times[day_key]["PM"])
                last_dt = datetime.combine(current_day, last_time)
                next_dt = last_dt + timedelta(minutes=BATCH_INTERVAL_MINUTES)
                next_time = next_dt.time()
                
                if next_time >= PM_END:
                    return None  # PM 时段已满
                return next_time
            
    def can_schedule_on_friday(self, constraints_list, current_day):
        """
        V9.9.11: 檢查批次是否可以排在週五
        
        規則：
        - 如果批次中任何 PN 的備註包含「不能放週五滴定」
        - 且 current_day 是週五（weekday == 4）
        - 則返回 False
        
        參數：
            constraints_list: 批次限制列表
            current_day: datetime 物件
            
        返回：
            (can_schedule, reason)
            - can_schedule: bool - 是否可排程
            - reason: str - 不可排程的原因（可選）
        """
        # 檢查是否為週五（0=週一, 4=週五）
        if current_day.weekday() != 4:
            return True, None  # 不是週五，可以排程
        
        # 遍歷批次中的所有 PN
        for c in constraints_list:
            remark = str(c.get("備註", "")).strip()
            pn = c.get("PN", "未知")
            
            # 檢查備註是否包含「不能放週五滴定」
            if "不能放週五滴定" in remark:
                reason = f"{pn} 備註限制：不能放週五滴定"
                print(f"        🔴 [{current_day.strftime('%Y-%m-%d')}] {reason}", file=sys.stderr)
                return False, reason
        
        # 批次中沒有週五限制
        return True, None
    
    def is_forbidden_end_time(self, end_datetime):
        """
        V9.9.12: 檢查預計結束時間是否落在禁止時段
        
        規則：
        - 禁止時段：03:00~08:00（凌晨時段）
        - 原因：避免半夜/清晨收藥
        
        參數：
            end_datetime: datetime 物件 - 預計結束時間
            
        返回：
            (is_forbidden, reason)
            - is_forbidden: bool - 是否在禁止時段
            - reason: str - 禁止原因（可選）
        """
        end_time = end_datetime.time()
    
        # 檢查是否在 03:00~08:00 之間
        if FORBIDDEN_END_TIME_START <= end_time <= FORBIDDEN_END_TIME_END:
            reason = f"預計結束時間 {end_time.strftime('%H:%M')} 落在禁止時段 (03:00~08:00)"
            print(f"        🔴 {reason}", file=sys.stderr)
            return True, reason
        
        return False, None
            
    def _calculate_ports_and_time(self, prod_qty, num_ports_str):
        """
        ✅ 修改：根據資料庫的 Port數 欄位計算
        - 移除 3-8 小時限制
        - 移除 PORT_OPTIONS = [4, 2] 限制
        """
        num_ports = 2 # 預設值
        
        try:
            if num_ports_str.upper() == "IVEK":
                # IVEK 任務, port 數量不重要 (假設為 2, 用於計算時間)
                num_ports = 2 
            else:
                num_ports = int(num_ports_str)
                if num_ports <= 0:
                    num_ports = 2
        except:
            print(f"    ⚠️ 警告: Port數欄位值 '{num_ports_str}' 無法解析, 自動設為 2", file=sys.stderr)
            num_ports = 2
        
        if prod_qty == 0:
            return num_ports, 0.5 # 預設 0.5 小時

        # ✅ 移除 3-8 小時限制，直接計算
        time_hrs = prod_qty / (DOSING_RATE_PER_HR * num_ports)
        
        return num_ports, time_hrs

    def calculate_person_priority(self, constraints_list):
        """計算批次的人員限制優先級"""
        all_people = set()
        for c in constraints_list:
            for key in ["配藥人-1", "配藥人-2", "配藥人-3"]:
                person = c.get(key)
                if person and str(person).strip():
                    all_people.add(str(person).strip().lower()) # ✅ 修正：使用小寫
        
        all_people.discard("")
        
        # ✅ 修正：還原為
        return len(all_people), list(all_people)

    def find_available_person_v10(self, constraints_list, delivery_dt, slot_name, priority=None):
        """V10: 簡化的人員查找"""
        # ✅ 修正：還原 find_available_person_v10 的完整邏輯
        day_key = delivery_dt.strftime('%Y-%m-%d')
        shift_key = f"{day_key}_{slot_name}"
        
        all_people = set()
        has_single_slot = False
        
        for c in constraints_list:
            for key in ["配藥人-1", "配藥人-2", "配藥人-3"]:
                person = c.get(key)
                if person and str(person).strip():
                    all_people.add(str(person).strip().lower()) # ✅ 修正：使用小寫
            
            if c.get("V9_Single_Slot", False):
                has_single_slot = True
        
        if not all_people:
            print(f"        🔴 [{day_key} {slot_name}] 批次無可配藥人員", file=sys.stderr)
            return None, False
        
        for person in sorted(all_people):
            # ========================================
            # V9.9.9: [新增] 檢查人員是否請假
            # ========================================
            if is_person_on_vacation(person, delivery_dt):
                print(f"        🔴 [{day_key}] {person} 請假，不可用", file=sys.stderr)
                continue
            can_use, reason = self.resources.can_use_person(
                person, shift_key, day_key, has_single_slot, priority
            )
            if can_use:
                print(f"        🟢 [{day_key} {slot_name}] 選中人員: {person}", file=sys.stderr)
                return person, has_single_slot
        
        print(f"        🔴 [{day_key} {slot_name}] 人員全被佔用: {list(all_people)}", file=sys.stderr)
        return None, False
    
    def find_available_dryer_v10(self, constraints_list, delivery_dt, priority=None, batch_tasks=None):
        """V10: 簡化的凍乾機查找 (✅ 新增 GLIPA/CK 規則檢查)"""
        day_key = delivery_dt.strftime('%Y-%m-%d')
        
        dryer_sets = [
            set(str(c.get("可用凍乾機", "")).split(',')) 
            for c in constraints_list
        ]
        common_dryers = set.intersection(*dryer_sets)
        common_dryers.discard("")
        
        if not common_dryers:
            print(f"        🔴 [{day_key}] 批次無共同凍乾機", file=sys.stderr)
            return None
            
        # ✅ 新增：檢查目前是否為 GLIPA
        current_batch_pns = set(t[2] for t in batch_tasks) if batch_tasks else set()
        is_glipa_batch = any(pn in self.RULE_GLIPA_PNs for pn in current_batch_pns)
        current_date = delivery_dt.date()
        is_p1_task = (priority == 1) # 檢查是否為 P1 任務

        for dryer in sorted(common_dryers):
            dryer = dryer.strip()
            if not dryer:
                continue
            
            # ========================================
            # ✅ 新增：GLIPA/CK 衝突規則檢查
            # ========================================
            if is_glipa_batch and dryer in self.dryer_block_until_date:
                blocked_until = self.dryer_block_until_date[dryer]
                
                # 檢查是否仍在 3 天的封鎖期內
                if current_date <= blocked_until:
                    if is_p1_task:
                        # 是 P1 任務, 例外允許排程
                        print(f"        ⚠️ [{day_key}] {dryer} 被 CK/K 規則封鎖, 但因 GLIPA 為 P1 任務例外允許", file=sys.stderr)
                        # (不 'continue', 繼續執行)
                    else:
                        # 不是 P1 任務, 嚴格禁止
                        print(f"        🔴 [{day_key}] {dryer} 被 CK/K 規則封鎖至 {blocked_until}, GLIPA (P{priority}) 不可用", file=sys.stderr)
                        continue # 跳過這台凍乾機
            # ========================================

            can_use, reason = self.resources.can_use_dryer(dryer, day_key, priority)
            if can_use:
                print(f"        🟢 [{day_key}] 選中凍乾機: {dryer}", file=sys.stderr)
                return dryer
        
        print(f"        🔴 [{day_key}] 凍乾機全被佔用: {list(common_dryers)}", file=sys.stderr)
        return None
    
    def check_availability(self, batch_tasks, constraints_list, delivery_dt, slot_name, priority=None):
        """檢查資源可用性"""
        day_key = delivery_dt.strftime('%Y-%m-%d')
        
        # ✅ 修改：從約束中讀取 Port數
        # 假設同批次的 Port數/IVEK 設定都相同
        num_ports_str = "2" # 預設值
        if constraints_list:
            num_ports_str = str(constraints_list[0].get("Port數", "2")).strip()

        # ✅ 修改：is_ivek_task 基於 Port數 欄位
        is_ivek_task = (num_ports_str.upper() == "IVEK")
        
        if is_ivek_task and self.resources.is_p0_strict_ivek_booked(day_key):
            print(f"    🔴 {day_key} IVEK 已被 P0 (非實驗) 佔用", file=sys.stderr)
            return None
        
        if priority is not None and priority >= 3:
            if is_ivek_task and self.resources.is_p0_experiment_ivek_booked(day_key):
                print(f"    🔴 {day_key} IVEK 已被 P0 (實驗) 佔用", file=sys.stderr)
                return None
        
        person_result = self.find_available_person_v10(constraints_list, delivery_dt, slot_name, priority)
        if not person_result or person_result[0] is None:
            return None
        
        person, is_single_slot = person_result

        # ✅ 修改：傳入 batch_tasks 參數, 讓 find_available_dryer_v10 知道現在排的是哪個任務
        dryer_id = self.find_available_dryer_v10(constraints_list, delivery_dt, priority, batch_tasks)
        if not dryer_id:
            return None

        try:
            dosing_start_dt = delivery_dt + timedelta(minutes=DOSING_PREP_TIME_MIN)
            
            qtys = [float(c.get("數量", 0)) for c in constraints_list if float(c.get("數量", 0)) > 0]
            max_qty = max(qtys) if qtys else 1000
            
            # ✅ 修改：傳入 num_ports_str, 並使用新的計算邏輯
            num_ports, dosing_hrs = self._calculate_ports_and_time(max_qty, num_ports_str)
            
            dosing_end_dt = add_hours_to_time(dosing_start_dt, dosing_hrs)
            
            freeze_duration = float(max(c.get("凍乾時間", 12) for c in constraints_list))
            final_end_dt = add_hours_to_time(dosing_end_dt, freeze_duration)
        
        except Exception as e:
            print(f"        🔴 時間計算失敗: {e}", file=sys.stderr)
            return None

        return {
            "person": person,
            "dryer": dryer_id, 
            "is_single_slot": is_single_slot,
            "times": {
                "delivery_dt": delivery_dt,
                "dosing_start_dt": dosing_start_dt,
                "dosing_end_dt": dosing_end_dt,
                "final_end_dt": final_end_dt,
                "freeze_duration_hr": freeze_duration,
                "dosing_hrs": dosing_hrs
            },
            "num_ports_per_pn": num_ports,
            "num_ports_str": num_ports_str, # ✅ 新增：傳遞 IVEK/"4" 標記
            'constraints_list': constraints_list
        }

    def book_resources(self, resources, delivery_dt, slot_name, priority=None):
        """V9.9.1: 統一使用 V10 資源預訂"""
        # ✅ 移除了未使用的 batch_tasks 參數
        person = resources["person"]
        dryer = resources["dryer"]
        is_single_slot = resources["is_single_slot"]

        day_key = delivery_dt.strftime('%Y-%m-%d')
        shift_key = f"{day_key}_{slot_name}"

        try:
            self.resources.book_person(person, shift_key, day_key, is_single_slot, priority)
        except Exception as e:
            # ✅ 修改：移除 return，只印出警告
            print(f"        ⚠️ 警告: 預訂人員 {person} 失敗: {e}", file=sys.stderr)

        try:
            self.resources.book_dryer(dryer, day_key, priority)
        except Exception as e:
            # ✅ 修改：移除 return，只印出警告
            print(f"        ⚠️ 警告: 預訂凍乾機 {dryer} 失敗: {e}", file=sys.stderr)
        
    def handle_p0_orders(self):
        """
        V9.9.5: 處理 P0 (插單/限制) 任務 - 增強診斷版
        """
        print("\n" + "="*70, file=sys.stderr)
        print(f"🔥 Step 3.0: 正在處理 P0 (插單/限制) 任務 ({CURRENT_VERSION})", file=sys.stderr)
        print("="*70, file=sys.stderr)
        
        df_p0_orders = self.db_constraints.get("P0_Orders", pd.DataFrame())
        
        if df_p0_orders.empty:
            print("  ⚠️ P0_Orders 表格為空", file=sys.stderr)
            return

        # 🔥 診斷 1：顯示讀取到的 P0 資料
        print(f"  📊 從資料庫讀取到 {len(df_p0_orders)} 筆 P0 任務", file=sys.stderr)
        print(f"  📋 P0 欄位: {df_p0_orders.columns.tolist()}", file=sys.stderr)
        
        # 🔥 修正：使用全域的排程起始日期而非今天
        global SCHEDULE_START_DATE
        
        if SCHEDULE_START_DATE:
            monday = SCHEDULE_START_DATE
            print(f"  📅 使用排程起始日期: {monday.strftime('%Y-%m-%d')}", file=sys.stderr)
        else:
            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            monday = (today + timedelta(days=-today.weekday(), weeks=0))
            print(f"  📅 使用今天推算的週一: {monday.strftime('%Y-%m-%d')}", file=sys.stderr)
        
        saturday = monday + timedelta(days=5)
        current_year = monday.year
        
        print(f"  🗓️  排程週範圍: {monday.strftime('%Y-%m-%d')} ~ {saturday.strftime('%Y-%m-%d')}", file=sys.stderr)
        
        # ========================================
        # ✅ 新增：處理凍乾機維修記錄
        # ========================================
        print("\n" + "="*70, file=sys.stderr)
        print("🔧 Step 3.0a: 處理凍乾機維修/故障/保養記錄", file=sys.stderr)
        print("="*70, file=sys.stderr)

        maintenance_count = 0

        for idx, row in df_p0_orders.iterrows():
            titrate = str(row.get("滴定機", "")).strip()
            marker = str(row.get("Marker", "")).strip()
            date_str = str(row.get("日期", "")).strip()
            dryer_str = str(row.get("凍乾機台", "")).strip()
            remark = str(row.get("備註", "")).strip()
            
            # 檢查條件
            if titrate != "" or marker != "":
                continue
            if not date_str:
                continue
            if not dryer_str:
                continue
            if not any(keyword in remark for keyword in ["維修", "故障", "保養"]):
                continue
            
            # 解析日期
            try:

                current_year = SCHEDULE_START_DATE.year if SCHEDULE_START_DATE else datetime.now().year
                
                date_str_clean = date_str.replace(' ', '').strip()
                task_date = None
                
                # 嘗試解析「11月24日」格式
                if re.search(r'(\d{1,2})月(\d{1,2})日', date_str_clean):
                    month = int(re.search(r'(\d{1,2})月', date_str_clean).group(1))
                    day = int(re.search(r'(\d{1,2})日', date_str_clean).group(1))
                    task_date = datetime(current_year, month, day)
                else:
                    # 其他格式
                    task_date = pd.to_datetime(date_str_clean, errors='coerce')
                    if pd.notna(task_date):
                        task_date = task_date.to_pydatetime()
                
                if task_date is None:
                    continue
                
                day_key = task_date.strftime('%Y-%m-%d')
                
                # 解析凍乾機編號
                dryer_num_match = re.search(r'\d+', dryer_str)
                if dryer_num_match:
                    dryer_num = int(dryer_num_match.group())
                    if 3 <= dryer_num <= 20:
                        dryer_name = str(dryer_num)
                        self.resources.mark_dryer_maintenance(dryer_name, day_key)
                        maintenance_count += 1
                        print(f"  ✅ {day_key}: {dryer_name} - {remark}", file=sys.stderr)
                
            except Exception as e:
                print(f"  ⚠️ 維修記錄解析失敗: {e}", file=sys.stderr)

        if maintenance_count > 0:
            print(f"\n  ✅ 共標記 {maintenance_count} 筆凍乾機維修記錄\n", file=sys.stderr)
        else:
            print(f"  ℹ️ 無凍乾機維修記錄\n", file=sys.stderr)
        
        p0_tasks_this_week = []
        p0_on_holiday = []
        p0_parse_failed = []
        p0_out_of_range = []
        
        # 🔥 診斷 2：逐筆處理並記錄每個任務的狀態
        for idx, row in df_p0_orders.iterrows():
            date_str = str(row.get("日期", "")).strip()
            marker = str(row.get("Marker", "")).strip()
            lot = str(row.get("Lot", "N/A")).strip()
            
            print(f"\n  ━━━ P0 任務 #{idx+1} ━━━", file=sys.stderr)
            print(f"      原始日期: '{date_str}'", file=sys.stderr)
            print(f"      Marker: '{marker}'", file=sys.stderr)
            print(f"      Lot: '{lot}'", file=sys.stderr)
            
            # 檢查必要欄位
            if not date_str or not marker:
                print(f"      ❌ 失敗: 缺少 {'日期' if not date_str else 'Marker'}", file=sys.stderr)
                continue
            
            # 🔥 修正：更強健的日期解析
            try:
                task_date = None
                date_str_clean = date_str.replace(' ', '').strip()
                
                # 嘗試多種日期格式
                parse_methods = [
                    # 格式1: YYYY-MM-DD 或 YYYY/MM/DD
                    lambda s: pd.to_datetime(s, format='%Y-%m-%d', errors='coerce'),
                    lambda s: pd.to_datetime(s, format='%Y/%m/%d', errors='coerce'),
                    
                    # 格式2: M月D日
                    lambda s: datetime(current_year, 
                                    int(re.search(r'(\d{1,2})月', s).group(1)),
                                    int(re.search(r'(\d{1,2})日', s).group(1))) 
                            if re.search(r'(\d{1,2})月(\d{1,2})日', s) else None,
                    
                    # 格式3: pandas 自動解析
                    lambda s: pd.to_datetime(s, errors='coerce')
                ]
                
                for method in parse_methods:
                    try:
                        result = method(date_str_clean)
                        if result is not None and not pd.isna(result):
                            task_date = result if isinstance(result, datetime) else result.to_pydatetime()
                            print(f"      ✅ 日期解析成功: {task_date.strftime('%Y-%m-%d')}", file=sys.stderr)
                            break
                    except:
                        continue
                
                if task_date is None:
                    raise ValueError(f"所有解析方法均失敗")
                
                # 🔥 診斷 3：檢查日期是否在範圍內
                if monday <= task_date <= saturday:
                    print(f"      ✅ 日期在本週範圍內", file=sys.stderr)
                    
                    # 檢查是否為休假日
                    if is_holiday(task_date):
                        p0_on_holiday.append({
                            'lot': lot,
                            'date': task_date.strftime('%Y-%m-%d'),
                            'date_obj': task_date
                        })
                        print(f"      ⚠️ 日期為休假日", file=sys.stderr)
                    else:
                        p0_tasks_this_week.append((task_date, row))
                        print(f"      ✅ 已加入本週任務列表", file=sys.stderr)
                else:
                    p0_out_of_range.append({
                        'lot': lot,
                        'date': task_date.strftime('%Y-%m-%d'),
                        'week_range': f"{monday.strftime('%Y-%m-%d')} ~ {saturday.strftime('%Y-%m-%d')}"
                    })
                    print(f"      ⚠️ 日期不在本週範圍（{task_date.strftime('%Y-%m-%d')}）", file=sys.stderr)
                        
            except Exception as e:
                p0_parse_failed.append({
                    'lot': lot,
                    'date_str': date_str,
                    'error': str(e)
                })
                print(f"      ❌ 日期解析失敗: {e}", file=sys.stderr)
                continue
        
        # 🔥 診斷 4：統計報告
        print(f"\n" + "="*70, file=sys.stderr)
        print(f"📊 P0 任務處理統計", file=sys.stderr)
        print(f"="*70, file=sys.stderr)
        print(f"  總計讀取: {len(df_p0_orders)} 筆", file=sys.stderr)
        print(f"  本週範圍內: {len(p0_tasks_this_week)} 筆", file=sys.stderr)
        print(f"  休假日衝突: {len(p0_on_holiday)} 筆", file=sys.stderr)
        print(f"  日期超出範圍: {len(p0_out_of_range)} 筆", file=sys.stderr)
        print(f"  日期解析失敗: {len(p0_parse_failed)} 筆", file=sys.stderr)
        
        # 顯示超出範圍的任務
        if p0_out_of_range:
            print(f"\n  ⚠️ 以下 P0 任務不在本週範圍：", file=sys.stderr)
            for item in p0_out_of_range:
                print(f"      • {item['lot']} ({item['date']}) - 本週範圍: {item['week_range']}", file=sys.stderr)
        
        # 顯示解析失敗的任務
        if p0_parse_failed:
            print(f"\n  ❌ 以下 P0 任務日期解析失敗：", file=sys.stderr)
            for item in p0_parse_failed:
                print(f"      • {item['lot']}: '{item['date_str']}' - {item['error']}", file=sys.stderr)
        
        print(f"="*70, file=sys.stderr)
        
        # 檢查休假日衝突
        if p0_on_holiday:
            print("\n" + "="*70, file=sys.stderr)
            print("❌ 排程終止：P0 任務日期衝突", file=sys.stderr)
            print("="*70, file=sys.stderr)
            print(f"以下 {len(p0_on_holiday)} 個 P0 任務的排程日期為休假日：", file=sys.stderr)
            for p0_info in p0_on_holiday:
                print(f"  • {p0_info['lot']} ({p0_info['date']})", file=sys.stderr)
            raise SystemExit("P0 任務與休假日衝突，排程已終止")
                
        if not p0_tasks_this_week:
            print(f"\n  ℹ️ 無本週 P0 任務需要處理", file=sys.stderr)
            return

        print(f"\n  🔄 開始處理 {len(p0_tasks_this_week)} 筆本週 P0 任務...", file=sys.stderr)
        
        # 處理每個 P0 任務
        processed_count = 0
        failed_count = 0
        
        for task_date, row in p0_tasks_this_week:
            day_key = task_date.strftime('%Y-%m-%d')
            lot = str(row.get("Lot", "N/A")).strip()
            staff = str(row.get("配藥同仁", "")).strip()
            dryer = str(row.get("凍乾機台", "")).strip()
            machine = str(row.get("滴定機", "")).strip().lower()
            rd_time_str = str(row.get("RD給藥時間", "09:00")).strip()
            remark = str(row.get("備註", "")).strip()
            
            print(f"\n  ━━━ 處理 P0: {lot} ({day_key}) ━━━", file=sys.stderr)
            
            # 檢查配藥人員
            if not staff:
                print(f"      ❌ 失敗: 未指定配藥同仁", file=sys.stderr)
                failed_count += 1
                continue
            
            # 檢查配藥人員是否請假
            if is_person_on_vacation(staff, task_date):
                print(f"      ❌ 失敗: 配藥人員 {staff} 請假", file=sys.stderr)
                failed_count += 1
                continue
            
            # 計算班次
            try:
                rd_hour = int(rd_time_str.split(':')[0])
                slot_name = "PM" if rd_hour >= 15 else "AM"
            except:
                slot_name = "AM"
                
            shift_key = f"{day_key}_{slot_name}"
            
            print(f"      配藥同仁: {staff}", file=sys.stderr)
            print(f"      班次: {slot_name}", file=sys.stderr)
            print(f"      滴定機: {machine}", file=sys.stderr)
            print(f"      凍乾機: {dryer}", file=sys.stderr)

            # 檢查是否無滴定機或凍乾機 → no_dryer
            if pd.isna(dryer) or dryer == "" or pd.isna(machine) or machine == "":
                print(f"      → 無滴定機或凍乾機，轉 no_dryer 處理", file=sys.stderr)
                
                constr_dict = {"配藥人-1": staff, "V9_Single_Slot": False}
                person_result = self._find_person_for_no_dryer(constr_dict, task_date)
                
                if person_result is None:
                    print(f"      ❌ 失敗: 無法預訂全天資源", file=sys.stderr)
                    failed_count += 1
                    continue
                
                person, _ = person_result
                
                self.no_dryer_records.append({
                    "日期": day_key,
                    "marker": row.get("Marker", ""),
                    "lot": lot,
                    "凍乾機台": "",
                    "數量": row.get("數量", 0),
                    "配藥同仁": person,
                    "RD給藥時間": rd_time_str,
                    "預計滴定時間": row.get("預計滴定時間", ""),
                    "預計結束": row.get("預計結束", ""),
                    "班次": slot_name,
                    "滴定機": machine,
                    "is_no_dryer": True,
                    "PN": row.get("PN", ""),
                    "備註": remark
                })
                self.scheduled_pns.add(lot)
                processed_count += 1
                print(f"      ✅ 成功: 已排入 no_dryer", file=sys.stderr)
                continue

            # 檢查是否為實驗任務
            is_experiment = "*實驗*" in remark
            
            if is_experiment:
                print(f"      → 實驗任務，將在 P2 後排程", file=sys.stderr)
                p0_data = row.to_dict()
                p0_data['P0_Parsed_Date'] = task_date
                p0_data['is_experiment'] = True
                self.p0_experiment_records.append(p0_data)
                self.scheduled_pns.add(lot)
                processed_count += 1
                print(f"      ✅ 成功: 已加入實驗 P0 列表", file=sys.stderr)
            else:
                print(f"      → 非實驗任務，優先排程", file=sys.stderr)
                
                try:
                    self.resources.book_p0_strict_person(staff, shift_key, day_key)
                    self.resources.book_p0_strict_dryer(dryer, day_key)
                    
                    if "ivek" in machine:
                        self.resources.book_p0_strict_ivek(day_key)
                        print(f"      → 已預訂 IVEK", file=sys.stderr)
                    elif "port" in machine:
                        try:
                            port_num = int(re.findall(r'\d+', machine)[0])
                            if 1 <= port_num <= 12:
                                self.resources.book_p0_strict_port_pair(shift_key, port_num)
                                print(f"      → 已預訂 Port{port_num}", file=sys.stderr)
                        except Exception as e:
                            print(f"      ⚠️ 警告: 無法解析滴定機編號: {e}", file=sys.stderr)
                    
                    p0_data = row.to_dict()
                    p0_data['P0_Parsed_Date'] = task_date
                    p0_data['is_experiment'] = False
                    self.p0_strict_records.append(p0_data)
                    
                    self.scheduled_pns.add(lot)
                    processed_count += 1
                    print(f"      ✅ 成功: 已強制排入非實驗 P0", file=sys.stderr)
                    
                except Exception as e:
                    print(f"      ❌ 失敗: {e}", file=sys.stderr)
                    traceback.print_exc(file=sys.stderr)
                    failed_count += 1

        # 最終統計
        print(f"\n" + "="*70, file=sys.stderr)
        print(f"✅ P0 任務處理完成", file=sys.stderr)
        print(f"="*70, file=sys.stderr)
        print(f"  成功處理: {processed_count} 筆", file=sys.stderr)
        print(f"  處理失敗: {failed_count} 筆", file=sys.stderr)
        print(f"  非實驗 P0: {len(self.p0_strict_records)} 筆", file=sys.stderr)
        print(f"  實驗 P0: {len(self.p0_experiment_records)} 筆", file=sys.stderr)
        print(f"  no_dryer: {len([r for r in self.no_dryer_records if r.get('is_no_dryer')])} 筆", file=sys.stderr)
        print(f"="*70 + "\n", file=sys.stderr)
        
    def schedule_p0_experiments(self, all_slots):
        """
        V9.9.1: 排程實驗 P0 (P2 後，P1/P2 允許衝突，P3-P5 不能衝突)
        """
        
        if not self.p0_experiment_records:
            print("  ... 無實驗 P0 任務需要排程", file=sys.stderr)
            return
        
        print(f"  開始排程 {len(self.p0_experiment_records)} 個實驗 P0 任務...", file=sys.stderr)
        
        for idx, p0_data in enumerate(self.p0_experiment_records, 1):
            # 步驟 1: 解析 P0 任務資料
            task_date = p0_data.get('P0_Parsed_Date')
            
            if task_date is None:
                print(f"  ❌ 實驗 P0 任務 #{idx}: 缺少日期資訊，跳過", file=sys.stderr)
                continue
            
            lot = str(p0_data.get("Lot", "N/A")).strip()
            marker = str(p0_data.get("Marker", "")).strip()
            staff = str(p0_data.get("配藥同仁", "")).strip()
            dryer = str(p0_data.get("凍乾機台", "")).strip()
            machine = str(p0_data.get("滴定機", "")).strip().lower()
            rd_time_str = str(p0_data.get("RD給藥時間", "09:00")).strip()
            remark = str(p0_data.get("備註", "")).strip()
            
            # 步驟 2: 驗證必要欄位
            if not staff:
                print(f"  ❌ 實驗 P0 任務 {lot}: 缺少配藥同仁，跳過", file=sys.stderr)
                continue
            
            if not dryer:
                print(f"  ❌ 實驗 P0 任務 {lot}: 缺少凍乾機台，跳過", file=sys.stderr)
                continue
            
            if not machine:
                print(f"  ❌ 實驗 P0 任務 {lot}: 缺少滴定機，跳過", file=sys.stderr)
                continue
            
            # 步驟 3: 計算班次
            try:
                rd_hour = int(rd_time_str.split(':')[0])
                slot_name = "PM" if rd_hour >= 15 else "AM"
            except Exception as e:
                print(f"  ⚠️ 實驗 P0 任務 {lot}: RD給藥時間解析失敗，預設為 AM", file=sys.stderr)
                slot_name = "AM"
            
            day_key = task_date.strftime('%Y-%m-%d')
            shift_key = f"{day_key}_{slot_name}"
            
            print(f"\n  🧪 [{idx}/{len(self.p0_experiment_records)}] 處理實驗 P0: {lot}", file=sys.stderr)
            print(f"      日期: {day_key} {slot_name}", file=sys.stderr)
            print(f"      人員: {staff}, 凍乾機: {dryer}, 滴定機: {machine}", file=sys.stderr)
            
            # 步驟 4: 預訂實驗 P0 資源
            try:
                # 4.1 預訂配藥人員
                self.resources.book_p0_experiment_person(staff, shift_key)
                
                # 4.2 預訂凍乾機
                self.resources.book_p0_experiment_dryer(dryer, day_key)
                
                # 4.3 預訂滴定機 (IVEK 或 Port)
                if "ivek" in machine:
                    self.resources.book_p0_experiment_ivek(day_key)
                    print(f"      ✓ 已預訂 IVEK", file=sys.stderr)
                    
                elif "port" in machine:
                    try:
                        port_numbers = re.findall(r'\d+', machine)
                        
                        if not port_numbers:
                            raise ValueError(f"無法從 '{machine}' 中提取 Port 編號")
                        
                        port_num = int(port_numbers[0])
                        
                        if not (1 <= port_num <= 12):
                            raise ValueError(f"Port 編號 {port_num} 超出範圍 (1-12)")
                        
                        self.resources.book_p0_experiment_port_pair(shift_key, port_num)
                        print(f"      ✓ 已預訂 Port{port_num} (含配對)", file=sys.stderr)
                        
                    except ValueError as ve:
                        print(f"      ⚠️ 警告: {ve}，將不預訂 Port", file=sys.stderr)
                    except Exception as e:
                        print(f"      ⚠️ 警告: 無法解析滴定機 '{machine}': {e}", file=sys.stderr)
                
                else:
                    print(f"      ⚠️ 警告: 滴定機類型 '{machine}' 無法識別", file=sys.stderr)
                
                # 步驟 5: 檢查是否與已排入的 P1/P2 衝突
                has_p1_p2_conflict = False
                conflict_details = []
                
                # 檢查人員衝突
                if shift_key in self.resources.person_usage:
                    p1_p5_people = self.resources.person_usage[shift_key]
                    if staff in p1_p5_people:
                        has_p1_p2_conflict = True
                        conflict_details.append(f"人員 {staff}")
                
                # 檢查凍乾機衝突
                if day_key in self.resources.dryer_usage:
                    p1_p5_dryers = self.resources.dryer_usage[day_key]
                    if dryer in p1_p5_dryers:
                        has_p1_p2_conflict = True
                        conflict_details.append(f"凍乾機 {dryer}")
                
                # 步驟 6: 更新 P0 資料標記
                p0_data['has_conflict'] = has_p1_p2_conflict
                p0_data['slot'] = slot_name
                p0_data['shift_key'] = shift_key
                p0_data['day_key'] = day_key
                
                if has_p1_p2_conflict:
                    print(f"      🔴 偵測到與 P1/P2 的資源衝突: {', '.join(conflict_details)}", file=sys.stderr)
                    print(f"      → Excel 將標記為淺紅色", file=sys.stderr)
                else:
                    print(f"      ✓ 無資源衝突", file=sys.stderr)
                
                print(f"      ✅ 實驗 P0 任務 {lot} 排程完成", file=sys.stderr)
                print(f"      → P3/P4/P5 將自動避開此任務的資源", file=sys.stderr)
                
            except Exception as e:
                print(f"      ❌ 實驗 P0 任務 {lot} 發生錯誤: {e}", file=sys.stderr)
                import traceback
                traceback.print_exc(file=sys.stderr)
                continue
        
        # 步驟 7: 排程完成統計
        print(f"\n{'='*70}", file=sys.stderr)
        print(f"🧪 實驗 P0 排程完成統計", file=sys.stderr)
        print(f"{'='*70}", file=sys.stderr)
        
        total = len(self.p0_experiment_records)
        has_conflict_count = sum(1 for p in self.p0_experiment_records if p.get('has_conflict', False))
        no_conflict_count = total - has_conflict_count
        
        print(f"  總計: {total} 個實驗 P0 任務", file=sys.stderr)
        print(f"  與 P1/P2 衝突 (淺紅色): {has_conflict_count} 個", file=sys.stderr)
        print(f"  無衝突 (標準格式): {no_conflict_count} 個", file=sys.stderr)
        print(f"  P3/P4/P5 將自動避開所有實驗 P0 的資源", file=sys.stderr)
        print(f"{'='*70}\n", file=sys.stderr)
    # ========================================
    # ✅ V9.7.14: 處理凍乾機為空任務
    # ========================================
    def handle_no_dryer_tasks(self):
        """
        處理凍乾機為空但有需求的任務
        規則: 例外 PN 排週三，其他排 Plan 前一天
        V9.9.2: 新增休假日檢查
        """
        print("\n" + "="*70, file=sys.stderr)
        print(f"🔥 Step 3.1: 處理凍乾機為空的任務（{CURRENT_VERSION}）", file=sys.stderr)
        print("="*70, file=sys.stderr)
        
        df_demand = self.data["demand"]
        df_prod_plan = self.db_constraints.get("production_Plan", pd.DataFrame())
        
        if df_prod_plan.empty:
            print("  ⚠️ production_Plan 為空", file=sys.stderr)
        
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        monday = (today + timedelta(days=-today.weekday(), weeks=0))
        w1_end = monday + timedelta(days=6)
        
        special_pn = "5714600102"
        this_wednesday = monday + timedelta(days=2)
        
        print(f"  W1 範圍: {monday.strftime('%Y-%m-%d')} ~ {w1_end.strftime('%Y-%m-%d')}", file=sys.stderr)
        print(f"  例外 PN ({special_pn}) 將排在: {this_wednesday.strftime('%Y-%m-%d')}", file=sys.stderr)
        
        date_cols_map = parse_date_columns(df_prod_plan)
        no_dryer_records_local = []
        
        for _, row in df_demand.iterrows():
            pn = str(row["料號"]).strip()
            
            if pn in self.scheduled_pns:
                continue
            
            w1_balance = row["Stock_plus_Dosing"] - row["第一周需求"]
            w1_shortage = abs(w1_balance) if w1_balance < 0 else 0
            
            if w1_shortage == 0:
                continue
            
            if pn not in self.constraints.index:
                continue
            
            constr_data = self.constraints.loc[pn]
            if isinstance(constr_data, pd.DataFrame):
                constr = constr_data.iloc[0]
            else:
                constr = constr_data
            
            dryer_value = str(constr.get("可用凍乾機", "")).strip()
            
            if dryer_value != "":
                continue
            
            print(f"  🔍 處理 {pn} (凍乾機為空, 需求量: {w1_shortage})", file=sys.stderr)
            
            schedule_date = None
            earliest_w1_date = None

            if pn not in df_prod_plan["PN"].values:
                if pn == special_pn:
                    print(f"    ⭐ {pn} (例外): 強制排入週三", file=sys.stderr)
                    schedule_date = this_wednesday
                else:
                    print(f"    ⚠️ {pn}: Plan 中無 W1 排程. 跳過.", file=sys.stderr)
                    continue
            else:
                pn_plan = df_prod_plan[df_prod_plan["PN"] == pn].iloc[0]
                w1_dates = []
                for col, col_date in date_cols_map.items():
                    if monday <= col_date <= w1_end:
                        qty = pn_plan.get(col, 0)
                        if pd.notna(qty) and float(qty) > 0:
                            w1_dates.append(col_date)
                
                if w1_dates:
                    earliest_w1_date = min(w1_dates)

                if earliest_w1_date is None:
                    if pn == special_pn:
                        print(f"    ⭐ {pn} (例外): 強制排入週三", file=sys.stderr)
                        schedule_date = this_wednesday
                    else:
                        print(f"    ⚠️ {pn}: Plan 中無 W1 排程. 跳過.", file=sys.stderr)
                        continue
                else:
                    schedule_date = earliest_w1_date - timedelta(days=1)
                    print(f"      {pn}: 排入前一天", file=sys.stderr)

            if schedule_date < monday:
                print(f"    ⚠️ {pn}: 排程日早於本週一，修正為週一", file=sys.stderr)
                schedule_date = monday
            
            # ========================================
            # V9.9.2: 檢查是否為休假日
            # ========================================
            if is_holiday(schedule_date):
                print(f"    ⚠️ {pn}: 原排程日 {schedule_date.strftime('%Y-%m-%d')} 為休假日", file=sys.stderr)
                
                # 嘗試往後找第一個非休假日
                max_search_days = 7
                found_alternative = False
                
                for offset in range(1, max_search_days + 1):
                    alternative_date = schedule_date + timedelta(days=offset)
                    
                    # 不超過本週末
                    if alternative_date > w1_end:
                        break
                    
                    if not is_holiday(alternative_date):
                        schedule_date = alternative_date
                        found_alternative = True
                        print(f"    ✅ {pn}: 調整至 {schedule_date.strftime('%Y-%m-%d')}", file=sys.stderr)
                        break
                
                if not found_alternative:
                    # 嘗試往前找
                    for offset in range(1, max_search_days + 1):
                        alternative_date = schedule_date - timedelta(days=offset)
                        
                        # 不早於週一
                        if alternative_date < monday:
                            break
                        
                        if not is_holiday(alternative_date):
                            schedule_date = alternative_date
                            found_alternative = True
                            print(f"    ✅ {pn}: 調整至 {schedule_date.strftime('%Y-%m-%d')}", file=sys.stderr)
                            break
                
                if not found_alternative:
                    print(f"    ❌ {pn}: 無法找到合適的非休假日，跳過", file=sys.stderr)
                    continue
            
            marker_name = row.get("BDC_Marker_Name", "")
            if not marker_name:
                print(f"    ⚠️ {pn} 無 Marker Name. 跳過.", file=sys.stderr)
                continue
            
            person_result = self._find_person_for_no_dryer(constr, schedule_date)
            
            if person_result is None:
                print(f"    ⚠️ {pn} 在 {schedule_date.strftime('%Y-%m-%d')} 無可用配藥人員. 跳過.", file=sys.stderr)
                continue
            
            person, is_single_slot = person_result
            
            print(f"    ✅ {pn} 排入 {schedule_date.strftime('%Y-%m-%d')}", file=sys.stderr)
            
            no_dryer_records_local.append({
                "日期": schedule_date.strftime("%Y-%m-%d"),
                "marker": marker_name,
                "lot": pn,
                "凍乾機台": "",
                "數量": w1_shortage,
                "配藥同仁": person,
                "RD給藥時間": "",
                "預計滴定時間": "",
                "預計結束": "",
                "班次": "PM",
                "滴定機": "",
                "is_no_dryer": True
            })
            
            self.scheduled_pns.add(pn)
        
        print(f"\n  ✅ 凍乾機為空任務完成 ({len(no_dryer_records_local)} 筆)", file=sys.stderr)
        if no_dryer_records_local:
            print("    這些配藥人員已占用整天，主排程將自動避開\n", file=sys.stderr)
        
        self.no_dryer_records.extend(no_dryer_records_local)

    def _find_person_for_no_dryer(self, constr, schedule_date):
        """為凍乾機為空的任務找配藥人員（占用全天）"""
        day_key = schedule_date.strftime('%Y-%m-%d')
        shift_key_am = f"{day_key}_AM"
        shift_key_pm = f"{day_key}_PM"
        
        all_people = set()
        is_single_slot = constr.get("V9_Single_Slot", False)
        
        for key in ["配藥人-1", "配藥人-2", "配藥人-3"]:
            person = constr.get(key)
            if person and str(person).strip():
                all_people.add(str(person).strip().lower()) # ✅ 修正：使用小寫
        
        all_people.discard("")
        
        if not all_people:
            print(f"        🔴 [{day_key}] 批次無可配藥人員", file=sys.stderr)
            return None
        
        for person in sorted(all_people):
            can_use_am, reason_am = self.resources.can_use_person(
                person, shift_key_am, day_key, is_single_slot
            )
            
            can_use_pm, reason_pm = self.resources.can_use_person(
                person, shift_key_pm, day_key, is_single_slot
            )
            
            if can_use_am and can_use_pm:
                self.resources.book_person(person, shift_key_am, day_key, is_single_slot)
                self.resources.book_person(person, shift_key_pm, day_key, is_single_slot)
                
                print(f"        🟢 [{day_key}] 選中人員: {person} (占用全天)", file=sys.stderr)
                return person, is_single_slot
            else:
                print(f"        🔴 [{day_key}] {person} 不可用", file=sys.stderr)
        
        print(f"        🔴 [{day_key}] 人員全被佔用", file=sys.stderr)
        return None
    
    """
    ==============================================
    V9.9.4: Scheduler.run() 完整修改版
    批次自動錯開 RD給藥時間
    ==============================================
    """

    def run(self):
        """V9.9.4: 主排程流程（批次自動錯開 RD給藥時間）"""
        print(f"\nStep 3: 開始排程 ({CURRENT_VERSION})...\n", file=sys.stderr)
        
        # ❌ 移除錯誤的日期計算
        # today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) 
        # monday = (today + timedelta(days=-today.weekday(), weeks=0)) 
        
        # ✅ 新增：使用 main 函數設定的全域變數
        global SCHEDULE_START_DATE
        if SCHEDULE_START_DATE is None:
             # 備用邏輯，以防萬一
            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            monday = (today + timedelta(days=-today.weekday(), weeks=0))
            print(f"  ⚠️ 警告: SCHEDULE_START_DATE 未設定，使用今天 {today.strftime('%Y-%m-%d')} 推算週一", file=sys.stderr)
        else:
            monday = SCHEDULE_START_DATE
        
        print(f"  排程起始: {monday.strftime('%Y-%m-%d')}\n", file=sys.stderr)
        
        saturday_date = monday + timedelta(days=5) 
        print(f"  P4/P5 截止日期 (週五): {(monday + timedelta(days=4)).strftime('%Y-%m-%d')}", file=sys.stderr)

        # 顯示休假日資訊
        if HOLIDAYS:
            print(f"\n  ⚠️   休假日設定:", file=sys.stderr)
            for h in sorted(HOLIDAYS):
                print(f"      {h.strftime('%Y-%m-%d (%A)')}", file=sys.stderr)
            print(file=sys.stderr)

        self.shift_port_counter = {}

        # ========================================
        # V9.9.4: 移除固定時間槽生成
        # ========================================
        # 改為在排程時動態分配時間
        
        # AM 和 PM 的起始時間
        AM_START = time(10, 30)
        PM_START = time(15, 30)
        PM_END = time(17, 30)
        
        # 每個批次間隔時間（分鐘）
        BATCH_INTERVAL_MINUTES = 30

        # 為每天準備資源追蹤
        for day_index in range(self.days_to_schedule + 1):
            if day_index == SCHEDULING_DAYS:
                w1_tasks_left = any(t[0] == 1 and t[2] not in self.scheduled_pns for t in self.task_queue)
                if not ALLOW_EXTEND_TO_6_DAYS or not w1_tasks_left:
                    break
                print("⚠️ 延長至第 6 天", file=sys.stderr)

            current_day = monday + timedelta(days=day_index)
            
            # V9.9.2: 檢查是否為休假日
            if is_holiday(current_day):
                print(f"  🗓️   跳過休假日: {current_day.strftime('%Y-%m-%d (%A)')}", file=sys.stderr)
                continue
            
            day_key = current_day.strftime('%Y-%m-%d')
            
            self.resources.reset_day(day_key)
            
            for slot_name in ["AM", "PM"]:
                shift_key = f"{day_key}_{slot_name}"
                self.resources.reset_shift(shift_key)
                self.shift_port_counter[shift_key] = 1

        print(f"生成 {self.days_to_schedule} 個排程日（已排除休假日）\n", file=sys.stderr)

        # ========================================
        # V9.9.2: P0 處理（內部已包含休假日檢查，若衝突會直接終止）
        # ========================================
        try:
            self.handle_p0_orders()
        except SystemExit as e:
            # P0 與休假日衝突，排程已終止
            print(f"\n{str(e)}", file=sys.stderr)
            return

        # ========================================
        # 處理凍乾機為空的任務（已包含休假日檢查）
        # ========================================
        self.handle_no_dryer_tasks()

        # ========================================
        # 主排程 (P1-P5)
        # ========================================
        print("="*70, file=sys.stderr)
        print("📋 Step 3.2: 開始主排程 (P1-P5)", file=sys.stderr)
        print("="*70, file=sys.stderr)

        print("計算任務優先級...", file=sys.stderr)
        task_priorities = []
        for task in self.task_queue:
            pn = task[2]
            
            if (pn in self.scheduled_pns) and (task[0] != 5):
                continue
            
            batch_tasks = self.find_batch(task)
            if not batch_tasks:
                continue
            
            constraints_list = self._normalize_constraints(batch_tasks)
            if not constraints_list:
                continue
            
            person_count, people_list = self.calculate_person_priority(constraints_list)
            
            task_priorities.append((task[0], task[1], person_count, task))
        
        # ✅ Rule 2: 按藥名 (x[3][5]) 排序
        task_priorities.sort(key=lambda x: (x[0], x[3][5], x[1][0], x[1][1], x[1][2], x[2]))
        
        # ✅ P6/P7 任務清單 (按庫存排序)
        p6_p7_fill_tasks_sorted = sorted(
            self.p6_p7_fill_tasks, 
            key=lambda x: (x[0], x[1][2]) # 1. P6 優先 2. 庫存 (stock_level) 
        )
        
        print(f"任務優先級排序完成 (共 {len(task_priorities)} 個任務)\n", file=sys.stderr)

        shift_stats = {}
        scheduled_count = 0
        failed_reasons = {}
        
        # === V9.9.1: 新增 P2 完成標記 ===
        p2_completed = False
        
        # ========================================
        # V9.9.2 (修正版): P1 任務追蹤
        # ========================================
        p1_total = sum(1 for t in task_priorities if t[0] == 1)
        # ✅ 新增：建立一個 P1 任務的 set, (group_name, pn)
        p1_tasks_set = set((t[3][5], t[3][2]) for t in task_priorities if t[0] == 1)
        p1_scheduled_count = 0
        
        # ========================================
        # V9.9.4: 追蹤每天每個班次已分配的批次時間
        # ========================================
        day_batch_times = {}  # {day_key: {"AM": [time1, time2, ...], "PM": [...]}}
        
        # ========================================
        # V9.9.4: 主排程循環（動態時間分配）
        # ========================================
        for week, subp, person_count, task in task_priorities:
            pn = task[2]
            
            # === V9.9.1: 在 P2 完成後插入實驗 P0 ===
            if not p2_completed and week >= 3:
                p2_completed = True
                if self.p0_experiment_records:
                    print("\n" + "="*70, file=sys.stderr)
                    print("🧪 插入 P0 (實驗) 任務 (P2 後排程)", file=sys.stderr)
                    print("="*70, file=sys.stderr)
                    self.schedule_p0_experiments([])  # 不需要 all_slots
            
            if (pn in self.scheduled_pns) and (week != 5):
                continue 

            batch_tasks = self.find_batch(task)
            if not batch_tasks:
                failed_reasons[pn] = "find_batch 失敗"
                continue 

            batch_pns = [t[2] for t in batch_tasks]
            batch_name = batch_tasks[0][5]
            
            constraints_list = self._normalize_constraints(batch_tasks)
            if not constraints_list:
                failed_reasons[pn] = "normalize_constraints 失敗"
                continue
            
            slot_prefs = constraints_list[0].get("V9_Slot_List", ["AM", "PM"])
            
            _, people_list = self.calculate_person_priority(constraints_list)
            print(f"  嘗試排程: {batch_name} ({batch_pns}) - (P{week})", file=sys.stderr)
            
            fail_reason = None
            found_slot = False
            
            # ========================================
            # V9.9.4: 動態尋找可用時間
            # ========================================
            
            # 遍歷排程天數
            for day_offset in range(self.days_to_schedule + 1):
                if found_slot:
                    break
                        
                current_day = monday + timedelta(days=day_offset)
                
                # ========================================
                # V9.9.2: 再次確認不是休假日（雙重保險）
                # ========================================
                if is_holiday(current_day):
                    continue
                
                # P4/P5 不排週六
                is_p4_or_p5 = (week >= 4)
                if is_p4_or_p5 and current_day.date() == saturday_date.date():
                    if not fail_reason:
                        fail_reason = "P4/P5 不排週六"
                    continue
                
                day_key = current_day.strftime('%Y-%m-%d')
                
                # 初始化該天的批次時間追蹤
                if day_key not in day_batch_times:
                    day_batch_times[day_key] = {"AM": [], "PM": []}
                
                # 嘗試兩個班次
                for slot_name in ["AM", "PM"]:
                    if found_slot:
                        break
                    
                    # 檢查班次偏好
                    if slot_name not in slot_prefs:
                        continue
                    
                    # ========================================
                    # V9.9.11: 檢查週五排程限制
                    # ========================================
                    can_schedule, friday_reason = self.can_schedule_on_friday(
                        constraints_list, current_day
                    )
                    
                    if not can_schedule:
                        if not fail_reason:
                            fail_reason = friday_reason
                        continue  # 跳過週五

                    # ========================================
                    # V9.9.10: 根据交藥時間决定 RD給藥時間
                    # ========================================
                    next_time = self.get_delivery_time_from_constraints(
                        constraints_list, slot_name, current_day, day_batch_times, day_key
                    )

                    if next_time is None:
                        if not fail_reason:
                            fail_reason = f"{day_key} {slot_name} 时段已滿"
                        continue

                    # 组合完整时间
                    delivery_dt = datetime.combine(current_day, next_time)
                    shift_key = f"{day_key}_{slot_name}"
                    
                    if shift_key not in shift_stats:
                        shift_stats[shift_key] = {"scheduled": 0, "failed": 0}
                    
                    # 檢查資源可用性
                    resources = self.check_availability(
                        batch_tasks, 
                        constraints_list, 
                        delivery_dt, 
                        slot_name, 
                        priority=week
                    )
                    
                    if not resources:
                        shift_stats[shift_key]["failed"] += 1
                        if not fail_reason:
                            fail_reason = f"check_availability 失敗"
                        continue
                    
                    # ========================================
                    # V9.9.12: 檢查預計結束時間是否在禁止時段
                    # ========================================
                    final_end_dt = resources["times"]["final_end_dt"]
                    is_forbidden, forbidden_reason = self.is_forbidden_end_time(final_end_dt)
                    
                    if is_forbidden:
                        if not fail_reason:
                            fail_reason = forbidden_reason
                        continue  # 跳過此時段

                    # 記錄排程
                    initial_len = len(self.schedule_df) if hasattr(self, 'schedule_df') and not self.schedule_df.empty else 0
                    
                    self.book_and_record(
                        resources, 
                        batch_tasks, 
                        delivery_dt.date(), 
                        slot_name, 
                        priority=week
                    )
                    
                    final_len = len(self.schedule_df) if hasattr(self, 'schedule_df') and not self.schedule_df.empty else 0
                    
                    if final_len > initial_len:
                        # 成功排程
                        self.book_resources(resources, delivery_dt, slot_name, priority=week)
                        # ========================================
                        # ✅ 新增：記錄 CK/PHOS/K 排程日期 (污染凍乾機)
                        # ========================================
                        scheduled_pns_in_batch = set(t[2] for t in batch_tasks)
                        is_blocking_batch = any(pn in self.RULE_BLOCKING_PNs for pn in scheduled_pns_in_batch)
                                
                        if is_blocking_batch:
                            dryer_id = resources.get("dryer")
                            current_date = delivery_dt.date()
                            blocked_until_date = current_date + timedelta(days=self.RULE_BLOCK_DAYS)
                                    
                                    # 更新或寫入封鎖記錄 (取較晚的日期)
                            if (dryer_id in self.dryer_block_until_date and 
                                blocked_until_date > self.dryer_block_until_date[dryer_id]):
                                self.dryer_block_until_date[dryer_id] = blocked_until_date
                            elif dryer_id not in self.dryer_block_until_date:
                                self.dryer_block_until_date[dryer_id] = blocked_until_date
                                        
                            print(f"  ✅ 記錄 {batch_name} 排程於 {current_date.strftime('%Y-%m-%d')}, 凍乾機 {dryer_id} 封鎖至 {blocked_until_date.strftime('%Y-%m-%d')}", file=sys.stderr)        
                        # ========================================
                        # V9.9.4: 記錄已分配的時間
                        # ========================================
                        day_batch_times[day_key][slot_name].append(next_time)
                        
                        if week != 5:
                            self.book_resources(resources, delivery_dt, slot_name, priority=week)
                                
                            # ========================================
                            # V9.9.2 (修正版): 追蹤 P1 排程成功數
                            # ========================================
                            for scheduled_task in batch_tasks:
                                scheduled_key = (scheduled_task[5], scheduled_task[2]) # (group_name, pn)
                                if scheduled_key in p1_tasks_set:
                                    p1_scheduled_count += 1
                                    p1_tasks_set.remove(scheduled_key) # 避免重複計算
                            
                        scheduled_count += 1
                        shift_stats[shift_key]["scheduled"] += 1
                        
                        print(f"  ✅ [{delivery_dt.strftime('%Y-%m-%d %H:%M')} {slot_name}] (P{week}) {batch_name}", file=sys.stderr)
                        
                        found_slot = True
                        break
                    else:
                        if not fail_reason or "Port" not in fail_reason:
                            fail_reason = f"Port 不足"
                
            if not found_slot:
                failed_reasons[pn] = fail_reason or f"所有時段均失敗"
                print(f"    ❌ (P{week}) {batch_name} - {fail_reason}", file=sys.stderr)
        
        # ========================================
        # ✅ 新增 (Rule 3)：P6/P7 填補空 Port 任務
        # ========================================
        print("\n" + "="*70, file=sys.stderr)
        print("🧩 Step 3.3: 開始 P6/P7 填補空 Port 任務", file=sys.stderr)
        print("="*70, file=sys.stderr)
        self.fill_gaps(monday, day_batch_times, p6_p7_fill_tasks_sorted)

        # ========================================
        # V9.9.2 (修正版): 檢查 P1 任務是否全部完成
        # ========================================
        if p1_total > 0 and p1_scheduled_count < p1_total:
            print("\n" + "="*70, file=sys.stderr)
            print("❌ 排程終止：P1 任務未全部滿足", file=sys.stderr)
            print("="*70, file=sys.stderr)
            print(f"P1 任務統計:", file=sys.stderr)
            print(f"  總計: {p1_total} 個任務", file=sys.stderr)
            print(f"  已排程: {p1_scheduled_count} 個任務", file=sys.stderr)
            print(f"  未排程: {p1_total - p1_scheduled_count} 個任務", file=sys.stderr)
            
            # ✅ 修改：從 p1_tasks_set (未被移除的) 來顯示
            for group_name, pn in p1_tasks_set:
                reason = failed_reasons.get(pn, "未知原因")
                print(f"  • {group_name} (PN: {pn}) - 失敗原因: {reason}", file=sys.stderr)
            
            print("\n可能的解決方案:", file=sys.stderr)
            print("  1. 調整休假日設定（若有設定休假日）", file=sys.stderr)
            print("  2. 調整 P1 任務的配藥限制（人員/凍乾機/時段）", file=sys.stderr)
            print("  3. 延長排程天數", file=sys.stderr)
            print("  4. 減少 P0 任務數量", file=sys.stderr)
            print("="*70, file=sys.stderr)
            return  # 終止排程

        # 統計
        print(f"\n{'='*70}", file=sys.stderr)
        print(f"📊 排程統計", file=sys.stderr)
        print(f"{'='*70}", file=sys.stderr)
        print(f"  凍乾機為空: {len(self.no_dryer_records)} 筆", file=sys.stderr)
        print(f"  P0 (非實驗): {len(self.p0_strict_records)} 筆", file=sys.stderr)
        print(f"  P0 (實驗): {len(self.p0_experiment_records)} 筆", file=sys.stderr)
        print(f"  P1-P5 成功: {scheduled_count} 批次", file=sys.stderr)
        print(f"    - P1: {p1_scheduled_count}/{p1_total} {'✅' if p1_scheduled_count == p1_total else '❌'}", file=sys.stderr)
        
        unmet_pns = [t[2] for t in self.task_queue if t[0] < 4 and t[2] not in self.scheduled_pns]
        print(f"  未排入 PN (P1-P3): {len(unmet_pns)}", file=sys.stderr)
        
        # ========================================
        # V9.9.4: 顯示時間分配統計
        # ========================================
        print(f"\n{'='*70}", file=sys.stderr)
        print(f"⏰ RD給藥時間分配統計", file=sys.stderr)
        print(f"{'='*70}", file=sys.stderr)
        for day_key in sorted(day_batch_times.keys()):
            print(f"\n  {day_key}:", file=sys.stderr)
            for slot_name in ["AM", "PM"]:
                times = sorted(day_batch_times[day_key][slot_name])
                if times:
                    time_strs = [t.strftime('%H:%M') for t in times]
                    print(f"    {slot_name}: {', '.join(time_strs)} (共 {len(times)} 批)", file=sys.stderr)
        
        print(f"\n✅ 所有排程完成", file=sys.stderr)

    # ==========================================================
    # ✅ 新增 (Rule 3)：P6/P7 填補空 Port 函數
    # ==========================================================
    def fill_gaps(self, monday, day_batch_times, p6_p7_tasks):
        """
        遍歷所有 Port (1-12)，找出 P1-P5 留下的空缺,
        並嘗試從 p6_p7_tasks 中找到任務來填補。
        """
        
        # 1. 建立一個 Port 使用狀況的字典
        port_usage = {}
        if not self.schedule_df.empty:
            for _, row in self.schedule_df.iterrows():
                if row.get('ports_list'):
                    day_key = row['日期']
                    slot_name = row['班次']
                    ports = str(row['ports_list']).split(',')
                    for p in ports:
                        if p.isdigit():
                            port_key = f"{day_key}_{slot_name}_Port{p}"
                            port_usage[port_key] = True

        # 2. 遍歷 P6/P7 任務 (已按庫存排過序)
        tasks_to_remove = [] # 記錄已排入的 P6/P7 任務
        
        for task_index, task in enumerate(p6_p7_tasks):
            week, subp, pn, short_qty, tag, group_name, marker_name, prod_qty = task
            
            num_ports_str = "1"
            if pn in self.constraints.index:
                constr = self.constraints.loc[pn]
                if isinstance(constr, pd.DataFrame): constr = constr.iloc[0]
                num_ports_str = str(constr.get("Port數", "1")).strip()
            
            try:
                ports_needed = int(num_ports_str)
            except:
                ports_needed = 1 
            
            if ports_needed not in [1, 2]:
                continue 
            
            print(f"  🧩 嘗試填補 P{week} 任務: {group_name} (需 {ports_needed} Port)", file=sys.stderr)

            # 3. 遍歷所有天和所有 Port，尋找空缺
            found_slot = False
            for day_offset in range(self.days_to_schedule):
                if found_slot: break
                current_day = monday + timedelta(days=day_offset)
                
                if is_holiday(current_day):
                    continue
                
                day_key = current_day.strftime('%Y-%m-%d')
                
                for slot_name in ["AM", "PM"]:
                    if found_slot: break
                    shift_key = f"{day_key}_{slot_name}"
                    
                    # 4. 尋找 'ports_needed' 數量的連續空 Port
                    start_port = -1
                    # ✅ (Rule 1) 允許 1-Port
                    for p_idx in range(1, MAX_PORTS + 2 - ports_needed):
                        is_available = True
                        ports_to_check = []
                        for i in range(ports_needed):
                            port_num = p_idx + i
                            if port_num > MAX_PORTS: # 超出 12
                                is_available = False
                                break
                            port_key = f"{shift_key}_Port{port_num}"
                            ports_to_check.append(port_num)
                            
                            if port_usage.get(port_key, False):
                                is_available = False 
                                break
                        
                        if is_available:
                            start_port = p_idx
                            break
                    
                    if start_port != -1:
                        # 5. 找到空 Port，嘗試排程
                        rd_time = time(10, 30) if slot_name == "AM" else time(15, 30)
                        delivery_dt = datetime.combine(current_day, rd_time)
                        
                        batch_tasks = [task] 
                        constraints_list = self._normalize_constraints(batch_tasks)

                        # 6. 檢查其他資源 (人員, 凍乾機, 凌晨規則)
                        resources = self.check_availability(batch_tasks, constraints_list, delivery_dt, slot_name, priority=week)
                        
                        if resources:
                            final_end_dt = resources["times"]["final_end_dt"]
                            is_forbidden, _ = self.is_forbidden_end_time(final_end_dt)
                            
                            if not is_forbidden:
                                # 8. 排程成功！
                                print(f"    ✅ 成功填補: {group_name} 於 {shift_key} Port {start_port}-{start_port + ports_needed - 1}", file=sys.stderr)
                                
                                self.book_resources(resources, delivery_dt, slot_name, priority=week)
                                
                                port_str = ",".join([f"Port{p}" for p in range(start_port, start_port + ports_needed)])
                                ports_list_str = ",".join([str(p) for p in range(start_port, start_port + ports_needed)])
                                
                                for t in batch_tasks:
                                    row = {
                                        "日期": current_day.strftime("%Y-%m-%d"),
                                        "marker": t[6], "滴定機": port_str,
                                        "凍乾機台": resources.get("dryer", "N/A"),
                                        "配藥同仁": resources.get("person", ""),
                                        "RD給藥時間": "N/A (填補)", "預計滴定時間": "N/A (填補)",
                                        "預計結束": final_end_dt.strftime("%H:%M"), "凍乾時間": f"{resources['times']['freeze_duration_hr']:.1f} hr",
                                        "數量": t[7], "lot": t[2], "班次": slot_name,
                                        "ports_list": ports_list_str
                                    }
                                    
                                    if any(pn in self.RULE_BLOCKING_PNs for pn in [t[2]]):
                                        dryer_id = resources.get("dryer")
                                        blocked_until_date = current_day.date() + timedelta(days=self.RULE_BLOCK_DAYS)
                                        if (dryer_id in self.dryer_block_until_date and blocked_until_date > self.dryer_block_until_date[dryer_id]):
                                            self.dryer_block_until_date[dryer_id] = blocked_until_date
                                        elif dryer_id not in self.dryer_block_until_date:
                                            self.dryer_block_until_date[dryer_id] = blocked_until_date
                                        print(f"      ✅ 記錄 {t[5]} 排程於 {current_day.strftime('%Y-%m-%d')}, 凍乾機 {dryer_id} 封鎖至 {blocked_until_date.strftime('%Y-%m-%d')}", file=sys.stderr)

                                    new_df = pd.DataFrame([row])
                                    if not hasattr(self, "schedule_df") or self.schedule_df.empty:
                                        self.schedule_df = new_df
                                    else:
                                        self.schedule_df = pd.concat([self.schedule_df, new_df], ignore_index=True)

                                for p in range(start_port, start_port + ports_needed):
                                    port_usage[f"{shift_key}_Port{p}"] = True
                                
                                found_slot = True
                                tasks_to_remove.append(task_index) # 標記此任務已被排程
                                break 
        
        # 9. 從 P6/P7 佇列中移除已排程的任務 (避免 P7 再次排程)
        # (倒序刪除, 避免 index 錯亂)
        for task_index in sorted(tasks_to_remove, reverse=True):
            del p6_p7_tasks[task_index]
        
        print(f"  🧩 填補空缺完成, {len(tasks_to_remove)} 個任務已排入, {len(p6_p7_tasks)} 個任務剩餘", file=sys.stderr)


    def book_and_record(self, resources, batch_tasks, current_day, slot_name, priority=None):
        """V9.9.1: Port 滿時不記錄，P3-P5 自動跳過實驗 P0 Port"""
        """✅ 修改：新增 IVEK 規則"""
        import pandas as pd

        times = resources["times"]
        dryer = resources.get("dryer", "N/A") 
        person = resources.get("person", "")
        freeze_duration = times.get("freeze_duration_hr", "N/A")
        
        day_key = current_day.strftime("%Y-%m-%d")
        shift_key = f"{day_key}_{slot_name}"
        
        # ✅ 新增：取得 num_ports_str (e.g., "IVEK", "4", "2")
        num_ports_str = resources.get("num_ports_str", "2")
        is_ivek_task = (num_ports_str.upper() == "IVEK")
        
        rows = []
        
        if is_ivek_task:
            # ========================================
            # ✅ 新增：IVEK 任務排程邏輯
            # ========================================
            print(f"    → {batch_tasks[0][5]}: IVEK 任務", file=sys.stderr)
            
            for task in batch_tasks:
                week, subp, pn, short_qty, tag, group_name, marker_name, prod_qty = task
                
                rows.append({
                    "日期": current_day.strftime("%Y-%m-%d"),
                    "marker": marker_name, 
                    "滴定機": "IVEK", # ✅ 固定寫入 IVEK
                    "凍乾機台": dryer,
                    "配藥同仁": person,
                    "RD給藥時間": times["delivery_dt"].strftime("%H:%M"), 
                    "預計滴定時間": times["dosing_start_dt"].strftime("%H:%M"),
                    "預計結束": times["final_end_dt"].strftime("%H:%M"),
                    "預冷時間": (times["dosing_end_dt"] + timedelta(hours=2)).strftime("%H:%M"), 
                    "凍乾時間": f"{freeze_duration:.1f} hr", 
                    "收藥時間": times["final_end_dt"].strftime("%H:%M"), 
                    "數量": prod_qty,
                    "lot": pn,
                    "班次": slot_name,
                    "ports_list": "IVEK" # ✅ 標記為 IVEK
                })
            
            # (IVEK 任務不佔用 Port 1-12, 不更新 self.shift_port_counter)

        else:
            # ========================================
            # ✅ 現有：Port 1-12 任務排程邏輯
            # ========================================
            num_ports_per_pn = resources.get("num_ports_per_pn", 2)
            
            if shift_key not in self.shift_port_counter:
                self.shift_port_counter[shift_key] = 1
            
            port_start = self.shift_port_counter[shift_key]

            # === V9.9: P1-P5 Port 分配自動跳過非實驗 P0 ===
            while self.resources.is_p0_strict_port_booked(shift_key, port_start):
                print(f"        ⚠️ P0 (非實驗) 佔用: Port{port_start}，自動跳過", file=sys.stderr)
                port_start += 1
            
            # === V9.9.1: P3-P5 Port 分配自動跳過實驗 P0 ===
            if priority is not None and priority >= 3: # P3, P4, P5
                while self.resources.is_p0_experiment_port_booked(shift_key, port_start):
                    print(f"        ⚠️ P0 (實驗) 佔用: Port{port_start}，P{priority} 自動跳過", file=sys.stderr)
                    port_start += 1
            
            self.shift_port_counter[shift_key] = port_start
            
            # 預先計算總 ports 需求
            total_ports_needed = 0
            for task in batch_tasks:
                actual_ports = num_ports_per_pn # ✅ 移除 MAX_PORTS_PER_PN 限制
                total_ports_needed += actual_ports
            
            # Port 不足時直接返回
            if port_start + total_ports_needed - 1 > MAX_PORTS:
                print(f"    ⚠️ {shift_key} Port 不足 (需要 {total_ports_needed}, 剩餘 {13 - port_start})", file=sys.stderr)
                return
            
            # Port 充足，開始分配
            current_port_idx = port_start
            
            for task in batch_tasks:
                week, subp, pn, short_qty, tag, group_name, marker_name, prod_qty = task
                
                actual_ports = min(num_ports_per_pn, MAX_PORTS_PER_PN)
                ports_for_this_pn = list(range(current_port_idx, current_port_idx + actual_ports))
                current_port_idx += actual_ports
                
                port_str = ",".join([f"Port{p}" for p in ports_for_this_pn])
                
                rows.append({
                    "日期": current_day.strftime("%Y-%m-%d"),
                    "marker": marker_name, 
                    "滴定機": port_str,
                    "凍乾機台": dryer,
                    "配藥同仁": person,
                    "RD給藥時間": times["delivery_dt"].strftime("%H:%M"), 
                    "預計滴定時間": times["dosing_start_dt"].strftime("%H:%M"),
                    "預計結束": times["final_end_dt"].strftime("%H:%M"),
                    "預冷時間": (times["dosing_end_dt"] + timedelta(hours=2)).strftime("%H:%M"), 
                    "凍乾時間": f"{freeze_duration:.1f} hr", 
                    "收藥時間": times["final_end_dt"].strftime("%H:%M"), 
                    "數量": prod_qty,
                    "lot": pn,
                    "班次": slot_name,
                    "ports_list": ",".join(map(str, ports_for_this_pn))
                })
            
            self.shift_port_counter[shift_key] = current_port_idx
            print(f"    → {batch_tasks[0][5]}: {len(rows)} PN, Port {port_start}-{current_port_idx-1}", file=sys.stderr)

        
        # === 共通邏輯：寫入 DataFrame ===
        new_df = pd.DataFrame(rows)
        if not hasattr(self, "schedule_df") or self.schedule_df.empty:
            self.schedule_df = new_df
        else:
            self.schedule_df = pd.concat([self.schedule_df, new_df], ignore_index=True)

        
    
    def _generate_dry_run_preview(self):
        """生成 Dry Run 預覽資料（JSON 格式）"""
        preview_data = []
        
        # P1-P5 排程
        if not self.schedule_df.empty:
            df_sorted = self.schedule_df.sort_values(by=["日期", "班次", "滴定機"])
            
            for _, row in df_sorted.iterrows():
                preview_data.append({
                    "date": row.get("日期", ""),
                    "titrate": row.get("滴定機", ""),
                    "freeze": row.get("凍乾機台", ""),
                    "pn": row.get("lot", ""),
                    "qty": row.get("數量", 0),
                    "staff": row.get("配藥同仁", ""),
                    "Name": row.get("marker", "")
                })
        
        # P0 (非實驗)
        for p0_data in self.p0_strict_records:
            task_date = p0_data.get('P0_Parsed_Date')
            if task_date:
                preview_data.append({
                    "date": task_date.strftime('%Y-%m-%d'),
                    "titrate": str(p0_data.get("滴定機", "")).strip(),
                    "freeze": str(p0_data.get("凍乾機台", "")).strip(),
                    "pn": str(p0_data.get("Lot", "")).strip(),
                    "qty": p0_data.get("數量", 0),
                    "staff": str(p0_data.get("配藥同仁", "")).strip(),
                    "Name": str(p0_data.get("Marker", "")).strip()
                })
        
        # P0 (實驗)
        for p0_data in self.p0_experiment_records:
            task_date = p0_data.get('P0_Parsed_Date')
            if task_date:
                preview_data.append({
                    "date": task_date.strftime('%Y-%m-%d'),
                    "titrate": str(p0_data.get("滴定機", "")).strip(),
                    "freeze": str(p0_data.get("凍乾機台", "")).strip(),
                    "pn": str(p0_data.get("Lot", "")).strip(),
                    "qty": p0_data.get("數量", 0),
                    "staff": str(p0_data.get("配藥同仁", "")).strip(),
                    "Name": str(p0_data.get("Marker", "")).strip()
                })
        
        # 凍乾機為空
        for record in self.no_dryer_records:
            preview_data.append({
                "date": record.get("日期", ""),
                "titrate": "",
                "freeze": "",
                "pn": record.get("lot", ""),
                "qty": record.get("數量", 0),
                "staff": record.get("配藥同仁", ""),
                "Name": record.get("marker", "")
            })
        
        return preview_data
    
    
    def save_records_to_excel(self, outdir):
        """
        V9.9.13: 固定格式 Excel 輸出
        固定結構：
        - IVEK: 2 行（固定顯示 IVEK 1-2）
        - 空行
        - AM Port1-12: 12 行（固定顯示）
        - 空行
        - PM Port1-12: 12 行（固定顯示）
        - 空行
        - no_dryer: 動態行數
        
        格式規則：
        - 滴定機欄位固定顯示（無論是否有排程）
        - no_dryer 背景淺黃色 (FFFFE0)
        - P0 實驗衝突背景淺紅色 (FFE0E0)
        """
        print("\n" + "="*70, file=sys.stderr)
        print(f"📊 Step 4: 輸出固定格式 Excel ({CURRENT_VERSION})", file=sys.stderr)
        print("="*70, file=sys.stderr)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(outdir, f"排程結果_{timestamp}.xlsx")

        # ========================================
        # 統一收集所有記錄
        # ========================================
        all_records = []
        
        global BATCH_START_NUMBER
        batch_start = BATCH_START_NUMBER
        work_order_counter = batch_start
        pn_work_order_map = {}
        pn_week_port_counter = {}
        
        # 1. 處理 P1-P5
        if not self.schedule_df.empty:
            print(f"  收集 P1-P5 排程: {len(self.schedule_df)} 筆", file=sys.stderr)
            
            df_sorted = self.schedule_df.sort_values(by=["日期", "班次", "滴定機"])
            
            for _, row in df_sorted.iterrows():
                pn = row.get("lot", "")
                date_str = row.get("日期", "")
                ports_list_str = row.get("ports_list", "")
                
                if ports_list_str:
                    ports = [int(p.strip()) for p in str(ports_list_str).split(',') if p.strip().isdigit()]
                else:
                    ports = []
                
                if not ports:
                    continue
                
                total_qty = row.get("數量", 0)
                qty_per_port = total_qty / len(ports) if len(ports) > 0 else 0
                
                # 生成工單號碼
                if pn not in pn_work_order_map:
                    try:
                        schedule_date = datetime.strptime(date_str, "%Y-%m-%d")
                        year_yy = schedule_date.strftime("%y")
                        month_char = chr(64 + schedule_date.month)
                        work_order_num = f"TMRA{year_yy}{month_char}{work_order_counter:03d}"
                        pn_work_order_map[pn] = work_order_num
                        work_order_counter += 1
                    except:
                        pn_work_order_map[pn] = f"TMRA25L{work_order_counter:03d}"
                        work_order_counter += 1
                
                work_order = pn_work_order_map[pn]
                
                # 計算 Batch
                try:
                    schedule_date = datetime.strptime(date_str, "%Y-%m-%d")
                    iso_year, iso_week, iso_day = schedule_date.isocalendar()
                    year_yy = schedule_date.strftime("%y")
                    pn_last3 = pn[-3:] if len(pn) >= 3 else pn.zfill(3)
                except:
                    iso_week = 1
                    year_yy = "25"
                    pn_last3 = "000"
                
                for port_num in ports:
                    pn_week_key = (pn, iso_week)
                    
                    if pn_week_key not in pn_week_port_counter:
                        pn_week_port_counter[pn_week_key] = 0
                    
                    pn_week_port_counter[pn_week_key] += 1
                    port_count = pn_week_port_counter[pn_week_key]
                    
                    if port_count <= 8:
                        count_char = str(port_count)
                    elif port_count == 9:
                        count_char = 'A'
                    else:
                        count_char = chr(65 + (port_count - 9))
                    
                    batch_code = f"{pn_last3}{year_yy}{iso_week:02d}{count_char}"
                    
                    rd_time = row.get("RD給藥時間", "")
                    if not rd_time:
                        slot_name = row.get("班次", "AM")
                        rd_time = "10:30" if slot_name == "AM" else "15:30"
                    
                    record = {
                        "滴定機": f"Port{port_num}",
                        "Marker": row.get("marker", ""),
                        "PN": pn,
                        "凍乾機台": row.get("凍乾機台", ""),
                        "數量": round(qty_per_port, 2),
                        "配藥同仁": row.get("配藥同仁", ""),
                        "日期": date_str,
                        "RD給藥時間": rd_time,
                        "預計滴定時間": row.get("預計滴定時間", ""),
                        "預計結束": row.get("預計結束", ""),
                        "工單號碼": work_order,
                        "Lot": batch_code,
                        "備註": "",
                        "班次": row.get("班次", "AM"),
                        "record_type": "P1-P5",
                        "has_conflict": False
                    }
                    
                    all_records.append(record)
        
        # 2. 處理 P0 (非實驗)
        if self.p0_strict_records:
            print(f"  收集 P0 (非實驗): {len(self.p0_strict_records)} 筆", file=sys.stderr)
            
            for p0_data in self.p0_strict_records:
                task_date = p0_data.get('P0_Parsed_Date')
                day_key = task_date.strftime('%Y-%m-%d')
                
                rd_time_str = str(p0_data.get("RD給藥時間", "09:00")).strip()
                try:
                    rd_hour = int(rd_time_str.split(':')[0])
                    slot_name = "PM" if rd_hour >= 15 else "AM"
                except:
                    slot_name = "AM"
                
                lot_value = str(p0_data.get("Lot", "")).strip()
                machine = str(p0_data.get("滴定機", "")).strip()
                
                record = {
                    "滴定機": machine,
                    "Marker": str(p0_data.get("Marker", "")).strip(),
                    "PN": str(p0_data.get("PN", "")).strip(),
                    "凍乾機台": str(p0_data.get("凍乾機台", "")).strip(),
                    "數量": str(p0_data.get("數量", "")).strip(),
                    "配藥同仁": str(p0_data.get("配藥同仁", "")).strip(),
                    "日期": day_key,
                    "RD給藥時間": rd_time_str,
                    "預計滴定時間": str(p0_data.get("預計滴定時間", "")).strip(),
                    "預計結束": str(p0_data.get("預計結束", "")).strip(),
                    "工單號碼": "",
                    "Lot": lot_value,
                    "備註": str(p0_data.get("備註", "")).strip(),
                    "班次": slot_name,
                    "record_type": "P0_strict",
                    "has_conflict": False
                }
                
                all_records.append(record)
        
        # 3. 處理 P0 (實驗)
        if self.p0_experiment_records:
            print(f"  收集 P0 (實驗): {len(self.p0_experiment_records)} 筆", file=sys.stderr)
            
            for p0_data in self.p0_experiment_records:
                task_date = p0_data.get('P0_Parsed_Date')
                day_key = task_date.strftime('%Y-%m-%d')
                
                rd_time_str = str(p0_data.get("RD給藥時間", "09:00")).strip()
                try:
                    rd_hour = int(rd_time_str.split(':')[0])
                    slot_name = "PM" if rd_hour >= 15 else "AM"
                except:
                    slot_name = "AM"
                
                lot_value = str(p0_data.get("Lot", "")).strip()
                machine = str(p0_data.get("滴定機", "")).strip()
                has_conflict = p0_data.get('has_conflict', False)
                
                record = {
                    "滴定機": machine,
                    "Marker": str(p0_data.get("Marker", "")).strip(),
                    "PN": str(p0_data.get("PN", "")).strip(),
                    "凍乾機台": str(p0_data.get("凍乾機台", "")).strip(),
                    "數量": str(p0_data.get("數量", "")).strip(),
                    "配藥同仁": str(p0_data.get("配藥同仁", "")).strip(),
                    "日期": day_key,
                    "RD給藥時間": rd_time_str,
                    "預計滴定時間": str(p0_data.get("預計滴定時間", "")).strip(),
                    "預計結束": str(p0_data.get("預計結束", "")).strip(),
                    "工單號碼": "",
                    "Lot": lot_value,
                    "備註": str(p0_data.get("備註", "")).strip(),
                    "班次": slot_name,
                    "record_type": "P0_experiment",
                    "has_conflict": has_conflict
                }
                
                all_records.append(record)
        
        # 4. 處理凍乾機為空
        no_dryer_records_list = []
        if self.no_dryer_records:
            print(f"  收集凍乾機為空任務: {len(self.no_dryer_records)} 筆", file=sys.stderr)
            
            for no_dryer_rec in self.no_dryer_records:
                lot_value = no_dryer_rec.get("lot", "")
                
                record = {
                    "滴定機": "",
                    "Marker": no_dryer_rec.get("marker", ""),
                    "PN": no_dryer_rec.get("PN", ""),
                    "凍乾機台": "",
                    "數量": no_dryer_rec.get("數量", ""),
                    "配藥同仁": no_dryer_rec.get("配藥同仁", ""),
                    "日期": no_dryer_rec.get("日期", ""),
                    "RD給藥時間": no_dryer_rec.get("RD給藥時間", ""),
                    "預計滴定時間": "",
                    "預計結束": "",
                    "工單號碼": "",
                    "Lot": lot_value,
                    "備註": no_dryer_rec.get("備註", "凍乾機為空"),
                    "班次": no_dryer_rec.get("班次", "PM"),
                    "record_type": "no_dryer",
                    "has_conflict": False
                }
                
                no_dryer_records_list.append(record)
        
        # ========================================
        # 按日期分組
        # ========================================
        from collections import defaultdict
        records_by_date = defaultdict(lambda: {
            "IVEK": [],
            "AM_Port": {},
            "PM_Port": {},
            "no_dryer": []
        })
        
        for record in all_records:
            date_key = record.get("日期", "")
            if not date_key:
                continue
            
            machine = record.get("滴定機", "")
            slot = record.get("班次", "AM")
            
            # IVEK
            if "IVEK" in machine.upper():
                records_by_date[date_key]["IVEK"].append(record)
            
            # Port
            elif "Port" in machine or re.search(r'[Pp]ort\s*\d+', machine):
                try:
                    port_num = int(re.search(r'\d+', machine).group())
                    if 1 <= port_num <= 12:
                        if slot == "AM":
                            records_by_date[date_key]["AM_Port"][port_num] = record
                        else:
                            records_by_date[date_key]["PM_Port"][port_num] = record
                except:
                    pass
        
        # no_dryer 記錄
        for record in no_dryer_records_list:
            date_key = record.get("日期", "")
            if date_key:
                records_by_date[date_key]["no_dryer"].append(record)
        
        print(f"\n  ✅ 共收集 {len(all_records) + len(no_dryer_records_list)} 筆記錄，分為 {len(records_by_date)} 天", file=sys.stderr)
        
        # ========================================
        # 建立 Excel 工作簿（固定格式）
        # ========================================
        wb = Workbook()
        wb.remove(wb.active)
        
        headers = [
            "滴定機", "Marker", "PN", "凍乾機台", "數量", "配藥同仁",
            "日期", "RD給藥時間", "預計滴定時間", "預計結束", 
            "工單號碼", "Lot", "備註"
        ]
        
        schedule_data_for_db = []
        
        # 顏色定義
        YELLOW_FILL = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # 淺黃色
        RED_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")  # 淺紅色
        
        # 按日期排序
        for date_key in sorted(records_by_date.keys()):
            day_data = records_by_date[date_key]
            
            # 建立工作表
            sheet_name = date_key
            ws = wb.create_sheet(title=sheet_name)
            
            # 寫入標題
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # ========================================
            # 固定格式輸出
            # ========================================
            current_row = 2
            
            standard_alignment = Alignment(horizontal="center", vertical="center")
            standard_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # ========================================
            # 1. IVEK 區塊（固定 2 行）
            # ========================================
            ivek_records = day_data["IVEK"][:2]  # 最多取 2 筆
            
            for i in range(2):
                # 固定顯示 IVEK
                ws.cell(row=current_row, column=1, value="IVEK")
                
                if i < len(ivek_records):
                    # 有資料，填入
                    record = ivek_records[i]
                    
                    for col_idx, header in enumerate(headers, 1):
                        if header == "滴定機":
                            value = "IVEK"
                        else:
                            value = record.get(header, "")
                        
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.alignment = standard_alignment
                        cell.border = standard_border
                        
                        # P0 實驗衝突 -> 淺紅色
                        if record.get("has_conflict", False):
                            cell.fill = RED_FILL
                    
                    # 準備資料庫記錄
                    db_record = {
                        "日期": record.get("日期", ""),
                        "marker": record.get("Marker", ""),
                        "滴定機": "IVEK",
                        "凍乾機台": record.get("凍乾機台", ""),
                        "配藥同仁": record.get("配藥同仁", ""),
                        "RD給藥時間": record.get("RD給藥時間", ""),
                        "預計滴定時間": record.get("預計滴定時間", ""),
                        "預計結束": record.get("預計結束", ""),
                        "預冷時間": "",
                        "凍乾時間": "",
                        "收藥時間": "",
                        "數量": record.get("數量", ""),
                        "PN": record.get("PN", ""),
                        "工單號碼": record.get("工單號碼", ""),
                        "Batch": record.get("Lot", ""),
                        "班次": record.get("班次", ""),
                        "record_type": record.get("record_type", ""),
                        "has_conflict": 1 if record.get("has_conflict", False) else 0
                    }
                    schedule_data_for_db.append(db_record)
                else:
                    # 無資料，只填滴定機欄位
                    for col_idx in range(1, len(headers) + 1):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.alignment = standard_alignment
                        cell.border = standard_border
                
                current_row += 1
            
            # 空行
            current_row += 1
            
            # ========================================
            # 2. AM Port1-12 區塊（固定 12 行）
            # ========================================
            am_ports = day_data["AM_Port"]
            
            for port_num in range(1, 13):
                # 固定顯示 Port
                ws.cell(row=current_row, column=1, value=f"Port{port_num}")
                
                if port_num in am_ports:
                    # 有資料，填入
                    record = am_ports[port_num]
                    
                    for col_idx, header in enumerate(headers, 1):
                        if header == "滴定機":
                            value = f"Port{port_num}"
                        else:
                            value = record.get(header, "")
                        
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.alignment = standard_alignment
                        cell.border = standard_border
                        
                        # P0 實驗衝突 -> 淺紅色
                        if record.get("has_conflict", False):
                            cell.fill = RED_FILL
                    
                    # 準備資料庫記錄
                    db_record = {
                        "日期": record.get("日期", ""),
                        "marker": record.get("Marker", ""),
                        "滴定機": f"Port{port_num}",
                        "凍乾機台": record.get("凍乾機台", ""),
                        "配藥同仁": record.get("配藥同仁", ""),
                        "RD給藥時間": record.get("RD給藥時間", ""),
                        "預計滴定時間": record.get("預計滴定時間", ""),
                        "預計結束": record.get("預計結束", ""),
                        "預冷時間": "",
                        "凍乾時間": "",
                        "收藥時間": "",
                        "數量": record.get("數量", ""),
                        "PN": record.get("PN", ""),
                        "工單號碼": record.get("工單號碼", ""),
                        "Batch": record.get("Lot", ""),
                        "班次": "AM",
                        "record_type": record.get("record_type", ""),
                        "has_conflict": 1 if record.get("has_conflict", False) else 0
                    }
                    schedule_data_for_db.append(db_record)
                else:
                    # 無資料，只填滴定機欄位
                    for col_idx in range(1, len(headers) + 1):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.alignment = standard_alignment
                        cell.border = standard_border
                
                current_row += 1
            
            # 空行
            current_row += 1
            
            # ========================================
            # 3. PM Port1-12 區塊（固定 12 行）
            # ========================================
            pm_ports = day_data["PM_Port"]
            
            for port_num in range(1, 13):
                # 固定顯示 Port
                ws.cell(row=current_row, column=1, value=f"Port{port_num}")
                
                if port_num in pm_ports:
                    # 有資料，填入
                    record = pm_ports[port_num]
                    
                    for col_idx, header in enumerate(headers, 1):
                        if header == "滴定機":
                            value = f"Port{port_num}"
                        else:
                            value = record.get(header, "")
                        
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.alignment = standard_alignment
                        cell.border = standard_border
                        
                        # P0 實驗衝突 -> 淺紅色
                        if record.get("has_conflict", False):
                            cell.fill = RED_FILL
                    
                    # 準備資料庫記錄
                    db_record = {
                        "日期": record.get("日期", ""),
                        "marker": record.get("Marker", ""),
                        "滴定機": f"Port{port_num}",
                        "凍乾機台": record.get("凍乾機台", ""),
                        "配藥同仁": record.get("配藥同仁", ""),
                        "RD給藥時間": record.get("RD給藥時間", ""),
                        "預計滴定時間": record.get("預計滴定時間", ""),
                        "預計結束": record.get("預計結束", ""),
                        "預冷時間": "",
                        "凍乾時間": "",
                        "收藥時間": "",
                        "數量": record.get("數量", ""),
                        "PN": record.get("PN", ""),
                        "工單號碼": record.get("工單號碼", ""),
                        "Batch": record.get("Lot", ""),
                        "班次": "PM",
                        "record_type": record.get("record_type", ""),
                        "has_conflict": 1 if record.get("has_conflict", False) else 0
                    }
                    schedule_data_for_db.append(db_record)
                else:
                    # 無資料，只填滴定機欄位
                    for col_idx in range(1, len(headers) + 1):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.alignment = standard_alignment
                        cell.border = standard_border
                
                current_row += 1
            
            # 空行
            current_row += 1
            
            # ========================================
            # 4. no_dryer 區塊（動態行數）
            # ========================================
            no_dryer_list = day_data["no_dryer"]
            
            if no_dryer_list:
                for record in no_dryer_list:
                    for col_idx, header in enumerate(headers, 1):
                        value = record.get(header, "")
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.alignment = standard_alignment
                        cell.border = standard_border
                        
                        # no_dryer -> 淺黃色
                        cell.fill = YELLOW_FILL
                    
                    # 準備資料庫記錄
                    db_record = {
                        "日期": record.get("日期", ""),
                        "marker": record.get("Marker", ""),
                        "滴定機": "",
                        "凍乾機台": "",
                        "配藥同仁": record.get("配藥同仁", ""),
                        "RD給藥時間": record.get("RD給藥時間", ""),
                        "預計滴定時間": "",
                        "預計結束": "",
                        "預冷時間": "",
                        "凍乾時間": "",
                        "收藥時間": "",
                        "數量": record.get("數量", ""),
                        "PN": record.get("PN", ""),
                        "工單號碼": "",
                        "Batch": record.get("Lot", ""),
                        "班次": record.get("班次", ""),
                        "record_type": "no_dryer",
                        "has_conflict": 0
                    }
                    schedule_data_for_db.append(db_record)
                    
                    current_row += 1
            
            # 設定欄寬
            column_widths = {
                "A": 12, "B": 20, "C": 15, "D": 12, "E": 10,
                "F": 12, "G": 12, "H": 12, "I": 12, "J": 12,
                "K": 15, "L": 15, "M": 20
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            ws.freeze_panes = "A2"
            
            print(f"  ✅ 工作表 '{sheet_name}' 已建立（固定格式）", file=sys.stderr)
        
        # ========================================
        # 儲存 Excel
        # ========================================
        try:
            wb.save(output_path)
            print(f"\n✅ Excel 輸出成功: {output_path}", file=sys.stderr)
            print(f"   共 {len(records_by_date)} 個工作表（固定格式）", file=sys.stderr)
        except Exception as e:
            print(f"❌ Excel 輸出失敗: {e}", file=sys.stderr)
            traceback.print_exc(file=sys.stderr)
            return None

        # ========================================
        # 寫入 SQLite 資料庫
        # ========================================
        if schedule_data_for_db:
            global SCHEDULE_START_DATE
            print(f"\n💾 準備寫入資料庫 ({len(schedule_data_for_db)} 筆)...", file=sys.stderr)

            try:
                if SCHEDULE_START_DATE:
                    week_obj = SCHEDULE_START_DATE
                else:
                    first_date = schedule_data_for_db[0].get("日期", "")
                    week_obj = datetime.strptime(first_date, "%Y-%m-%d")

                iso_year, iso_week, _ = week_obj.isocalendar()
                schedule_week = f"{iso_year}_W{iso_week:02d}"
                print(f"🗓️ 計算週次: {schedule_week}", file=sys.stderr)

            except Exception as e:
                print(f"⚠️ 無法解析週次: {e}", file=sys.stderr)
                schedule_week = "unknown"
                iso_year = datetime.now().year

            db_path = SCHEDULE_DB_PATH
            table_name = f"schedule_{iso_year}"

            try:
                conn = sqlite3.connect(db_path)
                cur = conn.cursor()

                cur.execute(f"""
                    CREATE TABLE IF NOT EXISTS {table_name} (
                        日期 TEXT,
                        marker TEXT,
                        滴定機 TEXT,
                        凍乾機台 TEXT,
                        配藥同仁 TEXT,
                        RD給藥時間 TEXT,
                        預計滴定時間 TEXT,
                        預計結束 TEXT,
                        預冷時間 TEXT,
                        凍乾時間 TEXT,
                        收藥時間 TEXT,
                        數量 TEXT,
                        PN TEXT,
                        工單號碼 TEXT,
                        Batch TEXT,
                        班次 TEXT,
                        record_type TEXT,
                        has_conflict INTEGER,
                        schedule_week TEXT
                    )
                """)

                cur.execute(f"DELETE FROM {table_name} WHERE schedule_week = ?", (schedule_week,))
                deleted_count = cur.rowcount

                for row in schedule_data_for_db:
                    row["schedule_week"] = schedule_week

                insert_sql = f"""
                    INSERT INTO {table_name} (
                        日期, marker, 滴定機, 凍乾機台, 配藥同仁,
                        RD給藥時間, 預計滴定時間, 預計結束, 預冷時間,
                        凍乾時間, 收藥時間, 數量, PN, 工單號碼,
                        Batch, 班次, record_type, has_conflict, schedule_week
                    ) VALUES (
                        :日期, :marker, :滴定機, :凍乾機台, :配藥同仁,
                        :RD給藥時間, :預計滴定時間, :預計結束, :預冷時間,
                        :凍乾時間, :收藥時間, :數量, :PN, :工單號碼,
                        :Batch, :班次, :record_type, :has_conflict, :schedule_week
                    )
                """
                
                cur.executemany(insert_sql, schedule_data_for_db)
                conn.commit()
                
                cur.execute(f"SELECT COUNT(*) FROM {table_name} WHERE schedule_week = ?", (schedule_week,))
                inserted_count = cur.fetchone()[0]
                
                cur.execute(f"""
                    SELECT COUNT(*) FROM {table_name} 
                    WHERE schedule_week = ? AND record_type IN ('P0_strict', 'P0_experiment')
                """, (schedule_week,))
                p0_count = cur.fetchone()[0]
                
                conn.close()

                print(f"✅ 成功寫入 {inserted_count} 筆資料到 {table_name}", file=sys.stderr)
                print(f"   - P0 任務: {p0_count} 筆", file=sys.stderr)
                print(f"   - 週次: {schedule_week}", file=sys.stderr)

            except Exception as e:
                print(f"❌ 資料庫寫入失敗: {e}", file=sys.stderr)
                traceback.print_exc(file=sys.stderr)

        return output_path
# ========================================
# Part 8: 主程序
# ========================================

def main():
    """主程序（前端整合版）"""
    
    # ========================================
    # V9.9.9: 解析命令行參數
    # ========================================
    args = parse_arguments()
    
    print(f"===== [{CURRENT_VERSION}] 排程開始 =====\n", file=sys.stderr)
    
    # 解析起始日期
    try:
        current_year = datetime.now().year
        parts = args.date.split('/')
        month, day = int(parts[0]), int(parts[1])
        start_date = datetime(current_year, month, day)
        
        # 確保是週一
        if start_date.weekday() != 0:
            start_date = start_date - timedelta(days=start_date.weekday())
        # ========================================
        # ✅ 新增：設定全域變數
        # ========================================
        global SCHEDULE_START_DATE
        SCHEDULE_START_DATE = start_date
        
        print(f"排程起始日期（週一）: {start_date.strftime('%Y-%m-%d')}\n", file=sys.stderr)
        
    except Exception as e:
        output = {"ok": False, "message": f"日期解析失敗: {e}"}
        print(json.dumps(output, ensure_ascii=False))
        sys.exit(1)
    
    # ========================================
    # 設定參數
    # ========================================
    
    # 設定休假日
    if args.holidays:
        print("="*70, file=sys.stderr)
        print("🗓️  設定休假日", file=sys.stderr)
        print("="*70, file=sys.stderr)
        setup_holidays_from_args(args.holidays, start_date)
        print(file=sys.stderr)
    
    # 設定批次編號
    if args.batch_numbers:
        print("="*70, file=sys.stderr)
        print("🔢  設定工單批次編號", file=sys.stderr)
        print("="*70, file=sys.stderr)
        setup_batch_start_from_args(args.batch_numbers)
        print(file=sys.stderr)
    
    # 設定休假人員
    if args.vacation_staff:
        print("="*70, file=sys.stderr)
        print("👥  設定休假人員", file=sys.stderr)
        print("="*70, file=sys.stderr)
        setup_vacation_staff_from_args(args.vacation_staff, start_date)
        print(file=sys.stderr)
    
    try:
        # ========================================
        # 載入資料
        # ========================================
        all_data = load_all_data(args.need)
        
        if not all_data or all_data["demand"] is None:
            output = {"ok": False, "message": "資料載入錯誤"}
            print(json.dumps(output, ensure_ascii=False))
            sys.exit(1)

        # ========================================
        # 執行排程
        # ========================================
        scheduler = Scheduler(all_data)
        scheduler.run()
        
        # ========================================
        # 輸出結果
        # ========================================
        if hasattr(scheduler, 'schedule_df') and not scheduler.schedule_df.empty:
            
            if args.dry_run:
                # Dry Run 模式：輸出預覽 JSON
                preview_data = scheduler._generate_dry_run_preview()
                output = {"ok": True, "preview": preview_data}
                print(json.dumps(output, ensure_ascii=False))
            else:
                # 正常模式：儲存 Excel
                # ✅ 修正 3: 接收 save_records_to_excel 返回的 output_path
                output_path = scheduler.save_records_to_excel(args.outdir)
                
                output = {"ok": True, "outPath": output_path}
                print(json.dumps(output, ensure_ascii=False))
                
            print(f"\n===== [{CURRENT_VERSION}] 排程完成 =====", file=sys.stderr)
        else:
            output = {"ok": False, "message": "排程已終止，未產生結果"}
            print(json.dumps(output, ensure_ascii=False))
            sys.exit(1)

    except Exception as e:
        output = {"ok": False, "message": f"排程失敗: {str(e)}"}
        print(json.dumps(output, ensure_ascii=False))
        traceback.print_exc(file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()