# -*- coding: utf-8 -*-
"""
plan_to_beads_schedule_1.py

依規則產生「6 天、每天最多 8 批」的生產排程；Dry Run 回傳 JSON，正式寫入空白排程.xlsm（保留 VBA）。

● Dry Run 預覽欄位：date, titrate, freeze, pn, qty, Name, staff
● 人力限制：同一人員單日最多 2 筆，超過遞延下一日

參數：
  --date MM/DD
  --need <需求Excel>（sheet: 滴定排程需求表, header=第2列）
  --limit <限制Excel>（sheet: 配藥限制, header=第1列）
  --template <空白排程.xlsm>
  --outdir <輸出資料夾>
  --dry-run
  --year YYYY（預設今年）
  --max-rows-per-day 8（每天最多筆數）
  --days 6（天數）

輸出：
  dry-run: {"ok": true, "preview": [ {date,titrate,freeze,pn,qty,Name,staff}, ... ], "overflow": N}
  正式:    {"ok": true, "outPath": "\\\\...\\beads排程_yyyymmdd_n.xlsm", "scheduled": M, "overflow": N}
"""

import argparse
import datetime as dt
import json
import os
import re
import sys
from collections import defaultdict, deque
from pathlib import Path
from shutil import copy2

try:
    import pandas as pd  # type: ignore
except Exception:
    pd = None

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:
    load_workbook = None

# === 不排程 PN 黑名單（canonical 字串；會用 canonical_pn 比對）===
PN_BLACKLIST = {
    "5714600131",
    "5714600141",
    "5714600142",
    "5714600143",
    "5714600146",
    "5714600145",
    "5711500264",
    "5711500265",
    "5711500266",
    "5711500272",
    "5711500273",
    "5714600102",
}

def is_blacklisted_pn(pn: str) -> bool:
    return canonical_pn(pn) in PN_BLACKLIST

# ---------- 基本工具 ----------

def fail(code: int, msg: str):
    print(msg, file=sys.stderr, flush=True)
    sys.exit(code)


MMDD_RE = re.compile(r"^(0[1-9]|1[0-2])/(0[1-9]|[12]\d|3[01])$")


def mmdd_to_date(mmdd: str, year: int) -> dt.date:
    if not MMDD_RE.match(mmdd):
        fail(2, f"Invalid --date '{mmdd}', must be MM/DD")
    m, d = map(int, mmdd.split("/"))
    return dt.date(year, m, d)


def ensure_exists(p: str, is_dir: bool = False) -> Path:
    path = Path(p)
    if is_dir:
        if not path.exists():
            path.mkdir(parents=True, exist_ok=True)
    else:
        if not path.exists():
            fail(3, f"Path not found: {p}")
    return path


def safe_int(x) -> int:
    try:
        if x is None:
            return 0
        s = str(x).strip().replace(",", "")
        if s == "":
            return 0
        return int(round(float(s)))
    except Exception:
        return 0

def parse_freeze_list(val) -> list[int]:
    """E 欄可用凍乾機：支援 3,4,5… 以及 3-10 / 3~10；只回傳整數機號。"""
    s = str(val or "").strip()
    if not s or s.lower() == "nan":
        return []
    out: list[int] = []
    for tok in re.split(r"[,\u3001、/／;；&\s]+", s):
        t = tok.strip()
        if not t:
            continue
        m = re.match(r"^(\d+)\s*[-~]\s*(\d+)$", t)  # 3-10 / 3~10
        if m:
            a, b = int(m.group(1)), int(m.group(2))
            lo, hi = sorted((a, b))
            out.extend(range(lo, hi + 1))
        elif t.isdigit():
            out.append(int(t))
    # 去重保序
    seen, res = set(), []
    for n in out:
        if n not in seen:
            seen.add(n); res.append(n)
    return res

def canonical_pn(x) -> str:
    """把 PN 穩定成純字串：去空白/逗號、去尾端 .0；避免科學記號造成對不到。"""
    s = str(x).strip()
    if not s or s.lower() in ("nan", "none"):
        return ""
    s = s.replace(" ", "").replace(",", "")
    if s.endswith(".0"):
        s = s[:-2]
    return s


def ceil_div_pos(a: int, b: int) -> int:
    """正整數天花板除法。"""
    return (a + b - 1) // b


# ---------- 讀需求 ----------

def read_need_excel(xlsx: Path):
    """
    sheet '滴定排程需求表'
    會自動偵測 header 列（嘗試 header=1,0,2,3），並允許欄位別名：
      料號: 料號 / PN / 品號
      品名: 品名 / 藥名 / 名稱
      庫存+滴定: 庫存+滴定 / 庫存＋滴定 / 庫存 / 目前庫存數 / 庫存數 / 庫存_含滴定
      第一周需求: 第一周需求 / 第一週需求 / 週1需求 / Week1需求 / 第1周需求 / 第1週需求
      第二周需求: 第二周需求 / 第二週需求 / 週2需求 / Week2需求 / 第2周需求 / 第2週需求 / 第二周週需求 / 第二週週需求 / 2週需求
      第三周需求: 第三周需求 / 第三週需求 / 週3需求 / Week3需求 / 第3周需求 / 第3週需求 / 第三周週需求 / 第三週週需求 / 3週需求
      凍乾數: 凍乾數 / 凍乾批量 / 批量 / 每批數量 / 單批數量 / 批次數量
    """
    if pd is None:
        fail(10, "pandas not installed. pip install pandas openpyxl")

    # 欄位別名表（左邊是程式內部使用的標準名）
    aliases = {
        "料號": ["料號", "pn", "PN", "品號", "料號/PN"],
        "品名": ["品名", "藥名", "名稱"],
        "庫存+滴定": ["庫存+滴定", "庫存＋滴定", "庫存", "目前庫存數", "庫存數", "庫存_含滴定"],
        "第一周需求": ["第一周需求", "第一週需求", "週1需求", "Week1需求", "第1周需求", "第1週需求", "1週需求"],
        "第二周需求": ["第二周需求", "第二週需求", "週2需求", "Week2需求", "第2周需求", "第2週需求", "第二周週需求", "第二週週需求", "2週需求"],
        "第三周需求": ["第三周需求", "第三週需求", "週3需求", "Week3需求", "第3周需求", "第3週需求", "第三周週需求", "第三週週需求", "3週需求"],
        "凍乾數": ["凍乾數", "凍乾批量", "批量", "每批數量", "單批數量", "批次數量"],
    }

    def norm(s: str) -> str:
        # 標準化：去空白、全形空白；週->周；全形＋->半形+；小寫
        t = str(s or "")
        t = t.replace("\u3000", " ").strip().replace(" ", "")
        t = t.replace("週", "周")
        t = t.replace("＋", "+")
        return t.lower()

    last_err = None
    for hdr in (1, 0, 2, 3):
        try:
            df = pd.read_excel(xlsx, sheet_name="滴定排程需求表", header=hdr)
        except Exception as e:
            last_err = e
            continue

        cols = [str(c).strip() for c in df.columns]
        norm_cols = [norm(c) for c in cols]

        # 建立 canonical -> 原始欄名 的映射
        colmap = {}
        for canon, cands in aliases.items():
            found = None
            for cand in cands:
                nc = norm(cand)
                if nc in norm_cols:
                    found = cols[norm_cols.index(nc)]
                    break
            if found:
                colmap[canon] = found

        needed = ["料號", "品名", "庫存+滴定", "第一周需求", "第二周需求", "第三周需求", "凍乾數"]
        missing = [k for k in needed if k not in colmap]
        if missing:
            last_err = f"header={hdr} 缺少欄位：{','.join(missing)}（實際欄位：{cols}）"
            continue

        # 轉 dict
        dic_INV = {}
        for _, r in df.fillna(0).iterrows():
            pn = canonical_pn(r[colmap["料號"]])
            if not pn:
                continue
            dic_INV[pn] = {
                "品名": str(r[colmap["品名"]]).strip(),
                "庫存+滴定": safe_int(r[colmap["庫存+滴定"]]),
                "第一周需求": safe_int(r[colmap["第一周需求"]]),
                "第二周需求": safe_int(r[colmap["第二周需求"]]),
                "第三周需求": safe_int(r[colmap["第三周需求"]]),
                "凍乾數": max(1, safe_int(r[colmap["凍乾數"]])),
            }

        # 附帶把實際對應情況印到 stderr（方便你在 Flask log 看到）
        mapping_info = {k: colmap[k] for k in needed}
        print(f"[read_need_excel] header={hdr} column mapping -> {mapping_info}", file=sys.stderr, flush=True)
        return dic_INV

    # 全部嘗試失敗
    if last_err:
        fail(11, f"需求檔欄位偵測失敗：{last_err}\n"
                 f"👉 請確認表頭是否在第2列（或第1/3列），且欄名為上述別名之一。")
    else:
        fail(11, "需求檔讀取失敗：無法開啟或找不到 sheet『滴定排程需求表』")



# ---------- 讀限制：抓可用凍乾機 & 配藥人員 ----------

_SEP = re.compile(r"[,\u3001、/／;；&\s]+")

def _split_names(s: str):
    names = [t.strip() for t in _SEP.split(str(s or "")) if t.strip()]
    # 去重且保序
    seen, out = set(), []
    for n in names:
        if n and n not in seen:
            seen.add(n)
            out.append(n)
    return out


_SEP = re.compile(r"[,\u3001、/／;；&\s]+")

def _split_names(s: str):
    names = [t.strip() for t in _SEP.split(str(s or "")) if t.strip()]
    seen, out = set(), []
    for n in names:
        if n and n not in seen:
            seen.add(n); out.append(n)
    return out

def read_limit_excel(xlsx: Path):
    """
    sheet『配藥限制』，header=第1列
      A 欄=PN（字串化）
      E 欄=可用凍乾機（例：3,4 或 3~10）
      I/J/K 欄=配藥主/副手
    回傳 dic_Prod[pn] = {"__freeze_list":[int...], "__staff_list":[str...]}
    """
    if pd is None:
        fail(10, "pandas not installed. pip install pandas openpyxl")

    try:
        df = pd.read_excel(xlsx, sheet_name="配藥限制", header=0, dtype=str, usecols="A:K")
    except Exception as e:
        fail(12, f"限制檔讀取失敗：{xlsx} | 配藥限制 | {e}")

    df = df.fillna("")
    cols = [str(c).strip() for c in df.columns]
    df.columns = cols

    dic_Prod = {}
    for _, row in df.iterrows():
        pn = canonical_pn(row.iloc[0])  # A 欄
        if not pn:
            continue

        # E 欄：可用凍乾機
        freeze_list = parse_freeze_list(row.iloc[4])  # 0-based index: A=0, E=4

        # I/J/K：配藥主/副手
        names = []
        for pos in (8, 9, 10):  # I/J/K
            if pos < len(row):
                val = str(row.iloc[pos]).strip()
                if val and val.lower() != "nan":
                    for t in re.split(r"[,\u3001、/／;；&\s]+", val):
                        t = t.strip()
                        if t and t not in names:
                            names.append(t)

        dic_Prod[pn] = {
            "__freeze_list": freeze_list,
            "__staff_list": names,
        }

    if not dic_Prod:
        fail(12, "限制檔讀到 0 筆（請確認 sheet=配藥限制、A=PN、E=可用凍乾機、I/J/K=人員）。")
    return dic_Prod



# ---------- 排程（三階段缺口 → 批次 → 人員限制分日） ----------

def compute_batches(dic_INV: dict, dic_Prod: dict, days: int = 6, per_day: int = 8):
    """
    依規則挑單，輸出「批次」清單（尚未分日分人）：
      item: {pn, qty, Name, freeze_list, staff_candidates}

    規則加強：
      - 若「第一/第二/第三周需求」三者皆 0（或空），**不排**（包含 stage0~3）
      - stage0：凍乾數<=0 且 庫存+滴定<=0，且有需求者先排；排量取最早有需求的週需求值
      - stage1~3：原缺口邏輯（以凍乾數為批量）
    """
    capacity = days * per_day
    out = []

    stock = {pn: safe_int(v.get("庫存+滴定", 0)) for pn, v in dic_INV.items()}
    batch = {pn: safe_int(v.get("凍乾數", 0)) for pn, v in dic_INV.items()}
    d1 = {pn: safe_int(v.get("第一周需求", 0)) for pn, v in dic_INV.items()}
    d2 = {pn: safe_int(v.get("第二周需求", 0)) for pn, v in dic_INV.items()}
    d3 = {pn: safe_int(v.get("第三周需求", 0)) for pn, v in dic_INV.items()}
    name_map = {pn: str(v.get("品名", "")) for pn, v in dic_INV.items()}

    # ★ 只有有需求者才納入後續判斷
    has_demand = {pn: (d1[pn] > 0 or d2[pn] > 0 or d3[pn] > 0) for pn in dic_INV.keys()}

    # ---------- stage 0: 凍乾數/庫存為空但有需求 -> 先排 ----------
    special = []
    for pn in dic_INV.keys():
         # ★ 黑名單直接略過
        if is_blacklisted_pn(pn):
            continue
        #「三周皆 0 不排」的規則，在這裡先擋掉
        if (d1[pn] == 0 and d2[pn] == 0 and d3[pn] == 0):
            continue
        if not has_demand[pn]:
            continue
        b = batch[pn]
        s = stock[pn]
        w1, w2, w3 = d1[pn], d2[pn], d3[pn]
        if b <= 0 and s <= 0:
            if any([w1, w2, w3]):
                # 取最早有需求的週
                if w1 > 0: qty, pri = w1, 0
                elif w2 > 0: qty, pri = w2, 1
                else: qty, pri = w3, 2
                freeze_list = (dic_Prod.get(pn, {}) or {}).get("__freeze_list", [])
                staff_candidates = (dic_Prod.get(pn, {}) or {}).get("__staff_list", [])
                special.append((pri, pn, qty, freeze_list, staff_candidates))
    # 依「越早需求越前」排序（同 pri 以數量大者先）
    special.sort(key=lambda t: (t[0], -t[2]))
    for _, pn, qty, freeze_list, staff_candidates in special:
        if len(out) >= capacity:
            break
        qty = max(1, qty)
        out.append({
            "pn": pn,
            "qty": qty,
            "Name": name_map[pn],
            "freeze_list": freeze_list,
            "staff_candidates": staff_candidates,
        })
        stock[pn] += qty  # 視為已補

    # ---------- stage 1~3: 原缺口邏輯（用凍乾數為批量） ----------
    def stage_demand(pn, stage):
        if stage == 1:
            return d1[pn]
        if stage == 2:
            return d1[pn] + d2[pn]
        return d1[pn] + d2[pn] + d3[pn]

    for stage in (1, 2, 3):
        cands = []
        for pn in dic_INV.keys():
            if is_blacklisted_pn(pn):
                continue
            if not has_demand[pn]:
                continue  # ★ 無需求者不納入
            b = max(1, safe_int(batch[pn]))  # 避免除以 0
            dem = stage_demand(pn, stage)
            cov = (stock[pn] - dem) / b
            if cov < 1:
                cands.append((cov, pn))

        cands.sort(key=lambda x: x[0])  # coverage 越不足越優先

        for _, pn in cands:
            if len(out) >= capacity:
                break
            b = max(1, safe_int(batch[pn]))
            dem = stage_demand(pn, stage)
            shortfall = dem + b - stock[pn]
            if shortfall <= 0:
                continue

            need_batches = ceil_div_pos(shortfall, b)
            freeze_list = (dic_Prod.get(pn, {}) or {}).get("__freeze_list", [])
            staff_candidates = (dic_Prod.get(pn, {}) or {}).get("__staff_list", [])

            for _ in range(need_batches):
                if len(out) >= capacity:
                    break
                out.append({
                    "pn": pn,
                    "qty": b,
                    "Name": name_map[pn],
                    "freeze_list": freeze_list,
                    "staff_candidates": staff_candidates,
                })
                stock[pn] += b

        if len(out) >= capacity:
            break

    overflow = max(0, len(out) - capacity)
    return out[:capacity], overflow



def assign_to_days_with_staff(
    batches: list,
    first_day: dt.date,
    days: int = 6,
    per_day: int = 8,
    per_staff_per_day: int = 2,
):
    """
    將 batches 依序排入 days 天：
      - 每天最多 per_day 筆
      - 有 staff_candidates 時：同一人員單日最多 per_staff_per_day 筆；超過遞延到下一天
      - 有 freeze_list 時：從允許機台中挑「當日目前最少被使用」的機台（輸出僅號碼）
      - *-D / *-U 同基底（如 K-D/K-U）盡量同機台
      - 同一天同 PN 的多批固定同機台（須在允許清單內）
    需要每筆 batch 具備：{'pn','qty','Name','freeze_list':[int...],'staff_candidates':[str...]}
    回傳 (preview, overflow)
    """
    q = deque(batches)
    preview = []

    def base_of(name: str) -> str | None:
        # 取 *-D / *-U 的基底（大小寫皆可）
        s = str(name or "").strip()
        m = re.match(r"(?i)^(.*?)[\-_]\s*([du])$", s)
        return m.group(1).strip().upper() if m else None

    for day_idx in range(days):
        day_str = (first_day + dt.timedelta(days=day_idx)).strftime("%m/%d")
        day_placed = 0

        # 當日統計
        staff_count = defaultdict(int)     # name -> used count today
        freeze_count = defaultdict(int)    # freeze_no(int) -> used count today
        base_freeze_map: dict[str, int] = {}  # 基底 -> 已用凍乾機
        pn_freeze_map: dict[str, int] = {}    # PN -> 已用凍乾機（同日固定）

        full_scan_budget = len(q) + 10

        while day_placed < per_day and q and full_scan_budget > 0:
            full_scan_budget -= 1
            progressed = False

            for _ in range(len(q)):
                b = q.popleft()

                # ---- 人員限制 ----
                staff = ""
                cands = [s for s in (b.get("staff_candidates") or []) if str(s).strip()]
                if cands:
                    under = [(nm, staff_count[nm]) for nm in cands if staff_count[nm] < per_staff_per_day]
                    if not under:
                        q.append(b)  # 今日人員額度已滿 → 明天
                        continue
                    under.sort(key=lambda t: t[1])
                    staff = under[0][0]
                    staff_count[staff] += 1

                # ---- 凍乾機指派（僅號碼）----
                freeze_txt = ""
                allowed = list(dict.fromkeys(b.get("freeze_list") or []))  # 去重保序
                chosen: int | None = None

                if allowed:
                    pn_key = canonical_pn(b.get("pn", ""))
                    base = base_of(b.get("Name", ""))

                    # 1) 同日同 PN：若已有機台且相容 → 直接沿用
                    if pn_key in pn_freeze_map and pn_freeze_map[pn_key] in allowed:
                        chosen = pn_freeze_map[pn_key]

                    # 2) *-D/*-U 同基底：若已有機台且相容 → 沿用
                    if chosen is None and base and base in base_freeze_map and base_freeze_map[base] in allowed:
                        chosen = base_freeze_map[base]

                    # 3) 若尚未有固定機台，嘗試與佇列中同基底/同 PN 找交集以便之後一致
                    if chosen is None:
                        inter: set[int] = set()
                        if base or pn_key:
                            for future in q:
                                same_pn = canonical_pn(future.get("pn", "")) == pn_key if pn_key else False
                                same_base = (base_of(future.get("Name", "")) == base) if base else False
                                if same_pn or same_base:
                                    fl = future.get("freeze_list") or []
                                    inter |= set(fl)
                        candidates = [n for n in allowed if (not inter or n in inter)] or allowed
                        chosen = sorted(candidates, key=lambda n: (freeze_count[n], n))[0]

                    # 記錄當日固定
                    pn_freeze_map[pn_key] = chosen
                    if base:
                        base_freeze_map.setdefault(base, chosen)

                    freeze_count[chosen] += 1
                    freeze_txt = str(chosen)

                # ---- 放入今天 ----
                preview.append({
                    "date": day_str,
                    "titrate": "T1",
                    "freeze": freeze_txt,   # 只寫號碼
                    "pn": b["pn"],
                    "qty": b["qty"],
                    "Name": b["Name"],
                    "staff": staff,
                })
                day_placed += 1
                progressed = True
                if day_placed >= per_day:
                    break

            if not progressed:
                break  # 今日受人員/機台限制卡住

    overflow = len(q)
    return preview, overflow





# ---------- 寫回 Excel（空白排程） ----------

def find_date_anchors(ws):
    anchors = []
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=2).value  # B
        if isinstance(v, str) and v.strip() == "日期":
            anchors.append(row)
    if len(anchors) < 6:
        fail(30, f"工作表B欄 '日期' 標籤不足 6 個（找到 {len(anchors)}）")
    return anchors[:6]


def clear_block(ws, r1, r2, c1=2, c2=17):
    if r2 < r1:
        return
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).value = None


def write_to_workbook(book_path: Path, first_day: dt.date, preview_by_day: list, per_day: int = 8):
    if load_workbook is None:
        fail(13, "openpyxl not installed. pip install openpyxl")

    wb = load_workbook(str(book_path), keep_vba=True)
    if "空白排程" not in wb.sheetnames:
        fail(31, "模板內找不到工作表：空白排程")
    ws = wb["空白排程"]

    anchors = find_date_anchors(ws)
    for day_idx in range(6):
        anchor_row = anchors[day_idx]
        the_date = (first_day + dt.timedelta(days=day_idx)).strftime("%m/%d")

        # 在「日期」右一格（C）寫入日期
        ws.cell(row=anchor_row, column=3).value = the_date

        data_start = anchor_row + 2
        data_end = (anchors[day_idx + 1] - 2) if day_idx + 1 < len(anchors) else ws.max_row

        # 先清 B:Q
        clear_block(ws, data_start, data_end, c1=2, c2=17)

        rows = [r for r in preview_by_day if r["date"] == the_date][:per_day]
        r = data_start
        for item in rows:
            if r > data_end:
                break
            ws.cell(row=r, column=2).value = item.get("titrate")     # B 滴定機
            ws.cell(row=r, column=3).value = item.get("pn")          # C Marker = PN
            ws.cell(row=r, column=4).value = item.get("freeze")      # D 凍乾機台
            ws.cell(row=r, column=5).value = item.get("qty")         # E 數量
            ws.cell(row=r, column=6).value = item.get("staff")       # F 配藥同仁
            ws.cell(row=r, column=7).value = the_date                # G 日期
            ws.cell(row=r, column=8).value = None                    # H RD給藥時間
            ws.cell(row=r, column=9).value = None                    # I 預計滴定時間
            ws.cell(row=r, column=10).value = None                   # J 預計結束
            ws.cell(row=r, column=11).value = None                   # K 預冷時間
            ws.cell(row=r, column=12).value = None                   # L 凍乾時間
            ws.cell(row=r, column=13).value = None                   # M 收藥時間
            ws.cell(row=r, column=14).value = None                   # N 工單編號
            ws.cell(row=r, column=15).value = None                   # O Lot
            ws.cell(row=r, column=16).value = item.get("Name")       # P 備註 = 品名
            ws.cell(row=r, column=17).value = None                   # Q
            r += 1

    wb.save(str(book_path))


def unique_outpath(outdir: Path, yyyymmdd: str, max_suffix: int = 99) -> Path:
    for n in range(1, max_suffix + 1):
        p = outdir / f"beads排程_{yyyymmdd}_{n}.xlsm"
        if not p.exists():
            return p
    fail(40, f"Too many existing files for {yyyymmdd}, exceeded _{max_suffix}")


# ---------- 主程式 ----------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--date", required=True, help="當週第一天，格式 MM/DD")
    ap.add_argument("--need", required=True, help="需求檔（含 凍乾數）")
    ap.add_argument("--limit", required=True, help="限制檔（含 可用凍乾機／配藥人員）")
    ap.add_argument("--template", required=True, help="空白排程.xlsm（sheet: 空白排程）")
    ap.add_argument("--outdir", required=True, help="輸出資料夾")
    ap.add_argument("--dry-run", action="store_true", help="只回傳預覽，不寫檔")
    ap.add_argument("--year", type=int, default=dt.date.today().year)
    ap.add_argument("--max-rows-per-day", type=int, default=8)
    ap.add_argument("--days", type=int, default=6)
    args = ap.parse_args()

    first_day = mmdd_to_date(args.date, args.year)
    need_xlsx = ensure_exists(args.need)
    limit_xlsx = ensure_exists(args.limit)
    template = ensure_exists(args.template)
    outdir = ensure_exists(args.outdir, is_dir=True)

    dic_INV = read_need_excel(need_xlsx)
    dic_Prod = read_limit_excel(limit_xlsx)

    # (1) 缺口 → 批次
    batches, _overflow_theoretical = compute_batches(dic_INV, dic_Prod, days=args.days, per_day=args.max_rows_per_day)
    # (2) 人員限制 → 分到每天
    preview, overflow = assign_to_days_with_staff(
        batches,
        first_day,
        days=args.days,
        per_day=args.max_rows_per_day,
        per_staff_per_day=2
    )

    if args.dry_run:
        print(json.dumps({"ok": True, "preview": preview, "overflow": overflow}, ensure_ascii=False), flush=True)
        return

    # 正式：複製模板為唯一檔名，寫入單檔（含 6 天）
    yyyymmdd = first_day.strftime("%Y%m%d")
    out_path = unique_outpath(Path(outdir), yyyymmdd)
    try:
        copy2(template, out_path)  # 保留 VBA
    except Exception as e:
        fail(20, f"Copy template failed: {e}")

    write_to_workbook(out_path, first_day, preview, per_day=args.max_rows_per_day)

    print(json.dumps({"ok": True, "outPath": str(out_path), "scheduled": len(preview), "overflow": overflow}, ensure_ascii=False), flush=True)


if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        raise
    except Exception as e:
        print(f"Unhandled error: {e}", file=sys.stderr, flush=True)
        sys.exit(99)
