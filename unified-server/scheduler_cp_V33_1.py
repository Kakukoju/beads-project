import pandas as pd
import sqlite3
import collections
import math
import openpyxl
import os
import shutil
import tempfile
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from ortools.sat.python import cp_model

# ==============================================================================
# 1. 基礎設定 (Configuration)
# ==============================================================================

REMOTE_DB_PATH = r"\\fls341\MBBU_FAB\MB_PD\生管自動化\滴定\資料庫\beads_sync.db"
START_HOUR = 8
MINS_PER_GRID = 30
DEFAULT_GRIDS = 35 
LUNCH_BREAK_GRIDS = [8, 9] 
PREP_ONLY_ROW_START = 38 
OUTPUT_FILENAME = "滴定排程結果_V33_12.xlsx"

# 全域工具：安全數值轉換
def safe_float(val):
    try:
        if pd.isna(val) or str(val).strip() == '': return 0.0
        return float(str(val).replace(',', ''))
    except: return 0.0

# ==============================================================================
# 2. 資料載入與解析
# ==============================================================================

def load_db_data_v33_12():
    print(f"📡 正在從網路路徑讀取資料庫...")
    if not os.path.exists(REMOTE_DB_PATH): 
        print(f"❌ 找不到資料庫路徑: {REMOTE_DB_PATH}")
        return None, None, None, None
    
    temp_db = os.path.join(tempfile.gettempdir(), "beads_v33_12.db")
    shutil.copy2(REMOTE_DB_PATH, temp_db)
    conn = sqlite3.connect(temp_db)
    df_rules = pd.read_sql_query("SELECT * FROM 配藥限制", conn)
    df_demand = pd.read_sql_query("SELECT * FROM BeadNeed WHERE date = (SELECT MAX(date) FROM BeadNeed)", conn)
    conn.close()
    if os.path.exists(temp_db): os.remove(temp_db)

    rules, reagent_groups, all_staff = {}, collections.defaultdict(list), set()
    for _, row in df_rules.iterrows():
        pn = str(row.get('PN', '')).strip().replace('.0', '')
        if not pn: continue
        name = str(row.get('Name', ''))
        reagent = name.split('-')[0].strip()
        
        # IVEK 與 Port 數判定
        port_val = str(row.get('Port數', '2')).strip().upper()
        num_ports = 0 if (port_val == 'IVEK' or name.upper().startswith('NA')) else (int(port_val) if port_val.isdigit() else 2)
        
        # 數量解析
        qty_raw = str(row.get('數量', '0')).strip()
        is_prep_only = (qty_raw == '' or qty_raw == '0')
        qty_opts = [safe_float(x.strip()) for x in qty_raw.lower().replace('or', ',').split(',') if x.strip()] if not is_prep_only else [0.0]
        
        staff = [str(row.get(f'配藥人-{i}', '')).strip() for i in range(1, 4)]
        staff = [s for s in staff if s and s.lower() != 'nan']
        for s in staff: all_staff.add(s)

        rules[pn] = {
            'pn': pn, 'name': name, 'reagent': reagent, 'num_ports': num_ports,
            'qty_options': sorted(qty_opts), 'is_prep_only': is_prep_only,
            'duration_freeze': int(safe_float(row.get('凍乾時間')) * 60) if row.get('凍乾時間') else 720,
            'compatible_dryers': [d.strip() for d in str(row.get('可用凍乾機', '')).split(',') if d.strip()],
            'compatible_staff': staff
        }
        if reagent: reagent_groups[reagent].append(pn)
    
    return rules, reagent_groups, sorted(list(all_staff)), df_demand

# ==============================================================================
# 3. 核心求解器 (V33.12：修正變數提取與 Port 邏輯)
# ==============================================================================

def run_solver_v33_12(jobs, reagent_groups, all_staff, horizon_days):
    model = cp_model.CpModel()
    grids_per_day = DEFAULT_GRIDS
    horizon_grids = horizon_days * grids_per_day
    all_dryers = sorted(list(set(d for j in jobs for d in j['compatible_dryers'] if not j['is_prep_only'])))
    
    job_vars = {}
    staff_ints = collections.defaultdict(list)
    dryer_exclusive_pool = collections.defaultdict(list)
    rects = []

    # 按試劑組處理
    reagent_to_jobs = collections.defaultdict(list)
    for j in jobs: reagent_to_jobs[j['reagent']].append(j)

    for reagent, group_jobs in reagent_to_jobs.items():
        # 群組同步變數
        base_s = model.NewIntVar(0, horizon_grids, f"s_{reagent}")
        base_day = model.NewIntVar(0, horizon_days - 1, f"d_{reagent}")
        base_st = model.NewIntVar(0, len(all_staff), f"st_{reagent}")
        base_dr = model.NewIntVar(0, len(all_dryers), f"dr_{reagent}")
        
        # Port 防汙：所有組別起始 Port 為偶數 (Port 1, 3, 5...)
        p_grp_start = model.NewIntVar(0, 11, f"p_start_{reagent}")
        model.AddModuloEquality(0, p_grp_start, 2)
        model.AddDivisionEquality(base_day, base_s, grids_per_day)

        # 1. 配藥人力合併
        is_prep_only_grp = all(j['is_prep_only'] for j in group_jobs)
        staff_pool = set()
        for j in group_jobs: staff_pool.update(j['compatible_staff'])
        
        st_bools = []
        for idx, sname in enumerate(all_staff):
            if sname in staff_pool:
                u = model.NewBoolVar(f"u_st_{reagent}_{idx}")
                st_bools.append(u)
                model.Add(base_st == idx).OnlyEnforceIf(u)
                if is_prep_only_grp:
                    staff_ints[idx].append(model.NewOptionalFixedSizeIntervalVar(base_s, 6, u, ''))
                else:
                    busy_s = model.NewIntVar(-4, horizon_grids, '')
                    model.Add(busy_s == base_s - 4)
                    staff_ints[idx].append(model.NewOptionalFixedSizeIntervalVar(busy_s, 4, u, ''))
        if st_bools: model.Add(sum(st_bools) == 1)

        # 2. 凍乾機日鎖定獨佔
        dr_bools = []
        for idx, dname in enumerate(all_dryers):
            if any(dname in j['compatible_dryers'] for j in group_jobs):
                ud = model.NewBoolVar(f"ud_dr_{reagent}_{idx}")
                dr_bools.append(ud)
                model.Add(base_dr == idx).OnlyEnforceIf(ud)
                lock_start = base_day * grids_per_day
                dryer_exclusive_pool[dname].append(model.NewOptionalFixedSizeIntervalVar(lock_start, grids_per_day, ud, ''))
        if dr_bools: model.Add(sum(dr_bools) == 1)

        # 3. 任務偏移與資源綁定
        current_p_offset = 0
        for j in group_jobs:
            job_vars[j['id']] = {'s': base_s, 'day': base_day, 'st': base_st, 'dr': base_dr, 'port': p_grp_start + current_p_offset}
            
            s_in_day = model.NewIntVar(0, grids_per_day, '')
            model.AddModuloEquality(s_in_day, base_s, grids_per_day)
            model.Add(s_in_day + j['grid_duration'] <= 34) 
            model.AddLinearExpressionInDomain(s_in_day, cp_model.Domain.FromIntervals([(0, 7), (10, 34)]))

            if not j['is_prep_only']:
                sf = model.NewIntVar(0, horizon_grids + 400, f"sf_{j['id']}")
                ef = model.NewIntVar(0, horizon_grids + 500, f"ef_{j['id']}")
                job_vars[j['id']].update({'sf': sf, 'ef': ef})
                
                model.Add(sf >= base_s + j['grid_duration'] + 1)
                f_g = math.ceil((j['duration_freeze'] + 120) / 30)
                model.Add(ef == sf + f_g)
                
                num_p = max(1, j['num_ports'])
                rects.append([model.NewFixedSizeIntervalVar(base_s, j['grid_duration'], ''), 
                              model.NewFixedSizeIntervalVar(p_grp_start + current_p_offset, num_p, '')])
                current_p_offset += num_p
        model.Add(p_grp_start + current_p_offset <= 12)

    if rects: model.AddNoOverlap2D([r[0] for r in rects], [r[1] for r in rects])
    for idx in staff_ints: model.AddNoOverlap(staff_ints[idx])
    for dname in dryer_exclusive_pool: model.AddNoOverlap(dryer_exclusive_pool[dname])

    model.Minimize(sum(v['day'] * 2000 + v['s'] for v in job_vars.values()))
    
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 60.0
    if solver.Solve(model) in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
        res = []
        for j in jobs:
            v = job_vars[j['id']]
            out = {**j, 'day': solver.Value(v['day']), 'grid_start': solver.Value(v['s']) % grids_per_day, 'assigned_staff': all_staff[solver.Value(v['st'])]}
            if not j['is_prep_only']:
                out.update({
                    'p_idx': solver.Value(v['port']), 
                    'assigned_dryer': all_dryers[solver.Value(v['dr'])], 
                    'ef_abs': solver.Value(v['ef']),
                    'sf_abs': solver.Value(v['sf'])
                })
            res.append(out)
        return res
    return None

# ==============================================================================
# 4. 報表產出
# ==============================================================================

def export_excel_v33_12(results, grids_per_day):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    by_day = collections.defaultdict(list)
    for j in results: by_day[j['day']].append(j)
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill_h = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    fill_prep = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
    colors = ["E6B8AF", "F4CCCC", "FCE5CD", "FFF2CC", "D9EAD3", "D0E0E3", "C9DAF8", "CFE2F3", "D9D2E9", "EAD1DC"]
    pn_colors = {}
    
    for d in sorted(by_day.keys()):
        ws = wb.create_sheet(f"Day {d+1}")
        headers = ["時間", "IVEK 1", "IVEK 2"] + [f"Port {i}" for i in range(1, 13)] + ["純配藥區"]
        for c, txt in enumerate(headers, 1):
            cell = ws.cell(1, c, txt)
            cell.fill, cell.border, cell.alignment = fill_h, border, Alignment('center', 'center')
            ws.column_dimensions[get_column_letter(c)].width = 17
            
        for i in range(grids_per_day):
            m = START_HOUR * 60 + i * 30
            txt = f"{m//60:02d}:{m%60:02d}" if m < 1440 else f"隔{(m-1440)//60:02d}:{(m-1440)%60:02d}"
            ws.cell(i+2, 1, txt).border = border

        p_idx_prep = 0
        for j in by_day[d]:
            if j['pn'] not in pn_colors: pn_colors[j['pn']] = colors[len(pn_colors) % len(colors)]
            c_fill = PatternFill(start_color=pn_colors[j['pn']], end_color=pn_colors[j['pn']], fill_type="solid")
            
            if j['is_prep_only']:
                c_s, r_s = len(headers), PREP_ONLY_ROW_START + p_idx_prep
                txt = f"{j['pn']}\n{j['name']}\n人:{j['assigned_staff']}"; p_idx_prep += 1
                cell = ws.cell(r_s, c_s, txt)
                cell.fill, cell.border, cell.alignment = fill_prep, border, Alignment('center', 'center', wrap_text=True)
            else:
                c_s, c_e = (2, 3) if j['num_ports'] == 0 else (j['p_idx'] + 4, j['p_idx'] + 4 + max(1, j['num_ports']) - 1)
                r_s, r_e = j['grid_start'] + 2, j['grid_start'] + 2 + j['grid_duration'] - 1
                f_m = (START_HOUR * 60) + (j.get('ef_abs', 0) * 30)
                txt = f"{j['pn']}\n{j['name']}\n量:{int(j['qty'])}\n人:{j['assigned_staff']}\n機:{j['assigned_dryer']}\nEnd {f_m%1440//60:02d}:{f_m%60:02d}"
                cell = ws.cell(r_s, c_s, txt)
                cell.fill, cell.border, cell.alignment = c_fill, border, Alignment('center', 'center', wrap_text=True)
                if r_e > r_s or c_e > c_s: 
                    try: ws.merge_cells(start_row=r_s, start_column=c_s, end_row=r_e, end_column=c_e)
                    except: pass
                for r in range(r_s, r_e+1):
                    for c in range(c_s, c_e+1): ws.cell(r, c).border = border
    wb.save(OUTPUT_FILENAME)

# ==============================================================================
# 5. 主程式
# ==============================================================================

def main():
    print("--------------------------------------------------")
    print("🚀 啟動 V33.12 生產排程系統 (修正 NameError & 產量比對)")
    print("--------------------------------------------------")
    # 核心修復：確保變數名稱在 main 函數內一致
    rules, groups, staff, df_demand = load_db_data_v33_12()
    if not rules: return
    
    tasks = []
    for _, row in df_demand.iterrows(): # 修正原本 df_d 的引用錯誤
        pn = str(row.get('pn', '')).strip().replace('.0', '')
        if pn not in rules: continue
        
        stock = safe_float(row.get('stock_unstock'))
        w1, w2, w3 = safe_float(row.get('w1')), safe_float(row.get('w2')), safe_float(row.get('w3'))
        
        # 決定優先級
        prio = 1 if stock < w1 else (2 if stock < (w1+w2) else (3 if stock < (w1+w2+w3) else 0))
        if prio > 0:
            tasks.append({
                'pn': pn,
                'prio': prio,
                'shortage': max(0.0, (w1+w2+w3)-stock)
            })

    jobs = []
    for t in tasks:
        r = rules[t['pn']]
        # 大容量選擇邏輯：選擇符合需求的最合適數量 (例如需求 9400 會選 11000)
        selected_qty = r['qty_options'][-1] if r['qty_options'] else 0.0
        for opt in r['qty_options']:
            if opt >= t['shortage']: 
                selected_qty = opt
                break
        
        # 滴定時間計算
        if r['is_prep_only']:
            grid_dur = 2
        else:
            divisor = 1500 if r['num_ports'] > 0 else 800
            grid_dur = math.ceil((selected_qty / max(1, r['num_ports']) / divisor * 60 + 30) / 30)
            
        jobs.append({
            **r,
            'id': f"P{t['prio']}_{t['pn']}",
            'qty': selected_qty,
            'grid_duration': grid_dur,
            'prio': t['prio']
        })

    # 多天數策略嘗試
    for days in [5, 7, 10, 14]:
        print(f"⏳ 嘗試在 {days} 天內尋找可行解...")
        res = run_solver_v33_12(jobs, groups, staff, days)
        if res:
            export_excel_v33_12(res, DEFAULT_GRIDS)
            print(f"🎉 產出成功！輸出檔案: {OUTPUT_FILENAME}")
            return
    print("❌ 失敗：無法在 14 天內找到解。")

if __name__ == '__main__':
    main()