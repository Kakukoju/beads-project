# schedule_with_win32com.py
# -*- coding: utf-8 -*-
import csv
import os
from pathlib import Path
from datetime import date, timedelta

import pandas as pd
from openpyxl import Workbook
import pythoncom
import win32com.client as win32

# --------------------- 小工具 ---------------------
def to_int(x) -> int:
    try:
        if x is None:
            return 0
        s = str(x).replace(",", "").strip()
        if s == "" or s.lower() in ("none", "nan"):
            return 0
        return int(round(float(s)))
    except Exception:
        return 0

def safe_str(x) -> str:
    return "" if x is None else str(x).strip()

def pick_file_dialog(title: str, initialdir: Path | None = None) -> Path | None:
    """彈出選檔視窗；若失敗則回傳 None。"""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        path = filedialog.askopenfilename(
            title=title,
            initialdir=str(initialdir or Path.cwd()),
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        root.update()
        root.destroy()
        return Path(path) if path else None
    except Exception as e:
        print(f"[警告] 無法開啟檔案選擇視窗：{e}")
        return None

# --------------------- Excel COM 讀取 ---------------------
def open_excel_calculated(path: Path):
    """用 COM 開啟並強制重算；回傳 (excel_app, workbook)。"""
    pythoncom.CoInitialize()
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(str(path))
    excel.CalculateFull()   # 強制全重算
    wb.Save()               # 寫回公式結果（更穩）
    return excel, wb

def close_excel(excel, wb):
    try:
        wb.Close(SaveChanges=False)
    except Exception:
        pass
    try:
        excel.Quit()
    except Exception:
        pass
    pythoncom.CoUninitialize()

def read_need_using_com(path_need: Path, sheet_name: str):
    """
    beads 需求模組.xlsx → 《滴定排程需求表》
    固定：標題列=第2列；資料列=第3列起
    欄位：B=PN，C=品名，L/M/N=第一/二/三週需求批次
    """
    excel, wb = open_excel_calculated(path_need)
    try:
        try:
            ws = wb.Worksheets(sheet_name)
        except Exception:
            ws = wb.Worksheets(1)

        rows = []
        last_row = ws.UsedRange.Rows(ws.UsedRange.Rows.Count).Row
        last_row = max(last_row, 200)  # 至少掃 200 列，避免提早停止
        blank_streak = 0
        for r in range(3, last_row + 1):
            pn  = safe_str(ws.Cells(r, 2).Value)   # B
            name= safe_str(ws.Cells(r, 3).Value)   # C
            w1  = to_int(ws.Cells(r, 12).Value)    # L
            w2  = to_int(ws.Cells(r, 13).Value)    # M
            w3  = to_int(ws.Cells(r, 14).Value)    # N
            if pn == "" and name == "" and w1 == 0 and w2 == 0 and w3 == 0:
                blank_streak += 1
                if blank_streak >= 25:
                    break
                continue
            blank_streak = 0
            if pn:
                rows.append({"PN": pn, "Name": name, "W1": w1, "W2": w2, "W3": w3})
        return rows
    finally:
        close_excel(excel, wb)

def read_limit_using_com(path_limit: Path, sheet_name: str):
    """
    滴定限制.xlsx → 《配藥限制》
    A=PN, I/J/K=人員, L=交藥時段, O=冷卻天數, R=同日群(Name)（本版僅讀）
    """
    excel, wb = open_excel_calculated(path_limit)
    try:
        try:
            ws = wb.Worksheets(sheet_name)
        except Exception:
            ws = wb.Worksheets(1)

        limits = {}
        last_row = ws.UsedRange.Rows(ws.UsedRange.Rows.Count).Row
        for r in range(2, last_row + 1):
            pn = safe_str(ws.Cells(r, 1).Value)     # A
            if not pn:
                continue
            staff_i  = safe_str(ws.Cells(r, 9).Value)   # I
            staff_j  = safe_str(ws.Cells(r,10).Value)   # J
            staff_k  = safe_str(ws.Cells(r,11).Value)   # K
            handover = safe_str(ws.Cells(r,12).Value)   # L
            cooldown = to_int(ws.Cells(r,15).Value)     # O
            same_day = safe_str(ws.Cells(r,18).Value)   # R

            staff = ",".join([s for s in (staff_i, staff_j, staff_k) if s])
            same_day_names = [x.strip() for x in same_day.split(",") if x.strip()] if same_day else []
            limits[pn] = {
                "staff": staff,
                "handover": handover,
                "cooldown": cooldown,
                "same_day_names": same_day_names
            }
        return limits
    finally:
        close_excel(excel, wb)

# --------------------- 排程邏輯 ---------------------
def build_tasks(need_rows):
    """W1→W2→W3 優先；各欄數字大的先排；每批展開為一筆。"""
    df = pd.DataFrame(need_rows)
    tasks = []

    def push(tag, col):
        sub = df[df[col] > 0].copy()
        if sub.empty:
            return
        sub = sub.sort_values(by=[col, "PN"], ascending=[False, True])
        for _, row in sub.iterrows():
            for _ in range(int(row[col])):
                tasks.append({
                    "PN": row["PN"],
                    "Name": row["Name"],
                    "WeekSource": tag,
                    "PriorityValue": int(row[col]),
                })
    push("W1", "W1"); push("W2", "W2"); push("W3", "W3")
    return tasks

def next_monday(from_day: date) -> date:
    offset = (0 - from_day.weekday() + 7) % 7
    if offset == 0:
        offset = 7
    return from_day + timedelta(days=offset)

def schedule_week(tasks, limits, start_monday: date, slots_per_day=8):
    days = [start_monday + timedelta(days=i) for i in range(5)]
    capacity = {d: 0 for d in days}
    last_day_by_pn = {}
    scheduled, unplaced = [], []
    wd = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]

    def allowed_time(pn): 
        t = limits.get(pn, {}).get("handover", "")
        return "AM/PM" if t == "" else t
    def staff_for(pn): 
        return limits.get(pn, {}).get("staff", "")
    def cooldown_for(pn): 
        return int(limits.get(pn, {}).get("cooldown", 0) or 0)

    for t in tasks:
        pn, nm = t["PN"], t["Name"]
        cd = cooldown_for(pn)
        placed = False
        for d in days:
            if capacity[d] >= slots_per_day:
                continue
            last = last_day_by_pn.get(pn)
            if last is not None and (d - last).days <= cd:
                continue
            capacity[d] += 1
            seq = capacity[d]
            scheduled.append({
                "Date": d.strftime("%Y-%m-%d"),
                "Weekday": wd[d.weekday()],
                "Seq": seq,
                "PN": pn,
                "Name": nm,
                "WeekSource": t["WeekSource"],
                "HandoverWindow": allowed_time(pn),
                "Staff": staff_for(pn),
                "CooldownDays": cd,
            })
            last_day_by_pn[pn] = d
            placed = True
            break
        if not placed:
            unplaced.append(t)
    return scheduled, unplaced, capacity

# --------------------- 輸出 ---------------------
def write_csv(path: Path, rows, header=None):
    path.parent.mkdir(parents=True, exist_ok=True)
    if isinstance(rows, list) and rows and isinstance(rows[0], dict):
        header = header or list(rows[0].keys())
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=header)
            w.writeheader(); w.writerows(rows)
    else:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            if header: w.writerow(header)
            if isinstance(rows, list):
                for r in rows: w.writerow(r)

def write_xlsx(out_xlsx: Path, scheduled, unplaced, capacity_dict):
    wb = Workbook()
    ws_cap = wb.active; ws_cap.title = "容量概覽"
    ws_cap.append(["Date","SlotsUsed","SlotsMax"])
    for d, used in capacity_dict.items():
        ws_cap.append([d.strftime("%Y-%m-%d"), used, 8])
    if scheduled:
        ws = wb.create_sheet("建議排程")
        cols = ["Date","Weekday","Seq","PN","Name","WeekSource","HandoverWindow","Staff","CooldownDays"]
        ws.append(cols)
        for r in scheduled: ws.append([r[c] for c in cols])
    if unplaced:
        ws = wb.create_sheet("未排入")
        cols = ["PN","Name","WeekSource","PriorityValue"]
        ws.append(cols)
        for t in unplaced: ws.append([t.get(c,"") for c in cols])
    wb.save(out_xlsx)

# --------------------- 主流程（支援選檔） ---------------------
def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--need", help="beads 需求模組.xlsx 路徑", default=None)
    ap.add_argument("--limit", help="滴定限制.xlsx 路徑", default=None)
    ap.add_argument("--sheet-need", default="滴定排程需求表")
    ap.add_argument("--sheet-limit", default="配藥限制")
    args = ap.parse_args()

    script_dir = Path(__file__).resolve().parent

    need_path = Path(args.need) if args.need else pick_file_dialog("選擇《beads 需求模組.xlsx》", script_dir)
    if not need_path:
        print("[中止] 未選取需求檔。"); return
    limit_path = Path(args.limit) if args.limit else pick_file_dialog("選擇《滴定限制.xlsx》", need_path.parent)
    if not limit_path:
        print("[中止] 未選取限制檔。"); return

    if not need_path.exists():
        print(f"[錯誤] 找不到需求檔：{need_path}"); return
    if not limit_path.exists():
        print(f"[錯誤] 找不到限制檔：{limit_path}"); return

    print(f"[info] 需求檔：{need_path}")
    print(f"[info] 限制檔：{limit_path}")

    need_rows = read_need_using_com(need_path, args.sheet_need)
    limits    = read_limit_using_com(limit_path, args.sheet_limit)
    tasks     = build_tasks(need_rows)

    start     = next_monday(date.today())
    scheduled, unplaced, capacity = schedule_week(tasks, limits, start, slots_per_day=8)

    out_dir  = need_path.parent
    csv_plan = out_dir / "schedule_plan_week1_v5.csv"
    csv_un   = out_dir / "schedule_unplaced_week1_v5.csv"
    csv_cap  = out_dir / "schedule_capacity_week1_v5.csv"
    xlsx_out = out_dir / f"每週生產排程-建議排程-{start.strftime('%Y%m%d')}-v5.xlsx"

    write_csv(csv_plan, scheduled, header=["Date","Weekday","Seq","PN","Name","WeekSource","HandoverWindow","Staff","CooldownDays"])
    write_csv(csv_un,   unplaced if unplaced else [["info","本週無未排入項目"]],
              header=["PN","Name","WeekSource","PriorityValue"] if unplaced else None)
    cap_rows = [{"Date": d.strftime("%Y-%m-%d"), "SlotsUsed": used, "SlotsMax": 8} for d, used in capacity.items()]
    write_csv(csv_cap, cap_rows, header=["Date","SlotsUsed","SlotsMax"])
    write_xlsx(xlsx_out, scheduled, unplaced, capacity)

    print("✅ Done")
    print(f"- 建議排程: {csv_plan}")
    print(f"- 未排入清單: {csv_un}")
    print(f"- 容量概覽: {csv_cap}")
    print(f"- 整合 Excel: {xlsx_out}")

if __name__ == "__main__":
    main()
