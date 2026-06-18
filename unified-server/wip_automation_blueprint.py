# -*- coding: utf-8 -*-
"""
WIP 自動化管理系統 Blueprint (跨年修正版 V2)
整合 WIP 報表監控和工單統計分析功能

修正重點：
1. 寫入邏輯：根據 Excel Sheet 名稱決定年份，而非系統時間。
2. 讀取邏輯：跨表查詢 (Union Query)，同時搜尋 2024/2025/2026 等所有存在的明細表。
"""

import hashlib
import os
import sqlite3
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Dict, List, Tuple, Set
import atexit
import time
import re

from flask import Blueprint, request, jsonify
import openpyxl
import pandas as pd
from watchdog.observers.polling import PollingObserver as Observer # 改用 Polling 避免網芳權限問題
from watchdog.events import FileSystemEventHandler

# 建立 Blueprint
wip_automation_bp = Blueprint('wip_automation', __name__)

# 設定日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ==================== 設定區 ====================
# WIP 報表監控設定
EXCEL_FILE_PATH = r"\\fls341\MBBU_FAB\MB_PD\生管自動化\工單入庫\Wip_program\WIP報表 2025-QR01 NEW (請勿亂動連結).xlsm"
DB_DIR = r"/opt/beadsops/data"
DB_NAME = "Bead_Sort_DB.db"
HEADER_ROW = 5  # Excel 的 header 在第 5 行

# 工單統計設定
FORMULATE_DB_PATH = r"D:\配藥表\資料庫\P01_formualte_schedule.db"

# 全域檔案監控器
file_observer: Optional[Observer] = None
last_sync_time = 0
SYNC_COOLDOWN = 10  # 同步冷卻時間（秒）


# ==================== WIP 監控類 ====================
class WIPMonitor:
    """WIP 報表監控類"""
    
    def __init__(self, excel_path: str, db_dir: str, db_name: str):
        self.excel_path = excel_path
        self.db_dir = db_dir
        self.db_name = db_name
        self.db_path = os.path.join(db_dir, db_name)
        self._ensure_db_directory()
    
    def _ensure_db_directory(self):
        try:
            Path(self.db_dir).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            logger.error(f"建立資料庫目錄失敗: {e}")
            raise
    
    def _get_target_sheet_info(self) -> Tuple[str, str]:
        """
        偵測 Excel 中的目標工作表名稱，並解析出年份
        回傳: (sheet_name, year_str)
        """
        current_year = datetime.now().year
        # 優先搜尋的年份列表 (今年 -> 去年 -> 明年)
        target_years = [current_year, current_year - 1, current_year + 1]
        
        try:
            wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            # 1. 精確比對 "YYYY 明細"
            for y in target_years:
                candidates = [f"{y} 明細", f"{y} 明細 ", f"{y}明細"]
                for cand in candidates:
                    if cand in sheet_names:
                        return cand, str(y)
            
            # 2. 模糊比對 (包含 "明細" 且有 4 位數字)
            for sheet in sheet_names:
                if "明細" in sheet:
                    match = re.search(r"20\d{2}", sheet)
                    if match:
                        return sheet, match.group(0)

            # 3. 如果都找不到，回傳預設值
            default_sheet = f"{current_year} 明細 "
            return default_sheet, str(current_year)

        except Exception as e:
            logger.warning(f"無法自動偵測工作表名稱: {e}")
            return f"{current_year} 明細 ", str(current_year)
    
    def _read_excel_data(self) -> Tuple[Optional[pd.DataFrame], str]:
        """讀取 Excel，A 欄由下往上找最後一筆有效 row"""

        try:
            if not os.path.exists(self.excel_path):
                logger.error(f"Excel 檔案不存在: {self.excel_path}")
                return None, ""

            sheet_name, year_str = self._get_target_sheet_info()
            logger.info(
                f"開始讀取 Excel: {self.excel_path}, "
                f"目標工作表: {sheet_name} (年份: {year_str})"
            )

            # --- 1️⃣ 先用 openpyxl 找最後一 row（A 欄） ---
            wb = openpyxl.load_workbook(self.excel_path, data_only=True, read_only=True)
            ws = wb[sheet_name]

            last_row = None
            for r in range(ws.max_row, HEADER_ROW, -1):
                v = ws[f"A{r}"].value
                if v not in (None, "", "nan"):
                    last_row = r
                    break

            wb.close()

            if not last_row or last_row < HEADER_ROW:
                logger.warning("⚠️ 找不到有效資料列")
                return None, ""

            logger.info(f"📌 偵測最後有效資料列: Row {last_row}")

            # --- 2️⃣ 再用 pandas 只讀有效範圍 ---
            nrows = last_row - (HEADER_ROW - 1)

            df = pd.read_excel(
                self.excel_path,
                sheet_name=sheet_name,
                header=HEADER_ROW - 1,
                usecols="A:U",
                nrows=nrows,
                engine="openpyxl",
                dtype=str
            )

            # 清理全空列（保險）
            df = df.dropna(how="all")

            df = self._convert_data_types(df)
            df["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            logger.info(f"✅ 成功讀取 {len(df)} 筆有效資料")
            return df, year_str

        except Exception as e:
            logger.error(f"❌ 讀取 Excel 失敗: {e}")
            return None, ""

    
    def _convert_data_types(self, df: pd.DataFrame) -> pd.DataFrame:
        text_cols = ['LOT NO', '料號', '工單號碼', '品名', '狀態', '備註', '差異說明']
        date_cols = ['滴定日期', '入庫日期', '警示日期', '藥劑效期']
        num_cols = ['工單數', '滴定數(扣除秤重)', '分裝數量', '實際入庫數量', '差異\n(負數為紅色)', '藥劑保存(月)', 'QA 檢驗', 'PE登記NG數']
        
        # 欄位正規化 (去除前後空白)
        df.columns = [str(c).strip() for c in df.columns]

        for col in date_cols:
            if col in df.columns:
                df[col] = self._format_date_column(df[col])
        for col in num_cols:
            if col in df.columns:
                df[col] = self._format_numeric_column(df[col])
        for col in text_cols:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip().replace('nan', '')
        if '料號' in df.columns:
            df['料號'] = df['料號'].apply(self._format_material_number)
        return df
    
    def _format_date_column(self, series: pd.Series) -> pd.Series:
        def format_date(value):
            if pd.isna(value) or value == '' or str(value).lower() == 'nan':
                return None

            try:
                # Excel serial number
                if isinstance(value, (int, float)):
                    dt = pd.to_datetime(
                        value,
                        unit='D',
                        origin='1899-12-30',
                        errors='coerce'
                    )
                else:
                    # 字串：只取日期部分，直接斬掉時間
                    s = str(value).strip()
                    s = s.split(" ")[0]   # ← 關鍵：砍掉 00:00:00
                    dt = pd.to_datetime(s, errors='coerce')

                if pd.isna(dt):
                    return None

                # 🔴 強制輸出 YYYY-MM-DD
                return dt.strftime('%Y-%m-%d')

            except Exception as e:
                logger.warning(f"日期轉換失敗: {value} ({e})")
                return None

        return series.apply(format_date)


    
    def _format_numeric_column(self, series: pd.Series) -> pd.Series:
        result = []
        for value in series:
            if pd.isna(value) or value == '' or str(value).lower() == 'nan':
                result.append(None)
            else:
                try:
                    # 嘗試處理 "1,234" 或 "1234.0"
                    val_str = str(value).replace(',', '')
                    result.append(int(round(float(val_str))))
                except:
                    result.append(None)
        return pd.Series(result, dtype='Int64')
    
    def _format_material_number(self, value) -> str:
        if pd.isna(value) or value == '' or str(value).lower() == 'nan': return ''
        s = str(value)
        return s[:-2] if s.endswith('.0') else s.strip()
    
    def _create_table_if_not_exists(self, conn: sqlite3.Connection, df: pd.DataFrame, table_name: str):
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        if cursor.fetchone() is None:
            logger.info(f"建立新資料表: {table_name}")
            # 使用 pandas 自動建表，但先寫入空資料
            df.head(0).to_sql(table_name, conn, if_exists='replace', index=False)

    def _create_sync_log_table(self, conn: sqlite3.Connection):
        conn.execute("""
            CREATE TABLE IF NOT EXISTS wip_sync_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                excel_file TEXT,
                sheet_name TEXT,
                year TEXT,
                row_count INTEGER,
                trigger_type TEXT,   -- auto / manual
                sync_time TEXT,
                status TEXT,         -- success / failed
                message TEXT
            )
        """)
        conn.commit()

    def _create_meta_table(self, conn):
        conn.execute("""
            CREATE TABLE IF NOT EXISTS wip_meta (
                key TEXT PRIMARY KEY,
                value TEXT,
                updated_at TEXT
            )
        """)
        conn.commit()


    def _generate_df_fingerprint(self, df: pd.DataFrame) -> str:
        """
        對 DataFrame 內容產生穩定 fingerprint
        """
        if df.empty:
            return ""

        # 選擇「真正代表業務意義」的欄位
        core_cols = [
            "LOT NO",
            "工單號碼",
            "狀態",
            "入庫日期",
            "實際入庫數量"
        ]

        exist_cols = [c for c in core_cols if c in df.columns]
        if not exist_cols:
            return ""

        # 只取核心欄位、排序，避免 row 順序影響
        stable_df = (
            df[exist_cols]
            .fillna("")
            .astype(str)
            .sort_values(by=exist_cols)
        )

        raw = "|".join(stable_df.to_numpy().flatten())
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()

    
    def _update_database(self, df: pd.DataFrame, year_str: str, trigger_type="auto"):
        if not year_str:
            year_str = str(datetime.now().year)

        table_name = "明細_2025"

        with sqlite3.connect(self.db_path) as conn:
            self._create_table_if_not_exists(conn, df, table_name)
            self._create_sync_log_table(conn)
            self._create_meta_table(conn)

            # 🔑 1. 產生 fingerprint
            new_fp = self._generate_df_fingerprint(df)

            cur = conn.cursor()
            cur.execute("SELECT value FROM wip_meta WHERE key='last_fingerprint'")
            row = cur.fetchone()
            old_fp = row[0] if row else None

            # 🚫 2. 如果沒變 → 直接跳過
            if old_fp == new_fp:
                logger.info("⏭️ Excel 內容未變，略過 DB 寫入")
                return "skipped"

            # ✅ 3. 有變 → 寫 DB
            df.to_sql(table_name, conn, if_exists="replace", index=False)

            # 更新 fingerprint
            conn.execute("""
                INSERT OR REPLACE INTO wip_meta (key, value, updated_at)
                VALUES (?, ?, ?)
            """, (
                "last_fingerprint",
                new_fp,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ))

            # 寫 sync log
            conn.execute("""
                INSERT INTO wip_sync_log (
                    excel_file, sheet_name, year, row_count,
                    trigger_type, sync_time, status, message
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                os.path.basename(self.excel_path),
                table_name,
                year_str,
                len(df),
                trigger_type,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "success",
                "DB updated (content changed)"
            ))

            conn.commit()
            logger.info(f"✅ DB 已更新（內容變更，{len(df)} 筆）")
            return "updated"

    
    def sync_data(self, trigger_type="auto"):
        try:
            logger.info(f"🔄 WIP sync start ({trigger_type})")
            df, year_str = self._read_excel_data()
            if df is None or df.empty:
                logger.warning("⚠️ 無資料，略過")
                return False

            result = self._update_database(df, year_str, trigger_type)

            if result == "skipped":
                logger.info("🟡 Excel 無實質變更，DB 未重寫")
            else:
                logger.info("🟢 Excel 有變更，DB 已同步")

            return True

        except Exception as e:
            logger.error(f"❌ 同步失敗: {e}")
            return False



    def _create_ignore_table_if_not_exists(self, conn: sqlite3.Connection):
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS ignored_orders (
                work_order TEXT PRIMARY KEY,
                reason TEXT,
                created_at TEXT
            )
        """)
        conn.commit()

    def get_ignored_orders(self) -> set:
        try:
            with sqlite3.connect(self.db_path) as conn:
                self._create_ignore_table_if_not_exists(conn)
                cursor = conn.cursor()
                cursor.execute("SELECT work_order FROM ignored_orders")
                return {row[0] for row in cursor.fetchall()}
        except Exception as e:
            logger.error(f"讀取忽略清單失敗: {e}")
            return set()

    def toggle_ignore_order(self, work_order: str, ignore: bool) -> bool:
        try:
            with sqlite3.connect(self.db_path) as conn:
                self._create_ignore_table_if_not_exists(conn)
                cursor = conn.cursor()
                if ignore:
                    cursor.execute("INSERT OR REPLACE INTO ignored_orders (work_order, created_at) VALUES (?, ?)", (work_order, datetime.now().isoformat()))
                else:
                    cursor.execute("DELETE FROM ignored_orders WHERE work_order = ?", (work_order,))
                conn.commit()
            return True
        except Exception as e:
            logger.error(f"切換忽略狀態失敗: {e}")
            return False

# ==================== 工單統計分析類 (修正版 V3) ====================
class WorkOrderAnalyzer:
    def __init__(self, formulate_db_path: str, wip_db_path: str):
        self.formulate_db_path = formulate_db_path
        self.wip_db_path = wip_db_path
    
    def get_all_wip_tables(self) -> List[str]:
        """取得資料庫中所有 '明細_YYYY' 格式的資料表"""
        try:
            if not os.path.exists(self.wip_db_path):
                return []
            
            with sqlite3.connect(self.wip_db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                tables = [row[0] for row in cursor.fetchall()]
                # 過濾出符合格式的表
                wip_tables = [t for t in tables if t.startswith("明細_") and t.split('_')[-1].isdigit()]
                return wip_tables
        except Exception as e:
            logger.error(f"讀取 WIP 資料表列表失敗: {e}")
            return []

    def get_produced_orders_from_formulate(self, days: int, end_date: datetime = None) -> Tuple[Dict[str, dict], datetime]:
        """
        [修正版] 取得生產工單
        修正重點：將日期過濾邏輯移至 Python 層，避免 SQL 字串比對錯誤 (例如 '2025/9' > '2025/12')
        """
        if end_date is None:
            end_date = datetime.now().replace(hour=23, minute=59, second=59, microsecond=999999)
        
        # 設定起始時間 (時分秒歸零)
        start_date = (end_date - timedelta(days=days)).replace(hour=0, minute=0, second=0, microsecond=0)
        
        logger.info(f"🔍 [Python Filter] 查詢範圍: {start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}")

        try:
            if not os.path.exists(self.formulate_db_path):
                logger.error(f"❌ 配藥表資料庫不存在")
                return {}, end_date

            with sqlite3.connect(self.formulate_db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                
                # 1. 撈取所有 TMRA 工單 (不設日期條件，避免 SQL 誤判)
                query = """
                    SELECT TRIM(WorkOrder) as WorkOrder, Lot, Marker, MAX(Date) as Date
                    FROM DropletSchedule
                    WHERE WorkOrder LIKE 'TMRA%'
                    GROUP BY TRIM(WorkOrder)
                """
                cursor.execute(query)
                rows = cursor.fetchall()
                
                result = {}
                count_total = 0
                count_valid = 0
                
                # 2. 在 Python 中進行精準日期比對
                for row in rows:
                    count_total += 1
                    wo = row['WorkOrder'].strip()
                    date_str = str(row['Date']).strip()
                    
                    if not wo or not date_str:
                        continue
                        
                    try:
                        # 嘗試解析多種日期格式
                        dt_obj = None
                        # 格式 A: 2025-01-01
                        if '-' in date_str:
                            dt_obj = datetime.strptime(date_str.split(' ')[0], "%Y-%m-%d")
                        # 格式 B: 2025/01/01
                        elif '/' in date_str:
                            dt_obj = datetime.strptime(date_str.split(' ')[0], "%Y/%m/%d")
                        # 格式 C: 20250101
                        elif len(date_str) == 8 and date_str.isdigit():
                            dt_obj = datetime.strptime(date_str, "%Y%m%d")
                            
                        if dt_obj:
                            # 關鍵判定：是否在範圍內
                            if start_date <= dt_obj <= end_date:
                                result[wo] = {
                                    'WorkOrder': wo, 
                                    'Lot': row['Lot'], 
                                    'Marker': row['Marker'], 
                                    'Date': date_str # 保持原始字串顯示
                                }
                                count_valid += 1
                    except ValueError:
                        continue # 日期格式錯誤則忽略

                logger.info(f"📊 日期過濾結果: 總數 {count_total} -> 符合區間 {count_valid} 筆")
                return result, end_date

        except Exception as e:
            logger.error(f"❌ 查詢配藥表資料庫失敗: {e}")
            raise

    def get_packaged_status_from_wip(self, work_orders: List[str]) -> List[str]:
        """跨表查詢：在所有 '明細_YYYY' 表中尋找工單是否入庫"""
        if not work_orders: return []
        
        wip_tables = self.get_all_wip_tables()
        if not wip_tables:
            logger.warning("⚠️ WIP 資料庫中沒有任何明細資料表")
            return []

        try:
            with sqlite3.connect(self.wip_db_path) as conn:
                cursor = conn.cursor()
                packaged_orders = set()
                
                # 分批查詢
                chunk_size = 900
                for i in range(0, len(work_orders), chunk_size):
                    chunk = work_orders[i:i + chunk_size]
                    placeholders = ','.join(['?'] * len(chunk))
                    
                    queries = []
                    for table in wip_tables:
                        queries.append(f"""
                            SELECT TRIM(工單號碼) FROM "{table}"
                            WHERE TRIM(工單號碼) IN ({placeholders})
                              AND "入庫日期" IS NOT NULL 
                              AND TRIM("入庫日期") != '' 
                              AND TRIM("入庫日期") != 'nan'
                        """)
                    
                    if not queries: break
                    final_query = " UNION ".join(queries)
                    params = chunk * len(wip_tables)
                    
                    cursor.execute(final_query, params)
                    packaged_orders.update([row[0] for row in cursor.fetchall()])
                
                return list(packaged_orders)
        except Exception as e:
            logger.error(f"❌ 查詢 WIP 入庫狀態失敗: {e}")
            return []

    def calculate_unpackaged_ratio(self, days: int, end_date: datetime = None) -> Dict:
        try:
            # 1. 生產 (Produced) - 從配藥表 (Python 過濾日期)
            formulate_data, _ = self.get_produced_orders_from_formulate(days, end_date)
            formulate_orders_list = list(formulate_data.keys())
            
            # 2. 入庫 (Packaged) - 從 WIP DB (跨年查詢)
            packaged_orders = self.get_packaged_status_from_wip(formulate_orders_list)
            
            formulate_set = set(formulate_orders_list)
            packaged_set = set(packaged_orders)
            
            # 3. 忽略 (Ignored)
            db_dir = os.path.dirname(self.wip_db_path) or '.'
            monitor = WIPMonitor(EXCEL_FILE_PATH, db_dir, os.path.basename(self.wip_db_path))
            ignored_orders = monitor.get_ignored_orders()
            
            # 4. 未入庫 (Produced - Packaged - Ignored)
            raw_unpackaged = formulate_set - packaged_set
            final_unpackaged = raw_unpackaged - ignored_orders
            
            unpackaged_details = []
            for wo in final_unpackaged:
                if wo in formulate_data: unpackaged_details.append(formulate_data[wo])
            
            # 按日期排序
            unpackaged_details.sort(key=lambda x: str(x.get('Date', '')))

            produced_count = len(formulate_set)
            packaged_count = len(packaged_set)
            unpackaged_count = len(final_unpackaged)
            ratio = (unpackaged_count / produced_count) if produced_count > 0 else 0.0

            return {
                'success': True,
                'statistics': {
                    'produced_count': produced_count,
                    'packaged_count': packaged_count,
                    'unpackaged_count': unpackaged_count,
                    'unpackaged_ratio': round(ratio, 4),
                    'unpackaged_percentage': round(ratio * 100, 2)
                },
                'work_orders': {
                    'unpackaged_details': unpackaged_details,
                    'ignored_count': len(raw_unpackaged) - len(final_unpackaged)
                }
            }
        except Exception as e:
            logger.error(f"❌ 計算統計資料失敗: {e}")
            raise
    def calculate_overall_wip_yield_tmr(self) -> Dict:
        """
        全部 WIP 整體 Yield
        規則：
        - 工單號碼 LIKE 'TMR%'
        - 唯一鍵 = LOT NO + 工單號碼
        - 狀態為空 = 製程中 (in-progress)，不計入良率分母
        - 狀態 = '入庫完成' 為 OK，其餘非空狀態為 Fail
        """
        wip_tables = self.get_all_wip_tables()
        if not wip_tables:
            return {
                "overall_yield": 0.0,
                "total": 0,
                "ok": 0,
                "fail": 0,
                "in_progress": 0,
                "base_date": None
            }

        ok_keys = set()
        fail_keys = set()
        in_progress_keys = set()
        latest_date = None

        with sqlite3.connect(self.wip_db_path) as conn:
            cur = conn.cursor()

            for table in wip_tables:
                cur.execute(f"""
                    SELECT
                        TRIM("LOT NO"),
                        TRIM("工單號碼"),
                        TRIM("狀態"),
                        "入庫日期"
                    FROM "{table}"
                    WHERE TRIM("工單號碼") LIKE 'TMR%'
                """)

                for lot, wo, status, inbound_date in cur.fetchall():
                    if not lot or not wo:
                        continue

                    key = f"{lot}__{wo}"
                    status = (status or "").strip()

                    if not status:
                        # 空狀態 = 製程中，不計入良率
                        in_progress_keys.add(key)
                    elif status == "入庫完成":
                        ok_keys.add(key)
                        if inbound_date:
                            if latest_date is None or inbound_date > latest_date:
                                latest_date = inbound_date
                    else:
                        fail_keys.add(key)

        ok = len(ok_keys)
        fail = len(fail_keys)
        total = ok + fail  # 只計算有明確結果的工單
        in_progress = len(in_progress_keys - ok_keys - fail_keys)
        yield_rate = (ok / total * 100) if total > 0 else 0.0

        return {
            "overall_yield": round(yield_rate, 2),
            "total": total,
            "ok": ok,
            "fail": fail,
            "in_progress": in_progress,
            "base_date": latest_date
        }


# ==================== 檔案監控 ====================
class ExcelFileHandler(FileSystemEventHandler):
    def __init__(self, excel_path: str, db_dir: str, db_name: str):
        self.excel_path = excel_path
        self.monitor = WIPMonitor(excel_path, db_dir, db_name)
    
    def on_modified(self, event):
        global last_sync_time
        if event.src_path.lower().endswith(('.xlsm', '.xlsx')):
            # 簡單判斷檔名是否相符
            if os.path.basename(event.src_path) == os.path.basename(self.excel_path):
                if time.time() - last_sync_time < SYNC_COOLDOWN: return
                logger.info(f"偵測到檔案變動: {event.src_path}")
                time.sleep(2) # 等待檔案寫入完成
                try:
                    if self.monitor.sync_data(trigger_type="auto"):
                        last_sync_time = time.time()
                except Exception as e:
                    logger.error(f"自動同步錯誤: {e}")

def start_file_monitoring():
    global file_observer
    try:
        excel_dir = os.path.dirname(EXCEL_FILE_PATH)
        if not excel_dir or not os.path.exists(excel_dir): return False
        
        observer = Observer() # 使用 PollingObserver
        observer.schedule(ExcelFileHandler(EXCEL_FILE_PATH, DB_DIR, DB_NAME), excel_dir, recursive=False)
        observer.start()
        file_observer = observer
        logger.info(f"✅ 檔案監控已啟動: {excel_dir}")
        return True
    except Exception as e:
        logger.error(f"啟動監控失敗: {e}")
        return False

def stop_file_monitoring():
    if file_observer: 
        logger.info("停止檔案監控...")
        file_observer.stop()
        file_observer.join()

def on_shutdown(): 
    stop_file_monitoring()


# ==================== Flask API ====================

@wip_automation_bp.route('/api/wip/sync', methods=['POST'])
def manual_sync():
    try:
        monitor = WIPMonitor(EXCEL_FILE_PATH, DB_DIR, DB_NAME)

        # 🔴 不要直接跑，丟到背景 thread
        import threading
        t = threading.Thread(
            target=monitor.sync_data,
            kwargs={"trigger_type": "manual"},
            daemon=True
        )
        t.start()

        return jsonify({
            "status": "accepted",
            "message": "sync triggered (background)"
        }), 202

    except Exception as e:
        logger.exception("manual sync crash")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@wip_automation_bp.route('/api/wip/status', methods=['GET'])
def get_wip_status():
    try:
        monitor = WIPMonitor(EXCEL_FILE_PATH, DB_DIR, DB_NAME)
        return jsonify({'status': 'running', 'database': monitor.db_path}), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@wip_automation_bp.route('/api/workorder/unpackaged-ratio', methods=['GET'])
def get_unpackaged_ratio():
    try:
        days = request.args.get('days', default=7, type=int)
        analyzer = WorkOrderAnalyzer(FORMULATE_DB_PATH, os.path.join(DB_DIR, DB_NAME))
        return jsonify(analyzer.calculate_unpackaged_ratio(days)), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@wip_automation_bp.route('/api/workorder/unpackaged-ratio-stats', methods=['GET'])
def get_unpackaged_ratio_stats():
    """取得未入庫工單比例統計 (解決 404 的關鍵路由)"""
    try:
        wip_db_path = os.path.join(DB_DIR, DB_NAME)
        analyzer = WorkOrderAnalyzer(FORMULATE_DB_PATH, wip_db_path)
        
        weekly = analyzer.calculate_unpackaged_ratio(7)
        monthly = analyzer.calculate_unpackaged_ratio(30)
        quarterly = analyzer.calculate_unpackaged_ratio(90)
        
        return jsonify({
            'success': True,
            'weekly': {
                'unpackaged': weekly['statistics']['unpackaged_count'],
                'produced': weekly['statistics']['produced_count'],
                'packaged': weekly['statistics']['packaged_count']
            },
            'monthly': {
                'unpackaged': monthly['statistics']['unpackaged_count'],
                'produced': monthly['statistics']['produced_count'],
                'packaged': monthly['statistics']['packaged_count']
            },
            'quarterly': {
                'unpackaged': quarterly['statistics']['unpackaged_count'],
                'produced': quarterly['statistics']['produced_count'],
                'packaged': quarterly['statistics']['packaged_count']
            },
            'timestamp': datetime.now().isoformat()
        }), 200
    except Exception as e:
        logger.error(f"統計 API 錯誤: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@wip_automation_bp.route('/api/workorder/ignore', methods=['POST'])
def toggle_workorder_ignore():
    try:
        data = request.json
        monitor = WIPMonitor(EXCEL_FILE_PATH, DB_DIR, DB_NAME)
        success = monitor.toggle_ignore_order(data.get('work_order'), data.get('ignore', True))
        return jsonify({'success': success})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    
# ==================== 工單 QR 追蹤 ====================
_WO_DB = "/opt/beadsops/data/work_orders.db"

_STATION_CFG = [
    ("收藥",    "時間_收藥",      "收藥_上傳者",    "收藥_照片"),
    ("滴定準備", "時間_滴定準備",  "滴定準備_上傳者", "滴定準備_照片"),
    ("滴定開始", "時間_滴定開始",  "滴定開始_上傳者", "滴定開始_照片"),
    ("滴定結束", "時間_滴定結束",  "滴定結束_上傳者", "滴定結束_照片"),
    ("凍乾準備", "時間_凍乾準備",  "凍乾準備_上傳者", "凍乾準備_照片"),
    ("凍乾開始", "時間_凍乾開始",  "凍乾開始_上傳者", "凍乾開始_照片"),
    ("凍乾結束", "時間_凍乾結束",  "凍乾結束_上傳者", "凍乾結束_照片"),
]

@wip_automation_bp.route('/api/workorder/qr-tracking', methods=['GET'])
def qr_tracking():
    try:
        start       = request.args.get('start', '')
        end         = request.args.get('end', '')
        wo_filter   = request.args.get('workOrder', '').strip()
        bead_filter = request.args.get('beadName', '').strip()
        incomplete_only = request.args.get('incompleteOnly', '') == '1'

        where, params = [], []
        if start:
            where.append("REPLACE(\"日期\", '/', '-') >= ?"); params.append(start)
        if end:
            where.append("REPLACE(\"日期\", '/', '-') <= ?"); params.append(end)
        if wo_filter:
            where.append("\"工單號\" LIKE ?"); params.append(f'%{wo_filter}%')
        if bead_filter:
            where.append("(bead_name LIKE ? OR \"PN\" LIKE ?)")
            params.extend([f'%{bead_filter}%', f'%{bead_filter}%'])

        sql = "SELECT \"工單號\", \"製令數量\", bead_name, \"PN\", \"日期\""
        for _, t, u, p in _STATION_CFG:
            sql += f", \"{t}\", \"{u}\", \"{p}\""
        sql += " FROM work_orders"
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY REPLACE(\"日期\", '/', '-') DESC, \"時間_收藥\" DESC"

        with sqlite3.connect(_WO_DB) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(sql, params).fetchall()

        result = []
        for row in rows:
            stations, progress = [], 0
            for i, (name, t_col, u_col, p_col) in enumerate(_STATION_CFG):
                t = row[t_col] or None
                u = row[u_col] or None
                photos = [x.strip() for x in (row[p_col] or '').split(';') if x.strip()]
                if t:
                    progress = i + 1
                stations.append({'name': name, 'time': t, 'uploader': u, 'photos': photos})

            if incomplete_only and progress == 7:
                continue

            result.append({
                'workOrder':       row['工單號'],
                'quantity':        row['製令數量'],
                'beadName':        row['bead_name'] or '',
                'marker':          row['PN'] or '',
                'date':            row['日期'],
                'stations':        stations,
                'progress':        progress,
                'progressPercent': int(progress / 7 * 100),
                'currentStation':  stations[progress - 1]['name'] if progress > 0 else '未開始',
            })

        return jsonify({'success': True, 'data': result, 'total': len(result)})
    except Exception as e:
        logger.error(f"qr-tracking error: {e}")
        return jsonify({'success': False, 'error': str(e), 'data': [], 'total': 0}), 500


@wip_automation_bp.route('/api/workorder/qr-stats', methods=['GET'])
def qr_stats():
    try:
        period = request.args.get('period', 'week')
        if period == 'week':
            grp = "strftime('%Y-W%W', REPLACE(\"日期\", '/', '-'))"
        elif period == 'month':
            grp = "strftime('%Y-%m', REPLACE(\"日期\", '/', '-'))"
        elif period == 'quarter':
            grp = ("strftime('%Y', REPLACE(\"日期\", '/', '-')) || '-Q' || "
                   "((CAST(strftime('%m', REPLACE(\"日期\", '/', '-')) AS INTEGER) + 2) / 3)")
        else:
            grp = "strftime('%Y', REPLACE(\"日期\", '/', '-'))"

        sql = (f"SELECT bead_name, {grp} AS period, COUNT(*) AS cnt "
               f"FROM work_orders "
               f"WHERE bead_name IS NOT NULL AND bead_name != '' "
               f"  AND \"日期\" IS NOT NULL AND \"日期\" != '' "
               f"GROUP BY bead_name, period ORDER BY period DESC, cnt DESC")

        with sqlite3.connect(_WO_DB) as conn:
            rows = conn.execute(sql).fetchall()

        bead_totals, period_set, cells = {}, set(), []
        for bead, p, cnt in rows:
            bead_totals[bead] = bead_totals.get(bead, 0) + cnt
            period_set.add(p)
            cells.append({'beadName': bead, 'period': p, 'count': cnt})

        top_beads = sorted(bead_totals, key=lambda b: bead_totals[b], reverse=True)[:24]
        top_set   = set(top_beads)
        periods   = sorted(period_set, reverse=True)[:12]
        per_set   = set(periods)

        return jsonify({
            'success': True,
            'periods': periods,
            'beads':   top_beads,
            'cells':   [c for c in cells if c['beadName'] in top_set and c['period'] in per_set],
        })
    except Exception as e:
        logger.error(f"qr-stats error: {e}")
        return jsonify({'success': False, 'error': str(e), 'periods': [], 'beads': [], 'cells': []}), 500


print("🔥 WIP yield route file loaded")
@wip_automation_bp.route('/api/wip/yield-overall-tmr', methods=['GET'])
def get_overall_wip_yield_tmr():
    try:
        analyzer = WorkOrderAnalyzer(
            FORMULATE_DB_PATH,
            os.path.join(DB_DIR, DB_NAME)
        )
        return jsonify({
            "success": True,
            **analyzer.calculate_overall_wip_yield_tmr()
        }), 200
    except Exception as e:
        logger.error(f"TMR overall yield error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    


# ==================== 初始化 ====================
def init_wip_automation(app):
    with app.app_context():
        try:
            start_file_monitoring()
            atexit.register(on_shutdown)
        except Exception as e:
            logger.error(f"WIP 初始化失敗: {e}")