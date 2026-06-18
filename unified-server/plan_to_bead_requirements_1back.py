# -*- coding: utf-8 -*-
import sqlite3, pandas as pd, openpyxl, logging, re, sys
from datetime import datetime, date, timedelta
from collections import defaultdict
from typing import Dict, List, Any, Optional, Tuple
from openpyxl.utils import column_index_from_string

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# ---------- 通用小工具 ----------
def _to_number(x):
    if x is None:
        return 0
    if isinstance(x, (int, float)):
        return x
    s = str(x).strip().replace(",", "").replace("\u00A0", "").replace("\u200b", "")
    if s == "":
        return 0
    try:
        return float(s) if "." in s else int(s)
    except Exception:
        return 0

def _normalize_pn(x: object) -> str:
    if x is None:
        return ""
    if isinstance(x, int):
        return str(x)
    if isinstance(x, float):
        return str(int(x)) if x.is_integer() else format(x, ".15g")
    s = str(x).strip().replace("\u00A0", "").replace("\u200b", "")
    m = re.fullmatch(r"(\d+)(?:\.0+)?", s)
    if m:
        return m.group(1)
    # 科學記號
    if re.fullmatch(r"[0-9]+(\.[0-9]+)?[eE][+\-]?[0-9]+", s):
        try:
            f = float(s)
            return str(int(f)) if f.is_integer() else format(f, ".15g")
        except Exception:
            pass
    return s

def try_parse_date_header(s: str) -> Optional[date]:
    if s is None:
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

def is_week_col(s: str) -> bool:
    return bool(re.fullmatch(r"(?i)WK\d{1,2}", str(s).strip())) if s is not None else False

# ---------- 解析 Panel 明細：成品 PN → 半品清單 ----------
def process_panel_details(file_path: str, sheet_name: str, start_col_letter: str = "D") -> Dict[str, List[Dict[str, Any]]]:
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"找不到工作表：{sheet_name}")
        ws = wb[sheet_name]

        start_col = column_index_from_string(start_col_letter)
        max_col, max_row = ws.max_column, ws.max_row

        panel_map: Dict[str, List[Dict[str, Any]]] = {}
        row = 1
        while row + 2 <= max_row:
            # 這三列：第一列 = 各半品 PN；第二列 = 名稱；第三列 = 數量
            # 同時 A 欄（第 1 欄）若有「成品 PN」，就把該 3 列對應的 bead 列表掛給那個成品 PN
            pns = [_normalize_pn(ws.cell(r, 1).value) for r in (row, row + 1, row + 2)]
            if not any(pns):
                row += 1
                continue

            beads: List[Dict[str, Any]] = []
            col = start_col
            while col <= max_col:
                bead_pn = ws.cell(row, col).value
                if bead_pn is None or str(bead_pn).strip() == "":
                    # 一旦遇到空白即視為本區塊結束
                    break
                bead_name = ws.cell(row + 1, col).value
                qty_val   = ws.cell(row + 2, col).value
                beads.append({
                    "bead_pn": str(bead_pn).strip(),
                    "bead_name": (str(bead_name).strip() if bead_name is not None else ""),
                    "quantity": _to_number(qty_val),
                })
                col += 1

            for pn in pns:
                if pn:
                    panel_map[pn] = beads

            row += 3

        logging.info(f"Panel 解析完成，產品數：{len(panel_map)}")
        return panel_map
    finally:
        wb.close()

# ---------- 從 DB 取得三週成品量（Plan='Plan'、TTL 規則、WKxx 排除、負數檢查） ----------
def get_weekly_production(db_path: str, table: str, start_mmdd: str, year: Optional[int] = None) -> Dict[str, Tuple[int,int,int]]:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # 欄位列表
    cur.execute(f"PRAGMA table_info({table})")
    col_rows = cur.fetchall()
    if not col_rows:
        raise RuntimeError(f"資料表不存在或無法讀取：{table}")
    col_names = [row["name"] for row in col_rows]

    # 找 PN 與 Plan 欄
    pn_col   = next((c for c in col_names if c.lower() == "pn"), None)
    plan_col = next((c for c in col_names if c.lower() == "plan"), None)
    if not pn_col or not plan_col:
        raise RuntimeError(f"缺 PN 或 Plan 欄；現有欄位：{col_names}")

    plan_idx = col_names.index(plan_col)

    # Plan 後方的「日期欄」：排除 WKxx，保留可 parse 的日期字串欄名
    date_cols = [c for c in col_names[plan_idx+1:] if not is_week_col(c) and try_parse_date_header(c)]
    if not date_cols:
        raise RuntimeError("Plan 後沒有可解析的日期欄")

    years = sorted({try_parse_date_header(c).year for c in date_cols})
    if not years:
        raise RuntimeError("無法由日期欄判斷年份")

    # 年份預設抓「最大年份」，通常是最新
    if year is None:
        year = max(years)
    logging.info(f"使用年份：{year}")

    mm, dd = map(int, start_mmdd.split("/"))
    start_date = date(year, mm, dd)
    w1, w2, w3, end_dt = start_date, start_date + timedelta(days=7), start_date + timedelta(days=14), start_date + timedelta(days=21)

    c2d = {c: try_parse_date_header(c) for c in date_cols}
    win1 = [c for c, d in c2d.items() if d and w1 <= d < w2]
    win2 = [c for c, d in c2d.items() if d and w2 <= d < w3]
    win3 = [c for c, d in c2d.items() if d and w3 <= d < end_dt]

    if not win1 or not win2 or not win3:
        raise RuntimeError(f"三週視窗有空集：win1={len(win1)}, win2={len(win2)}, win3={len(win3)}；請確認 start_mmdd={start_mmdd} 與年份={year} 及資料欄位。")

    cur.execute(f"SELECT rowid AS _rid, * FROM {table} ORDER BY rowid")
    rows = cur.fetchall()

    # 找到 TTL 切點（由下往上找最後一個 'plan ttl'）
    ttl_cut = None
    for r in reversed(rows):
        v = r[plan_col]
        if isinstance(v, str) and v.strip().lower() == "plan ttl":
            ttl_cut = r["_rid"]
            break

    # 只取 TTL 以上
    rows_above = [r for r in rows if (ttl_cut is None or r["_rid"] < ttl_cut)]

    # 跳過 TTL 行與其下兩行（保留你的原規則）
    skip_rids = set()
    if ttl_cut is not None:
        skip_rids.update({ttl_cut, ttl_cut + 1, ttl_cut + 2})

    def is_plan_row(r):
        v = r[plan_col]
        return isinstance(v, str) and v.strip().lower() == "plan"

    eligible = [r for r in rows_above if (r["_rid"] not in skip_rids and is_plan_row(r))]

    # 負數防呆
    for r in eligible:
        rd = dict(r)
        for c in (win1 + win2 + win3):
            v = _to_number(rd.get(c, 0))
            if v < 0:
                d = c2d.get(c)
                mmdd = d.strftime("%m/%d") if d else str(c)
                raise RuntimeError(f"資料錯誤：{mmdd} 的值 < 0")

    by_pn: Dict[str, List[float]] = defaultdict(lambda: [0.0, 0.0, 0.0])
    for r in eligible:
        rd = dict(r)
        pn = _normalize_pn(rd.get(pn_col))
        if not pn:
            continue
        by_pn[pn][0] += sum(_to_number(rd.get(c, 0)) for c in win1)
        by_pn[pn][1] += sum(_to_number(rd.get(c, 0)) for c in win2)
        by_pn[pn][2] += sum(_to_number(rd.get(c, 0)) for c in win3)

    plan3w = {
        pn: (int(round(v[0])), int(round(v[1])), int(round(v[2])))
        for pn, v in by_pn.items() if any(v)
    }
    logging.info(f"成品三週 PN 數：{len(plan3w)}（已過濾三週全 0）")
    conn.close()
    return plan3w

# ---------- 成品三週量 → 半品三週需求 ----------
def plan_to_bead_requirements(panel_map: Dict[str, List[Dict[str, Any]]],
                              plan3w: Dict[str, Tuple[int,int,int]]) -> Dict[str, Tuple[int,int,int]]:
    need_by_bead: Dict[str, List[float]] = defaultdict(lambda: [0.0, 0.0, 0.0])
    missing: List[str] = []

    for pn, (w1, w2, w3) in plan3w.items():
        beads = panel_map.get(pn)
        if not beads:
            missing.append(pn)
            continue
        for item in beads:
            bpn = str(item["bead_pn"]).strip()
            per = _to_number(item.get("quantity", 0))
            need_by_bead[bpn][0] += per * w1
            need_by_bead[bpn][1] += per * w2
            need_by_bead[bpn][2] += per * w3

    if missing:
        logging.warning("無 Panel 明細（略過）：" + ", ".join(missing[:30]) + (" ..." if len(missing) > 30 else ""))

    bead3w = {
        bpn: (int(round(v[0])), int(round(v[1])), int(round(v[2])))
        for bpn, v in need_by_bead.items() if any(v)
    }
    logging.info(f"半品需求筆數：{len(bead3w)}（已過濾三週全 0）")
    return bead3w

def get_bead_inventory_totals(db_path: str) -> Dict[str, int]:
    """
    回傳 { PN: 累計可用庫存(整數) }
    規則：
      1) 同 PN 以「可使用庫存」累加（逐列）
      2) 「可使用庫存」< 300 的列不計算（但 PN 在 allow_small_pns 例外名單就仍計算）
      3) 若「工單數 > 1000」且 入庫/累計領用/可使用庫存 三欄皆為 0 或空，視為已生產未入庫 → 用「工單數」當作可使用庫存累計
    """
    allow_small_pns = {"5714400220", "5714400221"}  # 例外一定計入

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("PRAGMA table_info(beads_inventory)")
    col_rows = cur.fetchall()
    if not col_rows:
        conn.close()
        raise RuntimeError("資料表 beads_inventory 不存在或無法讀取")
    cols = [r["name"] for r in col_rows]

    need = ["PN", "工單數", "入庫", "累計領用", "可使用庫存"]
    for k in need:
        if k not in cols:
            conn.close()
            raise RuntimeError(f"beads_inventory 少欄位：{k}；現有：{cols}")

    cur.execute('SELECT "PN","工單數","入庫","累計領用","可使用庫存" FROM beads_inventory')
    rows = cur.fetchall()
    conn.close()

    totals = defaultdict(float)

    def zeroish(x):
        return _to_number(x) == 0

    for r in rows:
        pn = _normalize_pn(r["PN"])
        if not pn:
            continue

        wo    = _to_number(r["工單數"])
        inb   = _to_number(r["入庫"])
        used  = _to_number(r["累計領用"])
        avail = _to_number(r["可使用庫存"])

        # 規則 3
        candidate = avail
        if wo > 1000 and zeroish(inb) and zeroish(used) and zeroish(avail):
            candidate = wo

        # 規則 2
        if candidate < 300 and pn not in allow_small_pns:
            continue

        totals[pn] += candidate

    return {pn: int(round(v)) for pn, v in totals.items()}

import os, time, tempfile, shutil
from pathlib import Path

def _safe_save_xlsx(wb, target_path: str, max_retry: int = 10, interval_sec: float = 2.0) -> str:
    """
    安全儲存到 target_path：
      1) 儲存到相同目錄的暫存檔
      2) os.replace 原子性覆蓋
      3) 若因為鎖檔/權限失敗 → 重試
      4) 全部失敗 → 存成備援檔名 (加時間戳) 並回傳該備援路徑
    回傳：實際寫出的檔案路徑（可能是 target_path 或備援檔）
    """
    target = Path(target_path)
    target_dir = target.parent
    target_dir.mkdir(parents=True, exist_ok=True)

    # 快速檢查目錄寫入權限
    if not os.access(str(target_dir), os.W_OK):
        # 目錄沒寫入權限 → 直接寫到使用者暫存區，最後嘗試複製回去
        tmp_fd, tmp_local = tempfile.mkstemp(suffix=".xlsx")
        os.close(tmp_fd)
        try:
            wb.save(tmp_local)
            # 嘗試複製回去（即使無法覆蓋，也至少留下本地暫存檔）
            shutil.copyfile(tmp_local, str(target))
            return str(target)
        except Exception:
            # 直接使用暫存路徑作為備援輸出
            return tmp_local
        finally:
            try: os.remove(tmp_local)
            except Exception: pass

    # 有寫入權限 → 暫存+原子置換
    for attempt in range(1, max_retry + 1):
        tmp_name = f".~save_{int(time.time())}_{os.getpid()}.xlsx"
        tmp_path = str(target_dir / tmp_name)
        try:
            wb.save(tmp_path)
            # 若能成功存暫存檔，試著原子覆蓋
            os.replace(tmp_path, str(target))
            return str(target)
        except PermissionError as e:
            # 可能被鎖，等一下重試
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass
            if attempt < max_retry:
                time.sleep(interval_sec)
                continue
            # 用備援檔名收尾
            ts = time.strftime("%Y%m%d-%H%M%S")
            fallback = str(target_dir / f"{target.stem}-OUTPUT-{ts}.xlsx")
            try:
                wb.save(fallback)
                return fallback
            except Exception:
                raise e  # 連備援都存不了，就把原錯拋出
        except Exception:
            # 其他錯誤也做重試
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass
            if attempt < max_retry:
                time.sleep(interval_sec)
                continue
            # 最後嘗試備援
            ts = time.strftime("%Y%m%d-%H%M%S")
            fallback = str(target_dir / f"{target.stem}-OUTPUT-{ts}.xlsx")
            wb.save(fallback)
            return fallback

def write_back_to_sheet(xlsx_path: str,
                        sheet_name: str,
                        bead3w: dict[str, tuple[int,int,int]],
                        inv_by_pn: dict[str, int]):
    """
    目標工作表：B 欄=半品料號
      - 先把 E3:Em、F3:Fm、G3:Gm、H3:Hm 清為 0
      - 依 B 欄料號比對回填：E=庫存，F/G/H=三週需求
    """
    import openpyxl, re
    from collections import defaultdict

    wb = openpyxl.load_workbook(xlsx_path)
    try:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"找不到工作表：{sheet_name}")
        ws = wb[sheet_name]

        max_row = ws.max_row
        pn_to_rows: dict[str, list[int]] = defaultdict(list)
        for r in range(1, max_row + 1):
            pn = _normalize_pn(ws.cell(r, 2).value)  # B 欄
            if pn and re.fullmatch(r"\d+", pn):
                pn_to_rows[pn].append(r)

        for r in range(3, max_row + 1):
            for c in ("E", "F", "G", "H"):
                cell = ws[f"{c}{r}"]
                cell.value = 0
                cell.number_format = "#,##0"

        for pn, inv in inv_by_pn.items():
            rows = pn_to_rows.get(_normalize_pn(pn), [])
            for r in rows:
                if r >= 3:
                    cell = ws[f"E{r}"]
                    cell.value = int(inv)
                    cell.number_format = "#,##0"

        for pn, (w1, w2, w3) in bead3w.items():
            rows = pn_to_rows.get(_normalize_pn(pn), [])
            for r in rows:
                if r >= 3:
                    ws[f"F{r}"].value = int(w1); ws[f"F{r}"].number_format = "#,##0"
                    ws[f"G{r}"].value = int(w2); ws[f"G{r}"].number_format = "#,##0"
                    ws[f"H{r}"].value = int(w3); ws[f"H{r}"].number_format = "#,##0"

        # ★ 使用安全儲存，避開共享鎖/權限問題
        out_path = _safe_save_xlsx(wb, xlsx_path, max_retry=15, interval_sec=2.0)
        logging.info(f"已回填：{sheet_name} → {out_path}")
        return out_path
    finally:
        # 確保關閉 Workbook，釋放檔案 Handle
        wb.close()


# ---------- 主流程 ----------
if __name__ == "__main__":
    # === 路徑/參數（依需求調整）===
    PANEL_XLSX   = r"\\fls341\MBBU_FAB\MB_PD\生管自動化\滴定\beads 需求模組.xlsx"
    PANEL_SHEET  = "Panel 明細"
    DB_PATH      = r"\\fls341\MBBU_FAB\MB_PD\生管自動化\滴定\資料庫\beads_sync.db"
    TABLE        = "production_plan"
    START_MMDD   = "08/14"
    YEAR         = 2025  # 若想自動推斷，可改為 None
    TARGET_SHEET = "滴定排程需求表"
    # ==========================

    # 1) Panel 明細 → 成品→半品對照
    panel_map = process_panel_details(PANEL_XLSX, PANEL_SHEET, start_col_letter="D")

    # 2) 成品 3 週量（Plan=Plan、TTL 規則、負數檢查）
    plan3w = get_weekly_production(DB_PATH, TABLE, START_MMDD, YEAR)

    # 3) 半品 3 週需求
    bead3w = plan_to_bead_requirements(panel_map, plan3w)

    # 4) 從 beads_inventory 匯總 PN 庫存
    inv_by_pn = get_bead_inventory_totals(DB_PATH)

    # 5) 回填：E=庫存、F/G/H=三週需求（E/F/G/H 清空從第 3 列開始）
    write_back_to_sheet(PANEL_XLSX, TARGET_SHEET, bead3w, inv_by_pn)
