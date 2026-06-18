import pandas as pd
import collections
import argparse
import sys
import os
import math
import openpyxl
import time
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from ortools.sat.python import cp_model

# ==============================================================================
# 1. 基礎設定 (Configuration)
# ==============================================================================

START_HOUR = 8
MINS_PER_GRID = 30
DEFAULT_GRIDS = 35  # 08:00 ~ 01:30
LUNCH_BREAK_GRIDS = [8, 9] # 12:00, 12:30
DEFAULT_RULEBOOK = "Rulebook.xlsx"
DEFAULT_DEMAND = "Demand.xlsx"
OUTPUT_FILENAME = "滴定排程結果_V32.xlsx"

def read_data_file(file_path):
    if not os.path.exists(file_path):
        print(f"❌ 錯誤: 找不到檔案: {file_path}")
        return None
    try:
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path, dtype=str)
        else:
            df = pd.read_excel(file_path, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        df.dropna(how='all', inplace=True)
        return df
    except Exception as e:
        print(f"❌ 讀取失敗: {e}")
        return None

def parse_allowed_grids(time_str, grids_per_day):
    full_day = [g for g in range(grids_per_day) if g not in LUNCH_BREAK_GRIDS]
    if pd.isna(time_str) or str(time_str).strip() == '' or str(time_str).lower() == 'nan':
        return full_day
    
    allowed = set()
    parts = str(time_str).replace('"', '').replace("'", "").split(',')
    has_specific = False
    
    for p in parts:
        try:
            if ':' not in p: continue
            h, m = map(int, p.strip().split(':'))
            start_min = (h * 60 + m) - (START_HOUR * 60)
            if start_min < 0: start_min = 0
            
            start_g = math.ceil(start_min / MINS_PER_GRID)
            for g in range(start_g, grids_per_day):
                if g not in LUNCH_BREAK_GRIDS:
                    allowed.add(g)
            has_specific = True
        except: pass
        
    return sorted(list(allowed)) if has_specific else full_day

def parse_qty_options(qty_str):
    if pd.isna(qty_str): return [0.0]
    try:
        parts = str(qty_str).lower().replace(',', '').split('or')
        opts = [float(p.strip()) for p in parts if float(p.strip()) > 0]
        return sorted(opts) if opts else [0.0]
    except: return [0.0]

# ==============================================================================
# 2. 載入資料
# ==============================================================================

def load_rulebook_v3(file_path, grids_per_day=DEFAULT_GRIDS, expand_window=False):
    print(f"📂 正在讀取規則書 (DayLimit={grids_per_day}格, 放寬={expand_window})...")
    df = read_data_file(file_path)
    if df is None: return None, None, None

    jobs = {} 
    reagent_groups = {} 
    all_staff_set = set() 
    
    for idx, row in df.iterrows():
        raw_pn = row.get('PN')
        if pd.isna(raw_pn): continue
        pn = str(raw_pn).strip().replace('.0', '')
        if not pn: continue

        reagent = str(row.get('Reagent', '')).strip()
        if reagent.lower() == 'nan': reagent = ''
        
        port_val = str(row.get('Port數', '2')).strip().upper()
        if port_val == 'IVEK': num_ports = 0 
        elif port_val.isdigit(): num_ports = int(port_val)
        else: num_ports = 2
        
        qty_options = parse_qty_options(row.get('數量', '0'))
        
        try:
            f_raw = row.get('凍乾時間')
            f_hrs = 12.0 if pd.isna(f_raw) or str(f_raw).strip()=='' else float(f_raw)
            if math.isnan(f_hrs): f_hrs = 12.0
        except: f_hrs = 12.0
        freeze_duration = int(f_hrs * 60)
        
        d_str = str(row.get('可用凍乾機', '')).strip().replace('"', '').replace("'", "")
        compatible_dryers = [d.strip() for d in d_str.split(',') if d.strip()]
        
        compatible_staff = []
        for col in ['配藥人-1', '配藥人-2', '配藥人-3']:
            s = str(row.get(col, '')).strip()
            if s and s.lower() != 'nan':
                compatible_staff.append(s)
                all_staff_set.add(s)
        
        time_str = row.get('交藥時間', '')
        if expand_window and '11:00' in str(time_str):
             time_str = str(time_str).replace('11:00', '10:00')
        
        allowed_grids = parse_allowed_grids(time_str, grids_per_day)
        
        split_flag = str(row.get('U,D 劑分開生產排程', '')).strip().upper()
        split_prod = (split_flag in ['V', 'TRUE', '1'])
        
        note = str(row.get('備註', '')).strip()
        no_friday = '不能放週五滴定' in note
        
        job = {
            'pn': pn, 'name': str(row.get('Name', '')), 'reagent': reagent,
            'num_ports': num_ports, 'qty_options': qty_options,
            'duration_freeze': freeze_duration, 'compatible_dryers': compatible_dryers,
            'compatible_staff': compatible_staff, 'allowed_grids': allowed_grids,
            'split_prod': split_prod, 'no_friday': no_friday
        }
        jobs[pn] = job
        if reagent:
            if reagent not in reagent_groups: reagent_groups[reagent] = []
            reagent_groups[reagent].append(pn)
            
    return jobs, reagent_groups, sorted(list(all_staff_set))

# ==============================================================================
# 3. 需求計算
# ==============================================================================

def calculate_priority_tasks(demand_file_path, rulebook_dict, reagent_groups):
    print(f"📉 正在讀取需求表...")
    df = read_data_file(demand_file_path)
    if df is None: return []

    if '料號' not in df.columns:
        try:
            if demand_file_path.endswith('.csv'):
                df = pd.read_csv(demand_file_path, header=1, dtype=str)
            else:
                df = pd.read_excel(demand_file_path, header=1, dtype=str)
            df.columns = [str(c).strip() for c in df.columns]
        except:
            return []

    pn_status = {}
    for idx, row in df.iterrows():
        raw_pn = row.get('料號')
        if pd.isna(raw_pn): continue
        pn = str(raw_pn).strip().replace('.0', '')
        if pn not in rulebook_dict: continue
        
        def pnum(v):
            try: return float(str(v).replace(',', ''))
            except: return 0.0
            
        stock = pnum(row.get('庫存+滴定', 0))
        w1 = pnum(row.get('第一周需求', 0))
        w2 = pnum(row.get('第二周週需求', 0))
        w3 = pnum(row.get('第三周週需求', 0))
        
        prio, shortage, deadline = 0, 0, None
        tmp = stock
        
        if tmp - w1 < 0:
            shortage = w1 - tmp; prio = 1; deadline = 3; tmp += shortage
        else: tmp -= w1
        
        if tmp - w2 < 0:
            shortage = w2 - tmp; prio = 2 if prio==0 else prio; tmp += shortage
        else: tmp -= w2
        
        if tmp - w3 < 0:
            shortage = w3 - tmp; prio = 3 if prio==0 else prio
            
        pn_status[pn] = {'shortage': shortage, 'priority': prio, 'deadline': deadline}

    final_tasks = []
    processed = set()
    
    for reagent, pns in reagent_groups.items():
        actives = [p for p in pns if p in pn_status]
        if not actives: continue
        
        if any(pn_status[p]['priority'] > 0 for p in actives):
            max_p = max(pn_status[p]['priority'] for p in actives)
            for p in actives:
                st = pn_status[p]
                task_p = max_p
                task_s = st['shortage']
                if st['priority'] == 0:
                    task_s = 0 
                
                final_tasks.append({
                    'pn': p, 'type': f"P{task_p}", 'shortage': task_s,
                    'deadline_days': st['deadline'], 'priority': task_p
                })
                processed.add(p)
    
    for pn, st in pn_status.items():
        if pn not in processed and st['priority'] > 0:
            final_tasks.append({
                'pn': pn, 'type': f"P{st['priority']}", 'shortage': st['shortage'],
                'deadline_days': st['deadline'], 'priority': st['priority']
            })
            
    return final_tasks

def prepare_jobs(tasks, rulebook, groups):
    jobs = []
    for t in tasks:
        pn = t['pn']
        rule = rulebook[pn]
        
        target = t['shortage'] if t['shortage'] > 0 else 0
        qty = rule['qty_options'][-1]
        for opt in rule['qty_options']:
            if opt >= target:
                qty = opt
                break
                
        ports = rule['num_ports']
        dur_min = int(math.ceil((qty / ports / 1500)*60 + 30)) if ports > 0 else 60
        grid_dur = math.ceil(dur_min / 30)
        fr_dur = rule['duration_freeze']
        
        job = {
            'id': f"{t['type']}_{pn}", 'pn': pn, 'name': rule['name'],
            'qty': qty, 'num_ports': ports, 'duration_dosing': dur_min,
            'grid_duration': grid_dur, 'duration_freeze': fr_dur,
            'compatible_dryers': rule['compatible_dryers'],
            'compatible_staff': rule['compatible_staff'],
            'allowed_grids': rule['allowed_grids'],
            'no_friday': rule['no_friday'], 'split_prod': rule['split_prod'],
            'reagent': rule['reagent'], 'deadline_days': t['deadline_days']
        }
        jobs.append(job)
        
    job_map = {j['pn']: j for j in jobs}
    for r, pns in groups.items():
        sub = [p for p in pns if p in job_map]
        if len(sub) > 1 and not any(job_map[p]['split_prod'] for p in sub):
            max_f = max(job_map[p]['duration_freeze'] for p in sub)
            for p in sub: job_map[p]['duration_freeze'] = max_f
            
    return jobs

# ==============================================================================
# 4. 求解器 (V32: 繼承 V31 邏輯)
# ==============================================================================

def run_solver(jobs, reagent_groups, all_staff, horizon_days, relax_ports, grids_per_day):
    mode_str = "寬鬆 (偶數Port)" if relax_ports else "嚴格 (單數Port)"
    print(f"🔄 排程運算: {horizon_days}天 | {mode_str} | 日工時:{grids_per_day}格")
    
    model = cp_model.CpModel()
    horizon_grids = horizon_days * grids_per_day
    
    all_dryers = sorted(list(set(d for j in jobs for d in j['compatible_dryers'])))
    dryer_intervals = collections.defaultdict(list)
    staff_intervals = collections.defaultdict(list)
    
    job_vars = {}
    pn_to_vars = {}
    rects = []
    
    for job in jobs:
        s_grid = model.NewIntVar(0, horizon_grids, f"start_{job['id']}")
        e_grid = model.NewIntVar(0, horizon_grids, f"end_{job['id']}")
        day = model.NewIntVar(0, horizon_days - 1, f"day_{job['id']}")
        
        d_idx = model.NewIntVar(0, len(all_dryers), f"dryer_{job['id']}")
        st_idx = model.NewIntVar(0, len(all_staff), f"staff_{job['id']}")
        p_start = model.NewIntVar(0, 11, f"port_{job['id']}")
        
        s_freeze = model.NewIntVar(0, horizon_grids + 100, f"freeze_start_{job['id']}")
        e_freeze = model.NewIntVar(0, horizon_grids + 200, f"freeze_end_{job['id']}")
        
        job_vars[job['id']] = {
            's': s_grid, 'e': e_grid, 'day': day, 
            'dryer': d_idx, 'staff': st_idx, 'port': p_start,
            's_freeze': s_freeze, 'e_freeze': e_freeze
        }
        pn_to_vars[job['pn']] = job_vars[job['id']]
        
        model.Add(e_grid == s_grid + job['grid_duration'])
        model.AddDivisionEquality(day, s_grid, grids_per_day)
        
        if job['num_ports'] > 0:
            x_int = model.NewFixedSizeIntervalVar(s_grid, job['grid_duration'], '')
            y_int = model.NewFixedSizeIntervalVar(p_start, job['num_ports'], '')
            rects.append([x_int, y_int])
            model.Add(p_start + job['num_ports'] <= 12)
            
        s_in_day = model.NewIntVar(0, grids_per_day, '')
        model.AddModuloEquality(s_in_day, s_grid, grids_per_day)
        model.Add(s_in_day + job['grid_duration'] <= grids_per_day)
        
        if job['allowed_grids']:
            valid = []
            for d in range(horizon_days):
                offset = d * grids_per_day
                for g in job['allowed_grids']:
                    valid.append(offset + g)
            if valid:
                model.AddLinearExpressionInDomain(s_grid, cp_model.Domain.FromValues(valid))
        
        # 凍乾邏輯
        model.Add(s_freeze >= e_grid)
        fr_grids = math.ceil((job['duration_freeze'] + 120) / 30)
        model.Add(e_freeze == s_freeze + fr_grids)
        
        used_dryers = []
        for idx, dname in enumerate(all_dryers):
            if dname in job['compatible_dryers']:
                is_used = model.NewBoolVar(f"u_{job['id']}_{dname}")
                used_dryers.append(is_used)
                model.Add(d_idx == idx).OnlyEnforceIf(is_used)
                
                occ = model.NewOptionalFixedSizeIntervalVar(s_freeze, fr_grids, is_used, '')
                dryer_intervals[dname].append({
                    'int': occ, 
                    'bool': is_used, 
                    'job': job['id']
                })
        model.Add(sum(used_dryers) == 1)
        
        # 人員
        used_staff = []
        for idx, sname in enumerate(all_staff):
            if sname in job['compatible_staff']:
                is_used = model.NewBoolVar(f"u_st_{job['id']}_{sname}")
                used_staff.append(is_used)
                model.Add(st_idx == idx).OnlyEnforceIf(is_used)
                gap = model.NewOptionalFixedSizeIntervalVar(s_grid, 5, is_used, '')
                staff_intervals[idx].append({'int': gap, 'job': job['id']})
        model.Add(sum(used_staff) == 1)

    if rects:
        model.AddNoOverlap2D([r[0] for r in rects], [r[1] for r in rects])

    # ==========================================================================
    # 配對邏輯 (Group Logic)
    # ==========================================================================
    slaves = set()
    partners = {}
    paired_secondary_jobs = set()
    
    for r, pns in reagent_groups.items():
        sub = [p for p in pns if p in pn_to_vars]
        if len(sub) < 2: continue
        
        is_glipa = any('220' in p for p in sub) and any('221' in p for p in sub)
        if is_glipa:
            pA = next((p for p in sub if '220' in p), None)
            pB = next((p for p in sub if '221' in p), None)
            if pA and pB:
                model.Add(pn_to_vars[pA]['s'] == pn_to_vars[pB]['s'])
                model.Add(pn_to_vars[pA]['dryer'] != pn_to_vars[pB]['dryer'])
                
                if not relax_ports:
                    model.AddModuloEquality(0, pn_to_vars[pA]['port'], 2)
                    model.Add(pn_to_vars[pB]['port'] == pn_to_vars[pA]['port'] + 2)
                
                idA = next(j['id'] for j in jobs if j['pn']==pA)
                idB = next(j['id'] for j in jobs if j['pn']==pB)
                partners[idA] = idB; partners[idB] = idA
                slaves.add(idB)
            continue
            
        base = sub[0]
        for other in sub[1:]:
            model.Add(pn_to_vars[other]['staff'] == pn_to_vars[base]['staff'])
            
        pA, pB = sub[0], sub[1]
        jobA = next(j for j in jobs if j['pn']==pA)
        jobB = next(j for j in jobs if j['pn']==pB)
        
        if not (jobA['split_prod'] or jobB['split_prod']):
            model.Add(pn_to_vars[pA]['s'] == pn_to_vars[pB]['s'])
            model.Add(pn_to_vars[pA]['dryer'] == pn_to_vars[pB]['dryer'])
            
            if not relax_ports:
                model.AddModuloEquality(0, pn_to_vars[pA]['port'], 2)
                model.Add(pn_to_vars[pB]['port'] == pn_to_vars[pA]['port'] + jobA['num_ports'])
                if jobA['num_ports'] % 2 != 0:
                    paired_secondary_jobs.add(jobB['id'])
            
            slaves.add(jobB['id'])
            partners[jobA['id']] = jobB['id']; partners[jobB['id']] = jobA['id']
        else:
            model.Add(pn_to_vars[pA]['day'] != pn_to_vars[pB]['day'])

    global_setup_intervals = []
    for job in jobs:
        if job['id'] in slaves: continue
        if job['reagent']:
            s_grid = job_vars[job['id']]['s']
            setup_int = model.NewFixedSizeIntervalVar(s_grid, 1, f"global_setup_{job['id']}")
            global_setup_intervals.append(setup_int)

    if global_setup_intervals:
        model.AddNoOverlap(global_setup_intervals)

    MAX_WAIT_GRIDS = 2 
    for jid, vars in job_vars.items():
        model.Add(vars['s_freeze'] <= vars['e'] + MAX_WAIT_GRIDS)

    if not relax_ports:
        for job in jobs:
            if job['num_ports'] > 0 and job['id'] not in paired_secondary_jobs:
                model.AddModuloEquality(0, pn_to_vars[job['pn']]['port'], 2)

    for dname, items in dryer_intervals.items():
        for i in range(len(items)):
            for j in range(i+1, len(items)):
                job_i = items[i]['job']
                job_j = items[j]['job']
                if partners.get(job_i) == job_j: continue
                model.AddNoOverlap([items[i]['int'], items[j]['int']])

    for idx, items in staff_intervals.items():
        for i in range(len(items)):
            for j in range(i+1, len(items)):
                if partners.get(items[i]['job']) == items[j]['job']: continue
                model.AddNoOverlap([items[i]['int'], items[j]['int']])

    makespan = model.NewIntVar(0, horizon_grids, 'makespan')
    model.AddMaxEquality(makespan, [v['e'] for v in job_vars.values()])
    model.Minimize(makespan)
    
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 30.0
    status = solver.Solve(model)
    
    if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
        results = []
        for jid, vars in job_vars.items():
            j = next(x for x in jobs if x['id'] == jid)
            j['grid_start'] = solver.Value(vars['s']) % grids_per_day
            j['day'] = solver.Value(vars['day'])
            j['start'] = solver.Value(vars['s'])
            j['assigned_port_idx'] = solver.Value(vars['port']) if j['num_ports'] > 0 else 0
            
            d_idx = solver.Value(vars['dryer'])
            st_idx = solver.Value(vars['staff'])
            j['assigned_dryer'] = all_dryers[d_idx]
            j['assigned_staff'] = all_staff[st_idx]
            
            s_freeze_val = solver.Value(vars['s_freeze'])
            e_freeze_val = solver.Value(vars['e_freeze'])
            j['freeze_start_abs'] = s_freeze_val
            j['freeze_end_abs'] = e_freeze_val
            
            results.append(j)
        return results
    return None

# ==============================================================================
# 5. 輸出
# ==============================================================================

def print_summary(results, grids_per_day):
    if results:
        actual_days = max(j['day'] for j in results) + 1
    else:
        actual_days = 0

    print(f"\n✅ 排程成功！工期: {actual_days} 天")
    print("-" * 155)
    print(f"{'Job ID':<16}| {'PN':<12}| {'Name':<12}| {'Day':<3}| {'Start':<8}| {'End':<8}| {'FrzEnd':<12}| {'Dryer':<6}| {'Qty':<8}| {'Staff':<8}| {'Port'}")
    print("-" * 155)
    
    results.sort(key=lambda x: (x['day'], x['start'], x.get('assigned_port_idx', 0)))
    
    for j in results:
        day = j['day'] + 1
        s_m = START_HOUR * 60 + j['grid_start'] * 30
        e_m = s_m + j['duration_dosing']
        
        total_mins = (START_HOUR * 60) + (j['freeze_end_abs'] * 30)
        fd_day = total_mins // (24 * 60) 
        fd_time = total_mins % (24 * 60)
        
        frz_str = f"{fd_time//60:02d}:{fd_time%60:02d}"
        if fd_day > (j['day']): 
             frz_str += f"(+{fd_day - j['day']})" 
        elif fd_day > 0: 
             frz_str += f"(d{fd_day+1})"

        def fmt(m):
            d = m // (24*60)
            m %= (24*60)
            t = f"{m//60:02d}:{m%60:02d}"
            return f"{t}(+{d})" if d > 0 else t
            
        p_str = "IVEK" if j['num_ports'] == 0 else str(j['assigned_port_idx'] + 1)
        if j['num_ports'] > 1: p_str += f"~{j['assigned_port_idx'] + j['num_ports']}"
        
        print(f"{j['id']:<16}| {j['pn']:<12}| {j['name']:<12}| {day:<3}| {fmt(s_m):<8}| {fmt(e_m):<8}| {frz_str:<12}| {j['assigned_dryer']:<6}| {j['qty']:<8.0f}| {j['assigned_staff']:<8}| {p_str}")
    print("-" * 155)

def export_excel_v32(results, grids_per_day, filename=OUTPUT_FILENAME):
    print(f"📊 正在準備 Excel 輸出: {filename} ...")
    
    save_name = filename
    try:
        f = open(filename, 'a+')
        f.close()
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        save_name = filename.replace(".xlsx", f"_{ts}.xlsx")
        print(f"⚠️ 原檔名被鎖定，自動另存為: {save_name}")
    except:
        pass

    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames: wb.remove(wb['Sheet'])
    by_day = collections.defaultdict(list)
    for j in results: by_day[j['day']].append(j)
    
    fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    colors = ["E6B8AF", "F4CCCC", "FCE5CD", "FFF2CC", "D9EAD3", "D0E0E3", "C9DAF8", "CFE2F3", "D9D2E9", "EAD1DC"]
    pn_colors = {}
    
    for d in sorted(by_day.keys()):
        ws = wb.create_sheet(f"Day {d+1}")
        headers = ["時間", "IVEK 1", "IVEK 2"] + [f"Port {i}" for i in range(1, 13)]
        for c, txt in enumerate(headers, 1):
            cell = ws.cell(1, c, txt)
            cell.fill = fill; cell.border = border; cell.alignment = Alignment('center', 'center')
            ws.column_dimensions[get_column_letter(c)].width = 14 if c > 1 else 10
            
        for i in range(grids_per_day):
            m = START_HOUR * 60 + i * 30
            txt = f"{m//60:02d}:{m%60:02d}"
            if m >= 1440: txt = f"隔{(m-1440)//60:02d}:{(m-1440)%60:02d}"
            cell = ws.cell(i+2, 1, txt)
            cell.border = border; cell.alignment = Alignment('center', 'center')
            
        for j in by_day[d]:
            if j['pn'] not in pn_colors: pn_colors[j['pn']] = colors[len(pn_colors)%len(colors)]
            c_fill = PatternFill(start_color=pn_colors[j['pn']], end_color=pn_colors[j['pn']], fill_type="solid")
            
            r_start = j['grid_start'] + 2
            r_end = r_start + j['grid_duration'] - 1
            
            c_start = 2 
            if j['num_ports'] > 0:
                c_start = 4 + j['assigned_port_idx']
            c_end = c_start + max(0, j['num_ports'] - 1)
            
            f_total_m = (START_HOUR * 60) + (j['freeze_end_abs'] * 30)
            f_hm = f"{f_total_m%1440//60:02d}:{f_total_m%60:02d}"
            f_d_offset = f_total_m // 1440 - j['day']
            f_note = f"(+{f_d_offset}d)" if f_d_offset > 0 else ""
            
            dryer_txt = f"Frz-{j['assigned_dryer']}\nEnd {f_hm}{f_note}"
            txt = f"{j['pn']}\n{j['name']}\n{int(j['qty']/max(1, j['num_ports']))}/port\n{dryer_txt}"
            
            try:
                is_merged = False
                for merged_range in ws.merged_cells.ranges:
                    if r_start >= merged_range.min_row and r_start <= merged_range.max_row and \
                       c_start >= merged_range.min_col and c_start <= merged_range.max_col:
                        is_merged = True
                        break
                if is_merged: continue
            except: pass
            
            cell = ws.cell(r_start, c_start, txt)
            cell.fill = c_fill; cell.border = border; cell.alignment = Alignment('center', 'center', wrap_text=True)
            
            if f_d_offset >= 1 and (f_total_m % 1440) > (1*60 + 30): 
                cell.font = Font(color="FF0000", bold=True)
            
            if r_end > r_start or c_end > c_start:
                try: ws.merge_cells(start_row=r_start, start_column=c_start, end_row=r_end, end_column=c_end)
                except: pass
                
            for r in range(r_start, r_end+1):
                for c in range(c_start, c_end+1): ws.cell(r, c).border = border
                
    try:
        wb.save(save_name)
        print(f"✅ Excel 已輸出至: {save_name}")
    except Exception as e:
        print(f"❌ 存檔失敗: {e}")

# ==============================================================================
# 6. 主程式 (V32)
# ==============================================================================

def main():
    print(f"🚀 啟動 V32 排程系統 (工期放寬策略)...")
    
    parser = argparse.ArgumentParser()
    parser.add_argument('--rulebook', help="規則書路徑", default=DEFAULT_RULEBOOK)
    parser.add_argument('--demand', help="需求表路徑", default=DEFAULT_DEMAND)
    args = parser.parse_args()
    
    rule_path = args.rulebook
    demand_path = args.demand

    if not os.path.exists(rule_path) or not os.path.exists(demand_path):
        print("❌ 找不到輸入檔案。")
        return

    try:
        rules, groups, staff = load_rulebook_v3(rule_path, DEFAULT_GRIDS)
        tasks = calculate_priority_tasks(demand_path, rules, groups)
        if not tasks: 
            print("✅ 無任務。")
            return
        jobs = prepare_jobs(tasks, rules, groups)
        
        # [V32] 修改策略：鎖定嚴格模式 (False)，逐漸增加天數 (3 -> 4 -> 5 -> 6 -> 7)
        strategies = [
            (3, False), (4, False), (5, False), (6, False), (7, False)
        ]
        
        final_res = None
        final_grids = DEFAULT_GRIDS

        for days, relax in strategies:
            res = run_solver(jobs, groups, staff, days, relax, DEFAULT_GRIDS)
            
            if res:
                # V32 保留 V29/V30 的優化邏輯 (若最後一天只有1個任務則嘗試壓縮)
                max_d = max(j['day'] for j in res)
                tasks_on_last_day = sum(1 for j in res if j['day'] == max_d)
                
                if tasks_on_last_day == 1 and max_d > 0:
                    print(f"\n⚠️ 偵測到 Day {max_d+1} 只有 1 個任務，嘗試優化 (延長加班 + 放寬限制)...")
                    
                    EXTENDED_GRIDS = 36 # 08:00 ~ 02:00
                    rules_opt, groups_opt, staff_opt = load_rulebook_v3(rule_path, EXTENDED_GRIDS, expand_window=True)
                    jobs_opt = prepare_jobs(tasks, rules_opt, groups_opt)
                    
                    res_opt = run_solver(jobs_opt, groups_opt, staff_opt, days, relax, EXTENDED_GRIDS)
                    
                    if res_opt:
                        new_max_d = max(j['day'] for j in res_opt)
                        if new_max_d < max_d:
                            print(f"🎉 優化成功！工期縮短為 {new_max_d+1} 天。")
                            final_res = res_opt
                            final_grids = EXTENDED_GRIDS
                            break 
                        else:
                            final_res = res
                    else:
                        final_res = res
                else:
                    final_res = res
                break 
        
        if final_res:
            print_summary(final_res, final_grids)
            export_excel_v32(final_res, final_grids)
        else:
            print("❌ 所有策略 (3~7天) 皆排程失敗。")

    except Exception as e:
        print(f"\n❌ 錯誤: {e}")
        import traceback
        traceback.print_exc()
    
    print("\n執行結束。")

if __name__ == '__main__':
    main()