import pandas as pd
import sqlite3
import os
import re
import math
import sys      # ✅ 新增
import json     # ✅ 新增
import argparse   # ✅ 新增
from datetime import datetime, timedelta, time
import traceback
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ====================================================================
# V9.9.2 修正版排程系統 (新增休假日功能)
# ====================================================================

# --------------------------------------------------------------------
# 1. 設定
# --------------------------------------------------------------------
MAIN_DB_PATH = r"\\fls341\MBBU_FAB\MB_PD\生管自動化\滴定\資料庫\beads_sync.db"
SCHEDULE_DB_PATH = r"\\fls341\MBBU_FAB\MB_PD\生管自動化\滴定\資料庫\Beads_Schedule.db"
output_dir = r"\\fls341\MBBU_FAB\MB_PD\生管自動化\滴定\beadsSchedule"
CURRENT_VERSION = "v9.9.12 (凌晨不收藥)"

SCHEDULING_DAYS = 5
ALLOW_EXTEND_TO_6_DAYS = True 
MAX_PORTS = 12 
DOSING_PREP_TIME_MIN = 30 
HARVEST_TIME_MIN = 30 
# ========================================
# V9.9.12: 禁止時段設定
# ========================================
FORBIDDEN_END_TIME_START = time(3, 0)   # 禁止結束時間起始：03:00
FORBIDDEN_END_TIME_END = time(8, 0)     # 禁止結束時間結束：08:00

# ===== MODIFIED RULES (2025-11-04) =====
DOSING_RATE_PER_HR = 1500
MAX_PORTS_PER_PN = 4
PORT_OPTIONS = [4, 2]
DOSING_TIME_MIN = 3.0
DOSING_TIME_MAX = 8.0
# ========================================

PREP_DURATION_MIN = 30 
WORK_START_TIME = time(8, 0)
WORK_END_TIME_LATE = time(3, 0)
WORK_END_TIME_ALERT = time(3, 0)
IVEK_COMPONENTS = ["5714400202", "5714400203", "5714400209", "5714400210"]
# ====================================================================
# 特殊案例 PN 定義
# ====================================================================
SPECIAL_CASE_PNS = {
    'tCREA': {
        'batch1': ['5714400180', '5714400181'],  # 同時交藥、同滴定、同凍乾機
        'batch2': ['5714400182']                  # 同天PM班、不同凍乾機
    },
    'Na_A': ['5714400202', '5714400203'],  # 連續兩天（擇一）
    'Na_B': ['5714400209', '5714400210'],  # 連續兩天（擇一）
    
    # ✅ V9.9.24 修改: GLIPA 改為分批次定義 (拆開排程以讀取正確 Port 數與機台)
    'GLIPA': {
        'batch1': ['5714400220'], # GLIPA-AD (2 Port)
        'batch2': ['5714400221']  # GLIPA-AU (1 Port)
    },
    # ✅ 新增：K系列邏輯 (116 為被動，226/117 為主動)
    'K_Series': {
        'K1': ['5714400226', '5714400116'], # 226 有需求時，帶上 116
        'K3': ['5714400117', '5714400116']  # 117 有需求時，帶上 116
    }
}

# 建立快速查詢集合
ALL_SPECIAL_PNS = set()
for group_name, group_data in SPECIAL_CASE_PNS.items():
    if isinstance(group_data, dict):  # tCREA, GLIPA, K_Series
        for batch in group_data.values():
            ALL_SPECIAL_PNS.update(batch)
    else:  # Na_A, Na_B
        ALL_SPECIAL_PNS.update(group_data)

print(f"特殊案例 PN: {ALL_SPECIAL_PNS}", file=sys.stderr)
# ========================================
# V9.9.2: 全域休假日設定
# ========================================
HOLIDAYS = set()  # 格式: datetime.date 物件
BATCH_START_NUMBER = 1  # 工單批次起始編號
VACATION_STAFF = {}  # {date_str: set(["人員1", "人員2", ...])}
# ✅ 新增：儲存排程起始日期
SCHEDULE_START_DATE = None

# --------------------------------------------------------------------
# 2. 輔助工具
# --------------------------------------------------------------------
def time_str_to_float(time_str: str) -> float:
    try:
        h, m = map(int, time_str.split(':'))
        return h + m / 60.0
    except Exception: 
        return 0.0

def float_to_time_str(hours: float) -> str:
    total_minutes = int(hours * 60)
    h = (total_minutes // 60) % 24 
    m = total_minutes % 60
    return f"{h:02d}:{m:02d}"

def add_hours_to_time(start_datetime: datetime, hours: float) -> datetime:
    return start_datetime + timedelta(minutes=int(hours * 60))

def parse_date_columns(df):
    """解析 Production Plan 的日期欄位"""
    date_cols = {}
    non_date_cols = [
        "PN", "料號", "品名", "Customer", "Description", "Plan", 
        "JAN", "FEB", "MAR", "APR", "MAY", "JUN", 
        "JUL", "AUG", "SEP", "OCT", "NOV", "DEC", "Unnamed: 0"
    ]
    non_date_cols.extend([f"WK{i:02d}" for i in range(1, 54)])
    
    date_regex = re.compile(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}')
        
    for col in df.columns:
        if col in non_date_cols:
            continue
        
        match = date_regex.match(col)
        if match:
            try:
                col_date = pd.to_datetime(match.group(0)) 
                date_cols[col] = col_date
                continue
            except (ValueError, TypeError):
                pass 

        try:
            col_clean = re.sub(r'\(.*\)', '', col).strip() 
            col_date = pd.to_datetime(col_clean)
            date_cols[col] = col_date
        except (ValueError, TypeError):
            pass
            
    return date_cols

def standardize_pn(pn_series: pd.Series) -> pd.Series:
    """標準化 PN"""
    return pn_series.astype(str).str.split('.').str[0]

# ========================================
# V9.9.2: 休假日輸入功能
# ========================================
def get_holidays_from_user():
    """
    讓使用者輸入休假日
    返回: set of datetime.date 物件
    """
    global HOLIDAYS
    
    print("\n" + "="*70, file=sys.stderr)
    print("🗓️  休假日設定", file=sys.stderr)
    print("="*70, file=sys.stderr)
    print("預計排程週是否有休假日？", file=sys.stderr)
    print("  [Y] 是，有休假日", file=sys.stderr)
    print("  [N] 否，無休假日（預設）", file=sys.stderr)
    
    has_holiday = input("\n請選擇 (Y/N，直接按 Enter 預設為 N): ").strip().upper()
    
    if has_holiday != 'Y':
        print("✅ 無休假日，繼續排程\n", file=sys.stderr)
        return set()
    
    print("\n請輸入休假日期:", file=sys.stderr)
    print("  格式: MM-DD 或 MM/DD", file=sys.stderr)
    print("  多日請用逗號分隔，例如: 11-11,11-12 或 11/11,11/12", file=sys.stderr)
    print("  年份將自動使用當前年份", file=sys.stderr)
    
    holiday_input = input("\n休假日期: ").strip()
    
    if not holiday_input:
        print("✅ 無休假日輸入，繼續排程\n", file=sys.stderr)
        return set()
    
    holidays = set()
    current_year = datetime.now().year
    
    # 解析輸入
    date_strings = [d.strip() for d in holiday_input.split(',') if d.strip()]
    
    for date_str in date_strings:
        try:
            # 替換 - 為 /
            date_str = date_str.replace('-', '/')
            
            # 嘗試解析 MM/DD
            parts = date_str.split('/')
            if len(parts) == 2:
                month = int(parts[0])
                day = int(parts[1])
                holiday_date = datetime(current_year, month, day).date()
                holidays.add(holiday_date)
                print(f"  ✅ 已加入休假日: {holiday_date.strftime('%Y-%m-%d')}", file=sys.stderr)
            else:
                print(f"  ⚠️ 格式錯誤，跳過: {date_str}", file=sys.stderr)
                
        except Exception as e:
            print(f"  ⚠️ 無法解析日期 '{date_str}': {e}", file=sys.stderr)
    
    if holidays:
        print(f"\n✅ 共設定 {len(holidays)} 個休假日:", file=sys.stderr)
        for h in sorted(holidays):
            print(f"      {h.strftime('%Y-%m-%d (%A)')}", file=sys.stderr)
    else:
        print("\n⚠️ 未成功設定任何休假日", file=sys.stderr)
    
    print("="*70 + "\n", file=sys.stderr)
    return holidays

def is_holiday(date_obj):
    """檢查日期是否為休假日"""
    if isinstance(date_obj, datetime):
        date_obj = date_obj.date()
    return date_obj in HOLIDAYS
# ========================================
# V9.9.9: 命令行參數解析
# ========================================
def parse_arguments():
    """解析命令行參數"""
    parser = argparse.ArgumentParser(description='Beads 排程系統（前端整合版）')
    
    # 必要參數
    parser.add_argument('--date', required=True, 
                        help='排程起始日期（週一），格式：MM/DD，例如：11/10')
    parser.add_argument('--need', required=True,
                        help='需求檔路徑（Excel 檔案）')
    
    # 可選參數
    parser.add_argument('--outdir', 
                        default=r"\\fls341\MBBU_FAB\MB_PD\生管自動化\滴定\beadsSchedule",
                        help='輸出資料夾路徑')
    parser.add_argument('--holidays', default='',
                        help='休假日，逗號分隔，格式：MM/DD，例如：11/10,11/11')
    parser.add_argument('--batch-numbers', default='',
                        help='生產工單批次編號起始值，例如：111')
    parser.add_argument('--vacation-staff', default='',
                        help='休假人員，格式：MM/DD-人名，例如：11/10-張三,11/11-李四')
    parser.add_argument('--dry-run', action='store_true',
                        help='Dry Run 模式（僅預覽，不輸出檔案）')
    
    return parser.parse_args()

def setup_holidays_from_args(holidays_str, start_date):
    """從參數設定休假日"""
    global HOLIDAYS
    
    if not holidays_str:
        return set()
    
    holidays = set()
    current_year = start_date.year
    
    date_strings = [d.strip() for d in holidays_str.split(',') if d.strip()]
    
    for date_str in date_strings:
        try:
            date_str = date_str.replace('-', '/')
            parts = date_str.split('/')
            
            if len(parts) == 2:
                month = int(parts[0])
                day = int(parts[1])
                holiday_date = datetime(current_year, month, day).date()
                holidays.add(holiday_date)
                print(f"  ✅ 已加入休假日: {holiday_date.strftime('%Y-%m-%d')}", file=sys.stderr)
            else:
                print(f"  ⚠️ 格式錯誤，跳過: {date_str}", file=sys.stderr)
                
        except Exception as e:
            print(f"  ⚠️ 無法解析日期 '{date_str}': {e}", file=sys.stderr)
    
    HOLIDAYS = holidays
    return holidays

def setup_holidays_from_args(holidays_str, start_date):
    """從參數設定休假日 (修正跨年問題)"""
    global HOLIDAYS
    
    if not holidays_str:
        return set()
    
    holidays = set()
    current_year = start_date.year
    start_month = start_date.month
    
    date_strings = [d.strip() for d in holidays_str.split(',') if d.strip()]
    
    print(f"  🔍 解析休假日 (基準年: {current_year}, 起始月: {start_month})...", file=sys.stderr)

    for date_str in date_strings:
        try:
            # 替換 - 為 /
            date_str = date_str.replace('-', '/')
            parts = date_str.split('/')
            
            if len(parts) == 2:
                month = int(parts[0])
                day = int(parts[1])
                
                # 🔥 跨年邏輯判定 🔥
                # 如果排程從 12月開始，但輸入的休假日是 1月，則該休假日應為明年
                if start_month == 12 and month == 1:
                    year = current_year + 1
                else:
                    year = current_year
                
                holiday_date = datetime(year, month, day).date()
                holidays.add(holiday_date)
                print(f"  ✅ 已加入休假日: {holiday_date.strftime('%Y-%m-%d')}", file=sys.stderr)
            else:
                print(f"  ⚠️ 格式錯誤，跳過: {date_str}", file=sys.stderr)
                
        except Exception as e:
            print(f"  ⚠️ 無法解析日期 '{date_str}': {e}", file=sys.stderr)
    
    HOLIDAYS = holidays
    return holidays

def setup_batch_start_from_args(batch_str):
    """從參數設定批次編號"""
    global BATCH_START_NUMBER
    
    if not batch_str:
        return 1
    
    try:
        batch_start = int(batch_str)
        if 1 <= batch_start <= 999:
            BATCH_START_NUMBER = batch_start
            print(f"  ✅ 工單批次起始編號: {batch_start}", file=sys.stderr)
            return batch_start
    except:
        pass
    
    print(f"  ⚠️ 批次編號無效，使用預設值 1", file=sys.stderr)
    return 1

def setup_vacation_staff_from_args(vacation_str, start_date):
    """從參數設定休假人員"""
    global VACATION_STAFF
    
    if not vacation_str:
        VACATION_STAFF = {}
        return {}
    
    vacation_staff = {}
    current_year = start_date.year
    
    entries = [e.strip() for e in vacation_str.split(',') if e.strip()]
    
    for entry in entries:
        try:
            if '-' not in entry:
                print(f"  ⚠️ 休假人員格式錯誤，跳過: {entry}", file=sys.stderr)
                print(f"      正確格式：MM/DD-人名，例如：11/10-張三", file=sys.stderr)
                continue
            
            date_part, person = entry.split('-', 1)
            date_part = date_part.strip().replace('/', '/')
            person = person.strip().lower() # ✅ 修正：儲存時轉為小寫
            
            parts = date_part.split('/')
            if len(parts) == 2:
                month = int(parts[0])
                day = int(parts[1])
                vacation_date = datetime(current_year, month, day).date()
                date_str = vacation_date.strftime('%Y-%m-%d')
                
                if date_str not in vacation_staff:
                    vacation_staff[date_str] = set()
                vacation_staff[date_str].add(person)
                
                print(f"  ✅ 已記錄休假: {date_str} - {person}", file=sys.stderr)
            else:
                print(f"  ⚠️ 日期格式錯誤，跳過: {date_part}", file=sys.stderr)
                
        except Exception as e:
            print(f"  ⚠️ 無法解析休假人員資訊 '{entry}': {e}", file=sys.stderr)
    
    VACATION_STAFF = vacation_staff
    
    if vacation_staff:
        print(f"\n  📊 休假人員統計:", file=sys.stderr)
        for date_str in sorted(vacation_staff.keys()):
            people = vacation_staff[date_str]
            print(f"      {date_str}: {', '.join(sorted(people))} (共 {len(people)} 人)", file=sys.stderr)
    
    return vacation_staff

def is_person_on_vacation(person, date_obj):
    """檢查人員在指定日期是否請假"""
    if isinstance(date_obj, datetime):
        date_obj = date_obj.date()
    
    date_str = date_obj.strftime('%Y-%m-%d')
    
    if date_str in VACATION_STAFF:
        return person.lower() in VACATION_STAFF[date_str] # ✅ 修正：比對時轉為小寫
    
    return False
# --------------------------------------------------------------------
# 3. 資料載入
# --------------------------------------------------------------------
def load_all_data(demand_xlsx_path):
    """載入所有資料"""
    print(f"Step 1: 正在載入資料 ({CURRENT_VERSION})...", file=sys.stderr)
    db_tables = {} 
    df_demand = pd.DataFrame()
    df_constraints = pd.DataFrame()

    # 載入 Dosing_Constraints
    try:
        with sqlite3.connect(MAIN_DB_PATH) as conn:
            df_constraints_raw = pd.read_sql("SELECT * FROM '配藥限制'", conn)
            
        df_constraints = df_constraints_raw.copy()
        
        def norm_col(s: str) -> str:
            return (str(s).strip().replace(" ", "").replace("_blank", "").lower())

        rename_map = {}
        for c in df_constraints.columns:
            nz = norm_col(c)
            if nz in ("可用凍乾機", "可用凍乾機台", "dryers", "dryer"):
                rename_map[c] = "可用凍乾機"
            elif nz == "pn":
                rename_map[c] = "PN"
            elif nz in ("配藥人-1","人員1","staff1"):
                rename_map[c] = "配藥人-1"
            elif nz in ("配藥人-2","人員2","staff2"):
                rename_map[c] = "配藥人-2"
            elif nz in ("配藥人-3","人員3","staff3"):
                rename_map[c] = "配藥人-3"
            elif nz in ("交藥時間","給藥時間","rd給藥時間"):
                rename_map[c] = "交藥時間"
            elif nz in ("凍乾時間", "freezehrs", "freezehour", "freezehr"):
                rename_map[c] = "凍乾時間"
            elif nz in ("滴定後凍乾時間差距(hr)", "滴定後凍乾時間差距", "waithr", "waithrs"):
                rename_map[c] = "滴定後凍乾時間差距 (HR)"

        df_constraints = df_constraints.rename(columns=rename_map)
        df_constraints = df_constraints[df_constraints['PN'].notna() & (df_constraints['PN'] != "")]
        df_constraints["PN"] = standardize_pn(df_constraints["PN"])
        
        num_cols = ["凍乾時間", "滴定後凍乾時間差距 (HR)"]
        for col in num_cols:
            if col in df_constraints.columns:
                df_constraints[col] = pd.to_numeric(df_constraints[col], errors='coerce').fillna(0)
            else:
                df_constraints[col] = 0 
        
        def count_csv(s):
            if not isinstance(s, str) or s == "": return 99
            return len([d for d in s.split(',') if d.strip()])
        
        df_constraints["V9_Dryer_Count"] = df_constraints["可用凍乾機"].apply(count_csv)

        def count_people(row):
            count = 0
            if isinstance(row.get("配藥人-1"), str) and row.get("配藥人-1", "").strip(): count += 1
            if isinstance(row.get("配藥人-2"), str) and row.get("配藥人-2", "").strip(): count += 1
            if isinstance(row.get("配藥人-3"), str) and row.get("配藥人-3", "").strip(): count += 1
            return count if count > 0 else 99
            
        df_constraints["V9_Person_Count"] = df_constraints.apply(count_people, axis=1)
        
        def get_slots(s):
            if not isinstance(s, str) or s == "": return ["AM", "PM"] 
            times = [t.strip() for t in s.split(',') if t.strip()]
            slots = set()
            
            for t in times:
                try:
                    hour = int(t.split(':')[0])
                    if 9 <= hour <= 14:
                        slots.add("AM")
                    elif 15 <= hour <= 23:
                        slots.add("PM")
                except:
                    pass

            return list(slots) if slots else ["AM", "PM"]

        df_constraints["V9_Slot_List"] = df_constraints["交藥時間"].apply(get_slots)
        df_constraints["V9_Single_Slot"] = df_constraints["交藥時間"].apply(
            lambda s: isinstance(s, str) and len(s.split(',')) == 1 and ":" in s
        )

        print("  ✅ '配藥限制' 載入完成", file=sys.stderr)

    except Exception as e:
        print(f"❌ 無法載入 '配藥限制': {e}", file=sys.stderr)
        return None 

    # 載入 BDC
    df_bdc = pd.DataFrame()
    try:
        with sqlite3.connect(MAIN_DB_PATH) as conn:
            df_bdc_raw = pd.read_sql("SELECT * FROM Beads_Dry_Count", conn)
            
            if len(df_bdc_raw.columns) < 4:
                raise ValueError(f"Beads_Dry_Count 欄位不足")
            
            col_map = {
                df_bdc_raw.columns[0]: "BDC_Group_Name", 
                df_bdc_raw.columns[1]: "料號",          
                df_bdc_raw.columns[2]: "BDC_Marker_Name",
                df_bdc_raw.columns[3]: "BDC_Prod_Qty"   
            }
            df_bdc = df_bdc_raw.rename(columns=col_map)
            df_bdc["料號"] = standardize_pn(df_bdc["料號"])
            df_bdc["BDC_Prod_Qty"] = pd.to_numeric(df_bdc["BDC_Prod_Qty"], errors='coerce').fillna(0)
            select_cols = ["BDC_Group_Name", "料號", "BDC_Marker_Name", "BDC_Prod_Qty"]
            df_bdc = df_bdc[select_cols]
            # ✅ 新增：標準化 BDC 的藥名 (去除空白)
            df_bdc["BDC_Group_Name"] = df_bdc["BDC_Group_Name"].astype(str).str.strip()
            print(f"  ✅ BDC 載入完成 ({len(df_bdc)} 筆)", file=sys.stderr)
            
    except Exception as e:
        print(f"❌ 無法載入 BDC: {e}", file=sys.stderr)
        return None 

    # 載入需求表
    try:
        print(f"  ⏳ 正在載入需求檔: {demand_xlsx_path}", file=sys.stderr)
        # ✅ 修正：使用 header=1 (因為標題在 Excel 的第二行)
        df_demand = pd.read_excel(demand_xlsx_path, sheet_name="滴定排程需求表", header=1, engine='openpyxl')
        
        def norm(s: str) -> str:
            if s is None: return ""
            return str(s).strip().replace(" ", "").replace("　", "").replace("＋", "").replace("週", "周")

        def find_col(*keywords):
            for raw in df_demand.columns:
                nz = norm(raw)
                if all(k in nz for k in keywords):
                    return raw
            return None

        col_stock = find_col("庫存", "滴定") or find_col("庫存")
        col_w1 = find_col("第一","周","需求") or "第一周需求"
        col_w2 = find_col("第二","周","需求") or "第二周需求"
        col_w3 = find_col("第三","周","需求") or "第三周需求"
        col_pn = find_col("料號") or "料號"
        col_group = find_col("藥名") or find_col("Group") or "BDC_Group_Name" # ✅ 新增

        df_demand[col_pn] = standardize_pn(df_demand[col_pn])
        for c in [col_stock, col_w1, col_w2, col_w3]:
            df_demand[c] = pd.to_numeric(df_demand[c], errors="coerce").fillna(0)

        df_demand = df_demand.rename(columns={
            col_pn: "料號", 
            col_group: "BDC_Group_Name", # ✅ 新增
            col_stock: "Stock_plus_Dosing",
            col_w1: "第一周需求", col_w2: "第二周需求", col_w3: "第三周需求"
        })
        # ✅ 修改：新增 BDC_Group_Name 到欄位列表中
        df_demand = df_demand[["料號", "BDC_Group_Name", "Stock_plus_Dosing", "第一周需求", "第二周需求", "第三周需求"]]
        # ✅ 新增：標準化藥名 (去除空白)
        df_demand["BDC_Group_Name"] = df_demand["BDC_Group_Name"].astype(str).str.strip()
        
        print(f"  ✅ 需求表載入完成 ({len(df_demand)} 筆)", file=sys.stderr)

    except Exception as e:
        print(f"❌ 無法載入需求表: {e}", file=sys.stderr)
        return None

    # 合併
    if df_demand is None or df_bdc is None or df_constraints is None:
        return None
        
    # ✅ 修改：使用 ["料號", "BDC_Group_Name"] 作為合併的 KEY
    df_merged = pd.merge(df_demand, df_bdc, on=["料號", "BDC_Group_Name"], how="left")
    
    df_constraints_v9 = df_constraints[[
        "PN", "V9_Dryer_Count", "V9_Person_Count", "V9_Slot_List", "V9_Single_Slot",
        "凍乾時間", "滴定後凍乾時間差距 (HR)", "可用凍乾機", "配藥人-1", "配藥人-2", "配藥人-3",
        "交藥時間",
        "備註",
        "Port數"  # ✅ 新增：讀取 Port 數欄位
        ]].rename(columns={"PN": "料號"})
    
    df_merged = pd.merge(df_merged, df_constraints_v9, on="料號", how="left")
    df_merged = df_merged.dropna(subset=["BDC_Group_Name"])
    
    print(f"  ✅ 合併完成 ({len(df_merged)} 筆有效需求)\n", file=sys.stderr)

    # 載入其他資料庫
    try:
        with sqlite3.connect(MAIN_DB_PATH) as conn:
            df_plan = pd.read_sql("SELECT * FROM production_Plan", conn)
            date_cols_map = parse_date_columns(df_plan)
            if date_cols_map:
                for col in date_cols_map.keys():
                    df_plan[col] = pd.to_numeric(df_plan[col], errors='coerce').fillna(0)

            db_tables["production_Plan"] = df_plan
            db_tables["BOM_Details"] = pd.read_sql("SELECT * FROM BOM_Details", conn)

        db_tables["BOM_Details"]["Finished_PartNo"] = standardize_pn(db_tables["BOM_Details"]["Finished_PartNo"])
        db_tables["BOM_Details"]["Component_No"] = standardize_pn(db_tables["BOM_Details"]["Component_No"])
        db_tables["production_Plan"]["PN"] = standardize_pn(db_tables["production_Plan"]["PN"])
        
        # ✅ 修改：不再使用 set_index
        db_tables["Beads_Dry_Count_Info"] = df_bdc.copy()
        db_tables["Dosing_Constraints"] = df_constraints.set_index("PN")

        # === V9.9: 載入 P0 (限制 OR 插單) ===
        try:
            df_p0_orders = pd.read_sql("SELECT * FROM '限制_OR_插單'", conn)
            print(f"  ... P0 原始欄位: {df_p0_orders.columns.to_list()}", file=sys.stderr)
            
            def clean_col_name(col):
                return str(col).strip().replace(':', '')
            
            df_p0_orders.columns = [clean_col_name(c) for c in df_p0_orders.columns]
            print(f"  ... P0 清理後欄位: {df_p0_orders.columns.to_list()}", file=sys.stderr)
            
            if "日期" in df_p0_orders.columns:
                df_p0_orders["日期"] = df_p0_orders["日期"].astype(str)
            
            db_tables["P0_Orders"] = df_p0_orders
            print(f"  ✅ '限制 OR 插單' 載入完成 ({len(df_p0_orders)} 筆)", file=sys.stderr)
        except Exception as e:
            print(f"  ⚠️ 警告: 無法載入 '限制 OR 插單' 表: {e}", file=sys.stderr)
            db_tables["P0_Orders"] = pd.DataFrame()

        print("  ✅ 資料庫載入完成\n", file=sys.stderr)

    except Exception as e:
        print(f"⚠️ 資料庫載入警告: {e}", file=sys.stderr)

    return {
        "demand": df_merged, 
        "db": db_tables,
        "output_template": pd.DataFrame() 
    }

# --------------------------------------------------------------------
# 4. 需求計算
# --------------------------------------------------------------------
def calculate_demand_queue(data: dict):
    print("Step 2: 計算 P1-P3 (缺料) 需求優先級...", file=sys.stderr)

    df = data["demand"] 
    db = data["db"]

    df_prod_plan = db.get("production_Plan", pd.DataFrame())
    bom = db.get("BOM_Details", pd.DataFrame())
    
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    monday = (today + timedelta(days=-today.weekday(), weeks=0)) 
    next_wednesday = monday + timedelta(days=2) 
    
    w1_urgent_wed_pn = set()
    w1_urgent_other_pn = set()
    
    if not df_prod_plan.empty and not bom.empty:
        date_cols_map = parse_date_columns(df_prod_plan)
        
        if date_cols_map and "PN" in df_prod_plan.columns:
            urgent_cols = [col for col, dt in date_cols_map.items() if dt <= next_wednesday]
            other_cols = [col for col, dt in date_cols_map.items() if dt > next_wednesday]
            
            df_plan_copy = df_prod_plan.set_index("PN")
            
            if urgent_cols:
                urgent_qty = df_plan_copy[urgent_cols].sum(axis=1)
                urgent_pns = set(urgent_qty[urgent_qty > 0].index)
                w1_urgent_wed_pn = set(bom[bom["Finished_PartNo"].isin(urgent_pns)]["Component_No"])
            
            if other_cols:
                other_qty = df_plan_copy[other_cols].sum(axis=1)
                other_pns = set(other_qty[other_qty > 0].index)
                w1_urgent_other_pn = set(bom[bom["Finished_PartNo"].isin(other_pns)]["Component_No"])

            print(f"  半品: {len(w1_urgent_wed_pn)} (週三前), {len(w1_urgent_other_pn)} (其他)", file=sys.stderr)

    # ==========================================================
    # ✅ 修正：按 BDC_Group_Name (藥名) 分組計算總需求 (1:1 邏輯)
    # ==========================================================
    
    agg_cols = {
        'Stock_plus_Dosing': 'min', 
        '第一周需求': 'max',       
        '第二周需求': 'max',       
        '第三周需求': 'max',       
        'BDC_Prod_Qty': 'first', 
        'BDC_Marker_Name': 'first',
        'V9_Dryer_Count': 'first',
        'V9_Person_Count': 'first'
    }
    
    def agg_pn_list(pn_series):
        return list(pn_series.unique())
    agg_cols['料號'] = agg_pn_list

    df_clean = df[df['BDC_Group_Name'].notna() & (df['BDC_Group_Name'] != "")]
    
    # ==========================================================
    # ✅ 移除 tCREA 特例，所有藥品使用統一的 1:1 邏輯
    # ==========================================================
    df_grouped = df_clean.groupby('BDC_Group_Name').agg(agg_cols).reset_index()

    # ==========================================================
    # 4. 遍歷分組後的 df_grouped (P1, P2, P3)
    # ==========================================================
    task_queue = [] # P1-P5 任務
    p6_p7_tasks = [] # P6/P7 任務 (填補空缺用)
    
    for _, r in df_grouped.iterrows():
        group_name = r["BDC_Group_Name"]
        prod_qty = r["BDC_Prod_Qty"] 
        marker_name = r["BDC_Marker_Name"] 
        dryer_prio = r.get("V9_Dryer_Count", 99)
        person_prio = r.get("V9_Person_Count", 99)
        pns_in_group = r.get("料號", [])

        # ========================================
        # ✅ 新增：強制忽略 5714400116 的獨立需求
        # ========================================
        # 116 是 K1/K3 的共用料，只由 226 或 117 觸發
        if "5714400116" in pns_in_group:
            # 如果這組只有 116，直接跳過
            if len(pns_in_group) == 1:
                print(f"  🚫 忽略獨立需求: 5714400116 (將隨 K1/K3 排程)", file=sys.stderr)
                continue

        if pd.isna(prod_qty) or prod_qty == 0: 
            print(f"  ⚠️ 警告: 群組 {group_name} 的 BDC_Prod_Qty 為 0 或 NaN (可能未在 BDC 資料庫中定義)，跳過此群組。", file=sys.stderr)
            continue

        w1_balance = r["Stock_plus_Dosing"] - r["第一周需求"]
        w1_short = abs(w1_balance) if w1_balance < 0 else 0
        w1_prod_count = math.ceil(w1_short / prod_qty) if w1_short > 0 else 0
        w1_planned = w1_prod_count * prod_qty

        w2_balance = w1_balance + w1_planned - r["第二周需求"]
        w2_short = abs(w2_balance) if w2_balance < 0 else 0
        w2_prod_count = math.ceil(w2_short / prod_qty) if w2_short > 0 else 0
        w2_planned = w2_prod_count * prod_qty

        w3_balance = w2_balance + w2_planned - r["第三周需求"]
        w3_short = abs(w3_balance) if w3_balance < 0 else 0
        
        has_urgent_wed = any(pn in w1_urgent_wed_pn for pn in pns_in_group)
        has_urgent_other = any(pn in w1_urgent_other_pn for pn in pns_in_group)
        representative_pn = pns_in_group[0] if pns_in_group else "UNKNOWN"
        
        if w1_short > 0:
            if has_urgent_wed: plan_prio = 0
            elif has_urgent_other: plan_prio = 1
            else: plan_prio = 2
            subp = (dryer_prio, person_prio, plan_prio)
            task_queue.append((1, subp, representative_pn, w1_short, "W1", group_name, marker_name, prod_qty))
        
        if w2_short > 0:
            subp = (dryer_prio, person_prio, 0)
            task_queue.append((2, subp, representative_pn, w2_short, "W2", group_name, marker_name, prod_qty))
        
        if w3_short > 0:
            subp = (dryer_prio, person_prio, 0)
            task_queue.append((3, subp, representative_pn, w3_short, "W3", group_name, marker_name, prod_qty))

    # ==========================================================
    # ✅ P3 特殊庫存規則
    # ==========================================================
    print("Step 2a: 檢查 P3 特殊庫存規則...", file=sys.stderr)
    existing_task_groups = set(t[5] for t in task_queue)
    df_all_demand_data = data["demand"].set_index('料號')
    
    def create_p3_stock_task(pn, group_name, marker_name, prod_qty, dryer_prio, person_prio):
        if pd.isna(prod_qty) or prod_qty == 0:
            print(f"  ⚠️ 警告: P3 庫存規則 {group_name} 凍乾數為 0 或 NaN，無法建立任務", file=sys.stderr)
            return None
        print(f"  ✅ 觸發 P3 庫存規則: {group_name} (PN: {pn})", file=sys.stderr)
        subp = (dryer_prio, person_prio, 0) # 0 = stock priority
        return (3, subp, pn, 0, "P3_Stock", group_name, marker_name, prod_qty)

    # --- 規則 1: CREA (5714400197, 5714400198) < 4000 ---
    rule1_pns = ['5714400197', '5714400198']
    rule1_threshold = 4000
    try:
        rule1_stocks = df_all_demand_data.loc[rule1_pns]['Stock_plus_Dosing']
        min_stock_r1 = rule1_stocks.min()
        if min_stock_r1 < rule1_threshold:
            pn_info_r1 = df_all_demand_data.loc['5714400197']
            rule1_group_name = pn_info_r1['BDC_Group_Name']
            if rule1_group_name not in existing_task_groups:
                task = create_p3_stock_task('5714400197', rule1_group_name, pn_info_r1['BDC_Marker_Name'], pn_info_r1['BDC_Prod_Qty'], pn_info_r1.get('V9_Dryer_Count', 99), pn_info_r1.get('V9_Person_Count', 99))
                if task: 
                    task_queue.append(task)
                    existing_task_groups.add(rule1_group_name)
    except KeyError:
        print(f"  ⚠️ 警告: P3 庫存規則 PN {rule1_pns} 不在需求檔中，跳過規則 1", file=sys.stderr)
    except Exception as e:
        print(f"  ⚠️ 警告: P3 庫存規則 1 失敗: {e}", file=sys.stderr)

    # --- 規則 2: IVEK-Na (5714400202, 5714400203) < 30000 ---
    rule2_pns = ['5714400202', '5714400203']
    rule2_threshold = 50000
    try:
        rule2_stocks = df_all_demand_data.loc[rule2_pns]['Stock_plus_Dosing']
        min_stock_r2 = rule2_stocks.min()
        if min_stock_r2 < rule2_threshold:
            pn_info_r2 = df_all_demand_data.loc['5714400202']
            rule2_group_name = pn_info_r2['BDC_Group_Name']
            if rule2_group_name not in existing_task_groups:
                task = create_p3_stock_task('5714400202', rule2_group_name, pn_info_r2['BDC_Marker_Name'], pn_info_r2['BDC_Prod_Qty'], pn_info_r2.get('V9_Dryer_Count', 99), pn_info_r2.get('V9_Person_Count', 99))
                if task: 
                    task_queue.append(task)
                    existing_task_groups.add(rule2_group_name)
    except KeyError:
        print(f"  ⚠️ 警告: P3 庫存規則 PN {rule2_pns} 不在需求檔中，跳過規則 2", file=sys.stderr)
    except Exception as e:
        print(f"  ⚠️ 警告: P3 庫存規則 2 失敗: {e}", file=sys.stderr)

    # ==========================================================
    # === P4/P5 補庫存任務 ===
    # ==========================================================
    print("Step 2b: 計算 P4/P5 (補庫存) 任務...", file=sys.stderr)
    constraints_df = data["db"].get("Dosing_Constraints", pd.DataFrame())
    IDLE_COL_NAME = "空閒產能時可優先安排生產"
    idle_pns = set()
    
    if IDLE_COL_NAME in constraints_df.columns:
        idle_filter = constraints_df[IDLE_COL_NAME].notna() & (constraints_df[IDLE_COL_NAME] != "")
        idle_pns = set(constraints_df[idle_filter].index.unique())
        print(f"  ✅ 找到 {len(idle_pns)} 個 P4/P5 補庫存標記", file=sys.stderr)
    else:
        print(f"⚠️ 找不到 '{IDLE_COL_NAME}' 欄位，跳過 P4/P5 任務", file=sys.stderr)

    added_groups_p4 = set()
    added_groups_p5 = set()
    df_all_demand_data_no_index = data["demand"] 

    if idle_pns:
        shortage_pns_in_groups = existing_task_groups 
        df_idle_tasks_data = df_all_demand_data_no_index[
            df_all_demand_data_no_index['料號'].isin(idle_pns) &
            ~df_all_demand_data_no_index['BDC_Group_Name'].isin(shortage_pns_in_groups)
        ]
        df_idle_tasks_data = df_idle_tasks_data.sort_values(by='Stock_plus_Dosing', ascending=True)
        
        new_idle_tasks_p4 = 0
        for _, r in df_idle_tasks_data.iterrows():
            group_name = r["BDC_Group_Name"]
            if group_name in added_groups_p4: continue
            pn = str(r["料號"]).strip()
            prod_qty = r["BDC_Prod_Qty"]
            if pd.isna(prod_qty) or prod_qty == 0: continue 
            marker_name = r["BDC_Marker_Name"]
            stock_level = r["Stock_plus_Dosing"]
            dryer_prio = r.get("V9_Dryer_Count", 99)
            person_prio = r.get("V9_Person_Count", 99)
            priority = 4; tag = "Idle"
            subp = (dryer_prio, person_prio, stock_level)
            task_queue.append((priority, subp, pn, 0, tag, group_name, marker_name, prod_qty))
            added_groups_p4.add(group_name)
            new_idle_tasks_p4 += 1
        print(f"  ✅ 新增 {new_idle_tasks_p4} 個 P4 任務", file=sys.stderr)

        new_idle_tasks_p5 = 0
        for _, r in df_idle_tasks_data.iterrows():
            group_name = r["BDC_Group_Name"]
            if group_name in added_groups_p5: continue
            pn = str(r["料號"]).strip()
            prod_qty = r["BDC_Prod_Qty"]
            if pd.isna(prod_qty) or prod_qty == 0: continue 
            marker_name = r["BDC_Marker_Name"]
            stock_level = r["Stock_plus_Dosing"]
            dryer_prio = r.get("V9_Dryer_Count", 99)
            person_prio = r.get("V9_Person_Count", 99)
            priority = 5; tag = "Idle_R2"
            subp = (dryer_prio, person_prio, stock_level)
            task_queue.append((priority, subp, pn, 0, tag, group_name, marker_name, prod_qty))
            added_groups_p5.add(group_name)
            new_idle_tasks_p5 += 1
        print(f"  ✅ 新增 {new_idle_tasks_p5} 個 P5 任務", file=sys.stderr)

    # ========================================
    # Step 2c: 計算 P6/P7 (自動補庫存) 任務
    # ========================================
    print("Step 2c: 計算 P6/P7 (自動補庫存) 任務...", file=sys.stderr)

    # 1. 取得所有已在 P1-P5 清單中的群組
    all_p1_p5_groups = existing_task_groups.union(added_groups_p4, added_groups_p5)

    p6_p7_tasks = []

    p6_rule_cols = ['Port數']
    if all(col in df_all_demand_data_no_index.columns for col in p6_rule_cols):
        
        # ========================================
        # 2. 篩選 P6/P7 候選 PNs（原有邏輯）
        # ========================================
        
        df_p6_p7_candidates = df_all_demand_data_no_index[
            (~df_all_demand_data_no_index['BDC_Group_Name'].isin(all_p1_p5_groups))
        ].copy()
        
        # 規則 2.1: Port數 必須為 1 或 2
        df_p6_p7_candidates = df_p6_p7_candidates[
            df_p6_p7_candidates['Port數'].astype(str).isin(['1', '2'])
        ]
        
        # ========================================
        # ✅ 修正：檢查「庫存安排最大量」
        # 使用 data 參數而非 self
        # ========================================
        
        print(f"  ℹ️ 開始檢查「庫存安排最大量」（從配藥限制表）", file=sys.stderr)
        
        # 確認欄位名稱
        possible_max_stock_cols = [
            '庫存安排最大量',
            '庫存最大量',
            '最大庫存量',
            '庫存安排',
            'Max_Stock'
        ]
        
        max_stock_col = None
        constraints = None
        
        # ✅ 從 data 中取得 constraints（而非 self.constraints）
        if 'constraints' in data and data['constraints'] is not None:
            constraints = data['constraints']
            
            for col_name in possible_max_stock_cols:
                if col_name in constraints.columns:
                    max_stock_col = col_name
                    print(f"  ✅ 找到欄位: {col_name}", file=sys.stderr)
                    break
        else:
            print(f"  ⚠️ 警告: data 中沒有 constraints 表", file=sys.stderr)
        
        if max_stock_col is None:
            print(f"  ⚠️ 警告: constraints 表中找不到「庫存安排最大量」欄位", file=sys.stderr)
            if constraints is not None:
                print(f"  ⚠️ constraints 可用欄位: {constraints.columns.tolist()}", file=sys.stderr)
        
        # 篩選：只保留「庫存安排最大量」有值的 PN
        original_count = len(df_p6_p7_candidates)
        filtered_pns = []
        excluded_pns = []
        
        for idx, row in df_p6_p7_candidates.iterrows():
            pn = str(row['料號']).strip()
            
            # ✅ 查詢 constraints 表
            should_include = False
            
            if max_stock_col is not None and constraints is not None and pn in constraints.index:
                c = constraints.loc[pn]
                if isinstance(c, pd.DataFrame):
                    c = c.iloc[0]
                
                # 檢查「庫存安排最大量」是否有值
                max_stock_value = c.get(max_stock_col)
                
                if pd.notna(max_stock_value) and str(max_stock_value).strip() != '':
                    should_include = True
            
            if should_include:
                filtered_pns.append(idx)
            else:
                excluded_pns.append(pn)
        
        # 更新候選清單
        df_p6_p7_candidates = df_p6_p7_candidates.loc[filtered_pns]
        
        filtered_count = len(df_p6_p7_candidates)
        excluded_count = original_count - filtered_count
        
        print(f"  ✅ 根據「{max_stock_col or '庫存安排最大量'}」篩選:", file=sys.stderr)
        print(f"     原始候選: {original_count} 個", file=sys.stderr)
        print(f"     有設定的: {filtered_count} 個", file=sys.stderr)
        print(f"     已排除的: {excluded_count} 個", file=sys.stderr)
        
        # 顯示被排除的 PN（前 10 個）
        if excluded_count > 0 and excluded_count <= 10:
            print(f"     排除的 PN: {', '.join(excluded_pns)}", file=sys.stderr)
        elif excluded_count > 10:
            print(f"     排除的 PN: {', '.join(excluded_pns[:10])} ... (共 {excluded_count} 個)", file=sys.stderr)
        
        # ========================================
        # 規則 2.2 & 2.3: 根據庫存分類 P6/P7
        # ========================================
        
        df_p6_tasks = df_p6_p7_candidates[
            df_p6_p7_candidates['Stock_plus_Dosing'] < 100000
        ]
        
        df_p7_tasks = df_p6_p7_candidates[
            df_p6_p7_candidates['Stock_plus_Dosing'] >= 100000
        ]

        # 3. 排序 - 庫存最低的優先
        df_p6_tasks = df_p6_tasks.sort_values(by='Stock_plus_Dosing', ascending=True)
        df_p7_tasks = df_p7_tasks.sort_values(by='Stock_plus_Dosing', ascending=True)

        # ========================================
        # 4. 加入 P6 任務
        # ========================================
        
        new_idle_tasks_p6 = 0
        for _, r in df_p6_tasks.iterrows():
            group_name = r["BDC_Group_Name"]
            pn = str(r["料號"]).strip()
            prod_qty = r["BDC_Prod_Qty"]
            
            if pd.isna(prod_qty) or prod_qty == 0:
                continue
            
            marker_name = r["BDC_Marker_Name"]
            stock_level = r["Stock_plus_Dosing"]
            dryer_prio = r.get("V9_Dryer_Count", 99)
            person_prio = r.get("V9_Person_Count", 99)
            priority = 6
            tag = "P6_Stock_L"
            subp = (dryer_prio, person_prio, stock_level)
            
            p6_p7_tasks.append((priority, subp, pn, 0, tag, group_name, marker_name, prod_qty))
            new_idle_tasks_p6 += 1
        
        print(f"  ✅ 新增 {new_idle_tasks_p6} 個 P6 任務 (<100k)", file=sys.stderr)

        # ========================================
        # 5. 加入 P7 任務
        # ========================================
        
        new_idle_tasks_p7 = 0
        for _, r in df_p7_tasks.iterrows():
            group_name = r["BDC_Group_Name"]
            pn = str(r["料號"]).strip()
            prod_qty = r["BDC_Prod_Qty"]
            
            if pd.isna(prod_qty) or prod_qty == 0:
                continue
            
            marker_name = r["BDC_Marker_Name"]
            stock_level = r["Stock_plus_Dosing"]
            dryer_prio = r.get("V9_Person_Count", 99)
            person_prio = r.get("V9_Person_Count", 99)
            priority = 7
            tag = "P7_Stock_H"
            subp = (dryer_prio, person_prio, stock_level)
            
            p6_p7_tasks.append((priority, subp, pn, 0, tag, group_name, marker_name, prod_qty))
            new_idle_tasks_p7 += 1
        
        print(f"  ✅ 新增 {new_idle_tasks_p7} 個 P7 任務 (>=100k, 填補空缺)", file=sys.stderr)

    else:
        print(f"  ⚠️ 警告: 找不到 'Port數' 欄位，跳過 P6/P7 任務", file=sys.stderr)

    
    # ==========================================================
    # ✅ P1-P5 任務清單 (主佇列)
    # ==========================================================
    task_queue.sort(key=lambda x: (x[0], x[5], x[1][0], x[1][1], x[1][2], x[2])) # ✅ Rule 2: 按藥名 (x[5]) 排序
    task_dict = {(t[5], t[2]): t for t in task_queue} 

    print(f"✅ 需求計算完成: {len(task_queue)} 個 P1-P5 任務, {len(p6_p7_tasks)} 個 P6/P7 任務\n", file=sys.stderr)

    return task_queue, task_dict, p6_p7_tasks

class ResourceManager:
    """V9.9.1: 資源管理器 (支援 P0 實驗/非實驗模式)"""
    
    def __init__(self):
        # P1-P5 資源
        self.person_usage = {}
        self.person_day_limit = {}
        self.dryer_usage = {}
        
        # V9.9: P0 專用資源 (非實驗模式：強制，P1-P5 都不能衝突)
        self.p0_strict_person_usage = {}
        self.p0_strict_dryer_usage = {}
        self.p0_strict_port_usage = {}
        self.p0_strict_ivek_usage = {}
        
        # V9.9.1: P0 實驗模式資源 (P1/P2 允許衝突，P3-P5 不能衝突)
        self.p0_experiment_person_usage = {}
        self.p0_experiment_dryer_usage = {}
        self.p0_experiment_port_usage = {}
        self.p0_experiment_ivek_usage = {}
        
        # ✅ 新增：凍乾機維修狀態追蹤
        self.dryer_maintenance = {}  # {day_key: set(['LD-3', 'LD-5', ...])}

    def reset_shift(self, shift_key):
        """重置某個班次的 P1-P5 資源"""
        self.person_usage[shift_key] = set()
        
    def reset_day(self, day_key):
        """重置某天的 P1-P5 資源"""
        self.person_day_limit[day_key] = set()
        self.dryer_usage[day_key] = set()

    # ========================================
    # V9.9: P0 非實驗任務預訂 (強制，P1-P5 都不能衝突)
    # ========================================

    def book_p0_strict_person(self, person, shift_key, day_key):
        """P0 (非實驗) 強制預訂人員"""
        person_lower = person.lower() # ✅ 修正：使用小寫
        if shift_key not in self.p0_strict_person_usage:
            self.p0_strict_person_usage[shift_key] = set()
        if shift_key not in self.person_usage:
            self.person_usage[shift_key] = set()

        if person_lower in self.p0_strict_person_usage[shift_key]: # ✅ 修正：使用小寫
            print(f"    ⚠️ P0 警告: 人員 {person} 在 {shift_key} 已被其他 P0 任務佔用", file=sys.stderr)
        
        if person_lower in self.person_usage[shift_key]: # ✅ 修正：使用小寫
            print(f"    ⚠️ P0 警告: 人員 {person} 在 {shift_key} 已被 P1-P5 任務佔用", file=sys.stderr)

        self.p0_strict_person_usage[shift_key].add(person_lower) # ✅ 修正：使用小寫
        self.person_usage[shift_key].add(person_lower) # ✅ 修正：使用小寫
        print(f"    🟢 P0 預訂人員: {person} ({shift_key})", file=sys.stderr) # Log 仍顯示原始大小寫

    def book_p0_strict_dryer(self, dryer, day_key):
        """P0 (非實驗) 強制預訂凍乾機"""
        if day_key not in self.p0_strict_dryer_usage:
            self.p0_strict_dryer_usage[day_key] = set()
        if day_key not in self.dryer_usage:
            self.dryer_usage[day_key] = set()
            
        if dryer in self.p0_strict_dryer_usage[day_key]:
            print(f"    ⚠️ P0 警告: 凍乾機 {dryer} 在 {day_key} 已被其他 P0 任務佔用", file=sys.stderr)
        
        if dryer in self.dryer_usage[day_key]:
            print(f"    ⚠️ P0 警告: 凍乾機 {dryer} 在 {day_key} 已被 P1-P5 任務佔用", file=sys.stderr)

        self.p0_strict_dryer_usage[day_key].add(dryer)
        self.dryer_usage[day_key].add(dryer)
        print(f"    🟢 P0 預訂凍乾機: {dryer} ({day_key})", file=sys.stderr)

    def book_p0_strict_port_pair(self, shift_key, port_num):
        """P0 (非實驗) 強制預訂 Port"""
        if shift_key not in self.p0_strict_port_usage:
            self.p0_strict_port_usage[shift_key] = set()

        if port_num % 2 == 0:
            pair = (port_num - 1, port_num)
        else:
            pair = (port_num, port_num + 1)
        
        port1, port2 = pair
        
        if (port1 in self.p0_strict_port_usage[shift_key]) or (port2 in self.p0_strict_port_usage[shift_key]):
             print(f"    ⚠️ P0 警告: Port {port1}/{port2} 在 {shift_key} 已被其他 P0 任務佔用", file=sys.stderr)
        
        self.p0_strict_port_usage[shift_key].add(port1)
        self.p0_strict_port_usage[shift_key].add(port2)
        print(f"    🟢 P0 預訂 Ports: {port1} 和 {port2} ({shift_key})", file=sys.stderr)

    def book_p0_strict_ivek(self, day_key):
        """P0 (非實驗) 強制預訂 IVEK"""
        if day_key not in self.p0_strict_ivek_usage:
            self.p0_strict_ivek_usage[day_key] = set()
            
        if "IVEK" in self.p0_strict_ivek_usage[day_key]:
            print(f"    ⚠️ P0 警告: IVEK 在 {day_key} 已被其他 P0 任務佔用", file=sys.stderr)
            
        self.p0_strict_ivek_usage[day_key].add("IVEK")
        print(f"    🟢 P0 預訂 IVEK ({day_key})", file=sys.stderr)

    
    # ========================================
    # V9.9.1: P0 實驗任務預訂 (P1/P2 允許衝突，P3-P5 不能衝突)
    # ========================================

    def book_p0_experiment_person(self, person, shift_key):
        """P0 (實驗) 預訂人員"""
        person_lower = person.lower() # ✅ 修正：使用小寫
        if shift_key not in self.p0_experiment_person_usage:
            self.p0_experiment_person_usage[shift_key] = set()
        
        if person_lower in self.p0_experiment_person_usage[shift_key]: # ✅ 修正：使用小寫
            print(f"    ⚠️ P0 (實驗) 警告: 人員 {person} 在 {shift_key} 已被其他實驗 P0 佔用", file=sys.stderr)
        
        self.p0_experiment_person_usage[shift_key].add(person_lower) # ✅ 修正：使用小寫
        print(f"    🧪 P0 (實驗) 預訂人員: {person} ({shift_key})", file=sys.stderr) # Log 仍顯示原始大小寫

    def book_p0_experiment_dryer(self, dryer, day_key):
        """P0 (實驗) 預訂凍乾機"""
        if day_key not in self.p0_experiment_dryer_usage:
            self.p0_experiment_dryer_usage[day_key] = set()
        
        if dryer in self.p0_experiment_dryer_usage[day_key]:
            print(f"    ⚠️ P0 (實驗) 警告: 凍乾機 {dryer} 在 {day_key} 已被其他實驗 P0 佔用", file=sys.stderr)
        
        self.p0_experiment_dryer_usage[day_key].add(dryer)
        print(f"    🧪 P0 (實驗) 預訂凍乾機: {dryer} ({day_key})", file=sys.stderr)

    def book_p0_experiment_port_pair(self, shift_key, port_num):
        """P0 (實驗) 預訂 Port"""
        if shift_key not in self.p0_experiment_port_usage:
            self.p0_experiment_port_usage[shift_key] = set()

        if port_num % 2 == 0:
            pair = (port_num - 1, port_num)
        else:
            pair = (port_num, port_num + 1)
        
        port1, port2 = pair
        
        if (port1 in self.p0_experiment_port_usage[shift_key]) or (port2 in self.p0_experiment_port_usage[shift_key]):
             print(f"    ⚠️ P0 (實驗) 警告: Port {port1}/{port2} 在 {shift_key} 已被其他實驗 P0 佔用", file=sys.stderr)
        
        self.p0_experiment_port_usage[shift_key].add(port1)
        self.p0_experiment_port_usage[shift_key].add(port2)
        print(f"    🧪 P0 (實驗) 預訂 Ports: {port1} 和 {port2} ({shift_key})", file=sys.stderr)

    def book_p0_experiment_ivek(self, day_key):
        """P0 (實驗) 預訂 IVEK"""
        if day_key not in self.p0_experiment_ivek_usage:
            self.p0_experiment_ivek_usage[day_key] = set()
            
        if "IVEK" in self.p0_experiment_ivek_usage[day_key]:
            print(f"    ⚠️ P0 (實驗) 警告: IVEK 在 {day_key} 已被其他實驗 P0 佔用", file=sys.stderr)
            
        self.p0_experiment_ivek_usage[day_key].add("IVEK")
        print(f"    🧪 P0 (實驗) 預訂 IVEK ({day_key})", file=sys.stderr)
        
    # ========================================
    # V9.9.1: P1-P5 任務檢查 (需考慮優先級)
    # ========================================

    def can_use_person(self, person, shift_key, day_key, is_single_slot, priority=None):
        """檢查人員是否可用 (P1-P5)"""
        # ========================================
        # V9.9.9: [新增] 檢查人員是否請假
        # ========================================
        try:
            date_obj = datetime.strptime(day_key, '%Y-%m-%d').date()
            if is_person_on_vacation(person, date_obj): # ✅ 修正：is_person_on_vacation 內部會處理 .lower()
                return False, f"{person} 在 {day_key} 請假"
        except Exception as e:
            print(f"              ⚠️ 無法檢查 {person} 的請假狀態: {e}", file=sys.stderr)
        
        person_lower = person.lower() # ✅ 修正：為所有比對建立小寫版本

        # 檢查非實驗 P0 衝突（所有 P1-P5 都不能衝突）
        if person_lower in self.p0_strict_person_usage.get(shift_key, set()): # ✅ 修正：使用小寫
            return False, f"{person} 被 P0 (非實驗) 佔用"
        
        # 檢查實驗 P0 衝突（僅 P3-P5 不能衝突）
        if priority is not None and priority >= 3:
            if person_lower in self.p0_experiment_person_usage.get(shift_key, set()): # ✅ 修正：使用小寫
                return False, f"{person} 被 P0 (實驗) 佔用"

        # 檢查 P1-P5 衝突
        if person_lower in self.person_usage.get(shift_key, set()): # ✅ 修正：使用小寫
            return False, f"{person} 在 {shift_key} 已被佔用"
        
        # 檢查單次限制
        if is_single_slot and person_lower in self.person_day_limit.get(day_key, set()): # ✅ 修正：使用小寫
            return False, f"{person} 在 {day_key} 已配過（單次限制）"
        
        return True, None
    
    def book_person(self, person, shift_key, day_key, is_single_slot, priority=None):
        """預訂人員 (P1-P5)"""
        can_use, reason = self.can_use_person(person, shift_key, day_key, is_single_slot, priority)
        
        if not can_use:
             raise Exception(reason)
        
        person_lower = person.lower() # ✅ 修正：使用小寫
        if shift_key not in self.person_usage:
            self.person_usage[shift_key] = set()
        self.person_usage[shift_key].add(person_lower) # ✅ 修正：使用小寫
        
        if is_single_slot:
            if day_key not in self.person_day_limit:
                self.person_day_limit[day_key] = set()
            self.person_day_limit[day_key].add(person_lower) # ✅ 修正：使用小寫
        
        print(f"        🟢 預訂人員: {person} ({shift_key})", file=sys.stderr) # Log 仍顯示原始大小寫

    def can_use_dryer(self, dryer, day_key, priority=None):
        """檢查凍乾機是否可用 (P1-P5)"""
        # ✅ 新增：檢查維修狀態（最優先檢查）
        if dryer in self.dryer_maintenance.get(day_key, set()):
            return False, f"{dryer} 在 {day_key} 維修中"
        
        # 檢查非實驗 P0 衝突（所有 P1-P5 都不能衝突）
        if dryer in self.p0_strict_dryer_usage.get(day_key, set()):
            return False, f"{dryer} 被 P0 (非實驗) 佔用"
        
        # 檢查實驗 P0 衝突（僅 P3-P5 不能衝突）
        if priority is not None and priority >= 3:
            if dryer in self.p0_experiment_dryer_usage.get(day_key, set()):
                return False, f"{dryer} 被 P0 (實驗) 佔用"

        # 檢查 P1-P5 衝突
        if dryer in self.dryer_usage.get(day_key, set()):
            return False, f"{dryer} 在 {day_key} 已被佔用"
            
        return True, None
    
    def is_dryer_available(self, dryer, day_key, priority=None):
        """
        輔助方法：檢查凍乾機是否可用 (僅回傳布林值)
        """
        can_use, _ = self.can_use_dryer(dryer, day_key, priority)
        return can_use

    def book_dryer(self, dryer, day_key, priority=None):
        """預訂凍乾機 (P1-P5)"""
        can_use, reason = self.can_use_dryer(dryer, day_key, priority)
        
        if not can_use:
            raise Exception(reason)
            
        if day_key not in self.dryer_usage:
            self.dryer_usage[day_key] = set()
        self.dryer_usage[day_key].add(dryer)
        print(f"        🟢 預訂凍乾機: {dryer} ({day_key})", file=sys.stderr)

    def mark_dryer_maintenance(self, dryer, day_key):
        """標記凍乾機在指定日期維修中"""
        if day_key not in self.dryer_maintenance:
            self.dryer_maintenance[day_key] = set()
        self.dryer_maintenance[day_key].add(dryer)
        print(f"    🔧 標記凍乾機維修: {dryer} ({day_key})", file=sys.stderr)
    
    # ========================================
    # V9.9.1: P1-P5 檢查 P0 Port/IVEK 佔用
    # ========================================

    def is_p0_strict_port_booked(self, shift_key, port_num):
        """P1-P5 檢查非實驗 P0 是否佔用此 Port"""
        return port_num in self.p0_strict_port_usage.get(shift_key, set())

    def is_p0_strict_ivek_booked(self, day_key):
        """P1-P5 檢查非實驗 P0 是否佔用 IVEK"""
        return "IVEK" in self.p0_strict_ivek_usage.get(day_key, set())

    def is_p0_experiment_port_booked(self, shift_key, port_num):
        """P3-P5 檢查實驗 P0 是否佔用此 Port"""
        return port_num in self.p0_experiment_port_usage.get(shift_key, set())

    def is_p0_experiment_ivek_booked(self, day_key):
        """P3-P5 檢查實驗 P0 是否佔用 IVEK"""
        return "IVEK" in self.p0_experiment_ivek_usage.get(day_key, set())

    def get_status(self, shift_key=None, day_key=None):
        """獲取資源使用狀態"""
        status = {}
        if shift_key:
            status['people'] = list(self.person_usage.get(shift_key, set()))
            status['people_p0_strict'] = list(self.p0_strict_person_usage.get(shift_key, set()))
            status['people_p0_experiment'] = list(self.p0_experiment_person_usage.get(shift_key, set()))
        if day_key:
            status['dryers'] = list(self.dryer_usage.get(day_key, set()))
            status['dryers_p0_strict'] = list(self.p0_strict_dryer_usage.get(day_key, set()))
            status['dryers_p0_experiment'] = list(self.p0_experiment_dryer_usage.get(day_key, set()))
            status['day_limited'] = list(self.person_day_limit.get(day_key, set()))
        return status
# --------------------------------------------------------------------
# 5. 排程器
# --------------------------------------------------------------------
class Scheduler:
    def __init__(self, data):
        self.data = data
        self.db_constraints = data["db"]
        self.schedule_df = pd.DataFrame()  
        # ✅ 修改：接收 P6/P7 任務清單
        self.task_queue, self.task_dict, self.p6_p7_fill_tasks = calculate_demand_queue(data)
        self.constraints = self.db_constraints.get("Dosing_Constraints", pd.DataFrame())
        self.beads_dry_info = self.db_constraints.get("Beads_Dry_Count_Info", pd.DataFrame())
        
        self.scheduled_pns = set() 
        self.days_to_schedule = SCHEDULING_DAYS
        self.w1_needs_met = False
        self.simulated_inventory = data["demand"].set_index('料號')['Stock_plus_Dosing'].to_dict()
        
        self.shift_port_counter = {}
        self.resources = ResourceManager()
        
        self.no_dryer_records = []
        self.p0_strict_records = []      # V9.9: 非實驗 P0
        self.p0_experiment_records = []  # V9.9: 實驗 P0
        
        # ========================================
        # ✅ V9.9.17: 儲存需求檔和 BDC（用於共用料號查詢）
        # ========================================
        self.demand_df = data["demand"]
        self.bdc_df = self.beads_dry_info  # 使用已經載入的 BDC
        
        # ✅ V9.9.17: 建立庫存查詢表 {PN: 庫存數量}
        self.stock_lookup = {}
        if not self.demand_df.empty:
            for _, row in self.demand_df.iterrows():
                pn = row.get('料號', '')
                stock = row.get('Stock_plus_Dosing', 0)
                if pn:
                    try:
                        self.stock_lookup[pn] = int(stock)
                    except:
                        self.stock_lookup[pn] = 0
        
        print(f"  ✅ 庫存查詢表建立完成 ({len(self.stock_lookup)} 筆)", file=sys.stderr)
        
        # ========================================
        # ✅ 新增：CK/PHOS/K 會 "污染" 凍乾機 3 天, GLIPA 不可使用
        # ========================================
        # "被限制的" 群組 (GLIPA)
        self.RULE_GLIPA_PNs = {'5714400220', '5714400221'} 
        # "執行限制的" 群組 (CK/PHOS/K...)
        self.RULE_BLOCKING_PNs = {
            '5714400222', '5714400199', # CK
            '5714400214', '5714400215', # PHOS (VB)
            '5714400116', '5714400226', '5714400216', '5714400117' # K1, K2, K3
        }
        self.RULE_BLOCK_DAYS = 2 # 污染 3 天 (Day 0, 1, 2)。第 3 天 (Day 3) 才可用。
        
        # 用來追蹤哪台凍乾機被污染到哪一天
        # 格式: { 'LD-5': date(2025, 11, 19), ... }
        self.dryer_block_until_date = {}

        # ✅ V9.9.15: 特殊案例追蹤
        self.na_scheduled_dates = {}      # {'Na_A': [date1, date2], 'Na_B': [...]}
        self.tcrea_scheduled = {
            'batch1': None,  # {'date': date, 'slot': 'AM'}
            'batch2': None
        }

    def _create_dummy_task(self, pn, group_name): # ✅ 修改參數
        try:
            # ✅ 修改查詢邏輯
            # 1. 根據 PN 查詢
            bdc_info_all = self.beads_dry_info[self.beads_dry_info["料號"] == pn]
            
            # 2. 根據 group_name 篩選出唯一一筆
            bdc_info_group = bdc_info_all[bdc_info_all["BDC_Group_Name"] == group_name]
            
            if bdc_info_group.empty:
                return None # 找不到對應的組合
            
            bdc_info = bdc_info_group.iloc[0]
            
            marker_name = bdc_info["BDC_Marker_Name"]
            prod_qty = bdc_info["BDC_Prod_Qty"]
            
            if pn in self.constraints.index:
                c_data = self.constraints.loc[pn]
                if isinstance(c_data, pd.DataFrame):
                    c = c_data.iloc[0]
                else:
                    c = c_data
                subp = (c.get("V9_Dryer_Count", 99), c.get("V9_Person_Count", 99), 99)
            else:
                subp = (99, 99, 99)
                
            return (99, subp, pn, 0, "NoDemand", group_name, marker_name, prod_qty)
        except Exception:
            return None
        
    def find_batch(self, current_task_info):
        """V9.9.17: 根據藥名分組，支援共用料號動態配對"""
        pn = current_task_info[2]
        group_name = current_task_info[5]
        task_priority = current_task_info[0]

        if not group_name or group_name == 'UNKNOWN':
            return [current_task_info]

        # ========================================
        # 檢查是否為特殊案例
        # ========================================
        if pn in ALL_SPECIAL_PNS:
            return self._find_special_batch(current_task_info)
        
        # ========================================
        # ✅ V9.9.17: 檢查是否為共用料號
        # ========================================
        possible_groups = self._get_possible_groups_for_pn(pn)
        
        if len(possible_groups) > 1:
            # 共用料號，選擇最佳 group_name
            group_name = self._find_best_pairing(pn, possible_groups)
            print(f"      → 使用藥名: {group_name}", file=sys.stderr)
        
        # ========================================
        # 一般案例：按藥名分組
        # ========================================
        try:
            pns_in_group = self.beads_dry_info[
                self.beads_dry_info["BDC_Group_Name"] == group_name
            ]["料號"].unique().tolist()
            
            # 移除特殊案例 PN
            pns_in_group = [p for p in pns_in_group if p not in ALL_SPECIAL_PNS]
            
        except:
            return [current_task_info]
        
        # ========================================
        # 為每個 PN 建立任務
        # ========================================
        batch_tasks = []
        for pn_in_group in pns_in_group:
            
            # 如果 PN 已排程，跳過（P5 除外）
            if (pn_in_group in self.scheduled_pns) and (task_priority != 5):
                continue 

            # 檢查 BDC_Prod_Qty
            bdc_rows = self.beads_dry_info[
                (self.beads_dry_info["料號"] == pn_in_group) &
                (self.beads_dry_info["BDC_Group_Name"] == group_name)
            ]
            
            if not bdc_rows.empty:
                prod_qty = bdc_rows.iloc[0]["BDC_Prod_Qty"]
                if prod_qty == 0:
                    continue
            else:
                continue 
            
            # 檢查是否需要分開排程
            if pn_in_group in self.constraints.index:
                constr = self.constraints.loc[pn_in_group]
                if isinstance(constr, pd.DataFrame):
                    constr = constr.iloc[0]
                
                separate_flag = str(constr.get('U,D 劑分開生產排程', '')).strip()
                if separate_flag:
                    continue
            
            # 從 task_dict 取得任務
            task_key = (group_name, pn_in_group)
            
            if task_key in self.task_dict:
                batch_tasks.append(self.task_dict[task_key])
            else:
                # 建立 dummy task
                dummy = self._create_dummy_task(pn_in_group, group_name)
                if dummy:
                    batch_tasks.append(dummy)
        
        # 去重（按 PN）
        unique_batch = list({t[2]: t for t in batch_tasks}.values())
        
        if not unique_batch:
            return []
        
        return unique_batch
    
    def _get_stock_level(self, pn):
        """
        ✅ V9.9.17: 查詢料號的庫存
        """
        return self.stock_lookup.get(pn, 0)

    def _get_possible_groups_for_pn(self, pn):
        """
        ✅ V9.9.17: 查詢一個 PN 在 BDC 中有哪些可能的藥名
        
        Returns:
            list: 例如 ['K1', 'K3']
        """
        possible_groups = []
        
        if not hasattr(self, 'bdc_df') or self.bdc_df.empty:
            return [None]
        
        # 從 BDC 查詢這個料號的所有藥名
        matching_rows = self.bdc_df[self.bdc_df['料號'] == pn]
        
        for _, row in matching_rows.iterrows():
            group_name = row.get('BDC_Group_Name', '')
            if group_name and group_name not in possible_groups:
                possible_groups.append(group_name)
        
        return possible_groups if possible_groups else [None]

    def _get_pns_for_group(self, group_name):
        """
        ✅ V9.9.17: 查詢一個藥名包含哪些料號
        
        Returns:
            list: 例如 ['5714400116', '5714400226']
        """
        pns = []
        
        if not hasattr(self, 'bdc_df') or self.bdc_df.empty:
            return pns
        
        # 從 BDC 查詢這個藥名的所有料號
        matching_rows = self.bdc_df[self.bdc_df['BDC_Group_Name'] == group_name]
        
        for _, row in matching_rows.iterrows():
            pn = row.get('料號', '')
            if pn and pn not in pns:
                pns.append(pn)
        
        return pns

    def _find_best_pairing(self, pn, possible_groups):
        """V9.9.17: 為共用料號選擇最佳配對（庫存較少的）"""
        best_group = None
        min_stock = float('inf')
        
        print(f"    🔄 共用料號 {pn} 可配對: {possible_groups}", file=sys.stderr)
        
        for group_name in possible_groups:
            group_pns = self._get_pns_for_group(group_name)
            
            if pn not in group_pns:
                continue
            
            pairing_pns = [p for p in group_pns if p != pn]
            
            if not pairing_pns:
                continue
            
            total_stock = sum(self._get_stock_level(p) for p in pairing_pns)
            
            print(f"      {group_name}: 配對料號 {pairing_pns}, 庫存={total_stock:,}", file=sys.stderr)
            
            if total_stock < min_stock:
                min_stock = total_stock
                best_group = group_name
        
        if not best_group:
            best_group = possible_groups[0]
    
        print(f"      ✅ 選擇配對: {best_group} (庫存={min_stock:,})", file=sys.stderr)
        
        # ✅ V9.9.17: 只返回 group_name，不返回 batch
        return best_group

    def _find_special_batch(self, current_task_info):
        """
        ✅ V9.9.17: 處理特殊案例批次
        """
        pn = current_task_info[2]
        
        # ========================================
        # tCREA: 3 個 PN，batch1 和 batch2 分開排
        # ========================================
        if pn in SPECIAL_CASE_PNS['tCREA']['batch1'] or pn in SPECIAL_CASE_PNS['tCREA']['batch2']:
            print(f"    🔥 特殊案例: tCREA", file=sys.stderr)
            
            # 判斷當前 PN 屬於哪個 batch
            if pn in SPECIAL_CASE_PNS['tCREA']['batch1']:
                # ✅ 返回 batch1，batch2 在 run() 中處理
                batch1_tasks = self._create_batch_from_pns(SPECIAL_CASE_PNS['tCREA']['batch1'])
                print(f"    → tCREA batch1 (同時段、同凍乾機): {[t[2] for t in batch1_tasks]}", file=sys.stderr)
                return batch1_tasks
            else:
                # batch2: 單獨排程
                batch2_tasks = self._create_batch_from_pns(SPECIAL_CASE_PNS['tCREA']['batch2'])
                print(f"    → tCREA batch2 (不同時段、不同凍乾機): {[t[2] for t in batch2_tasks]}", file=sys.stderr)
                return batch2_tasks
        
        # ========================================
        # ✅ 修正：GLIPA 應該一次返回兩個 PN（需要 2 台凍乾機）
        # ========================================
        elif pn in SPECIAL_CASE_PNS['GLIPA']['batch1'] or pn in SPECIAL_CASE_PNS['GLIPA']['batch2']:
            print(f"    🔥 特殊案例: GLIPA (需要 2 台不同凍乾機)", file=sys.stderr)
            
            # ✅ 合併兩個 batch，一次排程
            all_glipa_pns = SPECIAL_CASE_PNS['GLIPA']['batch1'] + SPECIAL_CASE_PNS['GLIPA']['batch2']
            glipa_tasks = self._create_batch_from_pns(all_glipa_pns)
            
            print(f"    → GLIPA 完整批次: {[t[2] for t in glipa_tasks]}", file=sys.stderr)
            print(f"       5714400220 (GLIPA-AD): 2 Port, 凍乾機 1", file=sys.stderr)
            print(f"       5714400221 (GLIPA-AU): 1 Port, 凍乾機 2", file=sys.stderr)
            
            return glipa_tasks
        
        # ========================================
        # Na_A / Na_B: 連續兩天，單獨處理
        # ========================================
        elif pn in SPECIAL_CASE_PNS['Na_A']:
            print(f"    🔥 特殊案例: Na_A (IVEK 連續兩天)", file=sys.stderr)
            return self._create_batch_from_pns(SPECIAL_CASE_PNS['Na_A'])
        
        elif pn in SPECIAL_CASE_PNS['Na_B']:
            print(f"    🔥 特殊案例: Na_B (IVEK 連續兩天)", file=sys.stderr)
            return self._create_batch_from_pns(SPECIAL_CASE_PNS['Na_B'])
        
        # ========================================
        # ✅ 新增：K 系列邏輯 (K1/K3)
        # ========================================
        # 狀況 A: K1 (226 觸發) -> 綁定 116
        elif pn == '5714400226':
            print(f"    🔥 特殊案例: K1 (226 帶動 116)", file=sys.stderr)
            # 強制建立 [226, 116]
            target_pns = SPECIAL_CASE_PNS['K_Series']['K1']
            batch_tasks = self._create_batch_from_pns(target_pns)
            return batch_tasks

        # 狀況 B: K3 (117 觸發) -> 綁定 116
        elif pn == '5714400117':
            print(f"    🔥 特殊案例: K3 (117 帶動 116)", file=sys.stderr)
            # 強制建立 [117, 116]
            target_pns = SPECIAL_CASE_PNS['K_Series']['K3']
            batch_tasks = self._create_batch_from_pns(target_pns)
            return batch_tasks

        # 狀況 C: 116 (理論上不應該單獨進入這裡，因為 calculate_demand_queue 已過濾)
        elif pn == '5714400116':
            print(f"    ⚠️ 警告: 發現單獨的 116 任務，忽略排程", file=sys.stderr)
            return [] # 回傳空列表，run 迴圈會跳過

        return [current_task_info]


    def _create_batch_from_pns(self, pn_list):
        """
        ✅ V9.9.15: 從 PN list 建立批次（即使需求檔沒有也建立）
        """
        batch_tasks = []
        
        for pn in pn_list:
            bdc_rows = self.beads_dry_info[self.beads_dry_info["料號"] == pn]
            
            if bdc_rows.empty:
                print(f"      ⚠️ {pn} 不在 BDC 中", file=sys.stderr)
                continue
            
            group_name = bdc_rows.iloc[0]["BDC_Group_Name"]
            task_key = (group_name, pn)
            
            if task_key in self.task_dict:
                batch_tasks.append(self.task_dict[task_key])
            else:
                dummy = self._create_dummy_task(pn, group_name)
                if dummy:
                    batch_tasks.append(dummy)
                    print(f"      ✅ 建立 dummy task: {pn} (需求檔沒有但仍需排程)", file=sys.stderr)
        
        return batch_tasks
    
    def _schedule_na_sequential(self, batch_tasks, start_day_limit, monday, day_batch_times):
        """
        ✅ V9.9.35: Na_A/Na_B 專用連續兩天排程邏輯
        規則:
        - PN1 排在 Day N (12000顆)
        - PN2 排在 Day N+1 (12000顆)
        - 兩天都必須有資源 (人員 + 凍乾機)
        """
        import sys
        from datetime import timedelta
        
        # 1. 確保批次依 PN 排序 (確保 209 在 210 前)
        batch_tasks.sort(key=lambda x: x[2]) 
        
        if len(batch_tasks) < 2:
            print(f"  ⚠️ Na Batch 任務數量不足 2 筆，無法執行連續排程", file=sys.stderr)
            return False

        task1 = batch_tasks[0] # Day 1
        task2 = batch_tasks[1] # Day 2
        
        pn1 = task1[2]
        pn2 = task2[2]
        
        print(f"  🔄 嘗試連續排程 Na 系列: Day 1 [{pn1}] -> Day 2 [{pn2}]", file=sys.stderr)
        
        # 2. 遍歷每一天作為 Day 1
        for day_offset in range(start_day_limit): 
            day1_date = monday + timedelta(days=day_offset)
            day2_date = monday + timedelta(days=day_offset + 1)
            
            # 檢查是否超過排程範圍 (若不允許週六，且 Day 2 是週六則跳過)
            if not ALLOW_EXTEND_TO_6_DAYS and day2_date.weekday() >= 5:
                continue
            
            # 檢查休假日
            if is_holiday(day1_date) or is_holiday(day2_date):
                continue
            
            # 準備 Key
            day1_key = day1_date.strftime('%Y-%m-%d')
            day2_key = day2_date.strftime('%Y-%m-%d')
            
            if day1_key not in day_batch_times: day_batch_times[day1_key] = {"AM": [], "PM": []}
            if day2_key not in day_batch_times: day_batch_times[day2_key] = {"AM": [], "PM": []}

            # --- 檢查 Day 1 (PN1) ---
            slot_name = "AM" # Na 系列通常在上午
            
            # 建立 Day 1 資源需求 (強制視為 IVEK)
            constrs1 = self._normalize_constraints([task1])
            
            # 動態時間分配
            next_time1 = self.get_delivery_time_from_constraints(constrs1, slot_name, day1_date, day_batch_times, day1_key)
            if not next_time1: continue 
            
            delivery_dt1 = datetime.combine(day1_date, next_time1)
            
            # 檢查資源 Day 1
            res1 = self.check_availability([task1], constrs1, delivery_dt1, slot_name, priority=task1[0])
            if not res1: continue
            
            # --- 檢查 Day 2 (PN2) ---
            constrs2 = self._normalize_constraints([task2])
            next_time2 = self.get_delivery_time_from_constraints(constrs2, slot_name, day2_date, day_batch_times, day2_key)
            if not next_time2: continue
            
            delivery_dt2 = datetime.combine(day2_date, next_time2)
            
            # 檢查資源 Day 2
            res2 = self.check_availability([task2], constrs2, delivery_dt2, slot_name, priority=task2[0])
            if not res2: continue
            
            # --- 兩天都成功！執行預訂 ---
            print(f"  ✅ Na 連續排程成功: {day1_key}({pn1}) -> {day2_key}({pn2})", file=sys.stderr)
            
            # 這裡有個關鍵技巧：我們修改 res 中的標記，確保 book_and_record 把它當 IVEK 寫入
            res1['num_ports_str'] = 'IVEK'
            res2['num_ports_str'] = 'IVEK'
            res1['is_ivek'] = True  # ← 新增
            res2['is_ivek'] = True 
            
            # 寫入 Day 1
            self.book_and_record(res1, [task1], day1_date, slot_name, priority=task1[0])
            self.book_resources(res1, delivery_dt1, slot_name, priority=task1[0])
            day_batch_times[day1_key][slot_name].append(next_time1)
            self.scheduled_pns.add(pn1)
            
            # 寫入 Day 2
            self.book_and_record(res2, [task2], day2_date, slot_name, priority=task2[0])
            self.book_resources(res2, delivery_dt2, slot_name, priority=task2[0])
            day_batch_times[day2_key][slot_name].append(next_time2)
            self.scheduled_pns.add(pn2)
            
            return True

        print(f"  ❌ Na 連續排程失敗: 無法找到連續兩天的空檔", file=sys.stderr)
        return False
    
    def _normalize_constraints(self, batch_tasks):
        """
        V9.9.17: 合併 batch 中所有 PN 的約束條件
        
        Returns:
            list: 包含一個合併後的 Series 的 list，如果沒有約束則返回空 list
        """
        constraints_list = []
        
        # 收集所有 PN 的約束
        for task in batch_tasks:
            pn = task[2]
            if pn in self.constraints.index:
                constr = self.constraints.loc[pn]
                if isinstance(constr, pd.DataFrame):
                    constr = constr.iloc[0]
                constraints_list.append(constr)
        
        # 如果沒有約束，檢查是否為 tCREA batch1
        if not constraints_list:
            if batch_tasks and batch_tasks[0][2] in SPECIAL_CASE_PNS['tCREA']['batch1']:
                print(f"    → tCREA batch1: 設定時段偏好為 AM", file=sys.stderr)
                return [pd.Series({
                    '時段偏好': 'AM',
                    'Port數': 2,
                    '凍乾時間_小時': 10.5
                })]
            
            return []  # ← 修改：返回空 list 而非 Series
        
        # 合併約束（使用第一個為基準）
        merged = constraints_list[0].copy()
        
        # 特殊處理：tCREA batch1 強制設為 AM
        if batch_tasks and batch_tasks[0][2] in SPECIAL_CASE_PNS['tCREA']['batch1']:
            merged['時段偏好'] = 'AM'
            print(f"    → tCREA batch1: 設定時段偏好為 AM", file=sys.stderr)
        
        # 合併其他約束（取最大值）
        if len(constraints_list) > 1:
            for constr in constraints_list[1:]:
                # 凍乾時間取最大
                if '凍乾時間_小時' in constr:
                    merged['凍乾時間_小時'] = max(
                        merged.get('凍乾時間_小時', 0),
                        constr['凍乾時間_小時']
                    )
                
                # Port數取總和（如果需要）
                # 注意：這裡通常不需要，因為 Port 數已在別處計算
        
        return [merged]  # ← 修改：返回包含 Series 的 list，而非單個 Series
    
    def get_delivery_time_from_constraints(self, constraints_list, slot_name, current_day, day_batch_times, day_key):
        """
        ✅ V9.9.16: 動態分配 RD 給藥時間，避免時段「已滿」誤判
        
        改進：
        1. 檢查實際 Port 使用情況，而非只看批次數
        2. 只有當剩餘 Port < 1 時才判定時段已滿
        3. 如果還有 Port，允許複用最後一個時間
        
        Args:
            constraints_list: 限制條件列表
            slot_name: "AM" 或 "PM"
            current_day: 當前日期
            day_batch_times: 已分配時間記錄 {day_key: {slot: [times]}}
            day_key: 日期字串 "YYYY-MM-DD"
        
        Returns:
            datetime.time 或 None
        """
        
        # 取得第一個限制來確定可用時間
        if not constraints_list:
            return None
        
        first_constr = constraints_list[0]
        if isinstance(first_constr, pd.DataFrame):
            first_constr = first_constr.iloc[0]
        
        # ========================================
        # 1. 解析可用時間
        # ========================================
        available_times_str = str(first_constr.get(
            'RD給藥時間', 
            '10:30, 11:00, 11:30, 12:00, 12:30, 15:30, 16:00, 16:30, 17:00'
        )).strip()
        
        try:
            time_slots = [t.strip() for t in available_times_str.split(',')]
        except:
            time_slots = ['10:30', '11:00', '11:30', '12:00', '12:30', 
                        '15:30', '16:00', '16:30', '17:00']
        
        # ========================================
        # 2. 根據班次過濾時間
        # ========================================
        if slot_name == "AM":
            candidate_times = [t for t in time_slots if int(t.split(':')[0]) < 15]
        else:  # PM
            candidate_times = [t for t in time_slots if int(t.split(':')[0]) >= 15]
        
        if not candidate_times:
            return None
        
        # ========================================
        # 3. 檢查已使用的時間
        # ========================================
        used_times = day_batch_times.get(day_key, {}).get(slot_name, [])
        
        # ========================================
        # 4. ✅ 檢查實際 Port 使用情況
        # ========================================
        shift_key = f"{day_key}_{slot_name}"
        used_ports_count = 0
        
        if hasattr(self, 'schedule_df') and not self.schedule_df.empty:
            shift_schedules = self.schedule_df[
                (self.schedule_df['日期'] == day_key) & 
                (self.schedule_df['班次'] == slot_name)
            ]
            
            for _, row in shift_schedules.iterrows():
                ports_str = row.get('ports_list', '')
                if ports_str:
                    try:
                        port_nums = [int(p.strip()) for p in str(ports_str).split(',') if p.strip()]
                        used_ports_count += len(port_nums)
                    except:
                        pass
        
        remaining_ports = MAX_PORTS - used_ports_count
        
        # ========================================
        # 5. ✅ 判斷時段是否已滿（基於 Port，而非批次數）
        # ========================================
        if remaining_ports < 1:
            print(f"    ⚠️ {shift_key} Port 已全部佔用 ({used_ports_count}/{MAX_PORTS})，時段已滿", file=sys.stderr)
            return None
        
        # ========================================
        # 6. ✅ 分配時間（優先未使用的時間）
        # ========================================
        # 優先使用未使用的時間
        for t_str in candidate_times:
            try:
                t = datetime.strptime(t_str, '%H:%M').time()
                if t not in used_times:
                    print(f"    ✅ {shift_key} 分配時間: {t_str} (剩餘 Port: {remaining_ports})", file=sys.stderr)
                    return t
            except:
                continue
        
        # ========================================
        # 7. ✅ 如果所有時間都用過，但還有 Port，複用最後一個時間
        # ========================================
        if candidate_times:
            last_time_str = candidate_times[-1]
            try:
                t = datetime.strptime(last_time_str, '%H:%M').time()
                print(f"    ✅ {shift_key} 複用時間: {last_time_str} (剩餘 Port: {remaining_ports})", file=sys.stderr)
                return t
            except:
                pass
        
        print(f"    ⚠️ {shift_key} 無可用時間 (Port: {used_ports_count}/{MAX_PORTS})", file=sys.stderr)
        return None
            
    def can_schedule_on_friday(self, constraints_list, current_day):
        """
        V9.9.11: 檢查批次是否可以排在週五
        
        規則：
        - 如果批次中任何 PN 的備註包含「不能放週五滴定」
        - 且 current_day 是週五（weekday == 4）
        - 則返回 False
        
        參數：
            constraints_list: 批次限制列表
            current_day: datetime 物件
            
        返回：
            (can_schedule, reason)
            - can_schedule: bool - 是否可排程
            - reason: str - 不可排程的原因（可選）
        """
        # 檢查是否為週五（0=週一, 4=週五）
        if current_day.weekday() != 4:
            return True, None  # 不是週五，可以排程
        
        # 遍歷批次中的所有 PN
        for c in constraints_list:
            remark = str(c.get("備註", "")).strip()
            pn = c.get("PN", "未知")
            
            # 檢查備註是否包含「不能放週五滴定」
            if "不能放週五滴定" in remark:
                reason = f"{pn} 備註限制：不能放週五滴定"
                print(f"        🔴 [{current_day.strftime('%Y-%m-%d')}] {reason}", file=sys.stderr)
                return False, reason
        
        # 批次中沒有週五限制
        return True, None
    
    def is_forbidden_end_time(self, end_datetime):
        """
        V9.9.12: 檢查預計結束時間是否落在禁止時段
        
        規則：
        - 禁止時段：03:00~08:00（凌晨時段）
        - 原因：避免半夜/清晨收藥
        
        參數：
            end_datetime: datetime 物件 - 預計結束時間
            
        返回：
            (is_forbidden, reason)
            - is_forbidden: bool - 是否在禁止時段
            - reason: str - 禁止原因（可選）
        """
        end_time = end_datetime.time()
    
        # 檢查是否在 03:00~08:00 之間
        if FORBIDDEN_END_TIME_START <= end_time <= FORBIDDEN_END_TIME_END:
            reason = f"預計結束時間 {end_time.strftime('%H:%M')} 落在禁止時段 (03:00~08:00)"
            print(f"        🔴 {reason}", file=sys.stderr)
            return True, reason
        
        return False, None
            
    def _calculate_ports_and_time(self, prod_qty, num_ports_str):
        """
        ✅ 修改：根據資料庫的 Port數 欄位計算
        - 移除 3-8 小時限制
        - 移除 PORT_OPTIONS = [4, 2] 限制
        """
        num_ports = 2 # 預設值
        
        try:
            if num_ports_str.upper() == "IVEK":
                # IVEK 任務, port 數量不重要 (假設為 2, 用於計算時間)
                num_ports = 2 
            else:
                num_ports = int(num_ports_str)
                if num_ports <= 0:
                    num_ports = 2
        except:
            print(f"    ⚠️ 警告: Port數欄位值 '{num_ports_str}' 無法解析, 自動設為 2", file=sys.stderr)
            num_ports = 2
        
        if prod_qty == 0:
            return num_ports, 0.5 # 預設 0.5 小時

        # ✅ 移除 3-8 小時限制，直接計算
        time_hrs = prod_qty / (DOSING_RATE_PER_HR * num_ports)
        
        return num_ports, time_hrs

    def calculate_person_priority(self, constraints_list):
        """計算批次的人員限制優先級"""
        all_people = set()
        for c in constraints_list:
            for key in ["配藥人-1", "配藥人-2", "配藥人-3"]:
                person = c.get(key)
                if person and str(person).strip():
                    all_people.add(str(person).strip().lower()) # ✅ 修正：使用小寫
        
        all_people.discard("")
        
        # ✅ 修正：還原為
        return len(all_people), list(all_people)

    def _is_dryer_safe_for_blocking_task(self, dryer_id, target_date, check_days=2):
        """
        ✅ V9.9.14: 檢查凍乾機的後 N 天是否已排程 GLIPA
        
        用於 K/PHOS/CK 排程時，確保不會污染未來的 GLIPA 任務
        
        Args:
            dryer_id: 凍乾機編號
            target_date: 當前排程日期
            check_days: 檢查未來幾天（預設 2 天）
        
        Returns:
            True: 後 N 天沒有 GLIPA，可以使用
            False: 後 N 天有 GLIPA，不可使用
        """
        dryer_str = str(dryer_id)
        
        print(f"          🔍 檢查凍乾機 {dryer_id} 後 {check_days} 天是否有 GLIPA", file=sys.stderr)
        
        # 檢查後 N 天
        for days_ahead in range(1, check_days + 1):
            future_date = target_date + timedelta(days=days_ahead)
            future_key = future_date.strftime('%Y-%m-%d')
            
            print(f"            → 檢查 {future_key} (後 {days_ahead} 天)", file=sys.stderr)
            
            # 檢查 schedule_df
            if not hasattr(self, 'schedule_df') or self.schedule_df.empty:
                print(f"            ℹ️ schedule_df 為空，跳過", file=sys.stderr)
                continue
            
            # 篩選該日期該凍乾機的記錄
            future_records = self.schedule_df[
                (self.schedule_df['日期'] == future_key) &
                (self.schedule_df['凍乾機台'] == dryer_str)
            ]
            
            if future_records.empty:
                print(f"            ✅ 凍乾機 {dryer_str} 在 {future_key} 沒有排程", file=sys.stderr)
                continue
            
            print(f"            ⚠️ 凍乾機 {dryer_str} 在 {future_key} 有 {len(future_records)} 筆排程", file=sys.stderr)
            
            # 檢查是否有 GLIPA
            for _, record in future_records.iterrows():
                pn = record.get('lot', '')
                marker = record.get('marker', '')
                
                print(f"              檢查: Marker='{marker}', PN='{pn}'", file=sys.stderr)
                
                # 檢查料號是否為 GLIPA
                if pn in self.RULE_GLIPA_PNs:
                    print(f"            ❌ 凍乾機 {dryer_id} 在 {future_key} 有 GLIPA ({pn})，不可用於 K/PHOS/CK", file=sys.stderr)
                    return False
                
                # 檢查藥名是否包含 GLIPA
                if 'GLIPA' in marker.upper():
                    print(f"            ❌ 凍乾機 {dryer_id} 在 {future_key} 有 GLIPA ({marker})，不可用於 K/PHOS/CK", file=sys.stderr)
                    return False
        
        print(f"          ✅ 凍乾機 {dryer_id} 後 {check_days} 天沒有 GLIPA，可安全使用", file=sys.stderr)
        return True

    def find_available_person_v10(self, constraints_list, delivery_dt, slot_name, priority=None):
        """V10: 簡化的人員查找"""
        # ✅ 修正：還原 find_available_person_v10 的完整邏輯
        day_key = delivery_dt.strftime('%Y-%m-%d')
        shift_key = f"{day_key}_{slot_name}"
        
        all_people = set()
        has_single_slot = False
        
        for c in constraints_list:
            for key in ["配藥人-1", "配藥人-2", "配藥人-3"]:
                person = c.get(key)
                if person and str(person).strip():
                    all_people.add(str(person).strip().lower()) # ✅ 修正：使用小寫
            
            if c.get("V9_Single_Slot", False):
                has_single_slot = True
        
        if not all_people:
            print(f"        🔴 [{day_key} {slot_name}] 批次無可配藥人員", file=sys.stderr)
            return None, False
        
        for person in sorted(all_people):
            # ========================================
            # V9.9.9: [新增] 檢查人員是否請假
            # ========================================
            if is_person_on_vacation(person, delivery_dt):
                print(f"        🔴 [{day_key}] {person} 請假，不可用", file=sys.stderr)
                continue
            can_use, reason = self.resources.can_use_person(
                person, shift_key, day_key, has_single_slot, priority
            )
            if can_use:
                print(f"        🟢 [{day_key} {slot_name}] 選中人員: {person}", file=sys.stderr)
                return person, has_single_slot
        
        print(f"        🔴 [{day_key} {slot_name}] 人員全被佔用: {list(all_people)}", file=sys.stderr)
        return None, False
    
    def calculate_dryer_priority(self, dryer_id):
        """
        ✅ V9.9.15: 計算凍乾機優先級（可用的藥越少越優先）
        """
        if dryer_id not in self.constraints.index:
            return 999  # 未知凍乾機，優先級最低
        
        # 計算有多少個 PN 可以使用這台凍乾機
        count = 0
        for idx, row in self.constraints.iterrows():
            available_dryers = str(row.get("可用凍乾機", "")).split(',')
            available_dryers = [d.strip() for d in available_dryers if d.strip()]
            
            if dryer_id in available_dryers:
                count += 1
        
        return count

    def find_available_dryer_v10_multi(self, constraints_list, delivery_dt, priority=None, batch_tasks=None, required_count=1):
        """
        ✅ V9.9.14: 尋找多台凍乾機
        - 週一/週二的 GLIPA 不受 CK/K 規則限制
        - K/PHOS/CK 排程時檢查後兩天是否有 GLIPA
        """
        day_key = delivery_dt.strftime('%Y-%m-%d')
        current_date = delivery_dt.date()
        
        # ========================================
        # ✅ 從 batch_tasks 提取 PN 列表
        # ========================================
        current_batch_pns = set()
        if batch_tasks:
            for task in batch_tasks:
                if len(task) >= 3:
                    current_batch_pns.add(task[2])  # task[2] 是 PN
        
        # 檢查是否為 GLIPA
        is_glipa_batch = bool(current_batch_pns & self.RULE_GLIPA_PNs)
        
        # 檢查是否為 K/PHOS/CK（會污染凍乾機的任務）
        is_blocking_batch = bool(current_batch_pns & self.RULE_BLOCKING_PNs)
        
        if is_blocking_batch:
            print(f"        🔍 偵測到 K/PHOS/CK 任務（{current_batch_pns & self.RULE_BLOCKING_PNs}），需檢查後兩天是否有 GLIPA", file=sys.stderr)
        
        # 檢查是否為週一/週二
        weekday = delivery_dt.weekday()
        is_monday_or_tuesday = (weekday in [0, 1])
        
        if is_glipa_batch and is_monday_or_tuesday:
            print(f"        ✅ GLIPA 排在週{'一' if weekday == 0 else '二'}，凍乾機選擇不受 CK/K 規則限制", file=sys.stderr)
        
        # 獲取所有 constraints 共同可用的凍乾機
        common_dryers = None
        for constraints in constraints_list:
            current_dryers = set(constraints.get("凍乾機台", []))
            if common_dryers is None:
                common_dryers = current_dryers
            else:
                common_dryers &= current_dryers
        
        if not common_dryers:
            print(f"        🔴 沒有共同可用的凍乾機", file=sys.stderr)
            return []
        
        # P1 任務可以例外使用被封鎖的凍乾機
        is_p1_task = (priority == 1)
        
        available_dryers = []
        
        for dryer in common_dryers:
            dryer_str = str(dryer)
            
            # 1. 檢查維修中
            if dryer in self.resources.dryer_maintenance.get(day_key, set()):
                print(f"        ⚠️ 凍乾機 {dryer} 在 {day_key} 維修中，跳過", file=sys.stderr)
                continue
            
            # ========================================
            # 2. GLIPA/CK 規則（週一/週二例外）
            # ========================================
            if is_glipa_batch and dryer in self.dryer_block_until_date:
                blocked_until = self.dryer_block_until_date[dryer]
                if current_date <= blocked_until:
                    # 週一/週二的 GLIPA 不受此限制
                    if is_monday_or_tuesday:
                        print(f"        💡 凍乾機 {dryer} 被封鎖至 {blocked_until}，但週{'一' if weekday == 0 else '二'} GLIPA 例外允許", file=sys.stderr)
                        # 不 continue，繼續使用
                    elif is_p1_task:
                        print(f"        💡 凍乾機 {dryer} 被封鎖至 {blocked_until}，但 P1 例外允許", file=sys.stderr)
                    else:
                        print(f"        ⚠️ 凍乾機 {dryer} 被 CK/K 規則封鎖至 {blocked_until}，跳過", file=sys.stderr)
                        continue  # 其他情況才跳過
            
            # ========================================
            # 3. ✅ 新增：K/PHOS/CK 檢查後兩天是否有 GLIPA
            # ========================================
            if is_blocking_batch:
                if not self._is_dryer_safe_for_blocking_task(dryer, delivery_dt):
                    print(f"        ❌ 凍乾機 {dryer} 的後兩天有 GLIPA 排程，不可用於 K/PHOS/CK", file=sys.stderr)
                    continue
            
            # 4. 檢查資源可用性
            if self.resources.is_dryer_available(dryer, day_key, priority):
                available_dryers.append(dryer)
            else:
                print(f"        ⚠️ 凍乾機 {dryer} 在 {day_key} 已被佔用（優先級 {priority}），跳過", file=sys.stderr)
        
        if not available_dryers:
            print(f"        🔴 [{day_key}] 沒有可用的凍乾機", file=sys.stderr)
            return []
        
        # 檢查是否有足夠數量的凍乾機
        if len(available_dryers) < required_count:
            print(f"        🔴 [{day_key}] 可用凍乾機不足（需要 {required_count}，可用 {len(available_dryers)}）", file=sys.stderr)
            return []
        
        # 按可用藥數優先級排序（從高到低）
        sorted_dryers = []
        for dryer in available_dryers:
            compatible_count = sum(
                1 for constraints in constraints_list
                if dryer in constraints.get("凍乾機台", [])
            )
            sorted_dryers.append((dryer, compatible_count))
        
        sorted_dryers.sort(key=lambda x: x[1], reverse=True)
        result = [dryer for dryer, _ in sorted_dryers[:required_count]]
        
        print(f"        🟢 [{day_key}] 選中凍乾機: {', '.join(map(str, result))} (按可用藥數優先級)", file=sys.stderr)
        return result


    def find_available_dryer_v10(self, constraints, delivery_dt, priority=None, batch_tasks=None):
        """
        ✅ V9.9.18: 尋找可用凍乾機
        - K/PHOS/CK 檢查後兩天是否有 GLIPA
        - 修正: 凍乾機被污染時，只阻擋 GLIPA，允許其他藥品使用
        """
        day_key = delivery_dt.strftime('%Y-%m-%d')
        current_date = delivery_dt.date()
        
        # ========================================
        # 1. 判斷當前任務屬性
        # ========================================
        current_batch_pns = set()
        if batch_tasks:
            for task in batch_tasks:
                if len(task) >= 3:
                    current_batch_pns.add(task[2])
        
        # 是否為 GLIPA (這是受害者，怕髒)
        is_current_glipa = bool(current_batch_pns & self.RULE_GLIPA_PNs)
        
        # 是否為 K/PHOS/CK (這是加害者，會弄髒機器)
        is_blocking_batch = bool(current_batch_pns & self.RULE_BLOCKING_PNs)
        
        if is_blocking_batch:
            # 如果我是髒藥，我要確保後兩天沒有 GLIPA 已經排在後面，不然我會害到它
            print(f"        🔍 偵測到 K/PHOS/CK 任務，檢查後續 GLIPA...", file=sys.stderr)

        # 獲取共同可用的凍乾機
        common_dryers = self._get_common_dryers(constraints, current_date)
        
        # P1 任務特權 (通常 P1 比較重要，有些規則可放寬，視需求而定)
        is_p1_task = (priority == 1)
        
        available_dryers = []
        
        for dryer in common_dryers:
            dryer_str = str(dryer)
            
            # ---------------------------------------------------
            # 檢查 1: 實體維修/佔用狀態 (所有人都要遵守)
            # ---------------------------------------------------
            # A. 維修中
            if dryer in self.resources.dryer_maintenance.get(day_key, set()):
                print(f"        ⚠️ 凍乾機 {dryer} 在 {day_key} 維修中，跳過", file=sys.stderr)
                continue
            
            # B. 已被其他任務佔用
            if not self.resources.is_dryer_available(dryer, day_key, priority):
                print(f"        ⚠️ 凍乾機 {dryer} 在 {day_key} 已被佔用，跳過", file=sys.stderr)
                continue

            # ---------------------------------------------------
            # 檢查 2: GLIPA 污染規則 (僅針對 GLIPA)
            # ---------------------------------------------------
            if dryer in self.dryer_block_until_date:
                blocked_until = self.dryer_block_until_date[dryer]
                
                if current_date <= blocked_until:
                    # 🔥 修正邏輯：只有當前任務是 GLIPA 才需要避開髒機器
                    if is_current_glipa:
                        if is_p1_task:
                             print(f"        💡 凍乾機 {dryer} 被封鎖，但 P1 GLIPA 例外允許", file=sys.stderr)
                        else:
                             print(f"        🚫 凍乾機 {dryer} 不乾淨 (封鎖至 {blocked_until})，GLIPA 跳過", file=sys.stderr)
                             continue
                    else:
                        # 其他藥品 (AMY, BUN...) 不怕髒，可以繼續
                        pass 

            # ---------------------------------------------------
            # 檢查 3: K/PHOS/CK 預判規則 (避免污染未來)
            # ---------------------------------------------------
            if is_blocking_batch:
                if not self._is_dryer_safe_for_blocking_task(dryer, delivery_dt):
                    print(f"        ❌ 凍乾機 {dryer} 後兩天有 GLIPA，K/PHOS/CK 不可使用", file=sys.stderr)
                    continue
            
            # 通過所有檢查
            available_dryers.append(dryer)
        
        if not available_dryers:
            print(f"        🔴 [{day_key}] 沒有可用的凍乾機", file=sys.stderr)
            return None
        
        # 排序選擇最佳凍乾機
        sorted_dryers = self._sort_dryers_by_compatibility(available_dryers, constraints)
        selected = sorted_dryers[0]
        
        print(f"        🟢 [{day_key}] 選中凍乾機: {selected}", file=sys.stderr)
        return selected
    
    def _get_common_dryers(self, constraints_list, current_date=None):
        """
        取得約束條件中共同可用的凍乾機 (取交集)
        """
        common_dryers = None

        for constr in constraints_list:
            # 嘗試獲取凍乾機欄位 (兼容 DB 的 '可用凍乾機' 和 P0 的 '凍乾機台')
            val = constr.get("可用凍乾機") or constr.get("凍乾機台") or ""
            
            current_set = set()
            
            if isinstance(val, str):
                # 解析 CSV 字串: "LD-3, LD-4" -> {'LD-3', 'LD-4'}
                current_set = {d.strip() for d in val.split(',') if d.strip()}
            elif isinstance(val, (list, set, tuple)):
                current_set = set(val)
            
            # 取交集
            if common_dryers is None:
                common_dryers = current_set
            else:
                common_dryers &= current_set
                
        if common_dryers is None:
            return []
            
        return list(common_dryers)

    def _sort_dryers_by_compatibility(self, available_dryers, constraints_list):
        """
        排序凍乾機：優先選擇「只能做較少藥品」的凍乾機 (保留彈性給其他難做的藥)
        """
        scored_dryers = []
        for dryer in available_dryers:
            # 計算這台凍乾機在整個 constraints table 中出現的次數 (越少代表越稀缺)
            # 這裡簡化邏輯：直接計算當前 batch 的兼容性，或者調用 calculate_dryer_priority
            priority_score = self.calculate_dryer_priority(dryer)
            scored_dryers.append((dryer, priority_score))
        
        # 分數越低 (能做的藥越少) 越優先 -> 升序排列
        scored_dryers.sort(key=lambda x: x[1])
        
        return [d[0] for d in scored_dryers]
    
    def _parse_quantity_with_options(self, qty_str, short_qty):
        """
        ✅ V9.9.13: 解析包含選項的數量字串（例如 "2700 or 11000"）
        
        規則：根據需求量選擇合適的數量
        - 需求量 > 9000 → 選擇較大值（11000）
        - 需求量 ≤ 9000 → 選擇較小值（2700）
        
        適用藥品：
        - HDL (5714400118, 5714400119)
        - TG-D (5714400187)
        - TG-U (5714400188)
        
        參數：
            qty_str: 數量字串，可能包含 "or"
            short_qty: 需求量（用於判斷選擇哪個值）
        
        回傳：
            float: 解析後的數量
        """
        qty_str = str(qty_str).strip()
        
        # 檢查是否包含 "or"
        if "or" in qty_str.lower():
            try:
                # 提取所有數字選項
                options = []
                for part in qty_str.split():
                    part_clean = part.strip()
                    if part_clean.lower() != "or":
                        try:
                            options.append(float(part_clean))
                        except ValueError:
                            pass
                
                if len(options) >= 2:
                    # 排序：從小到大
                    options.sort()
                    
                    # 根據需求量選擇（閾值 9000）
                    if short_qty > 9000:
                        selected = options[-1]  # 選擇較大值
                        print(f"        💡 數量選擇: '{qty_str}' → {selected} (需求 {short_qty} > 9000)", file=sys.stderr)
                    else:
                        selected = options[0]  # 選擇較小值
                        print(f"        💡 數量選擇: '{qty_str}' → {selected} (需求 {short_qty} ≤ 9000)", file=sys.stderr)
                    
                    return selected
                
                # 如果只有一個選項，回傳該選項
                elif len(options) == 1:
                    return options[0]
                    
            except Exception as e:
                print(f"        ⚠️ 數量解析失敗: '{qty_str}' - {e}", file=sys.stderr)
        
        # 一般情況：直接轉換為浮點數
        try:
            return float(qty_str)
        except (ValueError, TypeError):
            return 0
        
    def _is_dryer_clean_for_glipa(self, dryer_id, target_date, check_days=2):
        """
        ✅ V9.9.13: 檢查凍乾機前 N 天是否使用過 K/PHOS/CK 系列
        
        GLIPA 特殊規則：使用的兩台凍乾機前兩天不能用過：
        - K-* 系列（K-D, K2, K3 等所有 K 開頭）
        - PHOS-* (VB: 5714400214, HB: 5714400215)
        - CK-* (5714400222, 5714400199)
        
        參數：
            dryer_id: 凍乾機編號（字串或整數）
            target_date: 目標排程日期（datetime）
            check_days: 檢查前幾天（預設 2 天）
        
        回傳：
            bool: True 表示乾淨可用，False 表示有衝突
        """
        dryer_str = str(dryer_id)
        
        # 禁止的料號列表
        forbidden_pns = {
            '5714400214',  # PHOS (VB)
            '5714400215',  # PHOS (HB)
            '5714400222',  # CK
            '5714400199',  # CK
        }
        print(f"        🔍 檢查凍乾機 {dryer_id} (格式: '{dryer_str}') 前 {check_days} 天是否乾淨", file=sys.stderr)
    
        # 檢查前 N 天
        for days_back in range(1, check_days + 1):
            check_date = target_date - timedelta(days=days_back)
            day_key = check_date.strftime('%Y-%m-%d')
            
            # ✅ 修正：檢查 self.schedule_df（而不是 self.schedule）
            if not hasattr(self, 'schedule_df') or self.schedule_df.empty:
                continue
            
            # ========================================
            # ✅ 診斷：顯示該日期所有凍乾機的記錄
            # ========================================
            all_day_records = self.schedule_df[self.schedule_df['日期'] == day_key]
            
            if all_day_records.empty:
                print(f"          ℹ️ {day_key} 沒有任何排程記錄", file=sys.stderr)
            else:
                print(f"          📋 {day_key} 共有 {len(all_day_records)} 筆記錄：", file=sys.stderr)
                for idx, rec in all_day_records.iterrows():
                    print(f"              - 凍乾機: '{rec.get('凍乾機台', '')}', Marker: '{rec.get('marker', '')}', PN: '{rec.get('lot', '')}'", file=sys.stderr)
            # 篩選該日期該凍乾機的記錄
            day_records = self.schedule_df[
                (self.schedule_df['日期'] == day_key) &
                (self.schedule_df['凍乾機台'] == dryer_str)
            ]
            
            if day_records.empty:
                continue
            
            for _, record in day_records.iterrows():
                pn = record.get('lot', '')
                marker = record.get('marker', '')
                
                # 檢查料號是否在禁止列表
                if pn in forbidden_pns:
                    print(f"        ⚠️ 凍乾機 {dryer_id} 在 {day_key} 使用過料號 {pn}，不可用於 GLIPA", file=sys.stderr)
                    return False
                
                # 檢查藥名是否為 K 系列
                if marker.startswith('K-') or marker.startswith('K ') or marker in ['K2', 'K3', 'K1', 'K-D']:
                    print(f"        ⚠️ 凍乾機 {dryer_id} 在 {day_key} 使用過 {marker}，不可用於 GLIPA", file=sys.stderr)
                    return False
                
                # 檢查是否為 CK 或 PHOS
                if marker in ['CK', 'PHOS (VB)', 'PHOS (HB)', 'PHOS']:
                    print(f"        ⚠️ 凍乾機 {dryer_id} 在 {day_key} 使用過 {marker}，不可用於 GLIPA", file=sys.stderr)
                    return False
        
        return True
    def _is_dryer_safe_for_blocking_task(self, dryer_id, target_date, check_days=2):
        """
        ✅ V9.9.14: 檢查凍乾機的後 N 天是否已排程 GLIPA
        
        用於 K/PHOS/CK 排程時，確保不會污染未來的 GLIPA 任務
        
        Args:
            dryer_id: 凍乾機編號
            target_date: 當前排程日期
            check_days: 檢查未來幾天（預設 2 天）
        
        Returns:
            True: 後 N 天沒有 GLIPA，可以使用
            False: 後 N 天有 GLIPA，不可使用
        """
        dryer_str = str(dryer_id)
        
        print(f"          🔍 檢查凍乾機 {dryer_id} 後 {check_days} 天是否有 GLIPA", file=sys.stderr)
        
        # 檢查後 N 天
        for days_ahead in range(1, check_days + 1):
            future_date = target_date + timedelta(days=days_ahead)
            future_key = future_date.strftime('%Y-%m-%d')
            
            print(f"            → 檢查 {future_key} (後 {days_ahead} 天)", file=sys.stderr)
            
            # 檢查 schedule_df
            if not hasattr(self, 'schedule_df') or self.schedule_df.empty:
                print(f"            ℹ️ schedule_df 為空，跳過", file=sys.stderr)
                continue
            
            # 篩選該日期該凍乾機的記錄
            future_records = self.schedule_df[
                (self.schedule_df['日期'] == future_key) &
                (self.schedule_df['凍乾機台'] == dryer_str)
            ]
            
            if future_records.empty:
                print(f"            ✅ 凍乾機 {dryer_str} 在 {future_key} 沒有排程", file=sys.stderr)
                continue
            
            print(f"            ⚠️ 凍乾機 {dryer_str} 在 {future_key} 有 {len(future_records)} 筆排程", file=sys.stderr)
            
            # 檢查是否有 GLIPA
            for _, record in future_records.iterrows():
                pn = record.get('lot', '')
                marker = record.get('marker', '')
                
                print(f"              檢查: Marker='{marker}', PN='{pn}'", file=sys.stderr)
                
                # 檢查料號是否為 GLIPA
                if pn in self.RULE_GLIPA_PNs:
                    print(f"            ❌ 凍乾機 {dryer_id} 在 {future_key} 有 GLIPA ({pn})，不可用於 K/PHOS/CK", file=sys.stderr)
                    return False
                
                # 檢查藥名是否包含 GLIPA
                if 'GLIPA' in marker.upper():
                    print(f"            ❌ 凍乾機 {dryer_id} 在 {future_key} 有 GLIPA ({marker})，不可用於 K/PHOS/CK", file=sys.stderr)
                    return False
        
        print(f"          ✅ 凍乾機 {dryer_id} 後 {check_days} 天沒有 GLIPA，可安全使用", file=sys.stderr)
        return True
    def _calculate_aligned_ports(self, batch_tasks):
        """
        ✅ V9.9.32: 計算符合單雙數規則後的實際 Port 消耗量 (包含留空)
        修正: 
        1. 遇到 IVEK 不直接 return，改為設為 0 並繼續
        2. 在計算迴圈中處理 0 的情況
        """
        total_consumed = 0
        i = 0
        n = len(batch_tasks)
        
        # 1. 先取得每個 task 的 port 需求
        task_ports = []
        for task in batch_tasks:
            pn = task[2]
            p_count = 2 # 預設值
            
            if pn in self.constraints.index:
                c = self.constraints.loc[pn]
                if isinstance(c, pd.DataFrame): c = c.iloc[0]
                
                # 取得 Port 設定字串
                port_val = str(c.get("Port數", "2")).strip()
                
                # ✅ 修正 1: 如果是 IVEK，設為 0 (不要 return 0，否則會中斷整個 Batch)
                if port_val.upper() == "IVEK":
                    p_count = 0
                else:
                    try:
                        p_count = int(port_val)
                    except:
                        p_count = 2
            
            task_ports.append(p_count)

        # 2. 開始計算配對消耗
        while i < n:
            needed = task_ports[i]
            
            # ✅ 修正 2: 處理 IVEK (0 Port) 的情況
            if needed == 0:
                i += 1
                continue
            
            if needed >= 2:
                # 2 Ports: 佔用 2 個位置 (1對)
                total_consumed += 2
                i += 1
            elif needed == 1:
                # 1 Port: 檢查下一個是否也是 1 Port
                if i + 1 < n and task_ports[i+1] == 1:
                    # 找到配對 (ex. CRP-D + CRP-U) -> 兩人共用 2 個位置
                    total_consumed += 2
                    i += 2 # 跳過下一個 (因為已經配對了)
                else:
                    # 落單 (ex. GLIPA-AU 或 單獨 PN) -> 佔用 2 個位置 (1用 1空)
                    total_consumed += 2
                    i += 1
            else:
                # 異常狀況，預設 +2
                total_consumed += 2
                i += 1
                
        return total_consumed
    
    def check_availability(self, batch_tasks, constraints_list, delivery_dt, slot_name, priority=None):
        """
        檢查資源可用性 
        ✅ V9.9.13: 支援 GLIPA 轉址
        ✅ V9.9.30: 支援 Port 單雙數對齊預檢查
        """
        day_key = delivery_dt.strftime('%Y-%m-%d')
        shift_key = f"{day_key}_{slot_name}"
        
        # ========================================
        # 1. 判斷是否為 GLIPA (轉址邏輯)
        # ========================================
        current_batch_pns = set()
        if batch_tasks:
            for task in batch_tasks:
                if len(task) > 2:
                    current_batch_pns.add(task[2])

        is_glipa_batch = bool(current_batch_pns & self.RULE_GLIPA_PNs)

        if is_glipa_batch:
            # print(f"    🔀 [Redirect] 轉交 GLIPA 專用檢查: {current_batch_pns}", file=sys.stderr)
            return self._check_availability_glipa(batch_tasks, constraints_list, delivery_dt, slot_name, priority)

        # ========================================
        # 2. 一般任務 Port 容量檢查 (含對齊規則)
        # ========================================
        
        # 從約束中讀取 Port數 (判斷是否為 IVEK)
        num_ports_str = "2"
        if constraints_list:
            num_ports_str = str(constraints_list[0].get("Port數", "2")).strip()
        
        is_ivek_task = (num_ports_str.upper() == "IVEK")

        # 如果不是 IVEK，必須檢查 Port 1-12 是否夠用
        if not is_ivek_task:
            # 取得當前班次的 Port 指標
            if shift_key not in self.shift_port_counter:
                self.shift_port_counter[shift_key] = 1
            
            port_start = self.shift_port_counter[shift_key]

            # 模擬 P0 避讓與單數對齊 (不會真的修改 counter，只是預判)
            temp_start = port_start
            while True:
                is_booked = False
                # 檢查 P0 佔用
                if self.resources.is_p0_strict_port_booked(shift_key, temp_start): is_booked = True
                if priority is not None and priority >= 3 and self.resources.is_p0_experiment_port_booked(shift_key, temp_start): is_booked = True
                
                # 強制單數對齊 (Odd Alignment)
                if temp_start % 2 == 0:
                    temp_start += 1
                    continue

                if is_booked:
                    temp_start += 1
                else:
                    break
            
            # 🔥 這裡加入 total_ports_needed 的計算 🔥
            total_ports_needed = self._calculate_aligned_ports(batch_tasks)
            
            # 檢查是否超過 MAX_PORTS (12)
            if temp_start + total_ports_needed - 1 > MAX_PORTS:
                print(f"    ⚠️ {shift_key} Port 空間不足 (需 {total_ports_needed}, 起始 {temp_start})", file=sys.stderr)
                return None

        # ========================================
        # 3. IVEK 檢查
        # ========================================
        if is_ivek_task and self.resources.is_p0_strict_ivek_booked(day_key):
            print(f"    🔴 {day_key} IVEK 已被 P0 (非實驗) 佔用", file=sys.stderr)
            return None
        
        if priority is not None and priority >= 3:
            if is_ivek_task and self.resources.is_p0_experiment_ivek_booked(day_key):
                print(f"    🔴 {day_key} IVEK 已被 P0 (實驗) 佔用", file=sys.stderr)
                return None
        
        # ========================================
        # 4. 人員與凍乾機檢查 (原有邏輯)
        # ========================================
        person_result = self.find_available_person_v10(constraints_list, delivery_dt, slot_name, priority)
        if not person_result or person_result[0] is None:
            return None
        
        person, is_single_slot = person_result

        dryer_id = self.find_available_dryer_v10(constraints_list, delivery_dt, priority, batch_tasks)
        if not dryer_id:
            return None

        # ========================================
        # 5. 計算時間與回傳
        # ========================================
        try:
            dosing_start_dt = delivery_dt + timedelta(minutes=DOSING_PREP_TIME_MIN)
            
            total_short_qty = sum(t[3] for t in batch_tasks)
            qtys = []
            for c in constraints_list:
                qty_str = c.get("數量", 0)
                qty = self._parse_quantity_with_options(qty_str, total_short_qty)
                if qty > 0: qtys.append(qty)
            
            max_qty = max(qtys) if qtys else 1000
            
            # 計算 Port 數和時間
            num_ports, dosing_hrs = self._calculate_ports_and_time(max_qty, num_ports_str)
            
            dosing_end_dt = add_hours_to_time(dosing_start_dt, dosing_hrs)
            freeze_duration = float(max(c.get("凍乾時間", 12) for c in constraints_list))
            final_end_dt = add_hours_to_time(dosing_end_dt, freeze_duration)
        
        except Exception as e:
            print(f"        🔴 時間計算失敗: {e}", file=sys.stderr)
            return None

        return {
            "person": person,
            "dryer": dryer_id, 
            "is_single_slot": is_single_slot,
            "times": {
                "delivery_dt": delivery_dt,
                "dosing_start_dt": dosing_start_dt,
                "dosing_end_dt": dosing_end_dt,
                "final_end_dt": final_end_dt,
                "freeze_duration_hr": freeze_duration,
                "dosing_hrs": dosing_hrs
            },
            "num_ports_per_pn": num_ports,
            "num_ports_str": num_ports_str,
            "constraints_list": constraints_list,
            "is_glipa": False # 標記非 GLIPA
        }


    def _check_availability_glipa(self, batch_tasks, constraints_list, delivery_dt, slot_name, priority=None):
        """
        ✅ GLIPA 專用資源檢查（確保兩台不同凍乾機 + Port 空間預判）
        """
        import sys
        from datetime import timedelta
        
        day_key = delivery_dt.strftime('%Y-%m-%d')
        
        print(f"    🔥 [DEBUG] GLIPA 開始檢查資源", file=sys.stderr)
        
        # 1. 檢查人員
        person_result = self.find_available_person_v10(constraints_list, delivery_dt, slot_name, priority)
        if not person_result or person_result[0] is None:
            print(f"    🔴 GLIPA: 找不到可用人員", file=sys.stderr)
            return None
        
        person, is_single_slot = person_result
        print(f"    ✅ [DEBUG] GLIPA 選定人員: {person}", file=sys.stderr)
        
        # 2. 分別為兩個 PN 找凍乾機
        print(f"    🔍 [DEBUG] 開始為 GLIPA-AD 尋找凍乾機...", file=sys.stderr)
        
        dryer1 = None
        if len(batch_tasks) > 0 and len(constraints_list) > 0:
            dryer1 = self.find_available_dryer_v10(
                [constraints_list[0]],
                delivery_dt, 
                priority, 
                [batch_tasks[0]]
            )
        
        if not dryer1:
            print(f"    🔴 GLIPA: 找不到第一台凍乾機 (GLIPA-AD)", file=sys.stderr)
            return None
        
        if not self._is_dryer_clean_for_glipa(dryer1, delivery_dt):
            print(f"    🔴 GLIPA: 凍乾機 {dryer1} 前兩天不乾淨", file=sys.stderr)
            return None
        if dryer1 in self.dryer_block_until_date:
             blocked_until = self.dryer_block_until_date[dryer1]
             if delivery_dt.date() <= blocked_until:
                 print(f"    🔴 GLIPA: 凍乾機 {dryer1} 當週剛被污染 (至 {blocked_until})", file=sys.stderr)
                 return None
        
        print(f"    ✅ [DEBUG] GLIPA-AD 選定凍乾機: {dryer1}", file=sys.stderr)
        
        # 2.2 為第二個 PN 找凍乾機（必須與 dryer1 不同）
        print(f"    🔍 [DEBUG] 開始為 GLIPA-AU 尋找凍乾機（需與 {dryer1} 不同）...", file=sys.stderr)
        
        dryer2 = None
        # 取得約束條件
        if len(constraints_list) > 1:
            constr2 = constraints_list[1]
        elif len(constraints_list) == 1:
            constr2 = constraints_list[0]
        else:
            return None
        
        available_dryers_str = str(constr2.get("可用凍乾機", "")).strip()
        candidate_dryers = [d.strip() for d in available_dryers_str.split(',') if d.strip()]
        
        # 建議：可以加入排序邏輯，優先選比較難排的機器 (可選)
        # candidate_dryers.sort(key=lambda x: self.calculate_dryer_priority(x))

        for dryer in candidate_dryers:
            # 規則 1: 必須與第一台不同
            if dryer == dryer1: 
                continue
            
            # 規則 2: 資源必須可用 (未被佔用、未維修)
            if not self.resources.is_dryer_available(dryer, day_key, priority): 
                continue
            
            # 規則 3: 歷史紀錄必須乾淨 (檢查 schedule_df)
            if not self._is_dryer_clean_for_glipa(dryer, delivery_dt): 
                continue
            
            # 規則 4: [重要] 當週必須乾淨 (檢查 dryer_block_until_date)
            # GLIPA 是受害者，必須嚴格避開 CK/K/PHOS 污染
            if dryer in self.dryer_block_until_date:
                 blocked_until = self.dryer_block_until_date[dryer]
                 if delivery_dt.date() <= blocked_until:
                     print(f"    🚫 GLIPA: 凍乾機 {dryer} 當週剛被污染 (封鎖至 {blocked_until})", file=sys.stderr)
                     continue

            # 通過所有檢查
            dryer2 = dryer
            print(f"    ✅ [DEBUG] GLIPA-AU 選定凍乾機: {dryer2}", file=sys.stderr)
            break
        
        if not dryer2:
            print(f"    🔴 GLIPA: 找不到第二台合適的凍乾機", file=sys.stderr)
            return None
        
        # 3. 計算 Port 需求 (含對齊規則)
        # 🔥 這裡補上了 num_ports 的定義邏輯 🔥
        pn_port_map = {}
        total_ports_needed = self._calculate_aligned_ports(batch_tasks)
        
        print(f"    📊 [DEBUG] GLIPA Port 配置 (Total Need: {total_ports_needed}):", file=sys.stderr)
        
        for task in batch_tasks:
            pn = task[2]
            
            # === 讀取 Port 數邏輯開始 ===
            num_ports = 2
            if pn in self.constraints.index:
                c = self.constraints.loc[pn]
                if isinstance(c, pd.DataFrame): c = c.iloc[0]
                try:
                    num_ports = int(str(c.get("Port數", "2")).strip())
                except:
                    num_ports = 2
            # === 讀取 Port 數邏輯結束 ===
            
            pn_port_map[pn] = num_ports
            print(f"       {pn}: {num_ports} Port", file=sys.stderr)
            
        # 4. Port 容量預檢查
        shift_key = f"{day_key}_{slot_name}"
        if shift_key not in self.shift_port_counter:
            self.shift_port_counter[shift_key] = 1
        
        port_start = self.shift_port_counter[shift_key]
        
        # 模擬 P0 避讓與單數對齊
        temp_start = port_start
        while True:
            is_booked = False
            if self.resources.is_p0_strict_port_booked(shift_key, temp_start): is_booked = True
            if priority is not None and priority >= 3 and self.resources.is_p0_experiment_port_booked(shift_key, temp_start): is_booked = True
            
            if temp_start % 2 == 0:
                temp_start += 1
                continue

            if is_booked:
                temp_start += 1
            else:
                break
        
        if temp_start + total_ports_needed - 1 > MAX_PORTS:
             print(f"    🔴 GLIPA: Port 空間不足 (需 {total_ports_needed}, 起始 {temp_start})", file=sys.stderr)
             return None

        # 5. 計算時間
        try:
            dosing_start_dt = delivery_dt + timedelta(minutes=DOSING_PREP_TIME_MIN)
            total_short_qty = sum(t[3] for t in batch_tasks)
            qtys = []
            for c in constraints_list:
                qty_str = c.get("數量", 0)
                qty = self._parse_quantity_with_options(qty_str, total_short_qty)
                if qty > 0: qtys.append(qty)
            
            max_qty = max(qtys) if qtys else 2500
            # GLIPA 通常是 2 port 效率
            dosing_hrs = max_qty / (DOSING_RATE_PER_HR * 2)
            
            dosing_end_dt = add_hours_to_time(dosing_start_dt, dosing_hrs)
            freeze_duration = float(max(c.get("凍乾時間", 12) for c in constraints_list))
            final_end_dt = add_hours_to_time(dosing_end_dt, freeze_duration)
        except Exception as e:
            print(f"    🔴 GLIPA 時間計算失敗: {e}", file=sys.stderr)
            return None
        
        print(f"    ✅ [DEBUG] GLIPA 資源檢查完成！", file=sys.stderr)
        
        return {
            "person": person,
            "dryer": dryer1,
            "dryer2": dryer2,
            "is_single_slot": is_single_slot,
            "times": {
                "delivery_dt": delivery_dt,
                "dosing_start_dt": dosing_start_dt,
                "dosing_end_dt": dosing_end_dt,
                "final_end_dt": final_end_dt,
                "freeze_duration_hr": freeze_duration,
                "dosing_hrs": dosing_hrs
            },
            "pn_port_map": pn_port_map,
            "total_ports_needed": total_ports_needed,
            "num_ports_str": str(total_ports_needed),
            'constraints_list': constraints_list,
            "is_glipa": True
        }

    def book_resources(self, resources, delivery_dt, slot_name, priority=None):
        """V9.9.13: 統一使用 V10 資源預訂（支援 GLIPA 雙凍乾機）"""
        person = resources["person"]
        dryer = resources["dryer"]
        is_single_slot = resources["is_single_slot"]
        is_glipa = resources.get("is_glipa", False)  # ✅ 新增

        day_key = delivery_dt.strftime('%Y-%m-%d')
        shift_key = f"{day_key}_{slot_name}"

        # 預訂人員
        try:
            self.resources.book_person(person, shift_key, day_key, is_single_slot, priority)
        except Exception as e:
            print(f"        ⚠️ 警告: 預訂人員 {person} 失敗: {e}", file=sys.stderr)

        # 預訂第一台凍乾機
        try:
            self.resources.book_dryer(dryer, day_key, priority)
        except Exception as e:
            print(f"        ⚠️ 警告: 預訂凍乾機 {dryer} 失敗: {e}", file=sys.stderr)
    
        # ✅ 新增：GLIPA 需要預訂第二台凍乾機
        if is_glipa:
            dryer2 = resources.get("dryer2")
            if dryer2:
                try:
                    self.resources.book_dryer(dryer2, day_key, priority)
                    print(f"        🟢 預訂第二台凍乾機: {dryer2} ({day_key})", file=sys.stderr)
                except Exception as e:
                    print(f"        ⚠️ 警告: 預訂第二台凍乾機 {dryer2} 失敗: {e}", file=sys.stderr)
        
    def handle_p0_orders(self):
        """
        V9.9.5: 處理 P0 (插單/限制) 任務 - 增強診斷版
        """
        print("\n" + "="*70, file=sys.stderr)
        print(f"🔥 Step 3.0: 正在處理 P0 (插單/限制) 任務 ({CURRENT_VERSION})", file=sys.stderr)
        print("="*70, file=sys.stderr)
        
        df_p0_orders = self.db_constraints.get("P0_Orders", pd.DataFrame())
        
        if df_p0_orders.empty:
            print("  ⚠️ P0_Orders 表格為空", file=sys.stderr)
            return

        # 🔥 診斷 1：顯示讀取到的 P0 資料
        print(f"  📊 從資料庫讀取到 {len(df_p0_orders)} 筆 P0 任務", file=sys.stderr)
        print(f"  📋 P0 欄位: {df_p0_orders.columns.tolist()}", file=sys.stderr)
        
        # 🔥 修正：使用全域的排程起始日期而非今天
        global SCHEDULE_START_DATE
        
        if SCHEDULE_START_DATE:
            monday = SCHEDULE_START_DATE
            print(f"  📅 使用排程起始日期: {monday.strftime('%Y-%m-%d')}", file=sys.stderr)
        else:
            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            monday = (today + timedelta(days=-today.weekday(), weeks=0))
            print(f"  📅 使用今天推算的週一: {monday.strftime('%Y-%m-%d')}", file=sys.stderr)
        
        saturday = monday + timedelta(days=5)
        current_year = monday.year
        
        print(f"  🗓️  排程週範圍: {monday.strftime('%Y-%m-%d')} ~ {saturday.strftime('%Y-%m-%d')}", file=sys.stderr)
        
        # ========================================
        # ✅ 新增：處理凍乾機維修記錄
        # ========================================
        print("\n" + "="*70, file=sys.stderr)
        print("🔧 Step 3.0a: 處理凍乾機維修/故障/保養記錄", file=sys.stderr)
        print("="*70, file=sys.stderr)

        maintenance_count = 0

        for idx, row in df_p0_orders.iterrows():
            titrate = str(row.get("滴定機", "")).strip()
            marker = str(row.get("Marker", "")).strip()
            date_str = str(row.get("日期", "")).strip()
            dryer_str = str(row.get("凍乾機台", "")).strip()
            remark = str(row.get("備註", "")).strip()
            
            # 檢查條件
            if titrate != "" or marker != "":
                continue
            if not date_str:
                continue
            if not dryer_str:
                continue
            if not any(keyword in remark for keyword in ["維修", "故障", "保養"]):
                continue
            
            # 解析日期
            try:

                current_year = SCHEDULE_START_DATE.year if SCHEDULE_START_DATE else datetime.now().year
                
                date_str_clean = date_str.replace(' ', '').strip()
                task_date = None
                
                # 嘗試解析「11月24日」格式
                if re.search(r'(\d{1,2})月(\d{1,2})日', date_str_clean):
                    month = int(re.search(r'(\d{1,2})月', date_str_clean).group(1))
                    day = int(re.search(r'(\d{1,2})日', date_str_clean).group(1))
                    task_date = datetime(current_year, month, day)
                else:
                    # 其他格式
                    task_date = pd.to_datetime(date_str_clean, errors='coerce')
                    if pd.notna(task_date):
                        task_date = task_date.to_pydatetime()
                
                if task_date is None:
                    continue
                
                day_key = task_date.strftime('%Y-%m-%d')
                
                # 解析凍乾機編號
                dryer_num_match = re.search(r'\d+', dryer_str)
                if dryer_num_match:
                    dryer_num = int(dryer_num_match.group())
                    if 3 <= dryer_num <= 20:
                        dryer_name = str(dryer_num)
                        self.resources.mark_dryer_maintenance(dryer_name, day_key)
                        maintenance_count += 1
                        print(f"  ✅ {day_key}: {dryer_name} - {remark}", file=sys.stderr)
                
            except Exception as e:
                print(f"  ⚠️ 維修記錄解析失敗: {e}", file=sys.stderr)

        if maintenance_count > 0:
            print(f"\n  ✅ 共標記 {maintenance_count} 筆凍乾機維修記錄\n", file=sys.stderr)
        else:
            print(f"  ℹ️ 無凍乾機維修記錄\n", file=sys.stderr)
        
        p0_tasks_this_week = []
        p0_on_holiday = []
        p0_parse_failed = []
        p0_out_of_range = []
        
        # 🔥 診斷 2：逐筆處理並記錄每個任務的狀態
        for idx, row in df_p0_orders.iterrows():
            date_str = str(row.get("日期", "")).strip()
            marker = str(row.get("Marker", "")).strip()
            lot = str(row.get("Lot", "N/A")).strip()
            
            print(f"\n  ━━━ P0 任務 #{idx+1} ━━━", file=sys.stderr)
            print(f"      原始日期: '{date_str}'", file=sys.stderr)
            print(f"      Marker: '{marker}'", file=sys.stderr)
            print(f"      Lot: '{lot}'", file=sys.stderr)
            
            # 檢查必要欄位
            if not date_str or not marker:
                print(f"      ❌ 失敗: 缺少 {'日期' if not date_str else 'Marker'}", file=sys.stderr)
                continue
            
            # 🔥 修正：更強健的日期解析
            try:
                task_date = None
                date_str_clean = date_str.replace(' ', '').strip()
                
                # 嘗試多種日期格式
                parse_methods = [
                    # 格式1: YYYY-MM-DD 或 YYYY/MM/DD
                    lambda s: pd.to_datetime(s, format='%Y-%m-%d', errors='coerce'),
                    lambda s: pd.to_datetime(s, format='%Y/%m/%d', errors='coerce'),
                    
                    # 格式2: M月D日
                    lambda s: datetime(current_year, 
                                    int(re.search(r'(\d{1,2})月', s).group(1)),
                                    int(re.search(r'(\d{1,2})日', s).group(1))) 
                            if re.search(r'(\d{1,2})月(\d{1,2})日', s) else None,
                    
                    # 格式3: pandas 自動解析
                    lambda s: pd.to_datetime(s, errors='coerce')
                ]
                
                for method in parse_methods:
                    try:
                        result = method(date_str_clean)
                        if result is not None and not pd.isna(result):
                            task_date = result if isinstance(result, datetime) else result.to_pydatetime()
                            print(f"      ✅ 日期解析成功: {task_date.strftime('%Y-%m-%d')}", file=sys.stderr)
                            break
                    except:
                        continue
                
                if task_date is None:
                    raise ValueError(f"所有解析方法均失敗")
                
                # 🔥 診斷 3：檢查日期是否在範圍內
                if monday <= task_date <= saturday:
                    print(f"      ✅ 日期在本週範圍內", file=sys.stderr)
                    
                    # 檢查是否為休假日
                    if is_holiday(task_date):
                        p0_on_holiday.append({
                            'lot': lot,
                            'date': task_date.strftime('%Y-%m-%d'),
                            'date_obj': task_date
                        })
                        print(f"      ⚠️ 日期為休假日", file=sys.stderr)
                    else:
                        p0_tasks_this_week.append((task_date, row))
                        print(f"      ✅ 已加入本週任務列表", file=sys.stderr)
                else:
                    p0_out_of_range.append({
                        'lot': lot,
                        'date': task_date.strftime('%Y-%m-%d'),
                        'week_range': f"{monday.strftime('%Y-%m-%d')} ~ {saturday.strftime('%Y-%m-%d')}"
                    })
                    print(f"      ⚠️ 日期不在本週範圍（{task_date.strftime('%Y-%m-%d')}）", file=sys.stderr)
                        
            except Exception as e:
                p0_parse_failed.append({
                    'lot': lot,
                    'date_str': date_str,
                    'error': str(e)
                })
                print(f"      ❌ 日期解析失敗: {e}", file=sys.stderr)
                continue
        
        # 🔥 診斷 4：統計報告
        print(f"\n" + "="*70, file=sys.stderr)
        print(f"📊 P0 任務處理統計", file=sys.stderr)
        print(f"="*70, file=sys.stderr)
        print(f"  總計讀取: {len(df_p0_orders)} 筆", file=sys.stderr)
        print(f"  本週範圍內: {len(p0_tasks_this_week)} 筆", file=sys.stderr)
        print(f"  休假日衝突: {len(p0_on_holiday)} 筆", file=sys.stderr)
        print(f"  日期超出範圍: {len(p0_out_of_range)} 筆", file=sys.stderr)
        print(f"  日期解析失敗: {len(p0_parse_failed)} 筆", file=sys.stderr)
        
        # 顯示超出範圍的任務
        if p0_out_of_range:
            print(f"\n  ⚠️ 以下 P0 任務不在本週範圍：", file=sys.stderr)
            for item in p0_out_of_range:
                print(f"      • {item['lot']} ({item['date']}) - 本週範圍: {item['week_range']}", file=sys.stderr)
        
        # 顯示解析失敗的任務
        if p0_parse_failed:
            print(f"\n  ❌ 以下 P0 任務日期解析失敗：", file=sys.stderr)
            for item in p0_parse_failed:
                print(f"      • {item['lot']}: '{item['date_str']}' - {item['error']}", file=sys.stderr)
        
        print(f"="*70, file=sys.stderr)
        
        # 檢查休假日衝突
        if p0_on_holiday:
            print("\n" + "="*70, file=sys.stderr)
            print("❌ 排程終止：P0 任務日期衝突", file=sys.stderr)
            print("="*70, file=sys.stderr)
            print(f"以下 {len(p0_on_holiday)} 個 P0 任務的排程日期為休假日：", file=sys.stderr)
            for p0_info in p0_on_holiday:
                print(f"  • {p0_info['lot']} ({p0_info['date']})", file=sys.stderr)
            raise SystemExit("P0 任務與休假日衝突，排程已終止")
                
        if not p0_tasks_this_week:
            print(f"\n  ℹ️ 無本週 P0 任務需要處理", file=sys.stderr)
            return

        print(f"\n  🔄 開始處理 {len(p0_tasks_this_week)} 筆本週 P0 任務...", file=sys.stderr)
        
        # 處理每個 P0 任務
        processed_count = 0
        failed_count = 0
        
        for task_date, row in p0_tasks_this_week:
            day_key = task_date.strftime('%Y-%m-%d')
            lot = str(row.get("Lot", "N/A")).strip()
            staff = str(row.get("配藥同仁", "")).strip()
            dryer = str(row.get("凍乾機台", "")).strip()
            machine = str(row.get("滴定機", "")).strip().lower()
            rd_time_str = str(row.get("RD給藥時間", "09:00")).strip()
            remark = str(row.get("備註", "")).strip()
            
            print(f"\n  ━━━ 處理 P0: {lot} ({day_key}) ━━━", file=sys.stderr)
            
            # 檢查配藥人員
            if not staff:
                print(f"      ❌ 失敗: 未指定配藥同仁", file=sys.stderr)
                failed_count += 1
                continue
            
            # 檢查配藥人員是否請假
            if is_person_on_vacation(staff, task_date):
                print(f"      ❌ 失敗: 配藥人員 {staff} 請假", file=sys.stderr)
                failed_count += 1
                continue
            
            # 計算班次
            try:
                rd_hour = int(rd_time_str.split(':')[0])
                slot_name = "PM" if rd_hour >= 15 else "AM"
            except:
                slot_name = "AM"
                
            shift_key = f"{day_key}_{slot_name}"
            
            print(f"      配藥同仁: {staff}", file=sys.stderr)
            print(f"      班次: {slot_name}", file=sys.stderr)
            print(f"      滴定機: {machine}", file=sys.stderr)
            print(f"      凍乾機: {dryer}", file=sys.stderr)

            # 檢查是否無滴定機或凍乾機 → no_dryer
            if pd.isna(dryer) or dryer == "" or pd.isna(machine) or machine == "":
                print(f"      → 無滴定機或凍乾機，轉 no_dryer 處理", file=sys.stderr)
                
                constr_dict = {"配藥人-1": staff, "V9_Single_Slot": False}
                person_result = self._find_person_for_no_dryer(constr_dict, task_date)
                
                if person_result is None:
                    print(f"      ❌ 失敗: 無法預訂全天資源", file=sys.stderr)
                    failed_count += 1
                    continue
                
                person, _ = person_result
                
                self.no_dryer_records.append({
                    "日期": day_key,
                    "marker": row.get("Marker", ""),
                    "lot": lot,
                    "凍乾機台": "",
                    "數量": row.get("數量", 0),
                    "配藥同仁": person,
                    "RD給藥時間": rd_time_str,
                    "預計滴定時間": row.get("預計滴定時間", ""),
                    "預計結束": row.get("預計結束", ""),
                    "班次": slot_name,
                    "滴定機": machine,
                    "is_no_dryer": True,
                    "PN": row.get("PN", ""),
                    "備註": remark
                })
                self.scheduled_pns.add(lot)
                processed_count += 1
                print(f"      ✅ 成功: 已排入 no_dryer", file=sys.stderr)
                continue

            # 檢查是否為實驗任務
            is_experiment = "*實驗*" in remark
            
            if is_experiment:
                print(f"      → 實驗任務，將在 P2 後排程", file=sys.stderr)
                p0_data = row.to_dict()
                p0_data['P0_Parsed_Date'] = task_date
                p0_data['is_experiment'] = True
                self.p0_experiment_records.append(p0_data)
                self.scheduled_pns.add(lot)
                processed_count += 1
                print(f"      ✅ 成功: 已加入實驗 P0 列表", file=sys.stderr)
            else:
                print(f"      → 非實驗任務，優先排程", file=sys.stderr)
                
                try:
                    self.resources.book_p0_strict_person(staff, shift_key, day_key)
                    self.resources.book_p0_strict_dryer(dryer, day_key)
                    
                    if "ivek" in machine:
                        self.resources.book_p0_strict_ivek(day_key)
                        print(f"      → 已預訂 IVEK", file=sys.stderr)
                    elif "port" in machine:
                        try:
                            port_num = int(re.findall(r'\d+', machine)[0])
                            if 1 <= port_num <= 12:
                                self.resources.book_p0_strict_port_pair(shift_key, port_num)
                                print(f"      → 已預訂 Port{port_num}", file=sys.stderr)
                        except Exception as e:
                            print(f"      ⚠️ 警告: 無法解析滴定機編號: {e}", file=sys.stderr)
                    
                    p0_data = row.to_dict()
                    p0_data['P0_Parsed_Date'] = task_date
                    p0_data['is_experiment'] = False
                    self.p0_strict_records.append(p0_data)
                    
                    self.scheduled_pns.add(lot)
                    processed_count += 1
                    print(f"      ✅ 成功: 已強制排入非實驗 P0", file=sys.stderr)
                    
                except Exception as e:
                    print(f"      ❌ 失敗: {e}", file=sys.stderr)
                    traceback.print_exc(file=sys.stderr)
                    failed_count += 1

        # 最終統計
        print(f"\n" + "="*70, file=sys.stderr)
        print(f"✅ P0 任務處理完成", file=sys.stderr)
        print(f"="*70, file=sys.stderr)
        print(f"  成功處理: {processed_count} 筆", file=sys.stderr)
        print(f"  處理失敗: {failed_count} 筆", file=sys.stderr)
        print(f"  非實驗 P0: {len(self.p0_strict_records)} 筆", file=sys.stderr)
        print(f"  實驗 P0: {len(self.p0_experiment_records)} 筆", file=sys.stderr)
        print(f"  no_dryer: {len([r for r in self.no_dryer_records if r.get('is_no_dryer')])} 筆", file=sys.stderr)
        print(f"="*70 + "\n", file=sys.stderr)
        
    def schedule_p0_experiments(self, all_slots):
        """
        V9.9.1: 排程實驗 P0 (P2 後，P1/P2 允許衝突，P3-P5 不能衝突)
        """
        
        if not self.p0_experiment_records:
            print("  ... 無實驗 P0 任務需要排程", file=sys.stderr)
            return
        
        print(f"  開始排程 {len(self.p0_experiment_records)} 個實驗 P0 任務...", file=sys.stderr)
        
        for idx, p0_data in enumerate(self.p0_experiment_records, 1):
            # 步驟 1: 解析 P0 任務資料
            task_date = p0_data.get('P0_Parsed_Date')
            
            if task_date is None:
                print(f"  ❌ 實驗 P0 任務 #{idx}: 缺少日期資訊，跳過", file=sys.stderr)
                continue
            
            lot = str(p0_data.get("Lot", "N/A")).strip()
            marker = str(p0_data.get("Marker", "")).strip()
            staff = str(p0_data.get("配藥同仁", "")).strip()
            dryer = str(p0_data.get("凍乾機台", "")).strip()
            machine = str(p0_data.get("滴定機", "")).strip().lower()
            rd_time_str = str(p0_data.get("RD給藥時間", "09:00")).strip()
            remark = str(p0_data.get("備註", "")).strip()
            
            # 步驟 2: 驗證必要欄位
            if not staff:
                print(f"  ❌ 實驗 P0 任務 {lot}: 缺少配藥同仁，跳過", file=sys.stderr)
                continue
            
            if not dryer:
                print(f"  ❌ 實驗 P0 任務 {lot}: 缺少凍乾機台，跳過", file=sys.stderr)
                continue
            
            if not machine:
                print(f"  ❌ 實驗 P0 任務 {lot}: 缺少滴定機，跳過", file=sys.stderr)
                continue
            
            # 步驟 3: 計算班次
            try:
                rd_hour = int(rd_time_str.split(':')[0])
                slot_name = "PM" if rd_hour >= 15 else "AM"
            except Exception as e:
                print(f"  ⚠️ 實驗 P0 任務 {lot}: RD給藥時間解析失敗，預設為 AM", file=sys.stderr)
                slot_name = "AM"
            
            day_key = task_date.strftime('%Y-%m-%d')
            shift_key = f"{day_key}_{slot_name}"
            
            print(f"\n  🧪 [{idx}/{len(self.p0_experiment_records)}] 處理實驗 P0: {lot}", file=sys.stderr)
            print(f"      日期: {day_key} {slot_name}", file=sys.stderr)
            print(f"      人員: {staff}, 凍乾機: {dryer}, 滴定機: {machine}", file=sys.stderr)
            
            # 步驟 4: 預訂實驗 P0 資源
            try:
                # 4.1 預訂配藥人員
                self.resources.book_p0_experiment_person(staff, shift_key)
                
                # 4.2 預訂凍乾機
                self.resources.book_p0_experiment_dryer(dryer, day_key)
                
                # 4.3 預訂滴定機 (IVEK 或 Port)
                if "ivek" in machine:
                    self.resources.book_p0_experiment_ivek(day_key)
                    print(f"      ✓ 已預訂 IVEK", file=sys.stderr)
                    
                elif "port" in machine:
                    try:
                        port_numbers = re.findall(r'\d+', machine)
                        
                        if not port_numbers:
                            raise ValueError(f"無法從 '{machine}' 中提取 Port 編號")
                        
                        port_num = int(port_numbers[0])
                        
                        if not (1 <= port_num <= 12):
                            raise ValueError(f"Port 編號 {port_num} 超出範圍 (1-12)")
                        
                        self.resources.book_p0_experiment_port_pair(shift_key, port_num)
                        print(f"      ✓ 已預訂 Port{port_num} (含配對)", file=sys.stderr)
                        
                    except ValueError as ve:
                        print(f"      ⚠️ 警告: {ve}，將不預訂 Port", file=sys.stderr)
                    except Exception as e:
                        print(f"      ⚠️ 警告: 無法解析滴定機 '{machine}': {e}", file=sys.stderr)
                
                else:
                    print(f"      ⚠️ 警告: 滴定機類型 '{machine}' 無法識別", file=sys.stderr)
                
                # 步驟 5: 檢查是否與已排入的 P1/P2 衝突
                has_p1_p2_conflict = False
                conflict_details = []
                
                # 檢查人員衝突
                if shift_key in self.resources.person_usage:
                    p1_p5_people = self.resources.person_usage[shift_key]
                    if staff in p1_p5_people:
                        has_p1_p2_conflict = True
                        conflict_details.append(f"人員 {staff}")
                
                # 檢查凍乾機衝突
                if day_key in self.resources.dryer_usage:
                    p1_p5_dryers = self.resources.dryer_usage[day_key]
                    if dryer in p1_p5_dryers:
                        has_p1_p2_conflict = True
                        conflict_details.append(f"凍乾機 {dryer}")
                
                # 步驟 6: 更新 P0 資料標記
                p0_data['has_conflict'] = has_p1_p2_conflict
                p0_data['slot'] = slot_name
                p0_data['shift_key'] = shift_key
                p0_data['day_key'] = day_key
                
                if has_p1_p2_conflict:
                    print(f"      🔴 偵測到與 P1/P2 的資源衝突: {', '.join(conflict_details)}", file=sys.stderr)
                    print(f"      → Excel 將標記為淺紅色", file=sys.stderr)
                else:
                    print(f"      ✓ 無資源衝突", file=sys.stderr)
                
                print(f"      ✅ 實驗 P0 任務 {lot} 排程完成", file=sys.stderr)
                print(f"      → P3/P4/P5 將自動避開此任務的資源", file=sys.stderr)
                
            except Exception as e:
                print(f"      ❌ 實驗 P0 任務 {lot} 發生錯誤: {e}", file=sys.stderr)
                import traceback
                traceback.print_exc(file=sys.stderr)
                continue
        
        # 步驟 7: 排程完成統計
        print(f"\n{'='*70}", file=sys.stderr)
        print(f"🧪 實驗 P0 排程完成統計", file=sys.stderr)
        print(f"{'='*70}", file=sys.stderr)
        
        total = len(self.p0_experiment_records)
        has_conflict_count = sum(1 for p in self.p0_experiment_records if p.get('has_conflict', False))
        no_conflict_count = total - has_conflict_count
        
        print(f"  總計: {total} 個實驗 P0 任務", file=sys.stderr)
        print(f"  與 P1/P2 衝突 (淺紅色): {has_conflict_count} 個", file=sys.stderr)
        print(f"  無衝突 (標準格式): {no_conflict_count} 個", file=sys.stderr)
        print(f"  P3/P4/P5 將自動避開所有實驗 P0 的資源", file=sys.stderr)
        print(f"{'='*70}\n", file=sys.stderr)
    # ========================================
    # ✅ V9.7.14: 處理凍乾機為空任務
    # ========================================
    def handle_no_dryer_tasks(self):
        """
        處理凍乾機為空但有需求的任務
        規則: 例外 PN 排週三，其他排 Plan 前一天
        V9.9.2: 新增休假日檢查
        """
        print("\n" + "="*70, file=sys.stderr)
        print(f"🔥 Step 3.1: 處理凍乾機為空的任務（{CURRENT_VERSION}）", file=sys.stderr)
        print("="*70, file=sys.stderr)
        
        df_demand = self.data["demand"]
        df_prod_plan = self.db_constraints.get("production_Plan", pd.DataFrame())
        
        if df_prod_plan.empty:
            print("  ⚠️ production_Plan 為空", file=sys.stderr)
        
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        monday = (today + timedelta(days=-today.weekday(), weeks=0))
        w1_end = monday + timedelta(days=6)
        
        special_pn = "5714600102"
        this_wednesday = monday + timedelta(days=2)
        
        print(f"  W1 範圍: {monday.strftime('%Y-%m-%d')} ~ {w1_end.strftime('%Y-%m-%d')}", file=sys.stderr)
        print(f"  例外 PN ({special_pn}) 將排在: {this_wednesday.strftime('%Y-%m-%d')}", file=sys.stderr)
        
        date_cols_map = parse_date_columns(df_prod_plan)
        no_dryer_records_local = []
        
        for _, row in df_demand.iterrows():
            pn = str(row["料號"]).strip()
            
            if pn in self.scheduled_pns:
                continue
            
            w1_balance = row["Stock_plus_Dosing"] - row["第一周需求"]
            w1_shortage = abs(w1_balance) if w1_balance < 0 else 0
            
            if w1_shortage == 0:
                continue
            
            if pn not in self.constraints.index:
                continue
            
            constr_data = self.constraints.loc[pn]
            if isinstance(constr_data, pd.DataFrame):
                constr = constr_data.iloc[0]
            else:
                constr = constr_data
            
            dryer_value = str(constr.get("可用凍乾機", "")).strip()
            
            if dryer_value != "":
                continue
            
            print(f"  🔍 處理 {pn} (凍乾機為空, 需求量: {w1_shortage})", file=sys.stderr)
            
            schedule_date = None
            earliest_w1_date = None

            if pn not in df_prod_plan["PN"].values:
                if pn == special_pn:
                    print(f"    ⭐ {pn} (例外): 強制排入週三", file=sys.stderr)
                    schedule_date = this_wednesday
                else:
                    print(f"    ⚠️ {pn}: Plan 中無 W1 排程. 跳過.", file=sys.stderr)
                    continue
            else:
                pn_plan = df_prod_plan[df_prod_plan["PN"] == pn].iloc[0]
                w1_dates = []
                for col, col_date in date_cols_map.items():
                    if monday <= col_date <= w1_end:
                        qty = pn_plan.get(col, 0)
                        if pd.notna(qty) and float(qty) > 0:
                            w1_dates.append(col_date)
                
                if w1_dates:
                    earliest_w1_date = min(w1_dates)

                if earliest_w1_date is None:
                    if pn == special_pn:
                        print(f"    ⭐ {pn} (例外): 強制排入週三", file=sys.stderr)
                        schedule_date = this_wednesday
                    else:
                        print(f"    ⚠️ {pn}: Plan 中無 W1 排程. 跳過.", file=sys.stderr)
                        continue
                else:
                    schedule_date = earliest_w1_date - timedelta(days=1)
                    print(f"      {pn}: 排入前一天", file=sys.stderr)

            if schedule_date < monday:
                print(f"    ⚠️ {pn}: 排程日早於本週一，修正為週一", file=sys.stderr)
                schedule_date = monday
            
            # ========================================
            # V9.9.2: 檢查是否為休假日
            # ========================================
            if is_holiday(schedule_date):
                print(f"    ⚠️ {pn}: 原排程日 {schedule_date.strftime('%Y-%m-%d')} 為休假日", file=sys.stderr)
                
                # 嘗試往後找第一個非休假日
                max_search_days = 7
                found_alternative = False
                
                for offset in range(1, max_search_days + 1):
                    alternative_date = schedule_date + timedelta(days=offset)
                    
                    # 不超過本週末
                    if alternative_date > w1_end:
                        break
                    
                    if not is_holiday(alternative_date):
                        schedule_date = alternative_date
                        found_alternative = True
                        print(f"    ✅ {pn}: 調整至 {schedule_date.strftime('%Y-%m-%d')}", file=sys.stderr)
                        break
                
                if not found_alternative:
                    # 嘗試往前找
                    for offset in range(1, max_search_days + 1):
                        alternative_date = schedule_date - timedelta(days=offset)
                        
                        # 不早於週一
                        if alternative_date < monday:
                            break
                        
                        if not is_holiday(alternative_date):
                            schedule_date = alternative_date
                            found_alternative = True
                            print(f"    ✅ {pn}: 調整至 {schedule_date.strftime('%Y-%m-%d')}", file=sys.stderr)
                            break
                
                if not found_alternative:
                    print(f"    ❌ {pn}: 無法找到合適的非休假日，跳過", file=sys.stderr)
                    continue
            
            marker_name = row.get("BDC_Marker_Name", "")
            if not marker_name:
                print(f"    ⚠️ {pn} 無 Marker Name. 跳過.", file=sys.stderr)
                continue
            
            person_result = self._find_person_for_no_dryer(constr, schedule_date)
            
            if person_result is None:
                print(f"    ⚠️ {pn} 在 {schedule_date.strftime('%Y-%m-%d')} 無可用配藥人員. 跳過.", file=sys.stderr)
                continue
            
            person, is_single_slot = person_result
            
            print(f"    ✅ {pn} 排入 {schedule_date.strftime('%Y-%m-%d')}", file=sys.stderr)
            
            no_dryer_records_local.append({
                "日期": schedule_date.strftime("%Y-%m-%d"),
                "marker": marker_name,
                "lot": pn,
                "凍乾機台": "",
                "數量": w1_shortage,
                "配藥同仁": person,
                "RD給藥時間": "",
                "預計滴定時間": "",
                "預計結束": "",
                "班次": "PM",
                "滴定機": "",
                "is_no_dryer": True
            })
            
            self.scheduled_pns.add(pn)
        
        print(f"\n  ✅ 凍乾機為空任務完成 ({len(no_dryer_records_local)} 筆)", file=sys.stderr)
        if no_dryer_records_local:
            print("    這些配藥人員已占用整天，主排程將自動避開\n", file=sys.stderr)
        
        self.no_dryer_records.extend(no_dryer_records_local)

    def _find_person_for_no_dryer(self, constr, schedule_date):
        """為凍乾機為空的任務找配藥人員（占用全天）"""
        day_key = schedule_date.strftime('%Y-%m-%d')
        shift_key_am = f"{day_key}_AM"
        shift_key_pm = f"{day_key}_PM"
        
        all_people = set()
        is_single_slot = constr.get("V9_Single_Slot", False)
        
        for key in ["配藥人-1", "配藥人-2", "配藥人-3"]:
            person = constr.get(key)
            if person and str(person).strip():
                all_people.add(str(person).strip().lower()) # ✅ 修正：使用小寫
        
        all_people.discard("")
        
        if not all_people:
            print(f"        🔴 [{day_key}] 批次無可配藥人員", file=sys.stderr)
            return None
        
        for person in sorted(all_people):
            can_use_am, reason_am = self.resources.can_use_person(
                person, shift_key_am, day_key, is_single_slot
            )
            
            can_use_pm, reason_pm = self.resources.can_use_person(
                person, shift_key_pm, day_key, is_single_slot
            )
            
            if can_use_am and can_use_pm:
                self.resources.book_person(person, shift_key_am, day_key, is_single_slot)
                self.resources.book_person(person, shift_key_pm, day_key, is_single_slot)
                
                print(f"        🟢 [{day_key}] 選中人員: {person} (占用全天)", file=sys.stderr)
                return person, is_single_slot
            else:
                print(f"        🔴 [{day_key}] {person} 不可用", file=sys.stderr)
        
        print(f"        🔴 [{day_key}] 人員全被佔用", file=sys.stderr)
        return None
    
    def run(self):
        """
        ✅ V9.9.40: 主排程流程 (完整整合版)
        包含:
        1. Na_A/Na_B 連續兩天排程 (攔截邏輯)
        2. tCREA 同日綁定排程 (交易式檢查)
        3. CK/PHOS/K 凍乾機污染標記
        4. P1 任務計數修正
        5. tCREA 優先級提升
        """
        import sys
        from datetime import timedelta, datetime
        
        print(f"\nStep 3: 開始排程 ({CURRENT_VERSION})...\n", file=sys.stderr)
        
        # 設定全域起始日期
        global SCHEDULE_START_DATE
        if SCHEDULE_START_DATE is None:
            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            monday = (today + timedelta(days=-today.weekday(), weeks=0))
            print(f"  ⚠️ 警告: SCHEDULE_START_DATE 未設定，使用今天 {today.strftime('%Y-%m-%d')} 推算週一", file=sys.stderr)
        else:
            monday = SCHEDULE_START_DATE
        
        print(f"  排程起始: {monday.strftime('%Y-%m-%d')}\n", file=sys.stderr)
        
        saturday_date = monday + timedelta(days=5) 
        print(f"  P4/P5 截止日期 (週五): {(monday + timedelta(days=4)).strftime('%Y-%m-%d')}", file=sys.stderr)

        if HOLIDAYS:
            print(f"\n  ⚠️   休假日設定:", file=sys.stderr)
            for h in sorted(HOLIDAYS):
                print(f"      {h.strftime('%Y-%m-%d (%A)')}", file=sys.stderr)
            print(file=sys.stderr)

        self.shift_port_counter = {}

        # 改為在排程時動態分配時間 (移除固定時間槽生成)
        AM_START = time(10, 30)
        PM_START = time(15, 30)
        
        # 資源追蹤初始化
        day_batch_times = {}  # {day_key: {"AM": [], "PM": []}}

        for day_index in range(self.days_to_schedule + 1):
            if day_index == SCHEDULING_DAYS:
                w1_tasks_left = any(t[0] == 1 and t[2] not in self.scheduled_pns for t in self.task_queue)
                if not ALLOW_EXTEND_TO_6_DAYS or not w1_tasks_left:
                    break
                print("⚠️ 延長至第 6 天", file=sys.stderr)

            current_day = monday + timedelta(days=day_index)
            
            if is_holiday(current_day):
                print(f"  🗓️   跳過休假日: {current_day.strftime('%Y-%m-%d (%A)')}", file=sys.stderr)
                continue
            
            day_key = current_day.strftime('%Y-%m-%d')
            self.resources.reset_day(day_key)
            
            for slot_name in ["AM", "PM"]:
                shift_key = f"{day_key}_{slot_name}"
                self.resources.reset_shift(shift_key)
                self.shift_port_counter[shift_key] = 1

        print(f"生成 {self.days_to_schedule} 個排程日（已排除休假日）\n", file=sys.stderr)

        # P0 處理
        try:
            self.handle_p0_orders()
        except SystemExit as e:
            print(f"\n{str(e)}", file=sys.stderr)
            return

        self.handle_no_dryer_tasks()

        # ========================================
        # 主排程 (P1-P5)
        # ========================================
        print("="*70, file=sys.stderr)
        print("📋 Step 3.2: 開始主排程 (P1-P5)", file=sys.stderr)
        print("="*70, file=sys.stderr)

        print("計算任務優先級...", file=sys.stderr)

        task_priorities = []
        for task in self.task_queue:
            pn = task[2]
            
            if (pn in self.scheduled_pns) and (task[0] != 5):
                continue 
            
            batch_tasks = self.find_batch(task)
            if not batch_tasks:
                continue 
            
            constraints_list = self._normalize_constraints(batch_tasks)
            if not constraints_list:
                continue
            
            person_count, people_list = self.calculate_person_priority(constraints_list)
            
            task_priorities.append((task[0], task[1], person_count, task))
        
        # ==============================================================================
        # ✅ 修改排序邏輯: tCREA 至尊優先 (P0.5)
        # ==============================================================================
        TCREA_PNS = set(SPECIAL_CASE_PNS['tCREA']['batch1']) | set(SPECIAL_CASE_PNS['tCREA']['batch2'])

        task_priorities.sort(key=lambda x: (
            0.5 if x[3][2] in TCREA_PNS else x[0], # 1. tCREA 最優先
            0 if x[3][2] in self.RULE_GLIPA_PNs else 1, # 2. GLIPA VIP
            x[1][0], # 3. 機台數
            x[2],    # 4. 人數
            x[3][5], # 5. 藥名
            x[3][2]  # 6. PN
        ))
        
        # P6/P7 任務排序
        p6_p7_fill_tasks_sorted = sorted(
            self.p6_p7_fill_tasks, 
            key=lambda x: (x[0], x[1][2]) 
        )
        
        # 統計 P1-P5 需求
        priority_counts = {'P1': 0, 'P2': 0, 'P3': 0, 'P4': 0, 'P5': 0}
        for week, subp, person_count, task in task_priorities:
            priority_label = self._get_priority_label(week)
            if priority_label in priority_counts:
                priority_counts[priority_label] += 1
        
        print(f"\n📊 P1-P5 需求統計:", file=sys.stderr)
        for p, c in priority_counts.items():
            print(f"    {p}: {c} 批", file=sys.stderr)
        print(f"    總計: {sum(priority_counts.values())} 批\n", file=sys.stderr)
        
        print(f"任務優先級排序完成 (共 {len(task_priorities)} 個任務)\n", file=sys.stderr)

        shift_stats = {}
        scheduled_count = 0
        failed_reasons = {}
        p2_completed = False
        
        # P1 任務追蹤初始化
        p1_total = sum(1 for t in task_priorities if t[0] == 1)
        p1_tasks_set = set((t[3][5], t[3][2]) for t in task_priorities if t[0] == 1)
        p1_scheduled_count = 0
        
        # ========================================
        # 主排程循環
        # ========================================
        for week, subp, person_count, task in task_priorities:
            pn = task[2]
            
            # 插入實驗 P0
            if not p2_completed and week >= 3:
                p2_completed = True
                if self.p0_experiment_records:
                    print("\n" + "="*70, file=sys.stderr)
                    print("🧪 插入 P0 (實驗) 任務 (P2 後排程)", file=sys.stderr)
                    print("="*70, file=sys.stderr)
                    self.schedule_p0_experiments([]) 
            
            if (pn in self.scheduled_pns) and (week != 5):
                continue 

            batch_tasks = self.find_batch(task)
            if not batch_tasks:
                failed_reasons[pn] = "find_batch 失敗"
                continue 
            
            # ========================================
            # ✅ 攔截 Na_A / Na_B 進行連續排程
            # ========================================
            NA_A_PNS = {'5714400202', '5714400203'}
            NA_B_PNS = {'5714400209', '5714400210'}
            current_pns = set(t[2] for t in batch_tasks)
            
            if (current_pns & NA_A_PNS) or (current_pns & NA_B_PNS):
                if all(p in self.scheduled_pns for p in current_pns):
                    continue 
                
                success = self._schedule_na_sequential(batch_tasks, self.days_to_schedule, monday, day_batch_times)
                
                if success:
                    # 🔥 [修正] 使用外層變數更新 P1 計數
                    if week == 1:
                        scheduled_key = (task[5], task[2])
                        if scheduled_key in p1_tasks_set:
                            p1_scheduled_count += 1
                            p1_tasks_set.remove(scheduled_key)
                            print(f"  📊 [統計修正] P1 任務完成: {task[5]} - {task[2]}", file=sys.stderr)
                    continue 
                else:
                    failed_reasons[pn] = "Na 連續排程失敗"
                    continue

            # ========================================
            # 一般排程邏輯
            # ========================================
            batch_pns = [t[2] for t in batch_tasks]
            batch_name = batch_tasks[0][5]
            
            constraints_list = self._normalize_constraints(batch_tasks)
            if not constraints_list:
                failed_reasons[pn] = "normalize_constraints 失敗"
                continue
            
            slot_prefs = constraints_list[0].get("V9_Slot_List", ["AM", "PM"])
            _, people_list = self.calculate_person_priority(constraints_list)
            print(f"  嘗試排程: {batch_name} ({batch_pns}) - (P{week})", file=sys.stderr)
            
            fail_reason = None
            found_slot = False
            
            # 遍歷排程天數
            for day_offset in range(self.days_to_schedule + 1):
                if found_slot: break
                        
                current_day = monday + timedelta(days=day_offset)
                if is_holiday(current_day): continue
                
                # P4/P5 不排週六
                if (week >= 4) and current_day.date() == saturday_date.date():
                    if not fail_reason: fail_reason = "P4/P5 不排週六"
                    continue
                
                day_key = current_day.strftime('%Y-%m-%d')
                if day_key not in day_batch_times: day_batch_times[day_key] = {"AM": [], "PM": []}
                
                for slot_name in ["AM", "PM"]:
                    if found_slot: break
                    if slot_name not in slot_prefs: continue
                    
                    # 週五限制
                    can_schedule, friday_reason = self.can_schedule_on_friday(constraints_list, current_day)
                    if not can_schedule:
                        if not fail_reason: fail_reason = friday_reason
                        continue 

                    # 取得 RD 時間
                    next_time = self.get_delivery_time_from_constraints(
                        constraints_list, slot_name, current_day, day_batch_times, day_key
                    )
                    if next_time is None:
                        if not fail_reason: fail_reason = f"{day_key} {slot_name} 时段已滿"
                        continue

                    delivery_dt = datetime.combine(current_day, next_time)
                    shift_key = f"{day_key}_{slot_name}"
                    if shift_key not in shift_stats: shift_stats[shift_key] = {"scheduled": 0, "failed": 0}
                    
                    # =================================================================
                    # 1. 資源預檢查 (Batch 1)
                    # =================================================================
                    resources = self.check_availability(
                        batch_tasks, constraints_list, delivery_dt, slot_name, priority=week
                    )
                    
                    if not resources:
                        shift_stats[shift_key]["failed"] += 1
                        if not fail_reason: fail_reason = f"check_availability 失敗"
                        continue
                    
                    # 檢查禁止時段
                    final_end_dt = resources["times"]["final_end_dt"]
                    is_forbidden, forbidden_reason = self.is_forbidden_end_time(final_end_dt)
                    if is_forbidden:
                        if not fail_reason: fail_reason = forbidden_reason
                        continue

                    # =================================================================
                    # 2. ✅ tCREA "同日綁定" 強制檢查
                    # =================================================================
                    is_tcrea_batch1 = any(pn in SPECIAL_CASE_PNS['tCREA']['batch1'] for pn in batch_pns)
                    batch2_resources = None
                    batch2_tasks = []
                    batch2_delivery_dt = None
                    batch2_slot = None
                    can_schedule_batch2 = False
                    
                    if is_tcrea_batch1:
                        batch1_dryer = resources.get("dryer")
                        batch2_tasks = self._create_batch_from_pns(SPECIAL_CASE_PNS['tCREA']['batch2'])
                        if batch2_tasks:
                            batch2_constraints = self._normalize_constraints(batch2_tasks)
                            # 🔥 [新增] 強制從 Batch 2 的可用凍乾機中，剔除 Batch 1 用的那一台
                            if batch2_constraints:
                                for constr in batch2_constraints:
                                    available_str = str(constr.get("可用凍乾機", "")).strip()
                                    # 分割並過濾
                                    dryer_list = [d.strip() for d in available_str.split(',') if d.strip()]
                                    # 移除 batch1_dryer
                                    filtered_list = [d for d in dryer_list if d != batch1_dryer]
                                    
                                    if not filtered_list:
                                        print(f"  ⚠️ tCREA: 剔除 {batch1_dryer} 後，Batch 2 無可用凍乾機", file=sys.stderr)
                                    
                                    # 更新約束條件
                                    constr["可用凍乾機"] = ",".join(filtered_list)
                            
                            # 👆👆👆 [插入結束] 👆👆👆
                            batch2_slot = "AM" if slot_name == "PM" else "PM"
                            batch2_time = self.get_delivery_time_from_constraints(
                                batch2_constraints, batch2_slot, current_day, day_batch_times, day_key
                            )
                            if batch2_time:
                                batch2_delivery_dt = datetime.combine(current_day, batch2_time)
                                batch2_resources = self.check_availability(
                                    batch2_tasks, batch2_constraints, batch2_delivery_dt, batch2_slot, priority=week
                                )
                                if batch2_resources:
                                    if batch2_resources.get("dryer") != batch1_dryer:
                                        b2_end = batch2_resources["times"]["final_end_dt"]
                                        if not self.is_forbidden_end_time(b2_end)[0]:
                                            can_schedule_batch2 = True
                                    else:
                                         print(f"  ⚠️ tCREA: 兩批次凍乾機相同 ({batch1_dryer})", file=sys.stderr)

                        if not can_schedule_batch2:
                            print(f"  ⚠️ tCREA 同日綁定失敗: 跳過 {day_key}，嘗試下一天", file=sys.stderr)
                            if not fail_reason: fail_reason = "tCREA Batch 2 資源不足"
                            continue # 🔥 跳過這一天

                    # =================================================================
                    # 3. ✅ CK/PHOS/K 凍乾機污染標記 (移至 Commit 前)
                    # =================================================================
                    scheduled_pns_in_batch = set(t[2] for t in batch_tasks)
                    is_blocking_batch = any(pn in self.RULE_BLOCKING_PNs for pn in scheduled_pns_in_batch)
                            
                    if is_blocking_batch:
                        dryer_id = resources.get("dryer")
                        current_date = delivery_dt.date()
                        blocked_until_date = current_date + timedelta(days=self.RULE_BLOCK_DAYS)
                        
                        if (dryer_id in self.dryer_block_until_date and 
                            blocked_until_date > self.dryer_block_until_date[dryer_id]):
                            self.dryer_block_until_date[dryer_id] = blocked_until_date
                        elif dryer_id not in self.dryer_block_until_date:
                            self.dryer_block_until_date[dryer_id] = blocked_until_date
                                    
                        print(f"  ✅ 記錄 {batch_name} 排程於 {current_date}, 凍乾機 {dryer_id} 封鎖至 {blocked_until_date}", file=sys.stderr)

                    # =================================================================
                    # 4. 執行寫入 (Commit)
                    # =================================================================
                    
                    # 寫入 Batch 1
                    self.book_and_record(resources, batch_tasks, current_day, slot_name, priority=week)
                    self.book_resources(resources, delivery_dt, slot_name, priority=week)
                    day_batch_times[day_key][slot_name].append(next_time)
                    
                    for t in batch_tasks:
                        self.scheduled_pns.add(t[2])
                        if week == 1:
                            s_key = (t[5], t[2])
                            if s_key in p1_tasks_set:
                                p1_scheduled_count += 1
                                p1_tasks_set.remove(s_key)

                    scheduled_count += 1
                    shift_stats[shift_key]["scheduled"] += 1
                    print(f"  ✅ [{delivery_dt.strftime('%Y-%m-%d %H:%M')} {slot_name}] (P{week}) {batch_name}", file=sys.stderr)

                    # 寫入 Batch 2 (tCREA)
                    if is_tcrea_batch1 and batch2_resources:
                        self.book_and_record(batch2_resources, batch2_tasks, batch2_delivery_dt.date(), batch2_slot, priority=week)
                        self.book_resources(batch2_resources, batch2_delivery_dt, batch2_slot, priority=week)
                        day_batch_times[day_key][batch2_slot].append(batch2_time)
                        print(f"  ✅ tCREA batch2 連帶排程 ({day_key} {batch2_slot})", file=sys.stderr)
                        
                        for t in batch2_tasks:
                            self.scheduled_pns.add(t[2])
                            if week == 1:
                                s_key = (t[5], t[2])
                                if s_key in p1_tasks_set:
                                    p1_scheduled_count += 1
                                    p1_tasks_set.remove(s_key)

                    found_slot = True
                    break # 成功排入，跳出迴圈

            if not found_slot:
                failed_reasons[pn] = fail_reason or f"所有時段均失敗"
                print(f"    ❌ (P{week}) {batch_name} - {fail_reason}", file=sys.stderr)
        
        # ========================================
        # 收集未完成任務 & 填補空缺
        # ========================================
        unmet_tasks = []
        for t in self.task_queue:
            if t[0] <= 5 and t[2] not in self.scheduled_pns:
                unmet_tasks.append(t)

        print(f"\n  📊 統計: 共有 {len(unmet_tasks)} 個 P1-P5 任務未排入，將嘗試於填補階段優先重試。", file=sys.stderr)

        print("\n" + "="*70, file=sys.stderr)
        print("🧩 Step 3.3: 開始強力填補空缺 (優先重試 P1-P5 -> 再填 P6/P7)", file=sys.stderr)
        print("="*70, file=sys.stderr)
        
        all_fill_tasks = unmet_tasks + p6_p7_fill_tasks_sorted
        self.fill_gaps(monday, day_batch_times, all_fill_tasks)

        # ========================================
        # P1 完成檢查
        # ========================================
        if p1_total > 0 and p1_scheduled_count < p1_total:
            print("\n" + "="*70, file=sys.stderr)
            print("❌ 排程終止：P1 任務未全部滿足", file=sys.stderr)
            print("="*70, file=sys.stderr)
            print(f"P1 任務統計:", file=sys.stderr)
            print(f"  總計: {p1_total} 個任務", file=sys.stderr)
            print(f"  已排程: {p1_scheduled_count} 個任務", file=sys.stderr)
            print(f"  未排程: {p1_total - p1_scheduled_count} 個任務", file=sys.stderr)
            
            for group_name, pn in p1_tasks_set:
                reason = failed_reasons.get(pn, "未知原因")
                print(f"  • {group_name} (PN: {pn}) - 失敗原因: {reason}", file=sys.stderr)
            
            print("\n可能的解決方案:", file=sys.stderr)
            print("  1. 調整休假日設定", file=sys.stderr)
            print("  2. 調整 P1 任務的配藥限制", file=sys.stderr)
            print("  3. 延長排程天數", file=sys.stderr)
            print("  4. 減少 P0 任務數量", file=sys.stderr)
            print("="*70, file=sys.stderr)
            return 

        # 最終統計
        print(f"\n{'='*70}", file=sys.stderr)
        print(f"📊 排程統計", file=sys.stderr)
        print(f"{'='*70}", file=sys.stderr)
        print(f"  凍乾機為空: {len(self.no_dryer_records)} 筆", file=sys.stderr)
        print(f"  P0 (非實驗): {len(self.p0_strict_records)} 筆", file=sys.stderr)
        print(f"  P0 (實驗): {len(self.p0_experiment_records)} 筆", file=sys.stderr)
        print(f"  P1-P5 成功: {scheduled_count} 批次", file=sys.stderr)
        print(f"    - P1: {p1_scheduled_count}/{p1_total} {'✅' if p1_scheduled_count == p1_total else '❌'}", file=sys.stderr)
        
        unmet_pns = [t[2] for t in self.task_queue if t[0] < 4 and t[2] not in self.scheduled_pns]
        print(f"  未排入 PN (P1-P3): {len(unmet_pns)}", file=sys.stderr)
        
        print(f"\n{'='*70}", file=sys.stderr)
        print(f"⏰ RD給藥時間分配統計", file=sys.stderr)
        print(f"{'='*70}", file=sys.stderr)
        for day_key in sorted(day_batch_times.keys()):
            print(f"\n  {day_key}:", file=sys.stderr)
            for slot_name in ["AM", "PM"]:
                times = sorted(day_batch_times[day_key][slot_name])
                if times:
                    time_strs = [t.strftime('%H:%M') for t in times]
                    print(f"    {slot_name}: {', '.join(time_strs)} (共 {len(times)} 批)", file=sys.stderr)
        
        print(f"\n✅ 所有排程完成", file=sys.stderr)

    def fill_gaps(self, monday, day_batch_times, fill_tasks):
        """
        ✅ V9.9.31: 強力填補模式 (Slot-Centric)
        目標: 針對每個時段的空缺，優先重試 P1-P5，再填補 P6/P7
        邏輯: 鎖定一個班次 -> 算出可用 Port -> 掃描任務清單 -> 塞入 -> 重複直到滿
        """
        from datetime import timedelta
        import sys
        
        print("\n" + "="*80, file=sys.stderr)
        print("🧩 開始強力填補空缺流程 (Slot-Centric)", file=sys.stderr)
        print("="*80, file=sys.stderr)

        # 1. 建立快速查找表 (避免重複計算)
        # 從現有 schedule_df 重建 Port 佔用狀況，確保數據最新
        port_usage = {} 
        if hasattr(self, 'schedule_df') and not self.schedule_df.empty:
            for _, row in self.schedule_df.iterrows():
                d = row.get('日期')
                s = row.get('班次')
                p_list = row.get('ports_list', '')
                if d and s and p_list and str(p_list).upper() != 'IVEK':
                    try:
                        p_nums = [int(x) for x in str(p_list).split(',') if x.strip().isdigit()]
                        for pn in p_nums:
                            port_usage[f"{d}_{s}_Port{pn}"] = True
                    except: pass

        scheduled_count_in_gap = 0

        # 2. Slot-Centric 迴圈：遍歷每一天、每一個班次
        for day_offset in range(self.days_to_schedule):
            current_day = monday + timedelta(days=day_offset)
            
            # 跳過休假日
            if is_holiday(current_day): continue
            
            day_key = current_day.strftime('%Y-%m-%d')
            
            for slot_name in ["AM", "PM"]:
                shift_key = f"{day_key}_{slot_name}"
                
                # 初始化該班次的 Port 指標 (如果尚未存在)
                if shift_key not in self.shift_port_counter:
                    self.shift_port_counter[shift_key] = 1
                
                # === 內層迴圈：只要這個班次還有空間，就持續嘗試塞任務 ===
                while True:
                    # A. 計算當前可用的起始 Port (避開 P0 + 單數對齊)
                    port_start = self.shift_port_counter[shift_key]
                    
                    while True:
                        is_booked = False
                        # 檢查 P0 佔用
                        if self.resources.is_p0_strict_port_booked(shift_key, port_start): is_booked = True
                        # 檢查是否被主排程佔用 (Double Check)
                        if port_usage.get(f"{shift_key}_Port{port_start}", False): is_booked = True
                        
                        # 強制單數對齊 (Odd Alignment)
                        if port_start % 2 == 0:
                            port_start += 1
                            continue
                            
                        if is_booked:
                            port_start += 1
                        else:
                            break
                    
                    # 如果起始 Port 超過 12，代表這班次真的滿了，跳出
                    if port_start > MAX_PORTS:
                        break
                    
                    # 更新 counter
                    self.shift_port_counter[shift_key] = port_start
                    
                    # B. 遍歷任務清單，尋找能塞進去的「第一個」任務
                    scheduled_something = False
                    
                    for i, task in enumerate(fill_tasks):
                        pn = task[2]
                        group_name = task[5]
                        priority = task[0]
                        
                        # 1. 檢查是否已排程
                        if pn in self.scheduled_pns:
                            continue
                            
                        # 2. 排除特殊案例 (GLIPA/tCREA 等複雜配對，填補階段建議跳過，除非邏輯非常有把握)
                        # 若要開放，需確保 find_batch 能正確處理
                        if pn in ALL_SPECIAL_PNS:
                            continue

                        # 3. 取得 Batch (同藥名一起排)
                        batch_tasks = self.find_batch(task)
                        if not batch_tasks: continue
                        
                        # 再次檢查整批是否都乾淨 (未排程)
                        if any(t[2] in self.scheduled_pns for t in batch_tasks):
                            continue

                        # 4. 計算需要的 Port 空間 (含對齊)
                        ports_needed = self._calculate_aligned_ports(batch_tasks)
                        if ports_needed == 0: continue # IVEK 或異常
                        
                        # 5. 檢查空間是否足夠
                        if port_start + ports_needed - 1 > MAX_PORTS:
                            # 空間不夠，嘗試下一個比較小的任務
                            continue
                            
                        # 6. 準備約束條件
                        constraints_list = self._normalize_constraints(batch_tasks)
                        
                        # 7. 資源檢查 (check_availability)
                        # 獲取動態時間
                        next_time = self.get_delivery_time_from_constraints(
                            constraints_list, slot_name, current_day, day_batch_times, day_key
                        )
                        if not next_time: 
                            # 該時段的時間點已滿 (例如 RD 給藥時間都用光了)
                            break # 結束這個班次的嘗試
                            
                        delivery_dt = datetime.combine(current_day, next_time)
                        
                        # 執行核心檢查 (人員、凍乾機)
                        resources = self.check_availability(
                            batch_tasks, constraints_list, delivery_dt, slot_name, priority=priority
                        )
                        
                        if resources:
                            # 8. 檢查禁止時段 (03:00-08:00)
                            final_end_dt = resources["times"]["final_end_dt"]
                            is_forbidden, _ = self.is_forbidden_end_time(final_end_dt)
                            if is_forbidden:
                                continue

                            # === 9. 匹配成功！執行寫入 ===
                            print(f"  🧩 [填補成功] {day_key} {slot_name}: {group_name} (P{priority})", file=sys.stderr)
                            
                            # 寫入 Excel/DB 結構
                            self.book_and_record(resources, batch_tasks, current_day, slot_name, priority=priority)
                            # 佔用資源 (鎖定人員/凍乾機)
                            self.book_resources(resources, delivery_dt, slot_name, priority=priority)
                            
                            # 更新時間記錄
                            if day_key not in day_batch_times: day_batch_times[day_key] = {"AM": [], "PM": []}
                            day_batch_times[day_key][slot_name].append(next_time)
                            
                            # 標記已排程
                            for t in batch_tasks:
                                self.scheduled_pns.add(t[2])
                            
                            # 更新本地 Port Usage (防止同一個班次重複判斷)
                            # 注意：book_and_record 已經更新了 self.shift_port_counter
                            # 我們只需要標記成功，並跳出任務迴圈，重新進入 While 迴圈計算新的 port_start
                            scheduled_something = True
                            scheduled_count_in_gap += 1
                            break 
                    
                    # C. 如果掃描了一整輪任務，都沒法塞進目前的空位，那就結束這個班次
                    if not scheduled_something:
                        break
                        
        print("\n" + "="*80, file=sys.stderr)
        print(f"🧩 強力填補結束，共成功填入 {scheduled_count_in_gap} 批次", file=sys.stderr)
        print("="*80, file=sys.stderr)
          

    def book_and_record(self, resources, batch_tasks, current_day, slot_name, priority=None):
        """V9.9.13: Port 滿時不記錄，P3-P5 自動跳過實驗 P0 Port（支援 GLIPA）"""
        import pandas as pd

        # ========================================
        # ✅ V9.9.13: 檢查是否為 GLIPA
        # ========================================
        is_glipa = resources.get("is_glipa", False)
        
        if is_glipa:
            return self._book_and_record_glipa(resources, batch_tasks, current_day, slot_name, priority)
        
        # ========================================
        # 原有邏輯（一般任務）
        # ========================================
        times = resources["times"]
        dryer = resources.get("dryer", "N/A") 
        person = resources.get("person", "")
        freeze_duration = times.get("freeze_duration_hr", "N/A")
        
        day_key = current_day.strftime("%Y-%m-%d")
        shift_key = f"{day_key}_{slot_name}"
        
        # ✅ 修正：更精確判斷 IVEK 任務
        # 只要 constraints 中 Port數 寫 "IVEK"，或者 Port數 為 0，都視為 IVEK
        # ✅ 更可靠的 IVEK 判定邏輯
        num_ports_str = resources.get("num_ports_str", "2")
        is_ivek_explicit = resources.get("is_ivek", False)  # ← 新增明確標記
        is_ivek_task = (
            is_ivek_explicit or 
            str(num_ports_str).upper() == "IVEK" or 
            resources.get("num_ports_per_pn", 2) == 0
        )

        rows = []

        if is_ivek_task:
            # ========================================
            # ✅ IVEK 任務寫入 (拆成 2 筆)
            # ========================================
            print(f"    → {batch_tasks[0][5]}: IVEK 任務 (拆分兩筆顯示)", file=sys.stderr)
            
            for task in batch_tasks:
                week, subp, pn, short_qty, tag, group_name, marker_name, prod_qty = task
                
                # 🔥 關鍵修改：數量除以 2
                split_qty = prod_qty / 2  # 12000 -> 6000
                
                # 🔥 關鍵修改：寫入兩次
                for i in range(2):
                    rows.append({
                        "日期": current_day.strftime("%Y-%m-%d"),
                        "marker": marker_name, 
                        "滴定機": "IVEK",  # Excel 會顯示在最上方
                        "凍乾機台": dryer,
                        "配藥同仁": person,
                        "RD給藥時間": times["delivery_dt"].strftime("%H:%M"), 
                        "預計滴定時間": times["dosing_start_dt"].strftime("%H:%M"),
                        "預計結束": times["final_end_dt"].strftime("%H:%M"),
                        "預冷時間": (times["dosing_end_dt"] + timedelta(hours=2)).strftime("%H:%M"), 
                        "凍乾時間": f"{freeze_duration:.1f} hr", 
                        "收藥時間": times["final_end_dt"].strftime("%H:%M"), 
                        "數量": split_qty,  # 顯示 6000
                        "lot": pn,
                        "班次": slot_name,
                        "ports_list": "IVEK",
                        "record_type": "P1-P5",
                        "has_conflict": False
                    })
            
            # (IVEK 任務不佔用 Port 1-12, 不更新 self.shift_port_counter)

        else:
            # ========================================
            # ✅ V9.9.30: Port 1-12 排程 (含單雙數對齊邏輯)
            # ====================================
            
            if shift_key not in self.shift_port_counter:
                self.shift_port_counter[shift_key] = 1
            
            port_start = self.shift_port_counter[shift_key]

            # === P1-P5 Port 分配自動跳過被 P0 佔用的 Port ===
            # 這裡要確保 port_start 是單數 (Odd)，如果被佔用或目前是指向雙數，要往後找
            # 規則: 每個 Batch 總是從單數開始 (因為我們的邏輯是以 Pair 為單位)
            while True:
                is_booked = False
                # 檢查是否被佔用
                if self.resources.is_p0_strict_port_booked(shift_key, port_start): is_booked = True
                if priority is not None and priority >= 3 and self.resources.is_p0_experiment_port_booked(shift_key, port_start): is_booked = True
                
                # 強制從單數開始: 如果目前是雙數(偶數)，直接+1
                if port_start % 2 == 0:
                    port_start += 1
                    continue

                if is_booked:
                    print(f"        ⚠️ Port{port_start} 被 P0 佔用，跳過", file=sys.stderr)
                    port_start += 1 # 會變雙數，下個迴圈會再+1變單數
                else:
                    break

            self.shift_port_counter[shift_key] = port_start
            
            # 重新計算總需求 (Double Check)
            total_ports_needed = self._calculate_aligned_ports(batch_tasks)

            if port_start + total_ports_needed - 1 > MAX_PORTS:
                print(f"    ⚠️ {shift_key} Port 不足 (需要 {total_ports_needed}, 起始 {port_start})", file=sys.stderr)
                return

            # 開始分配
            current_port_idx = port_start
            
            # 我們需要手動遍歷 batch_tasks 來處理配對
            i = 0
            n = len(batch_tasks)
            
            while i < n:
                task = batch_tasks[i]
                # 解析 task 資訊
                week, subp, pn, short_qty, tag, group_name, marker_name, prod_qty = task
                
                # 取得該 PN 的 Port 數
                actual_ports = 2 # 預設
                if pn in self.constraints.index:
                    c = self.constraints.loc[pn]
                    if isinstance(c, pd.DataFrame): c = c.iloc[0]
                    try: actual_ports = int(str(c.get("Port數", "2")).strip())
                    except: actual_ports = 2
                
                assigned_ports = []
                
                if actual_ports >= 2:
                    assigned_ports = list(range(current_port_idx, current_port_idx + actual_ports))
                    current_port_idx += actual_ports
                    
                    # 確保下一個起始位置是單數
                    if current_port_idx % 2 == 0:
                        current_port_idx += 1
                    
                    i += 1
                    
                elif actual_ports == 1:
                    # 狀況 B: 1 Port
                    # 檢查下一個 PN 是否也是 1 Port
                    next_task = None
                    next_pn_ports = 0
                    if i + 1 < n:
                        next_task_data = batch_tasks[i+1]
                        # 取得下一個 PN 的 Port 數
                        try:
                            n_pn = next_task_data[2]
                            nc = self.constraints.loc[n_pn] if n_pn in self.constraints.index else {}
                            if isinstance(nc, pd.DataFrame): nc = nc.iloc[0]
                            next_pn_ports = int(str(nc.get("Port數", "2")).strip())
                        except: next_pn_ports = 2

                    if next_pn_ports == 1:
                        # 狀況 B-1: 配對成功 (CRP-D + CRP-U)
                        # 當前 Task -> Odd
                        assigned_ports = [current_port_idx]
                        
                        # 下一個 Task -> Even (立即處理)
                        next_task = batch_tasks[i+1]
                        n_week, n_subp, n_pn, n_short, n_tag, n_group, n_marker, n_prod = next_task
                        
                        # 寫入下一個 Task (雙數 Port)
                        next_assigned_ports = [current_port_idx + 1]
                        # ... (這裡需要呼叫寫入邏輯，為了避免重複代碼，我們下面統一寫入)
                        # 特殊處理: 在此迴圈內直接寫入下一個 task
                        
                        # 增加記錄 (Next Task)
                        rows.append({
                            "日期": current_day.strftime("%Y-%m-%d"),
                            "marker": n_marker, 
                            "滴定機": f"Port{next_assigned_ports[0]}",
                            "凍乾機台": dryer,
                            "配藥同仁": person,
                            "RD給藥時間": times["delivery_dt"].strftime("%H:%M"), 
                            "預計滴定時間": times["dosing_start_dt"].strftime("%H:%M"),
                            "預計結束": times["final_end_dt"].strftime("%H:%M"),
                            "預冷時間": (times["dosing_end_dt"] + timedelta(hours=2)).strftime("%H:%M"), 
                            "凍乾時間": f"{freeze_duration:.1f} hr", 
                            "收藥時間": times["final_end_dt"].strftime("%H:%M"), 
                            "數量": n_prod,
                            "lot": n_pn,
                            "班次": slot_name,
                            "ports_list": str(next_assigned_ports[0])
                        })
                        print(f"        [Pair] {n_pn} -> Port {next_assigned_ports[0]} (Even)", file=sys.stderr)
                        
                        current_port_idx += 2 # 跳過這對
                        i += 2 # 處理了兩個任務
                    else:
                        # 狀況 B-2: 落單 (GLIPA-AU 或 單獨) -> 佔用 Odd, Block Even
                        assigned_ports = [current_port_idx]
                        # Even port (current_port_idx + 1) 被留空
                        print(f"        [Block] Port {current_port_idx + 1} 保留為空", file=sys.stderr)
                        current_port_idx += 2
                        i += 1
                
                # 寫入當前 Task (Odd Port or 2 Ports)
                if assigned_ports:
                    port_str = ",".join([f"Port{p}" for p in assigned_ports])
                    rows.append({
                        "日期": current_day.strftime("%Y-%m-%d"),
                        "marker": marker_name, 
                        "滴定機": port_str,
                        "凍乾機台": dryer,
                        "配藥同仁": person,
                        "RD給藥時間": times["delivery_dt"].strftime("%H:%M"), 
                        "預計滴定時間": times["dosing_start_dt"].strftime("%H:%M"),
                        "預計結束": times["final_end_dt"].strftime("%H:%M"),
                        "預冷時間": (times["dosing_end_dt"] + timedelta(hours=2)).strftime("%H:%M"), 
                        "凍乾時間": f"{freeze_duration:.1f} hr", 
                        "收藥時間": times["final_end_dt"].strftime("%H:%M"), 
                        "數量": prod_qty,
                        "lot": pn,
                        "班次": slot_name,
                        "ports_list": ",".join(map(str, assigned_ports))
                    })
            
            # 更新 global counter
            self.shift_port_counter[shift_key] = current_port_idx
            print(f"    → {batch_tasks[0][5]}: {len(rows)} PN 排程完畢, Next Port: {current_port_idx}", file=sys.stderr)

        
        # === 共通邏輯：寫入 DataFrame ===
        new_df = pd.DataFrame(rows)
        if not hasattr(self, "schedule_df") or self.schedule_df.empty:
            self.schedule_df = new_df
        else:
            self.schedule_df = pd.concat([self.schedule_df, new_df], ignore_index=True)

    def _book_and_record_glipa(self, resources, batch_tasks, current_day, slot_name, priority=None):
        """
        ✅ GLIPA 專用記錄方法（使用兩台不同凍乾機 + 強制單雙數對齊 + AU留空）
        """
        import sys
        import pandas as pd
        from datetime import timedelta
        
        times = resources["times"]
        person = resources.get("person", "")
        dryer1 = resources.get("dryer", "N/A")
        dryer2 = resources.get("dryer2", "N/A")
        freeze_duration = times.get("freeze_duration_hr", "N/A")
        pn_port_map = resources.get("pn_port_map", {})
        # 注意: 這裡的 total_ports_needed 應該要是已經包含留空的數量 (例如 4)
        total_ports_needed = resources.get("total_ports_needed", 4)
        
        day_key = current_day.strftime("%Y-%m-%d")
        shift_key = f"{day_key}_{slot_name}"
        
        print(f"    📝 [DEBUG] GLIPA 開始記錄", file=sys.stderr)
        print(f"       凍乾機1: {dryer1}, 凍乾機2: {dryer2}", file=sys.stderr)
        
        if dryer1 == dryer2:
            print(f"    ❌ GLIPA: 兩台凍乾機相同 ({dryer1})，無法記錄", file=sys.stderr)
            return
        
        if shift_key not in self.shift_port_counter:
            self.shift_port_counter[shift_key] = 1
        
        port_start = self.shift_port_counter[shift_key]
        
        # === 步驟 1: 尋找起始 Port (避開 P0 + 強制單數對齊) ===
        while True:
            is_booked = False
            # 檢查是否被 P0 佔用
            if self.resources.is_p0_strict_port_booked(shift_key, port_start): is_booked = True
            if priority is not None and priority >= 3 and self.resources.is_p0_experiment_port_booked(shift_key, port_start): is_booked = True
            
            # 🔥 強制從單數開始 (Odd Alignment)
            if port_start % 2 == 0:
                port_start += 1
                continue

            if is_booked:
                print(f"    ⚠️ P0 佔用: Port{port_start}，自動跳過", file=sys.stderr)
                port_start += 1 # 變雙數，下個迴圈會再+1變單數
            else:
                break
        
        # 檢查空間是否足夠
        if port_start + total_ports_needed - 1 > MAX_PORTS:
            print(f"    ⚠️ {shift_key} Port 不足（GLIPA 需要 {total_ports_needed} Port, 起始 {port_start}）", file=sys.stderr)
            return
        
        rows = []
        current_port_idx = port_start
        
        # === 步驟 2: 遍歷任務並分配 Port (含留空邏輯) ===
        # 🔥 這裡補回了原本遺失的迴圈 🔥
        for idx, task in enumerate(batch_tasks):
            week, subp, pn, short_qty, tag, group_name, marker_name, prod_qty = task
            
            num_ports = pn_port_map.get(pn, 2)
            dryer = dryer1 if idx == 0 else dryer2
            
            assigned_ports = []
            
            if num_ports == 2:
                # GLIPA-AD: 佔用 2 個 [Odd, Even]
                assigned_ports = [current_port_idx, current_port_idx + 1]
                current_port_idx += 2
            elif num_ports == 1:
                # GLIPA-AU: 佔用 1 個 [Odd], 下一個 [Even] 留空
                assigned_ports = [current_port_idx]
                print(f"        [Block-GLIPA] Port {current_port_idx + 1} 保留為空", file=sys.stderr)
                current_port_idx += 2 # 🔥 強制跳過雙數，+2
            
            port_str = ",".join([f"Port{p}" for p in assigned_ports])
            ports_list_str = ",".join(map(str, assigned_ports))

            rows.append({
                "日期": current_day.strftime("%Y-%m-%d"),
                "marker": marker_name,
                "滴定機": port_str,
                "凍乾機台": dryer,
                "配藥同仁": person,
                "RD給藥時間": times["delivery_dt"].strftime("%H:%M"),
                "預計滴定時間": times["dosing_start_dt"].strftime("%H:%M"),
                "預計結束": times["final_end_dt"].strftime("%H:%M"),
                "預冷時間": (times["dosing_end_dt"] + timedelta(hours=2)).strftime("%H:%M"),
                "凍乾時間": f"{freeze_duration:.1f} hr",
                "收藥時間": times["final_end_dt"].strftime("%H:%M"),
                "數量": prod_qty,
                "lot": pn,
                "班次": slot_name,
                "ports_list": ports_list_str
            })
            
            print(f"       PN {pn} ({marker_name}):", file=sys.stderr)
            print(f"          Port: {port_str} ({num_ports} Port)", file=sys.stderr)
            print(f"          凍乾機: {dryer}", file=sys.stderr)
        
        # 更新全域計數器
        self.shift_port_counter[shift_key] = current_port_idx
        
        print(f"    ✅ [DEBUG] GLIPA 記錄完成:", file=sys.stderr)
        print(f"       Port 範圍: Port{port_start}-Port{current_port_idx-1}", file=sys.stderr)
        print(f"       凍乾機: {dryer1}, {dryer2} ✓不同", file=sys.stderr)
        
        new_df = pd.DataFrame(rows)
        if not hasattr(self, "schedule_df") or self.schedule_df.empty:
            self.schedule_df = new_df
        else:
            self.schedule_df = pd.concat([self.schedule_df, new_df], ignore_index=True)

    def _generate_dry_run_preview(self):
        """生成 Dry Run 預覽資料（JSON 格式）"""
        preview_data = []
        
        # P1-P5 排程
        if not self.schedule_df.empty:
            df_sorted = self.schedule_df.sort_values(by=["日期", "班次", "滴定機"])
            
            for _, row in df_sorted.iterrows():
                preview_data.append({
                    "date": row.get("日期", ""),
                    "titrate": row.get("滴定機", ""),
                    "freeze": row.get("凍乾機台", ""),
                    "pn": row.get("lot", ""),
                    "qty": row.get("數量", 0),
                    "staff": row.get("配藥同仁", ""),
                    "Name": row.get("marker", "")
                })
        
        # P0 (非實驗)
        for p0_data in self.p0_strict_records:
            task_date = p0_data.get('P0_Parsed_Date')
            if task_date:
                preview_data.append({
                    "date": task_date.strftime('%Y-%m-%d'),
                    "titrate": str(p0_data.get("滴定機", "")).strip(),
                    "freeze": str(p0_data.get("凍乾機台", "")).strip(),
                    "pn": str(p0_data.get("Lot", "")).strip(),
                    "qty": p0_data.get("數量", 0),
                    "staff": str(p0_data.get("配藥同仁", "")).strip(),
                    "Name": str(p0_data.get("Marker", "")).strip()
                })
        
        # P0 (實驗)
        for p0_data in self.p0_experiment_records:
            task_date = p0_data.get('P0_Parsed_Date')
            if task_date:
                preview_data.append({
                    "date": task_date.strftime('%Y-%m-%d'),
                    "titrate": str(p0_data.get("滴定機", "")).strip(),
                    "freeze": str(p0_data.get("凍乾機台", "")).strip(),
                    "pn": str(p0_data.get("Lot", "")).strip(),
                    "qty": p0_data.get("數量", 0),
                    "staff": str(p0_data.get("配藥同仁", "")).strip(),
                    "Name": str(p0_data.get("Marker", "")).strip()
                })
        
        # 凍乾機為空
        for record in self.no_dryer_records:
            preview_data.append({
                "date": record.get("日期", ""),
                "titrate": "",
                "freeze": "",
                "pn": record.get("lot", ""),
                "qty": record.get("數量", 0),
                "staff": record.get("配藥同仁", ""),
                "Name": record.get("marker", "")
            })
        
        return preview_data
    def _get_priority_label(self, week):
        """
        ✅ V9.9.16: 取得 priority 標籤
        
        Args:
            week: 週次 (1-5)
        
        Returns:
            str: "P1", "P2", "P3", "P4", "P5", 或 "Unknown"
        """
        priority_map = {
            1: 'P1',
            2: 'P2',
            3: 'P3',
            4: 'P4',
            5: 'P5',
        }
        return priority_map.get(week, 'Unknown')
    
    def save_records_to_excel(self, outdir):
        """
        ✅ V9.9.37: 固定格式 Excel 輸出 (修正 IVEK 顯示問題)
        修正點：當 ports_list 為 "IVEK" 時，不要過濾掉
        """
        import pandas as pd
        import os
        import sqlite3
        from datetime import datetime
        import re
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        import sys

        print("\n" + "="*70, file=sys.stderr)
        print(f"📊 Step 4: 輸出固定格式 Excel ({CURRENT_VERSION})", file=sys.stderr)
        print("="*70, file=sys.stderr)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(outdir, f"排程結果_{timestamp}.xlsx")

        # ========================================
        # 統一收集所有記錄
        # ========================================
        all_records = []
        
        global BATCH_START_NUMBER
        batch_start = BATCH_START_NUMBER
        work_order_counter = batch_start
        pn_work_order_map = {}
        pn_week_port_counter = {}
        
        # 1. 處理 P1-P5 排程 (包含 IVEK)
        if hasattr(self, 'schedule_df') and not self.schedule_df.empty:
            print(f"  收集 P1-P5 排程: {len(self.schedule_df)} 筆", file=sys.stderr)
            
            df_sorted = self.schedule_df.sort_values(by=["日期", "班次", "滴定機"])
            
            for _, row in df_sorted.iterrows():
                pn = row.get("lot", "")
                date_str = row.get("日期", "")
                ports_list_str = str(row.get("ports_list", "")).strip()
                
                # 🔥 [修正關鍵]：分辨 IVEK 與一般 Port
                is_ivek_row = (ports_list_str.upper() == "IVEK")
                
                ports = []
                if is_ivek_row:
                    ports = [0] # 用 0 代表 IVEK，避免被當成空清單過濾掉
                elif ports_list_str:
                    ports = [int(p.strip()) for p in ports_list_str.split(',') if p.strip().isdigit()]
                
                # 如果真的沒有 Port 也不是 IVEK，才跳過
                if not ports:
                    continue
                
                total_qty = row.get("數量", 0)
                # IVEK 不用除以 Port 數 (已經在 book_and_record 除過了)
                qty_per_port = total_qty if is_ivek_row else (total_qty / len(ports) if len(ports) > 0 else 0)
                
                # 生成工單號碼
                if pn not in pn_work_order_map:
                    try:
                        schedule_date = datetime.strptime(date_str, "%Y-%m-%d")
                        year_yy = schedule_date.strftime("%y")
                        month_char = chr(64 + schedule_date.month)
                        work_order_num = f"TMRA{year_yy}{month_char}{work_order_counter:03d}"
                        pn_work_order_map[pn] = work_order_num
                        work_order_counter += 1
                    except:
                        pn_work_order_map[pn] = f"TMRA25L{work_order_counter:03d}"
                        work_order_counter += 1
                
                work_order = pn_work_order_map[pn]
                
                # 計算 Batch Code
                try:
                    schedule_date = datetime.strptime(date_str, "%Y-%m-%d")
                    iso_year, iso_week, iso_day = schedule_date.isocalendar()
                    year_yy = schedule_date.strftime("%y")
                    pn_last3 = pn[-3:] if len(pn) >= 3 else pn.zfill(3)
                except:
                    iso_week = 1
                    year_yy = "25"
                    pn_last3 = "000"
                
                # 為每個 Port (或是 IVEK 本身) 建立記錄
                for port_num in ports:
                    
                    # Batch Code 邏輯
                    pn_week_key = (pn, iso_week)
                    if pn_week_key not in pn_week_port_counter:
                        pn_week_port_counter[pn_week_key] = 0
                    pn_week_port_counter[pn_week_key] += 1
                    port_count = pn_week_port_counter[pn_week_key]
                    
                    if port_count <= 8:
                        count_char = str(port_count)
                    elif port_count == 9:
                        count_char = 'A'
                    else:
                        count_char = chr(65 + (port_count - 9))
                    
                    batch_code = f"{pn_last3}{year_yy}{iso_week:02d}{count_char}"
                    
                    # 決定滴定機名稱
                    if is_ivek_row:
                        machine_name = "IVEK"
                    else:
                        machine_name = f"Port{port_num}"

                    rd_time = row.get("RD給藥時間", "")
                    if not rd_time:
                        slot_name = row.get("班次", "AM")
                        rd_time = "10:30" if slot_name == "AM" else "15:30"
                    
                    record = {
                        "滴定機": machine_name,
                        "Marker": row.get("marker", ""),
                        "PN": pn,
                        "凍乾機台": row.get("凍乾機台", ""),
                        "數量": round(qty_per_port, 2),
                        "配藥同仁": row.get("配藥同仁", ""),
                        "日期": date_str,
                        "RD給藥時間": rd_time,
                        "預計滴定時間": row.get("預計滴定時間", ""),
                        "預計結束": row.get("預計結束", ""),
                        "工單號碼": work_order,
                        "Lot": batch_code,
                        "備註": "",
                        "班次": row.get("班次", "AM"),
                        "record_type": "P1-P5",
                        "has_conflict": False
                    }
                    
                    all_records.append(record)

        # 2. 處理 P0 (非實驗)
        if self.p0_strict_records:
            print(f"  收集 P0 (非實驗): {len(self.p0_strict_records)} 筆", file=sys.stderr)
            for p0_data in self.p0_strict_records:
                task_date = p0_data.get('P0_Parsed_Date')
                day_key = task_date.strftime('%Y-%m-%d')
                rd_time_str = str(p0_data.get("RD給藥時間", "09:00")).strip()
                try:
                    rd_hour = int(rd_time_str.split(':')[0])
                    slot_name = "PM" if rd_hour >= 15 else "AM"
                except:
                    slot_name = "AM"
                
                record = {
                    "滴定機": str(p0_data.get("滴定機", "")).strip(),
                    "Marker": str(p0_data.get("Marker", "")).strip(),
                    "PN": str(p0_data.get("PN", "")).strip(),
                    "凍乾機台": str(p0_data.get("凍乾機台", "")).strip(),
                    "數量": str(p0_data.get("數量", "")).strip(),
                    "配藥同仁": str(p0_data.get("配藥同仁", "")).strip(),
                    "日期": day_key,
                    "RD給藥時間": rd_time_str,
                    "預計滴定時間": str(p0_data.get("預計滴定時間", "")).strip(),
                    "預計結束": str(p0_data.get("預計結束", "")).strip(),
                    "工單號碼": "",
                    "Lot": str(p0_data.get("Lot", "")).strip(),
                    "備註": str(p0_data.get("備註", "")).strip(),
                    "班次": slot_name,
                    "record_type": "P0_strict",
                    "has_conflict": False
                }
                all_records.append(record)

        # 3. 處理 P0 (實驗)
        if self.p0_experiment_records:
            print(f"  收集 P0 (實驗): {len(self.p0_experiment_records)} 筆", file=sys.stderr)
            for p0_data in self.p0_experiment_records:
                task_date = p0_data.get('P0_Parsed_Date')
                day_key = task_date.strftime('%Y-%m-%d')
                rd_time_str = str(p0_data.get("RD給藥時間", "09:00")).strip()
                try:
                    rd_hour = int(rd_time_str.split(':')[0])
                    slot_name = "PM" if rd_hour >= 15 else "AM"
                except:
                    slot_name = "AM"
                
                record = {
                    "滴定機": str(p0_data.get("滴定機", "")).strip(),
                    "Marker": str(p0_data.get("Marker", "")).strip(),
                    "PN": str(p0_data.get("PN", "")).strip(),
                    "凍乾機台": str(p0_data.get("凍乾機台", "")).strip(),
                    "數量": p0_data.get("數量", 0),
                    "配藥同仁": str(p0_data.get("配藥同仁", "")).strip(),
                    "日期": day_key,
                    "RD給藥時間": rd_time_str,
                    "預計滴定時間": str(p0_data.get("預計滴定時間", "")).strip(),
                    "預計結束": str(p0_data.get("預計結束", "")).strip(),
                    "工單號碼": "",
                    "Lot": str(p0_data.get("Lot", "")).strip(),
                    "備註": str(p0_data.get("備註", "")).strip(),
                    "班次": slot_name,
                    "record_type": "P0_experiment",
                    "has_conflict": p0_data.get('has_conflict', False)
                }
                all_records.append(record)

        # 4. 處理凍乾機為空
        if self.no_dryer_records:
            print(f"  收集凍乾機為空任務: {len(self.no_dryer_records)} 筆", file=sys.stderr)
            for no_dryer_rec in self.no_dryer_records:
                record = {
                    "滴定機": "",
                    "Marker": no_dryer_rec.get("marker", ""),
                    "PN": no_dryer_rec.get("PN", ""),
                    "凍乾機台": "",
                    "數量": no_dryer_rec.get("數量", ""),
                    "配藥同仁": no_dryer_rec.get("配藥同仁", ""),
                    "日期": no_dryer_rec.get("日期", ""),
                    "RD給藥時間": no_dryer_rec.get("RD給藥時間", ""),
                    "預計滴定時間": "",
                    "預計結束": "",
                    "工單號碼": "",
                    "Lot": no_dryer_rec.get("lot", ""),
                    "備註": no_dryer_rec.get("備註", "凍乾機為空"),
                    "班次": no_dryer_rec.get("班次", "PM"),
                    "record_type": "no_dryer",
                    "has_conflict": False
                }
                all_records.append(record)

        # ========================================
        # 按日期分組
        # ========================================
        from collections import defaultdict
        records_by_date = defaultdict(lambda: {
            "IVEK": [],
            "AM_Port": {},
            "PM_Port": {},
            "no_dryer": []
        })

        for record in all_records:
            date_key = record.get("日期", "")
            if not date_key: continue
            
            machine = str(record.get("滴定機", "")).upper()
            slot = record.get("班次", "AM")
            
            if "IVEK" in machine:
                records_by_date[date_key]["IVEK"].append(record)
            elif "PORT" in machine:
                try:
                    port_num = int(re.search(r'\d+', machine).group())
                    if 1 <= port_num <= 12:
                        if slot == "AM":
                            records_by_date[date_key]["AM_Port"][port_num] = record
                        else:
                            records_by_date[date_key]["PM_Port"][port_num] = record
                except:
                    pass
            elif record.get("record_type") == "no_dryer":
                records_by_date[date_key]["no_dryer"].append(record)
        
        print(f"\n  ✅ 共收集 {len(all_records)} 筆記錄，分為 {len(records_by_date)} 天", file=sys.stderr)

        # ========================================
        # 寫入 Excel
        # ========================================
        wb = Workbook()
        wb.remove(wb.active)
        
        headers = ["滴定機", "Marker", "PN", "凍乾機台", "數量", "配藥同仁", 
                   "日期", "RD給藥時間", "預計滴定時間", "預計結束", 
                   "工單號碼", "Lot", "備註"]
        
        schedule_data_for_db = []
        
        YELLOW_FILL = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
        RED_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        
        for date_key in sorted(records_by_date.keys()):
            day_data = records_by_date[date_key]
            ws = wb.create_sheet(title=date_key)
            
            # 標題
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            current_row = 2
            standard_alignment = Alignment(horizontal="center", vertical="center")
            standard_border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
            
            # 1. IVEK (固定 2 行)
            ivek_records = day_data["IVEK"][:2]
            for i in range(2):
                ws.cell(row=current_row, column=1, value="IVEK")
                if i < len(ivek_records):
                    record = ivek_records[i]
                    for col_idx, header in enumerate(headers, 1):
                        if header == "滴定機": value = "IVEK"
                        else: value = record.get(header, "")
                        
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.alignment = standard_alignment
                        cell.border = standard_border
                        if record.get("has_conflict", False): cell.fill = RED_FILL
                    
                    db_rec = record.copy()
                    db_rec["has_conflict"] = 1 if record.get("has_conflict") else 0
                    schedule_data_for_db.append(db_rec)
                else:
                    for col_idx in range(1, len(headers)+1):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.alignment = standard_alignment
                        cell.border = standard_border
                        if col_idx == 1: cell.value = "IVEK"
                current_row += 1
            
            current_row += 1 # 空行
            
            # 2. AM Port 1-12
            for port_num in range(1, 13):
                ws.cell(row=current_row, column=1, value=f"Port{port_num}")
                if port_num in day_data["AM_Port"]:
                    record = day_data["AM_Port"][port_num]
                    for col_idx, header in enumerate(headers, 1):
                        if header == "滴定機": value = f"Port{port_num}"
                        else: value = record.get(header, "")
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.alignment = standard_alignment
                        cell.border = standard_border
                        if record.get("has_conflict", False): cell.fill = RED_FILL
                    
                    db_rec = record.copy()
                    db_rec["has_conflict"] = 1 if record.get("has_conflict") else 0
                    schedule_data_for_db.append(db_rec)
                else:
                    for col_idx in range(1, len(headers)+1):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.alignment = standard_alignment
                        cell.border = standard_border
                        if col_idx == 1: cell.value = f"Port{port_num}"
                current_row += 1
                
            current_row += 1 # 空行

            # 3. PM Port 1-12
            for port_num in range(1, 13):
                ws.cell(row=current_row, column=1, value=f"Port{port_num}")
                if port_num in day_data["PM_Port"]:
                    record = day_data["PM_Port"][port_num]
                    for col_idx, header in enumerate(headers, 1):
                        if header == "滴定機": value = f"Port{port_num}"
                        else: value = record.get(header, "")
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.alignment = standard_alignment
                        cell.border = standard_border
                        if record.get("has_conflict", False): cell.fill = RED_FILL
                    
                    db_rec = record.copy()
                    db_rec["has_conflict"] = 1 if record.get("has_conflict") else 0
                    schedule_data_for_db.append(db_rec)
                else:
                    for col_idx in range(1, len(headers)+1):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.alignment = standard_alignment
                        cell.border = standard_border
                        if col_idx == 1: cell.value = f"Port{port_num}"
                current_row += 1

            # 4. no_dryer
            if day_data["no_dryer"]:
                current_row += 1
                for record in day_data["no_dryer"]:
                    for col_idx, header in enumerate(headers, 1):
                        value = record.get(header, "")
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.alignment = standard_alignment
                        cell.border = standard_border
                        cell.fill = YELLOW_FILL
                    
                    db_rec = record.copy()
                    db_rec["has_conflict"] = 0
                    schedule_data_for_db.append(db_rec)
                    current_row += 1

            # 調整欄寬
            column_widths = {"A": 12, "B": 20, "C": 15, "D": 12, "E": 10, "F": 12, "G": 12, "H": 12, "I": 12, "J": 12, "K": 15, "L": 15, "M": 20}
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            ws.freeze_panes = "A2"

        # 儲存
        try:
            wb.save(output_path)
            print(f"\n✅ Excel 輸出成功: {output_path}", file=sys.stderr)
            print(f"   共 {len(records_by_date)} 個工作表（固定格式）", file=sys.stderr)
        except Exception as e:
            print(f"❌ Excel 輸出失敗: {e}", file=sys.stderr)
            return None

        # 寫入 DB (保持不變)
        if schedule_data_for_db:
            self._write_to_db(schedule_data_for_db)

        return output_path
    
    def _write_to_db(self, data):
        """
        輔助函式：寫入資料庫
        修正: 統一欄位名稱 (:marker, :Batch) 並補齊 P0 缺失的欄位
        """
        import sqlite3
        from datetime import datetime
        
        # 引用全域變數
        global SCHEDULE_START_DATE
        
        print(f"\n💾 準備寫入資料庫 ({len(data)} 筆)...", file=sys.stderr)
        try:
            if SCHEDULE_START_DATE: 
                week_obj = SCHEDULE_START_DATE
            else: 
                try:
                    first_date = data[0].get("日期", "")
                    week_obj = datetime.strptime(first_date, "%Y-%m-%d")
                except:
                    week_obj = datetime.now()
            
            iso_year, iso_week, _ = week_obj.isocalendar()
            schedule_week = f"{iso_year}_W{iso_week:02d}"
            print(f"🗓️ 計算週次: {schedule_week}", file=sys.stderr)

            # ========================================
            # 🔥 資料正規化：建立符合 SQL 參數的字典列表
            # ========================================
            normalized_data = []
            for row in data:
                # 對應 SQL VALUES (:key) 的名稱
                new_row = {
                    "日期": row.get("日期", ""),
                    "marker": row.get("Marker", ""),     # Map: Marker -> marker
                    "滴定機": row.get("滴定機", ""),
                    "凍乾機台": row.get("凍乾機台", ""),
                    "配藥同仁": row.get("配藥同仁", ""),
                    "RD給藥時間": row.get("RD給藥時間", ""),
                    "預計滴定時間": row.get("預計滴定時間", ""),
                    "預計結束": row.get("預計結束", ""),
                    
                    # P0 任務可能沒有這些欄位，給預設空字串
                    "預冷時間": row.get("預冷時間", ""), 
                    "凍乾時間": row.get("凍乾時間", ""),
                    "收藥時間": row.get("收藥時間", ""),
                    
                    "數量": str(row.get("數量", "")), # 轉字串確保安全
                    "PN": row.get("PN", ""),
                    "工單號碼": row.get("工單號碼", ""),
                    "Batch": row.get("Lot", ""),        # Map: Lot -> Batch
                    "班次": row.get("班次", ""),
                    "record_type": row.get("record_type", ""),
                    "has_conflict": row.get("has_conflict", 0),
                    "schedule_week": schedule_week
                }
                normalized_data.append(new_row)

            db_path = SCHEDULE_DB_PATH
            table_name = f"schedule_{iso_year}"
            
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            
            # 建立表格
            cur.execute(f"""
                CREATE TABLE IF NOT EXISTS {table_name} (
                    日期 TEXT, marker TEXT, 滴定機 TEXT, 凍乾機台 TEXT, 配藥同仁 TEXT,
                    RD給藥時間 TEXT, 預計滴定時間 TEXT, 預計結束 TEXT, 預冷時間 TEXT,
                    凍乾時間 TEXT, 收藥時間 TEXT, 數量 TEXT, PN TEXT, 工單號碼 TEXT,
                    Batch TEXT, 班次 TEXT, record_type TEXT, has_conflict INTEGER,
                    schedule_week TEXT
                )
            """)
            
            # 清除舊資料
            cur.execute(f"DELETE FROM {table_name} WHERE schedule_week = ?", (schedule_week,))
            
            # 插入新資料
            insert_sql = f"""
                INSERT INTO {table_name} (
                    日期, marker, 滴定機, 凍乾機台, 配藥同仁, RD給藥時間, 預計滴定時間, 預計結束,
                    預冷時間, 凍乾時間, 收藥時間, 數量, PN, 工單號碼, Batch, 班次,
                    record_type, has_conflict, schedule_week
                ) VALUES (
                    :日期, :marker, :滴定機, :凍乾機台, :配藥同仁, :RD給藥時間, :預計滴定時間, :預計結束,
                    :預冷時間, :凍乾時間, :收藥時間, :數量, :PN, :工單號碼, :Batch, :班次,
                    :record_type, :has_conflict, :schedule_week
                )
            """
            cur.executemany(insert_sql, normalized_data)
            conn.commit()
            
            # 統計
            cur.execute(f"SELECT COUNT(*) FROM {table_name} WHERE schedule_week = ?", (schedule_week,))
            inserted_count = cur.fetchone()[0]
            
            conn.close()
            print(f"✅ 成功寫入資料到 {table_name} (共 {inserted_count} 筆)", file=sys.stderr)

        except Exception as e:
            print(f"❌ 資料庫寫入失敗: {e}", file=sys.stderr)
            import traceback
            traceback.print_exc(file=sys.stderr)
# ========================================
# Part 8: 主程序
# ========================================

# 定義雙向記錄器類別 (放在 main 前面或 auxiliary tools 區域)
class DualLogger:
    def __init__(self, filepath, original_stream):
        self.file = open(filepath, "w", encoding='utf-8')
        self.console = original_stream

    def write(self, message):
        self.console.write(message)
        self.file.write(message)
        self.file.flush() # 確保立即寫入檔案

    def flush(self):
        self.console.flush()
        self.file.flush()
        
    def close(self):
        self.file.close()
from datetime import datetime, timedelta, time

def simulate_min_production_days(task_queue, constraints_df, start_date):
    """
    模擬計算 P0~P3 所需的最少生產天數 (Port 起始時間 10:30)
    考慮：凍乾機台數限制、禁止收藥時段(03-08)、Port 產能
    """
    
    # ==========================
    # 1. 參數設定
    # ==========================
    START_HOUR = 10      # ✅ 修改: 起始時間 10點
    START_MINUTE = 30    # ✅ 修改: 起始時間 30分
    
    FORBIDDEN_START = 3  # 03:00 禁止收藥開始
    FORBIDDEN_END = 8    # 08:00 禁止收藥結束
    
    DOSING_SPEED = 1500  # 顆/hr
    PREP_TIME = 0.5      # 準備時間 (hr)
    
    # 2. 篩選 P0~P3 任務並排序
    # 排序邏輯：優先級高(0最優先) > 凍乾時間長(先排大石頭) > 數量大
    tasks = [t for t in task_queue if t[0] <= 3]
    
    def get_freeze_duration(pn):
        if pn in constraints_df.index:
            return float(constraints_df.loc[pn].get("凍乾時間", 12.0))
        return 12.0

    tasks.sort(key=lambda x: (x[0], -get_freeze_duration(x[2]), -x[3]))

    print(f"🎯 開始模擬 {len(tasks)} 個 P0-P3 任務的最少工期 (起始時間 {START_HOUR}:{START_MINUTE})...")

    # 3. 初始化資源狀態
    # ✅ 設定全域起始時間為 Start_Date 的 10:30
    base_time = datetime.combine(start_date, time(START_HOUR, START_MINUTE))
    
    # Port 1-12 的可用時間一開始都是 10:30
    port_availability = {i: base_time for i in range(1, 13)} 
    
    # 凍乾機可用時間 (動態建立)
    dryer_availability = {} 

    # 輔助：解析可用凍乾機
    def get_usable_dryers(pn):
        if pn not in constraints_df.index: return []
        val = str(constraints_df.loc[pn].get("可用凍乾機", "")).strip()
        if not val: return []
        return [d.strip() for d in val.split(',') if d.strip()]

    # 4. 開始模擬
    final_completion_time = base_time

    for task in tasks:
        priority, _, pn, qty_short, _, _, marker, bdc_qty = task
        
        # 決定生產量
        prod_qty = bdc_qty if bdc_qty > 0 else qty_short
        if prod_qty <= 0: continue

        # 讀取約束
        num_ports = 2
        freeze_hours = 12.0
        usable_dryers = []
        
        if pn in constraints_df.index:
            c = constraints_df.loc[pn]
            p_val = str(c.get("Port數", "2"))
            num_ports = 2 if p_val.upper() == "IVEK" else int(p_val)
            freeze_hours = float(c.get("凍乾時間", 12.0))
            usable_dryers = get_usable_dryers(pn)
        
        if not usable_dryers:
            print(f"  ⚠️ {pn} 無可用凍乾機設定，假設使用 'DEFAULT_DRYER'")
            usable_dryers = ["DEFAULT_DRYER"]

        # 初始化凍乾機時間
        for d in usable_dryers:
            if d not in dryer_availability:
                dryer_availability[d] = base_time

        # --- 計算耗時 ---
        # 滴定時間 (hr)
        dosing_hours = (prod_qty / (DOSING_SPEED * num_ports)) + PREP_TIME
        total_process_hours = dosing_hours + freeze_hours

        # --- 尋找最早開工時間 (Greedy) ---
        candidate_options = []

        for dryer in usable_dryers:
            dryer_ready_time = dryer_availability[dryer]
            
            # 找最早有空的連續 num_ports 個 Port (強制單數對齊 1, 3, 5...)
            earliest_port_time = datetime.max
            best_p_start = -1
            
            for p_start in range(1, 13, 2): 
                if p_start + num_ports - 1 > 12: continue
                
                # 該組 Port 的釋放時間 = max(各個 Port 的時間)
                ports_in_group = [p_start + k for k in range(num_ports)]
                group_ready_time = max(port_availability[p] for p in ports_in_group)
                
                if group_ready_time < earliest_port_time:
                    earliest_port_time = group_ready_time
                    best_p_start = p_start
            
            if best_p_start == -1: continue 

            # 暫定開工時間 = max(機台好, Port好)
            tentative_start = max(dryer_ready_time, earliest_port_time)
            # 如果暫定開工時間是休假日，就推遲到下一天的工作起始時間
            while is_holiday(tentative_start):
                # 推到隔天
                tentative_start += timedelta(days=1)
                # 重置時間為當天起始 (例如 10:30)
                tentative_start = tentative_start.replace(hour=START_HOUR, minute=START_MINUTE, second=0)
            
            # --- 🔥 檢查禁止收藥時段 (03:00 ~ 08:00) ---
            tentative_harvest = tentative_start + timedelta(hours=total_process_hours)
            th_hour = tentative_harvest.hour
            
            actual_start = tentative_start
            
            # 如果收藥時間落在禁止區間，往後推遲直到 08:01
            if FORBIDDEN_START <= th_hour < FORBIDDEN_END:
                # 設定目標收藥時間為當天 08:01
                target_harvest = tentative_harvest.replace(hour=8, minute=1, second=0, microsecond=0)
                
                # 如果原本是凌晨 4 點，改成早上 8 點，這還是在同一天，直接減
                # 如果原本是 23:00 變成 08:00 (跨日)，timedelta 會自動處理
                if target_harvest < tentative_harvest:
                    # 這種情況理論上不會發生在 th_hour 3~8 之間
                    # 但為了保險，如果是跨日導致的異常，加一天
                    target_harvest += timedelta(days=1)
                
                delay = target_harvest - tentative_harvest
                actual_start = tentative_start + delay
                # print(f"    🕒 (推遲) {pn} 因撞到凌晨收藥，推遲 {delay} 開工")

            candidate_options.append({
                'start': actual_start,
                'dryer': dryer,
                'p_start': best_p_start
            })

        # 選最早開工的方案
        if not candidate_options:
            continue
            
        best_option = min(candidate_options, key=lambda x: x['start'])
        
        # --- 執行資源更新 ---
        start_t = best_option['start']
        dosing_end_t = start_t + timedelta(hours=dosing_hours)
        harvest_t = start_t + timedelta(hours=total_process_hours)
        
        # 更新最後完工時間
        if harvest_t > final_completion_time:
            final_completion_time = harvest_t
            
        # 1. 更新 Port (滴定完即釋放，這是達成一日多次的關鍵)
        for k in range(num_ports):
            p_id = best_option['p_start'] + k
            port_availability[p_id] = dosing_end_t
            
        # 2. 更新凍乾機 (收藥才釋放)
        dryer_availability[best_option['dryer']] = harvest_t

    # 5. 計算總天數
    total_duration = final_completion_time - base_time
    days_needed = total_duration.days
    if total_duration.seconds > 0:
        days_needed += 1
        
    end_date = final_completion_time.date()
    days_span = (end_date - start_date.date()).days + 1
    
    print(f"\n📊 [10:30 起始] 模擬結果: P0-P3 完成時間: {final_completion_time.strftime('%Y-%m-%d %H:%M')}")
    print(f"   總計工期: {days_span} 天")
    
    return days_span, final_completion_time

def main():
    """主程序（前端整合版）"""
    
    # ========================================
    # V9.9.9: 解析命令行參數
    # ========================================
    args = parse_arguments()
    

    # ========================================
    # ✅ 新增: 設定 Log 輸出 (同時顯示於螢幕與寫入檔案)
    # ========================================
    try:
        # 建立 log 資料夾
        log_dir = os.path.join(args.outdir, "log")
        os.makedirs(log_dir, exist_ok=True)
        
        # 建立 log 檔案路徑 (time_log_YYYYMMDD_HHMMSS.txt)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_filename = f"time_log_{timestamp}.txt"
        log_path = os.path.join(log_dir, log_filename)
        
        # 重導向 sys.stderr (因為您的 print 都用 file=sys.stderr)
        # 如果您也有用 sys.stdout，可以依樣畫葫蘆
        sys.stderr = DualLogger(log_path, sys.stderr)
        
        print(f"📝 Log 記錄已啟動: {log_path}", file=sys.stderr)
        
    except Exception as e:
        print(f"⚠️ 無法建立 Log 檔案: {e}")


    print(f"===== [{CURRENT_VERSION}] 排程開始 =====\n", file=sys.stderr)
    
    # 解析起始日期
    try:
        current_year = datetime.now().year
        parts = args.date.split('/')
        month, day = int(parts[0]), int(parts[1])
        start_date = datetime(current_year, month, day)
        
        # 確保是週一
        if start_date.weekday() != 0:
            start_date = start_date - timedelta(days=start_date.weekday())
        # ========================================
        # ✅ 新增：設定全域變數
        # ========================================
        global SCHEDULE_START_DATE
        SCHEDULE_START_DATE = start_date
        
        print(f"排程起始日期（週一）: {start_date.strftime('%Y-%m-%d')}\n", file=sys.stderr)
        
    except Exception as e:
        output = {"ok": False, "message": f"日期解析失敗: {e}"}
        print(json.dumps(output, ensure_ascii=False))
        sys.exit(1)
    
    # ========================================
    # 設定參數
    # ========================================
    
    # 設定休假日
    if args.holidays:
        print("="*70, file=sys.stderr)
        print("🗓️  設定休假日", file=sys.stderr)
        print("="*70, file=sys.stderr)
        setup_holidays_from_args(args.holidays, start_date)
        print(file=sys.stderr)
    
    # 設定批次編號
    if args.batch_numbers:
        print("="*70, file=sys.stderr)
        print("🔢  設定工單批次編號", file=sys.stderr)
        print("="*70, file=sys.stderr)
        setup_batch_start_from_args(args.batch_numbers)
        print(file=sys.stderr)
    
    # 設定休假人員
    if args.vacation_staff:
        print("="*70, file=sys.stderr)
        print("👥  設定休假人員", file=sys.stderr)
        print("="*70, file=sys.stderr)
        setup_vacation_staff_from_args(args.vacation_staff, start_date)
        print(file=sys.stderr)
    
    try:
        # ========================================
        # 載入資料
        # ========================================
        all_data = load_all_data(args.need)
        
        if not all_data or all_data["demand"] is None:
            output = {"ok": False, "message": "資料載入錯誤"}
            print(json.dumps(output, ensure_ascii=False))
            sys.exit(1)

        # ========================================
        # 執行排程
        # ========================================
        scheduler = Scheduler(all_data)
        # 🔥🔥🔥 【插入點開始】 🔥🔥🔥
        # 在這裡呼叫模擬功能
        print("\n" + "="*60, file=sys.stderr)
        print("🔮 產能預測模擬 (P0~P3 最少工期計算)", file=sys.stderr)
        print("="*60, file=sys.stderr)

        # 呼叫函式
        min_required_days, est_finish_time = simulate_min_production_days(
            scheduler.task_queue,   # 任務清單
            scheduler.constraints,  # 限制條件(查凍乾機/Port數)
            SCHEDULE_START_DATE     # 排程起始日期 (全域變數)
        )

        print(f"\n💡 系統建議：", file=sys.stderr)
        print(f"   為了滿足 P0~P3 核心需求，", file=sys.stderr)
        print(f"   考量凍乾機數量與 03-08 禁止收藥限制，", file=sys.stderr)
        print(f"   理論最少需要工期: 【{min_required_days} 天】", file=sys.stderr)
        print(f"   目前設定排程天數: 【{SCHEDULING_DAYS} 天】", file=sys.stderr)

        if min_required_days > SCHEDULING_DAYS:
            print(f"   ⚠️ 警告：設定天數可能不足！建議延長至 {min_required_days} 天或啟用週六排程。", file=sys.stderr)
            # 檢查是否允許延長 (雖然模擬建議延長，但您可以設定上限，例如最多 7 天)
            MAX_ALLOWED_DAYS = 7 
            
            if min_required_days <= MAX_ALLOWED_DAYS:
                scheduler.days_to_schedule = min_required_days
                print(f"✅ 已將排程天數自動延長至 {scheduler.days_to_schedule} 天", file=sys.stderr)
            else:
                scheduler.days_to_schedule = MAX_ALLOWED_DAYS
                print(f"⚠️ 模擬天數 ({min_required_days}) 超過上限，強制設定為 {MAX_ALLOWED_DAYS} 天", file=sys.stderr)
        else:
            print(f"✅ 設定天數充足，無需調整。", file=sys.stderr)
        
        print("="*60 + "\n", file=sys.stderr)
        # 🔥🔥🔥 【插入點結束】 🔥🔥🔥
        scheduler.run()
        
        # ========================================
        # 輸出結果
        # ========================================
        if hasattr(scheduler, 'schedule_df') and not scheduler.schedule_df.empty:
            
            if args.dry_run:
                # Dry Run 模式：輸出預覽 JSON
                preview_data = scheduler._generate_dry_run_preview()
                output = {"ok": True, "preview": preview_data}
                print(json.dumps(output, ensure_ascii=False))
            else:
                # 正常模式：儲存 Excel
                # ✅ 修正 3: 接收 save_records_to_excel 返回的 output_path
                output_path = scheduler.save_records_to_excel(args.outdir)
                
                output = {"ok": True, "outPath": output_path}
                print(json.dumps(output, ensure_ascii=False))
                
            print(f"\n===== [{CURRENT_VERSION}] 排程完成 =====", file=sys.stderr)
        else:
            output = {"ok": False, "message": "排程已終止，未產生結果"}
            print(json.dumps(output, ensure_ascii=False))
            sys.exit(1)

    except Exception as e:
        output = {"ok": False, "message": f"排程失敗: {str(e)}"}
        print(json.dumps(output, ensure_ascii=False))
        traceback.print_exc(file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()