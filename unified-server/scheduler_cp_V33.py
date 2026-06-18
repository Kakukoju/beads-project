import pandas as pd
import sqlite3
import collections
import argparse
import sys
import os
import math
import openpyxl
import time
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from ortools.sat.python import cp_model

# ==============================================================================
# 1. 基礎設定 (Configuration)
# ==============================================================================

DB_PATH = r"\\fls341\MBBU_FAB\MB_PD\生管自動化\滴定\資料庫\beads_sync.db"
START_HOUR = 8
MINS_PER_GRID = 30
DEFAULT_GRIDS = 35  # 08:00 ~ 01:30
LUNCH_BREAK_GRIDS = [8, 9] # 12:00, 12:30
OUTPUT_FILENAME = "滴定排程結果_V33.xlsx"

# ==============================================================================
# 2. 資料庫與解析工具
# ==============================================================================

def load_data_from_db(db_path):
    if not os.path.exists(db_path):
        print(f"❌ 錯誤: 找不到資料庫檔案: {db_path}")
        return None, None
    try:
        conn = sqlite3.connect(db_path)
        # 讀取規則書
        df_rules = pd.read_sql_query("SELECT * FROM 配藥限制", conn)
        # 讀取最新日期的需求
        df_demand = pd.read_sql_query("""
            SELECT * FROM BeadNeed 
            WHERE date = (SELECT MAX(date) FROM BeadNeed)
        """, conn)
        conn.close()
        return df_rules, df_demand
    except Exception as e:
        print(f"❌ 資料庫讀取失敗: {e}")
        return None, None

def parse_allowed_grids(time_str, grids_per_day):
    full_day = [g for g in range(grids_per_day) if g not in LUNCH_BREAK_GRIDS]
    if pd.isna(time_str) or str(time_str).strip() == '' or str(time_str).lower() == 'nan':
        return full_day
    allowed = set()
    parts = str(time_str).replace('"', '').replace("'", "").split(',')
    has_specific = False
    for p in parts:
        try:
            if ':' not in p: continue
            h, m = map(int, p.strip().split(':'))
            start_min = (h * 60 + m) - (START_HOUR * 60)
            start_g = math.ceil(max(0, start_min) / MINS_PER_GRID)
            for g in range(start_g, grids_per_day):
                if g not in LUNCH_BREAK_GRIDS: allowed.add(g)
            has_specific = True
        except: pass
    return sorted(list(allowed)) if has_specific else full_day

def parse_qty_options(qty_str):
    if pd.isna(qty_str): return [0.0]
    try:
        parts = str(qty_str).lower().replace(',', '').split('or')
        opts = [float(p.strip()) for p in parts if float(p.strip()) > 0]
        return sorted(opts) if opts else [0.0]
    except: return [0.0]

# ==============================================================================
# 3. 資料預處理邏輯
# ==============================================================================

def prepare_rulebook_v33(df, grids_per_day, expand_window=False):
    jobs = {} 
    reagent_groups = {} 
    all_staff_set = set() 
    for _, row in df.iterrows():
        raw_pn = row.get('PN')
        if pd.isna(raw_pn): continue
        pn = str(raw_pn).strip().replace('.0', '')
        reagent = str(row.get('Reagent', '')).strip()
        if reagent.lower() == 'nan': reagent = ''
        
        port_val = str(row.get('Port數', '2')).strip().upper()
        num_ports = 0 if port_val == 'IVEK' else (int(port_val) if port_val.isdigit() else 2)
        qty_options = parse_qty_options(row.get('數量', '0'))
        
        try:
            f_hrs = float(row.get('凍乾時間', 12.0))
        except: f_hrs = 12.0
        
        d_str = str(row.get('可用凍乾機', '')).strip().replace('"', '').replace("'", "")
        compatible_dryers = [d.strip() for d in d_str.split(',') if d.strip()]
        
        compatible_staff = []
        for col in ['配藥人-1', '配藥人-2', '配藥人-3']:
            s = str(row.get(col, '')).strip()
            if s and s.lower() != 'nan':
                compatible_staff.append(s)
                all_staff_set.add(s)
        
        time_str = row.get('交藥時間', '')
        if expand_window and '11:00' in str(time_str):
            time_str = str(time_str).replace('11:00', '10:00')
        
        allowed_grids = parse_allowed_grids(time_str, grids_per_day)
        split_prod = str(row.get('U,D 劑分開生產排程', '')).strip().upper() in ['V', 'TRUE', '1']
        note = str(row.get('備註', '')).strip()
        
        job = {
            'pn': pn, 'name': str(row.get('Name', '')), 'reagent': reagent,
            'num_ports': num_ports, 'qty_options': qty_options,
            'duration_freeze': int(f_hrs * 60), 'compatible_dryers': compatible_dryers,
            'compatible_staff': compatible_staff, 'allowed_grids': allowed_grids,
            'split_prod': split_prod, 'no_friday': '不能放週五滴定' in note
        }
        jobs[pn] = job
        if reagent: reagent_groups.setdefault(reagent, []).append(pn)
    return jobs, reagent_groups, sorted(list(all_staff_set))

def calculate_priority_tasks_v33(df_demand, rulebook_dict, reagent_groups):
    pn_status = {}
    for _, row in df_demand.iterrows():
        pn = str(row.get('pn', '')).strip().replace('.0', '')
        if pn not in rulebook_dict: continue
        def pnum(v):
            try: return float(str(v).replace(',', ''))
            except: return 0.0
        stock = pnum(row.get('stock_unstock', 0))
        w1, w2, w3 = pnum(row.get('w1', 0)), pnum(row.get('w2', 0)), pnum(row.get('w3', 0))
        prio, shortage, deadline = 0, 0, None
        tmp = stock
        if tmp - w1 < 0:
            shortage = w1 - tmp; prio = 1; deadline = 3; tmp += shortage
        elif tmp - w1 - w2 < 0:
            shortage = (w1 + w2) - stock; prio = 2; tmp += shortage
        elif tmp - w1 - w2 - w3 < 0:
            shortage = (w1 + w2 + w3) - stock; prio = 3
        pn_status[pn] = {'shortage': shortage, 'priority': prio, 'deadline': deadline}

    final_tasks = []
    processed = set()
    for reagent, pns in reagent_groups.items():
        actives = [p for p in pns if p in pn_status]
        if not actives: continue
        if any(pn_status[p]['priority'] > 0 for p in actives):
            max_p = max(pn_status[p]['priority'] for p in actives)
            for p in actives:
                st = pn_status[p]
                final_tasks.append({
                    'pn': p, 'type': f"P{max_p}", 'shortage': st['shortage'] if st['priority'] > 0 else 0,
                    'deadline_days': st['deadline'], 'priority': max_p
                })
                processed.add(p)
    for pn, st in pn_status.items():
        if pn not in processed and st['priority'] > 0:
            final_tasks.append({'pn': pn, 'type': f"P{st['priority']}", 'shortage': st['shortage'], 'deadline_days': st['deadline'], 'priority': st['priority']})
    return final_tasks

def prepare_jobs(tasks, rulebook, groups):
    jobs = []
    for t in tasks:
        pn = t['pn']
        rule = rulebook[pn]
        target = t['shortage']
        qty = rule['qty_options'][-1]
        for opt in rule['qty_options']:
            if opt >= target:
                qty = opt
                break
        ports = rule['num_ports']
        dur_min = int(math.ceil((qty / ports / 1500)*60 + 30)) if ports > 0 else 60
        grid_dur = math.ceil(dur_min / 30)
        jobs.append({
            'id': f"{t['type']}_{pn}", 'pn': pn, 'name': rule['name'], 'qty': qty, 'num_ports': ports, 
            'duration_dosing': dur_min, 'grid_duration': grid_dur, 'duration_freeze': rule['duration_freeze'],
            'compatible_dryers': rule['compatible_dryers'], 'compatible_staff': rule['compatible_staff'],
            'allowed_grids': rule['allowed_grids'], 'no_friday': rule['no_friday'], 'split_prod': rule['split_prod'],
            'reagent': rule['reagent'], 'deadline_days': t['deadline_days']
        })
    # 同 reagent 組同步凍乾時間
    job_map = {j['pn']: j for j in jobs}
    for r, pns in groups.items():
        sub = [p for p in pns if p in job_map]
        if len(sub) > 1 and not any(job_map[p]['split_prod'] for p in sub):
            max_f = max(job_map[p]['duration_freeze'] for p in sub)
            for p in sub: job_map[p]['duration_freeze'] = max_f
    return jobs

# ==============================================================================
# 4. CP-SAT 核心求解器
# ==============================================================================

def run_solver(jobs, reagent_groups, all_staff, horizon_days, relax_ports, grids_per_day):
    model = cp_model.CpModel()
    horizon_grids = horizon_days * grids_per_day
    all_dryers = sorted(list(set(d for j in jobs for d in j['compatible_dryers'])))
    dryer_intervals, staff_intervals = collections.defaultdict(list), collections.defaultdict(list)
    job_vars, pn_to_vars, rects = {}, {}, []

    for job in jobs:
        s_grid = model.NewIntVar(0, horizon_grids, f"s_{job['id']}")
        e_grid = model.NewIntVar(0, horizon_grids, f"e_{job['id']}")
        day = model.NewIntVar(0, horizon_days - 1, f"d_{job['id']}")
        d_idx = model.NewIntVar(0, len(all_dryers) if all_dryers else 0, f"dr_{job['id']}")
        st_idx = model.NewIntVar(0, len(all_staff) if all_staff else 0, f"st_{job['id']}")
        p_start = model.NewIntVar(0, 11, f"p_{job['id']}")
        s_frz, e_frz = model.NewIntVar(0, horizon_grids + 100, f"sf_{job['id']}"), model.NewIntVar(0, horizon_grids + 200, f"ef_{job['id']}")

        job_vars[job['id']] = {'s': s_grid, 'e': e_grid, 'day': day, 'dryer': d_idx, 'staff': st_idx, 'port': p_start, 's_frz': s_frz, 'e_frz': e_frz}
        pn_to_vars[job['pn']] = job_vars[job['id']]

        model.Add(e_grid == s_grid + job['grid_duration'])
        model.AddDivisionEquality(day, s_grid, grids_per_day)
        
        if job['num_ports'] > 0:
            x_int = model.NewFixedSizeIntervalVar(s_grid, job['grid_duration'], '')
            y_int = model.NewFixedSizeIntervalVar(p_start, job['num_ports'], '')
            rects.append([x_int, y_int])
            model.Add(p_start + job['num_ports'] <= 12)
        
        s_in_day = model.NewIntVar(0, grids_per_day, '')
        model.AddModuloEquality(s_in_day, s_grid, grids_per_day)
        model.Add(s_in_day + job['grid_duration'] <= grids_per_day)
        if job['allowed_grids']:
            valid = [d * grids_per_day + g for d in range(horizon_days) for g in job['allowed_grids']]
            model.AddLinearExpressionInDomain(s_grid, cp_model.Domain.FromValues(valid))

        model.Add(s_frz >= e_grid)
        fr_grids = math.ceil((job['duration_freeze'] + 120) / 30)
        model.Add(e_frz == s_frz + fr_grids)

        for idx, dname in enumerate(all_dryers):
            if dname in job['compatible_dryers']:
                is_used = model.NewBoolVar(f"u_d_{job['id']}_{dname}")
                model.Add(d_idx == idx).OnlyEnforceIf(is_used)
                occ = model.NewOptionalFixedSizeIntervalVar(s_frz, fr_grids, is_used, '')
                dryer_intervals[dname].append({'int': occ, 'bool': is_used, 'job': job['id']})
        
        for idx, sname in enumerate(all_staff):
            if sname in job['compatible_staff']:
                is_used = model.NewBoolVar(f"u_s_{job['id']}_{sname}")
                model.Add(st_idx == idx).OnlyEnforceIf(is_used)
                gap = model.NewOptionalFixedSizeIntervalVar(s_grid, 5, is_used, '')
                staff_intervals[idx].append({'int': gap, 'job': job['id']})

    if rects: model.AddNoOverlap2D([r[0] for r in rects], [r[1] for r in rects])

    # 配對與 Port 邏輯
    slaves, partners, paired_secondary_jobs = set(), {}, set()
    for r, pns in reagent_groups.items():
        sub = [p for p in pns if p in pn_to_vars]
        if len(sub) < 2: continue
        if any('220' in p for p in sub) and any('221' in p for p in sub):
            pA, pB = next(p for p in sub if '220' in p), next(p for p in sub if '221' in p)
            model.Add(pn_to_vars[pA]['s'] == pn_to_vars[pB]['s'])
            model.Add(pn_to_vars[pA]['dryer'] != pn_to_vars[pB]['dryer'])
            if not relax_ports:
                model.AddModuloEquality(0, pn_to_vars[pA]['port'], 2)
                model.Add(pn_to_vars[pB]['port'] == pn_to_vars[pA]['port'] + 2)
            idA, idB = next(j['id'] for j in jobs if j['pn']==pA), next(j['id'] for j in jobs if j['pn']==pB)
            partners[idA], partners[idB], slaves.add(idB)
            continue
        pA, pB = sub[0], sub[1]
        jA, jB = next(j for j in jobs if j['pn']==pA), next(j for j in jobs if j['pn']==pB)
        if not (jA['split_prod'] or jB['split_prod']):
            model.Add(pn_to_vars[pA]['s'] == pn_to_vars[pB]['s'])
            model.Add(pn_to_vars[pA]['dryer'] == pn_to_vars[pB]['dryer'])
            if not relax_ports:
                model.AddModuloEquality(0, pn_to_vars[pA]['port'], 2)
                model.Add(pn_to_vars[pB]['port'] == pn_to_vars[pA]['port'] + jA['num_ports'])
                if jA['num_ports'] % 2 != 0: paired_secondary_jobs.add(jB['id'])
            slaves.add(jB['id']); partners[jA['id']] = jB['id']; partners[jB['id']] = jA['id']
        else: model.Add(pn_to_vars[pA]['day'] != pn_to_vars[pB]['day'])

    # 設置時間 (Setup) 與 Port 偶數限制
    global_setup = []
    for job in jobs:
        if job['id'] not in slaves and job['reagent']:
            setup_int = model.NewFixedSizeIntervalVar(job_vars[job['id']]['s'], 1, '')
            global_setup.append(setup_int)
        if not relax_ports and job['num_ports'] > 0 and job['id'] not in paired_secondary_jobs:
            model.AddModuloEquality(0, pn_to_vars[job['pn']]['port'], 2)
    if global_setup: model.AddNoOverlap(global_setup)

    for dname, items in dryer_intervals.items():
        for i in range(len(items)):
            for k in range(i+1, len(items)):
                if partners.get(items[i]['job']) != items[k]['job']: model.AddNoOverlap([items[i]['int'], items[k]['int']])

    makespan = model.NewIntVar(0, horizon_grids, 'makespan')
    model.AddMaxEquality(makespan, [v['e'] for v in job_vars.values()])
    model.Minimize(makespan)
    
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 30.0
    status = solver.Solve(model)
    if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
        res = []
        for jid, vars in job_vars.items():
            j = next(x for x in jobs if x['id'] == jid).copy()
            j.update({'grid_start': solver.Value(vars['s']) % grids_per_day, 'day': solver.Value(vars['day']), 'start': solver.Value(vars['s']),
                      'assigned_port_idx': solver.Value(vars['port']), 'assigned_dryer': all_dryers[solver.Value(vars['dryer'])],
                      'assigned_staff': all_staff[solver.Value(vars['staff'])], 'freeze_end_abs': solver.Value(vars['e_frz'])})
            res.append(j)
        return res
    return None

# ==============================================================================
# 5. Excel 輸出 (完整樣式)
# ==============================================================================

def export_excel_v33(results, grids_per_day, filename=OUTPUT_FILENAME):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    by_day = collections.defaultdict(list)
    for j in results: by_day[j['day']].append(j)
    
    fill_h = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    colors = ["E6B8AF", "F4CCCC", "FCE5CD", "FFF2CC", "D9EAD3", "D0E0E3", "C9DAF8", "CFE2F3", "D9D2E9", "EAD1DC"]
    pn_colors = {}
    
    for d in sorted(by_day.keys()):
        ws = wb.create_sheet(f"Day {d+1}")
        headers = ["時間", "IVEK 1", "IVEK 2"] + [f"Port {i}" for i in range(1, 13)]
        for c, txt in enumerate(headers, 1):
            cell = ws.cell(1, c, txt)
            cell.fill, cell.border, cell.alignment = fill_h, border, Alignment('center', 'center')
            ws.column_dimensions[get_column_letter(c)].width = 16 if c > 1 else 10
            
        for i in range(grids_per_day):
            m = START_HOUR * 60 + i * 30
            txt = f"{m//60:02d}:{m%60:02d}" if m < 1440 else f"隔{(m-1440)//60:02d}:{(m-1440)%60:02d}"
            cell = ws.cell(i+2, 1, txt)
            cell.border, cell.alignment = border, Alignment('center', 'center')
            
        for j in by_day[d]:
            if j['pn'] not in pn_colors: pn_colors[j['pn']] = colors[len(pn_colors)%len(colors)]
            c_fill = PatternFill(start_color=pn_colors[j['pn']], end_color=pn_colors[j['pn']], fill_type="solid")
            r_s, c_s = j['grid_start'] + 2, (4 + j['assigned_port_idx'] if j['num_ports'] > 0 else 2)
            r_e, c_e = r_s + j['grid_duration'] - 1, c_s + max(0, j['num_ports'] - 1)
            
            f_m = (START_HOUR * 60) + (j['freeze_end_abs'] * 30)
            f_txt = f"Frz-{j['assigned_dryer']}\nEnd {f_m%1440//60:02d}:{f_m%60:02d}" + (f"(+{f_m//1440 - j['day']}d)" if f_m//1440 > j['day'] else "")
            txt = f"{j['pn']}\n{j['name']}\n{int(j['qty']/max(1, j['num_ports']))}/port\n{f_txt}"
            
            cell = ws.cell(r_s, c_s, txt)
            cell.fill, cell.border, cell.alignment = c_fill, border, Alignment('center', 'center', wrap_text=True)
            if f_m // 1440 > j['day'] and (f_m % 1440) > 90: cell.font = Font(color="FF0000", bold=True)
            if r_e > r_s or c_e > c_s:
                try: ws.merge_cells(start_row=r_s, start_column=c_s, end_row=r_e, end_column=c_e)
                except: pass
            for r in range(r_s, r_e+1):
                for c in range(c_s, c_e+1): ws.cell(r, c).border = border
    wb.save(filename)
    print(f"✅ Excel 報表已存至: {filename}")

# ==============================================================================
# 6. 主程式 (Strategy & Optimization)
# ==============================================================================

def main():
    print("🚀 啟動 V33 排程系統...")
    df_rules, df_demand = load_data_from_db(DB_PATH)
    if df_rules is None or df_demand is None: return
    try:
        rules, groups, staff = prepare_rulebook_v33(df_rules, DEFAULT_GRIDS)
        tasks = calculate_priority_tasks_v33(df_demand, rules, groups)
        if not tasks: return print("✅ 無任務。")
        jobs = prepare_jobs(tasks, rules, groups)
        
        final_res, final_grids = None, DEFAULT_GRIDS
        for days in range(3, 8):
            res = run_solver(jobs, groups, staff, days, False, DEFAULT_GRIDS)
            if res:
                max_d = max(j['day'] for j in res)
                if sum(1 for j in res if j['day'] == max_d) == 1 and max_d > 0:
                    print(f"⚠️ Day {max_d+1} 只有 1 個任務，嘗試壓縮...")
                    rules_opt, groups_opt, staff_opt = prepare_rulebook_v33(df_rules, 36, True)
                    jobs_opt = prepare_jobs(tasks, rules_opt, groups_opt)
                    res_opt = run_solver(jobs_opt, groups_opt, staff_opt, days, False, 36)
                    if res_opt and max(j['day'] for j in res_opt) < max_d:
                        final_res, final_grids = res_opt, 36
                        break
                final_res = res
                break
        if final_res: export_excel_v33(final_res, final_grids)
        else: print("❌ 排程失敗。")
    except Exception as e:
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()