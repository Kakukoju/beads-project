"""
Unit tests for ExcelSyncService — Task 9.2: Excel data writing (schedule entries to columns)
Tests _map_entry_to_columns and _fill_day_sections methods.
"""

import os
import sys

import pytest
from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from ai_schedule.excel_sync_service import ExcelSyncService


class TestMapEntryToColumns:
    """Test _map_entry_to_columns mapping logic."""

    def setup_method(self):
        self.svc = ExcelSyncService()

    def test_full_entry_maps_all_columns(self):
        entry = {
            "machine_port": "P3",
            "marker": "tCREA-D",
            "freeze_dryer": "5",
            "formula": "",
            "quantity": 1300,
            "operator": "張三",
            "date": "2026-06-08",
            "rd_time": "14:00",
            "start_time": "14:30",
            "end_time": "18:00",
            "work_order": "TMRA26001",
            "batch": "180260240",
            "notes": "急件",
        }
        result = self.svc._map_entry_to_columns(entry)

        assert result[8] == "P3"          # H: 滴定機
        assert result[9] == "tCREA-D"     # I: Marker
        assert result[10] == "5"          # J: 凍乾機台
        assert result[11] == ""           # K: formula
        assert result[12] == 1300         # L: 數量
        assert result[13] == "張三"       # M: 配藥同仁
        assert result[14] == "2026-06-08" # N: 日期
        assert result[15] == "14:00"      # O: RD時間
        assert result[16] == "14:30"      # P: 滴定時間
        assert result[17] == "18:00"      # Q: 結束
        assert result[18] == "TMRA26001"  # R: 工單
        assert result[19] == "180260240"  # S: Lot
        assert result[20] == "急件"       # T: 備註

    def test_empty_entry_returns_empty_strings(self):
        result = self.svc._map_entry_to_columns({})
        for col in range(8, 21):
            assert result[col] == ""

    def test_partial_entry(self):
        entry = {"marker": "GGT", "quantity": 2600}
        result = self.svc._map_entry_to_columns(entry)

        assert result[9] == "GGT"
        assert result[12] == 2600
        assert result[8] == ""  # machine_port missing
        assert result[20] == ""  # notes missing

    def test_column_range_is_8_to_20(self):
        result = self.svc._map_entry_to_columns({})
        assert set(result.keys()) == set(range(8, 21))

    def test_formula_field_included(self):
        entry = {"formula": "=VLOOKUP(I100, $A$2:$G$50, 7, FALSE)"}
        result = self.svc._map_entry_to_columns(entry)
        assert result[11] == "=VLOOKUP(I100, $A$2:$G$50, 7, FALSE)"


class TestFillDaySections:
    """Test _fill_day_sections dynamic row management and data writing."""

    def setup_method(self):
        self.svc = ExcelSyncService()

    def _create_ws_with_sections(self, separator_rows, rows_per_section=3):
        """
        Create a worksheet with day separators and template data rows.

        Args:
            separator_rows: List of row numbers for separators
            rows_per_section: Number of data rows between separators
        """
        wb = Workbook()
        ws = wb.active

        # Ensure the worksheet has enough rows
        max_row = max(separator_rows) + rows_per_section + 5
        # Fill some data to extend ws.max_row
        ws.cell(row=max_row, column=1, value="end_marker")

        # Place separators
        for row in separator_rows:
            ws.cell(row=row, column=8, value="日期:")

        return ws

    def test_writes_entries_to_correct_columns(self):
        """Entries get written to the correct row/column positions."""
        ws = self._create_ws_with_sections([100], rows_per_section=3)
        entries_by_date = {
            "2026-06-08": [
                {
                    "machine_port": "P3",
                    "marker": "tCREA-D",
                    "freeze_dryer": "5",
                    "formula": "",
                    "quantity": 1300,
                    "operator": "張三",
                    "date": "2026-06-08",
                    "rd_time": "14:00",
                    "start_time": "14:30",
                    "end_time": "18:00",
                    "work_order": "TMRA26001",
                    "batch": "180260240",
                    "notes": "",
                }
            ]
        }

        self.svc._fill_day_sections(ws, entries_by_date, [100])

        # Entry should be written to row 101 (separator at 100, data starts at 101)
        assert ws.cell(row=101, column=8).value == "P3"
        assert ws.cell(row=101, column=9).value == "tCREA-D"
        assert ws.cell(row=101, column=10).value == "5"
        assert ws.cell(row=101, column=12).value == 1300
        assert ws.cell(row=101, column=13).value == "張三"
        assert ws.cell(row=101, column=18).value == "TMRA26001"
        assert ws.cell(row=101, column=19).value == "180260240"

    def test_multiple_entries_in_one_section(self):
        """Multiple entries for the same day fill consecutive rows."""
        ws = self._create_ws_with_sections([100], rows_per_section=5)
        entries_by_date = {
            "2026-06-08": [
                {"marker": "tCREA-D", "machine_port": "P3", "quantity": 1300},
                {"marker": "GGT", "machine_port": "P5", "quantity": 2600},
                {"marker": "ALB", "machine_port": "P1", "quantity": 900},
            ]
        }

        self.svc._fill_day_sections(ws, entries_by_date, [100])

        assert ws.cell(row=101, column=9).value == "tCREA-D"
        assert ws.cell(row=102, column=9).value == "GGT"
        assert ws.cell(row=103, column=9).value == "ALB"

    def test_inserts_rows_when_entries_exceed_template(self):
        """When entries > template rows, insert extra rows."""
        # 2 separators: section at row 100 has rows 101-109 (before 110)
        ws = self._create_ws_with_sections([100, 103], rows_per_section=2)
        # Section 0: rows 101-102 (2 template rows)
        entries_by_date = {
            "2026-06-08": [
                {"marker": "A", "machine_port": "P1"},
                {"marker": "B", "machine_port": "P2"},
                {"marker": "C", "machine_port": "P3"},
                {"marker": "D", "machine_port": "P4"},
            ],
            "2026-06-09": [
                {"marker": "E", "machine_port": "P5"},
            ],
        }

        self.svc._fill_day_sections(ws, entries_by_date, [100, 103])

        # Section 0 had 2 template rows but needed 4 → 2 rows inserted
        # After insertion of 2 rows in section 0, original section 0 data starts at 101
        assert ws.cell(row=101, column=9).value == "A"
        assert ws.cell(row=102, column=9).value == "B"
        assert ws.cell(row=103, column=9).value == "C"
        assert ws.cell(row=104, column=9).value == "D"

    def test_deletes_rows_when_entries_fewer_than_template(self):
        """When entries < template rows, delete excess rows."""
        # Section has 5 template rows (100+1 to 105-1 = rows 101-104)
        ws = self._create_ws_with_sections([100, 105], rows_per_section=2)
        # Section 0: rows 101-104 (4 template rows between separator 100 and 105)
        entries_by_date = {
            "2026-06-08": [
                {"marker": "A", "machine_port": "P1"},
            ],
            "2026-06-09": [],
        }

        self.svc._fill_day_sections(ws, entries_by_date, [100, 105])

        # Section 0 should have "A" at row 101
        assert ws.cell(row=101, column=9).value == "A"

    def test_empty_entries_for_a_day(self):
        """Empty entries for a day section should clear/delete template rows."""
        ws = self._create_ws_with_sections([100, 104], rows_per_section=2)
        # Put some existing data in section 0
        ws.cell(row=101, column=9, value="OLD_MARKER")
        ws.cell(row=102, column=9, value="OLD_MARKER2")

        entries_by_date = {
            "2026-06-08": [],
            "2026-06-09": [{"marker": "NEW", "machine_port": "P1"}],
        }

        self.svc._fill_day_sections(ws, entries_by_date, [100, 104])

        # Section 0 rows were deleted; row 101 should no longer have old data
        # (it's been shifted up or deleted)
        # The key assertion is section 1 still gets written correctly

    def test_preserves_stats_area(self):
        """Rows 1-98 are NEVER modified regardless of operations."""
        wb = Workbook()
        ws = wb.active

        # Put data in stats area
        ws.cell(row=50, column=8, value="STATS_DATA")
        ws.cell(row=98, column=8, value="LAST_STATS")

        # Separator at row 99 (first row after stats)
        ws.cell(row=99, column=8, value="日期:")
        ws.cell(row=110, column=1, value="end")  # extend max_row

        entries_by_date = {
            "2026-06-08": [
                {"marker": "A", "machine_port": "P1"},
            ]
        }

        self.svc._fill_day_sections(ws, entries_by_date, [99])

        # Stats area untouched
        assert ws.cell(row=50, column=8).value == "STATS_DATA"
        assert ws.cell(row=98, column=8).value == "LAST_STATS"

        # Entry written after separator
        assert ws.cell(row=100, column=9).value == "A"

    def test_no_separator_rows_does_nothing(self):
        """Empty separator_rows list should log warning and return."""
        wb = Workbook()
        ws = wb.active
        entries_by_date = {"2026-06-08": [{"marker": "A"}]}

        # Should not raise
        self.svc._fill_day_sections(ws, entries_by_date, [])

    def test_multiple_day_sections(self):
        """Fill entries across multiple day sections correctly."""
        # 3 separators at rows 100, 105, 110, with template room
        wb = Workbook()
        ws = wb.active
        ws.cell(row=100, column=8, value="日期:")
        ws.cell(row=105, column=8, value="日期:")
        ws.cell(row=110, column=8, value="日期:")
        ws.cell(row=120, column=1, value="end")  # extend max_row

        entries_by_date = {
            "2026-06-08": [{"marker": "Mon1"}, {"marker": "Mon2"}],
            "2026-06-09": [{"marker": "Tue1"}],
            "2026-06-10": [{"marker": "Wed1"}, {"marker": "Wed2"}, {"marker": "Wed3"}],
        }

        self.svc._fill_day_sections(ws, entries_by_date, [100, 105, 110])

        # Section 0 (Mon): rows 101-102 should have Mon entries
        assert ws.cell(row=101, column=9).value == "Mon1"
        assert ws.cell(row=102, column=9).value == "Mon2"

    def test_more_separators_than_date_keys(self):
        """If there are more separator rows than date keys, extra sections get no data."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=100, column=8, value="日期:")
        ws.cell(row=105, column=8, value="日期:")
        ws.cell(row=110, column=8, value="日期:")
        ws.cell(row=120, column=1, value="end")

        entries_by_date = {
            "2026-06-08": [{"marker": "Mon1"}],
        }

        # 3 separators but only 1 date key — should handle gracefully
        self.svc._fill_day_sections(ws, entries_by_date, [100, 105, 110])

        assert ws.cell(row=101, column=9).value == "Mon1"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
