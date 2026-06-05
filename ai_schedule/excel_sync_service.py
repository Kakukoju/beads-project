"""
Excel Sync Service — 獨立的 Excel 同步服務，與 RDS 交易解耦

Handles:
- Template sheet finding (closest week)
- Sheet copying and renaming
- Day date filling (ISO week dates + day names)
- Schedule data writing (Task 9.2)
- Full sync orchestration (Task 9.3)
"""

import re
import logging
from copy import copy
from dataclasses import dataclass
from datetime import date, timedelta
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# Pattern to match sheet names like "26排程表-wXX" possibly with " (N)" suffix
SHEET_NAME_PATTERN = re.compile(r'^26排程表-w(\d{1,2})(?:\s*\(\d+\))?$')


@dataclass
class SyncResult:
    """Result of an Excel sync operation.

    Attributes:
        status: 'success' or 'failed'
        error: Error message if status is 'failed', None otherwise.
    """
    status: str
    error: Optional[str] = None


class ExcelSyncService:
    """獨立的 Excel 同步服務，與 RDS 交易解耦"""

    EXCEL_PATH = "/home/ubuntu/beads-project/excelData/Excel_data/排程表week_2026.xlsm"
    STATS_AREA_END_ROW = 98
    DAY_SEPARATOR_COL = "H"  # Column 8
    DAY_SEPARATOR_VALUE = "日期:"

    # Chinese day names for Monday through Friday
    DAY_NAMES = ["一", "二", "三", "四", "五"]

    def __init__(self, excel_path: Optional[str] = None):
        """
        Initialize ExcelSyncService.

        Args:
            excel_path: Override path for the Excel file. Defaults to EXCEL_PATH.
        """
        if excel_path is not None:
            self.excel_path = excel_path
        else:
            self.excel_path = self.EXCEL_PATH

    def _find_closest_template_sheet(self, wb: Workbook, target_week: int) -> Optional[str]:
        """
        Find the sheet closest to target_week matching the '26排程表-wXX' pattern.

        Ignores sheets with '(*)' suffix duplicates. Returns the sheet name
        whose week number is closest to target_week.

        Args:
            wb: The openpyxl Workbook instance.
            target_week: The ISO week number to find the closest template for.

        Returns:
            The sheet name closest to target_week, or None if no matching sheets found.
        """
        candidates = []

        for name in wb.sheetnames:
            match = SHEET_NAME_PATTERN.match(name)
            if match:
                week_num = int(match.group(1))
                # Only consider sheets without suffix (no "(N)") as primary templates
                # But also include suffixed ones as fallback candidates
                has_suffix = '(' in name
                if not has_suffix:
                    candidates.append((name, week_num))

        if not candidates:
            # Fallback: include suffixed sheets if no clean ones found
            for name in wb.sheetnames:
                match = SHEET_NAME_PATTERN.match(name)
                if match:
                    week_num = int(match.group(1))
                    candidates.append((name, week_num))

        if not candidates:
            logger.warning("No template sheets matching '26排程表-wXX' pattern found.")
            return None

        # Sort by distance to target_week, prefer lower week numbers on tie
        candidates.sort(key=lambda x: (abs(x[1] - target_week), x[1]))
        closest_name = candidates[0][0]

        logger.info(
            f"Found closest template sheet '{closest_name}' for target week {target_week}"
        )
        return closest_name

    def _copy_and_rename_sheet(
        self, wb: Workbook, source_name: str, target_week: int
    ) -> Worksheet:
        """
        Copy the source sheet and rename the copy to '26排程表-w{XX}'.

        Args:
            wb: The openpyxl Workbook instance.
            source_name: Name of the source sheet to copy.
            target_week: The target ISO week number for naming.

        Returns:
            The newly created worksheet.
        """
        target_name = f"26排程表-w{target_week:02d}"

        # Check if target sheet already exists
        if target_name in wb.sheetnames:
            logger.info(f"Sheet '{target_name}' already exists, returning existing sheet.")
            return wb[target_name]

        source_ws = wb[source_name]
        new_ws = wb.copy_worksheet(source_ws)
        new_ws.title = target_name

        logger.info(f"Copied sheet '{source_name}' to '{target_name}'")
        return new_ws

    def _fill_day_dates(
        self, ws: Worksheet, target_week: int, year: int = 2026
    ) -> list[int]:
        """
        Find rows in H column containing '日期:' separator text and fill
        I column with the corresponding weekday date, J column with day name.

        Uses ISO week calculation to determine Monday-Friday dates for target_week.

        Args:
            ws: The worksheet to modify.
            target_week: The ISO week number.
            year: The year (defaults to 2026).

        Returns:
            List of row numbers where day separators were found.
        """
        # Calculate the Monday of the target ISO week
        monday = date.fromisocalendar(year, target_week, 1)

        # Generate weekday dates (Monday to Friday)
        weekday_dates = [monday + timedelta(days=i) for i in range(5)]

        # Find all day separator rows (H column == "日期:")
        separator_rows = []
        h_col = 8  # H column index

        for row in range(self.STATS_AREA_END_ROW + 1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=h_col).value
            if cell_value and str(cell_value).strip() == self.DAY_SEPARATOR_VALUE:
                separator_rows.append(row)

        if not separator_rows:
            logger.warning("No day separator rows ('日期:') found in worksheet.")
            return separator_rows

        # Fill dates for each separator row (up to 5 weekdays)
        i_col = 9   # I column index (date)
        j_col = 10  # J column index (day name)

        for idx, row_num in enumerate(separator_rows):
            if idx < len(weekday_dates):
                # Fill I column with the date
                ws.cell(row=row_num, column=i_col, value=weekday_dates[idx])
                # Fill J column with the Chinese day name
                ws.cell(row=row_num, column=j_col, value=self.DAY_NAMES[idx])

                logger.debug(
                    f"Row {row_num}: filled date={weekday_dates[idx]}, "
                    f"day={self.DAY_NAMES[idx]}"
                )
            else:
                # More separator rows than weekdays (e.g., Saturday section)
                # Fill with the corresponding date if within the week
                extra_day = monday + timedelta(days=idx)
                # Map weekday index to Chinese names (extend for weekend if needed)
                extended_day_names = ["一", "二", "三", "四", "五", "六", "日"]
                if idx < 7:
                    ws.cell(row=row_num, column=i_col, value=extra_day)
                    ws.cell(row=row_num, column=j_col, value=extended_day_names[idx])

        logger.info(
            f"Filled {min(len(separator_rows), len(weekday_dates))} day sections "
            f"for week {target_week} (year {year})"
        )
        return separator_rows

    def _map_entry_to_columns(self, entry: dict) -> dict[int, Any]:
        """
        Map a schedule entry dict to Excel column values (columns H through T).

        Column mapping:
            H (8):  滴定機 — machine_port
            I (9):  Marker — marker
            J (10): 凍乾機台 — freeze_dryer
            K (11): formula — derived from marker or empty
            L (12): 數量 — quantity
            M (13): 配藥同仁 — operator
            N (14): 日期 — date
            O (15): RD時間 — rd_time
            P (16): 滴定時間 — start_time
            Q (17): 結束 — end_time
            R (18): 工單 — work_order
            S (19): Lot (Batch) — batch
            T (20): 備註 — notes

        Args:
            entry: A schedule entry dict with keys matching generated_schedule columns.

        Returns:
            Dict mapping column index (8-20) to the value to write.
        """
        return {
            8: entry.get("machine_port", ""),
            9: entry.get("marker", ""),
            10: entry.get("freeze_dryer", ""),
            11: entry.get("formula", ""),  # K column: formula (often empty or derived)
            12: entry.get("quantity", ""),
            13: entry.get("operator", ""),
            14: entry.get("date", ""),
            15: entry.get("rd_time", ""),
            16: entry.get("start_time", ""),
            17: entry.get("end_time", ""),
            18: entry.get("work_order", ""),
            19: entry.get("batch", ""),
            20: entry.get("notes", ""),
        }

    def _fill_day_sections(
        self,
        ws: Worksheet,
        entries_by_date: dict[str, list[dict]],
        separator_rows: list[int],
    ) -> None:
        """
        Fill schedule entries into each day section between separator rows.

        For each day section (defined by consecutive separator_rows), writes the
        corresponding entries using _map_entry_to_columns. Dynamically inserts
        or deletes rows if the number of entries differs from the available
        template rows in that section.

        NEVER modifies rows 1-98 (statistics area).

        Args:
            ws: The worksheet to modify.
            entries_by_date: Dict mapping date keys (in order matching separator_rows)
                to lists of schedule entry dicts.
            separator_rows: Sorted list of row numbers where day separators are located.
                These are returned by _fill_day_dates().
        """
        if not separator_rows:
            logger.warning("No separator rows provided; cannot fill day sections.")
            return

        # Build list of date keys in order of separator_rows
        # We use the I column (9) value from each separator row to identify the date,
        # or fall back to positional matching with entries_by_date keys
        date_keys = list(entries_by_date.keys())

        # Process sections in reverse order so that row insertions/deletions
        # don't shift the row numbers of earlier sections
        for section_idx in range(len(separator_rows) - 1, -1, -1):
            sep_row = separator_rows[section_idx]

            # Determine the end boundary for this section
            if section_idx < len(separator_rows) - 1:
                next_sep_row = separator_rows[section_idx + 1]
            else:
                # Last section extends to the end of used rows
                next_sep_row = ws.max_row + 1

            # Data rows are between separator row + 1 and next separator - 1
            data_start = sep_row + 1
            data_end = next_sep_row - 1  # last available template row

            # Count existing template data rows in this section
            template_row_count = max(0, data_end - data_start + 1)

            # Get entries for this section by positional index
            if section_idx < len(date_keys):
                day_key = date_keys[section_idx]
                entries = entries_by_date.get(day_key, [])
            else:
                entries = []

            entry_count = len(entries)

            # Dynamic row management
            if entry_count > template_row_count:
                # Need to insert rows — insert after data_start so we have enough
                rows_to_insert = entry_count - template_row_count
                # Insert rows at the end of the current data section (before next sep)
                insert_at = data_start + template_row_count
                ws.insert_rows(insert_at, amount=rows_to_insert)

                # Update separator_rows for sections that come BEFORE this one
                # (processed later since we go in reverse) — but since we're in
                # reverse order, earlier sections have lower indices and lower row
                # numbers, so they are NOT affected by insertions below them.
                # However, we need to update next separator rows for correct boundary calc.
                # Since we process in reverse, subsequent iterations use original values
                # which are all above this point — no adjustment needed for reverse.

                logger.debug(
                    f"Section {section_idx} (row {sep_row}): inserted {rows_to_insert} "
                    f"rows at row {insert_at}"
                )

            elif entry_count < template_row_count and entry_count >= 0:
                # Delete excess rows from the end of this section
                rows_to_delete = template_row_count - entry_count
                # Delete from the bottom of the data section to preserve top rows
                delete_start = data_start + entry_count
                # Safety: never delete rows in the stats area
                if delete_start <= self.STATS_AREA_END_ROW:
                    logger.warning(
                        f"Skipping row deletion at row {delete_start} — "
                        f"would affect statistics area."
                    )
                    continue
                ws.delete_rows(delete_start, amount=rows_to_delete)

                logger.debug(
                    f"Section {section_idx} (row {sep_row}): deleted {rows_to_delete} "
                    f"rows starting at row {delete_start}"
                )

            # Write entries into the data rows
            for i, entry in enumerate(entries):
                row_num = data_start + i

                # Safety check: never write to stats area
                if row_num <= self.STATS_AREA_END_ROW:
                    logger.warning(
                        f"Skipping write to row {row_num} — statistics area."
                    )
                    continue

                col_values = self._map_entry_to_columns(entry)
                for col_idx, value in col_values.items():
                    ws.cell(row=row_num, column=col_idx, value=value)

        logger.info(
            f"Filled {sum(len(v) for v in entries_by_date.values())} entries "
            f"across {len(separator_rows)} day sections."
        )

    def sync_to_excel(
        self,
        schedule_entries: list[dict],
        target_week: int,
        year: int = 2026,
    ) -> SyncResult:
        """
        Orchestrate full Excel sync: open workbook → find template → copy →
        fill dates → fill schedule data → save.

        This method is decoupled from the RDS transaction. If the Excel file
        doesn't exist or any write step fails, it logs the error and returns
        a SyncResult with status='failed' — it does NOT raise exceptions to
        the caller.

        Args:
            schedule_entries: List of schedule entry dicts (as from GeneratedSchedule).
                Each entry should have keys: date, marker, machine_port,
                freeze_dryer, operator, rd_time, start_time, end_time,
                quantity, pn, batch, work_order, notes, formula (optional).
            target_week: The ISO week number to sync (e.g., 24 for W24).
            year: The year for ISO week calculation (defaults to 2026).

        Returns:
            SyncResult with status='success' if all steps completed,
            or status='failed' with error message if any step fails.
        """
        try:
            # Step 1: Open workbook (keep_vba=True to preserve macros in .xlsm)
            logger.info(
                f"Starting Excel sync for week {target_week} "
                f"({len(schedule_entries)} entries)"
            )
            wb = load_workbook(self.excel_path, keep_vba=True)

        except FileNotFoundError:
            error_msg = f"Excel file not found: {self.excel_path}"
            logger.error(f"[sync_to_excel] {error_msg}")
            return SyncResult(status="failed", error=error_msg)
        except Exception as e:
            error_msg = f"Failed to open Excel workbook: {e}"
            logger.error(f"[sync_to_excel] {error_msg}")
            return SyncResult(status="failed", error=error_msg)

        try:
            # Step 2: Find closest template sheet
            template_name = self._find_closest_template_sheet(wb, target_week)
            if template_name is None:
                error_msg = (
                    "No template sheet matching '26排程表-wXX' pattern found "
                    "in workbook."
                )
                logger.error(f"[sync_to_excel] {error_msg}")
                return SyncResult(status="failed", error=error_msg)

            # Step 3: Copy and rename to target week
            ws = self._copy_and_rename_sheet(wb, template_name, target_week)

            # Step 4: Fill day dates
            separator_rows = self._fill_day_dates(ws, target_week, year)

            # Step 5: Group entries by date and fill day sections
            entries_by_date = self._group_entries_by_date(schedule_entries)
            self._fill_day_sections(ws, entries_by_date, separator_rows)

            # Step 6: Save workbook
            wb.save(self.excel_path)
            logger.info(
                f"Excel sync completed successfully for week {target_week}"
            )
            return SyncResult(status="success")

        except Exception as e:
            error_msg = f"Excel sync failed during write: {e}"
            logger.error(f"[sync_to_excel] {error_msg}")
            return SyncResult(status="failed", error=error_msg)

    def _group_entries_by_date(
        self, schedule_entries: list[dict]
    ) -> dict[str, list[dict]]:
        """
        Group schedule entries by their date value, maintaining insertion order.

        Args:
            schedule_entries: List of entry dicts with a 'date' key.

        Returns:
            OrderedDict-like dict mapping date string keys to lists of entries
            for that date, ordered by first appearance.
        """
        from collections import OrderedDict

        grouped: dict[str, list[dict]] = OrderedDict()
        for entry in schedule_entries:
            entry_date = entry.get("date")
            # Normalize date to string key
            if hasattr(entry_date, 'isoformat'):
                date_key = entry_date.isoformat()
            elif entry_date is not None:
                date_key = str(entry_date)
            else:
                date_key = "unknown"

            if date_key not in grouped:
                grouped[date_key] = []
            grouped[date_key].append(entry)

        return grouped
