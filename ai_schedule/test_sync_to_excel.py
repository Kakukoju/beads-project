"""
Unit tests for ExcelSyncService.sync_to_excel() orchestrator — Task 9.3
"""

import os
import sys
import tempfile
from datetime import date, time

import pytest
from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from ai_schedule.excel_sync_service import ExcelSyncService, SyncResult


class TestSyncResult:
    """Test the SyncResult dataclass."""

    def test_success_result(self):
        result = SyncResult(status="success")
        assert result.status == "success"
        assert result.error is None

    def test_failed_result(self):
        result = SyncResult(status="failed", error="Something went wrong")
        assert result.status == "failed"
        assert result.error == "Something went wrong"

    def test_default_error_is_none(self):
        result = SyncResult(status="success")
        assert result.error is None


class TestSyncToExcel:
    """Test sync_to_excel orchestrator method."""

    def _create_test_workbook(self, tmp_path, sheet_name="26排程表-w22"):
        """Create a minimal .xlsm workbook with a template sheet for testing."""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Add day separators after stats area (row 98)
        separator_rows = [99, 115, 131, 147, 163]
        for row in separator_rows:
            ws.cell(row=row, column=8, value="日期:")

        # Ensure max_row is sufficient
        ws.cell(row=180, column=1, value="")

        filepath = os.path.join(tmp_path, "test_schedule.xlsm")
        wb.save(filepath)
        return filepath

    def _make_entries(self, target_week=24, year=2026, count=3):
        """Create sample schedule entries for a given week."""
        from datetime import timedelta
        monday = date.fromisocalendar(year, target_week, 1)
        entries = []
        for i in range(count):
            entries.append({
                "date": monday + timedelta(days=i % 5),
                "marker": f"Marker-{i+1}",
                "machine_port": f"P{i+1}",
                "freeze_dryer": f"FD{i+1}",
                "operator": f"Op{i+1}",
                "rd_time": time(8, 0),
                "start_time": time(9, 0),
                "end_time": time(12, 0),
                "quantity": 1000 + i * 100,
                "pn": f"PN-{i+1:03d}",
                "batch": f"B2026-{target_week:02d}-{i+1:03d}",
                "work_order": f"WO-{i+1}",
                "notes": f"Note {i+1}",
                "formula": "",
            })
        return entries

    def test_success_sync(self, tmp_path):
        """Test successful sync creates sheet and fills data."""
        filepath = self._create_test_workbook(str(tmp_path))
        svc = ExcelSyncService(excel_path=filepath)
        entries = self._make_entries(target_week=24)

        result = svc.sync_to_excel(entries, target_week=24, year=2026)

        assert result.status == "success"
        assert result.error is None

        # Verify the new sheet was created
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        assert "26排程表-w24" in wb.sheetnames

    def test_file_not_found(self, tmp_path):
        """Test graceful handling when Excel file doesn't exist."""
        svc = ExcelSyncService(excel_path="/nonexistent/path/file.xlsm")
        entries = self._make_entries()

        result = svc.sync_to_excel(entries, target_week=24)

        assert result.status == "failed"
        assert "not found" in result.error.lower() or "No such file" in result.error

    def test_no_template_sheet(self, tmp_path):
        """Test handling when no matching template sheet exists."""
        wb = Workbook()
        ws = wb.active
        ws.title = "SomeOtherSheet"
        filepath = os.path.join(str(tmp_path), "no_template.xlsm")
        wb.save(filepath)

        svc = ExcelSyncService(excel_path=filepath)
        entries = self._make_entries()

        result = svc.sync_to_excel(entries, target_week=24)

        assert result.status == "failed"
        assert "template" in result.error.lower() or "pattern" in result.error.lower()

    def test_empty_entries(self, tmp_path):
        """Test sync with empty entries list."""
        filepath = self._create_test_workbook(str(tmp_path))
        svc = ExcelSyncService(excel_path=filepath)

        result = svc.sync_to_excel([], target_week=24, year=2026)

        assert result.status == "success"
        assert result.error is None

    def test_existing_target_sheet_reused(self, tmp_path):
        """Test that existing target sheet is reused (not duplicated)."""
        # Create workbook with both template and target
        wb = Workbook()
        ws = wb.active
        ws.title = "26排程表-w22"
        ws.cell(row=99, column=8, value="日期:")
        ws.cell(row=180, column=1, value="")

        ws2 = wb.create_sheet("26排程表-w24")
        ws2.cell(row=99, column=8, value="日期:")
        ws2.cell(row=180, column=1, value="")

        filepath = os.path.join(str(tmp_path), "existing_target.xlsm")
        wb.save(filepath)

        svc = ExcelSyncService(excel_path=filepath)
        entries = self._make_entries(target_week=24)

        result = svc.sync_to_excel(entries, target_week=24, year=2026)
        assert result.status == "success"


class TestGroupEntriesByDate:
    """Test _group_entries_by_date helper."""

    def test_groups_by_date_object(self):
        svc = ExcelSyncService()
        entries = [
            {"date": date(2026, 6, 8), "marker": "A"},
            {"date": date(2026, 6, 8), "marker": "B"},
            {"date": date(2026, 6, 9), "marker": "C"},
        ]

        result = svc._group_entries_by_date(entries)

        assert len(result) == 2
        assert len(result["2026-06-08"]) == 2
        assert len(result["2026-06-09"]) == 1

    def test_groups_by_string_date(self):
        svc = ExcelSyncService()
        entries = [
            {"date": "2026-06-08", "marker": "A"},
            {"date": "2026-06-08", "marker": "B"},
        ]

        result = svc._group_entries_by_date(entries)
        assert len(result) == 1
        assert "2026-06-08" in result

    def test_none_date_grouped_as_unknown(self):
        svc = ExcelSyncService()
        entries = [
            {"date": None, "marker": "A"},
        ]

        result = svc._group_entries_by_date(entries)
        assert "unknown" in result

    def test_empty_entries(self):
        svc = ExcelSyncService()
        result = svc._group_entries_by_date([])
        assert result == {}

    def test_preserves_order(self):
        svc = ExcelSyncService()
        entries = [
            {"date": date(2026, 6, 9), "marker": "B"},
            {"date": date(2026, 6, 8), "marker": "A"},
            {"date": date(2026, 6, 10), "marker": "C"},
        ]

        result = svc._group_entries_by_date(entries)
        keys = list(result.keys())
        # First appearance order
        assert keys == ["2026-06-09", "2026-06-08", "2026-06-10"]


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
