"""
Unit tests for ExcelSyncService — Task 9.1: template finding and sheet operations
"""

import os
import shutil
import tempfile
from datetime import date

import pytest
from openpyxl import Workbook, load_workbook

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from ai_schedule.excel_sync_service import ExcelSyncService, SHEET_NAME_PATTERN


class TestSheetNamePattern:
    """Test the regex pattern for matching sheet names."""

    def test_matches_standard_format(self):
        assert SHEET_NAME_PATTERN.match("26排程表-w23") is not None

    def test_matches_zero_padded(self):
        assert SHEET_NAME_PATTERN.match("26排程表-w09") is not None

    def test_matches_single_digit(self):
        assert SHEET_NAME_PATTERN.match("26排程表-w5") is not None

    def test_matches_with_suffix(self):
        m = SHEET_NAME_PATTERN.match("26排程表-w23 (2)")
        assert m is not None
        assert int(m.group(1)) == 23

    def test_no_match_without_digits(self):
        assert SHEET_NAME_PATTERN.match("26排程表-w") is None

    def test_no_match_unrelated_sheet(self):
        assert SHEET_NAME_PATTERN.match("滴定") is None

    def test_extracts_week_number(self):
        m = SHEET_NAME_PATTERN.match("26排程表-w14")
        assert int(m.group(1)) == 14


class TestExcelSyncServiceInit:
    """Test __init__ with default and custom paths."""

    def test_default_path(self):
        svc = ExcelSyncService()
        assert svc.excel_path == ExcelSyncService.EXCEL_PATH

    def test_custom_path(self):
        svc = ExcelSyncService(excel_path="/tmp/custom.xlsm")
        assert svc.excel_path == "/tmp/custom.xlsm"


class TestFindClosestTemplateSheet:
    """Test _find_closest_template_sheet logic."""

    def _create_wb_with_sheets(self, sheet_names):
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        for name in sheet_names:
            wb.create_sheet(title=name)
        return wb

    def test_exact_match(self):
        svc = ExcelSyncService()
        wb = self._create_wb_with_sheets(["26排程表-w20", "26排程表-w22", "26排程表-w24"])
        result = svc._find_closest_template_sheet(wb, 22)
        assert result == "26排程表-w22"

    def test_closest_lower(self):
        svc = ExcelSyncService()
        wb = self._create_wb_with_sheets(["26排程表-w20", "26排程表-w22"])
        result = svc._find_closest_template_sheet(wb, 23)
        assert result == "26排程表-w22"

    def test_closest_higher(self):
        svc = ExcelSyncService()
        wb = self._create_wb_with_sheets(["26排程表-w20", "26排程表-w25"])
        result = svc._find_closest_template_sheet(wb, 23)
        assert result == "26排程表-w25"

    def test_ignores_suffix_sheets_when_clean_available(self):
        svc = ExcelSyncService()
        wb = self._create_wb_with_sheets([
            "26排程表-w22",
            "26排程表-w23 (2)",  # suffix - should be ignored
        ])
        result = svc._find_closest_template_sheet(wb, 23)
        assert result == "26排程表-w22"

    def test_uses_suffix_sheets_as_fallback(self):
        svc = ExcelSyncService()
        wb = self._create_wb_with_sheets([
            "26排程表-w23 (2)",
            "滴定",
        ])
        result = svc._find_closest_template_sheet(wb, 24)
        assert result == "26排程表-w23 (2)"

    def test_no_matching_sheets(self):
        svc = ExcelSyncService()
        wb = self._create_wb_with_sheets(["滴定", "其他Sheet"])
        result = svc._find_closest_template_sheet(wb, 24)
        assert result is None

    def test_empty_workbook(self):
        svc = ExcelSyncService()
        wb = Workbook()
        result = svc._find_closest_template_sheet(wb, 24)
        assert result is None


class TestCopyAndRenameSheet:
    """Test _copy_and_rename_sheet."""

    def test_creates_new_sheet(self):
        svc = ExcelSyncService()
        wb = Workbook()
        wb.active.title = "26排程表-w22"
        # Put some data in source
        wb.active.cell(row=1, column=1, value="test_data")

        new_ws = svc._copy_and_rename_sheet(wb, "26排程表-w22", 24)
        assert new_ws.title == "26排程表-w24"
        assert "26排程表-w24" in wb.sheetnames
        assert new_ws.cell(row=1, column=1).value == "test_data"

    def test_returns_existing_sheet(self):
        svc = ExcelSyncService()
        wb = Workbook()
        wb.active.title = "26排程表-w24"
        wb.active.cell(row=1, column=1, value="existing")

        result = svc._copy_and_rename_sheet(wb, "26排程表-w24", 24)
        assert result.title == "26排程表-w24"
        assert result.cell(row=1, column=1).value == "existing"

    def test_zero_pads_single_digit_week(self):
        svc = ExcelSyncService()
        wb = Workbook()
        wb.active.title = "26排程表-w02"

        new_ws = svc._copy_and_rename_sheet(wb, "26排程表-w02", 5)
        assert new_ws.title == "26排程表-w05"


class TestFillDayDates:
    """Test _fill_day_dates."""

    def _create_ws_with_separators(self, separator_rows):
        """Create a worksheet with day separators in H column."""
        wb = Workbook()
        ws = wb.active
        for row in separator_rows:
            ws.cell(row=row, column=8, value="日期:")
        return ws

    def test_fills_five_weekdays(self):
        svc = ExcelSyncService()
        separator_rows = [99, 120, 140, 160, 180]
        ws = self._create_ws_with_separators(separator_rows)

        result = svc._fill_day_dates(ws, 24, year=2026)
        assert result == separator_rows

        # Verify Monday
        assert ws.cell(row=99, column=9).value == date(2026, 6, 8)
        assert ws.cell(row=99, column=10).value == "一"

        # Verify Friday
        assert ws.cell(row=180, column=9).value == date(2026, 6, 12)
        assert ws.cell(row=180, column=10).value == "五"

    def test_fills_with_six_separators(self):
        """Template may have Saturday section — should handle gracefully."""
        svc = ExcelSyncService()
        separator_rows = [99, 120, 140, 160, 180, 200]
        ws = self._create_ws_with_separators(separator_rows)

        result = svc._fill_day_dates(ws, 24, year=2026)

        # Saturday gets filled with extended day name
        assert ws.cell(row=200, column=9).value == date(2026, 6, 13)
        assert ws.cell(row=200, column=10).value == "六"

    def test_no_separators(self):
        svc = ExcelSyncService()
        wb = Workbook()
        ws = wb.active

        result = svc._fill_day_dates(ws, 24)
        assert result == []

    def test_ignores_separators_in_stats_area(self):
        """Separators in rows 1-98 should be ignored."""
        svc = ExcelSyncService()
        wb = Workbook()
        ws = wb.active
        # Put separator in stats area (should be ignored)
        ws.cell(row=50, column=8, value="日期:")
        # Put separator after stats area
        ws.cell(row=99, column=8, value="日期:")

        result = svc._fill_day_dates(ws, 24, year=2026)
        assert result == [99]
        assert ws.cell(row=99, column=9).value == date(2026, 6, 8)

    def test_week_1_of_year(self):
        """Test edge case of week 1."""
        svc = ExcelSyncService()
        separator_rows = [99, 120, 140, 160, 180]
        ws = self._create_ws_with_separators(separator_rows)

        result = svc._fill_day_dates(ws, 1, year=2026)

        # Week 1 of 2026: Monday is 2025-12-29 (ISO week 1 can start in previous year)
        expected_monday = date.fromisocalendar(2026, 1, 1)
        assert ws.cell(row=99, column=9).value == expected_monday


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
